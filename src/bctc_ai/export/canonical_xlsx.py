from __future__ import annotations

import zipfile
from collections.abc import Sequence
from datetime import datetime
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any

from openpyxl.xml.functions import fromstring, tostring

CANONICAL_CORE_TIMESTAMP = datetime(2000, 1, 1)
CORE_PROPERTIES_MEMBER = "docProps/core.xml"
ZIP_MEMBER_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
EXCEL_CELL_TEXT_LIMIT = 32_767


class CanonicalXlsxError(ValueError):
    """Raised when an XLSX package cannot be serialized canonically."""


def set_literal_cell(cell: Any, value: object) -> None:
    """Set a cell while preventing source text from becoming an Excel formula."""

    if isinstance(value, str) and len(value) > EXCEL_CELL_TEXT_LIMIT:
        raise CanonicalXlsxError("source text exceeds Excel's cell text limit")
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def append_literal_row(sheet: Any, values: Sequence[object]) -> None:
    row = 1 if sheet.max_row == 1 and sheet["A1"].value is None else sheet.max_row + 1
    for column, value in enumerate(values, start=1):
        set_literal_cell(sheet.cell(row, column), value)


def workbook_has_formula(workbook: Any) -> bool:
    return any(
        cell.data_type == "f"
        or (isinstance(cell.value, str) and cell.value.startswith("=") and cell.data_type != "s")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def freeze_core_properties(workbook: Any, *, creator: str) -> bytes:
    if not isinstance(creator, str) or not creator or len(creator) > 255:
        raise CanonicalXlsxError("canonical XLSX creator is invalid")
    properties = workbook.properties
    properties.creator = creator
    properties.lastModifiedBy = creator
    properties.created = CANONICAL_CORE_TIMESTAMP
    properties.modified = CANONICAL_CORE_TIMESTAMP
    properties.version = "1"
    properties.revision = "1"
    properties.title = None
    properties.subject = None
    properties.description = None
    properties.identifier = None
    properties.language = None
    properties.keywords = None
    properties.category = None
    properties.contentStatus = None
    properties.lastPrinted = None
    return tostring(properties.to_tree())


def _validate_member_name(name: str) -> None:
    path = PurePosixPath(name)
    raw_parts = name.split("/")
    if (
        not name
        or "\\" in name
        or name.startswith("/")
        or path.is_absolute()
        or any(part in {"", ".", ".."} for part in raw_parts)
    ):
        raise CanonicalXlsxError(f"unsafe XLSX ZIP member name: {name!r}")


def canonicalize_xlsx_zip(payload: bytes, *, core_properties_xml: bytes) -> bytes:
    """Return an XLSX ZIP with stable member order, metadata, and core properties."""

    if not isinstance(payload, bytes) or not payload:
        raise CanonicalXlsxError("generated XLSX payload is empty")
    if not isinstance(core_properties_xml, bytes) or not core_properties_xml:
        raise CanonicalXlsxError("canonical XLSX core properties are empty")
    try:
        fromstring(core_properties_xml)
    except Exception as exc:
        raise CanonicalXlsxError("canonical XLSX core properties are invalid XML") from exc

    destination = BytesIO()
    try:
        with (
            zipfile.ZipFile(BytesIO(payload), "r") as source,
            zipfile.ZipFile(
                destination,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            ) as output,
        ):
            names = source.namelist()
            if len(names) != len(set(names)):
                raise CanonicalXlsxError("generated XLSX contains duplicate ZIP members")
            if names.count(CORE_PROPERTIES_MEMBER) != 1:
                raise CanonicalXlsxError(
                    "generated XLSX must contain exactly one core-properties member"
                )
            for name in names:
                _validate_member_name(name)
            for name in sorted(names):
                member = (
                    core_properties_xml if name == CORE_PROPERTIES_MEMBER else source.read(name)
                )
                info = zipfile.ZipInfo(name, date_time=ZIP_MEMBER_TIMESTAMP)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = 0o600 << 16
                info.flag_bits = 0x800
                info.extra = b""
                info.comment = b""
                output.writestr(info, member, compresslevel=9)
    except CanonicalXlsxError:
        raise
    except (OSError, RuntimeError, ValueError, zipfile.BadZipFile) as exc:
        raise CanonicalXlsxError("cannot canonicalize generated XLSX") from exc
    return destination.getvalue()


def deterministic_workbook_bytes(workbook: Any, *, creator: str) -> bytes:
    """Serialize an OpenPyXL workbook to deterministic XLSX bytes."""

    core_properties = freeze_core_properties(workbook, creator=creator)
    raw = BytesIO()
    try:
        workbook.save(raw)
    except Exception as exc:
        raise CanonicalXlsxError("cannot serialize OpenPyXL workbook") from exc
    return canonicalize_xlsx_zip(raw.getvalue(), core_properties_xml=core_properties)


__all__ = [
    "CANONICAL_CORE_TIMESTAMP",
    "CORE_PROPERTIES_MEMBER",
    "EXCEL_CELL_TEXT_LIMIT",
    "ZIP_MEMBER_TIMESTAMP",
    "CanonicalXlsxError",
    "append_literal_row",
    "canonicalize_xlsx_zip",
    "deterministic_workbook_bytes",
    "freeze_core_properties",
    "set_literal_cell",
    "workbook_has_formula",
]
