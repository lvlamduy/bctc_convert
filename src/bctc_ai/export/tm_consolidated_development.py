"""Fail-closed consolidated MBB TM development export.

The TM mapping layer is deliberately page-owned and its production results are
heterogeneous.  This module consumes those result objects without weakening
their page-specific contracts, normalizes only their export surface, and keeps
all non-observation schema states explicit.  It does not use the generic
``export.workbook`` evidence/value semantics.
"""

from __future__ import annotations

import hashlib
import importlib
import json
import math
import os
import re
import zipfile
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, fields, is_dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.xml.functions import fromstring, tostring

from bctc_ai.core.text import normalize_text
from bctc_ai.schema.registry import UNIVERSAL_TM_SCHEMA_ITEM_COUNT, SchemaItem

TM_CONSOLIDATED_POLICY_RELATIVE_PATH = Path("config/export/tm-consolidated-development-v1.yaml")
TM_CONSOLIDATED_SHEETS = (
    "TM",
    "OBSERVATIONS",
    "PROVENANCE",
    "VALIDATION",
    "RUN_METADATA",
)
TM_CONSOLIDATED_SCHEMA_COUNT = UNIVERSAL_TM_SCHEMA_ITEM_COUNT
TM_CONSOLIDATED_TEMPLATE_SHA256 = "8d9e76de0d42aa26591a87a5e2d522e7a69e089528047928b029ac4ed49f2b3c"
TM_CONSOLIDATED_SCHEMA_PROJECTION_SHA256 = (
    "194df64364a4dd2452252585770128697168feb7014961479dfcbd8db942b695"
)
TM_UNIVERSAL_SCHEMA_NAME = "UNIVERSAL_BANK_BCTC_SCHEMA"
TM_UNIVERSAL_SCHEMA_REVISION = "UNIVERSAL_BANK_BCTC_SCHEMA@6074"
TM_BASE_SCHEMA_COUNT = 1_593
TM_UNIVERSAL_SCHEMA_COUNT = 1_953
TM_DOCUMENT_NEW_REPORT_NORM_IDS = (
    *range(5_718, 6_034),
    *range(6_057, 6_075),
)
TM_LATEST_SCHEMA_BATCH_REPORT_NORM_IDS = (6_073, 6_074)
_PRODUCTION_SCHEMA_IDENTITY: dict[str, object] = {
    "schema_name": TM_UNIVERSAL_SCHEMA_NAME,
    "schema_revision": TM_UNIVERSAL_SCHEMA_REVISION,
    "base_schema": {
        "name": "BASE_SCHEMA",
        "item_count": TM_BASE_SCHEMA_COUNT,
        "statement_counts": {"CDKT": 77, "KQKD": 24, "LCTT": 107, "TM": 1_385},
        "ordered_canonical_projection_sha256": (
            "e63b77ebf99907843bea419cef32bc64cd709129813f89309f3b42fc818a1b10"
        ),
        "ordered_report_norm_ids_sha256": (
            "5cc0e9ea70b23af236ce43b920838299dbc91e9c0ef19d31165f4ce49eea4f9f"
        ),
    },
    "universal_schema": {
        "item_count": TM_UNIVERSAL_SCHEMA_COUNT,
        "statement_counts": {"CDKT": 99, "KQKD": 25, "LCTT": 110, "TM": 1_719},
        "high_watermark": 6_074,
        "ordered_canonical_projection_sha256": (
            "b6db1a5abe9cfe62c7cd55e43ebda6b7ca1bddd27d22c9761ea808a8eb3d8778"
        ),
        "ordered_report_norm_ids_sha256": (
            "749580743080c4b0336b9b9bc488a92df6ccdbd57e5fa9fcaf674c7f7729413a"
        ),
        "schema_graph_sha256": ("e3845a2f72995445d0519dac3036a0a34d9013c29154c71608c56a944310251b"),
        "universal_schema_sha256": (
            "e3845a2f72995445d0519dac3036a0a34d9013c29154c71608c56a944310251b"
        ),
    },
    "accepted_post_base_tm_additions": {
        "first_report_norm_id": 5_718,
        "last_report_norm_id": 6_074,
        "item_count": len(TM_DOCUMENT_NEW_REPORT_NORM_IDS),
    },
    "latest_schema_batch": {
        "first_report_norm_id": 6_073,
        "last_report_norm_id": 6_074,
        "item_count": len(TM_LATEST_SCHEMA_BATCH_REPORT_NORM_IDS),
    },
}

_FIXED_TIMESTAMP = datetime(2000, 1, 1)
_CORE_PROPERTIES_MEMBER = "docProps/core.xml"
_MODIFIED_PROPERTY_TAG = "{http://purl.org/dc/terms/}modified"
_SHA256 = re.compile(r"^[0-9a-f]{64}$")
_TYPE_PATH = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_OWNER_KEY = re.compile(r"^[a-z0-9][a-z0-9-]*$")
_ADAPTERS = {
    "auto",
    "page30",
    "mapped_assignments",
    "source_dispositions",
    "nested_assignments",
    "no_observations",
}
_CANONICAL_SCHEMA_STATUSES = {
    "MAPPED",
    "AMBIGUOUS",
    "UNRESOLVED",
    "NOT_OBSERVED",
    "NA",
}
_RAW_SCHEMA_STATUS = {
    "MAPPED": "MAPPED",
    "MAPPED_AUTOMATIC": "MAPPED",
    "MAPPED_AUTOMATIC_SCOPED": "MAPPED",
    "MAPPED_STRUCTURAL_SCOPED": "MAPPED",
    "AMBIGUOUS": "AMBIGUOUS",
    "AMBIGUOUS_MAPPING": "AMBIGUOUS",
    "UNRESOLVED": "UNRESOLVED",
    "UNRESOLVED_MAPPING": "UNRESOLVED",
    "NOT_OBSERVED": "NOT_OBSERVED",
    "NOT_OBSERVED_IN_THIS_PDF": "NOT_OBSERVED",
    "NA": "NA",
    "NOT_APPLICABLE": "NA",
    "SCHEMA_ITEM_NOT_APPLICABLE": "NA",
    "UNASSESSED": None,
}
_OBSERVATION_STATUSES = {"VALUE", "ZERO", "DASH", "BLANK"}
_PERIOD_TYPES = {"SNAPSHOT", "DURATION", "DURATION_PANEL", "FLOW"}
_DERIVATION_MARKERS = ("AGGREGAT", "CALCULAT", "DERIV", "IMPUT", "SUM")
_PAGE_VALIDATION_FAMILIES = (
    "accounting_checks",
    "percentage_checks",
    "hierarchy_checks",
    "catch_all_checks",
    "duplicate_checks",
    "validation_checks",
    "narrative_diagnostic",
)
_VALIDATION_ID_FIELDS = ("check_id", "diagnostic_id", "validation_id")


class TMConsolidatedDevelopmentExportError(ValueError):
    """Raised before export whenever a frozen TM invariant is not satisfied."""


@dataclass(frozen=True)
class TMConsolidatedOwnerPolicy:
    owner_key: str
    result_type: str
    adapter: str
    default_period_type: str | None = None
    default_period_roles: tuple[str, ...] = ()


@dataclass(frozen=True)
class TMConsolidatedExportPolicy:
    source_path: Path
    policy_sha256: str
    statement_type: str
    bank: str
    report_scope: str
    dataset_role: str
    schema_item_count: int
    schema_workbook_path: Path
    schema_workbook_sha256: str
    schema_projection_sha256: str
    output_sheets: tuple[str, ...]
    owners: tuple[TMConsolidatedOwnerPolicy, ...]
    schema_identity: Mapping[str, object] | None = None


@dataclass(frozen=True)
class TMConsolidatedOwnerInput:
    owner_key: str
    result: object


@dataclass(frozen=True)
class TMConsolidatedDevelopmentArtifacts:
    workbook_bytes: bytes
    provenance_bytes: bytes
    workbook_sha256: str
    provenance_sha256: str
    schema_item_count: int
    observation_count: int
    provenance_count: int
    status_counts: dict[str, int]
    fully_verified: bool = False


@dataclass(frozen=True)
class TMConsolidatedDevelopmentExportResult:
    workbook_path: str
    provenance_path: str
    workbook_sha256: str
    provenance_sha256: str
    workbook_size_bytes: int
    provenance_size_bytes: int
    schema_item_count: int
    observation_count: int
    provenance_count: int
    status_counts: dict[str, int]
    fully_verified: bool = False


@dataclass(frozen=True)
class _OwnerContext:
    policy: TMConsolidatedOwnerPolicy
    result: object
    result_type: str
    result_sha256: str
    authority_scope: str


@dataclass(frozen=True)
class _SchemaExportRow:
    report_norm_id: int
    display_order: int
    canonical_name: str
    schema_status: str
    raw_schema_status: str
    owner_key: str
    mapping_authority_scope: str
    report_scope: str
    source_ids: tuple[str, ...]
    reason: str
    observation_count: int = 0


@dataclass(frozen=True)
class _ObservationRecord:
    observation_key: str
    provenance_key: str
    report_norm_id: int
    canonical_name: str
    owner_key: str
    mapping_authority_scope: str
    schema_status: str
    value_status: str
    reported_value: Decimal | None
    canonical_value: Decimal | None
    period_start: str | None
    period_end: str | None
    period_type: str | None
    period_role: str
    unit: str
    unit_multiplier: int
    scope: str
    axis_key: str
    source_record_id: str
    source_ids: tuple[str, ...]
    observation_origin: str
    derivation_method: str | None
    derivation_source_ids: tuple[str, ...]
    derivation_component_report_norm_ids: tuple[int, ...]
    mapping_basis: str


@dataclass(frozen=True)
class _ProvenanceRecord:
    provenance_key: str
    observation_key: str
    owner_key: str
    report_norm_id: int
    page_tag: str
    page_number: int | None
    source_record_id: str
    source_ids: tuple[str, ...]
    raw_source_value: str | None
    source_bbox_json: str | None
    mapping_basis: str
    observation_origin: str
    derivation_method: str | None
    derivation_component_report_norm_ids: tuple[int, ...]
    mapping_result_sha256: str
    evidence_detail_sha256: str
    evidence_detail: object


@dataclass(frozen=True)
class _AdaptedAssignment:
    report_norm_id: int
    observation: str
    value: Decimal | None
    cell_index: int
    axis_key: str
    period_start: str | None = None
    period_end: str | None = None
    period_type: str | None = None
    period_role: str | None = None
    unit: str | None = None
    unit_multiplier: int | None = None
    mapping_basis: str = "DIRECT_SOURCE_DISPOSITION_ASSIGNMENT"


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _jsonable(value: object) -> object:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            raise TMConsolidatedDevelopmentExportError("non-finite float in TM evidence")
        return value
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise TMConsolidatedDevelopmentExportError("non-finite decimal in TM evidence")
        return format(value, "f")
    if isinstance(value, Enum):
        return _jsonable(value.value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Path):
        return value.name
    if is_dataclass(value) and not isinstance(value, type):
        result: dict[str, object] = {}
        for field in fields(value):
            field_value = getattr(value, field.name)
            if field.name.endswith("_path") and isinstance(field_value, (str, Path)):
                result[field.name] = Path(field_value).name
            else:
                result[field.name] = _jsonable(field_value)
        return result
    if isinstance(value, Mapping):
        result: dict[str, object] = {}
        for key in sorted(value, key=lambda candidate: str(candidate)):
            if not isinstance(key, (str, int)):
                raise TMConsolidatedDevelopmentExportError("unsupported TM evidence mapping key")
            result[str(key)] = _jsonable(value[key])
        return result
    if isinstance(value, (set, frozenset)):
        converted = [_jsonable(item) for item in value]
        return sorted(converted, key=lambda item: _canonical_json_bytes(item))
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return [_jsonable(item) for item in value]
    raise TMConsolidatedDevelopmentExportError(
        f"unsupported TM evidence type: {type(value).__module__}.{type(value).__qualname__}"
    )


def _canonical_json_bytes(payload: object) -> bytes:
    return (
        json.dumps(
            _jsonable(payload),
            ensure_ascii=False,
            allow_nan=False,
            indent=2,
            sort_keys=True,
        )
        + "\n"
    ).encode("utf-8")


def _compact_json(payload: object) -> str:
    return json.dumps(
        _jsonable(payload),
        ensure_ascii=False,
        allow_nan=False,
        separators=(",", ":"),
        sort_keys=True,
    )


def _required_string(payload: Mapping[str, object], key: str) -> str:
    value = payload.get(key)
    if not isinstance(value, str) or not value.strip():
        raise TMConsolidatedDevelopmentExportError(f"invalid TM export policy field: {key}")
    return value.strip()


def load_tm_consolidated_export_policy(path: Path) -> TMConsolidatedExportPolicy:
    """Load the pinned MBB consolidated TM export policy."""

    resolved = Path(path).resolve()
    try:
        raw_bytes = resolved.read_bytes()
        payload = yaml.safe_load(raw_bytes)
    except (OSError, yaml.YAMLError) as exc:
        raise TMConsolidatedDevelopmentExportError("cannot load TM export policy") from exc
    if not isinstance(payload, dict) or (
        payload.get("version") != 1
        or payload.get("policy") != "MBB_TM_CONSOLIDATED_DEVELOPMENT_EXPORT_V1"
        or payload.get("statement_type") != "TM"
        or payload.get("bank") != "MBB"
        or payload.get("report_scope") != "CONSOLIDATED"
        or payload.get("dataset_role") != "DEVELOPMENT"
    ):
        raise TMConsolidatedDevelopmentExportError("TM export policy identity drifted")
    schema = payload.get("schema")
    if not isinstance(schema, dict):
        raise TMConsolidatedDevelopmentExportError("TM export schema freeze is missing")
    schema_identity = payload.get("schema_identity")
    if schema_identity != _PRODUCTION_SCHEMA_IDENTITY:
        raise TMConsolidatedDevelopmentExportError("TM universal/base schema identity drifted")
    item_count = schema.get("item_count")
    workbook_path_text = _required_string(schema, "workbook_path")
    workbook_sha256 = _required_string(schema, "workbook_sha256")
    projection_sha256 = _required_string(schema, "projection_sha256")
    workbook_path = Path(workbook_path_text)
    if (
        item_count != TM_CONSOLIDATED_SCHEMA_COUNT
        or workbook_path.is_absolute()
        or ".." in workbook_path.parts
        or workbook_sha256 != TM_CONSOLIDATED_TEMPLATE_SHA256
        or projection_sha256 != TM_CONSOLIDATED_SCHEMA_PROJECTION_SHA256
        or _SHA256.fullmatch(workbook_sha256) is None
        or _SHA256.fullmatch(projection_sha256) is None
    ):
        raise TMConsolidatedDevelopmentExportError("TM export schema freeze drifted")
    output_sheets = payload.get("output_sheets")
    if not isinstance(output_sheets, list) or tuple(output_sheets) != TM_CONSOLIDATED_SHEETS:
        raise TMConsolidatedDevelopmentExportError("TM export sheet inventory drifted")
    raw_owners = payload.get("owners")
    if not isinstance(raw_owners, list) or not raw_owners:
        raise TMConsolidatedDevelopmentExportError("TM export owners are missing")
    owners: list[TMConsolidatedOwnerPolicy] = []
    for raw in raw_owners:
        if not isinstance(raw, dict):
            raise TMConsolidatedDevelopmentExportError("invalid TM export owner")
        owner_key = _required_string(raw, "owner_key")
        result_type = _required_string(raw, "result_type")
        adapter = _required_string(raw, "adapter")
        default_period_type = raw.get("default_period_type")
        roles = raw.get("default_period_roles", [])
        if (
            _OWNER_KEY.fullmatch(owner_key) is None
            or _TYPE_PATH.fullmatch(result_type) is None
            or adapter not in _ADAPTERS
            or (
                default_period_type is not None
                and default_period_type not in {"SNAPSHOT", "DURATION"}
            )
            or not isinstance(roles, list)
            or any(not isinstance(role, str) or not role for role in roles)
        ):
            raise TMConsolidatedDevelopmentExportError(f"invalid TM export owner: {owner_key}")
        owners.append(
            TMConsolidatedOwnerPolicy(
                owner_key=owner_key,
                result_type=result_type,
                adapter=adapter,
                default_period_type=default_period_type,
                default_period_roles=tuple(roles),
            )
        )
    owner_keys = [owner.owner_key for owner in owners]
    result_types = [owner.result_type for owner in owners]
    if (
        len(owner_keys) != len(set(owner_keys))
        or len(result_types) != len(set(result_types))
        or "page-0045" not in owner_keys
        or "residual" not in owner_keys
    ):
        raise TMConsolidatedDevelopmentExportError("TM export owner inventory drifted")
    return TMConsolidatedExportPolicy(
        source_path=resolved,
        policy_sha256=_sha256(raw_bytes),
        statement_type="TM",
        bank="MBB",
        report_scope="CONSOLIDATED",
        dataset_role="DEVELOPMENT",
        schema_item_count=item_count,
        schema_workbook_path=workbook_path,
        schema_workbook_sha256=workbook_sha256,
        schema_projection_sha256=projection_sha256,
        output_sheets=tuple(output_sheets),
        owners=tuple(owners),
        schema_identity=schema_identity,
    )


def audit_tm_consolidated_owner_result_contracts(
    policy: TMConsolidatedExportPolicy,
) -> dict[str, str]:
    """Import and audit every configured production mapping-result dataclass."""

    required_common = {
        "schema_item_count",
        "mapping_authority_scope",
        "mapping_authority_granted",
        "mapped_schema_count",
        "not_observed_schema_count",
        "unassessed_schema_count",
        "schema_dispositions",
    }
    required_by_adapter = {
        "page30": {"mapped_values", "source_dispositions"},
        "mapped_assignments": {"mapped_assignments", "source_dispositions"},
        "source_dispositions": {"source_dispositions"},
        "nested_assignments": {"source_dispositions"},
        "no_observations": set(),
        "auto": set(),
    }
    result: dict[str, str] = {}
    for owner in policy.owners:
        module_name, separator, class_name = owner.result_type.rpartition(".")
        if not separator:
            raise TMConsolidatedDevelopmentExportError(
                f"invalid configured TM result type: {owner.result_type}"
            )
        try:
            module = importlib.import_module(module_name)
            result_class = getattr(module, class_name)
        except (ImportError, AttributeError) as exc:
            raise TMConsolidatedDevelopmentExportError(
                f"cannot import TM result type: {owner.result_type}"
            ) from exc
        if not isinstance(result_class, type) or not is_dataclass(result_class):
            raise TMConsolidatedDevelopmentExportError(
                f"TM result type is not a dataclass: {owner.result_type}"
            )
        field_names = {field.name for field in fields(result_class)}
        required = required_common | required_by_adapter[owner.adapter]
        if owner.owner_key == "page-0046":
            required |= {"derived_assignment_count", "derived_assignments"}
        if not required <= field_names:
            raise TMConsolidatedDevelopmentExportError(
                f"TM result contract drifted for {owner.owner_key}: {sorted(required - field_names)}"
            )
        result[owner.owner_key] = owner.result_type
    return result


def bind_tm_consolidated_owner_results(
    owner_results: Mapping[str, object],
    policy: TMConsolidatedExportPolicy,
) -> tuple[TMConsolidatedOwnerInput, ...]:
    """Bind an unordered set of 27 actual mapper results to policy order."""

    if not isinstance(owner_results, Mapping) or any(
        not isinstance(key, str) or not key for key in owner_results
    ):
        raise TMConsolidatedDevelopmentExportError(
            "TM owner results must be a string-keyed mapping"
        )
    expected = {owner.owner_key for owner in policy.owners}
    actual = set(owner_results)
    if actual != expected:
        raise TMConsolidatedDevelopmentExportError(
            "TM owner result inventory incomplete; "
            f"missing={sorted(expected - actual)}, extra={sorted(actual - expected)}"
        )
    values = tuple(owner_results[owner.owner_key] for owner in policy.owners)
    if len({id(value) for value in values}) != len(values):
        raise TMConsolidatedDevelopmentExportError("duplicate TM mapping-result object")
    inputs = tuple(
        TMConsolidatedOwnerInput(owner.owner_key, value)
        for owner, value in zip(policy.owners, values, strict=True)
    )
    for owner, owner_input in zip(policy.owners, inputs, strict=True):
        if _type_path(owner_input.result) != owner.result_type:
            raise TMConsolidatedDevelopmentExportError(
                f"TM owner {owner.owner_key} result type drifted: {_type_path(owner_input.result)}"
            )
    return inputs


def _schema_projection(
    schema: Sequence[SchemaItem], policy: TMConsolidatedExportPolicy
) -> tuple[SchemaItem, ...]:
    tm = tuple(
        sorted(
            (item for item in schema if item.statement_type == policy.statement_type),
            key=lambda item: item.display_order,
        )
    )
    if (
        len(tm) != policy.schema_item_count
        or len({item.schema_id for item in tm}) != len(tm)
        or tuple(item.display_order for item in tm) != tuple(range(policy.schema_item_count))
        or any(not item.canonical_name for item in tm)
    ):
        raise TMConsolidatedDevelopmentExportError("TM schema denominator/identity drifted")
    payload = [
        (
            item.schema_id,
            item.display_order,
            item.canonical_name,
            item.parent_id,
            tuple(item.children),
        )
        for item in tm
    ]
    digest = _sha256(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    if digest != policy.schema_projection_sha256:
        raise TMConsolidatedDevelopmentExportError("TM schema hierarchy projection drifted")
    return tm


def _type_path(value: object) -> str:
    cls = type(value)
    return f"{cls.__module__}.{cls.__qualname__}"


def _as_tuple(value: object, *, field: str) -> tuple[Any, ...]:
    if not isinstance(value, (tuple, list)):
        raise TMConsolidatedDevelopmentExportError(f"TM mapping field {field} is not a sequence")
    return tuple(value)


def _schema_status(raw: object) -> str | None:
    if isinstance(raw, Enum):
        raw = raw.value
    if not isinstance(raw, str) or raw not in _RAW_SCHEMA_STATUS:
        raise TMConsolidatedDevelopmentExportError(f"unsupported TM schema status: {raw!r}")
    return _RAW_SCHEMA_STATUS[raw]


def _source_ids_from_disposition(disposition: object) -> tuple[str, ...]:
    for field_name in (
        "source_row_ids",
        "source_evidence_ids",
        "source_refs",
        "source_ids",
        "candidate_source_row_ids",
    ):
        if hasattr(disposition, field_name):
            raw = getattr(disposition, field_name)
            values = _as_tuple(raw, field=field_name)
            if any(not isinstance(value, str) or not value for value in values):
                raise TMConsolidatedDevelopmentExportError("invalid TM schema source provenance")
            return tuple(values)
    return ()


def _declared_count(result: object, name: str) -> int:
    raw = getattr(result, name, 0)
    if isinstance(raw, bool) or not isinstance(raw, int) or raw < 0:
        raise TMConsolidatedDevelopmentExportError(f"invalid TM mapping count: {name}")
    return raw


def _validate_declared_status_counts(result: object, counts: Counter[str | None]) -> None:
    declared = {
        "MAPPED": _declared_count(result, "mapped_schema_count"),
        "AMBIGUOUS": _declared_count(result, "ambiguous_schema_count"),
        "UNRESOLVED": _declared_count(result, "unresolved_schema_count"),
        "NOT_OBSERVED": _declared_count(result, "not_observed_schema_count"),
        "NA": _declared_count(result, "not_applicable_schema_count"),
        None: _declared_count(result, "unassessed_schema_count"),
    }
    if any(counts.get(status, 0) != count for status, count in declared.items()):
        raise TMConsolidatedDevelopmentExportError("TM owner schema status counts drifted")
    reconciled = sum(count for status, count in declared.items() if status is not None)
    for field_name in ("status_reconciled_schema_count", "assessed_schema_count"):
        if hasattr(result, field_name) and _declared_count(result, field_name) != reconciled:
            raise TMConsolidatedDevelopmentExportError(
                f"TM owner reconciled status count drifted: {field_name}"
            )


def _validate_owner_inputs(
    owner_inputs: Sequence[TMConsolidatedOwnerInput],
    policy: TMConsolidatedExportPolicy,
    schema: tuple[SchemaItem, ...],
) -> tuple[tuple[_OwnerContext, ...], tuple[_SchemaExportRow, ...]]:
    if isinstance(owner_inputs, (str, bytes)) or not isinstance(owner_inputs, Sequence):
        raise TMConsolidatedDevelopmentExportError("TM owner inputs must be an ordered sequence")
    inputs = tuple(owner_inputs)
    if any(not isinstance(item, TMConsolidatedOwnerInput) for item in inputs):
        raise TMConsolidatedDevelopmentExportError("invalid TM owner input")
    input_keys = [item.owner_key for item in inputs]
    expected_keys = [item.owner_key for item in policy.owners]
    if len(input_keys) != len(set(input_keys)):
        raise TMConsolidatedDevelopmentExportError("duplicate TM owner input")
    if set(input_keys) != set(expected_keys):
        missing = sorted(set(expected_keys) - set(input_keys))
        extra = sorted(set(input_keys) - set(expected_keys))
        raise TMConsolidatedDevelopmentExportError(
            f"TM owner inventory incomplete; missing={missing}, extra={extra}"
        )
    input_by_key = {item.owner_key: item.result for item in inputs}
    schema_by_id = {item.schema_id: item for item in schema}
    ownership: dict[int, _SchemaExportRow] = {}
    contexts: list[_OwnerContext] = []
    authority_scopes: set[str] = set()
    for owner_policy in policy.owners:
        result = input_by_key[owner_policy.owner_key]
        result_type = _type_path(result)
        if result_type != owner_policy.result_type:
            raise TMConsolidatedDevelopmentExportError(
                f"TM owner {owner_policy.owner_key} result type drifted: {result_type}"
            )
        if getattr(result, "statement_type", "TM") != "TM":
            raise TMConsolidatedDevelopmentExportError("non-TM result reached TM export")
        if getattr(result, "report_scope", policy.report_scope) != policy.report_scope:
            raise TMConsolidatedDevelopmentExportError("non-consolidated result reached TM export")
        if getattr(result, "mapping_authority_granted", False) is not True:
            raise TMConsolidatedDevelopmentExportError("TM mapping authority was not granted")
        if _declared_count(result, "schema_item_count") != policy.schema_item_count:
            raise TMConsolidatedDevelopmentExportError("TM owner schema denominator drifted")
        result_workbook_hash = getattr(result, "schema_workbook_sha256", None)
        if (
            result_workbook_hash is not None
            and result_workbook_hash != policy.schema_workbook_sha256
        ):
            raise TMConsolidatedDevelopmentExportError("TM owner schema workbook binding drifted")
        authority_scope = getattr(result, "mapping_authority_scope", None)
        if not isinstance(authority_scope, str) or not authority_scope:
            raise TMConsolidatedDevelopmentExportError("TM owner authority scope is missing")
        if authority_scope in authority_scopes:
            raise TMConsolidatedDevelopmentExportError("duplicate TM mapping authority owner")
        authority_scopes.add(authority_scope)
        dispositions = _as_tuple(
            getattr(result, "schema_dispositions", None), field="schema_dispositions"
        )
        if len(dispositions) != policy.schema_item_count:
            raise TMConsolidatedDevelopmentExportError(
                "TM owner schema disposition denominator drifted"
            )
        disposition_ids: set[int] = set()
        counts: Counter[str | None] = Counter()
        for disposition in dispositions:
            report_norm_id = getattr(disposition, "report_norm_id", None)
            if (
                isinstance(report_norm_id, bool)
                or not isinstance(report_norm_id, int)
                or report_norm_id in disposition_ids
            ):
                raise TMConsolidatedDevelopmentExportError("duplicate/invalid TM disposition ID")
            disposition_ids.add(report_norm_id)
            item = schema_by_id.get(report_norm_id)
            if item is None:
                raise TMConsolidatedDevelopmentExportError(
                    "TM disposition references unknown schema ID"
                )
            if getattr(disposition, "canonical_name", None) != item.canonical_name:
                raise TMConsolidatedDevelopmentExportError("TM disposition canonical name drifted")
            if (
                hasattr(disposition, "display_order")
                and disposition.display_order != item.display_order
            ):
                raise TMConsolidatedDevelopmentExportError("TM disposition display order drifted")
            raw_status = getattr(disposition, "status", None)
            status = _schema_status(raw_status)
            counts[status] += 1
            if status is None:
                continue
            if report_norm_id in ownership:
                raise TMConsolidatedDevelopmentExportError(
                    f"TM schema ID {report_norm_id} has duplicate owners"
                )
            reason = getattr(disposition, "reason", None)
            if not isinstance(reason, str) or not reason:
                raise TMConsolidatedDevelopmentExportError("TM disposition reason is missing")
            ownership[report_norm_id] = _SchemaExportRow(
                report_norm_id=report_norm_id,
                display_order=item.display_order,
                canonical_name=item.canonical_name,
                schema_status=status,
                raw_schema_status=str(raw_status),
                owner_key=owner_policy.owner_key,
                mapping_authority_scope=authority_scope,
                report_scope=policy.report_scope,
                source_ids=_source_ids_from_disposition(disposition),
                reason=reason,
            )
        if disposition_ids != set(schema_by_id):
            raise TMConsolidatedDevelopmentExportError("TM owner disposition identity set drifted")
        _validate_declared_status_counts(result, counts)
        result_payload = _canonical_json_bytes(result)
        contexts.append(
            _OwnerContext(
                policy=owner_policy,
                result=result,
                result_type=result_type,
                result_sha256=_sha256(result_payload),
                authority_scope=authority_scope,
            )
        )
    if set(ownership) != set(schema_by_id):
        missing = sorted(set(schema_by_id) - set(ownership))
        raise TMConsolidatedDevelopmentExportError(
            f"TM full-schema ownership is not exhaustive: {missing[:10]}"
        )
    return tuple(contexts), tuple(ownership[item.schema_id] for item in schema)


def _decimal(value: object, *, field: str) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise TMConsolidatedDevelopmentExportError(f"boolean TM numeric value: {field}")
    try:
        result = value if isinstance(value, Decimal) else Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise TMConsolidatedDevelopmentExportError(f"invalid TM numeric value: {field}") from exc
    if not result.is_finite():
        raise TMConsolidatedDevelopmentExportError(f"non-finite TM numeric value: {field}")
    return result


def _string_or_none(value: object, *, field: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if not isinstance(value, str):
        raise TMConsolidatedDevelopmentExportError(f"invalid TM text value: {field}")
    return value or None


def _page_number(page_tag: str) -> int | None:
    match = re.search(r"(?:page|pages)-(\d{4})", page_tag)
    return int(match.group(1)) if match else None


def _page_tag(context: _OwnerContext, assignment: object, source: object | None) -> str:
    for candidate in (
        getattr(assignment, "page_tag", None),
        getattr(source, "page_tag", None) if source is not None else None,
        getattr(context.result, "page_tag", None),
    ):
        if isinstance(candidate, str) and candidate:
            return candidate
    for candidate in (
        getattr(assignment, "page_number", None),
        getattr(source, "page_number", None) if source is not None else None,
        getattr(context.result, "page_number", None),
    ):
        if isinstance(candidate, int) and not isinstance(candidate, bool) and candidate > 0:
            return f"page-{candidate:04d}"
    return context.policy.owner_key


def _source_record_id(assignment: object, source: object | None) -> str:
    for field_name in ("source_row_id", "row_id", "source_id"):
        value = getattr(assignment, field_name, None)
        if isinstance(value, str) and value:
            return value
    if source is not None:
        for field_name in ("source_row_id", "row_id", "source_id"):
            value = getattr(source, field_name, None)
            if isinstance(value, str) and value:
                return value
    plural = getattr(assignment, "source_row_ids", None)
    if plural is None:
        plural = getattr(assignment, "component_row_ids", None)
    if (
        isinstance(plural, (tuple, list))
        and plural
        and all(isinstance(value, str) and value for value in plural)
    ):
        if len(plural) == 1:
            return plural[0]
        return f"aggregate:{_sha256(_canonical_json_bytes(tuple(plural)))}"
    raise TMConsolidatedDevelopmentExportError("TM observation has no source record ID")


def _component_source_ids(
    assignment: object,
    source: object | None,
    source_record_id: str,
) -> tuple[str, ...]:
    result: list[str] = []
    for candidate in (assignment, source):
        if candidate is None:
            continue
        for field_name in ("source_row_ids", "source_ids", "source_line_ids"):
            if not hasattr(candidate, field_name):
                continue
            values = _as_tuple(getattr(candidate, field_name), field=field_name)
            if any(not isinstance(value, str) or not value for value in values):
                raise TMConsolidatedDevelopmentExportError("invalid TM observation source IDs")
            result.extend(values)
        component_row_ids = getattr(candidate, "component_row_ids", None)
        if component_row_ids is not None:
            values = _as_tuple(component_row_ids, field="component_row_ids")
            if any(not isinstance(value, str) or not value for value in values):
                raise TMConsolidatedDevelopmentExportError(
                    "invalid TM derivation component row IDs"
                )
            result.extend(values)
    if not result:
        result.append(source_record_id)
    return tuple(dict.fromkeys(result))


def _first_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (str, int, float)) and not isinstance(value, bool):
        return str(value)
    return _compact_json(value)


def _raw_source_value(
    assignment: object, source: object | None, cell_index: int | None
) -> str | None:
    for field_name in (
        "reported_value",
        "raw_value",
        "raw_text",
        "raw_ocr_texts",
        "source_raw_values",
        "source_reported_values",
    ):
        if hasattr(assignment, field_name):
            return _first_text(getattr(assignment, field_name))
    if source is not None:
        for field_name in ("raw_values", "values"):
            if not hasattr(source, field_name):
                continue
            values = _as_tuple(getattr(source, field_name), field=field_name)
            if cell_index is not None and 0 <= cell_index < len(values):
                return _first_text(values[cell_index])
    return None


def _bbox_json(assignment: object, source: object | None, cell_index: int | None) -> str | None:
    for field_name in (
        "source_bbox",
        "value_bbox",
        "source_value_bboxes",
        "raw_ocr_bboxes",
    ):
        value = getattr(assignment, field_name, None)
        if value is not None and value != ():
            return _compact_json(value)
    if source is not None and hasattr(source, "visual_cell_evidence"):
        values = _as_tuple(source.visual_cell_evidence, field="visual_cell_evidence")
        if (
            cell_index is not None
            and 0 <= cell_index < len(values)
            and values[cell_index] is not None
        ):
            return _compact_json(values[cell_index])
    return None


def _period_role(assignment: object, axis_key: str, fallback: str | None) -> str:
    for field_name in ("period_role", "current_or_comparative"):
        value = getattr(assignment, field_name, None)
        if isinstance(value, str) and value:
            return value
    axis = getattr(assignment, "axis_role", None)
    if isinstance(axis, str) and axis in {"CURRENT", "COMPARATIVE"}:
        return axis
    return fallback or (axis_key if axis_key in {"CURRENT", "COMPARATIVE"} else "UNSPECIFIED")


def _axis_key(assignment: object, cell_index: int | None) -> str:
    for field_name in ("axis_id", "axis_key", "axis_role", "measure_role"):
        value = getattr(assignment, field_name, None)
        if isinstance(value, str) and value:
            return value
    return f"cell-{cell_index + 1}" if cell_index is not None else "UNSPECIFIED"


def _period(
    assignment: object,
    *,
    value_status: str,
    source: object | None,
    cell_index: int | None,
    policy: TMConsolidatedOwnerPolicy,
    axis_key: str,
) -> tuple[str | None, str | None, str | None, str]:
    start = _string_or_none(getattr(assignment, "period_start", None), field="period_start")
    end = _string_or_none(getattr(assignment, "period_end", None), field="period_end")
    period_type = _string_or_none(getattr(assignment, "period_type", None), field="period_type")
    role_fallback: str | None = None
    if source is not None and cell_index is not None:
        for singular, plural in (
            ("period_start", "period_starts"),
            ("period_end", "period_ends"),
            ("period_type", "period_types"),
        ):
            current = {"period_start": start, "period_end": end, "period_type": period_type}[
                singular
            ]
            if current is not None or not hasattr(source, plural):
                continue
            values = _as_tuple(getattr(source, plural), field=plural)
            if cell_index < len(values):
                converted = _string_or_none(values[cell_index], field=plural)
                if singular == "period_start":
                    start = converted
                elif singular == "period_end":
                    end = converted
                else:
                    period_type = converted
        if hasattr(source, "period_roles"):
            roles = _as_tuple(source.period_roles, field="period_roles")
            if cell_index < len(roles):
                role_fallback = _string_or_none(roles[cell_index], field="period_roles")
        elif cell_index < len(policy.default_period_roles):
            role_fallback = policy.default_period_roles[cell_index]
    if end is None:
        if value_status == "BLANK" and start is None and period_type is None:
            return None, None, None, _period_role(assignment, axis_key, role_fallback)
        raise TMConsolidatedDevelopmentExportError("TM observation period end is missing")
    if period_type is None:
        period_type = policy.default_period_type
    if period_type is None:
        period_type = "DURATION" if start is not None and start != end else "SNAPSHOT"
    if period_type not in _PERIOD_TYPES:
        raise TMConsolidatedDevelopmentExportError("unsupported TM observation period type")
    if start is None:
        if period_type != "SNAPSHOT":
            raise TMConsolidatedDevelopmentExportError(
                "TM duration observation has no period start"
            )
        start = end
    role = _period_role(assignment, axis_key, role_fallback)
    return start, end, period_type, role


def _unit(assignment: object, source: object | None) -> tuple[str, int]:
    raw_unit = getattr(assignment, "canonical_unit", None)
    if raw_unit is None:
        raw_unit = getattr(assignment, "unit", None)
    if raw_unit is None and source is not None:
        raw_unit = getattr(source, "unit", None)
    multiplier = getattr(assignment, "unit_multiplier", None)
    if multiplier is None and source is not None:
        multiplier = getattr(source, "unit_multiplier", None)
    if (
        not isinstance(raw_unit, str)
        or not raw_unit
        or isinstance(multiplier, bool)
        or not isinstance(multiplier, int)
        or multiplier <= 0
    ):
        raise TMConsolidatedDevelopmentExportError("TM observation unit binding is invalid")
    return raw_unit, multiplier


def _mapping_basis(assignment: object) -> str:
    for field_name in ("mapping_basis", "aggregation", "reason"):
        value = getattr(assignment, field_name, None)
        if isinstance(value, str) and value:
            return value
    return "DIRECT_PAGE_MAPPING_ASSIGNMENT"


def _derivation(
    assignment: object,
    mapping_basis: str,
    source_ids: tuple[str, ...],
) -> tuple[str | None, tuple[str, ...], tuple[int, ...]]:
    explicit = getattr(assignment, "derivation_method", None)
    if explicit is None:
        aggregation = getattr(assignment, "aggregation", None)
        if isinstance(aggregation, str) and aggregation not in {
            "DIRECT",
            "DIRECT_SOURCE_ROW",
            "NONE",
        }:
            explicit = aggregation
    marker_text = " ".join(
        str(value)
        for value in (
            explicit,
            mapping_basis,
            getattr(assignment, "value_origin", None),
        )
        if value is not None
    ).upper()
    explicitly_derived = bool(getattr(assignment, "is_derived", False))
    explicitly_imputed = bool(getattr(assignment, "is_imputed", False))
    marker_derived = any(marker in marker_text for marker in _DERIVATION_MARKERS)
    derived = explicitly_derived or marker_derived
    imputed = explicitly_imputed or "IMPUT" in marker_text
    if not derived and not imputed:
        return None, (), ()
    if (explicitly_derived or explicitly_imputed) and explicit is None and not marker_derived:
        raise TMConsolidatedDevelopmentExportError(
            "derived/imputed TM observation is not explicitly tagged with a method"
        )
    method = explicit if isinstance(explicit, str) and explicit else mapping_basis
    if not method or not source_ids:
        raise TMConsolidatedDevelopmentExportError(
            "derived/imputed TM observation lacks method or source provenance"
        )
    raw_components = getattr(assignment, "component_report_norm_ids", ())
    components = _as_tuple(raw_components, field="component_report_norm_ids")
    if any(isinstance(value, bool) or not isinstance(value, int) for value in components):
        raise TMConsolidatedDevelopmentExportError("invalid TM derivation component ReportNormIds")
    if components and len(components) != len(set(components)):
        raise TMConsolidatedDevelopmentExportError(
            "duplicate TM derivation component ReportNormIds"
        )
    return method, source_ids, tuple(components)


def _evidence_payload(assignment: object, source: object | None) -> object:
    return {
        "assignment": _jsonable(assignment),
        "source_disposition": _jsonable(source) if source is not None else None,
    }


def _make_observation(
    *,
    context: _OwnerContext,
    schema_rows: Mapping[int, _SchemaExportRow],
    assignment: object,
    source: object | None,
    cell_index: int | None,
    fallback_report_norm_id: int | None = None,
    fallback_observation: str | None = None,
    fallback_value: object = None,
) -> tuple[_ObservationRecord, _ProvenanceRecord]:
    report_norm_id = getattr(assignment, "report_norm_id", fallback_report_norm_id)
    if isinstance(report_norm_id, bool) or not isinstance(report_norm_id, int):
        raise TMConsolidatedDevelopmentExportError("TM observation ReportNormId is invalid")
    schema_row = schema_rows.get(report_norm_id)
    if schema_row is None or schema_row.owner_key != context.policy.owner_key:
        raise TMConsolidatedDevelopmentExportError("TM observation crosses its schema owner")
    if schema_row.schema_status != "MAPPED":
        raise TMConsolidatedDevelopmentExportError(
            "non-mapped TM schema status cannot carry observations"
        )
    observation = getattr(assignment, "observation", fallback_observation)
    if isinstance(observation, Enum):
        observation = observation.value
    if observation not in _OBSERVATION_STATUSES:
        raise TMConsolidatedDevelopmentExportError(
            f"unsupported TM observation status: {observation!r}"
        )
    raw_value = getattr(assignment, "value", fallback_value)
    if hasattr(assignment, "reported_value"):
        raw_value = assignment.reported_value
    reported_value = _decimal(raw_value, field="reported_value")
    if observation == "VALUE" and (reported_value is None or reported_value == 0):
        raise TMConsolidatedDevelopmentExportError("VALUE requires a non-zero numeric value")
    if observation == "ZERO" and reported_value != 0:
        raise TMConsolidatedDevelopmentExportError("ZERO requires numeric zero")
    if observation in {"DASH", "BLANK"} and reported_value is not None:
        raise TMConsolidatedDevelopmentExportError(f"{observation} cannot carry a value")
    unit, unit_multiplier = _unit(assignment, source)
    canonical_value = None if reported_value is None else reported_value * Decimal(unit_multiplier)
    explicit_canonical = getattr(assignment, "canonical_value_vnd", None)
    if (
        explicit_canonical is not None
        and _decimal(explicit_canonical, field="canonical_value") != canonical_value
    ):
        raise TMConsolidatedDevelopmentExportError("TM canonical value multiplier drifted")
    axis_key = _axis_key(assignment, cell_index)
    period_start, period_end, period_type, period_role = _period(
        assignment,
        value_status=observation,
        source=source,
        cell_index=cell_index,
        policy=context.policy,
        axis_key=axis_key,
    )
    scope = getattr(assignment, "scope", None) or getattr(
        context.result, "report_scope", "CONSOLIDATED"
    )
    if scope != "CONSOLIDATED":
        raise TMConsolidatedDevelopmentExportError("TM observation scope is not consolidated")
    source_record_id = _source_record_id(assignment, source)
    source_ids = _component_source_ids(assignment, source, source_record_id)
    mapping_basis = _mapping_basis(assignment)
    (
        derivation_method,
        derivation_source_ids,
        derivation_component_report_norm_ids,
    ) = _derivation(assignment, mapping_basis, source_ids)
    observation_origin = "DERIVED" if derivation_method is not None else "DIRECT"
    identity = {
        "owner_key": context.policy.owner_key,
        "report_norm_id": report_norm_id,
        "source_record_id": source_record_id,
        "source_ids": source_ids,
        "cell_index": cell_index,
        "axis_key": axis_key,
        "period_start": period_start,
        "period_end": period_end,
        "period_type": period_type,
        "period_role": period_role,
    }
    identity_digest = _sha256(_canonical_json_bytes(identity))
    observation_key = f"tmobs:{identity_digest}"
    provenance_key = f"tmprov:{identity_digest}"
    evidence_detail = _evidence_payload(assignment, source)
    evidence_detail_sha256 = _sha256(_canonical_json_bytes(evidence_detail))
    page_tag = _page_tag(context, assignment, source)
    observation_record = _ObservationRecord(
        observation_key=observation_key,
        provenance_key=provenance_key,
        report_norm_id=report_norm_id,
        canonical_name=schema_row.canonical_name,
        owner_key=context.policy.owner_key,
        mapping_authority_scope=context.authority_scope,
        schema_status=schema_row.schema_status,
        value_status=observation,
        reported_value=reported_value,
        canonical_value=canonical_value,
        period_start=period_start,
        period_end=period_end,
        period_type=period_type,
        period_role=period_role,
        unit=unit,
        unit_multiplier=unit_multiplier,
        scope=scope,
        axis_key=axis_key,
        source_record_id=source_record_id,
        source_ids=source_ids,
        observation_origin=observation_origin,
        derivation_method=derivation_method,
        derivation_source_ids=derivation_source_ids,
        derivation_component_report_norm_ids=derivation_component_report_norm_ids,
        mapping_basis=mapping_basis,
    )
    provenance_record = _ProvenanceRecord(
        provenance_key=provenance_key,
        observation_key=observation_key,
        owner_key=context.policy.owner_key,
        report_norm_id=report_norm_id,
        page_tag=page_tag,
        page_number=_page_number(page_tag),
        source_record_id=source_record_id,
        source_ids=source_ids,
        raw_source_value=_raw_source_value(assignment, source, cell_index),
        source_bbox_json=_bbox_json(assignment, source, cell_index),
        mapping_basis=mapping_basis,
        observation_origin=observation_origin,
        derivation_method=derivation_method,
        derivation_component_report_norm_ids=derivation_component_report_norm_ids,
        mapping_result_sha256=context.result_sha256,
        evidence_detail_sha256=evidence_detail_sha256,
        evidence_detail=evidence_detail,
    )
    return observation_record, provenance_record


def _source_index(result: object) -> dict[str, object]:
    dispositions = getattr(result, "source_dispositions", ())
    if not isinstance(dispositions, (tuple, list)):
        raise TMConsolidatedDevelopmentExportError("TM source dispositions are malformed")
    index: dict[str, object] = {}
    for source in dispositions:
        source_id = None
        for field_name in ("source_row_id", "row_id", "source_id"):
            candidate = getattr(source, field_name, None)
            if isinstance(candidate, str) and candidate:
                source_id = candidate
                break
        if source_id is None:
            continue
        if source_id in index:
            raise TMConsolidatedDevelopmentExportError("duplicate TM source disposition ID")
        index[source_id] = source
    return index


def _assignment_source(assignment: object, source_index: Mapping[str, object]) -> object | None:
    for field_name in ("source_row_id", "row_id", "source_id"):
        candidate = getattr(assignment, field_name, None)
        if isinstance(candidate, str) and candidate in source_index:
            return source_index[candidate]
    return None


def _validate_global_assignment_count(result: object, assignments: tuple[object, ...]) -> None:
    expected: int | None = None
    for field_name in ("mapped_status_assignment_count", "mapped_assignment_count"):
        if hasattr(result, field_name):
            expected = _declared_count(result, field_name)
            break
    if expected is None and hasattr(result, "mapped_value_assignment_count"):
        expected = _declared_count(result, "mapped_value_assignment_count") + _declared_count(
            result, "mapped_dash_assignment_count"
        )
    if expected is not None and expected != len(assignments):
        raise TMConsolidatedDevelopmentExportError("TM mapped assignment count drifted")


def _adapt_mapped_assignments(
    context: _OwnerContext,
    schema_rows: Mapping[int, _SchemaExportRow],
) -> tuple[list[_ObservationRecord], list[_ProvenanceRecord]]:
    assignments = _as_tuple(
        getattr(context.result, "mapped_assignments", None), field="mapped_assignments"
    )
    _validate_global_assignment_count(context.result, assignments)
    source_index = _source_index(context.result)
    observations: list[_ObservationRecord] = []
    provenance: list[_ProvenanceRecord] = []
    for assignment in assignments:
        cell_index = getattr(assignment, "cell_index", None)
        if isinstance(cell_index, bool) or (
            cell_index is not None and not isinstance(cell_index, int)
        ):
            raise TMConsolidatedDevelopmentExportError("invalid TM assignment cell index")
        pair = _make_observation(
            context=context,
            schema_rows=schema_rows,
            assignment=assignment,
            source=_assignment_source(assignment, source_index),
            cell_index=cell_index,
        )
        observations.append(pair[0])
        provenance.append(pair[1])
    if hasattr(context.result, "mapped_value_assignment_count"):
        numeric_count = sum(record.value_status in {"VALUE", "ZERO"} for record in observations)
        if numeric_count != _declared_count(context.result, "mapped_value_assignment_count"):
            raise TMConsolidatedDevelopmentExportError("TM numeric assignment count drifted")
    return observations, provenance


def _source_target_ids(source: object) -> tuple[int, ...]:
    plural = getattr(source, "report_norm_ids", ())
    if isinstance(plural, (tuple, list)) and plural:
        values = tuple(plural)
    else:
        singular = getattr(source, "report_norm_id", None)
        values = (singular,) if singular is not None else ()
    if any(isinstance(value, bool) or not isinstance(value, int) for value in values):
        raise TMConsolidatedDevelopmentExportError("invalid source-disposition ReportNormId")
    if len(values) != len(set(values)):
        raise TMConsolidatedDevelopmentExportError("duplicate source-disposition ReportNormId")
    return values


def _adapt_source_dispositions(
    context: _OwnerContext,
    schema_rows: Mapping[int, _SchemaExportRow],
) -> tuple[list[_ObservationRecord], list[_ProvenanceRecord]]:
    sources = _as_tuple(
        getattr(context.result, "source_dispositions", None), field="source_dispositions"
    )
    source_ids: set[str] = set()
    observations: list[_ObservationRecord] = []
    provenance: list[_ProvenanceRecord] = []
    for source in sources:
        row_id = getattr(source, "row_id", None)
        if not isinstance(row_id, str) or not row_id or row_id in source_ids:
            raise TMConsolidatedDevelopmentExportError("duplicate/invalid TM source row ID")
        source_ids.add(row_id)
        source_status = getattr(source, "status", None)
        if isinstance(source_status, Enum):
            source_status = source_status.value
        if source_status not in {"MAPPED_AUTOMATIC", "MAPPED_AUTOMATIC_SCOPED"}:
            continue
        targets = _source_target_ids(source)
        if not targets:
            raise TMConsolidatedDevelopmentExportError("mapped source disposition has no target")
        raw_observations = _as_tuple(
            getattr(source, "observations", None), field="source observations"
        )
        values = _as_tuple(getattr(source, "values", None), field="source values")
        if len(raw_observations) != len(values):
            raise TMConsolidatedDevelopmentExportError("TM source observation/value arity drifted")
        for target in targets:
            for cell_index, (observation, value) in enumerate(
                zip(raw_observations, values, strict=True)
            ):
                adapted = _AdaptedAssignment(
                    report_norm_id=target,
                    observation=str(observation),
                    value=_decimal(value, field="source value"),
                    cell_index=cell_index,
                    axis_key=f"cell-{cell_index + 1}",
                )
                pair = _make_observation(
                    context=context,
                    schema_rows=schema_rows,
                    assignment=adapted,
                    source=source,
                    cell_index=cell_index,
                )
                observations.append(pair[0])
                provenance.append(pair[1])
    direct_numeric_count = sum(record.value_status in {"VALUE", "ZERO"} for record in observations)
    declared_numeric_field = (
        "mapped_value_assignment_count"
        if hasattr(context.result, "mapped_value_assignment_count")
        else "mapped_value_count"
    )
    if hasattr(context.result, declared_numeric_field):
        if direct_numeric_count != _declared_count(context.result, declared_numeric_field):
            raise TMConsolidatedDevelopmentExportError(
                "TM source-disposition mapped value count drifted"
            )
    derived_assignments = getattr(context.result, "derived_assignments", ())
    if not isinstance(derived_assignments, (tuple, list)):
        raise TMConsolidatedDevelopmentExportError("TM derived assignments are malformed")
    if hasattr(context.result, "derived_assignment_count") and len(derived_assignments) != (
        _declared_count(context.result, "derived_assignment_count")
    ):
        raise TMConsolidatedDevelopmentExportError("TM derived assignment count drifted")
    for cell_index, assignment in enumerate(derived_assignments):
        pair = _make_observation(
            context=context,
            schema_rows=schema_rows,
            assignment=assignment,
            source=None,
            cell_index=cell_index,
        )
        if (
            pair[0].observation_origin != "DERIVED"
            or pair[0].derivation_method is None
            or not pair[0].derivation_source_ids
            or not pair[0].derivation_component_report_norm_ids
        ):
            raise TMConsolidatedDevelopmentExportError(
                "TM explicit derived assignment lost derivation provenance"
            )
        observations.append(pair[0])
        provenance.append(pair[1])
    return observations, provenance


def _adapt_nested_assignments(
    context: _OwnerContext,
    schema_rows: Mapping[int, _SchemaExportRow],
) -> tuple[list[_ObservationRecord], list[_ProvenanceRecord]]:
    sources = _as_tuple(
        getattr(context.result, "source_dispositions", None), field="source_dispositions"
    )
    seen_sources: set[str] = set()
    observations: list[_ObservationRecord] = []
    provenance: list[_ProvenanceRecord] = []
    for source in sources:
        row_id = getattr(source, "row_id", None)
        if not isinstance(row_id, str) or not row_id or row_id in seen_sources:
            raise TMConsolidatedDevelopmentExportError("duplicate/invalid nested TM source row")
        seen_sources.add(row_id)
        assignments = _as_tuple(
            getattr(source, "mapped_assignments", None), field="nested mapped_assignments"
        )
        for assignment in assignments:
            observation = getattr(assignment, "observation", None)
            if observation is None:
                if getattr(assignment, "value", None) is not None:
                    raise TMConsolidatedDevelopmentExportError(
                        "structural TM assignment unexpectedly carries a value"
                    )
                continue
            cell_index = getattr(assignment, "cell_index", None)
            if isinstance(cell_index, bool) or not isinstance(cell_index, int):
                raise TMConsolidatedDevelopmentExportError(
                    "nested TM observation lacks a concrete cell index"
                )
            pair = _make_observation(
                context=context,
                schema_rows=schema_rows,
                assignment=assignment,
                source=source,
                cell_index=cell_index,
            )
            observations.append(pair[0])
            provenance.append(pair[1])
    narrative_assignments = getattr(context.result, "narrative_assignments", ())
    if not isinstance(narrative_assignments, (tuple, list)):
        raise TMConsolidatedDevelopmentExportError("nested TM narrative assignments are malformed")
    for assignment in narrative_assignments:
        value_index = getattr(assignment, "value_index", None)
        if isinstance(value_index, bool) or not isinstance(value_index, int) or value_index < 0:
            raise TMConsolidatedDevelopmentExportError(
                "nested TM narrative observation lacks a concrete value index"
            )
        pair = _make_observation(
            context=context,
            schema_rows=schema_rows,
            assignment=assignment,
            source=None,
            cell_index=value_index,
        )
        observations.append(pair[0])
        provenance.append(pair[1])
    if hasattr(context.result, "mapped_value_count"):
        numeric_count = sum(record.value_status in {"VALUE", "ZERO"} for record in observations)
        if numeric_count != _declared_count(context.result, "mapped_value_count"):
            raise TMConsolidatedDevelopmentExportError("nested TM mapped value count drifted")
    return observations, provenance


def _adapt_page30(
    context: _OwnerContext,
    schema_rows: Mapping[int, _SchemaExportRow],
) -> tuple[list[_ObservationRecord], list[_ProvenanceRecord]]:
    mapped_values = _as_tuple(getattr(context.result, "mapped_values", None), field="mapped_values")
    if len(mapped_values) != _declared_count(context.result, "mapped_value_count"):
        raise TMConsolidatedDevelopmentExportError("page-30 mapped value count drifted")
    observations: list[_ObservationRecord] = []
    provenance: list[_ProvenanceRecord] = []
    for assignment in mapped_values:
        pair = _make_observation(
            context=context,
            schema_rows=schema_rows,
            assignment=assignment,
            source=None,
            cell_index=getattr(assignment, "axis_ordinal", None),
        )
        observations.append(pair[0])
        provenance.append(pair[1])
    axes: dict[int, object] = {}
    for assignment in mapped_values:
        ordinal = getattr(assignment, "axis_ordinal", None)
        if isinstance(ordinal, int):
            axes.setdefault(ordinal, assignment)
    if not axes:
        raise TMConsolidatedDevelopmentExportError("page-30 period axes are missing")
    sources = _as_tuple(
        getattr(context.result, "source_dispositions", None), field="source_dispositions"
    )
    for source in sources:
        if getattr(source, "value_presence", None) != "OBSERVED_STRUCTURAL_ITEM_WITH_BLANK_CELLS":
            continue
        targets = _as_tuple(
            getattr(source, "candidate_report_norm_ids", None),
            field="page-30 structural targets",
        )
        if len(targets) != 1 or isinstance(targets[0], bool) or not isinstance(targets[0], int):
            raise TMConsolidatedDevelopmentExportError("page-30 structural blank target drifted")
        for ordinal, axis in sorted(axes.items()):
            adapted = _AdaptedAssignment(
                report_norm_id=targets[0],
                observation="BLANK",
                value=None,
                cell_index=ordinal,
                axis_key=str(axis.axis_id),
                period_start=str(axis.period_start),
                period_end=str(axis.period_end),
                period_type=str(axis.period_type),
                period_role=str(axis.current_or_comparative),
                unit=str(axis.canonical_unit),
                unit_multiplier=axis.unit_multiplier,
                mapping_basis="VISIBLE_PAGE30_STRUCTURAL_BLANK_CELLS",
            )
            pair = _make_observation(
                context=context,
                schema_rows=schema_rows,
                assignment=adapted,
                source=source,
                cell_index=ordinal,
            )
            observations.append(pair[0])
            provenance.append(pair[1])
    return observations, provenance


def _adapt_no_observations(
    context: _OwnerContext,
    _schema_rows: Mapping[int, _SchemaExportRow],
) -> tuple[list[_ObservationRecord], list[_ProvenanceRecord]]:
    for field_name in ("mapped_values", "mapped_assignments"):
        value = getattr(context.result, field_name, ())
        if value:
            raise TMConsolidatedDevelopmentExportError(
                "no-observation TM owner unexpectedly carries assignments"
            )
    return [], []


def _resolved_adapter(context: _OwnerContext) -> str:
    adapter = context.policy.adapter
    if adapter != "auto":
        return adapter
    result = context.result
    if hasattr(result, "mapped_values"):
        return "page30"
    if hasattr(result, "mapped_assignments"):
        return "mapped_assignments"
    sources = getattr(result, "source_dispositions", ())
    if isinstance(sources, (tuple, list)) and any(
        hasattr(source, "mapped_assignments") for source in sources
    ):
        return "nested_assignments"
    if hasattr(result, "source_dispositions"):
        return "source_dispositions"
    raise TMConsolidatedDevelopmentExportError(
        f"cannot determine TM adapter for {context.policy.owner_key}"
    )


def _extract_observations(
    contexts: tuple[_OwnerContext, ...],
    schema_rows: tuple[_SchemaExportRow, ...],
) -> tuple[
    tuple[_SchemaExportRow, ...], tuple[_ObservationRecord, ...], tuple[_ProvenanceRecord, ...]
]:
    schema_by_id = {row.report_norm_id: row for row in schema_rows}
    adapters = {
        "page30": _adapt_page30,
        "mapped_assignments": _adapt_mapped_assignments,
        "source_dispositions": _adapt_source_dispositions,
        "nested_assignments": _adapt_nested_assignments,
        "no_observations": _adapt_no_observations,
    }
    observations: list[_ObservationRecord] = []
    provenance: list[_ProvenanceRecord] = []
    for context in contexts:
        adapter = _resolved_adapter(context)
        try:
            adapted_observations, adapted_provenance = adapters[adapter](context, schema_by_id)
        except TMConsolidatedDevelopmentExportError as exc:
            raise TMConsolidatedDevelopmentExportError(
                f"TM owner {context.policy.owner_key} observation adaptation failed: {exc}"
            ) from exc
        observations.extend(adapted_observations)
        provenance.extend(adapted_provenance)
    observation_keys = [item.observation_key for item in observations]
    provenance_keys = [item.provenance_key for item in provenance]
    if (
        len(observation_keys) != len(set(observation_keys))
        or len(provenance_keys) != len(set(provenance_keys))
        or {item.provenance_key for item in observations} != set(provenance_keys)
        or {item.observation_key for item in provenance} != set(observation_keys)
    ):
        raise TMConsolidatedDevelopmentExportError("duplicate or unlinked TM provenance")
    order_by_id = {row.report_norm_id: row.display_order for row in schema_rows}
    observations.sort(
        key=lambda item: (
            order_by_id[item.report_norm_id],
            item.period_end or "",
            item.period_role,
            item.axis_key,
            item.source_record_id,
            item.observation_key,
        )
    )
    provenance_by_key = {item.provenance_key: item for item in provenance}
    provenance = [provenance_by_key[item.provenance_key] for item in observations]
    observations_by_id: dict[int, list[_ObservationRecord]] = {}
    for item in observations:
        observations_by_id.setdefault(item.report_norm_id, []).append(item)
    updated_rows = []
    for row in schema_rows:
        row_observations = observations_by_id.get(row.report_norm_id, [])
        if (
            row.schema_status in {"AMBIGUOUS", "UNRESOLVED", "NOT_OBSERVED", "NA"}
            and row_observations
        ):
            raise TMConsolidatedDevelopmentExportError(
                "unresolved/not-observed/not-applicable TM row has observations"
            )
        source_ids = tuple(
            dict.fromkeys(
                (
                    *row.source_ids,
                    *(
                        source_id
                        for observation in row_observations
                        for source_id in (
                            observation.source_record_id,
                            *observation.source_ids,
                        )
                    ),
                )
            )
        )
        updated_rows.append(
            replace(row, source_ids=source_ids, observation_count=len(row_observations))
        )
    if sum(row.observation_count for row in updated_rows) != len(observations):
        raise TMConsolidatedDevelopmentExportError("TM observation denominator drifted")
    return tuple(updated_rows), tuple(observations), tuple(provenance)


def _status_text(value: object) -> str:
    if isinstance(value, Enum):
        value = value.value
    if not isinstance(value, str) or not value:
        raise TMConsolidatedDevelopmentExportError("TM source disposition status is missing")
    return value


def _source_target_report_norm_ids(source: object) -> tuple[int, ...]:
    result: list[int] = []
    for plural_name in (
        "report_norm_ids",
        "candidate_report_norm_ids",
        "mapped_report_norm_ids",
    ):
        raw = getattr(source, plural_name, ())
        if not isinstance(raw, (tuple, list)):
            raise TMConsolidatedDevelopmentExportError(
                f"TM source target field {plural_name} is malformed"
            )
        result.extend(raw)
    singular = getattr(source, "report_norm_id", None)
    if singular is not None:
        result.append(singular)
    if any(isinstance(value, bool) or not isinstance(value, int) for value in result):
        raise TMConsolidatedDevelopmentExportError("TM source target ReportNormId is invalid")
    return tuple(dict.fromkeys(result))


def _document_coverage(
    *,
    policy: TMConsolidatedExportPolicy,
    contexts: tuple[_OwnerContext, ...],
    schema_rows: tuple[_SchemaExportRow, ...],
    observations: tuple[_ObservationRecord, ...],
) -> dict[str, object]:
    """Derive document/source and universal-schema coverage from actual result objects."""

    source_status_counts: Counter[str] = Counter()
    source_category_counts: Counter[str] = Counter()
    visible_source_row_count = 0
    visible_source_value_status_slot_count = 0
    visible_value_status_slots_by_owner: dict[str, dict[str, object]] = {}
    mapped_new_source_row_count = 0
    mapped_reused_source_row_count = 0
    mapped_mixed_source_row_count = 0
    observations_by_owner = Counter(item.owner_key for item in observations)
    narrative_evidence_counts: Counter[str] = Counter()
    new_ids = set(TM_DOCUMENT_NEW_REPORT_NORM_IDS) if policy.schema_identity is not None else set()
    latest_batch_ids = (
        set(TM_LATEST_SCHEMA_BATCH_REPORT_NORM_IDS) if policy.schema_identity is not None else set()
    )

    for context in contexts:
        raw_sources = getattr(context.result, "source_dispositions", ())
        if not isinstance(raw_sources, (tuple, list)):
            raise TMConsolidatedDevelopmentExportError("TM source dispositions are malformed")
        sources = tuple(raw_sources)
        if hasattr(context.result, "source_row_count") and len(sources) != _declared_count(
            context.result, "source_row_count"
        ):
            raise TMConsolidatedDevelopmentExportError(
                f"TM visible source-row denominator drifted for {context.policy.owner_key}"
            )
        visible_source_row_count += len(sources)
        owner_mapped_count = 0
        for source in sources:
            raw_status = getattr(source, "status", None)
            status = (
                "MAPPED_NESTED_ASSIGNMENT_CONTAINER"
                if raw_status is None and hasattr(source, "mapped_assignments")
                else _status_text(raw_status)
            )
            source_status_counts[status] += 1
            if status.startswith(("MAPPED", "PARTIAL")):
                category = "MAPPED"
                owner_mapped_count += 1
                targets = set(_source_target_report_norm_ids(source))
                if targets & new_ids and targets - new_ids:
                    mapped_mixed_source_row_count += 1
                elif targets & new_ids:
                    mapped_new_source_row_count += 1
                else:
                    mapped_reused_source_row_count += 1
            elif "AMBIGUOUS" in status:
                category = "AMBIGUOUS"
            elif "QUESTION" in status or "UNRESOLVED" in status:
                category = "UNRESOLVED_OR_REVIEW"
            elif "SOURCE_ONLY" in status:
                category = "SOURCE_ONLY_VALIDATION"
            else:
                category = "OTHER_EXPLICIT_DISPOSITION"
            source_category_counts[category] += 1
        if hasattr(context.result, "mapped_source_row_count") and owner_mapped_count != (
            _declared_count(context.result, "mapped_source_row_count")
        ):
            raise TMConsolidatedDevelopmentExportError(
                f"TM mapped source-row denominator drifted for {context.policy.owner_key}"
            )
        if context.policy.owner_key != "residual":
            if policy.schema_identity is None:
                owner_visible_slots = observations_by_owner[context.policy.owner_key]
                slot_basis = "SYNTHETIC_TEST_EXPORTED_OBSERVATION_COUNT"
            elif hasattr(context.result, "financial_slot_count"):
                owner_visible_slots = _declared_count(context.result, "financial_slot_count")
                slot_basis = "MAPPER_DECLARED_FINANCIAL_SLOT_COUNT"
            elif context.policy.owner_key == "page-0030":
                owner_visible_slots = _declared_count(
                    context.result, "mapped_value_count"
                ) + _declared_count(context.result, "structural_blank_source_row_count")
                slot_basis = "PAGE30_MAPPED_VALUE_COUNT_PLUS_VISIBLE_STRUCTURAL_BLANK_ROW_COUNT"
            elif context.policy.owner_key == "page-0031":
                owner_visible_slots = _declared_count(context.result, "extracted_value_count")
                slot_basis = "PAGE31_EXTRACTED_VALUE_COUNT"
            else:
                raise TMConsolidatedDevelopmentExportError(
                    f"TM visible-slot denominator is unavailable for {context.policy.owner_key}"
                )
            visible_source_value_status_slot_count += owner_visible_slots
            visible_value_status_slots_by_owner[context.policy.owner_key] = {
                "count": owner_visible_slots,
                "basis": slot_basis,
            }
        for field_name in (
            "narrative_fact_count",
            "narrative_mapped_assignment_count",
            "narrative_value_count",
            "narrative_quantity_count",
            "narrative_record_count",
            "mapped_narrative_record_count",
        ):
            if hasattr(context.result, field_name):
                narrative_evidence_counts[field_name] += _declared_count(context.result, field_name)

    schema_status_counts = Counter(row.schema_status for row in schema_rows)
    schema_by_id = {row.report_norm_id: row for row in schema_rows}
    if not new_ids <= set(schema_by_id):
        raise TMConsolidatedDevelopmentExportError("TM document-new schema identity set drifted")
    mapped_new_schema_item_count = sum(
        schema_by_id[report_norm_id].schema_status == "MAPPED" for report_norm_id in new_ids
    )
    mapped_latest_schema_batch_item_count = sum(
        schema_by_id[report_norm_id].schema_status == "MAPPED"
        for report_norm_id in latest_batch_ids
    )
    not_observed_latest_schema_batch_item_count = sum(
        schema_by_id[report_norm_id].schema_status == "NOT_OBSERVED"
        for report_norm_id in latest_batch_ids
    )
    unresolved_latest_schema_batch_item_count = sum(
        schema_by_id[report_norm_id].schema_status == "UNRESOLVED"
        for report_norm_id in latest_batch_ids
    )
    mapped_schema_item_count = schema_status_counts["MAPPED"]
    mapped_reused_schema_item_count = mapped_schema_item_count - mapped_new_schema_item_count
    not_observed_new_schema_item_count = sum(
        schema_by_id[report_norm_id].schema_status == "NOT_OBSERVED" for report_norm_id in new_ids
    )
    unresolved_new_schema_item_count = sum(
        schema_by_id[report_norm_id].schema_status == "UNRESOLVED" for report_norm_id in new_ids
    )
    accounted_source_row_count = sum(source_category_counts.values())
    if accounted_source_row_count != visible_source_row_count:
        raise TMConsolidatedDevelopmentExportError("TM source-row coverage is not exhaustive")
    return {
        "visible_source_row_count": visible_source_row_count,
        "visible_source_cell_count": visible_source_value_status_slot_count,
        "visible_source_value_status_slot_count": visible_source_value_status_slot_count,
        "visible_value_status_slots_by_owner": visible_value_status_slots_by_owner,
        "narrative_evidence_counts": dict(sorted(narrative_evidence_counts.items())),
        "source_status_counts": dict(sorted(source_status_counts.items())),
        "source_category_counts": dict(sorted(source_category_counts.items())),
        "mapped_source_row_count": source_category_counts["MAPPED"],
        "mapped_new_source_row_count": mapped_new_source_row_count,
        "mapped_reused_source_row_count": mapped_reused_source_row_count,
        "mapped_mixed_source_row_count": mapped_mixed_source_row_count,
        "ambiguous_source_row_count": source_category_counts["AMBIGUOUS"],
        "unresolved_or_review_source_row_count": source_category_counts["UNRESOLVED_OR_REVIEW"],
        "source_only_validation_row_count": source_category_counts["SOURCE_ONLY_VALIDATION"],
        "accounted_source_row_count": accounted_source_row_count,
        "unaccounted_source_row_count": visible_source_row_count - accounted_source_row_count,
        "mapped_schema_item_count": mapped_schema_item_count,
        "mapped_new_schema_item_count": mapped_new_schema_item_count,
        "mapped_reused_schema_item_count": mapped_reused_schema_item_count,
        "not_observed_new_schema_item_count": not_observed_new_schema_item_count,
        "unresolved_new_schema_item_count": unresolved_new_schema_item_count,
        "new_schema_item_count": len(new_ids),
        "latest_schema_batch_item_count": len(latest_batch_ids),
        "mapped_latest_schema_batch_item_count": mapped_latest_schema_batch_item_count,
        "not_observed_latest_schema_batch_item_count": (
            not_observed_latest_schema_batch_item_count
        ),
        "unresolved_latest_schema_batch_item_count": (unresolved_latest_schema_batch_item_count),
        "ambiguous_schema_item_count": schema_status_counts["AMBIGUOUS"],
        "unresolved_schema_item_count": schema_status_counts["UNRESOLVED"],
        "not_observed_schema_item_count": schema_status_counts["NOT_OBSERVED"],
        "not_applicable_schema_item_count": schema_status_counts["NA"],
        "accounted_schema_item_count": sum(schema_status_counts.values()),
        "observed_output_cell_count": len(observations),
    }


def _has_formula(workbook: Any) -> bool:
    return any(
        cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def _load_template(
    template_bytes: bytes,
    schema: tuple[SchemaItem, ...],
    policy: TMConsolidatedExportPolicy,
) -> tuple[Any, tuple[tuple[tuple[object, str, int], ...], ...]]:
    if _sha256(template_bytes) != policy.schema_workbook_sha256:
        raise TMConsolidatedDevelopmentExportError("TM schema workbook SHA-256 drifted")
    try:
        workbook = load_workbook(BytesIO(template_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise TMConsolidatedDevelopmentExportError("cannot decode TM schema workbook") from exc
    sheet = workbook.active
    if (
        workbook.sheetnames != ["Sheet1"]
        or sheet.max_row != policy.schema_item_count + 1
        or sheet.max_column != 3
        or tuple(sheet.cell(1, column).value for column in range(1, 4))
        != (None, "ReportNormId", "ReportNormName")
        or _has_formula(workbook)
    ):
        workbook.close()
        raise TMConsolidatedDevelopmentExportError("TM schema workbook structure drifted")
    snapshot: list[tuple[tuple[object, str, int], ...]] = []
    for row_index in range(1, policy.schema_item_count + 2):
        cells = tuple(sheet.cell(row_index, column) for column in range(1, 4))
        snapshot.append(tuple((cell.value, cell.data_type, cell.style_id) for cell in cells))
        if row_index == 1:
            continue
        item = schema[row_index - 2]
        if (
            (cells[0].value, cells[1].value) != (item.display_order, item.schema_id)
            or not isinstance(cells[2].value, str)
            or normalize_text(cells[2].value) != item.canonical_name
        ):
            workbook.close()
            raise TMConsolidatedDevelopmentExportError("TM schema workbook/item binding drifted")
    sheet.title = "TM"
    return workbook, tuple(snapshot)


def _style_headers(sheet: Any, *, start_column: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(start_column, sheet.max_column + 1):
        cell = sheet.cell(1, column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _set_widths(sheet: Any, widths: Mapping[int, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _excel_number(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _write_tm_sheet(workbook: Any, rows: tuple[_SchemaExportRow, ...]) -> None:
    sheet = workbook["TM"]
    headers = (
        "SchemaStatus",
        "RawSchemaStatus",
        "OwnerKey",
        "MappingAuthorityScope",
        "ReportScope",
        "SourceIdsJson",
        "StatusReason",
        "ObservationCount",
    )
    for column, header in enumerate(headers, start=4):
        sheet.cell(1, column, header)
    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row_index, 4, row.schema_status)
        sheet.cell(row_index, 5, row.raw_schema_status)
        sheet.cell(row_index, 6, row.owner_key)
        sheet.cell(row_index, 7, row.mapping_authority_scope)
        sheet.cell(row_index, 8, row.report_scope)
        sheet.cell(row_index, 9, _compact_json(row.source_ids))
        sheet.cell(row_index, 10, row.reason)
        sheet.cell(row_index, 11, row.observation_count)
        sheet.cell(row_index, 10).alignment = Alignment(wrap_text=True, vertical="top")
    _style_headers(sheet, start_column=4)
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A1:K{len(rows) + 1}"
    _set_widths(
        sheet,
        {1: 12, 2: 16, 3: 70, 4: 18, 5: 32, 6: 22, 7: 52, 8: 18, 9: 52, 10: 80, 11: 18},
    )


def _write_observations(
    workbook: Any,
    observations: tuple[_ObservationRecord, ...],
) -> None:
    sheet = workbook.create_sheet("OBSERVATIONS")
    headers = (
        "ObservationKey",
        "ProvenanceKey",
        "ReportNormId",
        "CanonicalName",
        "OwnerKey",
        "MappingAuthorityScope",
        "SchemaStatus",
        "ValueStatus",
        "ReportedValue",
        "ReportedValueExact",
        "CanonicalValue",
        "CanonicalValueExact",
        "PeriodStart",
        "PeriodEnd",
        "PeriodType",
        "PeriodRole",
        "Unit",
        "UnitMultiplier",
        "Scope",
        "AxisKey",
        "SourceRecordId",
        "SourceIdsJson",
        "ObservationOrigin",
        "DerivationMethod",
        "DerivationSourceIdsJson",
        "DerivationComponentReportNormIdsJson",
        "MappingBasis",
    )
    sheet.append(headers)
    for item in observations:
        sheet.append(
            (
                item.observation_key,
                item.provenance_key,
                item.report_norm_id,
                item.canonical_name,
                item.owner_key,
                item.mapping_authority_scope,
                item.schema_status,
                item.value_status,
                _excel_number(item.reported_value),
                format(item.reported_value, "f") if item.reported_value is not None else None,
                _excel_number(item.canonical_value),
                format(item.canonical_value, "f") if item.canonical_value is not None else None,
                item.period_start,
                item.period_end,
                item.period_type,
                item.period_role,
                item.unit,
                item.unit_multiplier,
                item.scope,
                item.axis_key,
                item.source_record_id,
                _compact_json(item.source_ids),
                item.observation_origin,
                item.derivation_method,
                _compact_json(item.derivation_source_ids),
                _compact_json(item.derivation_component_report_norm_ids),
                item.mapping_basis,
            )
        )
    _style_headers(sheet)
    _set_widths(
        sheet,
        {
            1: 72,
            2: 72,
            3: 16,
            4: 55,
            5: 22,
            6: 52,
            8: 16,
            9: 22,
            10: 24,
            11: 22,
            12: 24,
            13: 16,
            14: 16,
            20: 28,
            21: 52,
            22: 60,
            23: 20,
            24: 40,
            25: 60,
            26: 52,
            27: 72,
        },
    )
    for row in sheet.iter_rows(min_row=2):
        row[8].number_format = "#,##0.############;[Red](#,##0.############);-"
        row[10].number_format = "#,##0.############;[Red](#,##0.############);-"


def _write_provenance(
    workbook: Any,
    provenance: tuple[_ProvenanceRecord, ...],
) -> None:
    sheet = workbook.create_sheet("PROVENANCE")
    headers = (
        "ProvenanceKey",
        "ObservationKey",
        "OwnerKey",
        "ReportNormId",
        "PageTag",
        "PageNumber",
        "SourceRecordId",
        "SourceIdsJson",
        "RawSourceValue",
        "SourceBboxJson",
        "MappingBasis",
        "ObservationOrigin",
        "DerivationMethod",
        "DerivationComponentReportNormIdsJson",
        "MappingResultSha256",
        "EvidenceDetailSha256",
    )
    sheet.append(headers)
    for item in provenance:
        sheet.append(
            (
                item.provenance_key,
                item.observation_key,
                item.owner_key,
                item.report_norm_id,
                item.page_tag,
                item.page_number,
                item.source_record_id,
                _compact_json(item.source_ids),
                item.raw_source_value,
                item.source_bbox_json,
                item.mapping_basis,
                item.observation_origin,
                item.derivation_method,
                _compact_json(item.derivation_component_report_norm_ids),
                item.mapping_result_sha256,
                item.evidence_detail_sha256,
            )
        )
    _style_headers(sheet)
    _set_widths(
        sheet,
        {
            1: 72,
            2: 72,
            3: 22,
            4: 16,
            5: 22,
            7: 52,
            8: 60,
            9: 55,
            10: 60,
            11: 72,
            12: 20,
            13: 40,
            14: 52,
            15: 68,
            16: 68,
        },
    )


def _safe_validation_status(status: object) -> str:
    if isinstance(status, Enum):
        status = status.value
    if not isinstance(status, str) or not status:
        raise TMConsolidatedDevelopmentExportError("TM validation status is missing")
    if not (status == "PASS" or status.startswith("PASS_") or status.startswith("NOT_TESTABLE")):
        raise TMConsolidatedDevelopmentExportError(f"unsafe TM page validation status: {status}")
    return status


def _status_bearing_records(value: object) -> tuple[object, ...]:
    if value is None:
        return ()
    candidates = tuple(value) if isinstance(value, (tuple, list)) else (value,)
    return tuple(item for item in candidates if hasattr(item, "status"))


def _declared_validation_count(result: object, field_name: str) -> int | None:
    if not hasattr(result, field_name):
        return None
    return _declared_count(result, field_name)


def _page_validation_rows(
    contexts: tuple[_OwnerContext, ...],
) -> tuple[dict[str, object], ...]:
    rows: list[dict[str, object]] = []
    for context in contexts:
        result = context.result
        records_by_family: dict[str, tuple[object, ...]] = {}
        for family in _PAGE_VALIDATION_FAMILIES:
            if not hasattr(result, family):
                continue
            raw = getattr(result, family)
            if family == "narrative_diagnostic":
                records = () if raw is None else (raw,)
            else:
                records = _as_tuple(raw, field=family)
            if any(not hasattr(record, "status") for record in records):
                raise TMConsolidatedDevelopmentExportError(
                    f"TM validation family {family} contains a statusless record"
                )
            records_by_family[family] = records
            for record in records:
                identifier = next(
                    (
                        getattr(record, field_name)
                        for field_name in _VALIDATION_ID_FIELDS
                        if isinstance(getattr(record, field_name, None), str)
                        and getattr(record, field_name)
                    ),
                    None,
                )
                if identifier is None:
                    raise TMConsolidatedDevelopmentExportError(
                        f"TM validation family {family} has no stable ID"
                    )
                status = _safe_validation_status(getattr(record, "status", None))
                rows.append(
                    {
                        "check_id": identifier,
                        "owner_key": context.policy.owner_key,
                        "validation_family": family,
                        "status": status,
                        "details": _jsonable(record),
                    }
                )

        if is_dataclass(result) and not isinstance(result, type):
            for field in fields(result):
                if field.name in _PAGE_VALIDATION_FAMILIES or not any(
                    marker in field.name for marker in ("check", "diagnostic", "validation")
                ):
                    continue
                if _status_bearing_records(getattr(result, field.name)):
                    raise TMConsolidatedDevelopmentExportError(
                        f"unsupported status-bearing TM validation family: {field.name}"
                    )
        count_bindings = {
            "accounting_checks": "accounting_check_count",
            "percentage_checks": "percentage_check_count",
            "hierarchy_checks": "hierarchy_check_count",
            "catch_all_checks": "catch_all_check_count",
            "duplicate_checks": "duplicate_check_count",
        }
        for family, count_field in count_bindings.items():
            declared = _declared_validation_count(result, count_field)
            if declared is not None and declared != len(records_by_family.get(family, ())):
                raise TMConsolidatedDevelopmentExportError(
                    f"TM page validation count drifted: {count_field}"
                )
        if hasattr(result, "validation_check_count"):
            validation_records = records_by_family.get("validation_checks")
            if validation_records is None:
                validation_records = (
                    *records_by_family.get("accounting_checks", ()),
                    *records_by_family.get("duplicate_checks", ()),
                )
            if _declared_count(result, "validation_check_count") != len(validation_records):
                raise TMConsolidatedDevelopmentExportError(
                    "TM page validation count drifted: validation_check_count"
                )

        status_count_bindings = {
            "accounting_checks": (
                "accounting_pass_count",
                "accounting_not_testable_count",
                "accounting_fail_count",
            ),
            "percentage_checks": ("percentage_pass_count", None, None),
            "hierarchy_checks": ("hierarchy_pass_count", None, None),
            "catch_all_checks": ("catch_all_pass_count", None, None),
            "duplicate_checks": ("duplicate_pass_count", None, None),
        }
        for family, field_names in status_count_bindings.items():
            family_records = records_by_family.get(family, ())
            statuses = tuple(str(record.status) for record in family_records)
            expected_counts = (
                sum(status == "PASS" or status.startswith("PASS_") for status in statuses),
                sum(status.startswith("NOT_TESTABLE") for status in statuses),
                sum(status == "FAIL" or status.startswith("FAIL_") for status in statuses),
            )
            for field_name, expected in zip(field_names, expected_counts, strict=True):
                declared = (
                    _declared_validation_count(result, field_name)
                    if field_name is not None
                    else None
                )
                if declared is not None and declared != expected:
                    raise TMConsolidatedDevelopmentExportError(
                        f"TM page validation status count drifted: {field_name}"
                    )
        if hasattr(result, "validation_pass_count"):
            validation_records = records_by_family.get("validation_checks")
            if validation_records is None:
                validation_records = (
                    *records_by_family.get("accounting_checks", ()),
                    *records_by_family.get("duplicate_checks", ()),
                )
            statuses = tuple(str(record.status) for record in validation_records)
            if _declared_count(result, "validation_pass_count") != sum(
                status == "PASS" or status.startswith("PASS_") for status in statuses
            ):
                raise TMConsolidatedDevelopmentExportError(
                    "TM page validation status count drifted: validation_pass_count"
                )
            declared_not_testable = _declared_validation_count(
                result, "validation_not_testable_count"
            )
            if declared_not_testable is not None and declared_not_testable != sum(
                status.startswith("NOT_TESTABLE") for status in statuses
            ):
                raise TMConsolidatedDevelopmentExportError(
                    "TM page validation status count drifted: validation_not_testable_count"
                )
    return tuple(rows)


def _validation_rows(
    *,
    policy: TMConsolidatedExportPolicy,
    contexts: tuple[_OwnerContext, ...],
    schema_rows: tuple[_SchemaExportRow, ...],
    observations: tuple[_ObservationRecord, ...],
    provenance: tuple[_ProvenanceRecord, ...],
    document_coverage: Mapping[str, object],
) -> tuple[dict[str, object], ...]:
    status_counts = Counter(row.schema_status for row in schema_rows)
    rows: list[dict[str, object]] = [
        {
            "check_id": "SCHEMA_FREEZE",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {
                "item_count": policy.schema_item_count,
                "projection_sha256": policy.schema_projection_sha256,
                "workbook_sha256": policy.schema_workbook_sha256,
            },
        },
        {
            "check_id": "OWNER_INVENTORY_AND_TYPE",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {
                "owner_count": len(contexts),
                "owner_keys": [context.policy.owner_key for context in contexts],
            },
        },
        {
            "check_id": "FULL_SCHEMA_DISJOINT_EXHAUSTIVE_OWNERSHIP",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {"status_counts": dict(sorted(status_counts.items()))},
        },
        {
            "check_id": "OBSERVATION_STATUS_AND_VALUE_CONSISTENCY",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {
                "observation_count": len(observations),
                "value_status_counts": dict(
                    sorted(Counter(item.value_status for item in observations).items())
                ),
            },
        },
        {
            "check_id": "PROVENANCE_ONE_TO_ONE",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {"provenance_count": len(provenance)},
        },
        {
            "check_id": "DERIVATION_AND_IMPUTATION_TAGS",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {
                "tagged_observation_count": sum(
                    item.derivation_method is not None for item in observations
                )
            },
        },
        {
            "check_id": "SOURCE_AND_SCHEMA_COVERAGE",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": dict(document_coverage),
        },
    ]
    rows.extend(
        {
            "check_id": "OWNER_RESULT_BINDING",
            "owner_key": context.policy.owner_key,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {
                "adapter": _resolved_adapter(context),
                "authority_scope": context.authority_scope,
                "result_sha256": context.result_sha256,
                "result_type": context.result_type,
            },
        }
        for context in contexts
    )
    rows.extend(_page_validation_rows(contexts))
    rows.append(
        {
            "check_id": "FORMULA_FREE",
            "owner_key": None,
            "validation_family": "EXPORTER_GUARD",
            "status": "PASS",
            "details": {"formula_count": 0},
        }
    )
    validation_keys: set[str] = set()
    for row in rows:
        validation_key = "tmval:" + _sha256(
            _canonical_json_bytes(
                {
                    "check_id": row["check_id"],
                    "details": row["details"],
                    "owner_key": row["owner_key"],
                    "status": row["status"],
                    "validation_family": row["validation_family"],
                }
            )
        )
        if validation_key in validation_keys:
            raise TMConsolidatedDevelopmentExportError("duplicate TM page validation record")
        validation_keys.add(validation_key)
        row["validation_key"] = validation_key
    return tuple(rows)


def _write_validation(workbook: Any, validations: tuple[dict[str, object], ...]) -> None:
    sheet = workbook.create_sheet("VALIDATION")
    sheet.append(
        (
            "ValidationKey",
            "CheckId",
            "OwnerKey",
            "ValidationFamily",
            "Status",
            "DetailsJson",
        )
    )
    for row in validations:
        sheet.append(
            (
                row["validation_key"],
                row["check_id"],
                row["owner_key"],
                row["validation_family"],
                row["status"],
                _compact_json(row["details"]),
            )
        )
    _style_headers(sheet)
    _set_widths(sheet, {1: 72, 2: 72, 3: 24, 4: 30, 5: 38, 6: 110})


def _write_run_metadata(workbook: Any, metadata: Mapping[str, object]) -> None:
    sheet = workbook.create_sheet("RUN_METADATA")
    sheet.append(("Key", "ValueJson"))
    for key in sorted(metadata):
        sheet.append((key, _compact_json(metadata[key])))
    _style_headers(sheet)
    _set_widths(sheet, {1: 48, 2: 110})


def _deterministic_workbook_bytes(workbook: Any) -> bytes:
    workbook.properties.creator = "bctc-ai"
    workbook.properties.lastModifiedBy = "bctc-ai"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP
    deterministic_core = tostring(workbook.properties.to_tree())
    raw = BytesIO()
    workbook.save(raw)
    target = BytesIO()
    with (
        zipfile.ZipFile(BytesIO(raw.getvalue()), "r") as archive,
        zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as output,
    ):
        names = archive.namelist()
        if len(names) != len(set(names)) or names.count(_CORE_PROPERTIES_MEMBER) != 1:
            raise TMConsolidatedDevelopmentExportError("generated TM workbook ZIP is invalid")
        generated_core = archive.read(_CORE_PROPERTIES_MEMBER)
        try:
            generated_root = fromstring(generated_core)
            deterministic_root = fromstring(deterministic_core)
        except Exception as exc:
            raise TMConsolidatedDevelopmentExportError(
                "generated TM workbook core properties are invalid"
            ) from exc
        generated_modified = generated_root.findall(_MODIFIED_PROPERTY_TAG)
        deterministic_modified = deterministic_root.findall(_MODIFIED_PROPERTY_TAG)
        if len(generated_modified) != 1 or len(deterministic_modified) != 1:
            raise TMConsolidatedDevelopmentExportError(
                "generated TM workbook core properties are invalid"
            )
        generated_modified[0].text = deterministic_modified[0].text
        if tostring(generated_root) != deterministic_core:
            raise TMConsolidatedDevelopmentExportError(
                "generated TM workbook core properties drifted"
            )
        for name in sorted(names):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            info.flag_bits = 0x800
            member = deterministic_core if name == _CORE_PROPERTIES_MEMBER else archive.read(name)
            output.writestr(info, member, compresslevel=9)
    return target.getvalue()


def _headers(sheet: Any) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def _verify_serialized_workbook(
    workbook_bytes: bytes,
    *,
    policy: TMConsolidatedExportPolicy,
    template_snapshot: tuple[tuple[tuple[object, str, int], ...], ...],
    schema_rows: tuple[_SchemaExportRow, ...],
    observations: tuple[_ObservationRecord, ...],
    provenance: tuple[_ProvenanceRecord, ...],
    validations: tuple[dict[str, object], ...],
) -> None:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    except Exception as exc:
        raise TMConsolidatedDevelopmentExportError(
            "serialized TM workbook cannot be reopened"
        ) from exc
    try:
        if tuple(workbook.sheetnames) != policy.output_sheets or _has_formula(workbook):
            raise TMConsolidatedDevelopmentExportError(
                "serialized TM workbook inventory/formula contract drifted"
            )
        main = workbook["TM"]
        actual_snapshot = tuple(
            tuple(
                (
                    main.cell(row_index, column).value,
                    main.cell(row_index, column).data_type,
                    main.cell(row_index, column).style_id,
                )
                for column in range(1, 4)
            )
            for row_index in range(1, policy.schema_item_count + 2)
        )
        if actual_snapshot != template_snapshot:
            raise TMConsolidatedDevelopmentExportError(
                "serialized TM workbook did not preserve template columns A:C"
            )
        main_headers = _headers(main)
        main_ids = [
            main.cell(row, main_headers["ReportNormId"]).value for row in range(2, main.max_row + 1)
        ]
        if len(main_ids) != policy.schema_item_count or len(main_ids) != len(set(main_ids)):
            raise TMConsolidatedDevelopmentExportError("serialized TM schema IDs drifted")
        expected_status_counts = Counter(row.schema_status for row in schema_rows)
        actual_status_counts = Counter(
            main.cell(row, main_headers["SchemaStatus"]).value for row in range(2, main.max_row + 1)
        )
        if actual_status_counts != expected_status_counts:
            raise TMConsolidatedDevelopmentExportError("serialized TM schema statuses drifted")
        observation_sheet = workbook["OBSERVATIONS"]
        observation_headers = _headers(observation_sheet)
        observation_keys = [
            observation_sheet.cell(row, observation_headers["ObservationKey"]).value
            for row in range(2, observation_sheet.max_row + 1)
        ]
        observation_provenance_keys = [
            observation_sheet.cell(row, observation_headers["ProvenanceKey"]).value
            for row in range(2, observation_sheet.max_row + 1)
        ]
        if (
            len(observation_keys) != len(observations)
            or len(observation_keys) != len(set(observation_keys))
            or len(observation_provenance_keys) != len(set(observation_provenance_keys))
        ):
            raise TMConsolidatedDevelopmentExportError(
                "serialized TM observation identities drifted"
            )
        for row in range(2, observation_sheet.max_row + 1):
            status = observation_sheet.cell(row, observation_headers["ValueStatus"]).value
            reported = observation_sheet.cell(row, observation_headers["ReportedValue"]).value
            canonical = observation_sheet.cell(row, observation_headers["CanonicalValue"]).value
            origin = observation_sheet.cell(row, observation_headers["ObservationOrigin"]).value
            derivation_method = observation_sheet.cell(
                row, observation_headers["DerivationMethod"]
            ).value
            component_ids = json.loads(
                observation_sheet.cell(
                    row,
                    observation_headers["DerivationComponentReportNormIdsJson"],
                ).value
            )
            if status == "VALUE" and (reported is None or reported == 0 or canonical is None):
                raise TMConsolidatedDevelopmentExportError("serialized VALUE drifted")
            if status == "ZERO" and (reported != 0 or canonical != 0):
                raise TMConsolidatedDevelopmentExportError("serialized ZERO drifted")
            if status in {"DASH", "BLANK"} and (reported is not None or canonical is not None):
                raise TMConsolidatedDevelopmentExportError(f"serialized {status} value drifted")
            if status not in _OBSERVATION_STATUSES:
                raise TMConsolidatedDevelopmentExportError(
                    "serialized TM observation status drifted"
                )
            if origin == "DERIVED" and not derivation_method:
                raise TMConsolidatedDevelopmentExportError(
                    "serialized derived TM observation lost its method/components"
                )
            if origin == "DIRECT" and (derivation_method is not None or component_ids):
                raise TMConsolidatedDevelopmentExportError(
                    "serialized direct TM observation carries derivation metadata"
                )
            if origin not in {"DIRECT", "DERIVED"}:
                raise TMConsolidatedDevelopmentExportError(
                    "serialized TM observation origin drifted"
                )
        provenance_sheet = workbook["PROVENANCE"]
        provenance_headers = _headers(provenance_sheet)
        provenance_keys = [
            provenance_sheet.cell(row, provenance_headers["ProvenanceKey"]).value
            for row in range(2, provenance_sheet.max_row + 1)
        ]
        if (
            len(provenance_keys) != len(provenance)
            or len(provenance_keys) != len(set(provenance_keys))
            or provenance_keys != observation_provenance_keys
        ):
            raise TMConsolidatedDevelopmentExportError("serialized TM provenance linkage drifted")
        provenance_origins = [
            provenance_sheet.cell(row, provenance_headers["ObservationOrigin"]).value
            for row in range(2, provenance_sheet.max_row + 1)
        ]
        observation_origins = [
            observation_sheet.cell(row, observation_headers["ObservationOrigin"]).value
            for row in range(2, observation_sheet.max_row + 1)
        ]
        if provenance_origins != observation_origins:
            raise TMConsolidatedDevelopmentExportError(
                "serialized TM provenance derivation origin drifted"
            )
        observation_count_by_id = Counter(item.report_norm_id for item in observations)
        for row_index, schema_row in enumerate(schema_rows, start=2):
            observed_count = main.cell(row_index, main_headers["ObservationCount"]).value
            if observed_count != observation_count_by_id[schema_row.report_norm_id]:
                raise TMConsolidatedDevelopmentExportError(
                    "serialized TM per-schema observation count drifted"
                )
            if schema_row.schema_status != "MAPPED" and observed_count != 0:
                raise TMConsolidatedDevelopmentExportError(
                    "serialized unresolved/NO/NA row has observations"
                )
        validation = workbook["VALIDATION"]
        validation_headers = _headers(validation)
        if validation.max_row - 1 != len(validations):
            raise TMConsolidatedDevelopmentExportError(
                "serialized TM validation denominator drifted"
            )
        for row_index, expected in enumerate(validations, start=2):
            actual = {
                "validation_key": validation.cell(
                    row_index, validation_headers["ValidationKey"]
                ).value,
                "check_id": validation.cell(row_index, validation_headers["CheckId"]).value,
                "owner_key": validation.cell(row_index, validation_headers["OwnerKey"]).value,
                "validation_family": validation.cell(
                    row_index, validation_headers["ValidationFamily"]
                ).value,
                "status": validation.cell(row_index, validation_headers["Status"]).value,
                "details": validation.cell(row_index, validation_headers["DetailsJson"]).value,
            }
            if actual != {
                "validation_key": expected["validation_key"],
                "check_id": expected["check_id"],
                "owner_key": expected["owner_key"],
                "validation_family": expected["validation_family"],
                "status": expected["status"],
                "details": _compact_json(expected["details"]),
            }:
                raise TMConsolidatedDevelopmentExportError(
                    "serialized TM validation record drifted"
                )
            _safe_validation_status(actual["status"])
        if any(str(item["status"]).startswith("FAIL") for item in validations):
            raise TMConsolidatedDevelopmentExportError("serialized TM validation drifted")
    finally:
        workbook.close()


def build_tm_consolidated_development_artifacts(
    *,
    template_path: Path,
    workbook_name: str,
    schema: Sequence[SchemaItem],
    owner_inputs: Sequence[TMConsolidatedOwnerInput],
    policy: TMConsolidatedExportPolicy,
    run_metadata: Mapping[str, object] | None = None,
) -> TMConsolidatedDevelopmentArtifacts:
    """Build deterministic XLSX and JSON provenance bytes without writing files."""

    if (
        not isinstance(workbook_name, str)
        or not workbook_name
        or Path(workbook_name).name != workbook_name
    ):
        raise TMConsolidatedDevelopmentExportError("invalid TM workbook artifact name")
    metadata = dict(run_metadata or {})
    if any(not isinstance(key, str) or not key for key in metadata):
        raise TMConsolidatedDevelopmentExportError("TM run metadata keys must be strings")
    _jsonable(metadata)
    try:
        template_bytes = Path(template_path).read_bytes()
    except OSError as exc:
        raise TMConsolidatedDevelopmentExportError("cannot read TM schema workbook") from exc
    projected_schema = _schema_projection(schema, policy)
    contexts, schema_rows = _validate_owner_inputs(owner_inputs, policy, projected_schema)
    schema_rows, observations, provenance = _extract_observations(contexts, schema_rows)
    status_counts = dict(sorted(Counter(row.schema_status for row in schema_rows).items()))
    document_coverage = _document_coverage(
        policy=policy,
        contexts=contexts,
        schema_rows=schema_rows,
        observations=observations,
    )
    if policy.schema_identity is not None:
        latest_batch_ids = set(TM_LATEST_SCHEMA_BATCH_REPORT_NORM_IDS)
        baseline_status_counts = dict(
            sorted(
                Counter(
                    row.schema_status
                    for row in schema_rows
                    if row.report_norm_id not in latest_batch_ids
                ).items()
            )
        )
        latest_status_counts = Counter(
            row.schema_status for row in schema_rows if row.report_norm_id in latest_batch_ids
        )
        if (
            baseline_status_counts != {"MAPPED": 890, "NA": 23, "NOT_OBSERVED": 804}
            or sum(latest_status_counts.values()) != len(latest_batch_ids)
            or set(latest_status_counts) - {"MAPPED", "NOT_OBSERVED", "UNRESOLVED"}
        ):
            raise TMConsolidatedDevelopmentExportError(
                "TM universal-schema production partition drifted"
            )
        if (
            document_coverage["new_schema_item_count"] != len(TM_DOCUMENT_NEW_REPORT_NORM_IDS)
            or document_coverage["mapped_new_schema_item_count"]
            + document_coverage["not_observed_new_schema_item_count"]
            + document_coverage["unresolved_new_schema_item_count"]
            != len(TM_DOCUMENT_NEW_REPORT_NORM_IDS)
            or document_coverage["mapped_latest_schema_batch_item_count"]
            + document_coverage["not_observed_latest_schema_batch_item_count"]
            + document_coverage["unresolved_latest_schema_batch_item_count"]
            != len(TM_LATEST_SCHEMA_BATCH_REPORT_NORM_IDS)
            or document_coverage["ambiguous_schema_item_count"] != 0
            or document_coverage["unresolved_schema_item_count"]
            != document_coverage["unresolved_latest_schema_batch_item_count"]
            or document_coverage["unaccounted_source_row_count"] != 0
            or document_coverage["visible_source_row_count"] != 553
            or document_coverage["visible_source_value_status_slot_count"] != 1_659
        ):
            raise TMConsolidatedDevelopmentExportError(
                "TM source-driven universal-schema coverage drifted"
            )
    validations = _validation_rows(
        policy=policy,
        contexts=contexts,
        schema_rows=schema_rows,
        observations=observations,
        provenance=provenance,
        document_coverage=document_coverage,
    )
    owner_bindings = tuple(
        {
            "owner_key": context.policy.owner_key,
            "adapter": _resolved_adapter(context),
            "authority_scope": context.authority_scope,
            "result_type": context.result_type,
            "result_sha256": context.result_sha256,
        }
        for context in contexts
    )
    provenance_payload = {
        "artifact_type": "MBB_TM_CONSOLIDATED_DEVELOPMENT_PROVENANCE",
        "bank": policy.bank,
        "dataset_role": policy.dataset_role,
        "format_version": 1,
        "fully_verified": False,
        "input_bindings": {
            "export_policy_sha256": policy.policy_sha256,
            "owner_results": owner_bindings,
            "schema_projection_sha256": policy.schema_projection_sha256,
            "schema_workbook_sha256": policy.schema_workbook_sha256,
        },
        "schema_identity": policy.schema_identity,
        "document_coverage": document_coverage,
        "observations": observations,
        "provenance": provenance,
        "report_scope": policy.report_scope,
        "run_metadata": metadata,
        "schema_rows": schema_rows,
        "statement_type": policy.statement_type,
        "summary": {
            "observation_count": len(observations),
            "owner_count": len(contexts),
            "provenance_count": len(provenance),
            "schema_item_count": len(schema_rows),
            "schema_status_counts": status_counts,
            "document_coverage": document_coverage,
            "validation_record_count": len(validations),
            "validation_status_counts": dict(
                sorted(Counter(str(item["status"]) for item in validations).items())
            ),
            "value_status_counts": dict(
                sorted(Counter(item.value_status for item in observations).items())
            ),
        },
        "validation": validations,
        "workbook_filename": workbook_name,
    }
    provenance_bytes = _canonical_json_bytes(provenance_payload)
    provenance_sha256 = _sha256(provenance_bytes)
    workbook, template_snapshot = _load_template(template_bytes, projected_schema, policy)
    try:
        _write_tm_sheet(workbook, schema_rows)
        _write_observations(workbook, observations)
        _write_provenance(workbook, provenance)
        _write_validation(workbook, validations)
        run_payload: dict[str, object] = {
            "artifact_type": "MBB_TM_CONSOLIDATED_DEVELOPMENT_WORKBOOK",
            "bank": policy.bank,
            "dataset_role": policy.dataset_role,
            "export_policy_sha256": policy.policy_sha256,
            "fully_verified": False,
            "observation_count": len(observations),
            "owner_count": len(contexts),
            "provenance_count": len(provenance),
            "provenance_sha256": provenance_sha256,
            "report_scope": policy.report_scope,
            "schema_item_count": len(schema_rows),
            "schema_identity": policy.schema_identity,
            "schema_projection_sha256": policy.schema_projection_sha256,
            "schema_status_counts": status_counts,
            "schema_workbook_sha256": policy.schema_workbook_sha256,
            "statement_type": policy.statement_type,
            "document_coverage": document_coverage,
            "validation_record_count": len(validations),
            "validation_status_counts": dict(
                sorted(Counter(str(item["status"]) for item in validations).items())
            ),
            "value_status_counts": dict(
                sorted(Counter(item.value_status for item in observations).items())
            ),
        }
        run_payload.update({f"run_metadata.{key}": value for key, value in metadata.items()})
        _write_run_metadata(workbook, run_payload)
        if _has_formula(workbook):
            raise TMConsolidatedDevelopmentExportError(
                "formulas are forbidden in TM development export"
            )
        workbook_bytes = _deterministic_workbook_bytes(workbook)
    finally:
        workbook.close()
    _verify_serialized_workbook(
        workbook_bytes,
        policy=policy,
        template_snapshot=template_snapshot,
        schema_rows=schema_rows,
        observations=observations,
        provenance=provenance,
        validations=validations,
    )
    return TMConsolidatedDevelopmentArtifacts(
        workbook_bytes=workbook_bytes,
        provenance_bytes=provenance_bytes,
        workbook_sha256=_sha256(workbook_bytes),
        provenance_sha256=provenance_sha256,
        schema_item_count=len(schema_rows),
        observation_count=len(observations),
        provenance_count=len(provenance),
        status_counts=status_counts,
    )


def _write_exclusive(path: Path, payload: bytes) -> tuple[int, int]:
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags, 0o644)
    except OSError as exc:
        raise TMConsolidatedDevelopmentExportError(
            f"TM output already exists or cannot be created: {path}"
        ) from exc
    try:
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        stat = path.stat(follow_symlinks=False)
        return stat.st_dev, stat.st_ino
    except Exception:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def _rollback_if_same(path: Path, identity: tuple[int, int]) -> None:
    try:
        stat = path.stat(follow_symlinks=False)
        if (stat.st_dev, stat.st_ino) == identity:
            path.unlink()
    except OSError:
        pass


def export_tm_consolidated_development(
    *,
    template_path: Path,
    workbook_path: Path,
    provenance_path: Path,
    schema: Sequence[SchemaItem],
    owner_inputs: Sequence[TMConsolidatedOwnerInput],
    policy: TMConsolidatedExportPolicy,
    run_metadata: Mapping[str, object] | None = None,
) -> TMConsolidatedDevelopmentExportResult:
    """Write paired TM development artifacts once, refusing any overwrite."""

    workbook_path = Path(workbook_path)
    provenance_path = Path(provenance_path)
    if workbook_path == provenance_path:
        raise TMConsolidatedDevelopmentExportError("TM output destinations must differ")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    provenance_path.parent.mkdir(parents=True, exist_ok=True)
    if os.path.lexists(workbook_path) or os.path.lexists(provenance_path):
        raise TMConsolidatedDevelopmentExportError(
            "TM development export refuses to overwrite outputs"
        )
    artifacts = build_tm_consolidated_development_artifacts(
        template_path=template_path,
        workbook_name=workbook_path.name,
        schema=schema,
        owner_inputs=owner_inputs,
        policy=policy,
        run_metadata=run_metadata,
    )
    workbook_identity = _write_exclusive(workbook_path, artifacts.workbook_bytes)
    try:
        _write_exclusive(provenance_path, artifacts.provenance_bytes)
    except Exception:
        _rollback_if_same(workbook_path, workbook_identity)
        raise
    return TMConsolidatedDevelopmentExportResult(
        workbook_path=workbook_path.as_posix(),
        provenance_path=provenance_path.as_posix(),
        workbook_sha256=artifacts.workbook_sha256,
        provenance_sha256=artifacts.provenance_sha256,
        workbook_size_bytes=len(artifacts.workbook_bytes),
        provenance_size_bytes=len(artifacts.provenance_bytes),
        schema_item_count=artifacts.schema_item_count,
        observation_count=artifacts.observation_count,
        provenance_count=artifacts.provenance_count,
        status_counts=artifacts.status_counts,
    )


__all__ = [
    "TM_CONSOLIDATED_POLICY_RELATIVE_PATH",
    "TM_CONSOLIDATED_SCHEMA_COUNT",
    "TM_CONSOLIDATED_SCHEMA_PROJECTION_SHA256",
    "TM_CONSOLIDATED_SHEETS",
    "TM_CONSOLIDATED_TEMPLATE_SHA256",
    "TM_BASE_SCHEMA_COUNT",
    "TM_DOCUMENT_NEW_REPORT_NORM_IDS",
    "TM_LATEST_SCHEMA_BATCH_REPORT_NORM_IDS",
    "TM_UNIVERSAL_SCHEMA_COUNT",
    "TM_UNIVERSAL_SCHEMA_NAME",
    "TM_UNIVERSAL_SCHEMA_REVISION",
    "TMConsolidatedDevelopmentArtifacts",
    "TMConsolidatedDevelopmentExportError",
    "TMConsolidatedDevelopmentExportResult",
    "TMConsolidatedExportPolicy",
    "TMConsolidatedOwnerInput",
    "TMConsolidatedOwnerPolicy",
    "audit_tm_consolidated_owner_result_contracts",
    "bind_tm_consolidated_owner_results",
    "build_tm_consolidated_development_artifacts",
    "export_tm_consolidated_development",
    "load_tm_consolidated_export_policy",
]
