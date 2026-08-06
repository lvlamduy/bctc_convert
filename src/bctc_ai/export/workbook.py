from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bctc_ai.core.atomic import atomic_write_bytes
from bctc_ai.core.contracts import EvidenceStatus, PipelineRecord, ValueStatus
from bctc_ai.schema.registry import SchemaItem, SchemaWorkbook, load_all

MAIN_SHEETS = ("CDKT", "KQKD", "LCTT", "TM")
SUPPORT_SHEETS = (
    "PROVENANCE",
    "REVIEW",
    "UNRESOLVED",
    "QUESTIONS",
    "SCHEMA_ADDITIONS",
    "RUN_METADATA",
)


@dataclass(frozen=True)
class WorkbookExportResult:
    path: str
    sha256: str
    sheet_names: tuple[str, ...]
    schema_counts: dict[str, int]
    exported_value_count: int
    review_count: int
    unresolved_count: int


def _read_jsonl(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.is_file():
        return []
    return [
        json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()
    ]


def _excel_value(value: str | None) -> str | int | float | None:
    if value is None:
        return None
    try:
        number = Decimal(value)
    except Exception:
        return value
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def _bbox_json(value: Any) -> str | None:
    if value is None:
        return None
    return json.dumps(asdict(value), ensure_ascii=False, sort_keys=True)


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _set_widths(sheet, widths: dict[int, float]) -> None:
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_main_sheet(
    workbook: Workbook,
    statement: str,
    items: list[SchemaItem],
    records: list[PipelineRecord],
) -> int:
    sheet = workbook.create_sheet(statement)
    headers = [
        None,
        "ReportNormId",
        "ReportNormName",
        "CurrentValue",
        "ComparativeValue",
        "CurrentObservation",
        "ComparativeObservation",
        "CurrentStatus",
        "ComparativeStatus",
        "CurrentValueStatus",
        "ComparativeValueStatus",
        "Unit",
        "CurrentPeriodEnd",
        "ComparativePeriodEnd",
    ]
    sheet.append(headers)
    by_id: dict[int, list[PipelineRecord]] = {}
    for record in records:
        if record.statement_type == statement and record.schema_id is not None:
            by_id.setdefault(record.schema_id, []).append(record)
    exported = 0
    accepted = {
        EvidenceStatus.AUTO_VERIFIED_HIGH,
        EvidenceStatus.AUTO_VERIFIED_MEDIUM,
    }
    exportable_value_statuses = {
        None,  # compatibility for historical records created before this contract
        ValueStatus.OBSERVED_VALUE,
        ValueStatus.OBSERVED_ZERO,
    }
    for item in sorted(items, key=lambda candidate: candidate.display_order):
        item_records = by_id.get(item.schema_id, [])
        current = next(
            (record for record in item_records if record.current_or_comparative == "CURRENT"),
            None,
        )
        comparative = next(
            (record for record in item_records if record.current_or_comparative == "COMPARATIVE"),
            None,
        )
        current_value = (
            _excel_value(current.normalized_value)
            if current is not None
            and current.status in accepted
            and current.value_status in exportable_value_statuses
            else None
        )
        comparative_value = (
            _excel_value(comparative.normalized_value)
            if comparative is not None
            and comparative.status in accepted
            and comparative.value_status in exportable_value_statuses
            else None
        )
        exported += int(current_value is not None) + int(comparative_value is not None)
        units = sorted({record.unit for record in item_records if record.unit})
        sheet.append(
            [
                item.display_order,
                item.schema_id,
                item.canonical_name,
                current_value,
                comparative_value,
                current.observation.value if current and current.observation else None,
                comparative.observation.value if comparative and comparative.observation else None,
                current.status.value if current else EvidenceStatus.NOT_OBSERVED.value,
                comparative.status.value if comparative else EvidenceStatus.NOT_OBSERVED.value,
                current.value_status.value
                if current and current.value_status
                else ValueStatus.NOT_OBSERVED.value
                if current is None
                else None,
                comparative.value_status.value
                if comparative and comparative.value_status
                else ValueStatus.NOT_OBSERVED.value
                if comparative is None
                else None,
                ";".join(units),
                current.period_end if current else None,
                comparative.period_end if comparative else None,
            ]
        )
    _style_header(sheet)
    _set_widths(
        sheet,
        {
            1: 10,
            2: 16,
            3: 72,
            4: 20,
            5: 20,
            6: 22,
            7: 22,
            8: 28,
            9: 28,
            10: 34,
            11: 34,
            12: 18,
            13: 20,
            14: 22,
        },
    )
    for row in sheet.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        row[3].number_format = "#,##0.############;[Red](#,##0.############);-"
        row[4].number_format = "#,##0.############;[Red](#,##0.############);-"
    return exported


def _write_provenance(workbook: Workbook, records: list[PipelineRecord]) -> None:
    sheet = workbook.create_sheet("PROVENANCE")
    headers = [
        "DocumentHash",
        "Page",
        "Table",
        "Row",
        "Column",
        "LabelBbox",
        "ValueBbox",
        "HeaderBbox",
        "UnitBbox",
        "SourceImageHash",
        "PrimaryOCR",
        "IndependentOCR",
        "SchemaId",
        "CanonicalName",
        "PDFLabel",
        "RawValue",
        "NormalizedValue",
        "Observation",
        "MappingReason",
        "Confidence",
        "Status",
        "ValueStatus",
        "PeriodStart",
        "PeriodEnd",
        "PeriodType",
        "CurrentOrComparative",
        "Unit",
        "Sign",
    ]
    sheet.append(headers)
    for record in records:
        provenance = record.provenance
        sheet.append(
            [
                provenance.document_hash if provenance else record.document_id,
                provenance.page if provenance else None,
                provenance.table_id if provenance else None,
                provenance.row_id if provenance else None,
                provenance.column_id if provenance else None,
                _bbox_json(provenance.label_bbox) if provenance else None,
                _bbox_json(provenance.value_bbox) if provenance else None,
                _bbox_json(provenance.header_bbox) if provenance else None,
                _bbox_json(provenance.unit_bbox) if provenance else None,
                provenance.source_image_hash if provenance else None,
                record.primary_ocr,
                record.independent_ocr,
                record.schema_id,
                record.canonical_name,
                record.pdf_label,
                record.raw_value,
                record.normalized_value,
                record.observation.value if record.observation else None,
                record.mapping_reason,
                record.confidence,
                record.status.value,
                record.value_status.value if record.value_status else None,
                record.period_start,
                record.period_end,
                record.period_type,
                record.current_or_comparative,
                record.unit,
                record.sign,
            ]
        )
    _style_header(sheet)
    _set_widths(
        sheet,
        {1: 68, 2: 10, 6: 38, 7: 38, 13: 14, 14: 55, 15: 55, 19: 60},
    )


def _write_record_queue(workbook: Workbook, name: str, records: list[PipelineRecord]) -> None:
    sheet = workbook.create_sheet(name)
    headers = [
        "DocumentId",
        "StatementType",
        "SchemaId",
        "CanonicalName",
        "PDFLabel",
        "RawValue",
        "NormalizedValue",
        "PeriodEnd",
        "Unit",
        "Sign",
        "Status",
        "ValueStatus",
        "Confidence",
        "RejectionReason",
        "CandidateList",
    ]
    sheet.append(headers)
    for record in records:
        sheet.append(
            [
                record.document_id,
                record.statement_type,
                record.schema_id,
                record.canonical_name,
                record.pdf_label,
                record.raw_value,
                record.normalized_value,
                record.period_end,
                record.unit,
                record.sign,
                record.status.value,
                record.value_status.value if record.value_status else None,
                record.confidence,
                record.rejection_reason,
                json.dumps(record.candidate_list, ensure_ascii=False, sort_keys=True),
            ]
        )
    _style_header(sheet)
    _set_widths(sheet, {1: 42, 4: 55, 5: 55, 13: 45, 14: 70})


def _write_mapping_records(workbook: Workbook, name: str, records: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(name)
    if not records:
        sheet.append(["Status"])
        sheet.append(["NO_RECORDS"])
        _style_header(sheet)
        return
    headers = list(records[0])
    sheet.append(headers)
    for record in records:
        sheet.append(
            [
                json.dumps(record.get(header), ensure_ascii=False, sort_keys=True)
                if isinstance(record.get(header), (dict, list))
                else record.get(header)
                for header in headers
            ]
        )
    _style_header(sheet)
    for index, _ in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = 24


def _write_metadata(workbook: Workbook, metadata: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("RUN_METADATA")
    sheet.append(["Key", "Value"])
    for key, value in sorted(metadata.items()):
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False, sort_keys=True)
        elif isinstance(value, (date, datetime)):
            value = value.isoformat()
        sheet.append([key, value])
    _style_header(sheet)
    _set_widths(sheet, {1: 34, 2: 100})


def export_workbook(
    project_root: Path,
    output_path: Path,
    records: list[PipelineRecord],
    *,
    run_metadata: dict[str, Any],
    questions_path: Path | None = None,
    schema_additions_path: Path | None = None,
) -> WorkbookExportResult:
    project_root = project_root.resolve()
    for record in records:
        record.validate()
    workbooks, items = load_all(project_root / "template", project_root)
    schema_counts = {workbook.statement_type: workbook.item_count for workbook in workbooks}
    output = Workbook()
    output.remove(output.active)
    exported = 0
    for statement in MAIN_SHEETS:
        exported += _write_main_sheet(
            output,
            statement,
            [item for item in items if item.statement_type == statement],
            records,
        )
    _write_provenance(output, records)
    review = [record for record in records if record.status is EvidenceStatus.REVIEW_REQUIRED]
    unresolved = [record for record in records if record.status is EvidenceStatus.UNRESOLVED]
    _write_record_queue(output, "REVIEW", review)
    _write_record_queue(output, "UNRESOLVED", unresolved)
    _write_mapping_records(output, "QUESTIONS", _read_jsonl(questions_path))
    _write_mapping_records(output, "SCHEMA_ADDITIONS", _read_jsonl(schema_additions_path))
    _write_metadata(
        output,
        {
            **run_metadata,
            "schema_source_hashes": {
                workbook.statement_type: workbook.sha256 for workbook in workbooks
            },
            "schema_counts": schema_counts,
            "exported_value_count": exported,
        },
    )
    buffer = BytesIO()
    output.save(buffer)
    digest = atomic_write_bytes(output_path, buffer.getvalue())
    return WorkbookExportResult(
        path=str(output_path.resolve()),
        sha256=digest,
        sheet_names=tuple(MAIN_SHEETS + SUPPORT_SHEETS),
        schema_counts=schema_counts,
        exported_value_count=exported,
        review_count=len(review),
        unresolved_count=len(unresolved),
    )


def verify_export(path: Path, expected_workbooks: list[SchemaWorkbook]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    expected_sheets = list(MAIN_SHEETS + SUPPORT_SHEETS)
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"unexpected sheet order: {workbook.sheetnames}")
    count_by_statement = {item.statement_type: item.item_count for item in expected_workbooks}
    for statement in MAIN_SHEETS:
        sheet = workbook[statement]
        identifiers = [sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)]
        if len(identifiers) != count_by_statement[statement]:
            raise ValueError(f"schema row count changed in {statement}")
        if len(identifiers) != len(set(identifiers)):
            raise ValueError(f"duplicate schema ID exported in {statement}")
    workbook.close()
