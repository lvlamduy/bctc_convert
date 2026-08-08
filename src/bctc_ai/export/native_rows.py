from __future__ import annotations

import hashlib
import json
import os
import re
import stat
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bctc_ai.export.canonical_xlsx import (
    CANONICAL_CORE_TIMESTAMP,
    EXCEL_CELL_TEXT_LIMIT,
    append_literal_row,
    deterministic_workbook_bytes,
    workbook_has_formula,
)
from bctc_ai.rows.native_statement import load_registered_native_statement_rows

EXPORT_POLICY_RELATIVE_PATH = Path("config/export/native-statement-rows-excel-v1.yaml")
SHEET_NAMES = ("PAGES", "ROWS", "CELLS", "HEADERS", "RUN_METADATA")

_EXPORT_POLICY = "REGISTERED_NATIVE_STATEMENT_ROWS_EXCEL_V1"
_EXPORT_CLAIM = "UNMAPPED_SOURCE_ROWS_AND_CELLS_ONLY_NO_SCHEMA_PROJECTION"
_EXPORT_STATUS = "ACCEPTED_SOURCE_FAITHFUL_NATIVE_ROWS_EXCEL_EXPORT"
_RECEIPT_TYPE = "REGISTERED_NATIVE_STATEMENT_ROWS_EXCEL_PROVENANCE_V1"
_SHA256 = re.compile(r"^[0-9a-f]{64}$")
_ROW_ORDINAL = re.compile(r":row-(?P<ordinal>[0-9]{4})$")
_EXCEL_TEXT_LIMIT = EXCEL_CELL_TEXT_LIMIT
_IMPLEMENTATION_PATHS = (
    "src/bctc_ai/export/canonical_xlsx.py",
    "src/bctc_ai/export/native_rows.py",
    "src/bctc_ai/rows/native_statement.py",
)
_STRICT_LOADER_INPUT_KINDS = (
    "REGISTERED_NATIVE_STATEMENT_ROWS_JSON",
    "SOURCE_PDF",
    "SOURCE_REGISTRY",
    "DATASET_ROLE_REGISTRY",
    "ACCEPTED_STATEMENT_DISCOVERY",
    "NATIVE_STATEMENT_ROWS_POLICY",
    "NATIVE_TEXT_QUALITY_CONFIG",
    "GEOMETRY_CONFIG",
    "PRODUCER_GIT_COMMIT",
    "THIS_EXPORT_POLICY",
    "EXPORT_IMPLEMENTATION",
)

PAGE_HEADERS = (
    "PageOrder",
    "Page",
    "StatementType",
    "Scope",
    "MappingEligible",
    "LocallyAccepted",
    "DiscoveryContractJson",
    "TextQuality",
    "CorruptionMarkersJson",
    "WidthPoints",
    "HeightPoints",
    "Rotation",
    "NativeWordCount",
    "NativeWordsSha256",
    "GeometryAuthority",
    "DataStartY",
    "DataEndY",
    "LabelRightBoundary",
    "EdgeTolerance",
    "GeometryAxesJson",
    "UnitRunIdsJson",
    "GeometryWarningsJson",
    "HeaderCount",
    "ReconstructedRowCount",
    "FinancialTableSpanRowCount",
    "OutsideFinancialTableSpanRowCount",
    "FinancialTableSpanCellCount",
    "OutsideFinancialTableSpanCellCount",
    "AllSourceCellCount",
)

ROW_HEADERS = (
    "PageOrder",
    "Page",
    "StatementType",
    "Scope",
    "MappingEligible",
    "SourceRowOrdinal",
    "SourceBucket",
    "WithinFinancialTableSpan",
    "RowId",
    "RowType",
    "SourceStatus",
    "RawLabel",
    "NormalizedLabel",
    "LabelBboxesJson",
    "RawNoteReference",
    "NoteReference",
    "NoteBboxJson",
    "Y0",
    "Y1",
    "Indentation",
    "CellCount",
    "WarningsJson",
    "ProvenanceJson",
)

CELL_HEADERS = (
    "PageOrder",
    "Page",
    "StatementType",
    "Scope",
    "MappingEligible",
    "SourceRowOrdinal",
    "SourceBucket",
    "WithinFinancialTableSpan",
    "RowId",
    "CellOrdinal",
    "AxisId",
    "RawText",
    "NormalizedText",
    "ValueText",
    "Observation",
    "SourceStatus",
    "SignEvidence",
    "ParseReason",
    "BboxJson",
    "RunId",
    "AxisDistance",
    "ProvenanceJson",
    "RawHeader",
    "HeaderBboxJson",
    "Unit",
    "UnitMultiplier",
    "UnitBboxJson",
    "PeriodStart",
    "PeriodEnd",
    "PeriodType",
    "DurationMonths",
    "CurrentOrComparative",
    "Restated",
    "HeaderConfidence",
    "HeaderEvidenceJson",
)

HEADER_HEADERS = (
    "PageOrder",
    "Page",
    "StatementType",
    "Scope",
    "MappingEligible",
    "HeaderOrdinal",
    "AxisId",
    "AxisGeometryJson",
    "RawHeader",
    "HeaderBboxJson",
    "Unit",
    "UnitMultiplier",
    "UnitBboxJson",
    "PeriodStart",
    "PeriodEnd",
    "PeriodType",
    "DurationMonths",
    "CurrentOrComparative",
    "Restated",
    "Confidence",
    "EvidenceJson",
)

METADATA_HEADERS = ("Key", "ValueJson")


class NativeRowsExcelExportError(RuntimeError):
    """Raised when native-row Excel validation or publication fails closed."""


@dataclass(frozen=True, slots=True)
class NativeRowsExcelPolicy:
    path: Path
    sha256: str
    name: str
    claim_boundary: str
    accepted_input: dict[str, object]
    allowed_roles: frozenset[str]
    forbidden_roles: frozenset[str]
    role_directories: dict[str, str]
    workbook_creator: str


@dataclass(frozen=True, slots=True)
class NativeRowsInputIdentity:
    path: str
    sha256: str
    size_bytes: int


@dataclass(frozen=True, slots=True)
class NativeRowsExcelArtifacts:
    workbook_bytes: bytes
    provenance_bytes: bytes
    workbook_sha256: str
    provenance_sha256: str
    summary: dict[str, object]


@dataclass(frozen=True, slots=True)
class NativeRowsExcelExportResult:
    workbook_path: Path
    provenance_path: Path
    workbook_sha256: str
    provenance_sha256: str
    workbook_size_bytes: int
    provenance_size_bytes: int
    summary: dict[str, object]


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _canonical_json_bytes(value: object) -> bytes:
    try:
        return (
            json.dumps(value, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise NativeRowsExcelExportError("native-row export value is not canonical JSON") from exc


def _json_text(value: object) -> str:
    try:
        rendered = json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError) as exc:
        raise NativeRowsExcelExportError("workbook JSON cell is not canonical") from exc
    if len(rendered) > _EXCEL_TEXT_LIMIT:
        raise NativeRowsExcelExportError("canonical JSON exceeds Excel's cell text limit")
    return rendered


def _number_text(value: object, label: str) -> str:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise NativeRowsExcelExportError(f"{label} must be numeric")
    rendered = _json_text(value)
    if rendered in {"NaN", "Infinity", "-Infinity"}:
        raise NativeRowsExcelExportError(f"{label} must be finite")
    return rendered


def _mapping(value: object, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise NativeRowsExcelExportError(f"{label} must be an object")
    return value


def _records(value: object, label: str) -> Sequence[Mapping[str, Any]]:
    if not isinstance(value, list) or any(not isinstance(item, Mapping) for item in value):
        raise NativeRowsExcelExportError(f"{label} must be an array of objects")
    return value


def _plain_filename(value: str, label: str) -> str:
    if not isinstance(value, str) or not value or Path(value).name != value:
        raise NativeRowsExcelExportError(f"{label} must be a plain filename")
    return value


def _project_path(project_root: Path, path: Path) -> Path:
    return path.resolve() if path.is_absolute() else (project_root / path).resolve()


def _relative(project_root: Path, path: Path, label: str) -> str:
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except ValueError as exc:
        raise NativeRowsExcelExportError(f"{label} must stay inside the project root") from exc


def _canonical_relative_posix(value: object, label: str) -> str:
    if not isinstance(value, str) or not value or "\\" in value:
        raise NativeRowsExcelExportError(f"{label} must be a canonical relative POSIX path")
    raw_parts = value.split("/")
    path = PurePosixPath(value)
    if (
        path.is_absolute()
        or re.match(r"^[A-Za-z]:", value)
        or any(part in {"", ".", ".."} for part in raw_parts)
        or path.as_posix() != value
    ):
        raise NativeRowsExcelExportError(f"{label} must be a canonical relative POSIX path")
    return value


def load_native_rows_excel_policy(path: Path, project_root: Path) -> NativeRowsExcelPolicy:
    project_root = project_root.resolve()
    path = _project_path(project_root, path)
    _relative(project_root, path, "native-row Excel policy")
    if path != (project_root / EXPORT_POLICY_RELATIVE_PATH).resolve():
        raise NativeRowsExcelExportError(
            f"native-row Excel requires canonical policy {EXPORT_POLICY_RELATIVE_PATH}"
        )
    try:
        encoded = path.read_bytes()
        payload = yaml.safe_load(encoded)
    except (OSError, yaml.YAMLError) as exc:
        raise NativeRowsExcelExportError(f"cannot load native-row Excel policy: {path}") from exc
    if not isinstance(payload, dict):
        raise NativeRowsExcelExportError("native-row Excel policy must be an object")
    if payload.get("version") != 1 or payload.get("policy") != _EXPORT_POLICY:
        raise NativeRowsExcelExportError("native-row Excel policy identity drifted")
    if payload.get("claim_boundary") != _EXPORT_CLAIM:
        raise NativeRowsExcelExportError("native-row Excel claim boundary drifted")
    accepted = payload.get("accepted_input")
    expected_accepted = {
        "format_version": "REGISTERED_NATIVE_STATEMENT_ROWS_RESULT_V1",
        "policy": "REGISTERED_NATIVE_STATEMENT_ROWS_V1",
        "claim_boundary": "UNMAPPED_SOURCE_ROWS_AND_CELLS_ONLY",
        "status": "ACCEPTED_NATIVE_STATEMENT_ROWS",
        "trusted_sha256_required": True,
    }
    if accepted != expected_accepted:
        raise NativeRowsExcelExportError("native-row Excel accepted-input contract drifted")
    allowed_roles = frozenset(payload.get("allowed_dataset_roles", ()))
    forbidden_roles = frozenset(payload.get("forbidden_dataset_roles", ()))
    expected_roles = frozenset(
        {"LOGIC_DEVELOPMENT", "CALIBRATION", "VALIDATION", "PRODUCTION_INPUT"}
    )
    if allowed_roles != expected_roles or forbidden_roles != {"UNTOUCHED_HOLDOUT"}:
        raise NativeRowsExcelExportError("native-row Excel role policy drifted")
    role_directories = payload.get("role_directories")
    expected_directories = {
        "LOGIC_DEVELOPMENT": "output/development",
        "CALIBRATION": "output/calibration",
        "VALIDATION": "output/validation",
        "PRODUCTION_INPUT": "output/production",
    }
    if role_directories != expected_directories:
        raise NativeRowsExcelExportError("native-row Excel role directories drifted")
    workbook = payload.get("workbook")
    expected_workbook = {
        "sheets": list(SHEET_NAMES),
        "creator": "bctc-ai/native-statement-rows-excel-v1",
        "deterministic_xlsx": True,
        "formulas_allowed": False,
        "schema_projection_allowed": False,
        "template_inputs_allowed": False,
        "dash_coercion_to_zero_allowed": False,
        "preserve_outside_financial_table_span_rows": True,
        "bind_cell_axis_context": True,
    }
    if workbook != expected_workbook:
        raise NativeRowsExcelExportError("native-row Excel workbook policy drifted")
    if payload.get("publication") != {
        "paired_workbook_and_provenance": True,
        "workbook_written_before_provenance_completion_marker": True,
        "exclusive_no_overwrite": True,
        "sibling_outputs_required": True,
        "absolute_project_paths_allowed": False,
        "canonical_provenance_json": True,
    }:
        raise NativeRowsExcelExportError("native-row Excel publication policy drifted")
    if payload.get("role_isolation") != {
        "runtime_input_allowlist": list(_STRICT_LOADER_INPUT_KINDS),
        "strict_source_contract_revalidated": True,
        "prior_answer_artifacts_allowed": False,
        "historical_values_allowed": False,
        "role_a_outputs_allowed": False,
        "schema_inputs_allowed": False,
        "template_inputs_allowed": False,
        "cross_role_inputs_allowed": False,
    }:
        raise NativeRowsExcelExportError("native-row Excel isolation policy drifted")
    return NativeRowsExcelPolicy(
        path=path,
        sha256=_sha256(encoded),
        name=_EXPORT_POLICY,
        claim_boundary=_EXPORT_CLAIM,
        accepted_input=dict(accepted),
        allowed_roles=allowed_roles,
        forbidden_roles=forbidden_roles,
        role_directories=dict(role_directories),
        workbook_creator=str(workbook["creator"]),
    )


def _validate_input_identity(identity: NativeRowsInputIdentity) -> None:
    if (
        not isinstance(identity.sha256, str)
        or _SHA256.fullmatch(identity.sha256) is None
        or isinstance(identity.size_bytes, bool)
        or not isinstance(identity.size_bytes, int)
        or identity.size_bytes < 1
    ):
        raise NativeRowsExcelExportError("native-row input identity is invalid")
    _canonical_relative_posix(identity.path, "native-row input identity path")


def _validate_implementation_ledger(
    ledger: Sequence[Mapping[str, object]],
) -> tuple[dict[str, object], ...]:
    if isinstance(ledger, (str, bytes, bytearray)):
        raise NativeRowsExcelExportError("export implementation ledger is invalid")
    normalized: list[dict[str, object]] = []
    for record in ledger:
        if not isinstance(record, Mapping) or set(record) != {"path", "sha256", "size_bytes"}:
            raise NativeRowsExcelExportError("export implementation identity fields are invalid")
        path = _canonical_relative_posix(record["path"], "export implementation path")
        digest = record["sha256"]
        size = record["size_bytes"]
        if (
            not isinstance(digest, str)
            or _SHA256.fullmatch(digest) is None
            or isinstance(size, bool)
            or not isinstance(size, int)
            or size < 0
        ):
            raise NativeRowsExcelExportError("export implementation identity is invalid")
        normalized.append({"path": path, "sha256": digest, "size_bytes": size})
    paths = [str(record["path"]) for record in normalized]
    if len(paths) != len(set(paths)) or paths != sorted(paths):
        raise NativeRowsExcelExportError(
            "export implementation ledger must be unique and path-sorted"
        )
    return tuple(normalized)


def _validate_payload_identity(
    payload: Mapping[str, Any], policy: NativeRowsExcelPolicy
) -> tuple[Mapping[str, Any], str]:
    for key, expected in policy.accepted_input.items():
        if key == "trusted_sha256_required":
            continue
        if payload.get(key) != expected:
            raise NativeRowsExcelExportError(f"native-row input {key} is not accepted")
    source = _mapping(payload.get("source"), "native-row source")
    role = source.get("dataset_role")
    if not isinstance(role, str) or role not in policy.allowed_roles:
        raise NativeRowsExcelExportError("native-row input dataset role is not exportable")
    authority = _mapping(payload.get("authority"), "native-row authority")
    if authority.get("schema_mapper") is not None:
        raise NativeRowsExcelExportError("native-row Excel forbids schema mapping authority")
    isolation = _mapping(payload.get("isolation"), "native-row isolation")
    for key in (
        "prior_answer_artifacts_loaded",
        "historical_values_loaded",
        "role_a_outputs_loaded",
        "schema_inputs_loaded",
        "template_inputs_loaded",
    ):
        if isolation.get(key) is not False:
            raise NativeRowsExcelExportError("native-row input isolation was weakened")
    summary = _mapping(payload.get("summary"), "native-row summary")
    if summary.get("schema_items_created") != 0 or summary.get("schema_items_mapped") != 0:
        raise NativeRowsExcelExportError("native-row input contains schema projection")
    run_id = payload.get("run_id")
    if not isinstance(run_id, str) or not run_id:
        raise NativeRowsExcelExportError("native-row input run_id is invalid")
    return source, role


def _row_ordinal(row: Mapping[str, Any]) -> int:
    row_id = row.get("row_id")
    match = _ROW_ORDINAL.search(row_id) if isinstance(row_id, str) else None
    if match is None:
        raise NativeRowsExcelExportError("native-row source row ID lacks a stable ordinal")
    return int(match.group("ordinal"))


def _ordered_page_rows(
    page: Mapping[str, Any],
) -> list[tuple[int, str, bool, Mapping[str, Any]]]:
    inside = _records(page.get("rows"), "financial-table-span rows")
    outside = _records(
        page.get("outside_financial_table_span_rows"),
        "outside-financial-table-span rows",
    )
    combined: list[tuple[int, str, bool, Mapping[str, Any]]] = []
    for bucket, expected, records in (
        ("FINANCIAL_TABLE_SPAN", True, inside),
        ("OUTSIDE_FINANCIAL_TABLE_SPAN", False, outside),
    ):
        for row in records:
            if row.get("within_financial_table_span") is not expected:
                raise NativeRowsExcelExportError("native-row span bucket/flag is inconsistent")
            combined.append((_row_ordinal(row), bucket, expected, row))
    combined.sort(key=lambda item: item[0])
    if [item[0] for item in combined] != list(range(1, len(combined) + 1)):
        raise NativeRowsExcelExportError("native-row source ordinals are not contiguous")
    row_ids = [item[3].get("row_id") for item in combined]
    if len(row_ids) != len(set(row_ids)):
        raise NativeRowsExcelExportError("native-row source IDs are duplicated")
    if len(inside) != page.get("financial_table_span_row_count"):
        raise NativeRowsExcelExportError("financial-table-span row denominator drifted")
    if len(outside) != page.get("outside_financial_table_span_row_count"):
        raise NativeRowsExcelExportError("outside-span row denominator drifted")
    if len(combined) != page.get("reconstructed_row_count"):
        raise NativeRowsExcelExportError("all-source row denominator drifted")
    return combined


def _record(headers: Sequence[str], values: Sequence[object]) -> dict[str, object]:
    if len(headers) != len(values):
        raise NativeRowsExcelExportError("internal workbook row width drifted")
    if any(isinstance(value, str) and len(value) > _EXCEL_TEXT_LIMIT for value in values):
        raise NativeRowsExcelExportError("source text exceeds Excel's cell text limit")
    return dict(zip(headers, values, strict=True))


def _project_records(
    payload: Mapping[str, Any],
) -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    dict[str, object],
]:
    pages = _records(payload.get("pages"), "native-row pages")
    page_numbers = [page.get("page") for page in pages]
    if (
        not page_numbers
        or any(isinstance(page, bool) or not isinstance(page, int) for page in page_numbers)
        or page_numbers != sorted(set(page_numbers))
    ):
        raise NativeRowsExcelExportError("native-row pages are not uniquely source-ordered")
    page_records: list[dict[str, object]] = []
    row_records: list[dict[str, object]] = []
    cell_records: list[dict[str, object]] = []
    header_records: list[dict[str, object]] = []
    row_type_counts: dict[str, Counter[str]] = {
        "FINANCIAL_TABLE_SPAN": Counter(),
        "OUTSIDE_FINANCIAL_TABLE_SPAN": Counter(),
    }
    observation_counts: dict[str, Counter[str]] = {
        "FINANCIAL_TABLE_SPAN": Counter(),
        "OUTSIDE_FINANCIAL_TABLE_SPAN": Counter(),
    }
    source_status_counts: dict[str, Counter[str]] = {
        "FINANCIAL_TABLE_SPAN": Counter(),
        "OUTSIDE_FINANCIAL_TABLE_SPAN": Counter(),
    }
    statement_scope_counts: Counter[str] = Counter()
    section_counts: Counter[str] = Counter()
    unlabeled_counts: Counter[str] = Counter()
    global_row_ids: set[str] = set()
    inside_rows = 0
    outside_rows = 0
    inside_cells = 0
    outside_cells = 0
    header_count = 0

    for page_order, page in enumerate(pages, start=1):
        page_number = page.get("page")
        statement_type = page.get("statement_type")
        scope = page.get("scope")
        if (
            isinstance(page_number, bool)
            or not isinstance(page_number, int)
            or not isinstance(statement_type, str)
            or not isinstance(scope, str)
        ):
            raise NativeRowsExcelExportError("native-row page identity is invalid")
        contract = _mapping(page.get("discovery_contract"), "page discovery contract")
        mapping_eligible = contract.get("mapping_eligible")
        locally_accepted = contract.get("locally_accepted")
        if type(mapping_eligible) is not bool or type(locally_accepted) is not bool:
            raise NativeRowsExcelExportError("page discovery eligibility is invalid")
        geometry = _mapping(page.get("geometry"), "page geometry")
        axes = _records(geometry.get("axes"), "page geometry axes")
        axes_by_id: dict[str, Mapping[str, Any]] = {}
        for axis in axes:
            axis_id = axis.get("axis_id")
            if not isinstance(axis_id, str) or not axis_id or axis_id in axes_by_id:
                raise NativeRowsExcelExportError("page geometry axis identity is invalid")
            axes_by_id[axis_id] = axis
        headers = _records(page.get("headers"), "page headers")
        headers_by_axis: dict[str, Mapping[str, Any]] = {}
        for header_order, header in enumerate(headers, start=1):
            axis_id = header.get("axis_id")
            if (
                not isinstance(axis_id, str)
                or axis_id not in axes_by_id
                or axis_id in headers_by_axis
            ):
                raise NativeRowsExcelExportError("header-to-axis binding is invalid")
            headers_by_axis[axis_id] = header
            header_records.append(
                _record(
                    HEADER_HEADERS,
                    (
                        page_order,
                        page_number,
                        statement_type,
                        scope,
                        mapping_eligible,
                        header_order,
                        axis_id,
                        _json_text(axes_by_id[axis_id]),
                        header.get("raw_header"),
                        _json_text(header.get("header_bbox")),
                        header.get("unit"),
                        header.get("unit_multiplier"),
                        _json_text(header.get("unit_bbox")),
                        header.get("period_start"),
                        header.get("period_end"),
                        header.get("period_type"),
                        header.get("duration_months"),
                        header.get("current_or_comparative"),
                        header.get("restated"),
                        _number_text(header.get("confidence"), "header confidence"),
                        _json_text(header.get("evidence")),
                    ),
                )
            )
        value_axis_ids = {
            axis_id for axis_id, axis in axes_by_id.items() if axis.get("role") == "VALUE"
        }
        if set(headers_by_axis) != value_axis_ids:
            raise NativeRowsExcelExportError(
                "headers do not bind every source value axis exactly once"
            )
        ordered_rows = _ordered_page_rows(page)
        page_inside_cells = 0
        page_outside_cells = 0
        for ordinal, bucket, within_span, row in ordered_rows:
            row_id = row.get("row_id")
            if not isinstance(row_id, str) or row_id in global_row_ids:
                raise NativeRowsExcelExportError("native-row source ID is invalid or duplicated")
            global_row_ids.add(row_id)
            cells = _records(row.get("cells"), "native-row cells")
            warnings = row.get("warnings")
            if not isinstance(warnings, list) or any(
                not isinstance(item, str) for item in warnings
            ):
                raise NativeRowsExcelExportError("native-row warnings are invalid")
            row_type = row.get("row_type")
            if not isinstance(row_type, str):
                raise NativeRowsExcelExportError("native-row type is invalid")
            if row_type == "SECTION_HEADER" and cells:
                raise NativeRowsExcelExportError("section header unexpectedly contains cells")
            unlabeled_numeric = not row.get("normalized_label") and bool(cells)
            if unlabeled_numeric and "numeric row has no attached label" not in warnings:
                raise NativeRowsExcelExportError("unlabeled numeric row lacks its explicit warning")
            row_type_counts[bucket][row_type] += 1
            section_counts[bucket] += int(row_type == "SECTION_HEADER")
            unlabeled_counts[bucket] += int(unlabeled_numeric)
            row_records.append(
                _record(
                    ROW_HEADERS,
                    (
                        page_order,
                        page_number,
                        statement_type,
                        scope,
                        mapping_eligible,
                        ordinal,
                        bucket,
                        within_span,
                        row_id,
                        row_type,
                        row.get("source_status"),
                        row.get("raw_label"),
                        row.get("normalized_label"),
                        _json_text(row.get("label_bboxes")),
                        row.get("raw_note_reference"),
                        row.get("note_reference"),
                        _json_text(row.get("note_bbox")),
                        _number_text(row.get("y0"), "row y0"),
                        _number_text(row.get("y1"), "row y1"),
                        _number_text(row.get("indentation"), "row indentation"),
                        len(cells),
                        _json_text(warnings),
                        _json_text(row.get("provenance")),
                    ),
                )
            )
            for cell_order, cell in enumerate(cells, start=1):
                axis_id = cell.get("axis_id")
                if not isinstance(axis_id, str) or axis_id not in headers_by_axis:
                    raise NativeRowsExcelExportError(
                        "source cell does not have exactly one header binding"
                    )
                header = headers_by_axis[axis_id]
                observation = cell.get("observation")
                source_status = cell.get("source_status")
                if not isinstance(observation, str) or not isinstance(source_status, str):
                    raise NativeRowsExcelExportError("source cell status is invalid")
                if observation == "DASH" and cell.get("value") is not None:
                    raise NativeRowsExcelExportError("DASH source cell cannot carry a value")
                if cell.get("value") is not None and not isinstance(cell.get("value"), str):
                    raise NativeRowsExcelExportError("source numeric value must remain text")
                observation_counts[bucket][observation] += 1
                source_status_counts[bucket][source_status] += 1
                cell_records.append(
                    _record(
                        CELL_HEADERS,
                        (
                            page_order,
                            page_number,
                            statement_type,
                            scope,
                            mapping_eligible,
                            ordinal,
                            bucket,
                            within_span,
                            row_id,
                            cell_order,
                            axis_id,
                            cell.get("raw_text"),
                            cell.get("normalized_text"),
                            cell.get("value"),
                            observation,
                            source_status,
                            cell.get("sign_evidence"),
                            cell.get("parse_reason"),
                            _json_text(cell.get("bbox")),
                            cell.get("run_id"),
                            _number_text(cell.get("axis_distance"), "cell axis distance"),
                            _json_text(cell.get("provenance")),
                            header.get("raw_header"),
                            _json_text(header.get("header_bbox")),
                            header.get("unit"),
                            header.get("unit_multiplier"),
                            _json_text(header.get("unit_bbox")),
                            header.get("period_start"),
                            header.get("period_end"),
                            header.get("period_type"),
                            header.get("duration_months"),
                            header.get("current_or_comparative"),
                            header.get("restated"),
                            _number_text(header.get("confidence"), "header confidence"),
                            _json_text(header.get("evidence")),
                        ),
                    )
                )
            if within_span:
                inside_rows += 1
                inside_cells += len(cells)
                page_inside_cells += len(cells)
            else:
                outside_rows += 1
                outside_cells += len(cells)
                page_outside_cells += len(cells)
        header_count += len(headers)
        statement_scope_counts[f"{statement_type}:{scope}"] += 1
        page_records.append(
            _record(
                PAGE_HEADERS,
                (
                    page_order,
                    page_number,
                    statement_type,
                    scope,
                    mapping_eligible,
                    locally_accepted,
                    _json_text(contract),
                    page.get("text_quality"),
                    _json_text(page.get("corruption_markers")),
                    _number_text(page.get("width_points"), "page width"),
                    _number_text(page.get("height_points"), "page height"),
                    page.get("rotation"),
                    page.get("native_word_count"),
                    page.get("native_words_sha256"),
                    geometry.get("authority"),
                    _number_text(geometry.get("data_start_y"), "geometry data start"),
                    _number_text(geometry.get("data_end_y"), "geometry data end"),
                    _number_text(geometry.get("label_right_boundary"), "geometry label boundary"),
                    _number_text(geometry.get("edge_tolerance"), "geometry edge tolerance"),
                    _json_text(axes),
                    _json_text(geometry.get("unit_run_ids")),
                    _json_text(geometry.get("warnings")),
                    len(headers),
                    len(ordered_rows),
                    page.get("financial_table_span_row_count"),
                    page.get("outside_financial_table_span_row_count"),
                    page_inside_cells,
                    page_outside_cells,
                    page_inside_cells + page_outside_cells,
                ),
            )
        )

    source_summary = _mapping(payload.get("summary"), "native-row source summary")
    if (
        source_summary.get("page_count") != len(page_records)
        or source_summary.get("financial_table_span_row_count") != inside_rows
        or source_summary.get("cell_count") != inside_cells
        or source_summary.get("section_header_count") != section_counts["FINANCIAL_TABLE_SPAN"]
        or source_summary.get("unlabeled_numeric_row_count")
        != unlabeled_counts["FINANCIAL_TABLE_SPAN"]
        or source_summary.get("cell_source_status_counts")
        != dict(sorted(source_status_counts["FINANCIAL_TABLE_SPAN"].items()))
    ):
        raise NativeRowsExcelExportError("native-row in-span source summary drifted")

    summary: dict[str, object] = {
        "page_count": len(page_records),
        "reconstructed_row_count": inside_rows + outside_rows,
        "financial_table_span_row_count": inside_rows,
        "outside_financial_table_span_row_count": outside_rows,
        "all_source_row_count": inside_rows + outside_rows,
        "financial_table_span_cell_count": inside_cells,
        "outside_financial_table_span_cell_count": outside_cells,
        "all_source_cell_count": inside_cells + outside_cells,
        "header_binding_count": header_count,
        "financial_table_span_section_header_count": section_counts["FINANCIAL_TABLE_SPAN"],
        "all_source_section_header_count": sum(section_counts.values()),
        "financial_table_span_unlabeled_numeric_row_count": unlabeled_counts[
            "FINANCIAL_TABLE_SPAN"
        ],
        "all_source_unlabeled_numeric_row_count": sum(unlabeled_counts.values()),
        "row_type_counts_by_bucket": {
            key: dict(sorted(value.items())) for key, value in row_type_counts.items()
        },
        "observation_counts_by_bucket": {
            key: dict(sorted(value.items())) for key, value in observation_counts.items()
        },
        "source_status_counts_by_bucket": {
            key: dict(sorted(value.items())) for key, value in source_status_counts.items()
        },
        "statement_scope_page_counts": dict(sorted(statement_scope_counts.items())),
        "exact_row_denominator_reconciled": True,
        "exact_cell_denominator_reconciled": True,
    }
    return page_records, row_records, cell_records, header_records, summary


def _style_sheet(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        header = str(sheet.cell(1, column).value)
        width = 18
        if header in {
            "RawLabel",
            "NormalizedLabel",
            "RawText",
            "NormalizedText",
            "RawHeader",
            "WarningsJson",
            "ProvenanceJson",
            "ValueJson",
        }:
            width = 42
        elif header.endswith("Json"):
            width = 34
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_records(
    workbook: Workbook,
    sheet_name: str,
    headers: Sequence[str],
    records: Sequence[Mapping[str, object]],
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    append_literal_row(sheet, headers)
    for record in records:
        append_literal_row(sheet, [record[header] for header in headers])
    _style_sheet(sheet)


def _metadata_records(
    payload: Mapping[str, Any],
    *,
    identity: NativeRowsInputIdentity,
    summary: Mapping[str, object],
) -> list[dict[str, object]]:
    values: dict[str, object] = {
        "export.artifact_type": "REGISTERED_NATIVE_STATEMENT_ROWS_EXCEL_V1",
        "export.claim_boundary": _EXPORT_CLAIM,
        "export.format_version": 1,
        "export.input_filename": Path(identity.path).name,
        "export.input_sha256": identity.sha256,
        "export.input_size_bytes": identity.size_bytes,
        "export.no_schema_mapping": True,
        "export.sheet_names": list(SHEET_NAMES),
        "export.summary": dict(summary),
    }
    for key in (
        "format_version",
        "policy",
        "claim_boundary",
        "status",
        "run_id",
        "source",
        "statement_discovery",
        "code",
        "authority",
        "isolation",
        "inputs",
        "selection",
        "summary",
    ):
        values[f"input.{key}"] = payload.get(key)
    return [_record(METADATA_HEADERS, (key, _json_text(values[key]))) for key in sorted(values)]


def _verify_workbook(
    payload: bytes,
    *,
    creator: str,
    expected_records: Mapping[str, Sequence[Mapping[str, object]]],
) -> None:
    try:
        workbook = load_workbook(BytesIO(payload), read_only=False, data_only=False)
    except Exception as exc:
        raise NativeRowsExcelExportError(
            "serialized native-row workbook cannot be reopened"
        ) from exc
    try:
        if tuple(workbook.sheetnames) != SHEET_NAMES or workbook_has_formula(workbook):
            raise NativeRowsExcelExportError("native-row workbook sheet/formula contract drifted")
        expected_headers = {
            "PAGES": PAGE_HEADERS,
            "ROWS": ROW_HEADERS,
            "CELLS": CELL_HEADERS,
            "HEADERS": HEADER_HEADERS,
            "RUN_METADATA": METADATA_HEADERS,
        }
        for name, headers in expected_headers.items():
            sheet = workbook[name]
            observed = tuple(sheet.cell(1, column).value for column in range(1, len(headers) + 1))
            if observed != tuple(headers) or sheet.max_column != len(headers):
                raise NativeRowsExcelExportError(f"native-row workbook {name} headers drifted")
            records = expected_records[name]
            if sheet.max_row != len(records) + 1:
                raise NativeRowsExcelExportError(f"native-row workbook {name} count drifted")
            for row_index, record in enumerate(records, start=2):
                for column, header in enumerate(headers, start=1):
                    expected = record[header]
                    # XLSX has no durable representation for an empty string;
                    # the row contract and explicit warnings distinguish it
                    # from nullable source fields.
                    roundtrip_expected = None if expected == "" else expected
                    cell = sheet.cell(row_index, column)
                    numeric_expected = type(roundtrip_expected) in {int, float}
                    numeric_actual = type(cell.value) in {int, float}
                    if numeric_expected and numeric_actual:
                        equal = Decimal(str(cell.value)) == Decimal(str(roundtrip_expected))
                        type_matches = cell.data_type == "n"
                    else:
                        equal = cell.value == roundtrip_expected
                        type_matches = roundtrip_expected is None or type(cell.value) is type(
                            roundtrip_expected
                        )
                    if not equal or not type_matches:
                        raise NativeRowsExcelExportError(
                            f"native-row workbook {name}!{cell.coordinate} cell roundtrip drifted"
                        )
                    if isinstance(roundtrip_expected, str) and cell.data_type != "s":
                        raise NativeRowsExcelExportError(
                            f"native-row workbook {name} source text is not literal"
                        )
        properties = workbook.properties
        if (
            properties.creator != creator
            or properties.lastModifiedBy != creator
            or properties.created != CANONICAL_CORE_TIMESTAMP
            or properties.modified != CANONICAL_CORE_TIMESTAMP
            or properties.version != "1"
            or properties.revision != "1"
        ):
            raise NativeRowsExcelExportError("native-row workbook core properties drifted")
    finally:
        workbook.close()


def build_native_rows_excel_artifacts(
    payload: Mapping[str, Any],
    *,
    input_identity: NativeRowsInputIdentity,
    workbook_filename: str,
    policy: NativeRowsExcelPolicy,
    implementation_ledger: Sequence[Mapping[str, object]],
) -> NativeRowsExcelArtifacts:
    """Build deterministic workbook/receipt bytes without publishing either file."""

    _validate_input_identity(input_identity)
    canonical_input = _canonical_json_bytes(payload)
    if (
        len(canonical_input) != input_identity.size_bytes
        or _sha256(canonical_input) != input_identity.sha256
    ):
        raise NativeRowsExcelExportError(
            "native-row payload does not match its bound input identity"
        )
    implementation_ledger = _validate_implementation_ledger(implementation_ledger)
    workbook_filename = _plain_filename(workbook_filename, "workbook filename")
    if not workbook_filename.casefold().endswith(".xlsx"):
        raise NativeRowsExcelExportError("native-row workbook filename must end in .xlsx")
    source, role = _validate_payload_identity(payload, policy)
    page_records, row_records, cell_records, header_records, summary = _project_records(payload)
    metadata_records = _metadata_records(
        payload,
        identity=input_identity,
        summary=summary,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    try:
        _write_records(workbook, "PAGES", PAGE_HEADERS, page_records)
        _write_records(workbook, "ROWS", ROW_HEADERS, row_records)
        _write_records(workbook, "CELLS", CELL_HEADERS, cell_records)
        _write_records(workbook, "HEADERS", HEADER_HEADERS, header_records)
        _write_records(workbook, "RUN_METADATA", METADATA_HEADERS, metadata_records)
        if workbook_has_formula(workbook):
            raise NativeRowsExcelExportError("formulas are forbidden in native-row workbook")
        workbook_bytes = deterministic_workbook_bytes(
            workbook,
            creator=policy.workbook_creator,
        )
    finally:
        workbook.close()
    _verify_workbook(
        workbook_bytes,
        creator=policy.workbook_creator,
        expected_records={
            "PAGES": page_records,
            "ROWS": row_records,
            "CELLS": cell_records,
            "HEADERS": header_records,
            "RUN_METADATA": metadata_records,
        },
    )
    workbook_sha256 = _sha256(workbook_bytes)
    receipt = {
        "artifact_type": _RECEIPT_TYPE,
        "claim_boundary": policy.claim_boundary,
        "code": {
            "exporter_implementation": [dict(record) for record in implementation_ledger],
            "source_capture": payload.get("code"),
        },
        "dataset_role": role,
        "format_version": 1,
        "input": {
            "claim_boundary": payload.get("claim_boundary"),
            "format_version": payload.get("format_version"),
            "path": input_identity.path,
            "policy": payload.get("policy"),
            "sha256": input_identity.sha256,
            "size_bytes": input_identity.size_bytes,
            "status": payload.get("status"),
        },
        "isolation": {
            "cross_role_inputs_loaded": False,
            "historical_values_loaded": False,
            "prior_answer_artifacts_loaded": False,
            "role_a_outputs_loaded": False,
            "runtime_input_allowlist": list(_STRICT_LOADER_INPUT_KINDS),
            "schema_inputs_loaded": False,
            "source_capture": payload.get("isolation"),
            "strict_source_contract_revalidated": True,
            "template_inputs_loaded": False,
            "trusted_input_sha256_required": True,
        },
        "policy": {
            "dash_coercion_to_zero": False,
            "export_policy": policy.name,
            "export_policy_sha256": policy.sha256,
            "input_policy": payload.get("policy"),
            "schema_projection": False,
        },
        "run_id": payload.get("run_id"),
        "run_metadata": {
            "source": source,
            "statement_discovery": payload.get("statement_discovery"),
        },
        "status": _EXPORT_STATUS,
        "summary": summary,
        "workbook": {
            "filename": workbook_filename,
            "formula_count": 0,
            "sha256": workbook_sha256,
            "sheet_names": list(SHEET_NAMES),
            "size_bytes": len(workbook_bytes),
        },
    }
    provenance_bytes = _canonical_json_bytes(receipt)
    return NativeRowsExcelArtifacts(
        workbook_bytes=workbook_bytes,
        provenance_bytes=provenance_bytes,
        workbook_sha256=workbook_sha256,
        provenance_sha256=_sha256(provenance_bytes),
        summary=summary,
    )


def _implementation_ledger(project_root: Path) -> tuple[dict[str, object], ...]:
    records: list[dict[str, object]] = []
    for relative in _IMPLEMENTATION_PATHS:
        path = (project_root / relative).resolve()
        if not path.is_file():
            raise NativeRowsExcelExportError(f"export implementation is absent: {relative}")
        encoded = path.read_bytes()
        records.append(
            {
                "path": relative,
                "sha256": _sha256(encoded),
                "size_bytes": len(encoded),
            }
        )
    return tuple(records)


def _same_regular_inode(left: os.stat_result, right: os.stat_result) -> bool:
    return (
        stat.S_ISREG(left.st_mode)
        and stat.S_ISREG(right.st_mode)
        and (left.st_dev, left.st_ino) == (right.st_dev, right.st_ino)
    )


def _fsync_directory(directory: Path) -> None:
    descriptor = os.open(directory, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _rollback_if_same(path: Path, identity: os.stat_result) -> None:
    try:
        current = path.stat(follow_symlinks=False)
    except FileNotFoundError:
        return
    if not _same_regular_inode(identity, current):
        raise NativeRowsExcelExportError(f"refusing unsafe rollback of changed output: {path}")
    path.unlink()
    _fsync_directory(path.parent)


def _write_exclusive(path: Path, payload: bytes) -> os.stat_result:
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_CLOEXEC", 0)
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags, 0o644)
    except OSError as exc:
        raise NativeRowsExcelExportError(
            f"native-row output already exists or cannot be created: {path}"
        ) from exc
    created: os.stat_result | None = None
    descriptor_open = True
    try:
        created = os.fstat(descriptor)
        view = memoryview(payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                raise NativeRowsExcelExportError(f"short native-row output write: {path}")
            view = view[written:]
        os.fsync(descriptor)
        final_descriptor = os.fstat(descriptor)
        os.close(descriptor)
        descriptor_open = False
        linked = path.stat(follow_symlinks=False)
        if (
            not _same_regular_inode(created, final_descriptor)
            or not _same_regular_inode(final_descriptor, linked)
            or linked.st_size != len(payload)
            or path.read_bytes() != payload
        ):
            raise NativeRowsExcelExportError(f"native-row output identity drifted: {path}")
        return linked
    except BaseException:
        if descriptor_open:
            os.close(descriptor)
        if created is not None:
            try:
                _rollback_if_same(path, created)
            except FileNotFoundError:
                pass
        raise


def _assert_no_symlink_components(project_root: Path, directory: Path) -> None:
    relative = directory.relative_to(project_root)
    current = project_root
    for part in relative.parts:
        current = current / part
        if os.path.lexists(current) and current.is_symlink():
            raise NativeRowsExcelExportError(f"output directory contains a symlink: {current}")


def _publish_pair(
    workbook_path: Path,
    provenance_path: Path,
    workbook_bytes: bytes,
    provenance_bytes: bytes,
) -> None:
    if workbook_path.parent != provenance_path.parent or workbook_path == provenance_path:
        raise NativeRowsExcelExportError("native-row outputs must be distinct siblings")
    if os.path.lexists(workbook_path) or os.path.lexists(provenance_path):
        raise NativeRowsExcelExportError("native-row export refuses to overwrite output pair")
    workbook_identity = _write_exclusive(workbook_path, workbook_bytes)
    provenance_identity: os.stat_result | None = None
    try:
        provenance_identity = _write_exclusive(provenance_path, provenance_bytes)
        _fsync_directory(workbook_path.parent)
        if workbook_path.read_bytes() != workbook_bytes:
            raise NativeRowsExcelExportError("published native-row workbook bytes drifted")
        if provenance_path.read_bytes() != provenance_bytes:
            raise NativeRowsExcelExportError("published native-row provenance bytes drifted")
        if not _same_regular_inode(
            workbook_identity, workbook_path.stat(follow_symlinks=False)
        ) or not _same_regular_inode(
            provenance_identity, provenance_path.stat(follow_symlinks=False)
        ):
            raise NativeRowsExcelExportError("published native-row pair inode identity drifted")
    except BaseException as publication_error:
        rollback_errors: list[BaseException] = []
        if provenance_identity is not None:
            try:
                _rollback_if_same(provenance_path, provenance_identity)
            except BaseException as exc:
                rollback_errors.append(exc)
        try:
            _rollback_if_same(workbook_path, workbook_identity)
        except BaseException as exc:
            rollback_errors.append(exc)
        if rollback_errors:
            raise NativeRowsExcelExportError(
                "native-row pair publication rollback was incomplete: "
                + "; ".join(str(error) for error in rollback_errors)
            ) from publication_error
        raise


def export_registered_native_rows_excel(
    *,
    project_root: Path,
    rows_path: Path,
    expected_sha256: str,
    workbook_path: Path,
    provenance_path: Path,
    policy_path: Path = EXPORT_POLICY_RELATIVE_PATH,
    rows_policy_path: Path | None = None,
) -> NativeRowsExcelExportResult:
    """Validate one registered row artifact and exclusively publish its Excel pair."""

    project_root = project_root.resolve()
    rows_path = _project_path(project_root, rows_path)
    workbook_path = _project_path(project_root, workbook_path)
    provenance_path = _project_path(project_root, provenance_path)
    policy_path = _project_path(project_root, policy_path)
    if rows_policy_path is not None:
        rows_policy_path = _project_path(project_root, rows_policy_path)
    if _SHA256.fullmatch(expected_sha256) is None:
        raise NativeRowsExcelExportError("trusted native-row SHA-256 is invalid")
    if workbook_path.parent != provenance_path.parent or workbook_path == provenance_path:
        raise NativeRowsExcelExportError("native-row outputs must be distinct siblings")
    if workbook_path.suffix.casefold() != ".xlsx" or provenance_path.suffix.casefold() != ".json":
        raise NativeRowsExcelExportError("native-row output extensions must be .xlsx and .json")
    if os.path.lexists(workbook_path) or os.path.lexists(provenance_path):
        raise NativeRowsExcelExportError("native-row export refuses to overwrite output pair")

    policy = load_native_rows_excel_policy(policy_path, project_root)
    try:
        initial_input_bytes = rows_path.read_bytes()
    except OSError as exc:
        raise NativeRowsExcelExportError("cannot read trusted native-row input") from exc
    if _sha256(initial_input_bytes) != expected_sha256:
        raise NativeRowsExcelExportError("native-row input does not match trusted SHA-256")
    payload = load_registered_native_statement_rows(
        rows_path,
        project_root=project_root,
        expected_sha256=expected_sha256,
        policy_path=rows_policy_path,
    )
    source = _mapping(payload.get("source"), "native-row source")
    role = source.get("dataset_role")
    if not isinstance(role, str) or role not in policy.allowed_roles:
        raise NativeRowsExcelExportError("native-row dataset role is not exportable")
    role_directory = (project_root / policy.role_directories[role]).resolve()
    _relative(project_root, role_directory, "native-row role directory")
    for path in (rows_path, workbook_path, provenance_path):
        try:
            path.relative_to(role_directory)
        except ValueError as exc:
            raise NativeRowsExcelExportError(
                f"native-row {role} input/output must stay under {policy.role_directories[role]}"
            ) from exc
    _assert_no_symlink_components(project_root, workbook_path.parent)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    _assert_no_symlink_components(project_root, workbook_path.parent)
    implementation = _implementation_ledger(project_root)
    identity = NativeRowsInputIdentity(
        path=_relative(project_root, rows_path, "native-row input"),
        sha256=expected_sha256,
        size_bytes=len(initial_input_bytes),
    )
    artifacts = build_native_rows_excel_artifacts(
        payload,
        input_identity=identity,
        workbook_filename=workbook_path.name,
        policy=policy,
        implementation_ledger=implementation,
    )
    try:
        final_input_bytes = rows_path.read_bytes()
        final_policy_sha256 = _sha256(policy_path.read_bytes())
    except OSError as exc:
        raise NativeRowsExcelExportError(
            "native-row export input disappeared during build"
        ) from exc
    if final_input_bytes != initial_input_bytes or final_policy_sha256 != policy.sha256:
        raise NativeRowsExcelExportError("native-row export inputs changed during build")
    if _implementation_ledger(project_root) != implementation:
        raise NativeRowsExcelExportError("native-row export implementation changed during build")
    _publish_pair(
        workbook_path,
        provenance_path,
        artifacts.workbook_bytes,
        artifacts.provenance_bytes,
    )
    return NativeRowsExcelExportResult(
        workbook_path=workbook_path,
        provenance_path=provenance_path,
        workbook_sha256=artifacts.workbook_sha256,
        provenance_sha256=artifacts.provenance_sha256,
        workbook_size_bytes=len(artifacts.workbook_bytes),
        provenance_size_bytes=len(artifacts.provenance_bytes),
        summary=artifacts.summary,
    )


__all__ = [
    "CELL_HEADERS",
    "EXPORT_POLICY_RELATIVE_PATH",
    "HEADER_HEADERS",
    "METADATA_HEADERS",
    "PAGE_HEADERS",
    "ROW_HEADERS",
    "SHEET_NAMES",
    "NativeRowsExcelArtifacts",
    "NativeRowsExcelExportError",
    "NativeRowsExcelExportResult",
    "NativeRowsExcelPolicy",
    "NativeRowsInputIdentity",
    "build_native_rows_excel_artifacts",
    "export_registered_native_rows_excel",
    "load_native_rows_excel_policy",
]
