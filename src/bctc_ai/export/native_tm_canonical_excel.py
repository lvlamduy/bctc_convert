from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import shutil
import stat
import struct
import subprocess
import sys
import tempfile
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bctc_ai.document_phase import native_tm_document_artifact as _publication
from bctc_ai.export.canonical_xlsx import (
    CANONICAL_CORE_TIMESTAMP,
    EXCEL_CELL_TEXT_LIMIT,
    append_literal_row,
    deterministic_workbook_bytes,
    workbook_has_formula,
)
from bctc_ai.mapping.native_tm_canonical import (
    load_registered_native_tm_canonical_mapping,
)
from bctc_ai.rows.native_tm_observations import (
    load_registered_native_tm_observations,
)

EXPORT_POLICY_RELATIVE_PATH = Path("config/export/native-tm-canonical-excel-v1.yaml")

SHEET_NAMES = (
    "CANONICAL_OBSERVATIONS",
    "SCHEMA_DISPOSITIONS",
    "SOURCE_DISPOSITIONS",
    "SOURCE_OBJECTS",
    "VALIDATION",
    "RUN_METADATA",
)

_EXPORT_POLICY = "REGISTERED_NATIVE_TM_CANONICAL_EXCEL_V1"
_EXPORT_CLAIM = "RECEIPT_BOUND_NATIVE_TM_CANONICAL_OBSERVATIONS_AND_COMPLETE_ACCOUNTING_EXCEL"
_EXPORT_STATUS = "COMPLETE_REGISTERED_NATIVE_TM_CANONICAL_EXCEL_EXPORT"
_RECEIPT_TYPE = "REGISTERED_NATIVE_TM_CANONICAL_EXCEL_PROVENANCE_V1"
_MAPPING_FORMAT = "REGISTERED_NATIVE_TM_CANONICAL_MAPPING_RESULT_V1"
_MAPPING_POLICY = "REGISTERED_NATIVE_TM_CANONICAL_MAPPING_V1"
_MAPPING_CLAIM = "BOUNDED_SOURCE_EVIDENCE_TM_CANONICAL_MAPPING_ONLY"
_MAPPING_STATUS = "COMPLETE_NATIVE_TM_CANONICAL_DISPOSITION_ACCOUNTING"
_OBSERVATIONS_FORMAT = "REGISTERED_NATIVE_TM_OBSERVATIONS_RESULT_V1"
_OBSERVATIONS_POLICY = "REGISTERED_NATIVE_TM_OBSERVATIONS_V1"
_OBSERVATIONS_CLAIM = "SOURCE_ONLY_NATIVE_TM_OBSERVATION_FLATTENING"
_OBSERVATIONS_STATUS = "COMPLETE_NATIVE_TM_SOURCE_OBJECT_ACCOUNTING"
_DATASET_ROLE = "LOGIC_DEVELOPMENT"
_OUTPUT_DIRECTORY = "output/development"
_TM_SCHEMA_COUNT = 1710
_TERMINAL_OUTCOMES = (
    "OBSERVED_VALUE",
    "OBSERVED_ZERO",
    "DASH",
    "BLANK",
    "NOT_OBSERVED",
    "NOT_APPLICABLE",
    "AMBIGUOUS",
    "UNRESOLVED",
)
_IMPLEMENTATION_PATHS = (
    "pyproject.toml",
    "src/bctc_ai/__init__.py",
    "src/bctc_ai/axes/__init__.py",
    "src/bctc_ai/axes/header_binding.py",
    "src/bctc_ai/core/__init__.py",
    "src/bctc_ai/core/contracts.py",
    "src/bctc_ai/core/hashing.py",
    "src/bctc_ai/core/text.py",
    "src/bctc_ai/document_phase/__init__.py",
    "src/bctc_ai/document_phase/native_tm_document_artifact.py",
    "src/bctc_ai/export/__init__.py",
    "src/bctc_ai/export/canonical_xlsx.py",
    "src/bctc_ai/export/native_tm_canonical_excel.py",
    "src/bctc_ai/mapping/__init__.py",
    "src/bctc_ai/mapping/native_tm_canonical.py",
    "src/bctc_ai/ocr/__init__.py",
    "src/bctc_ai/ocr/native_text_quality_v2.py",
    "src/bctc_ai/ocr/pdf_text.py",
    "src/bctc_ai/rows/__init__.py",
    "src/bctc_ai/rows/native_tm_observations.py",
    "src/bctc_ai/rows/pdf_statement.py",
    "src/bctc_ai/tables/__init__.py",
    "src/bctc_ai/tables/geometry.py",
    "src/bctc_ai/tables/native_tm_regions.py",
    "uv.lock",
)
_SHA256 = re.compile(r"^[0-9a-f]{64}$")
_GIT_COMMIT = re.compile(r"^[0-9a-f]{40}$")
_JSON_CHUNK_SIZE = 30_000
_REPLAY_MAGIC = b"BCTC_AI_NATIVE_TM_CANONICAL_EXCEL_PAIR_V1\x00"


CANONICAL_OBSERVATION_HEADERS = (
    "ObservationOrder",
    "ObservationId",
    "ReportNormId",
    "CanonicalName",
    "SchemaDisplayOrder",
    "TerminalOutcome",
    "ReportedValueText",
    "CanonicalValueText",
    "Unit",
    "UnitMultiplierText",
    "PeriodType",
    "PeriodStart",
    "PeriodEnd",
    "AsOfDate",
    "PresentationScope",
    "RowId",
    "DimensionId",
    "ContextId",
    "MatchBasis",
    "SourceRecordSha256",
    "SourcePage",
    "SourceRawLabel",
    "SourceRawLabelJson",
    "SourceNormalizedLabel",
    "SourceObservationSha256",
    "SourceRowSha256",
    "SourceDimensionSha256",
    "SourceObservationJson",
    "SourceRowJson",
    "SourceDimensionJson",
    "ObservationRecordSha256",
    "ObservationJson",
)

SCHEMA_DISPOSITION_HEADERS = (
    "SchemaOrder",
    "ReportNormId",
    "CanonicalName",
    "SchemaDisplayOrder",
    "ParentReportNormId",
    "HierarchyLevel",
    "NotesSection",
    "ChildrenJson",
    "ContextStatus",
    "MappingEligible",
    "NoteFamilyRootId",
    "MappingDisposition",
    "TerminalOutcome",
    "IsObserved",
    "IsNotObserved",
    "IsUnresolved",
    "IsAmbiguous",
    "Reason",
    "SourceRowIdsJson",
    "SourceObservationIdsJson",
    "SchemaItemSha256",
    "SchemaDispositionSha256",
    "SchemaItemJson",
    "SchemaDispositionJson",
)

SOURCE_DISPOSITION_HEADERS = (
    "SourceOrder",
    "SourceObjectType",
    "SourceObjectId",
    "UpstreamSourceDisposition",
    "MappingDisposition",
    "IsMappedExistingItem",
    "IsUnresolved",
    "IsAmbiguous",
    "Reason",
    "ContextId",
    "RowId",
    "DimensionId",
    "TargetReportNormId",
    "CandidateReportNormIdsJson",
    "MatchBasis",
    "MatchedRetrievalKey",
    "AliasAuthorityType",
    "AliasAuthorityEvidenceSha256",
    "SourceObjectSha256",
    "SourceObjectJsonPartCount",
    "UpstreamSourceDispositionSha256",
    "UpstreamSourceDispositionJson",
    "SourceDispositionSha256",
    "SourceDispositionJson",
)

SOURCE_OBJECT_HEADERS = (
    "SourceObjectOrder",
    "SourceObjectType",
    "SourceObjectId",
    "PartNumber",
    "PartCount",
    "SourceObjectSha256",
    "SourceObjectJsonPart",
)

VALIDATION_HEADERS = (
    "ValidationOrder",
    "RecordType",
    "RecordIndex",
    "Status",
    "ReportNormId",
    "ContextId",
    "PartNumber",
    "PartCount",
    "RecordSha256",
    "RecordJsonPart",
)

RUN_METADATA_HEADERS = (
    "Key",
    "PartNumber",
    "PartCount",
    "ValueSha256",
    "ValueJsonPart",
)


class NativeTMCanonicalExcelExportError(RuntimeError):
    """A native-TM canonical workbook or its provenance failed closed."""


@dataclass(frozen=True, slots=True)
class NativeTMCanonicalExcelPolicy:
    path: Path
    sha256: str
    name: str
    claim_boundary: str
    accepted_mapping: dict[str, object]
    accepted_observations: dict[str, object]
    dataset_role: str
    output_directory: str
    schema_disposition_count: int
    workbook_creator: str


@dataclass(frozen=True, slots=True)
class ArtifactIdentity:
    path: str
    sha256: str
    size_bytes: int


@dataclass(frozen=True, slots=True)
class NativeTMCanonicalExcelArtifacts:
    workbook_bytes: bytes
    provenance_bytes: bytes
    workbook_sha256: str
    provenance_sha256: str
    summary: dict[str, object]


@dataclass(frozen=True, slots=True)
class NativeTMCanonicalExcelExportResult:
    workbook_path: Path
    provenance_path: Path
    workbook_sha256: str
    provenance_sha256: str
    workbook_size_bytes: int
    provenance_size_bytes: int
    summary: dict[str, object]


@dataclass(slots=True)
class _HeldFile:
    path: Path
    relative_path: str
    payload: bytes
    guard: Any
    identity: os.stat_result


@dataclass(slots=True)
class _InputBundle:
    mapping: _HeldFile
    observations: _HeldFile
    native_document: _HeldFile
    source_pdf: _HeldFile
    statement_discovery: _HeldFile
    mapping_payload: dict[str, Any]
    observations_payload: dict[str, Any]
    transitive_identities: tuple[ArtifactIdentity, ...]

    @property
    def guards(self) -> tuple[_HeldFile, ...]:
        return (
            self.mapping,
            self.observations,
            self.native_document,
            self.source_pdf,
            self.statement_discovery,
        )


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _canonical_json_bytes(value: object) -> bytes:
    try:
        rendered = json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            indent=2,
            sort_keys=True,
        )
    except (TypeError, ValueError) as exc:
        raise NativeTMCanonicalExcelExportError("export data is not canonical JSON") from exc
    return (rendered + "\n").encode("utf-8")


def _compact_json_unbounded(value: object) -> str:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError) as exc:
        raise NativeTMCanonicalExcelExportError("workbook data is not JSON-safe") from exc


def _compact_json(value: object) -> str:
    rendered = _compact_json_unbounded(value)
    if len(rendered) > EXCEL_CELL_TEXT_LIMIT:
        raise NativeTMCanonicalExcelExportError("workbook record exceeds Excel cell limit")
    return rendered


def _record_sha256(value: object) -> str:
    return _sha256(_compact_json_unbounded(value).encode("utf-8"))


def _decimal_text(value: Decimal) -> str:
    return "0" if value == 0 else format(value, "f")


def _mapping(value: object, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise NativeTMCanonicalExcelExportError(f"{label} must be an object")
    return value


def _records(value: object, label: str) -> Sequence[Mapping[str, Any]]:
    if not isinstance(value, list) or any(not isinstance(item, Mapping) for item in value):
        raise NativeTMCanonicalExcelExportError(f"{label} must be an array of objects")
    return value


def _canonical_relative_path(value: object, label: str) -> str:
    if not isinstance(value, str) or not value or "\\" in value:
        raise NativeTMCanonicalExcelExportError(f"{label} is not a canonical relative path")
    candidate = PurePosixPath(value)
    if (
        candidate.is_absolute()
        or value != candidate.as_posix()
        or re.match(r"^[A-Za-z]:", value)
        or any(part in {"", ".", ".."} for part in candidate.parts)
    ):
        raise NativeTMCanonicalExcelExportError(f"{label} is not a canonical relative path")
    return value


def _project_path(project_root: Path, path: Path, label: str) -> tuple[str, Path]:
    project_root = project_root.resolve()
    candidate = Path(path)
    if candidate.is_absolute():
        normalized = Path(os.path.normpath(str(candidate)))
        try:
            relative = normalized.relative_to(project_root).as_posix()
        except ValueError as exc:
            raise NativeTMCanonicalExcelExportError(
                f"{label} must stay inside the project root"
            ) from exc
    else:
        relative = candidate.as_posix()
    relative = _canonical_relative_path(relative, label)
    return relative, project_root.joinpath(*PurePosixPath(relative).parts)


def _same_regular_inode(left: os.stat_result, right: os.stat_result) -> bool:
    return (
        stat.S_ISREG(left.st_mode)
        and stat.S_ISREG(right.st_mode)
        and (left.st_dev, left.st_ino) == (right.st_dev, right.st_ino)
    )


def _open_held_file(
    project_root: Path,
    path: Path,
    label: str,
) -> _HeldFile:
    relative, absolute = _project_path(project_root, path, label)
    try:
        native_guard = _publication._open_artifact_read_guard(
            project_root.resolve(), absolute, relative
        )
    except (OSError, _publication.NativeTMDocumentArtifactError) as exc:
        raise NativeTMCanonicalExcelExportError(f"cannot open {label}") from exc
    return _HeldFile(
        path=absolute,
        relative_path=relative,
        payload=bytes(native_guard.payload),
        guard=native_guard,
        identity=native_guard.identity,
    )


def _assert_held_stable(guard: _HeldFile, label: str) -> None:
    try:
        _publication._revalidate_artifact_read_guard(guard.guard)
    except _publication.NativeTMDocumentArtifactError as exc:
        raise NativeTMCanonicalExcelExportError(f"{label} changed during export") from exc


def _close_held(guard: _HeldFile) -> None:
    _publication._close_artifact_read_guard(guard.guard)


def _parse_canonical_json(payload: bytes, label: str) -> dict[str, Any]:
    try:
        value = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise NativeTMCanonicalExcelExportError(f"{label} is not JSON") from exc
    if not isinstance(value, dict) or payload != _canonical_json_bytes(value):
        raise NativeTMCanonicalExcelExportError(f"{label} is not canonical JSON")
    return value


def _validate_identity(identity: ArtifactIdentity, label: str) -> None:
    _canonical_relative_path(identity.path, f"{label} path")
    if (
        not isinstance(identity.sha256, str)
        or _SHA256.fullmatch(identity.sha256) is None
        or isinstance(identity.size_bytes, bool)
        or not isinstance(identity.size_bytes, int)
        or identity.size_bytes < 1
    ):
        raise NativeTMCanonicalExcelExportError(f"{label} identity is invalid")


def _identity_from_record(value: object, label: str) -> ArtifactIdentity:
    record = _mapping(value, label)
    identity = ArtifactIdentity(
        path=_canonical_relative_path(record.get("path"), f"{label} path"),
        sha256=record.get("sha256"),
        size_bytes=record.get("size_bytes"),
    )
    _validate_identity(identity, label)
    return identity


def _expected_policy_payload() -> dict[str, object]:
    return {
        "version": 1,
        "policy": _EXPORT_POLICY,
        "claim_boundary": _EXPORT_CLAIM,
        "accepted_input": {
            "mapping": {
                "format_version": _MAPPING_FORMAT,
                "policy": _MAPPING_POLICY,
                "claim_boundary": _MAPPING_CLAIM,
                "status": _MAPPING_STATUS,
            },
            "native_tm_observations": {
                "format_version": _OBSERVATIONS_FORMAT,
                "policy": _OBSERVATIONS_POLICY,
                "claim_boundary": _OBSERVATIONS_CLAIM,
                "status": _OBSERVATIONS_STATUS,
            },
            "dataset_role": _DATASET_ROLE,
            "trusted_mapping_sha256_required": True,
            "observations_are_receipt_bound_not_public_input": True,
            "strict_mapping_producer_commit_replay_required": True,
            "strict_observations_producer_commit_replay_required": True,
        },
        "denominators": {
            "tm_schema_dispositions": _TM_SCHEMA_COUNT,
            "source_object_dispositions": "EXACT_POSITIVE_MAPPING_DECLARED_DENOMINATOR",
            "exactly_one_schema_disposition_per_report_norm_id": True,
            "exactly_one_mapping_disposition_per_source_object": True,
        },
        "workbook": {
            "sheets": list(SHEET_NAMES),
            "creator": "bctc-ai/native-tm-canonical-excel-v1",
            "deterministic_xlsx": True,
            "formulas_allowed": False,
            "imputation_allowed": False,
            "force_mapping_allowed": False,
            "current_mutable_schema_inputs_allowed": False,
            "embedded_producer_schema_snapshot_is_authority": True,
            "preserve_full_record_json": True,
            "explicit_terminal_outcome_columns": True,
            "explicit_unresolved_and_not_observed_rows": True,
        },
        "publication": {
            "output_directory": _OUTPUT_DIRECTORY,
            "paired_workbook_and_provenance": True,
            "workbook_written_before_provenance_completion_marker": True,
            "exclusive_no_overwrite": True,
            "sibling_outputs_required": True,
            "absolute_project_paths_in_receipts_allowed": False,
            "canonical_provenance_json": True,
            "rollback_after_failed_strict_replay": True,
        },
        "isolation": {
            "public_inputs": [
                "REGISTERED_NATIVE_TM_CANONICAL_MAPPING_PATH",
                "TRUSTED_MAPPING_SHA256",
            ],
            "internally_authenticated_transitive_inputs": [
                "REGISTERED_NATIVE_TM_OBSERVATIONS_ARTIFACT",
                "REGISTERED_NATIVE_TM_DOCUMENT_ARTIFACT",
                "SOURCE_PDF",
                "ACCEPTED_STATEMENT_DISCOVERY",
            ],
            "fd_held_nofollow_stability_guards_required": True,
            "producer_snapshots_replayed_without_current_schema": True,
            "current_mutable_schema_loaded": False,
            "template_inputs_loaded": False,
            "historical_values_loaded": False,
            "role_a_outputs_loaded": False,
            "human_review_outputs_loaded": False,
        },
    }


def _policy_from_bytes(
    payload: bytes,
    *,
    path: Path,
) -> NativeTMCanonicalExcelPolicy:
    try:
        value = yaml.safe_load(payload)
    except yaml.YAMLError as exc:
        raise NativeTMCanonicalExcelExportError("native-TM Excel policy is invalid YAML") from exc
    if value != _expected_policy_payload():
        raise NativeTMCanonicalExcelExportError("native-TM Excel policy drifted")
    accepted = value["accepted_input"]
    denominators = value["denominators"]
    return NativeTMCanonicalExcelPolicy(
        path=path,
        sha256=_sha256(payload),
        name=_EXPORT_POLICY,
        claim_boundary=_EXPORT_CLAIM,
        accepted_mapping=dict(accepted["mapping"]),
        accepted_observations=dict(accepted["native_tm_observations"]),
        dataset_role=_DATASET_ROLE,
        output_directory=_OUTPUT_DIRECTORY,
        schema_disposition_count=int(denominators["tm_schema_dispositions"]),
        workbook_creator=str(value["workbook"]["creator"]),
    )


def load_native_tm_canonical_excel_policy(
    path: Path,
    project_root: Path,
) -> NativeTMCanonicalExcelPolicy:
    project_root = project_root.resolve()
    relative, absolute = _project_path(project_root, path, "native-TM Excel policy")
    if relative != EXPORT_POLICY_RELATIVE_PATH.as_posix():
        raise NativeTMCanonicalExcelExportError(
            f"native-TM Excel requires {EXPORT_POLICY_RELATIVE_PATH.as_posix()}"
        )
    guard = _open_held_file(project_root, absolute, "native-TM Excel policy")
    try:
        policy = _policy_from_bytes(guard.payload, path=absolute)
        _assert_held_stable(guard, "native-TM Excel policy")
        return policy
    finally:
        _close_held(guard)


def _accepted_contract(
    payload: Mapping[str, Any],
    expected: Mapping[str, object],
    label: str,
) -> None:
    for key, value in expected.items():
        if payload.get(key) != value:
            raise NativeTMCanonicalExcelExportError(f"{label} {key} is not accepted")


def _open_identity_guard(
    project_root: Path,
    identity: ArtifactIdentity,
    label: str,
) -> _HeldFile:
    guard = _open_held_file(project_root, Path(identity.path), label)
    if (
        guard.relative_path != identity.path
        or len(guard.payload) != identity.size_bytes
        or _sha256(guard.payload) != identity.sha256
    ):
        _close_held(guard)
        raise NativeTMCanonicalExcelExportError(f"{label} identity drifted")
    return guard


def _expected_upstream_envelope(
    identity: ArtifactIdentity,
    payload: Mapping[str, Any],
) -> dict[str, object]:
    code = _mapping(payload.get("code"), "upstream producer code")
    return {
        "path": identity.path,
        "sha256": identity.sha256,
        "size_bytes": identity.size_bytes,
        "format_version": payload.get("format_version"),
        "policy": payload.get("policy"),
        "claim_boundary": payload.get("claim_boundary"),
        "status": payload.get("status"),
        "run_id": payload.get("run_id"),
        "producer_git_commit": code.get("commit"),
    }


def _strict_load_input_bundle(
    *,
    project_root: Path,
    mapping_path: Path,
    mapping_expected_sha256: str,
    policy: NativeTMCanonicalExcelPolicy,
) -> _InputBundle:
    """Authenticate the public mapping before following any receipt-bound path."""

    if (
        not isinstance(mapping_expected_sha256, str)
        or _SHA256.fullmatch(mapping_expected_sha256) is None
    ):
        raise NativeTMCanonicalExcelExportError("trusted mapping SHA-256 is invalid")
    mapping_guard = _open_held_file(project_root, mapping_path, "native-TM canonical mapping")
    guards: list[_HeldFile] = [mapping_guard]
    try:
        if not mapping_guard.relative_path.startswith(f"{policy.output_directory}/"):
            raise NativeTMCanonicalExcelExportError(
                "native-TM canonical mapping must stay in the configured output directory"
            )
        if _sha256(mapping_guard.payload) != mapping_expected_sha256:
            raise NativeTMCanonicalExcelExportError(
                "native-TM canonical mapping does not match trusted SHA-256"
            )
        try:
            authenticated_mapping = load_registered_native_tm_canonical_mapping(
                mapping_guard.path,
                project_root=project_root,
                expected_sha256=mapping_expected_sha256,
            )
        except Exception as exc:
            raise NativeTMCanonicalExcelExportError(
                "native-TM canonical mapping failed authoritative producer replay"
            ) from exc
        if (
            not isinstance(authenticated_mapping, dict)
            or _canonical_json_bytes(authenticated_mapping) != mapping_guard.payload
        ):
            raise NativeTMCanonicalExcelExportError(
                "authenticated mapping differs from fd-held bytes"
            )
        _accepted_contract(authenticated_mapping, policy.accepted_mapping, "mapping input")

        observations_envelope = _mapping(
            authenticated_mapping.get("native_tm_observations"),
            "authenticated mapping observation receipt",
        )
        observations_identity = _identity_from_record(
            observations_envelope, "receipt-bound native-TM observations"
        )
        if not observations_identity.path.startswith(f"{policy.output_directory}/"):
            raise NativeTMCanonicalExcelExportError(
                "receipt-bound observations leave the configured output directory"
            )
        observations_guard = _open_identity_guard(
            project_root,
            observations_identity,
            "receipt-bound native-TM observations",
        )
        guards.append(observations_guard)
        try:
            authenticated_observations = load_registered_native_tm_observations(
                observations_guard.path,
                project_root=project_root,
                expected_sha256=observations_identity.sha256,
            )
        except Exception as exc:
            raise NativeTMCanonicalExcelExportError(
                "receipt-bound observations failed authoritative producer replay"
            ) from exc
        if (
            not isinstance(authenticated_observations, dict)
            or _canonical_json_bytes(authenticated_observations) != observations_guard.payload
        ):
            raise NativeTMCanonicalExcelExportError(
                "authenticated observations differ from fd-held bytes"
            )
        _accepted_contract(
            authenticated_observations,
            policy.accepted_observations,
            "receipt-bound observations",
        )
        if observations_envelope != _expected_upstream_envelope(
            observations_identity, authenticated_observations
        ):
            raise NativeTMCanonicalExcelExportError(
                "mapping does not bind the observations envelope exactly"
            )
        if authenticated_mapping.get("source") != authenticated_observations.get("source"):
            raise NativeTMCanonicalExcelExportError(
                "mapping and observations source envelopes differ"
            )
        source = _mapping(authenticated_mapping.get("source"), "mapping source")
        if source.get("dataset_role") != policy.dataset_role:
            raise NativeTMCanonicalExcelExportError(
                "native-TM mapping dataset role is not exportable"
            )

        native_envelope = _mapping(
            authenticated_observations.get("native_tm_document"),
            "authenticated observations native-document receipt",
        )
        native_identity = _identity_from_record(native_envelope, "receipt-bound native-TM document")
        if not native_identity.path.startswith(f"{policy.output_directory}/"):
            raise NativeTMCanonicalExcelExportError(
                "receipt-bound native-TM document leaves the output directory"
            )
        native_guard = _open_identity_guard(
            project_root, native_identity, "receipt-bound native-TM document"
        )
        guards.append(native_guard)
        native_payload = _parse_canonical_json(
            native_guard.payload, "receipt-bound native-TM document"
        )
        if native_envelope != _expected_upstream_envelope(native_identity, native_payload):
            raise NativeTMCanonicalExcelExportError(
                "observations do not bind the native-TM document envelope exactly"
            )
        if native_payload.get("source") != authenticated_observations.get("source"):
            raise NativeTMCanonicalExcelExportError(
                "native document and observations source envelopes differ"
            )
        source_identity_record = _mapping(native_payload.get("source"), "native document source")
        source_identity = ArtifactIdentity(
            path=_canonical_relative_path(
                source_identity_record.get("relative_path"), "source PDF path"
            ),
            sha256=source_identity_record.get("sha256"),
            size_bytes=source_identity_record.get("size_bytes"),
        )
        _validate_identity(source_identity, "source PDF")
        source_guard = _open_identity_guard(project_root, source_identity, "source PDF")
        guards.append(source_guard)
        discovery_record = _mapping(
            native_payload.get("statement_discovery"), "statement discovery receipt"
        )
        discovery_identity = _identity_from_record(discovery_record, "accepted statement discovery")
        if not discovery_identity.path.startswith(f"{policy.output_directory}/"):
            raise NativeTMCanonicalExcelExportError(
                "accepted statement discovery leaves the output directory"
            )
        discovery_guard = _open_identity_guard(
            project_root, discovery_identity, "accepted statement discovery"
        )
        guards.append(discovery_guard)
        _parse_canonical_json(discovery_guard.payload, "accepted statement discovery")
        inode_keys = {(item.identity.st_dev, item.identity.st_ino) for item in guards}
        if len(inode_keys) != len(guards):
            raise NativeTMCanonicalExcelExportError(
                "mapping lineage artifacts must be distinct regular files"
            )
        bundle = _InputBundle(
            mapping=mapping_guard,
            observations=observations_guard,
            native_document=native_guard,
            source_pdf=source_guard,
            statement_discovery=discovery_guard,
            mapping_payload=authenticated_mapping,
            observations_payload=authenticated_observations,
            transitive_identities=(native_identity, source_identity, discovery_identity),
        )
        _assert_bundle_stable(bundle)
        return bundle
    except BaseException:
        for guard in reversed(guards):
            _close_held(guard)
        raise


def _close_input_bundle(bundle: _InputBundle) -> None:
    for guard in reversed(bundle.guards):
        _close_held(guard)


def _assert_bundle_stable(bundle: _InputBundle) -> None:
    for guard, label in (
        (bundle.mapping, "native-TM canonical mapping"),
        (bundle.observations, "receipt-bound native-TM observations"),
        (bundle.native_document, "receipt-bound native-TM document"),
        (bundle.source_pdf, "source PDF"),
        (bundle.statement_discovery, "accepted statement discovery"),
    ):
        _assert_held_stable(guard, label)


def _snapshot_records(
    mapping_payload: Mapping[str, Any],
    *,
    policy: NativeTMCanonicalExcelPolicy,
) -> tuple[
    Sequence[Mapping[str, Any]],
    Sequence[Mapping[str, Any]],
]:
    snapshots = _mapping(mapping_payload.get("producer_snapshots"), "producer snapshots")
    schema_snapshot = _mapping(snapshots.get("tm_schema"), "producer TM schema snapshot")
    context_snapshot = _mapping(snapshots.get("tm_context"), "producer TM context snapshot")
    schema = _records(schema_snapshot.get("records"), "producer TM schema records")
    contexts = _records(context_snapshot.get("records"), "producer TM context records")
    if (
        schema_snapshot.get("record_count") != len(schema)
        or schema_snapshot.get("payload_sha256") != _record_sha256(schema)
        or context_snapshot.get("record_count") != len(contexts)
        or context_snapshot.get("payload_sha256") != _record_sha256(contexts)
        or len(schema) != policy.schema_disposition_count
        or len(contexts) != len(schema)
    ):
        raise NativeTMCanonicalExcelExportError(
            "producer TM schema/context snapshot denominator or hash drifted"
        )
    schema_ids = [item.get("schema_id") for item in schema]
    context_ids = [item.get("report_norm_id") for item in contexts]
    if (
        any(isinstance(item, bool) or not isinstance(item, int) for item in schema_ids)
        or len(set(schema_ids)) != len(schema_ids)
        or schema_ids != context_ids
        or [item.get("display_order") for item in schema]
        != list(range(policy.schema_disposition_count))
        or [item.get("display_order") for item in contexts]
        != list(range(policy.schema_disposition_count))
    ):
        raise NativeTMCanonicalExcelExportError("producer TM schema/context snapshot order drifted")
    return schema, contexts


def _validate_input_pair(
    mapping_payload: Mapping[str, Any],
    observations_payload: Mapping[str, Any],
    *,
    mapping_identity: ArtifactIdentity,
    observations_identity: ArtifactIdentity,
    policy: NativeTMCanonicalExcelPolicy,
) -> Mapping[str, Any]:
    _validate_identity(mapping_identity, "mapping input")
    _validate_identity(observations_identity, "observations input")
    _accepted_contract(mapping_payload, policy.accepted_mapping, "mapping input")
    _accepted_contract(observations_payload, policy.accepted_observations, "observations input")
    mapping_bytes = _canonical_json_bytes(mapping_payload)
    observations_bytes = _canonical_json_bytes(observations_payload)
    if (
        len(mapping_bytes) != mapping_identity.size_bytes
        or _sha256(mapping_bytes) != mapping_identity.sha256
        or len(observations_bytes) != observations_identity.size_bytes
        or _sha256(observations_bytes) != observations_identity.sha256
    ):
        raise NativeTMCanonicalExcelExportError(
            "strict-loaded payload differs from trusted input identity"
        )
    observations_envelope = _mapping(
        mapping_payload.get("native_tm_observations"), "mapping observations envelope"
    )
    if observations_envelope != _expected_upstream_envelope(
        observations_identity, observations_payload
    ):
        raise NativeTMCanonicalExcelExportError(
            "mapping and receipt-bound observations are not an exact pair"
        )
    if mapping_payload.get("source") != observations_payload.get("source"):
        raise NativeTMCanonicalExcelExportError("mapping/observations source pair drifted")
    source = _mapping(mapping_payload.get("source"), "mapping source")
    if source.get("dataset_role") != policy.dataset_role:
        raise NativeTMCanonicalExcelExportError("mapping dataset role is not exportable")
    return source


def _schema_projection(
    mapping_payload: Mapping[str, Any],
    *,
    policy: NativeTMCanonicalExcelPolicy,
) -> tuple[
    list[dict[str, object]],
    dict[int, Mapping[str, Any]],
    dict[int, Mapping[str, Any]],
]:
    schema, contexts = _snapshot_records(mapping_payload, policy=policy)
    dispositions = _records(mapping_payload.get("schema_dispositions"), "TM schema dispositions")
    if len(dispositions) != policy.schema_disposition_count:
        raise NativeTMCanonicalExcelExportError("TM schema disposition denominator drifted")
    schema_by_id = {int(item["schema_id"]): item for item in schema}
    disposition_by_id: dict[int, Mapping[str, Any]] = {}
    projected: list[dict[str, object]] = []
    claimed_row_ids: set[str] = set()
    claimed_observation_ids: set[str] = set()
    for order, (item, context, disposition) in enumerate(
        zip(schema, contexts, dispositions, strict=True), start=1
    ):
        identifier = item["schema_id"]
        if (
            disposition.get("report_norm_id") != identifier
            or disposition.get("canonical_name") != item.get("canonical_name")
            or disposition.get("display_order") != item.get("display_order")
            or disposition.get("context_status") != context.get("context_status")
            or disposition.get("mapping_eligible") != context.get("mapping_eligible")
            or disposition.get("parent_report_norm_id") != context.get("parent_report_norm_id")
            or disposition.get("note_family_root_id") != context.get("note_family_root_id")
            or not isinstance(disposition.get("mapping_disposition"), str)
            or disposition.get("terminal_outcome") not in _TERMINAL_OUTCOMES
            or not isinstance(disposition.get("reason"), str)
            or not isinstance(disposition.get("source_row_ids"), list)
            or not isinstance(disposition.get("source_observation_ids"), list)
        ):
            raise NativeTMCanonicalExcelExportError(
                "TM schema disposition does not join its producer snapshot"
            )
        source_row_ids = disposition.get("source_row_ids")
        source_observation_ids = disposition.get("source_observation_ids")
        mapping_disposition = disposition.get("mapping_disposition")
        terminal_outcome = disposition.get("terminal_outcome")
        if (
            any(not isinstance(value, str) or not value for value in source_row_ids)
            or any(not isinstance(value, str) or not value for value in source_observation_ids)
            or len(source_row_ids) != len(set(source_row_ids))
            or len(source_observation_ids) != len(set(source_observation_ids))
            or claimed_row_ids.intersection(source_row_ids)
            or claimed_observation_ids.intersection(source_observation_ids)
            or (
                mapping_disposition == "EXISTING_ITEM"
                and (
                    terminal_outcome not in {"OBSERVED_VALUE", "OBSERVED_ZERO"}
                    or len(source_row_ids) != 1
                    or not source_observation_ids
                )
            )
            or (
                mapping_disposition == "LOCALLY_NOT_OBSERVED"
                and (terminal_outcome != "NOT_OBSERVED" or source_row_ids or source_observation_ids)
            )
            or (
                mapping_disposition == "UNRESOLVED"
                and (terminal_outcome != "UNRESOLVED" or source_row_ids or source_observation_ids)
            )
            or mapping_disposition not in {"EXISTING_ITEM", "LOCALLY_NOT_OBSERVED", "UNRESOLVED"}
        ):
            raise NativeTMCanonicalExcelExportError(
                "TM schema source citations are not globally one-to-one"
            )
        claimed_row_ids.update(source_row_ids)
        claimed_observation_ids.update(source_observation_ids)
        if identifier in disposition_by_id:
            raise NativeTMCanonicalExcelExportError("duplicate TM schema disposition")
        disposition_by_id[int(identifier)] = disposition
        outcome = str(disposition["terminal_outcome"])
        projected.append(
            dict(
                zip(
                    SCHEMA_DISPOSITION_HEADERS,
                    (
                        order,
                        identifier,
                        item.get("canonical_name"),
                        item.get("display_order"),
                        item.get("parent_id"),
                        item.get("hierarchy_level"),
                        item.get("notes_section"),
                        _compact_json(item.get("children")),
                        context.get("context_status"),
                        context.get("mapping_eligible"),
                        context.get("note_family_root_id"),
                        disposition.get("mapping_disposition"),
                        outcome,
                        outcome in {"OBSERVED_VALUE", "OBSERVED_ZERO"},
                        outcome == "NOT_OBSERVED",
                        outcome == "UNRESOLVED",
                        outcome == "AMBIGUOUS",
                        disposition.get("reason"),
                        _compact_json(disposition.get("source_row_ids")),
                        _compact_json(disposition.get("source_observation_ids")),
                        _record_sha256(item),
                        _record_sha256(disposition),
                        _compact_json(item),
                        _compact_json(disposition),
                    ),
                    strict=True,
                )
            )
        )
    coverage = _mapping(mapping_payload.get("coverage"), "mapping coverage")
    terminal_counts = Counter(item["terminal_outcome"] for item in dispositions)
    expected_counts = {status: terminal_counts[status] for status in _TERMINAL_OUTCOMES}
    if (
        coverage.get("statement_type") != "TM"
        or coverage.get("schema_item_count") != policy.schema_disposition_count
        or coverage.get("schema_disposition_count") != policy.schema_disposition_count
        or coverage.get("terminal_outcome_counts") != expected_counts
        or coverage.get("reason_counts")
        != dict(sorted(Counter(item["reason"] for item in dispositions).items()))
        or coverage.get("exactly_one_terminal_outcome_per_schema_id") is not True
        or coverage.get("workbook_display_order_complete") is not True
    ):
        raise NativeTMCanonicalExcelExportError("TM schema coverage receipt drifted")
    return projected, schema_by_id, disposition_by_id


def _source_projection(
    mapping_payload: Mapping[str, Any],
    observations_payload: Mapping[str, Any],
    *,
    policy: NativeTMCanonicalExcelPolicy,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    mapping_records = _records(
        mapping_payload.get("source_dispositions"), "mapping source dispositions"
    )
    upstream_records = _records(
        observations_payload.get("source_dispositions"), "upstream source dispositions"
    )
    accounting = _mapping(mapping_payload.get("source_accounting"), "source accounting")
    declared_count = accounting.get("mapping_source_disposition_count")
    if (
        isinstance(declared_count, bool)
        or not isinstance(declared_count, int)
        or declared_count < 1
        or accounting.get("upstream_source_object_count") != declared_count
        or len(mapping_records) != declared_count
        or len(upstream_records) != declared_count
    ):
        raise NativeTMCanonicalExcelExportError("native-TM source disposition denominator drifted")

    source_objects, upstream_by_id = _source_object_index(observations_payload)
    if len(source_objects) != declared_count or len(upstream_by_id) != declared_count:
        raise NativeTMCanonicalExcelExportError("native-TM source object partition drifted")
    observations_accounting = _mapping(
        observations_payload.get("source_accounting"), "observations source accounting"
    )
    observations_counts = _mapping(
        observations_accounting.get("counts"), "observations source accounting counts"
    )
    if (
        observations_counts.get("source_disposition_count") != declared_count
        or observations_accounting.get("source_object_accounting_complete") is not True
    ):
        raise NativeTMCanonicalExcelExportError(
            "observations source-accounting denominator drifted"
        )
    projected: list[dict[str, object]] = []
    source_object_rows: list[dict[str, object]] = []
    seen: set[str] = set()
    previous_key: tuple[str, str] | None = None
    required = {
        "source_object_type",
        "source_object_id",
        "upstream_source_disposition",
        "mapping_disposition",
        "reason",
        "context_id",
        "row_id",
        "dimension_id",
        "target_report_norm_id",
        "candidate_report_norm_ids",
        "match_basis",
        "matched_retrieval_key",
        "alias_authority_type",
        "alias_authority_evidence_sha256",
    }
    for order, record in enumerate(mapping_records, start=1):
        if not required <= set(record):
            raise NativeTMCanonicalExcelExportError(
                "mapping source disposition fields are incomplete"
            )
        identifier = record.get("source_object_id")
        object_type = record.get("source_object_type")
        if (
            not isinstance(identifier, str)
            or not identifier
            or identifier in seen
            or not isinstance(object_type, str)
            or not object_type
        ):
            raise NativeTMCanonicalExcelExportError(
                "mapping source disposition identity is invalid"
            )
        key = (object_type, identifier)
        if previous_key is not None and key < previous_key:
            raise NativeTMCanonicalExcelExportError(
                "mapping source dispositions are not in producer order"
            )
        previous_key = key
        seen.add(identifier)
        upstream = upstream_by_id.get(identifier)
        if (
            upstream is None
            or upstream.get("source_object_type") != object_type
            or upstream.get("source_disposition") != record.get("upstream_source_disposition")
            or not isinstance(record.get("mapping_disposition"), str)
            or not isinstance(record.get("reason"), str)
            or not isinstance(record.get("candidate_report_norm_ids"), list)
        ):
            raise NativeTMCanonicalExcelExportError(
                "mapping disposition does not join its upstream source object"
            )
        source_object = source_objects.get(identifier)
        if source_object is None or source_object.get("record_type") != object_type:
            raise NativeTMCanonicalExcelExportError(
                "mapping disposition does not join its receipt-bound source object"
            )
        source_object_sha256, source_object_parts = _chunks(source_object)
        upstream_sha256 = _record_sha256(upstream)
        disposition = str(record["mapping_disposition"])
        projected.append(
            dict(
                zip(
                    SOURCE_DISPOSITION_HEADERS,
                    (
                        order,
                        object_type,
                        identifier,
                        record.get("upstream_source_disposition"),
                        disposition,
                        disposition == "MAPPED_EXISTING_ITEM",
                        disposition == "UNRESOLVED",
                        disposition == "AMBIGUOUS",
                        record.get("reason"),
                        record.get("context_id"),
                        record.get("row_id"),
                        record.get("dimension_id"),
                        record.get("target_report_norm_id"),
                        _compact_json(record.get("candidate_report_norm_ids")),
                        record.get("match_basis"),
                        record.get("matched_retrieval_key"),
                        record.get("alias_authority_type"),
                        record.get("alias_authority_evidence_sha256"),
                        source_object_sha256,
                        len(source_object_parts),
                        upstream_sha256,
                        _compact_json(upstream),
                        _record_sha256(record),
                        _compact_json(record),
                    ),
                    strict=True,
                )
            )
        )
        for part_number, part in enumerate(source_object_parts, start=1):
            source_object_rows.append(
                dict(
                    zip(
                        SOURCE_OBJECT_HEADERS,
                        (
                            order,
                            object_type,
                            identifier,
                            part_number,
                            len(source_object_parts),
                            source_object_sha256,
                            part,
                        ),
                        strict=True,
                    )
                )
            )
    if seen != set(upstream_by_id):
        raise NativeTMCanonicalExcelExportError(
            "mapping source dispositions do not exhaust upstream source objects"
        )
    if accounting != {
        "upstream_source_object_count": declared_count,
        "mapping_source_disposition_count": declared_count,
        "upstream_source_dispositions_sha256": _record_sha256(upstream_records),
        "mapping_source_dispositions_sha256": _record_sha256(mapping_records),
        "exactly_one_mapping_disposition_per_upstream_source_object": True,
        "source_object_accounting_complete": True,
    }:
        raise NativeTMCanonicalExcelExportError("mapping source-accounting receipt drifted")
    return projected, source_object_rows


def _source_object_index(
    observations_payload: Mapping[str, Any],
) -> tuple[dict[str, Mapping[str, Any]], dict[str, Mapping[str, Any]]]:
    """Reconstruct the mapper's exact primary/evidence source-object partition."""

    upstream_by_id: dict[str, Mapping[str, Any]] = {}
    for record in _records(
        observations_payload.get("source_dispositions"), "upstream source dispositions"
    ):
        if set(record) != {"source_object_id", "source_object_type", "source_disposition"}:
            raise NativeTMCanonicalExcelExportError("upstream source disposition fields drifted")
        identifier = record.get("source_object_id")
        if not isinstance(identifier, str) or not identifier or identifier in upstream_by_id:
            raise NativeTMCanonicalExcelExportError(
                "upstream source disposition identity is invalid"
            )
        if not isinstance(record.get("source_object_type"), str) or not isinstance(
            record.get("source_disposition"), str
        ):
            raise NativeTMCanonicalExcelExportError(
                "upstream source disposition identity is invalid"
            )
        upstream_by_id[identifier] = record

    source_objects: dict[str, Mapping[str, Any]] = {}

    def register(record: Mapping[str, Any], expected_type: str, label: str) -> None:
        identifier = record.get("source_object_id")
        if (
            not isinstance(identifier, str)
            or not identifier
            or identifier in source_objects
            or record.get("record_type") != expected_type
            or upstream_by_id.get(identifier)
            != {
                "source_object_id": identifier,
                "source_object_type": expected_type,
                "source_disposition": record.get("source_disposition"),
            }
        ):
            raise NativeTMCanonicalExcelExportError(f"{label} source identity is not exact")
        source_objects[identifier] = record

    for collection, record_type, identity_field in (
        ("page_inventory", "PAGE_CONTEXT", "page_id"),
        ("contexts", "CONTEXT", "context_id"),
        ("rows", "ROW", "row_id"),
        ("dimensions", "DIMENSION", "dimension_id"),
        ("observations", "OBSERVATION", "observation_id"),
    ):
        natural_ids: set[str] = set()
        for record in _records(observations_payload.get(collection), collection):
            natural_id = record.get(identity_field)
            if not isinstance(natural_id, str) or not natural_id or natural_id in natural_ids:
                raise NativeTMCanonicalExcelExportError(
                    f"{collection} natural identity is not exact"
                )
            natural_ids.add(natural_id)
            register(record, record_type, f"{collection} record")

    evidence = _mapping(observations_payload.get("source_evidence"), "source evidence")

    def register_evidence(value: object, label: str) -> None:
        if isinstance(value, list):
            for ordinal, item in enumerate(value):
                register_evidence(item, f"{label}[{ordinal}]")
            return
        if not isinstance(value, Mapping):
            return
        if "source_object_id" in value:
            record_type = value.get("record_type")
            if not isinstance(record_type, str) or not record_type:
                raise NativeTMCanonicalExcelExportError(f"{label} source object type is invalid")
            register(value, record_type, label)
        for key, item in value.items():
            if key != "source_object_id":
                register_evidence(item, f"{label}.{key}")

    for category, raw_records in evidence.items():
        if not isinstance(category, str) or not category:
            raise NativeTMCanonicalExcelExportError("source evidence category is invalid")
        for ordinal, record in enumerate(_records(raw_records, f"{category} evidence")):
            if "source_object_id" not in record:
                raise NativeTMCanonicalExcelExportError(
                    "source evidence record has no source object identity"
                )
            register_evidence(record, f"{category} evidence record {ordinal}")

    if set(source_objects) != set(upstream_by_id):
        raise NativeTMCanonicalExcelExportError("native-TM source object partition drifted")
    return source_objects, upstream_by_id


def _canonical_observation_projection(
    mapping_payload: Mapping[str, Any],
    observations_payload: Mapping[str, Any],
    *,
    schema_by_id: Mapping[int, Mapping[str, Any]],
    disposition_by_id: Mapping[int, Mapping[str, Any]],
) -> list[dict[str, object]]:
    observations = _records(mapping_payload.get("canonical_observations"), "canonical observations")
    report_scope_binding = _mapping(
        observations_payload.get("report_scope_binding"), "observations report scope binding"
    )
    report_scope = report_scope_binding.get("scope")
    if report_scope_binding.get(
        "binding_status"
    ) != "RESOLVED_UNANIMOUS_SOURCE_VISIBLE_TM_HEADERS" or report_scope not in {
        "CONSOLIDATED",
        "SEPARATE",
    }:
        raise NativeTMCanonicalExcelExportError("observations report scope is unresolved")
    source_observations = _unique_index(
        observations_payload.get("observations"), "observation_id", "upstream observations"
    )
    source_contexts = _unique_index(
        observations_payload.get("contexts"), "context_id", "upstream contexts"
    )
    source_rows = _unique_index(observations_payload.get("rows"), "row_id", "upstream rows")
    source_dimensions = _unique_index(
        observations_payload.get("dimensions"), "dimension_id", "upstream dimensions"
    )
    mapping_source_records = _records(
        mapping_payload.get("source_dispositions"), "mapping source dispositions"
    )
    mapping_source_dispositions = _unique_index(
        list(mapping_source_records),
        "source_object_id",
        "mapping source dispositions",
    )
    mapped_observation_ids = {
        str(record.get("source_object_id"))
        for record in mapping_source_records
        if record.get("source_object_type") == "OBSERVATION"
        and record.get("mapping_disposition") == "MAPPED_EXISTING_ITEM"
    }
    mapped_row_records: dict[str, Mapping[str, Any]] = {}
    for record in mapping_source_records:
        if (
            record.get("source_object_type") != "ROW"
            or record.get("mapping_disposition") != "MAPPED_EXISTING_ITEM"
        ):
            continue
        row_identifier = record.get("row_id")
        if (
            not isinstance(row_identifier, str)
            or not row_identifier
            or row_identifier in mapped_row_records
        ):
            raise NativeTMCanonicalExcelExportError(
                "mapped row source disposition identity is invalid"
            )
        mapped_row_records[row_identifier] = record
    required = {
        "observation_id",
        "row_id",
        "dimension_id",
        "context_id",
        "report_norm_id",
        "terminal_outcome",
        "reported_value",
        "unit",
        "unit_multiplier",
        "canonical_value",
        "period_type",
        "period_start",
        "period_end",
        "as_of_date",
        "presentation_scope",
        "match_basis",
        "source_record_sha256",
    }
    expected_observation_ids = {
        str(identifier)
        for disposition in disposition_by_id.values()
        for identifier in disposition.get("source_observation_ids", [])
    }
    expected_row_ids = {
        str(identifier)
        for disposition in disposition_by_id.values()
        for identifier in disposition.get("source_row_ids", [])
    }
    seen: set[str] = set()
    seen_rows: set[str] = set()
    producer_order_keys: list[tuple[int, int, str]] = []
    projected: list[dict[str, object]] = []
    for order, observation in enumerate(observations, start=1):
        if not required <= set(observation):
            raise NativeTMCanonicalExcelExportError("canonical observation fields are incomplete")
        observation_id = observation.get("observation_id")
        identifier = observation.get("report_norm_id")
        if (
            not isinstance(observation_id, str)
            or not observation_id
            or observation_id in seen
            or isinstance(identifier, bool)
            or not isinstance(identifier, int)
            or identifier not in schema_by_id
            or observation.get("terminal_outcome") not in {"OBSERVED_VALUE", "OBSERVED_ZERO"}
            or _SHA256.fullmatch(str(observation.get("source_record_sha256", ""))) is None
        ):
            raise NativeTMCanonicalExcelExportError("canonical observation identity is invalid")
        schema_item = schema_by_id[identifier]
        disposition = disposition_by_id[identifier]
        row_id = observation.get("row_id")
        dimension_id = observation.get("dimension_id")
        context_id = observation.get("context_id")
        source_observation = source_observations.get(str(observation_id))
        source_context = source_contexts.get(str(context_id))
        source_row = source_rows.get(str(row_id))
        source_dimension = source_dimensions.get(str(dimension_id))
        mapping_source = mapping_source_dispositions.get(str(observation_id))
        mapping_row = mapped_row_records.get(str(row_id))
        parsed = (
            source_observation.get("parsed") if isinstance(source_observation, Mapping) else None
        )
        period_materialization = (
            source_dimension.get("period_materialization")
            if isinstance(source_dimension, Mapping)
            else None
        )
        unit_materialization = (
            source_dimension.get("unit_materialization")
            if isinstance(source_dimension, Mapping)
            else None
        )
        source_status = (
            source_observation.get("source_status")
            if isinstance(source_observation, Mapping)
            else None
        )
        expected_outcome = {
            "OBSERVED_VALUE": "OBSERVED_VALUE",
            "OBSERVED_ZERO": "OBSERVED_ZERO",
        }.get(str(source_status))
        try:
            source_value = (
                Decimal(str(parsed["value"]))
                if isinstance(parsed, Mapping)
                and "value" in parsed
                and not isinstance(parsed.get("value"), bool)
                else None
            )
        except (InvalidOperation, ValueError):
            source_value = None
        try:
            source_value_text = Decimal(
                str(source_observation.get("value_text"))
                if isinstance(source_observation, Mapping)
                else "NaN"
            )
            normalized_value_text = (
                Decimal(str(parsed.get("normalized_text"))) if isinstance(parsed, Mapping) else None
            )
        except (InvalidOperation, ValueError, TypeError):
            source_value_text = None
            normalized_value_text = None
        multiplier = source_dimension.get("unit_multiplier") if source_dimension else None
        expected_canonical_value = (
            _decimal_text(source_value * Decimal(multiplier))
            if source_value is not None
            and source_value.is_finite()
            and isinstance(multiplier, int)
            and not isinstance(multiplier, bool)
            and multiplier > 0
            else None
        )
        expected_source_status = (
            "OBSERVED_ZERO"
            if source_value is not None and source_value.is_finite() and source_value == 0
            else (
                "OBSERVED_VALUE" if source_value is not None and source_value.is_finite() else None
            )
        )
        if (
            disposition.get("mapping_disposition") != "EXISTING_ITEM"
            or disposition.get("terminal_outcome") not in {"OBSERVED_VALUE", "OBSERVED_ZERO"}
            or row_id not in disposition.get("source_row_ids", [])
            or observation_id not in disposition.get("source_observation_ids", [])
            or source_observation is None
            or source_context is None
            or source_row is None
            or source_dimension is None
            or mapping_source is None
            or mapping_source.get("source_object_type") != "OBSERVATION"
            or mapping_source.get("mapping_disposition") != "MAPPED_EXISTING_ITEM"
            or mapping_source.get("target_report_norm_id") != identifier
            or mapping_source.get("row_id") != row_id
            or mapping_source.get("dimension_id") != dimension_id
            or mapping_source.get("context_id") != context_id
            or mapping_source.get("match_basis") != observation.get("match_basis")
            or mapping_row is None
            or mapping_row.get("target_report_norm_id") != identifier
            or mapping_row.get("context_id") != context_id
            or mapping_row.get("dimension_id") is not None
            or mapping_row.get("match_basis") != observation.get("match_basis")
            or mapping_row.get("matched_retrieval_key")
            != mapping_source.get("matched_retrieval_key")
            or expected_outcome != observation.get("terminal_outcome")
            or source_value is None
            or not source_value.is_finite()
            or source_status != expected_source_status
            or str(observation.get("reported_value")) != _decimal_text(source_value)
            or source_value_text != source_value
            or not isinstance(parsed, Mapping)
            or normalized_value_text != source_value
            or source_dimension.get("binding_status") != "RESOLVED"
            or not isinstance(period_materialization, Mapping)
            or period_materialization.get("resolution_status") != "SOURCE_BINDING_RESOLVED"
            or not isinstance(unit_materialization, Mapping)
            or unit_materialization.get("resolution_status") != "SOURCE_BINDING_RESOLVED"
            or observation.get("unit") != source_dimension.get("unit")
            or observation.get("unit_multiplier") != multiplier
            or observation.get("period_type") != source_dimension.get("period_type")
            or observation.get("period_start") != source_dimension.get("period_start")
            or observation.get("period_end") != source_dimension.get("period_end")
            or observation.get("as_of_date")
            != (
                source_dimension.get("period_end")
                if source_dimension.get("period_type") == "SNAPSHOT"
                else None
            )
            or observation.get("presentation_scope") != report_scope
            or str(observation.get("canonical_value")) != expected_canonical_value
            or source_observation.get("row_id") != row_id
            or source_observation.get("dimension_id") != dimension_id
            or source_observation.get("context_id") != context_id
            or source_row.get("context_id") != context_id
            or source_dimension.get("context_id") != context_id
            or row_id not in source_context.get("row_ids", [])
            or dimension_id not in source_context.get("dimension_ids", [])
            or observation_id not in source_context.get("observation_ids", [])
            or observation_id not in source_row.get("observation_ids", [])
            or len(
                {
                    source_context.get("page"),
                    source_row.get("page"),
                    source_dimension.get("page"),
                    source_observation.get("page"),
                }
            )
            != 1
            or len(
                {
                    source_context.get("source_table_id"),
                    source_row.get("source_table_id"),
                    source_dimension.get("source_table_id"),
                    source_observation.get("source_table_id"),
                }
            )
            != 1
            or source_context.get("page") is None
            or source_context.get("source_table_id") is None
            or observation.get("source_record_sha256")
            != source_observation.get(
                "source_slot_record_sha256", _record_sha256(source_observation)
            )
        ):
            raise NativeTMCanonicalExcelExportError(
                "canonical observation does not join its schema disposition"
            )
        display_order = int(schema_item["display_order"])
        axis_ordinal = source_dimension.get("axis_ordinal")
        if isinstance(axis_ordinal, bool) or not isinstance(axis_ordinal, int) or axis_ordinal < 0:
            raise NativeTMCanonicalExcelExportError(
                "canonical observation dimension axis ordinal is invalid"
            )
        producer_order_keys.append((display_order, axis_ordinal, observation_id))
        seen.add(observation_id)
        seen_rows.add(str(row_id))
        projected.append(
            dict(
                zip(
                    CANONICAL_OBSERVATION_HEADERS,
                    (
                        order,
                        observation_id,
                        identifier,
                        schema_item.get("canonical_name"),
                        schema_item.get("display_order"),
                        observation.get("terminal_outcome"),
                        str(observation.get("reported_value")),
                        str(observation.get("canonical_value")),
                        observation.get("unit"),
                        str(observation.get("unit_multiplier")),
                        observation.get("period_type"),
                        observation.get("period_start"),
                        observation.get("period_end"),
                        observation.get("as_of_date"),
                        observation.get("presentation_scope"),
                        row_id,
                        dimension_id,
                        context_id,
                        observation.get("match_basis"),
                        observation.get("source_record_sha256"),
                        source_row.get("page"),
                        source_row.get("label") or None,
                        _compact_json(source_row.get("label")),
                        mapping_source.get("matched_retrieval_key"),
                        _record_sha256(source_observation),
                        _record_sha256(source_row),
                        _record_sha256(source_dimension),
                        _compact_json(source_observation),
                        _compact_json(source_row),
                        _compact_json(source_dimension),
                        _record_sha256(observation),
                        _compact_json(observation),
                    ),
                    strict=True,
                )
            )
        )
    if (
        seen != expected_observation_ids
        or seen != mapped_observation_ids
        or seen_rows != expected_row_ids
        or seen_rows != set(mapped_row_records)
        or producer_order_keys != sorted(producer_order_keys)
    ):
        raise NativeTMCanonicalExcelExportError(
            "canonical observations do not exhaust mapped schema source citations"
        )
    return projected


def _unique_index(value: object, key: str, label: str) -> dict[str, Mapping[str, Any]]:
    indexed: dict[str, Mapping[str, Any]] = {}
    for record in _records(value, label):
        identifier = record.get(key)
        if not isinstance(identifier, str) or not identifier or identifier in indexed:
            raise NativeTMCanonicalExcelExportError(f"{label} identity is invalid")
        indexed[identifier] = record
    return indexed


def _chunks(value: object) -> tuple[str, list[str]]:
    rendered = _compact_json_unbounded(value)
    digest = _sha256(rendered.encode("utf-8"))
    parts = [
        rendered[index : index + _JSON_CHUNK_SIZE]
        for index in range(0, len(rendered), _JSON_CHUNK_SIZE)
    ]
    return digest, parts or [""]


def _validation_projection(
    mapping_payload: Mapping[str, Any],
) -> tuple[list[dict[str, object]], int]:
    root_assessments = _records(mapping_payload.get("root_assessments"), "root_assessments")
    accepted_subtrees = _records(mapping_payload.get("accepted_subtrees"), "accepted_subtrees")
    equation_checks = _records(mapping_payload.get("equation_checks"), "equation_checks")
    roots_by_id: dict[int, Mapping[str, Any]] = {}
    accepted_roots: dict[int, Mapping[str, Any]] = {}
    referenced_by_roots: dict[str, int] = {}
    for record in root_assessments:
        identifier = record.get("inferred_root_report_norm_id")
        if (
            isinstance(identifier, bool)
            or not isinstance(identifier, int)
            or identifier in roots_by_id
        ):
            raise NativeTMCanonicalExcelExportError("root-assessment identity is invalid")
        equation_ids = record.get("equation_check_ids")
        if (
            not isinstance(equation_ids, list)
            or any(not isinstance(item, str) or not item for item in equation_ids)
            or len(equation_ids) != len(set(equation_ids))
            or any(item in referenced_by_roots for item in equation_ids)
        ):
            raise NativeTMCanonicalExcelExportError(
                "root-assessment equation references are invalid"
            )
        roots_by_id[identifier] = record
        referenced_by_roots.update({item: identifier for item in equation_ids})
        if record.get("status") == "ACCEPTED":
            accepted_roots[identifier] = record
    subtrees_by_root: dict[int, Mapping[str, Any]] = {}
    for record in accepted_subtrees:
        identifier = record.get("inferred_root_report_norm_id")
        if (
            isinstance(identifier, bool)
            or not isinstance(identifier, int)
            or identifier in subtrees_by_root
            or record.get("local_completeness_status") != "COMPLETE_BOUNDED_TABLE_SUBTREE"
        ):
            raise NativeTMCanonicalExcelExportError("accepted subtree identity is invalid")
        subtrees_by_root[identifier] = record
    equation_by_id: dict[str, Mapping[str, Any]] = {}
    for record in equation_checks:
        identifier = record.get("equation_check_id")
        if (
            not isinstance(identifier, str)
            or not identifier
            or identifier in equation_by_id
            or record.get("status") not in {"EXACT", "MISMATCH"}
            or record.get("used_as_post_lineage_veto") is not True
            or record.get("used_for_target_selection") is not False
        ):
            raise NativeTMCanonicalExcelExportError("equation-check receipt is invalid")
        equation_by_id[identifier] = record
    if set(accepted_roots) != set(subtrees_by_root):
        raise NativeTMCanonicalExcelExportError("accepted root/subtree accounting drifted")
    for identifier, root in accepted_roots.items():
        subtree = subtrees_by_root[identifier]
        root_equations = root.get("equation_check_ids")
        subtree_equations = subtree.get("equation_check_ids")
        if (
            not isinstance(root_equations, list)
            or root_equations != subtree_equations
            or any(not isinstance(item, str) or not item for item in root_equations)
            or len(root_equations) != len(set(root_equations))
            or root.get("accepted_context_id") != subtree.get("context_id")
            or root.get("terminal_row_id") != subtree.get("terminal_row_id")
        ):
            raise NativeTMCanonicalExcelExportError("accepted root/subtree receipt drifted")
        for equation_id in root_equations:
            equation = equation_by_id.get(equation_id)
            if (
                equation is None
                or equation.get("status") != "EXACT"
                or equation.get("inferred_root_report_norm_id") != identifier
                or equation.get("context_id") != subtree.get("context_id")
            ):
                raise NativeTMCanonicalExcelExportError(
                    "equation check does not join its accepted subtree"
                )
    for equation_id, equation in equation_by_id.items():
        owner_id = referenced_by_roots.get(equation_id)
        if owner_id is None or equation.get("inferred_root_report_norm_id") != owner_id:
            raise NativeTMCanonicalExcelExportError(
                "equation check does not join its owning root assessment"
            )
    completion = _mapping(mapping_payload.get("completion"), "mapping completion")
    if (
        set(referenced_by_roots) != set(equation_by_id)
        or completion.get("accepted_root_count") != len(accepted_roots)
        or completion.get("source_accounting_complete") is not True
        or completion.get("tm_schema_disposition_accounting_complete") is not True
        or not isinstance(completion.get("document_complete"), bool)
    ):
        raise NativeTMCanonicalExcelExportError("mapping validation accounting drifted")

    projected: list[dict[str, object]] = []
    logical_count = 0
    for record_type, records in (
        ("ROOT_ASSESSMENT", root_assessments),
        ("ACCEPTED_SUBTREE", accepted_subtrees),
        ("EQUATION_CHECK", equation_checks),
    ):
        for index, record in enumerate(records, start=1):
            logical_count += 1
            digest, parts = _chunks(record)
            status = record.get("status") or record.get("local_completeness_status")
            report_norm_id = record.get("inferred_root_report_norm_id")
            context_id = record.get("accepted_context_id") or record.get("context_id")
            for part_number, part in enumerate(parts, start=1):
                projected.append(
                    dict(
                        zip(
                            VALIDATION_HEADERS,
                            (
                                len(projected) + 1,
                                record_type,
                                index,
                                status,
                                report_norm_id,
                                context_id,
                                part_number,
                                len(parts),
                                digest,
                                part,
                            ),
                            strict=True,
                        )
                    )
                )
    return projected, logical_count


def _metadata_projection(values: Mapping[str, object]) -> list[dict[str, object]]:
    projected: list[dict[str, object]] = []
    for key in sorted(values):
        digest, parts = _chunks(values[key])
        for part_number, part in enumerate(parts, start=1):
            projected.append(
                dict(
                    zip(
                        RUN_METADATA_HEADERS,
                        (key, part_number, len(parts), digest, part),
                        strict=True,
                    )
                )
            )
    return projected


def _identity_record(identity: ArtifactIdentity, payload: Mapping[str, Any]) -> dict[str, object]:
    return {
        "path": identity.path,
        "sha256": identity.sha256,
        "size_bytes": identity.size_bytes,
        "format_version": payload.get("format_version"),
        "policy": payload.get("policy"),
        "claim_boundary": payload.get("claim_boundary"),
        "status": payload.get("status"),
        "run_id": payload.get("run_id"),
    }


def _transitive_identity_records(
    identities: Sequence[ArtifactIdentity],
) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    kinds = (
        "REGISTERED_NATIVE_TM_DOCUMENT_ARTIFACT",
        "SOURCE_PDF",
        "ACCEPTED_STATEMENT_DISCOVERY",
    )
    if len(identities) != len(kinds):
        raise NativeTMCanonicalExcelExportError("transitive input identity count drifted")
    for kind, identity in zip(kinds, identities, strict=True):
        _validate_identity(identity, kind)
        result.append(
            {
                "kind": kind,
                "path": identity.path,
                "sha256": identity.sha256,
                "size_bytes": identity.size_bytes,
            }
        )
    return result


def _content_ledger(
    mapping_payload: Mapping[str, Any],
    observations_payload: Mapping[str, Any],
) -> dict[str, object]:
    mapping_sections = {}
    for key in (
        "canonical_observations",
        "schema_dispositions",
        "source_dispositions",
        "root_assessments",
        "accepted_subtrees",
        "equation_checks",
        "coverage",
        "completion",
        "producer_snapshots",
    ):
        value = mapping_payload.get(key)
        mapping_sections[key] = {
            "record_count": len(value) if isinstance(value, list) else None,
            "sha256": _record_sha256(value),
        }
    source_objects, upstream_by_id = _source_object_index(observations_payload)
    source_object_projection = [
        {
            "source_object_id": identifier,
            "source_object_type": upstream_by_id[identifier]["source_object_type"],
            "source_object_sha256": _record_sha256(source_objects[identifier]),
        }
        for identifier in sorted(
            source_objects,
            key=lambda item: (
                str(upstream_by_id[item]["source_object_type"]),
                item,
            ),
        )
    ]
    observation_sections: dict[str, object] = {}
    for key in (
        "page_inventory",
        "contexts",
        "rows",
        "dimensions",
        "observations",
        "source_evidence",
        "source_dispositions",
        "source_accounting",
    ):
        value = observations_payload.get(key)
        if isinstance(value, list):
            top_level_count: int | None = len(value)
            payload_kind = "RECORD_LIST"
        elif key == "source_evidence" and isinstance(value, Mapping):
            top_level_count = sum(len(item) for item in value.values() if isinstance(item, list))
            payload_kind = "RECORD_LIST_MAPPING"
        elif isinstance(value, Mapping):
            top_level_count = None
            payload_kind = "OBJECT"
        else:
            top_level_count = None
            payload_kind = "INVALID"
        observation_sections[key] = {
            "payload_kind": payload_kind,
            "top_level_record_count": top_level_count,
            "sha256": _record_sha256(value),
        }
    return {
        "mapping_sections": mapping_sections,
        "observations_sections": observation_sections,
        "reconstructed_source_objects": {
            "record_count": len(source_object_projection),
            "projection_sha256": _record_sha256(source_object_projection),
        },
    }


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        header = str(sheet.cell(1, column).value or "")
        width = 18
        if header.endswith("Json") or header.endswith("JsonPart"):
            width = 52
        elif "Name" in header or "Reason" in header:
            width = 34
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_records(
    workbook: Workbook,
    name: str,
    headers: Sequence[str],
    records: Sequence[Mapping[str, object]],
) -> None:
    sheet = workbook.create_sheet(name)
    append_literal_row(sheet, headers)
    for record in records:
        if set(record) != set(headers):
            raise NativeTMCanonicalExcelExportError(f"{name} projection fields drifted")
        append_literal_row(sheet, [record[header] for header in headers])
    _style_sheet(sheet)


def _verify_chunk_records(
    records: Sequence[Mapping[str, object]],
    *,
    key_field: str,
    index_field: str | None,
    part_field: str,
    count_field: str,
    hash_field: str,
    value_field: str,
) -> None:
    grouped: dict[tuple[object, ...], list[Mapping[str, object]]] = {}
    for record in records:
        key = (
            (record[key_field],)
            if index_field is None
            else (record[key_field], record[index_field])
        )
        grouped.setdefault(key, []).append(record)
    for parts in grouped.values():
        parts.sort(key=lambda item: int(item[part_field]))
        count = len(parts)
        if (
            [item[part_field] for item in parts] != list(range(1, count + 1))
            or any(item[count_field] != count for item in parts)
            or len({item[hash_field] for item in parts}) != 1
        ):
            raise NativeTMCanonicalExcelExportError("chunked workbook record is malformed")
        rendered = "".join(str(item[value_field]) for item in parts)
        if _sha256(rendered.encode("utf-8")) != parts[0][hash_field]:
            raise NativeTMCanonicalExcelExportError("chunked workbook record hash drifted")
        try:
            json.loads(rendered)
        except json.JSONDecodeError as exc:
            raise NativeTMCanonicalExcelExportError("chunked workbook record is not JSON") from exc


def _verify_workbook(
    payload: bytes,
    *,
    creator: str,
    records_by_sheet: Mapping[str, Sequence[Mapping[str, object]]],
) -> None:
    try:
        workbook = load_workbook(BytesIO(payload), read_only=False, data_only=False)
    except Exception as exc:
        raise NativeTMCanonicalExcelExportError("canonical workbook cannot be reopened") from exc
    try:
        if tuple(workbook.sheetnames) != SHEET_NAMES or workbook_has_formula(workbook):
            raise NativeTMCanonicalExcelExportError(
                "canonical workbook sheet or formula contract drifted"
            )
        headers_by_sheet = {
            "CANONICAL_OBSERVATIONS": CANONICAL_OBSERVATION_HEADERS,
            "SCHEMA_DISPOSITIONS": SCHEMA_DISPOSITION_HEADERS,
            "SOURCE_DISPOSITIONS": SOURCE_DISPOSITION_HEADERS,
            "SOURCE_OBJECTS": SOURCE_OBJECT_HEADERS,
            "VALIDATION": VALIDATION_HEADERS,
            "RUN_METADATA": RUN_METADATA_HEADERS,
        }
        for name in SHEET_NAMES:
            sheet = workbook[name]
            headers = headers_by_sheet[name]
            if tuple(cell.value for cell in sheet[1]) != tuple(headers):
                raise NativeTMCanonicalExcelExportError(f"{name} headers drifted")
            records = records_by_sheet[name]
            if sheet.max_row != len(records) + 1:
                raise NativeTMCanonicalExcelExportError(f"{name} row count drifted")
            for row_number, expected in enumerate(records, start=2):
                actual = tuple(
                    sheet.cell(row_number, column).value for column in range(1, len(headers) + 1)
                )
                wanted = tuple(expected[header] for header in headers)
                if actual != wanted:
                    raise NativeTMCanonicalExcelExportError(f"{name} content drifted")
        _verify_chunk_records(
            records_by_sheet["SOURCE_OBJECTS"],
            key_field="SourceObjectId",
            index_field=None,
            part_field="PartNumber",
            count_field="PartCount",
            hash_field="SourceObjectSha256",
            value_field="SourceObjectJsonPart",
        )
        _verify_chunk_records(
            records_by_sheet["VALIDATION"],
            key_field="RecordType",
            index_field="RecordIndex",
            part_field="PartNumber",
            count_field="PartCount",
            hash_field="RecordSha256",
            value_field="RecordJsonPart",
        )
        _verify_chunk_records(
            records_by_sheet["RUN_METADATA"],
            key_field="Key",
            index_field=None,
            part_field="PartNumber",
            count_field="PartCount",
            hash_field="ValueSha256",
            value_field="ValueJsonPart",
        )
        properties = workbook.properties
        if (
            properties.creator != creator
            or properties.lastModifiedBy != creator
            or properties.created != CANONICAL_CORE_TIMESTAMP
            or properties.modified != CANONICAL_CORE_TIMESTAMP
            or properties.version != "1"
            or properties.revision != "1"
        ):
            raise NativeTMCanonicalExcelExportError("canonical workbook metadata drifted")
    finally:
        workbook.close()


def _validate_implementation_ledger(
    value: Sequence[Mapping[str, object]],
) -> tuple[dict[str, object], ...]:
    if isinstance(value, (str, bytes, bytearray)):
        raise NativeTMCanonicalExcelExportError("export implementation ledger is invalid")
    records: list[dict[str, object]] = []
    for record in value:
        if not isinstance(record, Mapping) or set(record) != {"path", "sha256", "size_bytes"}:
            raise NativeTMCanonicalExcelExportError("export implementation identity is malformed")
        path = _canonical_relative_path(record.get("path"), "implementation path")
        digest = record.get("sha256")
        size = record.get("size_bytes")
        if (
            not isinstance(digest, str)
            or _SHA256.fullmatch(digest) is None
            or isinstance(size, bool)
            or not isinstance(size, int)
            or size < 0
        ):
            raise NativeTMCanonicalExcelExportError("export implementation identity is invalid")
        records.append({"path": path, "sha256": digest, "size_bytes": size})
    if [record["path"] for record in records] != sorted({record["path"] for record in records}):
        raise NativeTMCanonicalExcelExportError(
            "export implementation ledger must be unique and sorted"
        )
    return tuple(records)


def _git(project_root: Path, *arguments: str, binary: bool = False) -> str | bytes:
    try:
        result = subprocess.run(
            ["git", *arguments],
            cwd=project_root,
            check=True,
            capture_output=True,
            text=not binary,
        )
    except (OSError, subprocess.CalledProcessError) as exc:
        raise NativeTMCanonicalExcelExportError(
            "cannot inspect native-TM Excel producer Git state"
        ) from exc
    return result.stdout


def _current_git_state(project_root: Path) -> dict[str, object]:
    commit = str(_git(project_root, "rev-parse", "HEAD")).strip()
    dirty = bool(str(_git(project_root, "status", "--porcelain", "--untracked-files=all")).strip())
    if _GIT_COMMIT.fullmatch(commit) is None or dirty:
        raise NativeTMCanonicalExcelExportError(
            "native-TM Excel producer requires a clean Git HEAD"
        )
    return {"commit": commit, "dirty": False}


def _git_file_bytes(project_root: Path, commit: str, relative: str) -> bytes:
    if not isinstance(commit, str) or _GIT_COMMIT.fullmatch(commit) is None:
        raise NativeTMCanonicalExcelExportError("exporter producer commit is invalid")
    relative = _canonical_relative_path(relative, "producer implementation path")
    tree_entry = _git(
        project_root,
        "ls-tree",
        "-z",
        commit,
        "--",
        relative,
        binary=True,
    )
    if not isinstance(tree_entry, bytes):
        raise NativeTMCanonicalExcelExportError("producer Git tree read was not binary")
    entries = [item for item in tree_entry.split(b"\x00") if item]
    if len(entries) != 1 or b"\t" not in entries[0]:
        raise NativeTMCanonicalExcelExportError("producer Git file is not uniquely tracked")
    metadata, raw_path = entries[0].split(b"\t", 1)
    fields = metadata.split()
    if (
        len(fields) != 3
        or fields[0] not in {b"100644", b"100755"}
        or fields[1] != b"blob"
        or raw_path != relative.encode("utf-8")
    ):
        raise NativeTMCanonicalExcelExportError(
            "producer Git implementation is not a regular tracked blob"
        )
    payload = _git(project_root, "show", f"{commit}:{relative}", binary=True)
    if not isinstance(payload, bytes):
        raise NativeTMCanonicalExcelExportError("producer Git file read was not binary")
    return payload


def _implementation_ledger_at_commit(
    project_root: Path, commit: str
) -> tuple[dict[str, object], ...]:
    return tuple(
        {
            "path": relative,
            "sha256": _sha256(payload),
            "size_bytes": len(payload),
        }
        for relative in sorted(_IMPLEMENTATION_PATHS)
        for payload in (_git_file_bytes(project_root, commit, relative),)
    )


def _implementation_ledger(project_root: Path) -> tuple[dict[str, object], ...]:
    records: list[dict[str, object]] = []
    for relative in _IMPLEMENTATION_PATHS:
        guard = _open_held_file(
            project_root,
            Path(relative),
            f"export implementation {relative}",
        )
        try:
            records.append(
                {
                    "path": relative,
                    "sha256": _sha256(guard.payload),
                    "size_bytes": len(guard.payload),
                }
            )
            _assert_held_stable(guard, f"export implementation {relative}")
        finally:
            _close_held(guard)
    records.sort(key=lambda item: str(item["path"]))
    return tuple(records)


def _current_exporter_producer(project_root: Path) -> dict[str, object]:
    state = _current_git_state(project_root)
    commit = str(state["commit"])
    current = _implementation_ledger(project_root)
    committed = _implementation_ledger_at_commit(project_root, commit)
    if current != committed:
        raise NativeTMCanonicalExcelExportError(
            "native-TM Excel implementation differs from clean producer commit"
        )
    return {
        "commit": commit,
        "dirty": False,
        "implementation": [dict(record) for record in committed],
    }


def _validate_exporter_producer(
    value: object,
    *,
    project_root: Path | None = None,
) -> dict[str, object]:
    producer = _mapping(value, "exporter producer")
    if set(producer) != {"commit", "dirty", "implementation"}:
        raise NativeTMCanonicalExcelExportError("exporter producer fields drifted")
    commit = producer.get("commit")
    if (
        not isinstance(commit, str)
        or _GIT_COMMIT.fullmatch(commit) is None
        or producer.get("dirty") is not False
    ):
        raise NativeTMCanonicalExcelExportError("exporter producer identity is invalid")
    implementation = _validate_implementation_ledger(
        _records(producer.get("implementation"), "exporter implementation ledger")
    )
    normalized = {
        "commit": commit,
        "dirty": False,
        "implementation": [dict(record) for record in implementation],
    }
    if project_root is not None:
        committed = _implementation_ledger_at_commit(project_root, commit)
        if implementation != committed:
            raise NativeTMCanonicalExcelExportError(
                "exporter implementation ledger differs from producer commit"
            )
    return normalized


def _preflight_committed_exporter_producer(
    value: object,
    *,
    project_root: Path,
) -> dict[str, object]:
    """Validate receipt code against Git and the frozen V1 manifest."""

    producer = _mapping(value, "exporter producer")
    if set(producer) != {"commit", "dirty", "implementation"}:
        raise NativeTMCanonicalExcelExportError("exporter producer fields drifted")
    commit = producer.get("commit")
    if (
        not isinstance(commit, str)
        or _GIT_COMMIT.fullmatch(commit) is None
        or producer.get("dirty") is not False
    ):
        raise NativeTMCanonicalExcelExportError("exporter producer identity is invalid")
    implementation = _validate_implementation_ledger(
        _records(producer.get("implementation"), "exporter implementation ledger")
    )
    if tuple(str(record["path"]) for record in implementation) != tuple(
        sorted(_IMPLEMENTATION_PATHS)
    ):
        raise NativeTMCanonicalExcelExportError(
            "exporter producer ledger differs from the frozen V1 manifest"
        )
    _git(project_root, "merge-base", "--is-ancestor", commit, "HEAD")
    for record in implementation:
        payload = _git_file_bytes(project_root, commit, str(record["path"]))
        if _sha256(payload) != record["sha256"] or len(payload) != record["size_bytes"]:
            raise NativeTMCanonicalExcelExportError(
                "exporter implementation ledger differs from producer commit"
            )
    return {
        "commit": commit,
        "dirty": False,
        "implementation": [dict(record) for record in implementation],
    }


_PRODUCER_REPLAY_BOOTSTRAP = r"""
import json
import os
import pathlib
import stat
import struct
import subprocess
import sys

source_tree = pathlib.Path(sys.argv[1]).resolve()
repository = pathlib.Path(sys.argv[2]).resolve()
expected_commit = sys.argv[3]
head = subprocess.run(
    ["git", "rev-parse", "HEAD"], cwd=source_tree, check=True, capture_output=True, text=True
).stdout.strip()
dirty = subprocess.run(
    ["git", "status", "--porcelain", "--untracked-files=all"],
    cwd=source_tree,
    check=True,
    capture_output=True,
).stdout.strip()
if head != expected_commit or dirty:
    raise RuntimeError("exporter replay source tree is not the clean producer commit")
expected_module = source_tree / "src/bctc_ai/export/native_tm_canonical_excel.py"
expected_module.relative_to(source_tree)
cursor = source_tree
for part in expected_module.relative_to(source_tree).parts:
    cursor = cursor / part
    identity = os.lstat(cursor)
    if stat.S_ISLNK(identity.st_mode):
        raise RuntimeError("exporter replay module path contains a symlink")
if not stat.S_ISREG(os.lstat(expected_module).st_mode):
    raise RuntimeError("exporter replay module is not a regular file")
sys.path.insert(0, str(source_tree / "src"))
from bctc_ai.export import native_tm_canonical_excel as producer

if pathlib.Path(producer.__file__).resolve() != expected_module.resolve():
    raise RuntimeError("exporter replay imported outside its isolated producer tree")
request = json.loads(sys.stdin.buffer.read())
artifacts = producer._rebuild_pair_at_producer_commit(
    project_root=repository,
    mapping_relative_path=request["mapping_relative_path"],
    mapping_expected_sha256=request["mapping_expected_sha256"],
    workbook_relative_path=request["workbook_relative_path"],
    provenance_relative_path=request["provenance_relative_path"],
    expected_exporter_producer=request["exporter_producer"],
)
head = subprocess.run(
    ["git", "rev-parse", "HEAD"], cwd=source_tree, check=True, capture_output=True, text=True
).stdout.strip()
dirty = subprocess.run(
    ["git", "status", "--porcelain", "--untracked-files=all"],
    cwd=source_tree,
    check=True,
    capture_output=True,
).stdout.strip()
if head != expected_commit or dirty:
    raise RuntimeError("exporter replay source tree changed during rebuild")
sys.stdout.buffer.write(producer._frame_replayed_pair(
    artifacts.workbook_bytes, artifacts.provenance_bytes
))
"""


def _isolated_subprocess_environment() -> dict[str, str]:
    return {
        "PATH": os.environ.get("PATH", "/usr/bin:/bin"),
        "LANG": "C.UTF-8",
        "LC_ALL": "C.UTF-8",
        "GIT_CONFIG_NOSYSTEM": "1",
        "GIT_CONFIG_GLOBAL": os.devnull,
    }


def _run_checked_process(
    arguments: Sequence[str],
    *,
    cwd: Path,
    environment: Mapping[str, str],
    stdin: bytes | None = None,
    timeout: int = 1800,
) -> bytes:
    try:
        result = subprocess.run(
            list(arguments),
            cwd=cwd,
            env=dict(environment),
            input=stdin,
            capture_output=True,
            check=False,
            timeout=timeout,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise NativeTMCanonicalExcelExportError(
            "native-TM Excel producer-commit replay process failed"
        ) from exc
    if result.returncode != 0:
        raise NativeTMCanonicalExcelExportError(
            "native-TM Excel producer-commit replay process failed"
        )
    return result.stdout


def _frame_replayed_pair(workbook_bytes: bytes, provenance_bytes: bytes) -> bytes:
    if not workbook_bytes or not provenance_bytes:
        raise NativeTMCanonicalExcelExportError("producer replay returned an empty pair")
    return (
        _REPLAY_MAGIC
        + struct.pack(">QQ", len(workbook_bytes), len(provenance_bytes))
        + workbook_bytes
        + provenance_bytes
    )


def _parse_replayed_pair(payload: bytes) -> tuple[bytes, bytes]:
    header_size = len(_REPLAY_MAGIC) + 16
    if len(payload) < header_size or not payload.startswith(_REPLAY_MAGIC):
        raise NativeTMCanonicalExcelExportError("producer replay pair frame is invalid")
    workbook_size, provenance_size = struct.unpack(">QQ", payload[len(_REPLAY_MAGIC) : header_size])
    if workbook_size < 1 or provenance_size < 1:
        raise NativeTMCanonicalExcelExportError("producer replay pair frame is empty")
    if len(payload) != header_size + workbook_size + provenance_size:
        raise NativeTMCanonicalExcelExportError("producer replay pair frame length drifted")
    workbook_end = header_size + workbook_size
    return payload[header_size:workbook_end], payload[workbook_end:]


def _rebuild_pair_at_producer_commit(
    *,
    project_root: Path,
    mapping_relative_path: str,
    mapping_expected_sha256: str,
    workbook_relative_path: str,
    provenance_relative_path: str,
    expected_exporter_producer: Mapping[str, object],
) -> NativeTMCanonicalExcelArtifacts:
    """Historical entry point executed only from a detached producer checkout."""

    project_root = project_root.resolve()
    producer = _validate_exporter_producer(expected_exporter_producer, project_root=project_root)
    commit = str(producer["commit"])
    policy_bytes = _git_file_bytes(project_root, commit, EXPORT_POLICY_RELATIVE_PATH.as_posix())
    policy = _policy_from_bytes(
        policy_bytes,
        path=project_root.joinpath(*EXPORT_POLICY_RELATIVE_PATH.parts),
    )
    bundle = _strict_load_input_bundle(
        project_root=project_root,
        mapping_path=Path(mapping_relative_path),
        mapping_expected_sha256=mapping_expected_sha256,
        policy=policy,
    )
    try:
        mapping_identity = ArtifactIdentity(
            path=bundle.mapping.relative_path,
            sha256=mapping_expected_sha256,
            size_bytes=len(bundle.mapping.payload),
        )
        observations_identity = ArtifactIdentity(
            path=bundle.observations.relative_path,
            sha256=_sha256(bundle.observations.payload),
            size_bytes=len(bundle.observations.payload),
        )
        artifacts = _build_prevalidated_native_tm_canonical_excel_artifacts(
            bundle.mapping_payload,
            bundle.observations_payload,
            mapping_identity=mapping_identity,
            observations_identity=observations_identity,
            transitive_identities=bundle.transitive_identities,
            workbook_relative_path=workbook_relative_path,
            provenance_relative_path=provenance_relative_path,
            policy=policy,
            exporter_producer=producer,
        )
        _assert_bundle_stable(bundle)
        return artifacts
    finally:
        _close_input_bundle(bundle)


def _producer_commit_replay(
    *,
    project_root: Path,
    exporter_producer: Mapping[str, object],
    mapping_identity: ArtifactIdentity,
    workbook_relative_path: str,
    provenance_relative_path: str,
) -> tuple[bytes, bytes]:
    producer = _preflight_committed_exporter_producer(exporter_producer, project_root=project_root)
    commit = str(producer["commit"])
    temporary_root = Path(tempfile.mkdtemp(prefix="native-tm-excel-replay-"))
    environment = _isolated_subprocess_environment()
    try:
        clone_root = temporary_root / "producer"
        _run_checked_process(
            (
                "git",
                "clone",
                "--quiet",
                "--no-checkout",
                "--no-hardlinks",
                "--",
                str(project_root),
                str(clone_root),
            ),
            cwd=temporary_root,
            environment=environment,
        )
        _run_checked_process(
            ("git", "checkout", "--quiet", "--detach", commit),
            cwd=clone_root,
            environment=environment,
        )
        request = _canonical_json_bytes(
            {
                "exporter_producer": producer,
                "mapping_expected_sha256": mapping_identity.sha256,
                "mapping_relative_path": mapping_identity.path,
                "provenance_relative_path": provenance_relative_path,
                "workbook_relative_path": workbook_relative_path,
            }
        )
        framed = _run_checked_process(
            (
                sys.executable,
                "-I",
                "-B",
                "-c",
                _PRODUCER_REPLAY_BOOTSTRAP,
                str(clone_root),
                str(project_root),
                commit,
            ),
            cwd=temporary_root,
            environment=environment,
            stdin=request,
        )
        return _parse_replayed_pair(framed)
    finally:
        shutil.rmtree(temporary_root, ignore_errors=True)


def _receipt_isolation() -> dict[str, object]:
    return {
        "current_mutable_schema_loaded": False,
        "fd_held_nofollow_stability_guards": True,
        "historical_values_loaded": False,
        "human_review_outputs_loaded": False,
        "observations_are_receipt_bound_not_public_input": True,
        "producer_snapshots_replayed_without_current_schema": True,
        "role_a_outputs_loaded": False,
        "template_inputs_loaded": False,
    }


def _receipt_policy(policy: NativeTMCanonicalExcelPolicy) -> dict[str, object]:
    return {
        "export_policy": policy.name,
        "export_policy_sha256": policy.sha256,
        "force_mapping": False,
        "imputation": False,
        "producer_schema_snapshot_authoritative": True,
    }


def _build_prevalidated_native_tm_canonical_excel_artifacts(
    mapping_payload: Mapping[str, Any],
    observations_payload: Mapping[str, Any],
    *,
    mapping_identity: ArtifactIdentity,
    observations_identity: ArtifactIdentity,
    transitive_identities: Sequence[ArtifactIdentity],
    workbook_relative_path: str,
    provenance_relative_path: str,
    policy: NativeTMCanonicalExcelPolicy,
    exporter_producer: Mapping[str, object],
) -> NativeTMCanonicalExcelArtifacts:
    """Build bytes only from strict-loaded payloads and embedded producer snapshots."""

    source = _validate_input_pair(
        mapping_payload,
        observations_payload,
        mapping_identity=mapping_identity,
        observations_identity=observations_identity,
        policy=policy,
    )
    producer = _validate_exporter_producer(exporter_producer)
    workbook_relative_path = _canonical_relative_path(workbook_relative_path, "workbook path")
    provenance_relative_path = _canonical_relative_path(provenance_relative_path, "provenance path")
    workbook_location = PurePosixPath(workbook_relative_path)
    provenance_location = PurePosixPath(provenance_relative_path)
    if (
        workbook_location.suffix.casefold() != ".xlsx"
        or provenance_location.suffix.casefold() != ".json"
        or workbook_location.parent != provenance_location.parent
        or workbook_location == provenance_location
    ):
        raise NativeTMCanonicalExcelExportError(
            "workbook and provenance must be distinct .xlsx/.json siblings"
        )
    for location in (workbook_location, provenance_location):
        try:
            location.relative_to(PurePosixPath(policy.output_directory))
        except ValueError as exc:
            raise NativeTMCanonicalExcelExportError(
                "export output leaves configured output directory"
            ) from exc
    transitive_records = _transitive_identity_records(transitive_identities)
    schema_records, schema_by_id, disposition_by_id = _schema_projection(
        mapping_payload, policy=policy
    )
    source_records, source_object_records = _source_projection(
        mapping_payload, observations_payload, policy=policy
    )
    observation_records = _canonical_observation_projection(
        mapping_payload,
        observations_payload,
        schema_by_id=schema_by_id,
        disposition_by_id=disposition_by_id,
    )
    validation_records, validation_logical_count = _validation_projection(mapping_payload)
    terminal_counts = Counter(record["TerminalOutcome"] for record in schema_records)
    mapping_disposition_counts = Counter(record["MappingDisposition"] for record in source_records)
    summary: dict[str, object] = {
        "canonical_observation_count": len(observation_records),
        "tm_schema_disposition_count": len(schema_records),
        "source_object_disposition_count": len(source_records),
        "source_object_logical_count": len(source_records),
        "source_object_sheet_row_count": len(source_object_records),
        "schema_terminal_outcome_counts": {
            outcome: terminal_counts[outcome] for outcome in _TERMINAL_OUTCOMES
        },
        "source_mapping_disposition_counts": dict(sorted(mapping_disposition_counts.items())),
        "validation_logical_record_count": validation_logical_count,
        "validation_sheet_row_count": len(validation_records),
        "formula_count": 0,
        "imputed_value_count": 0,
        "producer_snapshot_schema_item_count": len(schema_by_id),
    }
    declared_source_count = _mapping(
        mapping_payload.get("source_accounting"), "source accounting"
    ).get("mapping_source_disposition_count")
    if (
        summary["tm_schema_disposition_count"] != policy.schema_disposition_count
        or summary["source_object_disposition_count"] != declared_source_count
        or summary["source_object_logical_count"] != declared_source_count
    ):
        raise NativeTMCanonicalExcelExportError("export accounting denominator drifted")
    content_ledger = _content_ledger(mapping_payload, observations_payload)
    metadata_values = {
        "ACCEPTED_SUBTREES": mapping_payload.get("accepted_subtrees"),
        "COMPLETION": mapping_payload.get("completion"),
        "CONTENT_LEDGER": content_ledger,
        "COVERAGE": mapping_payload.get("coverage"),
        "EQUATION_CHECKS": mapping_payload.get("equation_checks"),
        "EXPORT_SUMMARY": summary,
        "INPUT_IDENTITIES": {
            "mapping": _identity_record(mapping_identity, mapping_payload),
            "native_tm_observations": _identity_record(observations_identity, observations_payload),
            "transitive_replay_inputs": transitive_records,
        },
        "MAPPING_AUTHORITY": mapping_payload.get("authority"),
        "MAPPING_CODE": mapping_payload.get("code"),
        "MAPPING_INPUTS": mapping_payload.get("inputs"),
        "MAPPING_SCHEMA_IDENTITY": mapping_payload.get("schema"),
        "PRODUCER_SNAPSHOTS": mapping_payload.get("producer_snapshots"),
        "ROOT_ASSESSMENTS": mapping_payload.get("root_assessments"),
        "ROUTING_CONTRACT": mapping_payload.get("routing_contract"),
        "SOURCE": source,
    }
    metadata_records = _metadata_projection(metadata_values)
    records_by_sheet: dict[str, Sequence[Mapping[str, object]]] = {
        "CANONICAL_OBSERVATIONS": observation_records,
        "SCHEMA_DISPOSITIONS": schema_records,
        "SOURCE_DISPOSITIONS": source_records,
        "SOURCE_OBJECTS": source_object_records,
        "VALIDATION": validation_records,
        "RUN_METADATA": metadata_records,
    }
    workbook = Workbook()
    workbook.remove(workbook.active)
    try:
        for name, headers in (
            ("CANONICAL_OBSERVATIONS", CANONICAL_OBSERVATION_HEADERS),
            ("SCHEMA_DISPOSITIONS", SCHEMA_DISPOSITION_HEADERS),
            ("SOURCE_DISPOSITIONS", SOURCE_DISPOSITION_HEADERS),
            ("SOURCE_OBJECTS", SOURCE_OBJECT_HEADERS),
            ("VALIDATION", VALIDATION_HEADERS),
            ("RUN_METADATA", RUN_METADATA_HEADERS),
        ):
            _write_records(workbook, name, headers, records_by_sheet[name])
        if workbook_has_formula(workbook):
            raise NativeTMCanonicalExcelExportError("formulas are forbidden in workbook")
        workbook_bytes = deterministic_workbook_bytes(workbook, creator=policy.workbook_creator)
    finally:
        workbook.close()
    _verify_workbook(
        workbook_bytes,
        creator=policy.workbook_creator,
        records_by_sheet=records_by_sheet,
    )
    workbook_sha256 = _sha256(workbook_bytes)
    receipt = {
        "artifact_type": _RECEIPT_TYPE,
        "claim_boundary": policy.claim_boundary,
        "code": {
            "commit": producer["commit"],
            "dirty": False,
            "implementation": copy.deepcopy(producer["implementation"]),
            "mapping_producer": copy.deepcopy(mapping_payload.get("code")),
            "observations_producer": copy.deepcopy(observations_payload.get("code")),
        },
        "content_ledger": content_ledger,
        "dataset_role": policy.dataset_role,
        "format_version": 1,
        "inputs": {
            "mapping": _identity_record(mapping_identity, mapping_payload),
            "native_tm_observations": _identity_record(observations_identity, observations_payload),
            "transitive_replay_inputs": transitive_records,
            "exact_receipt_lineage_verified": True,
        },
        "isolation": _receipt_isolation(),
        "policy": _receipt_policy(policy),
        "provenance": {
            "filename": provenance_location.name,
            "path": provenance_relative_path,
        },
        "run_id": mapping_payload.get("run_id"),
        "source": copy.deepcopy(source),
        "status": _EXPORT_STATUS,
        "summary": summary,
        "workbook": {
            "filename": workbook_location.name,
            "formula_count": 0,
            "path": workbook_relative_path,
            "sha256": workbook_sha256,
            "sheet_names": list(SHEET_NAMES),
            "size_bytes": len(workbook_bytes),
        },
    }
    provenance_bytes = _canonical_json_bytes(receipt)
    return NativeTMCanonicalExcelArtifacts(
        workbook_bytes=workbook_bytes,
        provenance_bytes=provenance_bytes,
        workbook_sha256=workbook_sha256,
        provenance_sha256=_sha256(provenance_bytes),
        summary=summary,
    )


@dataclass(slots=True)
class _PairPublicationGuard:
    workbook: Any
    provenance: Any


def _rollback_pair(guard: _PairPublicationGuard, cause: BaseException) -> None:
    errors: list[BaseException] = []
    for item in (guard.provenance, guard.workbook):
        try:
            _publication._rollback_publication(item, cause)
        except BaseException as exc:
            errors.append(exc)
    if errors:
        raise NativeTMCanonicalExcelExportError(
            "native-TM Excel publication rollback was incomplete: "
            + "; ".join(str(item) for item in errors)
        ) from cause


def _close_pair(guard: _PairPublicationGuard) -> None:
    _publication._close_guard_best_effort(guard.provenance)
    _publication._close_guard_best_effort(guard.workbook)


def _revalidate_published_pair(
    project_root: Path,
    guard: _PairPublicationGuard,
    workbook_bytes: bytes,
    provenance_bytes: bytes,
) -> None:
    try:
        parents = [
            os.fstat(guard.workbook.parent_descriptor),
            os.fstat(guard.provenance.parent_descriptor),
        ]
        if (parents[0].st_dev, parents[0].st_ino) != (
            parents[1].st_dev,
            parents[1].st_ino,
        ):
            raise NativeTMCanonicalExcelExportError(
                "published native-TM Excel pair parent identity drifted"
            )
        output_inodes: list[tuple[int, int]] = []
        for item, expected in (
            (guard.workbook, workbook_bytes),
            (guard.provenance, provenance_bytes),
        ):
            descriptor_identity = os.fstat(item.output_descriptor)
            linked_identity = os.stat(
                item.basename,
                dir_fd=item.parent_descriptor,
                follow_symlinks=False,
            )
            current_parent_descriptor = _publication._open_parent_directory(
                project_root,
                PurePosixPath(item.relative_path).parent,
                create=False,
            )
            try:
                current_parent = os.fstat(current_parent_descriptor)
                current_link = os.stat(
                    item.basename,
                    dir_fd=current_parent_descriptor,
                    follow_symlinks=False,
                )
            finally:
                os.close(current_parent_descriptor)
            if (
                not _same_regular_inode(item.identity, descriptor_identity)
                or not _same_regular_inode(descriptor_identity, linked_identity)
                or not _same_regular_inode(descriptor_identity, current_link)
                or (current_parent.st_dev, current_parent.st_ino)
                != (parents[0].st_dev, parents[0].st_ino)
                or descriptor_identity.st_size != len(expected)
                or _publication._read_output_descriptor(item.output_descriptor, len(expected))
                != expected
            ):
                raise NativeTMCanonicalExcelExportError(
                    "published native-TM Excel pair identity or bytes drifted"
                )
            output_inodes.append((descriptor_identity.st_dev, descriptor_identity.st_ino))
        if len(set(output_inodes)) != 2:
            raise NativeTMCanonicalExcelExportError(
                "published native-TM Excel outputs must be distinct regular inodes"
            )
    except NativeTMCanonicalExcelExportError:
        raise
    except (OSError, _publication.NativeTMDocumentArtifactError) as exc:
        raise NativeTMCanonicalExcelExportError(
            "published native-TM Excel pair changed during strict replay"
        ) from exc


def _publish_pair(
    project_root: Path,
    workbook_path: Path,
    provenance_path: Path,
    workbook_bytes: bytes,
    provenance_bytes: bytes,
) -> _PairPublicationGuard:
    if workbook_path.parent != provenance_path.parent or workbook_path == provenance_path:
        raise NativeTMCanonicalExcelExportError("outputs must be distinct siblings")
    try:
        workbook_guard = _publication._write_exclusive(project_root, workbook_path, workbook_bytes)
    except (OSError, _publication.NativeTMDocumentArtifactError) as exc:
        raise NativeTMCanonicalExcelExportError(
            "native-TM Excel workbook cannot be exclusively published"
        ) from exc
    provenance_guard: Any | None = None
    try:
        provenance_guard = _publication._write_exclusive(
            project_root, provenance_path, provenance_bytes
        )
        pair = _PairPublicationGuard(
            workbook=workbook_guard,
            provenance=provenance_guard,
        )
        workbook_parent = os.fstat(workbook_guard.parent_descriptor)
        provenance_parent = os.fstat(provenance_guard.parent_descriptor)
        if (workbook_parent.st_dev, workbook_parent.st_ino) != (
            provenance_parent.st_dev,
            provenance_parent.st_ino,
        ):
            raise NativeTMCanonicalExcelExportError(
                "native-TM Excel outputs were published into different physical parents"
            )
        return pair
    except BaseException as cause:
        if provenance_guard is None:
            try:
                _publication._rollback_publication(workbook_guard, cause)
            except BaseException as rollback_error:
                raise NativeTMCanonicalExcelExportError(
                    "native-TM Excel workbook rollback failed after provenance publication error"
                ) from rollback_error
            finally:
                _publication._close_guard_best_effort(workbook_guard)
            raise NativeTMCanonicalExcelExportError(
                "native-TM Excel provenance cannot be exclusively published"
            ) from cause
        pair = _PairPublicationGuard(workbook=workbook_guard, provenance=provenance_guard)
        try:
            _rollback_pair(pair, cause)
        finally:
            _close_pair(pair)
        raise


def _output_paths(
    *,
    project_root: Path,
    workbook_path: Path,
    provenance_path: Path,
    policy: NativeTMCanonicalExcelPolicy,
    create_parent: bool,
) -> tuple[str, Path, str, Path]:
    if policy.output_directory != _OUTPUT_DIRECTORY:
        raise NativeTMCanonicalExcelExportError("export output policy directory drifted")
    return _preflight_output_paths(
        project_root=project_root,
        workbook_path=workbook_path,
        provenance_path=provenance_path,
        create_parent=create_parent,
    )


def _preflight_output_paths(
    *,
    project_root: Path,
    workbook_path: Path,
    provenance_path: Path,
    create_parent: bool,
) -> tuple[str, Path, str, Path]:
    workbook_relative, workbook_absolute = _project_path(
        project_root, workbook_path, "workbook output"
    )
    provenance_relative, provenance_absolute = _project_path(
        project_root, provenance_path, "provenance output"
    )
    if (
        workbook_absolute.suffix.casefold() != ".xlsx"
        or provenance_absolute.suffix.casefold() != ".json"
        or workbook_absolute.parent != provenance_absolute.parent
        or workbook_absolute == provenance_absolute
    ):
        raise NativeTMCanonicalExcelExportError(
            "outputs must be distinct sibling .xlsx and .json files"
        )
    output_root = PurePosixPath(_OUTPUT_DIRECTORY)
    for relative in (workbook_relative, provenance_relative):
        try:
            PurePosixPath(relative).relative_to(output_root)
        except ValueError as exc:
            raise NativeTMCanonicalExcelExportError(
                "outputs leave configured output directory"
            ) from exc
    # Publication creates/traverses parents only through retained no-follow dirfds.
    # Completed-pair loading opens the same path through retained read guards.
    _ = create_parent
    return (
        workbook_relative,
        workbook_absolute,
        provenance_relative,
        provenance_absolute,
    )


def _receipt_mapping_identity(value: object, label: str) -> ArtifactIdentity:
    record = _mapping(value, label)
    if set(record) != {
        "path",
        "sha256",
        "size_bytes",
        "format_version",
        "policy",
        "claim_boundary",
        "status",
        "run_id",
    }:
        raise NativeTMCanonicalExcelExportError(f"{label} fields drifted")
    identity = _identity_from_record(record, label)
    if not isinstance(record.get("run_id"), str) or not record.get("run_id"):
        raise NativeTMCanonicalExcelExportError(f"{label} run_id is invalid")
    return identity


def _preflight_receipt(
    receipt: object,
    *,
    project_root: Path,
    workbook_relative: str,
    provenance_relative: str,
    workbook_sha256: str,
    workbook_size_bytes: int,
) -> tuple[
    ArtifactIdentity,
    ArtifactIdentity,
    tuple[dict[str, object], ...],
    dict[str, object],
    dict[str, object],
]:
    record = _mapping(receipt, "native-TM Excel provenance")
    expected_keys = {
        "artifact_type",
        "claim_boundary",
        "code",
        "content_ledger",
        "dataset_role",
        "format_version",
        "inputs",
        "isolation",
        "policy",
        "provenance",
        "run_id",
        "source",
        "status",
        "summary",
        "workbook",
    }
    if (
        set(record) != expected_keys
        or record.get("artifact_type") != _RECEIPT_TYPE
        or record.get("claim_boundary") != _EXPORT_CLAIM
        or record.get("dataset_role") != _DATASET_ROLE
        or record.get("format_version") != 1
        or record.get("status") != _EXPORT_STATUS
        or record.get("isolation") != _receipt_isolation()
        or not isinstance(record.get("run_id"), str)
        or not record.get("run_id")
    ):
        raise NativeTMCanonicalExcelExportError("native-TM Excel provenance envelope drifted")
    workbook = _mapping(record.get("workbook"), "provenance workbook")
    provenance = _mapping(record.get("provenance"), "provenance self receipt")
    if workbook != {
        "filename": PurePosixPath(workbook_relative).name,
        "formula_count": 0,
        "path": workbook_relative,
        "sha256": workbook_sha256,
        "sheet_names": list(SHEET_NAMES),
        "size_bytes": workbook_size_bytes,
    } or provenance != {
        "filename": PurePosixPath(provenance_relative).name,
        "path": provenance_relative,
    }:
        raise NativeTMCanonicalExcelExportError("provenance output receipt drifted")
    inputs = _mapping(record.get("inputs"), "provenance inputs")
    if (
        set(inputs)
        != {
            "mapping",
            "native_tm_observations",
            "transitive_replay_inputs",
            "exact_receipt_lineage_verified",
        }
        or inputs.get("exact_receipt_lineage_verified") is not True
    ):
        raise NativeTMCanonicalExcelExportError("provenance input receipt drifted")
    mapping_receipt = _mapping(inputs.get("mapping"), "mapping receipt")
    observations_receipt = _mapping(inputs.get("native_tm_observations"), "observations receipt")
    mapping_identity = _receipt_mapping_identity(mapping_receipt, "mapping receipt")
    observations_identity = _receipt_mapping_identity(observations_receipt, "observations receipt")
    if {
        key: mapping_receipt.get(key)
        for key in ("format_version", "policy", "claim_boundary", "status")
    } != {
        "format_version": _MAPPING_FORMAT,
        "policy": _MAPPING_POLICY,
        "claim_boundary": _MAPPING_CLAIM,
        "status": _MAPPING_STATUS,
    } or {
        key: observations_receipt.get(key)
        for key in ("format_version", "policy", "claim_boundary", "status")
    } != {
        "format_version": _OBSERVATIONS_FORMAT,
        "policy": _OBSERVATIONS_POLICY,
        "claim_boundary": _OBSERVATIONS_CLAIM,
        "status": _OBSERVATIONS_STATUS,
    }:
        raise NativeTMCanonicalExcelExportError("provenance upstream contract drifted")
    transitive = _records(
        inputs.get("transitive_replay_inputs"), "transitive replay input receipts"
    )
    if len(transitive) != 3:
        raise NativeTMCanonicalExcelExportError("transitive replay receipt count drifted")
    transitive_records: list[dict[str, object]] = []
    expected_kinds = (
        "REGISTERED_NATIVE_TM_DOCUMENT_ARTIFACT",
        "SOURCE_PDF",
        "ACCEPTED_STATEMENT_DISCOVERY",
    )
    for item, expected_kind in zip(transitive, expected_kinds, strict=True):
        if set(item) != {"kind", "path", "sha256", "size_bytes"}:
            raise NativeTMCanonicalExcelExportError("transitive replay receipt fields drifted")
        identity = _identity_from_record(item, "transitive replay receipt")
        if item.get("kind") != expected_kind:
            raise NativeTMCanonicalExcelExportError("transitive replay receipt order drifted")
        transitive_records.append(
            {
                "kind": item["kind"],
                "path": identity.path,
                "sha256": identity.sha256,
                "size_bytes": identity.size_bytes,
            }
        )
    if (
        not mapping_identity.path.startswith(f"{_OUTPUT_DIRECTORY}/")
        or not observations_identity.path.startswith(f"{_OUTPUT_DIRECTORY}/")
        or not str(transitive_records[0]["path"]).startswith(f"{_OUTPUT_DIRECTORY}/")
        or not str(transitive_records[2]["path"]).startswith(f"{_OUTPUT_DIRECTORY}/")
        or len(
            {
                mapping_identity.path,
                observations_identity.path,
                *(str(item["path"]) for item in transitive_records),
            }
        )
        != 5
    ):
        raise NativeTMCanonicalExcelExportError("provenance input paths are not isolated")
    code = _mapping(record.get("code"), "provenance code")
    if set(code) != {
        "commit",
        "dirty",
        "implementation",
        "mapping_producer",
        "observations_producer",
    }:
        raise NativeTMCanonicalExcelExportError("provenance code receipt drifted")
    _mapping(code.get("mapping_producer"), "mapping producer receipt")
    _mapping(code.get("observations_producer"), "observations producer receipt")
    exporter_producer = _preflight_committed_exporter_producer(
        {
            "commit": code.get("commit"),
            "dirty": code.get("dirty"),
            "implementation": code.get("implementation"),
        },
        project_root=project_root,
    )
    receipt_policy = _mapping(record.get("policy"), "provenance policy")
    if (
        set(receipt_policy)
        != {
            "export_policy",
            "export_policy_sha256",
            "force_mapping",
            "imputation",
            "producer_schema_snapshot_authoritative",
        }
        or receipt_policy.get("export_policy") != _EXPORT_POLICY
        or not isinstance(receipt_policy.get("export_policy_sha256"), str)
        or _SHA256.fullmatch(str(receipt_policy.get("export_policy_sha256"))) is None
        or receipt_policy.get("force_mapping") is not False
        or receipt_policy.get("imputation") is not False
        or receipt_policy.get("producer_schema_snapshot_authoritative") is not True
    ):
        raise NativeTMCanonicalExcelExportError("provenance export policy receipt drifted")
    producer_policy_bytes = _git_file_bytes(
        project_root,
        str(exporter_producer["commit"]),
        EXPORT_POLICY_RELATIVE_PATH.as_posix(),
    )
    if _sha256(producer_policy_bytes) != receipt_policy["export_policy_sha256"]:
        raise NativeTMCanonicalExcelExportError(
            "provenance export policy differs from producer commit"
        )
    return (
        mapping_identity,
        observations_identity,
        tuple(transitive_records),
        exporter_producer,
        copy.deepcopy(dict(record)),
    )


def load_registered_native_tm_canonical_excel(
    *,
    project_root: Path,
    workbook_path: Path,
    workbook_expected_sha256: str,
    provenance_path: Path,
    provenance_expected_sha256: str,
) -> NativeTMCanonicalExcelExportResult:
    """Strictly replay a pair with its historical exporter producer."""

    project_root = project_root.resolve()
    if (
        not isinstance(workbook_expected_sha256, str)
        or _SHA256.fullmatch(workbook_expected_sha256) is None
    ):
        raise NativeTMCanonicalExcelExportError("trusted workbook SHA-256 is invalid")
    if (
        not isinstance(provenance_expected_sha256, str)
        or _SHA256.fullmatch(provenance_expected_sha256) is None
    ):
        raise NativeTMCanonicalExcelExportError("trusted provenance SHA-256 is invalid")
    workbook_guard: _HeldFile | None = None
    provenance_guard: _HeldFile | None = None
    lineage_guards: list[_HeldFile] = []
    try:
        (
            workbook_relative,
            workbook_absolute,
            provenance_relative,
            provenance_absolute,
        ) = _preflight_output_paths(
            project_root=project_root,
            workbook_path=workbook_path,
            provenance_path=provenance_path,
            create_parent=False,
        )
        workbook_guard = _open_held_file(
            project_root, workbook_absolute, "native-TM canonical workbook"
        )
        provenance_guard = _open_held_file(
            project_root, provenance_absolute, "native-TM canonical provenance"
        )
        workbook_parent = workbook_guard.guard.parent_identity
        provenance_parent = provenance_guard.guard.parent_identity
        if (
            (workbook_parent.st_dev, workbook_parent.st_ino)
            != (provenance_parent.st_dev, provenance_parent.st_ino)
            or not stat.S_ISREG(workbook_guard.identity.st_mode)
            or not stat.S_ISREG(provenance_guard.identity.st_mode)
            or (workbook_guard.identity.st_dev, workbook_guard.identity.st_ino)
            == (provenance_guard.identity.st_dev, provenance_guard.identity.st_ino)
        ):
            raise NativeTMCanonicalExcelExportError(
                "completed native-TM Excel pair must be distinct regular-file siblings"
            )
        if (
            _sha256(workbook_guard.payload) != workbook_expected_sha256
            or _sha256(provenance_guard.payload) != provenance_expected_sha256
        ):
            raise NativeTMCanonicalExcelExportError(
                "completed native-TM Excel pair does not match trusted SHA-256"
            )
        receipt = _parse_canonical_json(provenance_guard.payload, "native-TM canonical provenance")
        (
            mapping_identity,
            observations_identity,
            transitive_records,
            exporter_producer,
            normalized_receipt,
        ) = _preflight_receipt(
            receipt,
            project_root=project_root,
            workbook_relative=workbook_relative,
            provenance_relative=provenance_relative,
            workbook_sha256=workbook_expected_sha256,
            workbook_size_bytes=len(workbook_guard.payload),
        )
        replayed_workbook, replayed_provenance = _producer_commit_replay(
            project_root=project_root,
            mapping_identity=mapping_identity,
            exporter_producer=exporter_producer,
            workbook_relative_path=workbook_relative,
            provenance_relative_path=provenance_relative,
        )
        if (
            replayed_workbook != workbook_guard.payload
            or replayed_provenance != provenance_guard.payload
        ):
            raise NativeTMCanonicalExcelExportError(
                "completed native-TM Excel pair differs from deterministic replay"
            )
        lineage_identities = (
            (mapping_identity, "native-TM canonical mapping"),
            (observations_identity, "receipt-bound native-TM observations"),
            *(
                (
                    ArtifactIdentity(
                        path=str(item["path"]),
                        sha256=str(item["sha256"]),
                        size_bytes=int(item["size_bytes"]),
                    ),
                    str(item["kind"]),
                )
                for item in transitive_records
            ),
        )
        for identity, label in lineage_identities:
            lineage_guards.append(_open_identity_guard(project_root, identity, label))
        inode_keys = {(item.identity.st_dev, item.identity.st_ino) for item in lineage_guards}
        if len(inode_keys) != len(lineage_guards):
            raise NativeTMCanonicalExcelExportError(
                "completed-pair lineage inputs must be distinct regular files"
            )
        for item, (_identity, label) in zip(lineage_guards, lineage_identities, strict=True):
            _assert_held_stable(item, label)
        _assert_held_stable(workbook_guard, "native-TM canonical workbook")
        _assert_held_stable(provenance_guard, "native-TM canonical provenance")
        summary = _mapping(normalized_receipt.get("summary"), "provenance summary")
        return NativeTMCanonicalExcelExportResult(
            workbook_path=workbook_absolute,
            provenance_path=provenance_absolute,
            workbook_sha256=workbook_expected_sha256,
            provenance_sha256=provenance_expected_sha256,
            workbook_size_bytes=len(replayed_workbook),
            provenance_size_bytes=len(replayed_provenance),
            summary=copy.deepcopy(dict(summary)),
        )
    finally:
        for guard in reversed(lineage_guards):
            _close_held(guard)
        if provenance_guard is not None:
            _close_held(provenance_guard)
        if workbook_guard is not None:
            _close_held(workbook_guard)


def export_registered_native_tm_canonical_excel(
    *,
    project_root: Path,
    mapping_path: Path,
    mapping_expected_sha256: str,
    workbook_path: Path,
    provenance_path: Path,
) -> NativeTMCanonicalExcelExportResult:
    """Publish Excel from one trusted mapping; all other inputs are receipt-bound."""

    project_root = project_root.resolve()
    policy_relative, policy_absolute = _project_path(
        project_root, EXPORT_POLICY_RELATIVE_PATH, "native-TM Excel policy"
    )
    if policy_relative != EXPORT_POLICY_RELATIVE_PATH.as_posix():
        raise NativeTMCanonicalExcelExportError(
            f"native-TM Excel requires {EXPORT_POLICY_RELATIVE_PATH.as_posix()}"
        )
    policy_guard = _open_held_file(project_root, policy_absolute, "native-TM Excel policy")
    bundle: _InputBundle | None = None
    publication_guard: _PairPublicationGuard | None = None
    try:
        policy = _policy_from_bytes(policy_guard.payload, path=policy_absolute)
        exporter_producer = _current_exporter_producer(project_root)
        producer_commit = str(exporter_producer["commit"])
        if policy_guard.payload != _git_file_bytes(
            project_root, producer_commit, EXPORT_POLICY_RELATIVE_PATH.as_posix()
        ):
            raise NativeTMCanonicalExcelExportError(
                "native-TM Excel policy differs from producer commit"
            )
        (
            workbook_relative,
            workbook_absolute,
            provenance_relative,
            provenance_absolute,
        ) = _output_paths(
            project_root=project_root,
            workbook_path=workbook_path,
            provenance_path=provenance_path,
            policy=policy,
            create_parent=True,
        )
        bundle = _strict_load_input_bundle(
            project_root=project_root,
            mapping_path=mapping_path,
            mapping_expected_sha256=mapping_expected_sha256,
            policy=policy,
        )
        mapping_identity = ArtifactIdentity(
            path=bundle.mapping.relative_path,
            sha256=mapping_expected_sha256,
            size_bytes=len(bundle.mapping.payload),
        )
        observations_identity = ArtifactIdentity(
            path=bundle.observations.relative_path,
            sha256=_sha256(bundle.observations.payload),
            size_bytes=len(bundle.observations.payload),
        )
        artifacts = _build_prevalidated_native_tm_canonical_excel_artifacts(
            bundle.mapping_payload,
            bundle.observations_payload,
            mapping_identity=mapping_identity,
            observations_identity=observations_identity,
            transitive_identities=bundle.transitive_identities,
            workbook_relative_path=workbook_relative,
            provenance_relative_path=provenance_relative,
            policy=policy,
            exporter_producer=exporter_producer,
        )
        _assert_bundle_stable(bundle)
        _assert_held_stable(policy_guard, "native-TM Excel policy")
        if _current_exporter_producer(project_root) != exporter_producer:
            raise NativeTMCanonicalExcelExportError("export producer changed during build")
        publication_guard = _publish_pair(
            project_root,
            workbook_absolute,
            provenance_absolute,
            artifacts.workbook_bytes,
            artifacts.provenance_bytes,
        )
        try:
            result = load_registered_native_tm_canonical_excel(
                project_root=project_root,
                workbook_path=Path(workbook_relative),
                workbook_expected_sha256=artifacts.workbook_sha256,
                provenance_path=Path(provenance_relative),
                provenance_expected_sha256=artifacts.provenance_sha256,
            )
            _assert_bundle_stable(bundle)
            _assert_held_stable(policy_guard, "native-TM Excel policy")
            if _current_exporter_producer(project_root) != exporter_producer:
                raise NativeTMCanonicalExcelExportError(
                    "export producer changed during strict replay"
                )
            _revalidate_published_pair(
                project_root,
                publication_guard,
                artifacts.workbook_bytes,
                artifacts.provenance_bytes,
            )
            _close_pair(publication_guard)
            publication_guard = None
            return result
        except BaseException as replay_error:
            try:
                _rollback_pair(publication_guard, replay_error)
            finally:
                _close_pair(publication_guard)
                publication_guard = None
            raise NativeTMCanonicalExcelExportError(
                "published native-TM Excel pair failed strict replay and was rolled back"
            ) from replay_error
    finally:
        if publication_guard is not None:
            _close_pair(publication_guard)
        if bundle is not None:
            _close_input_bundle(bundle)
        _close_held(policy_guard)


__all__ = [
    "CANONICAL_OBSERVATION_HEADERS",
    "EXPORT_POLICY_RELATIVE_PATH",
    "RUN_METADATA_HEADERS",
    "SCHEMA_DISPOSITION_HEADERS",
    "SHEET_NAMES",
    "SOURCE_DISPOSITION_HEADERS",
    "VALIDATION_HEADERS",
    "ArtifactIdentity",
    "NativeTMCanonicalExcelArtifacts",
    "NativeTMCanonicalExcelExportError",
    "NativeTMCanonicalExcelExportResult",
    "NativeTMCanonicalExcelPolicy",
    "export_registered_native_tm_canonical_excel",
    "load_native_tm_canonical_excel_policy",
    "load_registered_native_tm_canonical_excel",
]
