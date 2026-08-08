"""Coverage-first KQKD development workbook export.

The exporter deliberately keeps three evidence scopes separate: item mapping,
numeric transcription, and accounting checks.  A successful export is useful
development evidence, but it is never promoted to a fully verified artifact.
"""

from __future__ import annotations

import hashlib
import json
import os
import zipfile
from collections.abc import Mapping
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bctc_ai.mapping.kqkd_item_mapping import (
    KQKDItemMappingResult,
    KQKDSchemaStatus,
    KQKDSourceRowStatus,
    validate_kqkd_mapping_result,
)
from bctc_ai.reconciliation.kqkd_numeric import KQKDNumericVerificationResult
from bctc_ai.tables.kqkd_word_box import ParsedKQKDWordBoxPage

KQKD_DEVELOPMENT_SHEETS = (
    "KQKD",
    "PROVENANCE",
    "VALIDATION_DIAGNOSTICS",
    "RUN_METADATA",
)
KQKD_DEVELOPMENT_SCHEMA_COUNT = 24
KQKD_DEVELOPMENT_SOURCE_ROW_COUNT = 22
KQKD_DEVELOPMENT_TARGET_VALUE_COUNT = 42
KQKD_DEVELOPMENT_ACCOUNTING_CHECK_COUNT = 32

_FIXED_TIMESTAMP = datetime(2000, 1, 1, tzinfo=UTC).replace(tzinfo=None)
_JSON_TYPES = (str, int, float, bool, list, tuple, dict, type(None))


class KQKDDevelopmentExportError(ValueError):
    """Raised when development-export inputs or destinations drift."""


@dataclass(frozen=True)
class KQKDDevelopmentArtifacts:
    workbook_bytes: bytes
    coverage_bytes: bytes
    workbook_sha256: str
    coverage_sha256: str
    target_value_count: int
    accounting_check_count: int
    fully_verified: bool = False


@dataclass(frozen=True)
class KQKDDevelopmentExportResult:
    workbook_path: str
    coverage_path: str
    workbook_sha256: str
    coverage_sha256: str
    workbook_size_bytes: int
    coverage_size_bytes: int
    target_value_count: int
    accounting_check_count: int
    fully_verified: bool = False


@dataclass(frozen=True)
class _ValidatedInputs:
    target_axes: tuple[Any, Any]
    provenance_axes: tuple[Any, Any]
    rows_by_id: dict[str, Any]
    source_by_id: dict[str, Any]
    schema_by_id: dict[int, Any]
    numeric_by_key: dict[tuple[str, str], Any]


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _canonical_json_bytes(payload: object) -> bytes:
    return (
        json.dumps(payload, ensure_ascii=False, allow_nan=False, sort_keys=True, indent=2) + "\n"
    ).encode("utf-8")


def _canonical_json_cell(payload: object) -> str:
    return json.dumps(
        payload,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _normalize_run_metadata(metadata: Mapping[str, object] | None) -> dict[str, object]:
    if metadata is None:
        return {}
    if not isinstance(metadata, Mapping):
        raise KQKDDevelopmentExportError("run metadata must be a mapping")
    normalized: dict[str, object] = {}
    for key, value in metadata.items():
        if not isinstance(key, str) or not key or len(key) > 128:
            raise KQKDDevelopmentExportError("run metadata key is invalid")
        if not isinstance(value, _JSON_TYPES):
            raise KQKDDevelopmentExportError("run metadata must contain JSON values")
        try:
            normalized[key] = json.loads(
                json.dumps(value, ensure_ascii=False, allow_nan=False, sort_keys=True)
            )
        except (TypeError, ValueError) as exc:
            raise KQKDDevelopmentExportError("run metadata is not canonical JSON") from exc
    return dict(sorted(normalized.items()))


def _decimal_integer(value: str, *, field: str) -> Decimal:
    try:
        parsed = Decimal(value)
    except (InvalidOperation, TypeError) as exc:
        raise KQKDDevelopmentExportError(f"{field} is not a decimal integer") from exc
    if not parsed.is_finite() or parsed != parsed.to_integral_value():
        raise KQKDDevelopmentExportError(f"{field} is not a decimal integer")
    return parsed


def _canonical_vnd(cell: Any, axis: Any) -> int:
    value = _decimal_integer(
        cell.pdf_text_signed_integer_reported_unit,
        field="PDF-text reported-unit value",
    )
    multiplier = axis.unit_multiplier
    if isinstance(multiplier, bool) or not isinstance(multiplier, int) or multiplier <= 0:
        raise KQKDDevelopmentExportError("KQKD unit multiplier is invalid")
    canonical = value * multiplier
    if canonical != canonical.to_integral_value():
        raise KQKDDevelopmentExportError("canonical VND value is not integral")
    return int(canonical)


def _validate_inputs(
    parsed: ParsedKQKDWordBoxPage,
    mapping: KQKDItemMappingResult,
    numeric: KQKDNumericVerificationResult,
) -> _ValidatedInputs:
    if not isinstance(parsed, ParsedKQKDWordBoxPage):
        raise KQKDDevelopmentExportError("parsed KQKD page model is required")
    if not isinstance(mapping, KQKDItemMappingResult):
        raise KQKDDevelopmentExportError("KQKD item mapping model is required")
    if not isinstance(numeric, KQKDNumericVerificationResult):
        raise KQKDDevelopmentExportError("KQKD numeric verification model is required")
    try:
        validate_kqkd_mapping_result(mapping)
    except ValueError as exc:
        raise KQKDDevelopmentExportError("KQKD mapping result is invalid") from exc

    if (
        mapping.statement_type != "KQKD"
        or mapping.status != "RESOLVED"
        or not mapping.automatic_selection_allowed
        or mapping.schema_item_count != KQKD_DEVELOPMENT_SCHEMA_COUNT
        or mapping.mapped_schema_count != 21
        or mapping.not_observed_schema_count != 3
        or mapping.ambiguous_schema_count != 0
        or mapping.source_row_count != KQKD_DEVELOPMENT_SOURCE_ROW_COUNT
        or mapping.mapped_source_row_count != 21
        or mapping.source_only_row_count != 1
        or mapping.ambiguous_source_row_count != 0
    ):
        raise KQKDDevelopmentExportError("KQKD mapping coverage is not exportable")
    if len(parsed.rows) != KQKD_DEVELOPMENT_SOURCE_ROW_COUNT or len(parsed.axes) != 4:
        raise KQKDDevelopmentExportError("KQKD parser denominator drifted")

    target_axes = tuple(parsed.schema_export_axes)
    provenance_axes = tuple(parsed.provenance_only_axes)
    if (
        len(target_axes) != 2
        or len(provenance_axes) != 2
        or {axis.current_or_comparative for axis in target_axes} != {"CURRENT", "COMPARATIVE"}
        or any(axis.group.value != "QUARTER" for axis in target_axes)
        or any(axis.group.value != "YTD" for axis in provenance_axes)
        or any(axis.canonical_unit != "VND" for axis in parsed.axes)
    ):
        raise KQKDDevelopmentExportError("KQKD target/provenance axis contract drifted")
    if len({axis.axis_id for axis in parsed.axes}) != 4 or tuple(
        axis.ordinal for axis in parsed.axes
    ) != (1, 2, 3, 4):
        raise KQKDDevelopmentExportError("KQKD axis identity drifted")

    rows_by_id = {row.row_id: row for row in parsed.rows}
    if len(rows_by_id) != len(parsed.rows) or tuple(row.ordinal for row in parsed.rows) != tuple(
        range(1, KQKD_DEVELOPMENT_SOURCE_ROW_COUNT + 1)
    ):
        raise KQKDDevelopmentExportError("KQKD source row identity drifted")
    if any(len(row.row.cells) != 4 or len(row.value_bboxes) != 4 for row in parsed.rows):
        raise KQKDDevelopmentExportError("KQKD source row cell denominator drifted")

    source_by_id = {item.row_id: item for item in mapping.source_dispositions}
    schema_by_id = {item.report_norm_id: item for item in mapping.schema_dispositions}
    if set(source_by_id) != set(rows_by_id) or len(schema_by_id) != 24:
        raise KQKDDevelopmentExportError("KQKD mapping/parser identity binding drifted")
    for row in parsed.rows:
        disposition = source_by_id[row.row_id]
        if disposition.order != row.ordinal:
            raise KQKDDevelopmentExportError("KQKD mapping/parser order binding drifted")

    if numeric.source_ocr_sha256 != parsed.source_sha256:
        raise KQKDDevelopmentExportError("numeric verification is bound to another OCR input")
    if (
        numeric.cell_count != 88
        or numeric.observed_cell_count != 88
        or numeric.numeric_verified_cell_count != 88
        or not numeric.numeric_verified
        or numeric.mapping_authority
        or numeric.fully_verified
        or numeric.accounting_equation_count != 8
        or numeric.accounting_check_count != KQKD_DEVELOPMENT_ACCOUNTING_CHECK_COUNT
        or numeric.accounting_passed_check_count != KQKD_DEVELOPMENT_ACCOUNTING_CHECK_COUNT
        or not numeric.accounting_verified
    ):
        raise KQKDDevelopmentExportError("KQKD numeric/accounting verification drifted")
    numeric_by_key = numeric.cell_by_key()
    expected_keys = {(row.row_id, axis.axis_id) for row in parsed.rows for axis in parsed.axes}
    if len(numeric_by_key) != 88 or set(numeric_by_key) != expected_keys:
        raise KQKDDevelopmentExportError("KQKD numeric cell identity drifted")
    for row in parsed.rows:
        for axis, parsed_cell in zip(parsed.axes, row.row.cells, strict=True):
            cell = numeric_by_key[(row.row_id, axis.axis_id)]
            if (
                cell.row_ordinal != row.ordinal
                or cell.axis_ordinal != axis.ordinal
                or cell.ppocr_raw_text != parsed_cell.raw_text
                or not cell.observed
                or not cell.numeric_verified
                or not cell.pdf_text_tokens
                or cell.pdf_text_joined != "".join(token.text for token in cell.pdf_text_tokens)
            ):
                raise KQKDDevelopmentExportError("KQKD numeric cell evidence drifted")
            ppocr = _decimal_integer(
                cell.ppocr_signed_integer_reported_unit,
                field="PP-OCR reported-unit value",
            )
            pdf_text = _decimal_integer(
                cell.pdf_text_signed_integer_reported_unit,
                field="PDF-text reported-unit value",
            )
            if parsed_cell.value != ppocr or ppocr != pdf_text:
                raise KQKDDevelopmentExportError("KQKD independent numeric evidence disagrees")

    if len({equation.equation_id for equation in numeric.accounting_equations}) != 8:
        raise KQKDDevelopmentExportError("KQKD accounting equation identity drifted")
    for equation in numeric.accounting_equations:
        if (
            not equation.passed
            or len(equation.residuals_by_axis) != 4
            or any(residual != "0" for residual in equation.residuals_by_axis)
        ):
            raise KQKDDevelopmentExportError("KQKD accounting equation failed")
        if not 1 <= equation.target_row_ordinal <= 22 or any(
            not 1 <= operand.row_ordinal <= 22 for operand in equation.operands
        ):
            raise KQKDDevelopmentExportError("KQKD accounting equation cross-link drifted")

    return _ValidatedInputs(
        target_axes=(target_axes[0], target_axes[1]),
        provenance_axes=(provenance_axes[0], provenance_axes[1]),
        rows_by_id=rows_by_id,
        source_by_id=source_by_id,
        schema_by_id=schema_by_id,
        numeric_by_key=numeric_by_key,
    )


def _style_header_range(sheet: Any, *, start_column: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(start_column, sheet.max_column + 1):
        cell = sheet.cell(1, column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _bbox_payload(bbox: Any) -> dict[str, object] | None:
    return None if bbox is None else asdict(bbox)


def _load_template(template_bytes: bytes, mapping: KQKDItemMappingResult) -> tuple[Any, tuple]:
    try:
        workbook = load_workbook(BytesIO(template_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise KQKDDevelopmentExportError("cannot load KQKD workbook template") from exc
    if workbook.sheetnames != ["Sheet1"]:
        workbook.close()
        raise KQKDDevelopmentExportError("KQKD workbook template sheet inventory drifted")
    sheet = workbook["Sheet1"]
    if sheet.max_row != 25 or sheet.max_column != 3:
        workbook.close()
        raise KQKDDevelopmentExportError("KQKD workbook template denominator drifted")
    if tuple(sheet.cell(1, column).value for column in range(1, 4)) != (
        None,
        "ReportNormId",
        "ReportNormName",
    ):
        workbook.close()
        raise KQKDDevelopmentExportError("KQKD workbook template header drifted")

    schema_by_order = {item.display_order: item for item in mapping.schema_dispositions}
    ac_snapshot = []
    for row_index in range(1, 26):
        cells = tuple(sheet.cell(row_index, column) for column in range(1, 4))
        ac_snapshot.append(tuple((cell.value, cell.data_type, cell.style_id) for cell in cells))
        if row_index == 1:
            continue
        display_order, report_norm_id, canonical_name = (cell.value for cell in cells)
        disposition = schema_by_order.get(display_order)
        if disposition is None or (
            disposition.report_norm_id,
            disposition.canonical_name,
        ) != (report_norm_id, canonical_name):
            workbook.close()
            raise KQKDDevelopmentExportError("mapping/template schema binding drifted")
    for candidate in workbook.worksheets:
        for row in candidate.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    workbook.close()
                    raise KQKDDevelopmentExportError("formulas are forbidden in development export")
    sheet.title = "KQKD"
    return workbook, tuple(ac_snapshot)


def _write_main_sheet(workbook: Any, context: _ValidatedInputs) -> int:
    sheet = workbook["KQKD"]
    headers = (
        "CurrentValueVND",
        "CurrentStatus",
        "CurrentPeriodStart",
        "CurrentPeriodEnd",
        "CurrentUnit",
        "ComparativeValueVND",
        "ComparativeStatus",
        "ComparativePeriodStart",
        "ComparativePeriodEnd",
        "ComparativeUnit",
    )
    for column, header in enumerate(headers, start=4):
        sheet.cell(1, column, header)
    _style_header_range(sheet, start_column=4)
    sheet.freeze_panes = "D2"

    by_role = {axis.current_or_comparative: axis for axis in context.target_axes}
    exported = 0
    for row_index in range(2, 26):
        report_norm_id = sheet.cell(row_index, 2).value
        disposition = context.schema_by_id[report_norm_id]
        for role, value_column in (("CURRENT", 4), ("COMPARATIVE", 9)):
            axis = by_role[role]
            status_column = value_column + 1
            sheet.cell(row_index, value_column + 2, axis.period_start.isoformat())
            sheet.cell(row_index, value_column + 3, axis.period_end.isoformat())
            sheet.cell(row_index, value_column + 4, axis.canonical_unit)
            if disposition.status == KQKDSchemaStatus.NOT_OBSERVED_IN_THIS_PDF.value:
                sheet.cell(row_index, status_column, disposition.status)
                continue
            if disposition.status != KQKDSchemaStatus.MAPPED.value or not disposition.source_row_id:
                raise KQKDDevelopmentExportError("ambiguous KQKD item reached workbook export")
            cell = context.numeric_by_key[(disposition.source_row_id, axis.axis_id)]
            if not cell.numeric_verified:
                raise KQKDDevelopmentExportError("unverified numeric cell reached workbook export")
            sheet.cell(row_index, value_column, _canonical_vnd(cell, axis))
            sheet.cell(row_index, status_column, "VALUE")
            sheet.cell(row_index, value_column).number_format = "#,##0;[Red](#,##0);-"
            exported += 1
    for column in range(4, 14):
        sheet.column_dimensions[get_column_letter(column)].width = 22
    sheet.auto_filter.ref = "A1:M25"
    if exported != KQKD_DEVELOPMENT_TARGET_VALUE_COUNT:
        raise KQKDDevelopmentExportError("KQKD target value count drifted")
    return exported


def _write_provenance(
    workbook: Any,
    parsed: ParsedKQKDWordBoxPage,
    context: _ValidatedInputs,
) -> None:
    sheet = workbook.create_sheet("PROVENANCE")
    headers = (
        "RowId",
        "RowOrdinal",
        "SourceLabel",
        "NoteReference",
        "SourceRowStatus",
        "ReportNormId",
        "AxisId",
        "AxisOrdinal",
        "AxisGroup",
        "SchemaExportCandidate",
        "CurrentOrComparative",
        "PeriodStart",
        "PeriodEnd",
        "PeriodType",
        "CanonicalUnit",
        "UnitMultiplier",
        "PPOCRRawText",
        "PPOCRSignedIntegerReportedUnit",
        "PDFTextTokensJson",
        "PDFTextJoined",
        "PDFTextSignedIntegerReportedUnit",
        "CanonicalValueVND",
        "Observed",
        "NumericVerified",
        "LabelBboxJson",
        "ValueBboxJson",
        "LabelLineIndicesJson",
        "ValueLineIndicesJson",
        "MappingReason",
    )
    sheet.append(headers)
    for row in parsed.rows:
        source = context.source_by_id[row.row_id]
        for axis_index, axis in enumerate(parsed.axes):
            cell = context.numeric_by_key[(row.row_id, axis.axis_id)]
            tokens = [
                {
                    "token_index": token.token_index,
                    "text": token.text,
                    "pdf_bbox": _bbox_payload(token.pdf_bbox),
                    "render_bbox": _bbox_payload(token.render_bbox),
                }
                for token in cell.pdf_text_tokens
            ]
            sheet.append(
                (
                    row.row_id,
                    row.ordinal,
                    row.row.label,
                    row.row.note_reference,
                    source.status,
                    source.report_norm_id,
                    axis.axis_id,
                    axis.ordinal,
                    axis.group.value,
                    axis.schema_export_candidate,
                    axis.current_or_comparative,
                    axis.period_start.isoformat(),
                    axis.period_end.isoformat(),
                    axis.period_type,
                    axis.canonical_unit,
                    axis.unit_multiplier,
                    cell.ppocr_raw_text,
                    cell.ppocr_signed_integer_reported_unit,
                    _canonical_json_cell(tokens),
                    cell.pdf_text_joined,
                    cell.pdf_text_signed_integer_reported_unit,
                    _canonical_vnd(cell, axis),
                    cell.observed,
                    cell.numeric_verified,
                    _canonical_json_cell(_bbox_payload(row.label_bbox)),
                    _canonical_json_cell(_bbox_payload(row.value_bboxes[axis_index])),
                    _canonical_json_cell(list(row.label_line_indices)),
                    _canonical_json_cell(list(row.value_line_indices[axis_index])),
                    source.reason,
                )
            )
    _style_header_range(sheet)
    for column in (3, 19, 25, 26, 29):
        sheet.column_dimensions[get_column_letter(column)].width = 52


def _write_diagnostics(
    workbook: Any,
    parsed: ParsedKQKDWordBoxPage,
    numeric: KQKDNumericVerificationResult,
    context: _ValidatedInputs,
) -> int:
    sheet = workbook.create_sheet("VALIDATION_DIAGNOSTICS")
    headers = (
        "DiagnosticType",
        "EquationId",
        "AxisId",
        "AxisGroup",
        "CurrentOrComparative",
        "TargetRowOrdinal",
        "TargetRowId",
        "TargetReportNormId",
        "OperandsJson",
        "ResidualReportedUnit",
        "Status",
        "AuthorityScope",
    )
    sheet.append(headers)
    rows_by_ordinal = {row.ordinal: row for row in parsed.rows}
    count = 0
    for equation in numeric.accounting_equations:
        target_row = rows_by_ordinal[equation.target_row_ordinal]
        target_source = context.source_by_id[target_row.row_id]
        operands = [
            {
                "row_ordinal": operand.row_ordinal,
                "row_id": rows_by_ordinal[operand.row_ordinal].row_id,
                "coefficient": operand.coefficient,
            }
            for operand in equation.operands
        ]
        for axis, residual in zip(parsed.axes, equation.residuals_by_axis, strict=True):
            sheet.append(
                (
                    "ACCOUNTING_EQUATION",
                    equation.equation_id,
                    axis.axis_id,
                    axis.group.value,
                    axis.current_or_comparative,
                    equation.target_row_ordinal,
                    target_row.row_id,
                    target_source.report_norm_id,
                    _canonical_json_cell(operands),
                    residual,
                    "PASS" if equation.passed and residual == "0" else "FAIL",
                    "ACCOUNTING_ONLY",
                )
            )
            count += 1
    _style_header_range(sheet)
    sheet.column_dimensions["I"].width = 72
    if count != KQKD_DEVELOPMENT_ACCOUNTING_CHECK_COUNT:
        raise KQKDDevelopmentExportError("KQKD accounting diagnostic count drifted")
    return count


def _write_metadata(
    workbook: Any,
    *,
    parsed: ParsedKQKDWordBoxPage,
    mapping: KQKDItemMappingResult,
    numeric: KQKDNumericVerificationResult,
    run_metadata: Mapping[str, object],
    target_value_count: int,
    accounting_check_count: int,
) -> None:
    sheet = workbook.create_sheet("RUN_METADATA")
    sheet.append(("Key", "Value"))
    standard: dict[str, object] = {
        "accounting.authority_scope": "ACCOUNTING_EQUATIONS_ONLY",
        "accounting.check_count": accounting_check_count,
        "accounting.passed_check_count": numeric.accounting_passed_check_count,
        "accounting.verified": numeric.accounting_verified,
        "dataset_role": "DEVELOPMENT",
        "fully_verified": False,
        "mapping.authority_scope": "SCHEMA_ROW_SELECTION_ONLY",
        "mapping.automatic_selection_allowed": mapping.automatic_selection_allowed,
        "mapping.status": mapping.status,
        "numeric.authority_scope": "NUMERIC_TRANSCRIPTION_ONLY",
        "numeric.mapping_authority": numeric.mapping_authority,
        "numeric.verified_cell_count": numeric.numeric_verified_cell_count,
        "numeric.verified": numeric.numeric_verified,
        "parser.source_sha256": parsed.source_sha256,
        "schema.reconciled_count": mapping.schema_item_count,
        "schema.mapped_count": mapping.mapped_schema_count,
        "schema.not_observed_count": mapping.not_observed_schema_count,
        "source.row_count": mapping.source_row_count,
        "source.source_only_count": mapping.source_only_row_count,
        "statement_type": "KQKD",
        "target.value_count": target_value_count,
    }
    for key, value in sorted(standard.items()):
        sheet.append((key, value))
    for key, value in run_metadata.items():
        rendered = (
            value
            if isinstance(value, (str, int, float, bool)) or value is None
            else (_canonical_json_cell(value))
        )
        sheet.append((f"run.{key}", rendered))
    _style_header_range(sheet)
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 72


def _has_formula(workbook: Any) -> bool:
    return any(
        cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def _deterministic_workbook_bytes(workbook: Any) -> bytes:
    workbook.properties.creator = "bctc-ai"
    workbook.properties.lastModifiedBy = "bctc-ai"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP
    raw = BytesIO()
    workbook.save(raw)
    source = BytesIO(raw.getvalue())
    target = BytesIO()
    with (
        zipfile.ZipFile(source, "r") as archive,
        zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as output,
    ):
        for name in sorted(archive.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            info.flag_bits = 0x800
            output.writestr(info, archive.read(name), compresslevel=9)
    return target.getvalue()


def _verify_serialized_workbook(
    workbook_bytes: bytes,
    *,
    ac_snapshot: tuple,
    target_value_count: int,
) -> None:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    except Exception as exc:
        raise KQKDDevelopmentExportError("serialized KQKD workbook cannot be reopened") from exc
    try:
        if tuple(workbook.sheetnames) != KQKD_DEVELOPMENT_SHEETS or _has_formula(workbook):
            raise KQKDDevelopmentExportError("serialized KQKD workbook contract drifted")
        main = workbook["KQKD"]
        actual_snapshot = tuple(
            tuple(
                (
                    main.cell(row_index, column).value,
                    main.cell(row_index, column).data_type,
                    main.cell(row_index, column).style_id,
                )
                for column in range(1, 4)
            )
            for row_index in range(1, 26)
        )
        if actual_snapshot != ac_snapshot:
            raise KQKDDevelopmentExportError("template columns A:C were not preserved")
        exported = sum(
            main.cell(row_index, column).value is not None
            for row_index in range(2, 26)
            for column in (4, 9)
        )
        if exported != target_value_count:
            raise KQKDDevelopmentExportError("serialized KQKD target count drifted")
    finally:
        workbook.close()


def build_kqkd_development_artifacts(
    *,
    template_path: Path,
    workbook_name: str,
    parsed: ParsedKQKDWordBoxPage,
    mapping: KQKDItemMappingResult,
    numeric: KQKDNumericVerificationResult,
    run_metadata: Mapping[str, object] | None = None,
) -> KQKDDevelopmentArtifacts:
    """Build deterministic workbook/coverage bytes without writing to disk."""

    if (
        not isinstance(workbook_name, str)
        or not workbook_name
        or Path(workbook_name).name != workbook_name
        or not workbook_name.lower().endswith(".xlsx")
    ):
        raise KQKDDevelopmentExportError("workbook name must be a local .xlsx filename")
    template_path = Path(template_path)
    try:
        template_bytes = template_path.read_bytes()
    except OSError as exc:
        raise KQKDDevelopmentExportError("cannot read KQKD workbook template") from exc
    metadata = _normalize_run_metadata(run_metadata)
    context = _validate_inputs(parsed, mapping, numeric)
    workbook, ac_snapshot = _load_template(template_bytes, mapping)
    try:
        target_value_count = _write_main_sheet(workbook, context)
        _write_provenance(workbook, parsed, context)
        accounting_check_count = _write_diagnostics(workbook, parsed, numeric, context)
        _write_metadata(
            workbook,
            parsed=parsed,
            mapping=mapping,
            numeric=numeric,
            run_metadata=metadata,
            target_value_count=target_value_count,
            accounting_check_count=accounting_check_count,
        )
        if _has_formula(workbook):
            raise KQKDDevelopmentExportError("formulas are forbidden in development export")
        workbook_bytes = _deterministic_workbook_bytes(workbook)
    finally:
        workbook.close()
    _verify_serialized_workbook(
        workbook_bytes,
        ac_snapshot=ac_snapshot,
        target_value_count=target_value_count,
    )
    workbook_sha256 = _sha256(workbook_bytes)

    not_observed_ids = sorted(
        item.report_norm_id
        for item in mapping.schema_dispositions
        if item.status == KQKDSchemaStatus.NOT_OBSERVED_IN_THIS_PDF.value
    )
    source_only_rows = sorted(
        (
            {
                "row_id": item.row_id,
                "row_ordinal": item.order,
                "source_label": context.rows_by_id[item.row_id].row.label,
            }
            for item in mapping.source_dispositions
            if item.status == KQKDSourceRowStatus.SOURCE_ONLY_PDF_ROW.value
        ),
        key=lambda item: item["row_ordinal"],
    )
    mapped_ids = {
        item.row_id
        for item in mapping.source_dispositions
        if item.status == KQKDSourceRowStatus.MAPPED.value
    }
    mapped_ytd_count = sum(
        context.numeric_by_key[(row_id, axis.axis_id)].numeric_verified
        for row_id in mapped_ids
        for axis in context.provenance_axes
    )
    coverage = {
        "artifact_type": "KQKD_DEVELOPMENT_COVERAGE",
        "authority": {
            "accounting": {
                "passed": numeric.accounting_verified,
                "scope": "ACCOUNTING_EQUATIONS_ONLY",
            },
            "fully_verified": False,
            "mapping": {
                "automatic_selection_allowed": mapping.automatic_selection_allowed,
                "passed": mapping.status == "RESOLVED",
                "scope": "SCHEMA_ROW_SELECTION_ONLY",
            },
            "numeric": {
                "mapping_authority": numeric.mapping_authority,
                "passed": numeric.numeric_verified,
                "scope": "NUMERIC_TRANSCRIPTION_ONLY",
            },
        },
        "coverage": {
            "accounting_check_count": accounting_check_count,
            "accounting_passed_check_count": numeric.accounting_passed_check_count,
            "mapped_schema_count": mapping.mapped_schema_count,
            "mapped_ytd_provenance_value_count": mapped_ytd_count,
            "not_observed_schema_count": mapping.not_observed_schema_count,
            "numeric_verified_cell_count": numeric.numeric_verified_cell_count,
            "provenance_cell_count": numeric.cell_count,
            "reconciled_schema_count": mapping.schema_item_count,
            "source_only_row_count": mapping.source_only_row_count,
            "source_row_count": mapping.source_row_count,
            "target_value_count": target_value_count,
            "ytd_provenance_cell_count": len(parsed.rows) * len(context.provenance_axes),
        },
        "dataset_role": "DEVELOPMENT",
        "format_version": 1,
        "fully_verified": False,
        "input_bindings": {
            "accounting_payload_sha256": numeric.accounting_payload_sha256,
            "mapping_payload_sha256": _sha256(_canonical_json_bytes(mapping.to_dict())),
            "mapping_policy_sha256": mapping.policy_sha256,
            "numeric_verification_payload_sha256": numeric.verification_payload_sha256,
            "parser_source_sha256": parsed.source_sha256,
            "schema_projection_sha256": mapping.schema_projection_sha256,
            "source_pdf_sha256": numeric.source_pdf_sha256,
            "source_render_sha256": numeric.source_render_sha256,
            "template_sha256": _sha256(template_bytes),
        },
        "not_observed_report_norm_ids": not_observed_ids,
        "run_metadata": metadata,
        "source_only_rows": source_only_rows,
        "statement_type": "KQKD",
        "workbook": {
            "filename": workbook_name,
            "sha256": workbook_sha256,
            "size_bytes": len(workbook_bytes),
        },
    }
    coverage_bytes = _canonical_json_bytes(coverage)
    return KQKDDevelopmentArtifacts(
        workbook_bytes=workbook_bytes,
        coverage_bytes=coverage_bytes,
        workbook_sha256=workbook_sha256,
        coverage_sha256=_sha256(coverage_bytes),
        target_value_count=target_value_count,
        accounting_check_count=accounting_check_count,
    )


def _write_exclusive(path: Path, payload: bytes) -> tuple[int, int]:
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags, 0o644)
    except OSError as exc:
        raise KQKDDevelopmentExportError(
            f"output already exists or cannot be created: {path}"
        ) from exc
    try:
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        stat = path.stat(follow_symlinks=False)
        return stat.st_dev, stat.st_ino
    except Exception:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def _rollback_if_same(path: Path, identity: tuple[int, int]) -> None:
    try:
        stat = path.stat(follow_symlinks=False)
        if (stat.st_dev, stat.st_ino) == identity:
            path.unlink()
    except OSError:
        pass


def export_kqkd_development(
    *,
    template_path: Path,
    workbook_path: Path,
    coverage_path: Path,
    parsed: ParsedKQKDWordBoxPage,
    mapping: KQKDItemMappingResult,
    numeric: KQKDNumericVerificationResult,
    run_metadata: Mapping[str, object] | None = None,
) -> KQKDDevelopmentExportResult:
    """Write the paired development artifacts once, refusing any overwrite."""

    workbook_path = Path(workbook_path)
    coverage_path = Path(coverage_path)
    if workbook_path == coverage_path:
        raise KQKDDevelopmentExportError("workbook and coverage destinations must differ")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    coverage_path.parent.mkdir(parents=True, exist_ok=True)
    if os.path.lexists(workbook_path) or os.path.lexists(coverage_path):
        raise KQKDDevelopmentExportError("development export refuses to overwrite outputs")

    artifacts = build_kqkd_development_artifacts(
        template_path=template_path,
        workbook_name=workbook_path.name,
        parsed=parsed,
        mapping=mapping,
        numeric=numeric,
        run_metadata=run_metadata,
    )
    workbook_identity = _write_exclusive(workbook_path, artifacts.workbook_bytes)
    try:
        _write_exclusive(coverage_path, artifacts.coverage_bytes)
    except Exception:
        _rollback_if_same(workbook_path, workbook_identity)
        raise
    return KQKDDevelopmentExportResult(
        workbook_path=workbook_path.as_posix(),
        coverage_path=coverage_path.as_posix(),
        workbook_sha256=artifacts.workbook_sha256,
        coverage_sha256=artifacts.coverage_sha256,
        workbook_size_bytes=len(artifacts.workbook_bytes),
        coverage_size_bytes=len(artifacts.coverage_bytes),
        target_value_count=artifacts.target_value_count,
        accounting_check_count=artifacts.accounting_check_count,
    )


__all__ = [
    "KQKD_DEVELOPMENT_ACCOUNTING_CHECK_COUNT",
    "KQKD_DEVELOPMENT_SCHEMA_COUNT",
    "KQKD_DEVELOPMENT_SHEETS",
    "KQKD_DEVELOPMENT_SOURCE_ROW_COUNT",
    "KQKD_DEVELOPMENT_TARGET_VALUE_COUNT",
    "KQKDDevelopmentArtifacts",
    "KQKDDevelopmentExportError",
    "KQKDDevelopmentExportResult",
    "build_kqkd_development_artifacts",
    "export_kqkd_development",
]
