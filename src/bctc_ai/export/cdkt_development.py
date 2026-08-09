"""Coverage-first CDKT workbook update from the frozen E0041 result.

This module preserves the sealed E0041 files and writes a separate development
workbook against the current business-schema template.  It applies only the
four user-authorized CDKT changes: the new total-equity item, the two repeated
structural headings, the 5699 identity decision, and the confirmed 2.320 value.
"""

from __future__ import annotations

import hashlib
import json
import os
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.xml.functions import fromstring, tostring

from bctc_ai.core.contracts import ObservationKind
from bctc_ai.tables.cdkt_off_balance_page5 import (
    CDKT_OFF_BALANCE_PAGE5_POLICY_RELATIVE_PATH,
    ParsedCDKTOffBalancePage5,
    load_cdkt_off_balance_page5_policy,
    parse_cdkt_off_balance_page5,
)

CDKT_SCHEMA_COUNT = 99
CDKT_MAPPED_SCHEMA_COUNT = 73
CDKT_NOT_OBSERVED_SCHEMA_COUNT = 26
CDKT_UNRESOLVED_SCHEMA_COUNT = 0
CDKT_SOURCE_ROW_COUNT = 75
CDKT_PHYSICAL_CELL_COUNT = 150
CDKT_OBSERVED_VALUE_COUNT = 132
CDKT_DERIVED_VALUE_COUNT = 2
CDKT_EXPORTED_VALUE_COUNT = 134
CDKT_TOTAL_EQUITY_ID = 5712
CDKT_VPB_SCHEMA_IDS = frozenset((*range(6035, 6054), 6055, 6056))
CDKT_VPB_OFF_BALANCE_SCHEMA_IDS = frozenset((*range(6038, 6054), 6055, 6056))
CDKT_MBB_MAPPED_OFF_BALANCE_SCHEMA_IDS = frozenset(range(6038, 6049))
CDKT_MBB_NOT_OBSERVED_OFF_BALANCE_SCHEMA_IDS = frozenset((*range(6049, 6054), 6055, 6056))
CDKT_TEMPLATE_MAX_ROW = CDKT_SCHEMA_COUNT + 1
CDKT_SCHEMA_EVOLUTION_OWNERSHIP = {
    6035: (4347, "page-0003-row-006-label", (-34_663_000_000, -39_393_000_000)),
    6036: (4352, "page-0003-row-017-label", (-561_362_000_000, -232_739_000_000)),
}

_CURRENT_PERIOD_END = "2026-03-31"
_COMPARATIVE_PERIOD_END = "2025-12-31"
_CURRENT_TOTAL_EQUITY_VND = 149_745_325_000_000
_COMPARATIVE_TOTAL_EQUITY_VND = 142_022_525_000_000
_USER_CONFIRMED_4363_CURRENT_VND = 2_320_000_000
_FIXED_TIMESTAMP = datetime(2000, 1, 1)
_CORE_PROPERTIES_MEMBER = "docProps/core.xml"
_MODIFIED_PROPERTY_TAG = "{http://purl.org/dc/terms/}modified"

_NOT_OBSERVED_IDS = frozenset(
    {
        4303,
        4306,
        4309,
        4326,
        4329,
        4333,
        4340,
        4341,
        4344,
        4345,
        4347,
        4359,
        4352,
        4360,
        4369,
        4370,
        4373,
        4374,
        6037,
        *CDKT_MBB_NOT_OBSERVED_OFF_BALANCE_SCHEMA_IDS,
    }
)
_NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES_IDS = _NOT_OBSERVED_IDS - {
    4306,
    6037,
    *CDKT_MBB_NOT_OBSERVED_OFF_BALANCE_SCHEMA_IDS,
}
_CROSS_STATEMENT_NOTE_LINKS = (
    (4344, 576, "RELATED_NOTE_DETAIL_NOT_PROVEN_IDENTICAL"),
    (4326, 585, "RELATED_NOTE_DISCLOSURE_NO_BACKFILL"),
    (4345, 5718, "BROADER_COMBINED_PROVISION_NOT_EQUIVALENT"),
)
_EQUITY_COMPONENT_IDS = (4364, 4365, 4342, 4341, 4343, 5699)


class CDKTDevelopmentExportError(ValueError):
    """Raised when frozen input, schema, or output invariants drift."""


@dataclass(frozen=True)
class CDKTDevelopmentArtifacts:
    workbook_bytes: bytes
    coverage_bytes: bytes
    workbook_sha256: str
    coverage_sha256: str
    observed_value_count: int
    derived_value_count: int
    exported_value_count: int
    fully_verified: bool = False


@dataclass(frozen=True)
class CDKTDevelopmentExportResult:
    workbook_path: str
    coverage_path: str
    workbook_sha256: str
    coverage_sha256: str
    workbook_size_bytes: int
    coverage_size_bytes: int
    observed_value_count: int
    derived_value_count: int
    exported_value_count: int
    fully_verified: bool = False


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _canonical_json_bytes(payload: object) -> bytes:
    return (
        json.dumps(payload, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")


def _has_formula(workbook: Any) -> bool:
    return any(
        cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def _deterministic_workbook_bytes(workbook: Any) -> bytes:
    workbook.properties.creator = "bctc-ai"
    workbook.properties.lastModifiedBy = "bctc-ai"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP
    deterministic_core = tostring(workbook.properties.to_tree())
    raw = BytesIO()
    workbook.save(raw)
    target = BytesIO()
    with (
        zipfile.ZipFile(BytesIO(raw.getvalue()), "r") as archive,
        zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as output,
    ):
        names = archive.namelist()
        if len(names) != len(set(names)) or names.count(_CORE_PROPERTIES_MEMBER) != 1:
            raise CDKTDevelopmentExportError("generated workbook core properties are invalid")
        generated_core = archive.read(_CORE_PROPERTIES_MEMBER)
        try:
            generated_root = fromstring(generated_core)
            deterministic_root = fromstring(deterministic_core)
        except Exception as exc:
            raise CDKTDevelopmentExportError(
                "generated workbook core properties are invalid"
            ) from exc
        generated_modified = generated_root.findall(_MODIFIED_PROPERTY_TAG)
        deterministic_modified = deterministic_root.findall(_MODIFIED_PROPERTY_TAG)
        if len(generated_modified) != 1 or len(deterministic_modified) != 1:
            raise CDKTDevelopmentExportError("generated workbook core properties are invalid")
        generated_modified[0].text = deterministic_modified[0].text
        if tostring(generated_root) != deterministic_core:
            raise CDKTDevelopmentExportError("generated workbook core properties drifted")
        for name in sorted(names):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            info.flag_bits = 0x800
            member = deterministic_core if name == _CORE_PROPERTIES_MEMBER else archive.read(name)
            output.writestr(info, member, compresslevel=9)
    return target.getvalue()


def _load_inputs(template_bytes: bytes, sealed_workbook_bytes: bytes) -> tuple[Any, Any]:
    try:
        template = load_workbook(BytesIO(template_bytes), data_only=False, keep_links=False)
        sealed = load_workbook(BytesIO(sealed_workbook_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise CDKTDevelopmentExportError("cannot decode CDKT workbook inputs") from exc
    if template.sheetnames != ["Sheet1"] or template.active.max_row != CDKT_TEMPLATE_MAX_ROW:
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("CDKT v2 template denominator drifted")
    if template.active.max_column != 3 or tuple(
        template.active.cell(1, column).value for column in range(1, 4)
    ) != (None, "ReportNormId", "ReportNormName"):
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("CDKT v2 template header drifted")
    template_ids = [
        template.active.cell(row, 2).value for row in range(2, CDKT_TEMPLATE_MAX_ROW + 1)
    ]
    template_by_id = {
        template.active.cell(row, 2).value: template.active.cell(row, 3).value
        for row in range(2, CDKT_TEMPLATE_MAX_ROW + 1)
    }
    if (
        len(template_ids) != CDKT_SCHEMA_COUNT
        or len(set(template_ids)) != CDKT_SCHEMA_COUNT
        or template_ids.count(CDKT_TOTAL_EQUITY_ID) != 1
        or template_by_id.get(CDKT_TOTAL_EQUITY_ID) != "TỔNG VỐN CHỦ SỞ HỮU"
        or template_by_id.get(4350) != "Chứng khoán đầu tư sẵn sàng để bán"
        or set(CDKT_VPB_SCHEMA_IDS) - set(template_ids)
    ):
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("CDKT v2 business-schema identities drifted")
    if sealed.sheetnames != [
        "Sheet1",
        "PROVENANCE",
        "VALIDATION_DIAGNOSTICS",
        "RUN_METADATA",
    ]:
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("sealed CDKT workbook inventory drifted")
    main = sealed["Sheet1"]
    if main.max_row != 78 or main.max_column != 11:
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("sealed CDKT main-sheet denominator drifted")
    sealed_ids = [main.cell(row, 2).value for row in range(2, 79)]
    if len(set(sealed_ids)) != 77 or set(sealed_ids) != set(template_ids) - {
        CDKT_TOTAL_EQUITY_ID,
        *CDKT_VPB_SCHEMA_IDS,
    }:
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("sealed/template CDKT identity linkage drifted")
    if _has_formula(template) or _has_formula(sealed):
        template.close()
        sealed.close()
        raise CDKTDevelopmentExportError("formulas are forbidden in CDKT development inputs")
    return template, sealed


def _copy_template_columns(template: Any, target: Any) -> tuple:
    source = template.active
    snapshot = []
    for row in range(1, CDKT_TEMPLATE_MAX_ROW + 1):
        row_snapshot = []
        for column in range(1, 4):
            original = source.cell(row, column)
            cell = target.cell(row, column)
            cell.value = original.value
            cell._style = copy(original._style)
            cell.number_format = original.number_format
            cell.alignment = copy(original.alignment)
            cell.font = copy(original.font)
            cell.fill = copy(original.fill)
            cell.border = copy(original.border)
            cell.protection = copy(original.protection)
            row_snapshot.append((cell.value, cell.data_type, cell.style_id))
        snapshot.append(tuple(row_snapshot))
    return tuple(snapshot)


def _row_by_id(sheet: Any) -> dict[int, int]:
    result = {sheet.cell(row, 2).value: row for row in range(2, sheet.max_row + 1)}
    if len(result) != sheet.max_row - 1:
        raise CDKTDevelopmentExportError("duplicate CDKT ReportNormId in workbook")
    return result


def _set_period_metadata(sheet: Any, row: int) -> None:
    sheet.cell(row, 8, _CURRENT_PERIOD_END)
    sheet.cell(row, 9, _COMPARATIVE_PERIOD_END)
    sheet.cell(row, 10, "VND")
    sheet.cell(row, 11, "CONSOLIDATED")


def _update_main_sheet(
    template: Any,
    workbook: Any,
    off_balance: ParsedCDKTOffBalancePage5,
) -> tuple[tuple, dict[int, int]]:
    sheet = workbook["Sheet1"]
    sealed_rows = {
        sheet.cell(row, 2).value: tuple(
            (sheet.cell(row, column).value, copy(sheet.cell(row, column)._style))
            for column in range(4, 12)
        )
        for row in range(2, sheet.max_row + 1)
    }
    default_styles = tuple(
        copy(sheet.cell(sheet.max_row, column)._style) for column in range(4, 12)
    )
    template_ids = [
        template.active.cell(row, 2).value for row in range(2, CDKT_TEMPLATE_MAX_ROW + 1)
    ]
    for row, report_norm_id in enumerate(template_ids, start=2):
        source_report_norm_id = CDKT_SCHEMA_EVOLUTION_OWNERSHIP.get(
            report_norm_id, (report_norm_id, "", ())
        )[0]
        source_cells = sealed_rows.get(source_report_norm_id)
        for column in range(4, 12):
            cell = sheet.cell(row, column)
            if source_cells is None:
                cell.value = None
                cell._style = copy(default_styles[column - 4])
            else:
                value, style = source_cells[column - 4]
                cell.value = value
                cell._style = copy(style)
    snapshot = _copy_template_columns(template, sheet)
    sheet.title = "CDKT"
    sheet.auto_filter.ref = f"A1:K{CDKT_TEMPLATE_MAX_ROW}"
    rows = _row_by_id(sheet)

    for report_norm_id in _NOT_OBSERVED_IDS:
        row = rows[report_norm_id]
        status = (
            "NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES"
            if report_norm_id in _NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES_IDS
            else "NOT_OBSERVED_IN_THIS_PDF"
        )
        sheet.cell(row, 4).value = None
        sheet.cell(row, 5).value = None
        sheet.cell(row, 6, status)
        sheet.cell(row, 7, status)
        _set_period_metadata(sheet, row)

    if (
        off_balance.scope != "CONSOLIDATED"
        or off_balance.unit != "VND"
        or off_balance.unit_multiplier != 1_000_000
        or {row.report_norm_id for row in off_balance.rows}
        != CDKT_MBB_MAPPED_OFF_BALANCE_SCHEMA_IDS
    ):
        raise CDKTDevelopmentExportError("CDKT page-5 off-balance source binding drifted")
    for source_row in off_balance.rows:
        row = rows[source_row.report_norm_id]
        for cell, (value_column, status_column) in zip(
            source_row.cells, ((4, 6), (5, 7)), strict=True
        ):
            if cell.period_end.isoformat() not in {_CURRENT_PERIOD_END, _COMPARATIVE_PERIOD_END}:
                raise CDKTDevelopmentExportError("CDKT page-5 period binding drifted")
            if cell.observation is ObservationKind.BLANK:
                sheet.cell(row, value_column).value = None
                sheet.cell(row, status_column, "BLANK")
            elif cell.observation in {ObservationKind.VALUE, ObservationKind.ZERO}:
                if cell.canonical_value is None:
                    raise CDKTDevelopmentExportError("CDKT page-5 value binding is absent")
                sheet.cell(row, value_column, cell.canonical_value)
                sheet.cell(row, status_column, cell.observation.value)
                sheet.cell(row, value_column).number_format = "#,##0;[Red](#,##0);-"
            else:
                raise CDKTDevelopmentExportError("CDKT page-5 observation is unsupported")
        _set_period_metadata(sheet, row)

    for target_id, (
        _source_id,
        _source_row_id,
        expected_values,
    ) in CDKT_SCHEMA_EVOLUTION_OWNERSHIP.items():
        row = rows[target_id]
        if tuple(sheet.cell(row, column).value for column in (4, 5)) != expected_values or tuple(
            sheet.cell(row, column).value for column in (6, 7)
        ) != ("VALUE", "VALUE"):
            raise CDKTDevelopmentExportError(
                f"CDKT schema-evolution ownership values drifted for {target_id}"
            )
        _set_period_metadata(sheet, row)

    row = rows[4363]
    sheet.cell(row, 4, _USER_CONFIRMED_4363_CURRENT_VND)
    sheet.cell(row, 6, "VALUE")
    _set_period_metadata(sheet, row)

    component_values: dict[str, list[int]] = {"CURRENT": [], "COMPARATIVE": []}
    for report_norm_id in _EQUITY_COMPONENT_IDS:
        component_row = rows[report_norm_id]
        for role, column in (("CURRENT", 4), ("COMPARATIVE", 5)):
            value = sheet.cell(component_row, column).value
            if report_norm_id == 4341:
                if value is not None:
                    raise CDKTDevelopmentExportError("not-observed 4341 unexpectedly has a value")
                continue
            if isinstance(value, bool) or not isinstance(value, int):
                raise CDKTDevelopmentExportError("CDKT equity component value drifted")
            component_values[role].append(value)
    current_equity = sum(component_values["CURRENT"])
    comparative_equity = sum(component_values["COMPARATIVE"])
    if (current_equity, comparative_equity) != (
        _CURRENT_TOTAL_EQUITY_VND,
        _COMPARATIVE_TOTAL_EQUITY_VND,
    ):
        raise CDKTDevelopmentExportError("CDKT total-equity component equation failed")

    row = rows[4325]
    sheet.cell(row, 4, current_equity)
    sheet.cell(row, 5, comparative_equity)
    sheet.cell(row, 6, "DERIVED_FROM_COMPONENTS")
    sheet.cell(row, 7, "DERIVED_FROM_COMPONENTS")
    _set_period_metadata(sheet, row)

    row = rows[CDKT_TOTAL_EQUITY_ID]
    sheet.cell(row, 4, _CURRENT_TOTAL_EQUITY_VND)
    sheet.cell(row, 5, _COMPARATIVE_TOTAL_EQUITY_VND)
    sheet.cell(row, 6, "VALUE")
    sheet.cell(row, 7, "VALUE")
    _set_period_metadata(sheet, row)
    for column in (4, 5):
        sheet.cell(row, column).number_format = "#,##0;[Red](#,##0);-"

    liabilities = sheet.cell(rows[4304], 4).value, sheet.cell(rows[4304], 5).value
    grand_total = sheet.cell(rows[4305], 4).value, sheet.cell(rows[4305], 5).value
    if (
        liabilities[0] + current_equity,
        liabilities[1] + comparative_equity,
    ) != grand_total:
        raise CDKTDevelopmentExportError("CDKT liabilities-plus-equity equation failed")

    statuses = [
        sheet.cell(row, column).value
        for row in range(2, CDKT_TEMPLATE_MAX_ROW + 1)
        for column in (6, 7)
    ]
    expected = {
        "BLANK": 7,
        "DASH": 5,
        "DERIVED_FROM_COMPONENTS": 2,
        "NOT_OBSERVED_IN_THIS_PDF": 18,
        "NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES": 34,
        "VALUE": 132,
    }
    actual = {status: statuses.count(status) for status in set(statuses)}
    if actual != expected:
        raise CDKTDevelopmentExportError(f"CDKT target status counts drifted: {actual}")
    return snapshot, rows


def _bbox_json(bbox: Any) -> str:
    return json.dumps(
        [bbox.x0, bbox.y0, bbox.x1, bbox.y1],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _append_off_balance_provenance(
    sheet: Any,
    headers: dict[str, int],
    off_balance: ParsedCDKTOffBalancePage5,
) -> None:
    validation_refs = list(off_balance.accounting_checks_passed)
    for source_row in off_balance.rows:
        for cell in source_row.cells:
            evidence_bbox = cell.value_bbox or source_row.label_bbox
            source_lines = [f"page-0005:line-{source_row.label_line_index:04d}"]
            if cell.value_line_index is not None:
                source_lines.append(f"page-0005:line-{cell.value_line_index:04d}")
            status = cell.observation.value
            record = {
                "CellId": f"{source_row.row_id}-axis-{cell.axis_ordinal + 1}",
                "RowId": source_row.row_id,
                "Page": off_balance.page_number,
                "RowOrdinal": source_row.ordinal,
                "AxisOrdinal": cell.axis_ordinal,
                "PeriodRole": cell.period_role,
                "PeriodEnd": cell.period_end.isoformat(),
                "Scope": off_balance.scope,
                "SourceRowIds": json.dumps(source_lines, separators=(",", ":")),
                "SourceLabel": source_row.visible_label,
                "SourceObservation": status,
                "SourceNumericStatus": status,
                "VisibleRawValue": cell.raw_text or None,
                "EvidenceDisplayedValue": cell.displayed_value,
                "EvidenceCanonicalValue": cell.canonical_value,
                "ExportedCanonicalValue": cell.canonical_value,
                "Status": status,
                "MappingStatus": (
                    "SOURCE_VISIBLE_STRUCTURAL_SCHEMA_MAPPING"
                    if cell.observation is ObservationKind.BLANK
                    else "SOURCE_VISIBLE_EXACT_SCHEMA_MAPPING"
                ),
                "ReportNormId": source_row.report_norm_id,
                "ProposedReportNormId": source_row.report_norm_id,
                "CandidateReportNormIds": f"[{source_row.report_norm_id}]",
                "CropPath": off_balance.source_render_path,
                "CropBbox": _bbox_json(evidence_bbox),
                "CropSha256": off_balance.source_render_sha256,
                "SourceRenderSha256": off_balance.source_render_sha256,
                "SourceOcrSha256": off_balance.source_ocr_sha256,
                "Unit": off_balance.unit,
                "UnitMultiplier": off_balance.unit_multiplier,
                "NumericVerificationStatus": (
                    "SOURCE_VISIBLE_STRUCTURAL_BLANK"
                    if cell.observation is ObservationKind.BLANK
                    else "SOURCE_OCR_STRICT_GROUPING_VERIFIED"
                ),
                "ValidationRefs": json.dumps(
                    ["PAGE5_SOURCE_VISIBLE_BINDING", *validation_refs],
                    separators=(",", ":"),
                ),
            }
            row = sheet.max_row + 1
            for name, column in headers.items():
                sheet.cell(row, column, record.get(name))


def _update_provenance(workbook: Any, off_balance: ParsedCDKTOffBalancePage5) -> None:
    sheet = workbook["PROVENANCE"]
    headers = {sheet.cell(1, column).value: column for column in range(1, sheet.max_column + 1)}
    required = {
        "CellId",
        "RowId",
        "Page",
        "RowOrdinal",
        "AxisOrdinal",
        "PeriodRole",
        "PeriodEnd",
        "Scope",
        "EvidenceDisplayedValue",
        "EvidenceCanonicalValue",
        "ExportedCanonicalValue",
        "Status",
        "MappingStatus",
        "ReportNormId",
        "ProposedReportNormId",
        "CandidateReportNormIds",
        "NumericVerificationStatus",
        "ValidationRefs",
    }
    if not required <= set(headers) or sheet.max_row != 129:
        raise CDKTDevelopmentExportError("sealed CDKT provenance schema drifted")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, headers["Scope"], "CONSOLIDATED")
        row_id = sheet.cell(row, headers["RowId"]).value
        axis = sheet.cell(row, headers["AxisOrdinal"]).value
        ownership = next(
            (
                (source_id, target_id, expected_values)
                for target_id, (source_id, source_row_id, expected_values) in (
                    CDKT_SCHEMA_EVOLUTION_OWNERSHIP.items()
                )
                if source_row_id == row_id
            ),
            None,
        )
        if ownership is not None:
            source_id, target_id, expected_values = ownership
            if (
                axis not in (0, 1)
                or sheet.cell(row, headers["ReportNormId"]).value != source_id
                or sheet.cell(row, headers["ExportedCanonicalValue"]).value != expected_values[axis]
                or sheet.cell(row, headers["Status"]).value != "VALUE"
            ):
                raise CDKTDevelopmentExportError(
                    f"sealed CDKT schema-evolution provenance drifted for {target_id}"
                )
            updates = {
                "MappingStatus": "SCHEMA_EVOLUTION_REASSIGNED_FROM_NARROWER_ITEM",
                "ReportNormId": target_id,
                "ProposedReportNormId": target_id,
                "CandidateReportNormIds": f"[{target_id}]",
            }
        elif row_id == "page-0004-row-000-label":
            updates = {
                "Status": "BLANK",
                "MappingStatus": "STRUCTURAL_REPEAT_OF_MAPPED_ITEM",
                "ReportNormId": 4304,
                "ProposedReportNormId": 4304,
                "CandidateReportNormIds": "[4304]",
                "NumericVerificationStatus": "USER_CONFIRMED_STRUCTURAL_BLANK",
            }
        elif row_id == "page-0004-row-013-label":
            updates = {
                "Status": "BLANK",
                "MappingStatus": "STRUCTURAL_REPEAT_OF_MAPPED_ITEM",
                "ReportNormId": CDKT_TOTAL_EQUITY_ID,
                "ProposedReportNormId": CDKT_TOTAL_EQUITY_ID,
                "CandidateReportNormIds": f"[{CDKT_TOTAL_EQUITY_ID}]",
                "NumericVerificationStatus": "USER_CONFIRMED_STRUCTURAL_BLANK",
            }
        elif row_id == "page-0004-row-023-label":
            value = _CURRENT_TOTAL_EQUITY_VND if axis == 0 else _COMPARATIVE_TOTAL_EQUITY_VND
            updates = {
                "Status": "VALUE",
                "MappingStatus": "USER_CONFIRMED_SCHEMA_MAPPING",
                "ReportNormId": CDKT_TOTAL_EQUITY_ID,
                "ProposedReportNormId": CDKT_TOTAL_EQUITY_ID,
                "CandidateReportNormIds": f"[{CDKT_TOTAL_EQUITY_ID}]",
                "ExportedCanonicalValue": value,
                "NumericVerificationStatus": "VERIFIED_OBSERVED_VALUE",
            }
        elif row_id == "page-0004-row-011-label" and axis == 0:
            updates = {
                "EvidenceDisplayedValue": "2320",
                "EvidenceCanonicalValue": _USER_CONFIRMED_4363_CURRENT_VND,
                "ExportedCanonicalValue": _USER_CONFIRMED_4363_CURRENT_VND,
                "Status": "VALUE",
                "NumericVerificationStatus": "USER_CONFIRMED_VISIBLE_VALUE",
            }
        else:
            continue
        for name, value in updates.items():
            sheet.cell(row, headers[name], value)

    _append_off_balance_provenance(sheet, headers, off_balance)
    statuses = [sheet.cell(row, headers["Status"]).value for row in range(2, sheet.max_row + 1)]
    expected = {"BLANK": 13, "DASH": 5, "VALUE": 132}
    actual = {status: statuses.count(status) for status in set(statuses)}
    if actual != expected:
        raise CDKTDevelopmentExportError(f"CDKT physical status counts drifted: {actual}")


def _append_business_diagnostics(workbook: Any, off_balance: ParsedCDKTOffBalancePage5) -> None:
    sheet = workbook["VALIDATION_DIAGNOSTICS"]
    if sheet.max_row != 37:
        raise CDKTDevelopmentExportError("sealed CDKT diagnostic denominator drifted")
    rows = (
        (
            "BUSINESS_4325_CURRENT",
            "BUSINESS_SCHEMA_FORMULA",
            "CURRENT",
            "BUSINESS_SCHEMA_FORMULA",
            "4325 = 4364 + 4365 + 4342 + 4341 + 4343 + 5699; 4341 not observed",
            "PASS",
            str(_CURRENT_TOTAL_EQUITY_VND),
            str(_CURRENT_TOTAL_EQUITY_VND),
        ),
        (
            "BUSINESS_4325_COMPARATIVE",
            "BUSINESS_SCHEMA_FORMULA",
            "COMPARATIVE",
            "BUSINESS_SCHEMA_FORMULA",
            "4325 = 4364 + 4365 + 4342 + 4341 + 4343 + 5699; 4341 not observed",
            "PASS",
            str(_COMPARATIVE_TOTAL_EQUITY_VND),
            str(_COMPARATIVE_TOTAL_EQUITY_VND),
        ),
        (
            "BUSINESS_5712_CURRENT",
            "BUSINESS_SCHEMA_FORMULA",
            "CURRENT",
            "VISIBLE_TOTAL_AND_FORMULA",
            "5712 = 4325 + 4306; 4306 not observed as a separate line",
            "PASS",
            str(_CURRENT_TOTAL_EQUITY_VND),
            str(_CURRENT_TOTAL_EQUITY_VND),
        ),
        (
            "BUSINESS_5712_COMPARATIVE",
            "BUSINESS_SCHEMA_FORMULA",
            "COMPARATIVE",
            "VISIBLE_TOTAL_AND_FORMULA",
            "5712 = 4325 + 4306; 4306 not observed as a separate line",
            "PASS",
            str(_COMPARATIVE_TOTAL_EQUITY_VND),
            str(_COMPARATIVE_TOTAL_EQUITY_VND),
        ),
    )
    for values in rows:
        sheet.append(values)
    by_id = {row.report_norm_id: row for row in off_balance.rows}
    for axis, period_role in enumerate(("CURRENT", "COMPARATIVE")):
        target = by_id[6041].cells[axis].canonical_value
        components = [
            by_id[report_norm_id].cells[axis].canonical_value
            for report_norm_id in range(6042, 6046)
        ]
        if target is None or any(value is None for value in components):
            raise CDKTDevelopmentExportError("CDKT page-5 diagnostic values are absent")
        component_sum = sum(value for value in components if value is not None)
        sheet.append(
            (
                f"OFF_BALANCE_6041_{period_role}",
                "SOURCE_VISIBLE_ACCOUNTING_CHECK",
                period_role,
                "PAGE5_OCR_AND_PIXEL_EVIDENCE",
                "6041 = 6042 + 6043 + 6044 + 6045",
                "PASS" if target == component_sum else "FAIL",
                str(target),
                str(component_sum),
            )
        )
        if target != component_sum:
            raise CDKTDevelopmentExportError("CDKT page-5 diagnostic equation failed")


def _replace_metadata(workbook: Any, off_balance: ParsedCDKTOffBalancePage5) -> None:
    sheet = workbook["RUN_METADATA"]
    sheet.delete_rows(1, sheet.max_row)
    rows = (
        ("Key", "Value"),
        ("artifact_type", "CDKT_BUSINESS_SCHEMA_DEVELOPMENT_EXCEL"),
        ("dataset_role", "DEVELOPMENT"),
        ("statement_type", "CDKT"),
        ("schema.total", CDKT_SCHEMA_COUNT),
        ("schema.reconciled", CDKT_SCHEMA_COUNT),
        ("schema.mapped", CDKT_MAPPED_SCHEMA_COUNT),
        ("schema.not_observed", CDKT_NOT_OBSERVED_SCHEMA_COUNT),
        (
            "schema.not_observed_on_target_statement_pages",
            len(_NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES_IDS),
        ),
        ("source.rows", CDKT_SOURCE_ROW_COUNT),
        ("source.physical_cells", CDKT_PHYSICAL_CELL_COUNT),
        ("target.observed_values", CDKT_OBSERVED_VALUE_COUNT),
        ("target.derived_values", CDKT_DERIVED_VALUE_COUNT),
        ("target.exported_values", CDKT_EXPORTED_VALUE_COUNT),
        ("scope", "CONSOLIDATED"),
        ("target.statement_pages", "3-5"),
        ("source.page5.pdf_sha256", off_balance.source_pdf_sha256),
        ("source.page5.ocr_sha256", off_balance.source_ocr_sha256),
        ("source.page5.render_sha256", off_balance.source_render_sha256),
        ("source.page5.policy_sha256", off_balance.policy_sha256),
        ("cross_statement_note_linkage.value_backfill_allowed", False),
        ("fully_verified", False),
        ("production_approved", False),
    )
    for values in rows:
        sheet.append(values)


def _verify_serialized(workbook_bytes: bytes, template_snapshot: tuple) -> None:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise CDKTDevelopmentExportError("serialized CDKT workbook cannot be reopened") from exc
    try:
        if workbook.sheetnames != [
            "CDKT",
            "PROVENANCE",
            "VALIDATION_DIAGNOSTICS",
            "RUN_METADATA",
        ] or _has_formula(workbook):
            raise CDKTDevelopmentExportError("serialized CDKT workbook contract drifted")
        sheet = workbook["CDKT"]
        actual_snapshot = tuple(
            tuple(
                (
                    sheet.cell(row, column).value,
                    sheet.cell(row, column).data_type,
                    sheet.cell(row, column).style_id,
                )
                for column in range(1, 4)
            )
            for row in range(1, CDKT_TEMPLATE_MAX_ROW + 1)
        )
        if actual_snapshot != template_snapshot:
            raise CDKTDevelopmentExportError("CDKT template A:C columns were not preserved")
        if (
            sum(
                sheet.cell(row, column).value is not None
                for row in range(2, CDKT_TEMPLATE_MAX_ROW + 1)
                for column in (4, 5)
            )
            != CDKT_EXPORTED_VALUE_COUNT
            or workbook["PROVENANCE"].max_row != 151
        ):
            raise CDKTDevelopmentExportError("serialized CDKT coverage drifted")
    finally:
        workbook.close()


def _resolve_project_root(*input_paths: Path) -> Path:
    roots = {
        parent
        for input_path in input_paths
        for parent in (Path(input_path).resolve().parent, *Path(input_path).resolve().parents)
        if (parent / CDKT_OFF_BALANCE_PAGE5_POLICY_RELATIVE_PATH).is_file()
    }
    if len(roots) != 1:
        raise CDKTDevelopmentExportError("cannot resolve unique CDKT project root")
    return roots.pop()


def build_cdkt_development_artifacts(
    *,
    template_path: Path,
    sealed_workbook_path: Path,
    workbook_name: str,
) -> CDKTDevelopmentArtifacts:
    """Build deterministic development workbook and coverage bytes in memory."""

    if Path(workbook_name).name != workbook_name or not workbook_name.lower().endswith(".xlsx"):
        raise CDKTDevelopmentExportError("workbook name must be a local .xlsx filename")
    project_root = _resolve_project_root(template_path, sealed_workbook_path)
    off_balance = parse_cdkt_off_balance_page5(
        load_cdkt_off_balance_page5_policy(
            project_root / CDKT_OFF_BALANCE_PAGE5_POLICY_RELATIVE_PATH,
            project_root,
        )
    )
    try:
        template_bytes = Path(template_path).read_bytes()
        sealed_bytes = Path(sealed_workbook_path).read_bytes()
    except OSError as exc:
        raise CDKTDevelopmentExportError("cannot read CDKT development inputs") from exc
    template, workbook = _load_inputs(template_bytes, sealed_bytes)
    try:
        template_snapshot, _rows = _update_main_sheet(template, workbook, off_balance)
        _update_provenance(workbook, off_balance)
        _append_business_diagnostics(workbook, off_balance)
        _replace_metadata(workbook, off_balance)
        workbook_bytes = _deterministic_workbook_bytes(workbook)
    finally:
        template.close()
        workbook.close()
    _verify_serialized(workbook_bytes, template_snapshot)
    workbook_sha256 = _sha256(workbook_bytes)
    coverage = {
        "artifact_type": "CDKT_BUSINESS_SCHEMA_DEVELOPMENT_COVERAGE",
        "authority": {
            "components": {
                "off_balance_page_5": (
                    "HASH_BOUND_SOURCE_PDF_RENDER_OCR_PLUS_AUDITED_SCHEMA_ALIASES"
                ),
                "statement_pages_3_4": (
                    "USER_CONFIRMED_MAPPING_AND_VISIBLE_VALUE_PLUS_ACCOUNTING_DERIVATION"
                ),
            },
            "fully_verified": False,
            "production_approved": False,
            "scope": "MIXED_USER_CONFIRMED_AND_AUDITED_SOURCE_EVIDENCE",
        },
        "coverage": {
            "derived_value_count": CDKT_DERIVED_VALUE_COUNT,
            "exported_value_count": CDKT_EXPORTED_VALUE_COUNT,
            "mapped_schema_count": CDKT_MAPPED_SCHEMA_COUNT,
            "not_observed_schema_count": CDKT_NOT_OBSERVED_SCHEMA_COUNT,
            "observed_value_count": CDKT_OBSERVED_VALUE_COUNT,
            "physical_cell_count": CDKT_PHYSICAL_CELL_COUNT,
            "physical_status_counts": {"BLANK": 13, "DASH": 5, "VALUE": 132},
            "reconciled_schema_count": CDKT_SCHEMA_COUNT,
            "source_row_count": CDKT_SOURCE_ROW_COUNT,
            "unresolved_schema_count": CDKT_UNRESOLVED_SCHEMA_COUNT,
        },
        "decisions": {
            "Q001": {"report_norm_id": 5712, "status": "RESOLVED_BY_USER"},
            "Q002": {"report_norm_id": 4304, "status": "STRUCTURAL_REPEAT"},
            "Q003": {"report_norm_id": 5712, "status": "STRUCTURAL_REPEAT"},
            "Q004": {"report_norm_id": 5699, "status": "RESOLVED_BY_USER"},
            "Q005": {
                "status": "TARGET_STATEMENT_BOUNDARY_CLARIFIED_NO_VALUE_BACKFILL",
                "target_status": "NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES",
            },
            "Q006": {"report_norm_id": 4350, "status": "SCHEMA_LABEL_CORRECTED"},
            "Q010": {
                "canonical_value_vnd": _USER_CONFIRMED_4363_CURRENT_VND,
                "report_norm_id": 4363,
                "status": "RESOLVED_BY_USER",
            },
            "MBB_PROVISION_SCHEMA_EVOLUTION": {
                "status": "SOURCE_VISIBLE_BROADER_ITEMS_REASSIGNED",
                "reassignments": [
                    {
                        "from_report_norm_id": source_id,
                        "source_row_id": source_row_id,
                        "to_report_norm_id": target_id,
                    }
                    for target_id, (source_id, source_row_id, _values) in sorted(
                        CDKT_SCHEMA_EVOLUTION_OWNERSHIP.items()
                    )
                ],
            },
        },
        "cross_statement_note_links": [
            {
                "cdkt_report_norm_id": cdkt_id,
                "note_pdf_page": 30,
                "note_report_norm_id": note_id,
                "note_statement": "TM",
                "relation": relation,
                "value_backfill_allowed": False,
            }
            for cdkt_id, note_id, relation in _CROSS_STATEMENT_NOTE_LINKS
        ],
        "off_balance_page5": {
            "accounting_checks_passed": list(off_balance.accounting_checks_passed),
            "schema_accounting_checks": [
                {
                    "target_report_norm_id": check.target_report_norm_id,
                    "physical_component_report_norm_ids": list(check.component_report_norm_ids),
                    "schema_component_report_norm_ids": list(
                        check.schema_component_report_norm_ids
                    ),
                    "compressed_unobserved_subtotal": {
                        "report_norm_id": check.compressed_unobserved_subtotal.report_norm_id,
                        "leaf_report_norm_ids": list(
                            check.compressed_unobserved_subtotal.leaf_report_norm_ids
                        ),
                        "imputed": False,
                    },
                    "operator": check.operator,
                }
                for check in off_balance.accounting_checks
            ],
            "mapped_report_norm_ids": sorted(CDKT_MBB_MAPPED_OFF_BALANCE_SCHEMA_IDS),
            "physical_cell_count": off_balance.physical_cell_count,
            "policy_path": off_balance.policy_path,
            "policy_sha256": off_balance.policy_sha256,
            "source_ocr_path": off_balance.source_ocr_path,
            "source_ocr_sha256": off_balance.source_ocr_sha256,
            "source_pdf_path": off_balance.source_pdf_path,
            "source_pdf_sha256": off_balance.source_pdf_sha256,
            "source_render_path": off_balance.source_render_path,
            "source_render_sha256": off_balance.source_render_sha256,
            "source_row_count": off_balance.source_row_count,
        },
        "formula_checks": {
            "4325_components": {"passed": True, "status": "DERIVED_FROM_COMPONENTS"},
            "5712_equals_4325_plus_4306": {
                "passed": True,
                "status": "VISIBLE_TOTAL_WITH_4306_NOT_OBSERVED",
            },
            "4305_equals_4304_plus_5712": {"passed": True},
        },
        "not_observed_report_norm_ids": sorted(_NOT_OBSERVED_IDS),
        "not_observed_on_target_statement_pages_report_norm_ids": sorted(
            _NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES_IDS
        ),
        "workbook": {
            "filename": workbook_name,
            "sha256": workbook_sha256,
            "size_bytes": len(workbook_bytes),
        },
    }
    coverage_bytes = _canonical_json_bytes(coverage)
    return CDKTDevelopmentArtifacts(
        workbook_bytes=workbook_bytes,
        coverage_bytes=coverage_bytes,
        workbook_sha256=workbook_sha256,
        coverage_sha256=_sha256(coverage_bytes),
        observed_value_count=CDKT_OBSERVED_VALUE_COUNT,
        derived_value_count=CDKT_DERIVED_VALUE_COUNT,
        exported_value_count=CDKT_EXPORTED_VALUE_COUNT,
    )


def _fsync_directory(directory: Path) -> None:
    flags = os.O_RDONLY | getattr(os, "O_DIRECTORY", 0) | getattr(os, "O_CLOEXEC", 0)
    descriptor = os.open(directory, flags)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _write_exclusive(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if os.path.lexists(path):
        raise CDKTDevelopmentExportError(f"refusing to overwrite CDKT output: {path}")
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    created = False
    try:
        descriptor = os.open(path, flags, 0o644)
        created = True
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(data)
            stream.flush()
            os.fsync(stream.fileno())
        _fsync_directory(path.parent)
    except Exception:
        if created:
            path.unlink(missing_ok=True)
            _fsync_directory(path.parent)
        raise


def export_cdkt_development(
    *,
    template_path: Path,
    sealed_workbook_path: Path,
    workbook_path: Path,
    coverage_path: Path,
) -> CDKTDevelopmentExportResult:
    """Write the two development outputs, refusing overwrite of either."""

    workbook_path = Path(workbook_path)
    coverage_path = Path(coverage_path)
    if workbook_path.parent != coverage_path.parent or workbook_path == coverage_path:
        raise CDKTDevelopmentExportError("CDKT outputs must be distinct siblings")
    if os.path.lexists(workbook_path) or os.path.lexists(coverage_path):
        raise CDKTDevelopmentExportError("CDKT development output already exists")
    artifacts = build_cdkt_development_artifacts(
        template_path=template_path,
        sealed_workbook_path=sealed_workbook_path,
        workbook_name=workbook_path.name,
    )
    _write_exclusive(workbook_path, artifacts.workbook_bytes)
    try:
        _write_exclusive(coverage_path, artifacts.coverage_bytes)
    except Exception:
        workbook_path.unlink(missing_ok=True)
        _fsync_directory(workbook_path.parent)
        raise
    return CDKTDevelopmentExportResult(
        workbook_path=workbook_path.as_posix(),
        coverage_path=coverage_path.as_posix(),
        workbook_sha256=artifacts.workbook_sha256,
        coverage_sha256=artifacts.coverage_sha256,
        workbook_size_bytes=len(artifacts.workbook_bytes),
        coverage_size_bytes=len(artifacts.coverage_bytes),
        observed_value_count=artifacts.observed_value_count,
        derived_value_count=artifacts.derived_value_count,
        exported_value_count=artifacts.exported_value_count,
    )


__all__ = [
    "CDKTDevelopmentArtifacts",
    "CDKTDevelopmentExportError",
    "CDKTDevelopmentExportResult",
    "CDKT_TEMPLATE_MAX_ROW",
    "CDKT_VPB_SCHEMA_IDS",
    "build_cdkt_development_artifacts",
    "export_cdkt_development",
]
