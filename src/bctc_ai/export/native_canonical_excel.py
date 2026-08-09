from __future__ import annotations

import hashlib
import json
import os
import re
import stat
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bctc_ai.export.canonical_xlsx import (
    CANONICAL_CORE_TIMESTAMP,
    EXCEL_CELL_TEXT_LIMIT,
    append_literal_row,
    deterministic_workbook_bytes,
    workbook_has_formula,
)
from bctc_ai.mapping.native_canonical import load_registered_native_canonical_mapping
from bctc_ai.rows.native_statement import load_registered_native_statement_rows

EXPORT_POLICY_RELATIVE_PATH = Path("config/export/native-canonical-excel-v1.yaml")
SHEET_NAMES = (
    "SOURCE_ROWS",
    "CELLS",
    "SCHEMA_COVERAGE",
    "NEW_ITEM_PROPOSALS",
    "VALIDATION",
    "RUN_METADATA",
)

_EXPORT_POLICY = "REGISTERED_NATIVE_CANONICAL_EXCEL_V1"
_EXPORT_CLAIM = "SOURCE_FAITHFUL_CANONICAL_DISPOSITIONS_AND_SCHEMA_COVERAGE_EXCEL"
_EXPORT_STATUS = "ACCEPTED_NATIVE_CANONICAL_EXCEL_EXPORT"
_RECEIPT_TYPE = "REGISTERED_NATIVE_CANONICAL_EXCEL_PROVENANCE_V1"
_MAPPING_FORMAT = "REGISTERED_NATIVE_CANONICAL_MAPPING_RESULT_V1"
_MAPPING_POLICY = "REGISTERED_NATIVE_CANONICAL_MAPPING_V1"
_MAPPING_CLAIM = "SOURCE_ROW_CANONICAL_DISPOSITIONS_AND_SCHEMA_GAP_PROPOSALS"
_MAPPING_STATUS = "ACCEPTED_NATIVE_CANONICAL_MAPPING"
_ROWS_FORMAT = "REGISTERED_NATIVE_STATEMENT_ROWS_RESULT_V1"
_ROWS_POLICY = "REGISTERED_NATIVE_STATEMENT_ROWS_V1"
_ROWS_CLAIM = "UNMAPPED_SOURCE_ROWS_AND_CELLS_ONLY"
_ROWS_STATUS = "ACCEPTED_NATIVE_STATEMENT_ROWS"
_DISPOSITIONS = (
    "EXISTING_ITEM",
    "NEW_ITEM_PROPOSAL",
    "AMBIGUOUS",
    "UNRESOLVED",
    "STRUCTURAL",
)
_TERMINAL_OUTCOMES = (
    "OBSERVED_VALUE",
    "OBSERVED_ZERO",
    "DASH",
    "BLANK",
    "NOT_OBSERVED",
    "NOT_APPLICABLE",
    "AMBIGUOUS",
    "UNRESOLVED",
)
_ALLOWED_ROLE_SEQUENCE = (
    "LOGIC_DEVELOPMENT",
    "CALIBRATION",
    "VALIDATION",
    "PRODUCTION_INPUT",
)
_ALLOWED_ROLES = frozenset(_ALLOWED_ROLE_SEQUENCE)
_FORBIDDEN_ROLES = frozenset({"UNTOUCHED_HOLDOUT"})
_ROLE_DIRECTORIES = {
    "LOGIC_DEVELOPMENT": "output/development",
    "CALIBRATION": "output/calibration",
    "VALIDATION": "output/validation",
    "PRODUCTION_INPUT": "output/production",
}
_RUNTIME_INPUT_ALLOWLIST = (
    "REGISTERED_NATIVE_CANONICAL_MAPPING_JSON",
    "REGISTERED_NATIVE_STATEMENT_ROWS_JSON",
    "PRODUCER_VERSIONED_SNAPSHOTS_EMBEDDED_IN_MAPPING",
    "SOURCE_PDF",
    "SOURCE_REGISTRY",
    "DATASET_ROLE_REGISTRY",
    "MAPPING_STRICT_LOADER",
    "NATIVE_ROWS_STRICT_LOADER",
    "THIS_EXPORT_POLICY",
    "EXPORT_IMPLEMENTATION",
)
_IMPLEMENTATION_PATHS = (
    "src/bctc_ai/export/canonical_xlsx.py",
    "src/bctc_ai/export/native_canonical_excel.py",
    "src/bctc_ai/mapping/native_canonical.py",
    "src/bctc_ai/rows/native_statement.py",
)
_SHA256 = re.compile(r"^[0-9a-f]{64}$")
_ROW_ORDINAL = re.compile(r":row-(?P<ordinal>[0-9]{4})$")
_JSON_CHUNK_SIZE = 30_000

SOURCE_ROW_HEADERS = (
    "SourceOrder",
    "Page",
    "PageRowOrder",
    "Statement",
    "PresentationScope",
    "SourceBucket",
    "WithinFinancialTableSpan",
    "RowId",
    "RowType",
    "SourceRowStatus",
    "RawLabel",
    "NormalizedLabel",
    "RawNoteReference",
    "NoteReference",
    "IndentationText",
    "CellCount",
    "Disposition",
    "SelectedReportNormId",
    "SelectedCanonicalName",
    "CandidateReportNormIdsJson",
    "MatchBasis",
    "MatchedSchemaLabel",
    "AliasAuthorityType",
    "AliasAuthorityEvidenceSha256",
    "SourceParentRowId",
    "SchemaParentReportNormId",
    "SchemaHierarchyLevel",
    "SchemaDisplayOrder",
    "EquationEvidenceIndexesJson",
    "ConflictEvidenceIndexesJson",
    "Reason",
    "SourceRowSha256",
    "SourceCellsSha256",
    "SourceCellIdsJson",
    "WarningsJson",
    "SourceProvenanceJson",
    "SourceRowEnvelopeJson",
    "SourceDispositionJson",
)

CELL_HEADERS = (
    "SourceOrder",
    "Page",
    "PageRowOrder",
    "Statement",
    "PresentationScope",
    "WithinFinancialTableSpan",
    "RowId",
    "Disposition",
    "SelectedReportNormId",
    "CellOrder",
    "CellId",
    "AxisId",
    "RawText",
    "NormalizedText",
    "ValueText",
    "ValueJson",
    "Observation",
    "SourceStatus",
    "SignEvidenceJson",
    "ParseReason",
    "BboxJson",
    "RunId",
    "AxisDistanceText",
    "CellProvenanceJson",
    "RawHeader",
    "Unit",
    "UnitMultiplierJson",
    "PeriodStart",
    "PeriodEnd",
    "PeriodType",
    "DurationMonthsJson",
    "CurrentOrComparative",
    "RestatedJson",
    "HeaderConfidenceText",
    "HeaderJson",
    "CellSha256",
    "CellJson",
)

SCHEMA_COVERAGE_HEADERS = (
    "SchemaOrder",
    "ReportNormId",
    "CanonicalName",
    "Statement",
    "Section",
    "ParentReportNormId",
    "HierarchyLevel",
    "DisplayOrder",
    "ItemTypeOrScopeJson",
    "ApplicableScopeJson",
    "CashFlowBranch",
    "DocumentReportingScope",
    "SourceStatementScope",
    "BlockExhaustive",
    "TerminalOutcome",
    "ObservationBasis",
    "SourceRowId",
    "CandidateSourceRowIdsJson",
    "BlockEvidenceSha256",
    "SourceScopeEvidenceJson",
    "ChildrenJson",
    "PreviousReportNormId",
    "NextReportNormId",
    "SchemaItemSha256",
    "SchemaDispositionSha256",
    "SchemaItemJson",
    "SchemaDispositionJson",
)

NEW_ITEM_PROPOSAL_HEADERS = (
    "ProposalOrder",
    "ProposalKey",
    "CanonicalLabel",
    "Statement",
    "Section",
    "ParentKind",
    "ParentReportNormId",
    "ParentProposalKey",
    "HierarchyLevel",
    "InsertAfterReportNormId",
    "InsertBeforeReportNormId",
    "SourceRowId",
    "SourcePage",
    "VisibleSourceLabel",
    "ReasonExistingItemsInsufficient",
    "AliasesJson",
    "EquationEvidenceJson",
    "AllocatedReportNormId",
    "ProposalSha256",
    "ProposalJson",
)

VALIDATION_HEADERS = (
    "ValidationOrder",
    "RecordType",
    "RecordIndex",
    "Statement",
    "Status",
    "AffectedRowIdsJson",
    "PartNumber",
    "PartCount",
    "RecordSha256",
    "RecordJsonPart",
)

RUN_METADATA_HEADERS = (
    "Key",
    "PartNumber",
    "PartCount",
    "ValueSha256",
    "ValueJsonPart",
)


class NativeCanonicalExcelExportError(RuntimeError):
    """Raised when the paired canonical-mapping Excel export fails closed."""


@dataclass(frozen=True, slots=True)
class NativeCanonicalExcelPolicy:
    path: Path
    sha256: str
    name: str
    claim_boundary: str
    accepted_mapping: dict[str, object]
    accepted_rows: dict[str, object]
    allowed_roles: frozenset[str]
    forbidden_roles: frozenset[str]
    role_directories: dict[str, str]
    workbook_creator: str


@dataclass(frozen=True, slots=True)
class ArtifactIdentity:
    path: str
    sha256: str
    size_bytes: int


@dataclass(frozen=True, slots=True)
class NativeCanonicalExcelArtifacts:
    workbook_bytes: bytes
    provenance_bytes: bytes
    workbook_sha256: str
    provenance_sha256: str
    summary: dict[str, object]


@dataclass(frozen=True, slots=True)
class NativeCanonicalExcelExportResult:
    workbook_path: Path
    provenance_path: Path
    workbook_sha256: str
    provenance_sha256: str
    workbook_size_bytes: int
    provenance_size_bytes: int
    summary: dict[str, object]


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _canonical_json_bytes(value: object) -> bytes:
    try:
        return (
            json.dumps(value, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise NativeCanonicalExcelExportError("export input is not canonical JSON data") from exc


def _compact_json(value: object) -> str:
    try:
        rendered = json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError) as exc:
        raise NativeCanonicalExcelExportError("workbook value is not canonical JSON") from exc
    if len(rendered) > EXCEL_CELL_TEXT_LIMIT:
        raise NativeCanonicalExcelExportError("canonical record exceeds Excel's cell text limit")
    return rendered


def _record_hash(value: object) -> str:
    return _sha256(_compact_json_unbounded(value).encode("utf-8"))


def _compact_json_unbounded(value: object) -> str:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError) as exc:
        raise NativeCanonicalExcelExportError("workbook value is not canonical JSON") from exc


def _mapping(value: object, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise NativeCanonicalExcelExportError(f"{label} must be an object")
    return value


def _exact_mapping(
    value: object,
    expected_keys: set[str],
    label: str,
) -> Mapping[str, Any]:
    record = _mapping(value, label)
    if set(record) != expected_keys:
        raise NativeCanonicalExcelExportError(f"{label} fields drifted")
    return record


def _records(value: object, label: str) -> Sequence[Mapping[str, Any]]:
    if not isinstance(value, list) or any(not isinstance(record, Mapping) for record in value):
        raise NativeCanonicalExcelExportError(f"{label} must be an array of objects")
    return value


def _project_path(project_root: Path, path: Path) -> Path:
    return path.resolve() if path.is_absolute() else (project_root / path).resolve()


def _relative(project_root: Path, path: Path, label: str) -> str:
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except ValueError as exc:
        raise NativeCanonicalExcelExportError(f"{label} must stay inside the project root") from exc


def _canonical_relative_path(value: object, label: str) -> str:
    if not isinstance(value, str) or not value or "\\" in value:
        raise NativeCanonicalExcelExportError(f"{label} must be a canonical relative POSIX path")
    path = PurePosixPath(value)
    parts = value.split("/")
    if (
        path.is_absolute()
        or re.match(r"^[A-Za-z]:", value)
        or any(part in {"", ".", ".."} for part in parts)
        or path.as_posix() != value
    ):
        raise NativeCanonicalExcelExportError(f"{label} must be a canonical relative POSIX path")
    return value


def _plain_filename(value: str, label: str) -> str:
    if not isinstance(value, str) or not value or Path(value).name != value:
        raise NativeCanonicalExcelExportError(f"{label} must be a plain filename")
    return value


def _validate_identity(identity: ArtifactIdentity, label: str) -> None:
    _canonical_relative_path(identity.path, f"{label} path")
    if (
        not isinstance(identity.sha256, str)
        or _SHA256.fullmatch(identity.sha256) is None
        or isinstance(identity.size_bytes, bool)
        or not isinstance(identity.size_bytes, int)
        or identity.size_bytes < 1
    ):
        raise NativeCanonicalExcelExportError(f"{label} identity is invalid")


def _expected_policy_payload() -> dict[str, object]:
    return {
        "version": 1,
        "policy": _EXPORT_POLICY,
        "claim_boundary": _EXPORT_CLAIM,
        "accepted_inputs": {
            "mapping": {
                "format_version": _MAPPING_FORMAT,
                "policy": _MAPPING_POLICY,
                "claim_boundary": _MAPPING_CLAIM,
                "status": _MAPPING_STATUS,
            },
            "native_rows": {
                "format_version": _ROWS_FORMAT,
                "policy": _ROWS_POLICY,
                "claim_boundary": _ROWS_CLAIM,
                "status": _ROWS_STATUS,
            },
            "trusted_sha256_required": True,
            "exact_mapping_to_native_rows_pair_required": True,
        },
        "allowed_dataset_roles": list(_ALLOWED_ROLE_SEQUENCE),
        "forbidden_dataset_roles": sorted(_FORBIDDEN_ROLES),
        "role_directories": _ROLE_DIRECTORIES,
        "workbook": {
            "sheets": list(SHEET_NAMES),
            "creator": "bctc-ai/native-canonical-excel-v1",
            "deterministic_xlsx": True,
            "formulas_allowed": False,
            "imputation_allowed": False,
            "force_mapping_allowed": False,
            "report_norm_id_allocation_allowed": False,
            "source_cell_values_as_json_text": True,
            "embedded_producer_schema_snapshot_is_authority": True,
            "current_mutable_schema_inputs_allowed": False,
            "preserve_full_record_json": True,
        },
        "publication": {
            "paired_workbook_and_provenance": True,
            "workbook_written_before_provenance_completion_marker": True,
            "exclusive_no_overwrite": True,
            "sibling_outputs_required": True,
            "absolute_project_paths_allowed": False,
            "canonical_provenance_json": True,
        },
        "role_isolation": {
            "runtime_input_allowlist": list(_RUNTIME_INPUT_ALLOWLIST),
            "strict_mapping_contract_revalidated": True,
            "strict_native_rows_contract_revalidated": True,
            "current_mutable_schema_loaded": False,
            "template_inputs_loaded": False,
            "historical_values_loaded": False,
            "role_a_outputs_loaded": False,
            "human_review_outputs_loaded": False,
        },
    }


def load_native_canonical_excel_policy(
    path: Path, project_root: Path
) -> NativeCanonicalExcelPolicy:
    project_root = project_root.resolve()
    path = _project_path(project_root, path)
    _relative(project_root, path, "native canonical Excel policy")
    if path != (project_root / EXPORT_POLICY_RELATIVE_PATH).resolve():
        raise NativeCanonicalExcelExportError(
            f"native canonical Excel requires canonical policy {EXPORT_POLICY_RELATIVE_PATH}"
        )
    try:
        encoded = path.read_bytes()
        payload = yaml.safe_load(encoded)
    except (OSError, yaml.YAMLError) as exc:
        raise NativeCanonicalExcelExportError(
            f"cannot load native canonical Excel policy: {path}"
        ) from exc
    if payload != _expected_policy_payload():
        raise NativeCanonicalExcelExportError("native canonical Excel policy drifted")
    accepted = payload["accepted_inputs"]
    workbook = payload["workbook"]
    return NativeCanonicalExcelPolicy(
        path=path,
        sha256=_sha256(encoded),
        name=_EXPORT_POLICY,
        claim_boundary=_EXPORT_CLAIM,
        accepted_mapping=dict(accepted["mapping"]),
        accepted_rows=dict(accepted["native_rows"]),
        allowed_roles=_ALLOWED_ROLES,
        forbidden_roles=_FORBIDDEN_ROLES,
        role_directories=dict(_ROLE_DIRECTORIES),
        workbook_creator=str(workbook["creator"]),
    )


def _validate_implementation_ledger(
    ledger: Sequence[Mapping[str, object]],
) -> tuple[dict[str, object], ...]:
    if isinstance(ledger, (str, bytes, bytearray)):
        raise NativeCanonicalExcelExportError("export implementation ledger is invalid")
    result: list[dict[str, object]] = []
    for record in ledger:
        if not isinstance(record, Mapping) or set(record) != {"path", "sha256", "size_bytes"}:
            raise NativeCanonicalExcelExportError("export implementation identity is malformed")
        path = _canonical_relative_path(record["path"], "export implementation path")
        digest = record["sha256"]
        size = record["size_bytes"]
        if (
            not isinstance(digest, str)
            or _SHA256.fullmatch(digest) is None
            or isinstance(size, bool)
            or not isinstance(size, int)
            or size < 0
        ):
            raise NativeCanonicalExcelExportError("export implementation identity is invalid")
        result.append({"path": path, "sha256": digest, "size_bytes": size})
    paths = [str(record["path"]) for record in result]
    if paths != sorted(set(paths)):
        raise NativeCanonicalExcelExportError(
            "export implementation ledger must be unique and path-sorted"
        )
    return tuple(result)


def _validate_payload_identity(
    mapping_payload: Mapping[str, Any],
    rows_payload: Mapping[str, Any],
    *,
    mapping_identity: ArtifactIdentity,
    rows_identity: ArtifactIdentity,
    policy: NativeCanonicalExcelPolicy,
) -> tuple[Mapping[str, Any], str]:
    for key, expected in policy.accepted_mapping.items():
        if mapping_payload.get(key) != expected:
            raise NativeCanonicalExcelExportError(f"mapping input {key} is not accepted")
    for key, expected in policy.accepted_rows.items():
        if rows_payload.get(key) != expected:
            raise NativeCanonicalExcelExportError(f"native-row input {key} is not accepted")
    if mapping_payload.get("source") != rows_payload.get("source"):
        raise NativeCanonicalExcelExportError("mapping/native-row source envelopes differ")
    source = _mapping(rows_payload.get("source"), "native-row source")
    role = source.get("dataset_role")
    if not isinstance(role, str) or role not in policy.allowed_roles:
        raise NativeCanonicalExcelExportError("paired input dataset role is not exportable")
    native_rows = _mapping(mapping_payload.get("native_rows"), "mapping native-row envelope")
    expected_pair = {
        "path": rows_identity.path,
        "sha256": rows_identity.sha256,
        "size_bytes": rows_identity.size_bytes,
        "format_version": rows_payload.get("format_version"),
        "policy": rows_payload.get("policy"),
        "claim_boundary": rows_payload.get("claim_boundary"),
        "status": rows_payload.get("status"),
        "run_id": rows_payload.get("run_id"),
        "producer_git_commit": _mapping(rows_payload.get("code"), "native-row code").get("commit"),
        "denominator": "ALL_RECONSTRUCTED_SOURCE_ROWS",
    }
    if native_rows != expected_pair:
        raise NativeCanonicalExcelExportError(
            "mapping does not bind the supplied trusted native-row artifact exactly"
        )
    if mapping_identity.path == rows_identity.path:
        raise NativeCanonicalExcelExportError("mapping and native-row paths must be distinct")
    return source, role


def _receipt_input_record(
    identity: ArtifactIdentity,
    payload: Mapping[str, Any],
) -> dict[str, object]:
    return {
        "path": identity.path,
        "sha256": identity.sha256,
        "size_bytes": identity.size_bytes,
        "claim_boundary": payload.get("claim_boundary"),
        "format_version": payload.get("format_version"),
        "policy": payload.get("policy"),
        "run_id": payload.get("run_id"),
        "status": payload.get("status"),
    }


def _receipt_isolation() -> dict[str, object]:
    return {
        "current_mutable_schema_loaded": False,
        "historical_values_loaded": False,
        "human_review_outputs_loaded": False,
        "role_a_outputs_loaded": False,
        "runtime_input_allowlist": list(_RUNTIME_INPUT_ALLOWLIST),
        "strict_mapping_contract_revalidated": True,
        "strict_native_rows_contract_revalidated": True,
        "template_inputs_loaded": False,
    }


def _receipt_policy(policy: NativeCanonicalExcelPolicy) -> dict[str, object]:
    return {
        "export_policy": policy.name,
        "export_policy_sha256": policy.sha256,
        "force_mapping": False,
        "imputation": False,
        "producer_schema_snapshot_authoritative": True,
        "report_norm_id_allocation": False,
    }


def _row_ordinal(row: Mapping[str, Any]) -> int:
    row_id = row.get("row_id")
    match = _ROW_ORDINAL.search(row_id) if isinstance(row_id, str) else None
    if match is None:
        raise NativeCanonicalExcelExportError("source row ID lacks a stable ordinal")
    return int(match.group("ordinal"))


@dataclass(frozen=True, slots=True)
class _JoinedSourceRow:
    source_order: int
    page: int
    page_row_order: int
    statement: str
    scope: str
    within_span: bool
    row: Mapping[str, Any]
    page_record: Mapping[str, Any]


def _source_rows(rows_payload: Mapping[str, Any]) -> tuple[_JoinedSourceRow, ...]:
    pages = _records(rows_payload.get("pages"), "native-row pages")
    if not pages:
        raise NativeCanonicalExcelExportError("native-row input has no pages")
    joined: list[_JoinedSourceRow] = []
    seen: set[str] = set()
    source_order = 0
    for page_record in pages:
        page = page_record.get("page")
        statement = page_record.get("statement_type")
        scope = page_record.get("scope")
        if (
            isinstance(page, bool)
            or not isinstance(page, int)
            or not isinstance(statement, str)
            or not isinstance(scope, str)
        ):
            raise NativeCanonicalExcelExportError("native-row page identity is malformed")
        inside = _records(page_record.get("rows"), "financial-table-span rows")
        outside = _records(
            page_record.get("outside_financial_table_span_rows"),
            "outside-financial-table-span rows",
        )
        merged = [(row, True) for row in inside] + [(row, False) for row in outside]
        merged.sort(key=lambda item: _row_ordinal(item[0]))
        if len(merged) != page_record.get("reconstructed_row_count"):
            raise NativeCanonicalExcelExportError("native-row page denominator drifted")
        if len(inside) != page_record.get("financial_table_span_row_count"):
            raise NativeCanonicalExcelExportError("financial-table-span denominator drifted")
        if len(outside) != page_record.get("outside_financial_table_span_row_count"):
            raise NativeCanonicalExcelExportError("outside-span denominator drifted")
        ordinals = [_row_ordinal(row) for row, _ in merged]
        if ordinals != list(range(1, len(merged) + 1)):
            raise NativeCanonicalExcelExportError("page source-row ordinals are not contiguous")
        for row, within_span in merged:
            row_id = row.get("row_id")
            if not isinstance(row_id, str) or row_id in seen:
                raise NativeCanonicalExcelExportError("source row ID is invalid or duplicated")
            seen.add(row_id)
            if row.get("within_financial_table_span") is not within_span:
                raise NativeCanonicalExcelExportError("source row span bucket/flag drifted")
            if row.get("page") != page:
                raise NativeCanonicalExcelExportError("source row page binding drifted")
            source_order += 1
            joined.append(
                _JoinedSourceRow(
                    source_order=source_order,
                    page=page,
                    page_row_order=_row_ordinal(row),
                    statement=statement,
                    scope=scope,
                    within_span=within_span,
                    row=row,
                    page_record=page_record,
                )
            )
    return tuple(joined)


def _source_cell_join(row: _JoinedSourceRow, rows_sha256: str) -> dict[str, object]:
    cells = _records(row.row.get("cells"), "source row cells")
    cell_ids: list[str] = []
    for cell in cells:
        provenance = _mapping(cell.get("provenance"), "source cell provenance")
        axis_id = cell.get("axis_id")
        if provenance.get("row_id") != row.row.get("row_id"):
            raise NativeCanonicalExcelExportError("source cell provenance cannot join its row")
        if not isinstance(axis_id, str) or not axis_id:
            raise NativeCanonicalExcelExportError("source cell axis identity is invalid")
        cell_ids.append(f"{row.row['row_id']}:{axis_id}")
    if len(cell_ids) != len(set(cell_ids)):
        raise NativeCanonicalExcelExportError("source row repeats a cell axis")
    return {
        "native_rows_sha256": rows_sha256,
        "row_id": row.row["row_id"],
        "source_row_sha256": _record_hash(row.row),
        "source_cells_sha256": _record_hash(list(cells)),
        "source_cell_ids": cell_ids,
        "source_cell_count": len(cells),
    }


def _record(headers: Sequence[str], values: Sequence[object]) -> dict[str, object]:
    if len(headers) != len(values):
        raise NativeCanonicalExcelExportError("internal workbook row width drifted")
    if any(isinstance(value, str) and len(value) > EXCEL_CELL_TEXT_LIMIT for value in values):
        raise NativeCanonicalExcelExportError("workbook source text exceeds Excel's cell limit")
    return dict(zip(headers, values, strict=True))


def _json_or_none(value: object) -> str:
    return _compact_json(value)


def _scalar_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return _compact_json(value)


def _project_source_and_cells(
    mapping_payload: Mapping[str, Any],
    rows_payload: Mapping[str, Any],
    *,
    rows_sha256: str,
) -> tuple[list[dict[str, object]], list[dict[str, object]], tuple[_JoinedSourceRow, ...]]:
    source_rows = _source_rows(rows_payload)
    dispositions = _records(mapping_payload.get("source_dispositions"), "source dispositions")
    if len(dispositions) != len(source_rows):
        raise NativeCanonicalExcelExportError("source disposition denominator drifted")
    source_records: list[dict[str, object]] = []
    cell_records: list[dict[str, object]] = []
    seen_dispositions: set[str] = set()
    for source_row, disposition in zip(source_rows, dispositions, strict=True):
        row = source_row.row
        row_id = row.get("row_id")
        expected_fields = {
            "row_id": row_id,
            "source_order": source_row.source_order,
            "page": source_row.page,
            "page_row_order": source_row.page_row_order,
            "statement": source_row.statement,
            "scope": source_row.scope,
            "within_financial_table_span": source_row.within_span,
            "row_type": row.get("row_type"),
            "raw_label": row.get("raw_label"),
            "normalized_label": row.get("normalized_label"),
            "indentation": row.get("indentation"),
        }
        if any(disposition.get(key) != value for key, value in expected_fields.items()):
            raise NativeCanonicalExcelExportError("source disposition-to-row binding drifted")
        if not isinstance(row_id, str) or row_id in seen_dispositions:
            raise NativeCanonicalExcelExportError("source row received multiple dispositions")
        seen_dispositions.add(row_id)
        status = disposition.get("disposition")
        if status not in _DISPOSITIONS:
            raise NativeCanonicalExcelExportError("unknown source disposition")
        selected_id = disposition.get("selected_report_norm_id")
        if (status == "EXISTING_ITEM") != (type(selected_id) is int):
            raise NativeCanonicalExcelExportError(
                "source disposition ReportNormId authority drifted"
            )
        join = _source_cell_join(source_row, rows_sha256)
        if disposition.get("source_cell_join") != join:
            raise NativeCanonicalExcelExportError("source disposition cell hash join drifted")
        cells = _records(row.get("cells"), "source row cells")
        row_envelope = dict(row)
        row_envelope.pop("cells", None)
        source_records.append(
            _record(
                SOURCE_ROW_HEADERS,
                (
                    source_row.source_order,
                    source_row.page,
                    source_row.page_row_order,
                    source_row.statement,
                    source_row.scope,
                    (
                        "FINANCIAL_TABLE_SPAN"
                        if source_row.within_span
                        else "OUTSIDE_FINANCIAL_TABLE_SPAN"
                    ),
                    source_row.within_span,
                    row_id,
                    row.get("row_type"),
                    row.get("source_status"),
                    row.get("raw_label"),
                    row.get("normalized_label"),
                    row.get("raw_note_reference"),
                    row.get("note_reference"),
                    _scalar_text(row.get("indentation")),
                    len(cells),
                    status,
                    selected_id,
                    disposition.get("selected_canonical_name"),
                    _json_or_none(disposition.get("candidate_report_norm_ids")),
                    disposition.get("match_basis"),
                    disposition.get("matched_schema_label"),
                    disposition.get("alias_authority_type"),
                    disposition.get("alias_authority_evidence_sha256"),
                    disposition.get("source_parent_row_id"),
                    disposition.get("schema_parent_report_norm_id"),
                    disposition.get("schema_hierarchy_level"),
                    disposition.get("schema_display_order"),
                    _json_or_none(disposition.get("equation_evidence_indexes")),
                    _json_or_none(disposition.get("conflict_evidence_indexes")),
                    disposition.get("reason"),
                    join["source_row_sha256"],
                    join["source_cells_sha256"],
                    _json_or_none(join["source_cell_ids"]),
                    _json_or_none(row.get("warnings")),
                    _json_or_none(row.get("provenance")),
                    _compact_json(row_envelope),
                    _compact_json(disposition),
                ),
            )
        )
        headers = _records(source_row.page_record.get("headers"), "page headers")
        headers_by_axis: dict[str, Mapping[str, Any]] = {}
        for header in headers:
            axis = header.get("axis_id")
            if not isinstance(axis, str) or not axis or axis in headers_by_axis:
                raise NativeCanonicalExcelExportError("page header axis identity is invalid")
            headers_by_axis[axis] = header
        for cell_order, cell in enumerate(cells, start=1):
            axis_id = cell.get("axis_id")
            header = headers_by_axis.get(axis_id) if isinstance(axis_id, str) else None
            if header is None:
                raise NativeCanonicalExcelExportError("source cell lacks a unique header binding")
            cell_id = f"{row_id}:{axis_id}"
            cell_records.append(
                _record(
                    CELL_HEADERS,
                    (
                        source_row.source_order,
                        source_row.page,
                        source_row.page_row_order,
                        source_row.statement,
                        source_row.scope,
                        source_row.within_span,
                        row_id,
                        status,
                        selected_id,
                        cell_order,
                        cell_id,
                        axis_id,
                        cell.get("raw_text"),
                        cell.get("normalized_text"),
                        _scalar_text(cell.get("value")),
                        _json_or_none(cell.get("value")),
                        cell.get("observation"),
                        cell.get("source_status"),
                        _json_or_none(cell.get("sign_evidence")),
                        cell.get("parse_reason"),
                        _json_or_none(cell.get("bbox")),
                        cell.get("run_id"),
                        _scalar_text(cell.get("axis_distance")),
                        _json_or_none(cell.get("provenance")),
                        header.get("raw_header"),
                        header.get("unit"),
                        _json_or_none(header.get("unit_multiplier")),
                        header.get("period_start"),
                        header.get("period_end"),
                        header.get("period_type"),
                        _json_or_none(header.get("duration_months")),
                        header.get("current_or_comparative"),
                        _json_or_none(header.get("restated")),
                        _scalar_text(header.get("confidence")),
                        _compact_json(header),
                        _record_hash(cell),
                        _compact_json(cell),
                    ),
                )
            )
    return source_records, cell_records, source_rows


def _project_schema_coverage(
    mapping_payload: Mapping[str, Any],
) -> tuple[list[dict[str, object]], Sequence[Mapping[str, Any]]]:
    snapshots = _mapping(mapping_payload.get("producer_snapshots"), "producer snapshots")
    schema_snapshot = _mapping(snapshots.get("schema"), "producer schema snapshot")
    schema_items = _records(schema_snapshot.get("items"), "producer schema items")
    if schema_snapshot.get("items_sha256") != _record_hash(list(schema_items)):
        raise NativeCanonicalExcelExportError("producer schema snapshot hash drifted")
    dispositions = _records(mapping_payload.get("schema_dispositions"), "schema dispositions")
    if len(dispositions) != len(schema_items):
        raise NativeCanonicalExcelExportError("schema disposition denominator drifted")
    records: list[dict[str, object]] = []
    seen_ids: set[int] = set()
    for order, (item, disposition) in enumerate(
        zip(schema_items, dispositions, strict=True), start=1
    ):
        schema_id = item.get("schema_id")
        if type(schema_id) is not int or schema_id in seen_ids:
            raise NativeCanonicalExcelExportError("producer schema ReportNormId is invalid")
        seen_ids.add(schema_id)
        if (
            disposition.get("report_norm_id") != schema_id
            or disposition.get("canonical_name") != item.get("canonical_name")
            or disposition.get("statement") != item.get("statement_type")
            or disposition.get("display_order") != item.get("display_order")
            or disposition.get("parent_report_norm_id") != item.get("parent_id")
            or disposition.get("hierarchy_level") != item.get("hierarchy_level")
        ):
            raise NativeCanonicalExcelExportError(
                "schema disposition differs from its producer schema snapshot"
            )
        terminal = disposition.get("terminal_outcome")
        if terminal not in _TERMINAL_OUTCOMES:
            raise NativeCanonicalExcelExportError("unknown schema terminal outcome")
        records.append(
            _record(
                SCHEMA_COVERAGE_HEADERS,
                (
                    order,
                    schema_id,
                    item.get("canonical_name"),
                    item.get("statement_type"),
                    item.get("notes_section"),
                    item.get("parent_id"),
                    item.get("hierarchy_level"),
                    item.get("display_order"),
                    _json_or_none(item.get("scope")),
                    _json_or_none(disposition.get("applicable_scope")),
                    item.get("cash_flow_branch"),
                    disposition.get("document_reporting_scope"),
                    disposition.get("source_statement_scope"),
                    disposition.get("block_exhaustive"),
                    terminal,
                    disposition.get("observation_basis"),
                    disposition.get("source_row_id"),
                    _json_or_none(disposition.get("candidate_source_row_ids")),
                    disposition.get("block_evidence_sha256"),
                    _json_or_none(disposition.get("source_scope_evidence")),
                    _json_or_none(item.get("children")),
                    item.get("previous_id"),
                    item.get("next_id"),
                    _record_hash(item),
                    _record_hash(disposition),
                    _compact_json(item),
                    _compact_json(disposition),
                ),
            )
        )
    source_dispositions = _records(
        mapping_payload.get("source_dispositions"), "source dispositions"
    )
    selected = [
        (disposition.get("row_id"), disposition.get("selected_report_norm_id"))
        for disposition in source_dispositions
        if disposition.get("disposition") == "EXISTING_ITEM"
    ]
    selected_ids = [schema_id for _, schema_id in selected]
    if any(type(schema_id) is not int for schema_id in selected_ids) or len(selected_ids) != len(
        set(selected_ids)
    ):
        raise NativeCanonicalExcelExportError(
            "existing source dispositions do not have unique ReportNormIds"
        )
    candidate_ids: set[int] = set()
    for disposition in source_dispositions:
        raw_candidates = disposition.get("candidate_report_norm_ids")
        if not isinstance(raw_candidates, list) or any(
            type(candidate) is not int for candidate in raw_candidates
        ):
            raise NativeCanonicalExcelExportError("source candidate ReportNormIds are malformed")
        candidate_ids.update(raw_candidates)
    if not set(selected_ids).issubset(seen_ids) or not candidate_ids.issubset(seen_ids):
        raise NativeCanonicalExcelExportError("source disposition references unknown schema items")
    schema_disposition_by_id = {int(record["report_norm_id"]): record for record in dispositions}
    for row_id, schema_id in selected:
        schema_disposition = schema_disposition_by_id[int(schema_id)]
        if schema_disposition.get("source_row_id") != row_id or schema_disposition.get(
            "terminal_outcome"
        ) not in {"OBSERVED_VALUE", "OBSERVED_ZERO", "DASH", "BLANK"}:
            raise NativeCanonicalExcelExportError(
                "mapped source ownership differs from schema coverage"
            )
    return records, schema_items


def _project_new_item_proposals(
    mapping_payload: Mapping[str, Any], source_row_ids: set[str]
) -> list[dict[str, object]]:
    proposals = _records(mapping_payload.get("new_item_proposals"), "new-item proposals")
    proposal_keys: set[str] = set()
    records: list[dict[str, object]] = []
    for order, proposal in enumerate(proposals, start=1):
        proposal_key = proposal.get("proposal_key")
        if not isinstance(proposal_key, str) or not proposal_key or proposal_key in proposal_keys:
            raise NativeCanonicalExcelExportError("new-item proposal key is invalid or duplicated")
        proposal_keys.add(proposal_key)
        if proposal.get("report_norm_id") is not None:
            raise NativeCanonicalExcelExportError("new-item proposal allocated a ReportNormId")
        parent = _mapping(proposal.get("parent"), "new-item proposal parent")
        anchors = _mapping(
            proposal.get("display_order_anchors"), "new-item proposal display anchors"
        )
        source_evidence = _mapping(
            proposal.get("source_evidence"), "new-item proposal source evidence"
        )
        source_row_id = source_evidence.get("row_id")
        if source_row_id not in source_row_ids:
            raise NativeCanonicalExcelExportError("new-item proposal source evidence is unbound")
        records.append(
            _record(
                NEW_ITEM_PROPOSAL_HEADERS,
                (
                    order,
                    proposal_key,
                    proposal.get("canonical_label"),
                    proposal.get("statement"),
                    proposal.get("section"),
                    parent.get("kind"),
                    parent.get("report_norm_id"),
                    parent.get("proposal_key"),
                    proposal.get("hierarchy_level"),
                    anchors.get("insert_after_report_norm_id"),
                    anchors.get("insert_before_report_norm_id"),
                    source_row_id,
                    source_evidence.get("page"),
                    source_evidence.get("visible_label"),
                    proposal.get("reason_existing_items_are_insufficient"),
                    _json_or_none(proposal.get("possible_aliases")),
                    _json_or_none(proposal.get("equation_evidence")),
                    proposal.get("report_norm_id"),
                    _record_hash(proposal),
                    _compact_json(proposal),
                ),
            )
        )
    proposal_dispositions = {
        disposition.get("row_id")
        for disposition in _records(
            mapping_payload.get("source_dispositions"), "source dispositions"
        )
        if disposition.get("disposition") == "NEW_ITEM_PROPOSAL"
    }
    proposal_rows = {
        _mapping(proposal.get("source_evidence"), "new-item proposal source evidence").get("row_id")
        for proposal in proposals
    }
    if proposal_rows != proposal_dispositions:
        raise NativeCanonicalExcelExportError("new-item proposal/disposition ownership drifted")
    return records


def _chunk_record(
    *,
    order_start: int,
    record_type: str,
    record_index: str,
    statement: object,
    status: object,
    affected_row_ids: object,
    value: object,
) -> tuple[list[dict[str, object]], int]:
    rendered = _compact_json_unbounded(value)
    digest = _sha256(rendered.encode("utf-8"))
    parts = [
        rendered[index : index + _JSON_CHUNK_SIZE]
        for index in range(0, len(rendered), _JSON_CHUNK_SIZE)
    ] or [""]
    records: list[dict[str, object]] = []
    order = order_start
    for part_number, part in enumerate(parts, start=1):
        records.append(
            _record(
                VALIDATION_HEADERS,
                (
                    order,
                    record_type,
                    record_index,
                    statement,
                    status,
                    _json_or_none(affected_row_ids),
                    part_number,
                    len(parts),
                    digest,
                    part,
                ),
            )
        )
        order += 1
    return records, order


def _validation_items(
    mapping_payload: Mapping[str, Any], rows_payload: Mapping[str, Any]
) -> list[tuple[str, str, object]]:
    items: list[tuple[str, str, object]] = [
        ("LCTT_METHOD", "1", mapping_payload.get("lctt_method")),
    ]
    path_summaries = _mapping(mapping_payload.get("path_summaries"), "path summaries")
    items.extend(
        ("PATH_SUMMARY", str(statement), summary) for statement, summary in path_summaries.items()
    )
    for key, record_type in (
        ("alias_proposals", "ALIAS_PROPOSAL"),
        ("equations", "EQUATION"),
        ("conflicts", "CONFLICT"),
    ):
        items.extend(
            (record_type, str(index), record)
            for index, record in enumerate(_records(mapping_payload.get(key), key), start=1)
        )
    items.append(("MANDATORY_SEARCH", "1", mapping_payload.get("mandatory_search")))
    snapshots = _mapping(mapping_payload.get("producer_snapshots"), "producer snapshots")
    policy_snapshot = _mapping(snapshots.get("policy"), "producer policy snapshot")
    schema_snapshot = _mapping(snapshots.get("schema"), "producer schema snapshot")
    aliases_snapshot = _mapping(
        snapshots.get("accepted_aliases"), "producer accepted-alias snapshot"
    )
    coverage_snapshot = _mapping(snapshots.get("coverage"), "producer coverage snapshot")
    cash_snapshot = _mapping(snapshots.get("cash_flow_rules"), "producer cash-flow-rules snapshot")
    items.extend(
        [
            ("PRODUCER_POLICY_SNAPSHOT", "1", policy_snapshot),
            (
                "PRODUCER_SCHEMA_SNAPSHOT_IDENTITY",
                "1",
                {
                    "items_sha256": schema_snapshot.get("items_sha256"),
                    "item_count": len(
                        _records(schema_snapshot.get("items"), "producer schema items")
                    ),
                },
            ),
            (
                "PRODUCER_ACCEPTED_ALIAS_SNAPSHOT_IDENTITY",
                "1",
                {
                    "records_sha256": aliases_snapshot.get("records_sha256"),
                    "record_count": len(
                        _records(
                            aliases_snapshot.get("records"),
                            "producer accepted-alias records",
                        )
                    ),
                },
            ),
            ("PRODUCER_COVERAGE_SNAPSHOT", "1", coverage_snapshot),
            ("PRODUCER_CASH_FLOW_RULES_SNAPSHOT", "1", cash_snapshot),
        ]
    )
    alias_records = _records(aliases_snapshot.get("records"), "producer accepted-alias records")
    if aliases_snapshot.get("records_sha256") != _record_hash(list(alias_records)):
        raise NativeCanonicalExcelExportError("producer accepted-alias snapshot hash drifted")
    items.extend(
        ("ACCEPTED_ALIAS_AUTHORITY", str(index), record)
        for index, record in enumerate(alias_records, start=1)
    )
    for index, page in enumerate(_records(rows_payload.get("pages"), "native-row pages"), start=1):
        page_context = dict(page)
        page_context.pop("rows", None)
        page_context.pop("outside_financial_table_span_rows", None)
        items.append(("NATIVE_PAGE_CONTEXT", str(index), page_context))
    return items


def _project_validation(
    mapping_payload: Mapping[str, Any], rows_payload: Mapping[str, Any]
) -> tuple[list[dict[str, object]], int]:
    records: list[dict[str, object]] = []
    order = 1
    logical_count = 0
    for record_type, record_index, value in _validation_items(mapping_payload, rows_payload):
        record = _mapping(value, f"{record_type} validation record")
        statement = record.get("statement") or record.get("statement_type")
        status = record.get("status") or record.get("resolution") or record.get("method")
        affected = record.get("affected_row_ids")
        if affected is None:
            row_id = record.get("row_id") or record.get("source_row_id")
            affected = [] if row_id is None else [row_id]
        chunks, order = _chunk_record(
            order_start=order,
            record_type=record_type,
            record_index=record_index,
            statement=statement,
            status=status,
            affected_row_ids=affected,
            value=record,
        )
        records.extend(chunks)
        logical_count += 1
    return records, logical_count


def _metadata_chunks(values: Mapping[str, object]) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for key in sorted(values):
        rendered = _compact_json_unbounded(values[key])
        digest = _sha256(rendered.encode("utf-8"))
        parts = [
            rendered[index : index + _JSON_CHUNK_SIZE]
            for index in range(0, len(rendered), _JSON_CHUNK_SIZE)
        ] or [""]
        for part_number, part in enumerate(parts, start=1):
            records.append(
                _record(
                    RUN_METADATA_HEADERS,
                    (key, part_number, len(parts), digest, part),
                )
            )
    return records


def _content_ledger(
    mapping_payload: Mapping[str, Any], rows_payload: Mapping[str, Any]
) -> dict[str, object]:
    keys = (
        "source_dispositions",
        "new_item_proposals",
        "alias_proposals",
        "equations",
        "conflicts",
        "schema_dispositions",
    )
    ledger = {
        f"mapping.{key}": {
            "count": len(_records(mapping_payload.get(key), key)),
            "sha256": _record_hash(mapping_payload.get(key)),
        }
        for key in keys
    }
    ledger["native_rows.pages"] = {
        "count": len(_records(rows_payload.get("pages"), "native-row pages")),
        "sha256": _record_hash(rows_payload.get("pages")),
    }
    ledger["producer_schema.items"] = {
        "count": len(
            _records(
                _mapping(
                    _mapping(mapping_payload.get("producer_snapshots"), "producer snapshots").get(
                        "schema"
                    ),
                    "producer schema snapshot",
                ).get("items"),
                "producer schema items",
            )
        ),
        "sha256": _mapping(
            _mapping(mapping_payload.get("producer_snapshots"), "producer snapshots").get("schema"),
            "producer schema snapshot",
        ).get("items_sha256"),
    }
    return ledger


def _metadata_values(
    mapping_payload: Mapping[str, Any],
    rows_payload: Mapping[str, Any],
    *,
    mapping_identity: ArtifactIdentity,
    rows_identity: ArtifactIdentity,
    policy: NativeCanonicalExcelPolicy,
    implementation_ledger: Sequence[Mapping[str, object]],
    summary: Mapping[str, object],
    content_ledger: Mapping[str, object],
) -> dict[str, object]:
    values: dict[str, object] = {
        "export.artifact_type": "REGISTERED_NATIVE_CANONICAL_EXCEL_V1",
        "export.claim_boundary": policy.claim_boundary,
        "export.formulas_allowed": False,
        "export.force_mapping_allowed": False,
        "export.imputation_allowed": False,
        "export.current_mutable_schema_loaded": False,
        "export.producer_schema_snapshot_authoritative": True,
        "export.mapping_input": {
            "path": mapping_identity.path,
            "sha256": mapping_identity.sha256,
            "size_bytes": mapping_identity.size_bytes,
        },
        "export.native_rows_input": {
            "path": rows_identity.path,
            "sha256": rows_identity.sha256,
            "size_bytes": rows_identity.size_bytes,
        },
        "export.policy": {
            "path": policy.path.name,
            "name": policy.name,
            "sha256": policy.sha256,
        },
        "export.implementation": [dict(record) for record in implementation_ledger],
        "export.sheet_names": list(SHEET_NAMES),
        "export.summary": dict(summary),
        "export.content_ledger": dict(content_ledger),
    }
    for key in (
        "format_version",
        "policy",
        "claim_boundary",
        "status",
        "run_id",
        "source",
        "native_rows",
        "schema",
        "code",
        "authority",
        "isolation",
        "inputs",
        "summary",
    ):
        values[f"mapping.{key}"] = mapping_payload.get(key)
    for key in (
        "format_version",
        "policy",
        "claim_boundary",
        "status",
        "run_id",
        "source",
        "statement_discovery",
        "code",
        "authority",
        "isolation",
        "inputs",
        "selection",
        "summary",
    ):
        values[f"native_rows.{key}"] = rows_payload.get(key)
    return values


def _validate_mapping_summary(
    mapping_payload: Mapping[str, Any],
    *,
    disposition_counts: Counter[str],
    source_row_count: int,
    schema_item_count: int,
    proposal_count: int,
) -> None:
    summary = _mapping(mapping_payload.get("summary"), "mapping summary")
    if "new_schema_items_discovered" in summary:
        raise NativeCanonicalExcelExportError(
            "mapping summary uses deprecated new_schema_items_discovered"
        )
    if "new_schema_item_proposals" not in summary:
        raise NativeCanonicalExcelExportError(
            "mapping summary must report new_schema_item_proposals"
        )
    expected_counts = {status: disposition_counts[status] for status in _DISPOSITIONS}
    optional_expectations = {
        "visible_source_items": source_row_count,
        "source_items_successfully_accounted_for": source_row_count,
        "source_item_accounting_denominator": source_row_count,
        "mapped_to_existing_canonical_items": disposition_counts["EXISTING_ITEM"],
        "ambiguous": disposition_counts["AMBIGUOUS"],
        "unresolved": disposition_counts["UNRESOLVED"],
        "structural": disposition_counts["STRUCTURAL"],
        "universal_schema_item_count": schema_item_count,
        "new_schema_item_proposals": proposal_count,
    }
    for key, expected in optional_expectations.items():
        if key in summary and summary[key] != expected:
            raise NativeCanonicalExcelExportError(f"mapping summary {key} drifted")
    if (
        "source_disposition_counts" in summary
        and summary["source_disposition_counts"] != expected_counts
    ):
        raise NativeCanonicalExcelExportError("mapping source-disposition summary drifted")


def _style_sheet(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        width = max(
            len(str(sheet.cell(row, column).value or ""))
            for row in range(1, min(sheet.max_row, 200) + 1)
        )
        sheet.column_dimensions[letter].width = min(max(width + 2, 10), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_records(
    workbook: Workbook,
    sheet_name: str,
    headers: Sequence[str],
    records: Sequence[Mapping[str, object]],
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    append_literal_row(sheet, headers)
    for record in records:
        if tuple(record) != tuple(headers):
            raise NativeCanonicalExcelExportError(f"{sheet_name} record column order drifted")
        append_literal_row(sheet, [record[header] for header in headers])
    _style_sheet(sheet)


def _verify_sheet(
    workbook: Any,
    name: str,
    headers: Sequence[str],
    records: Sequence[Mapping[str, object]],
) -> None:
    sheet = workbook[name]
    observed_headers = tuple(sheet.cell(1, column).value for column in range(1, len(headers) + 1))
    if observed_headers != tuple(headers) or sheet.max_column != len(headers):
        raise NativeCanonicalExcelExportError(f"canonical workbook {name} headers drifted")
    if sheet.max_row != len(records) + 1:
        raise NativeCanonicalExcelExportError(f"canonical workbook {name} denominator drifted")
    for row_index, record in enumerate(records, start=2):
        for column, header in enumerate(headers, start=1):
            expected = record[header]
            roundtrip_expected = None if expected == "" else expected
            cell = sheet.cell(row_index, column)
            if cell.value != roundtrip_expected or (
                roundtrip_expected is not None and type(cell.value) is not type(roundtrip_expected)
            ):
                raise NativeCanonicalExcelExportError(
                    f"canonical workbook {name}!{cell.coordinate} roundtrip drifted"
                )
            if isinstance(roundtrip_expected, str) and cell.data_type != "s":
                raise NativeCanonicalExcelExportError(
                    f"canonical workbook {name}!{cell.coordinate} text is not literal"
                )


def _verify_chunk_sheet(
    workbook: Any,
    name: str,
    *,
    key_header: str,
    part_header: str,
    count_header: str,
    hash_header: str,
    payload_header: str,
) -> None:
    sheet = workbook[name]
    headers = {
        str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
    }
    groups: dict[tuple[object, ...], list[tuple[int, int, str, str]]] = {}
    for row_index in range(2, sheet.max_row + 1):
        if name == "VALIDATION":
            key = (
                sheet.cell(row_index, headers["RecordType"]).value,
                sheet.cell(row_index, headers["RecordIndex"]).value,
            )
        else:
            key = (sheet.cell(row_index, headers[key_header]).value,)
        groups.setdefault(key, []).append(
            (
                sheet.cell(row_index, headers[part_header]).value,
                sheet.cell(row_index, headers[count_header]).value,
                sheet.cell(row_index, headers[hash_header]).value,
                sheet.cell(row_index, headers[payload_header]).value or "",
            )
        )
    for chunks in groups.values():
        chunks.sort(key=lambda item: item[0])
        counts = {item[1] for item in chunks}
        digests = {item[2] for item in chunks}
        if (
            len(counts) != 1
            or len(digests) != 1
            or next(iter(counts)) != len(chunks)
            or [item[0] for item in chunks] != list(range(1, len(chunks) + 1))
        ):
            raise NativeCanonicalExcelExportError(f"canonical workbook {name} chunks drifted")
        rendered = "".join(item[3] for item in chunks)
        if _sha256(rendered.encode("utf-8")) != next(iter(digests)):
            raise NativeCanonicalExcelExportError(f"canonical workbook {name} chunk hash drifted")
        try:
            json.loads(rendered)
        except json.JSONDecodeError as exc:
            raise NativeCanonicalExcelExportError(
                f"canonical workbook {name} chunk JSON drifted"
            ) from exc


def _verify_workbook(
    payload: bytes,
    *,
    creator: str,
    expected_records: Mapping[str, Sequence[Mapping[str, object]]],
) -> None:
    try:
        workbook = load_workbook(BytesIO(payload), read_only=False, data_only=False)
    except Exception as exc:
        raise NativeCanonicalExcelExportError("canonical workbook cannot be reopened") from exc
    try:
        if tuple(workbook.sheetnames) != SHEET_NAMES or workbook_has_formula(workbook):
            raise NativeCanonicalExcelExportError(
                "canonical workbook sheet/formula contract drifted"
            )
        headers = {
            "SOURCE_ROWS": SOURCE_ROW_HEADERS,
            "CELLS": CELL_HEADERS,
            "SCHEMA_COVERAGE": SCHEMA_COVERAGE_HEADERS,
            "NEW_ITEM_PROPOSALS": NEW_ITEM_PROPOSAL_HEADERS,
            "VALIDATION": VALIDATION_HEADERS,
            "RUN_METADATA": RUN_METADATA_HEADERS,
        }
        for name, expected_headers in headers.items():
            _verify_sheet(workbook, name, expected_headers, expected_records[name])
        _verify_chunk_sheet(
            workbook,
            "VALIDATION",
            key_header="RecordType",
            part_header="PartNumber",
            count_header="PartCount",
            hash_header="RecordSha256",
            payload_header="RecordJsonPart",
        )
        _verify_chunk_sheet(
            workbook,
            "RUN_METADATA",
            key_header="Key",
            part_header="PartNumber",
            count_header="PartCount",
            hash_header="ValueSha256",
            payload_header="ValueJsonPart",
        )
        properties = workbook.properties
        if (
            properties.creator != creator
            or properties.lastModifiedBy != creator
            or properties.created != CANONICAL_CORE_TIMESTAMP
            or properties.modified != CANONICAL_CORE_TIMESTAMP
            or properties.version != "1"
            or properties.revision != "1"
        ):
            raise NativeCanonicalExcelExportError("canonical workbook core properties drifted")
    finally:
        workbook.close()


def _build_prevalidated_native_canonical_excel_artifacts(
    mapping_payload: Mapping[str, Any],
    rows_payload: Mapping[str, Any],
    *,
    mapping_identity: ArtifactIdentity,
    rows_identity: ArtifactIdentity,
    workbook_relative_path: str,
    provenance_relative_path: str,
    policy: NativeCanonicalExcelPolicy,
    implementation_ledger: Sequence[Mapping[str, object]],
) -> NativeCanonicalExcelArtifacts:
    """Build accepted bytes after callers have completed both strict input loads."""

    _validate_identity(mapping_identity, "mapping input")
    _validate_identity(rows_identity, "native-row input")
    implementation = _validate_implementation_ledger(implementation_ledger)
    workbook_relative_path = _canonical_relative_path(
        workbook_relative_path, "workbook relative path"
    )
    provenance_relative_path = _canonical_relative_path(
        provenance_relative_path, "provenance relative path"
    )
    workbook_location = PurePosixPath(workbook_relative_path)
    provenance_location = PurePosixPath(provenance_relative_path)
    workbook_filename = _plain_filename(workbook_location.name, "workbook filename")
    provenance_filename = _plain_filename(provenance_location.name, "provenance filename")
    if Path(workbook_filename).suffix.casefold() != ".xlsx":
        raise NativeCanonicalExcelExportError("workbook filename must end in .xlsx")
    if Path(provenance_filename).suffix.casefold() != ".json":
        raise NativeCanonicalExcelExportError("provenance filename must end in .json")
    if workbook_filename == provenance_filename:
        raise NativeCanonicalExcelExportError("workbook and provenance filenames must differ")
    if workbook_location.parent != provenance_location.parent:
        raise NativeCanonicalExcelExportError("workbook and provenance paths must be siblings")
    mapping_bytes = _canonical_json_bytes(mapping_payload)
    rows_bytes = _canonical_json_bytes(rows_payload)
    if (
        len(mapping_bytes) != mapping_identity.size_bytes
        or _sha256(mapping_bytes) != mapping_identity.sha256
    ):
        raise NativeCanonicalExcelExportError("mapping payload differs from its trusted identity")
    if len(rows_bytes) != rows_identity.size_bytes or _sha256(rows_bytes) != rows_identity.sha256:
        raise NativeCanonicalExcelExportError(
            "native-row payload differs from its trusted identity"
        )
    source, role = _validate_payload_identity(
        mapping_payload,
        rows_payload,
        mapping_identity=mapping_identity,
        rows_identity=rows_identity,
        policy=policy,
    )
    source_records, cell_records, source_rows = _project_source_and_cells(
        mapping_payload, rows_payload, rows_sha256=rows_identity.sha256
    )
    schema_records, schema_items = _project_schema_coverage(mapping_payload)
    source_row_ids = {str(row.row["row_id"]) for row in source_rows}
    proposal_records = _project_new_item_proposals(mapping_payload, source_row_ids)
    validation_records, validation_logical_count = _project_validation(
        mapping_payload, rows_payload
    )
    disposition_counts = Counter(record["Disposition"] for record in source_records)
    terminal_counts = Counter(record["TerminalOutcome"] for record in schema_records)
    _validate_mapping_summary(
        mapping_payload,
        disposition_counts=disposition_counts,
        source_row_count=len(source_records),
        schema_item_count=len(schema_items),
        proposal_count=len(proposal_records),
    )
    summary: dict[str, object] = {
        "source_row_count": len(source_records),
        "source_cell_count": len(cell_records),
        "source_disposition_counts": {
            status: disposition_counts[status] for status in _DISPOSITIONS
        },
        "producer_schema_item_count": len(schema_items),
        "schema_terminal_outcome_counts": {
            status: terminal_counts[status] for status in _TERMINAL_OUTCOMES
        },
        "new_item_proposal_count": len(proposal_records),
        "validation_logical_record_count": validation_logical_count,
        "validation_sheet_row_count": len(validation_records),
        "formula_count": 0,
        "imputed_cell_count": 0,
        "allocated_report_norm_id_count": 0,
    }
    content_ledger = _content_ledger(mapping_payload, rows_payload)
    metadata_records = _metadata_chunks(
        _metadata_values(
            mapping_payload,
            rows_payload,
            mapping_identity=mapping_identity,
            rows_identity=rows_identity,
            policy=policy,
            implementation_ledger=implementation,
            summary=summary,
            content_ledger=content_ledger,
        )
    )
    records_by_sheet: dict[str, Sequence[Mapping[str, object]]] = {
        "SOURCE_ROWS": source_records,
        "CELLS": cell_records,
        "SCHEMA_COVERAGE": schema_records,
        "NEW_ITEM_PROPOSALS": proposal_records,
        "VALIDATION": validation_records,
        "RUN_METADATA": metadata_records,
    }
    workbook = Workbook()
    workbook.remove(workbook.active)
    try:
        for name, headers in (
            ("SOURCE_ROWS", SOURCE_ROW_HEADERS),
            ("CELLS", CELL_HEADERS),
            ("SCHEMA_COVERAGE", SCHEMA_COVERAGE_HEADERS),
            ("NEW_ITEM_PROPOSALS", NEW_ITEM_PROPOSAL_HEADERS),
            ("VALIDATION", VALIDATION_HEADERS),
            ("RUN_METADATA", RUN_METADATA_HEADERS),
        ):
            _write_records(workbook, name, headers, records_by_sheet[name])
        if workbook_has_formula(workbook):
            raise NativeCanonicalExcelExportError("formulas are forbidden in canonical workbook")
        workbook_bytes = deterministic_workbook_bytes(workbook, creator=policy.workbook_creator)
    finally:
        workbook.close()
    _verify_workbook(
        workbook_bytes,
        creator=policy.workbook_creator,
        expected_records=records_by_sheet,
    )
    workbook_sha256 = _sha256(workbook_bytes)
    receipt = {
        "artifact_type": _RECEIPT_TYPE,
        "claim_boundary": policy.claim_boundary,
        "code": {
            "exporter_implementation": [dict(record) for record in implementation],
            "mapping_producer": mapping_payload.get("code"),
            "native_rows_producer": rows_payload.get("code"),
        },
        "content_ledger": content_ledger,
        "dataset_role": role,
        "format_version": 1,
        "inputs": {
            "mapping": _receipt_input_record(mapping_identity, mapping_payload),
            "native_rows": _receipt_input_record(rows_identity, rows_payload),
            "exact_pair_verified": True,
        },
        "isolation": _receipt_isolation(),
        "policy": _receipt_policy(policy),
        "run_id": mapping_payload.get("run_id"),
        "source": source,
        "status": _EXPORT_STATUS,
        "summary": summary,
        "provenance": {
            "filename": provenance_filename,
            "path": provenance_relative_path,
        },
        "workbook": {
            "filename": workbook_filename,
            "formula_count": 0,
            "path": workbook_relative_path,
            "sha256": workbook_sha256,
            "sheet_names": list(SHEET_NAMES),
            "size_bytes": len(workbook_bytes),
        },
    }
    provenance_bytes = _canonical_json_bytes(receipt)
    return NativeCanonicalExcelArtifacts(
        workbook_bytes=workbook_bytes,
        provenance_bytes=provenance_bytes,
        workbook_sha256=workbook_sha256,
        provenance_sha256=_sha256(provenance_bytes),
        summary=summary,
    )


def _implementation_ledger(project_root: Path) -> tuple[dict[str, object], ...]:
    records: list[dict[str, object]] = []
    for relative in _IMPLEMENTATION_PATHS:
        canonical = _canonical_relative_path(relative, "export implementation path")
        path = project_root.joinpath(*PurePosixPath(canonical).parts)
        _assert_no_symlink_components(project_root, path.parent)
        snapshot = _read_regular_snapshot(path, f"export implementation {relative}")
        records.append(
            {
                "path": relative,
                "sha256": _sha256(snapshot.payload),
                "size_bytes": len(snapshot.payload),
            }
        )
    return tuple(sorted(records, key=lambda record: str(record["path"])))


def _same_regular_inode(left: os.stat_result, right: os.stat_result) -> bool:
    return (
        stat.S_ISREG(left.st_mode)
        and stat.S_ISREG(right.st_mode)
        and (left.st_dev, left.st_ino) == (right.st_dev, right.st_ino)
    )


def _fsync_directory(directory: Path) -> None:
    descriptor = os.open(directory, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _rollback_if_same(path: Path, identity: os.stat_result) -> None:
    try:
        current = path.stat(follow_symlinks=False)
    except FileNotFoundError:
        return
    if not _same_regular_inode(identity, current):
        raise NativeCanonicalExcelExportError(f"refusing unsafe rollback: {path}")
    path.unlink()
    _fsync_directory(path.parent)


def _write_exclusive(path: Path, payload: bytes) -> os.stat_result:
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_CLOEXEC", 0)
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags, 0o644)
    except OSError as exc:
        raise NativeCanonicalExcelExportError(
            f"canonical Excel output already exists or cannot be created: {path}"
        ) from exc
    created: os.stat_result | None = None
    descriptor_open = True
    try:
        created = os.fstat(descriptor)
        view = memoryview(payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                raise NativeCanonicalExcelExportError(f"short output write: {path}")
            view = view[written:]
        os.fsync(descriptor)
        final_descriptor = os.fstat(descriptor)
        os.close(descriptor)
        descriptor_open = False
        linked = path.stat(follow_symlinks=False)
        if (
            not _same_regular_inode(created, final_descriptor)
            or not _same_regular_inode(final_descriptor, linked)
            or linked.st_size != len(payload)
            or path.read_bytes() != payload
        ):
            raise NativeCanonicalExcelExportError(f"published output identity drifted: {path}")
        return linked
    except BaseException:
        if descriptor_open:
            os.close(descriptor)
        if created is not None:
            _rollback_if_same(path, created)
        raise


def _publish_pair(
    workbook_path: Path,
    provenance_path: Path,
    workbook_bytes: bytes,
    provenance_bytes: bytes,
) -> tuple[os.stat_result, os.stat_result]:
    if workbook_path.parent != provenance_path.parent or workbook_path == provenance_path:
        raise NativeCanonicalExcelExportError("export outputs must be distinct siblings")
    if os.path.lexists(workbook_path) or os.path.lexists(provenance_path):
        raise NativeCanonicalExcelExportError("canonical Excel export refuses to overwrite")
    workbook_identity = _write_exclusive(workbook_path, workbook_bytes)
    provenance_identity: os.stat_result | None = None
    try:
        provenance_identity = _write_exclusive(provenance_path, provenance_bytes)
        _fsync_directory(workbook_path.parent)
        if (
            workbook_path.read_bytes() != workbook_bytes
            or provenance_path.read_bytes() != provenance_bytes
            or not _same_regular_inode(workbook_identity, workbook_path.stat(follow_symlinks=False))
            or not _same_regular_inode(
                provenance_identity, provenance_path.stat(follow_symlinks=False)
            )
        ):
            raise NativeCanonicalExcelExportError("published canonical Excel pair drifted")
        return workbook_identity, provenance_identity
    except BaseException as publication_error:
        rollback_errors: list[BaseException] = []
        if provenance_identity is not None:
            try:
                _rollback_if_same(provenance_path, provenance_identity)
            except BaseException as exc:
                rollback_errors.append(exc)
        try:
            _rollback_if_same(workbook_path, workbook_identity)
        except BaseException as exc:
            rollback_errors.append(exc)
        if rollback_errors:
            raise NativeCanonicalExcelExportError(
                "canonical Excel pair rollback was incomplete: "
                + "; ".join(str(error) for error in rollback_errors)
            ) from publication_error
        raise


def _rollback_owned_pair(
    workbook_path: Path,
    provenance_path: Path,
    identities: tuple[os.stat_result, os.stat_result],
    cause: BaseException,
) -> None:
    rollback_errors: list[BaseException] = []
    for path, identity in (
        (provenance_path, identities[1]),
        (workbook_path, identities[0]),
    ):
        try:
            _rollback_if_same(path, identity)
        except BaseException as exc:
            rollback_errors.append(exc)
    if rollback_errors:
        raise NativeCanonicalExcelExportError(
            "canonical Excel replay rollback was incomplete: "
            + "; ".join(str(error) for error in rollback_errors)
        ) from cause


def _assert_no_symlink_components(project_root: Path, directory: Path) -> None:
    relative = directory.relative_to(project_root)
    current = project_root
    for part in relative.parts:
        current = current / part
        if os.path.lexists(current):
            identity = current.stat(follow_symlinks=False)
            if stat.S_ISLNK(identity.st_mode):
                raise NativeCanonicalExcelExportError(
                    f"artifact path contains a symlink: {current}"
                )
            if not stat.S_ISDIR(identity.st_mode):
                raise NativeCanonicalExcelExportError(
                    f"artifact path component is not a directory: {current}"
                )


@dataclass(frozen=True, slots=True)
class _RegularFileSnapshot:
    path: Path
    payload: bytes
    identity: os.stat_result


def _strict_relative_artifact_path(
    project_root: Path,
    path: Path,
    label: str,
) -> tuple[str, Path]:
    candidate = Path(path)
    relative = _canonical_relative_path(candidate.as_posix(), label)
    absolute = project_root.joinpath(*PurePosixPath(relative).parts)
    _assert_no_symlink_components(project_root, absolute.parent)
    return relative, absolute


def _read_regular_snapshot(path: Path, label: str) -> _RegularFileSnapshot:
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags)
    except OSError as exc:
        raise NativeCanonicalExcelExportError(f"cannot open {label} as a regular file") from exc
    try:
        before = os.fstat(descriptor)
        if not stat.S_ISREG(before.st_mode):
            raise NativeCanonicalExcelExportError(f"{label} must be a regular file")
        chunks: list[bytes] = []
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            chunks.append(chunk)
        after = os.fstat(descriptor)
    finally:
        os.close(descriptor)
    try:
        linked = path.stat(follow_symlinks=False)
    except OSError as exc:
        raise NativeCanonicalExcelExportError(f"{label} disappeared during read") from exc
    payload = b"".join(chunks)
    stable_metadata = (
        before.st_size,
        before.st_mtime_ns,
        before.st_ctime_ns,
        before.st_mode,
    ) == (
        after.st_size,
        after.st_mtime_ns,
        after.st_ctime_ns,
        after.st_mode,
    )
    if (
        not stable_metadata
        or not _same_regular_inode(before, after)
        or not _same_regular_inode(after, linked)
        or after.st_size != len(payload)
    ):
        raise NativeCanonicalExcelExportError(f"{label} changed during read")
    return _RegularFileSnapshot(path=path, payload=payload, identity=linked)


def _same_snapshot_identity(left: os.stat_result, right: os.stat_result) -> bool:
    return _same_regular_inode(left, right) and (
        left.st_size,
        left.st_mtime_ns,
        left.st_ctime_ns,
        left.st_mode,
    ) == (
        right.st_size,
        right.st_mtime_ns,
        right.st_ctime_ns,
        right.st_mode,
    )


def _assert_snapshot_stable(snapshot: _RegularFileSnapshot, label: str) -> None:
    current = _read_regular_snapshot(snapshot.path, label)
    if current.payload != snapshot.payload or not _same_snapshot_identity(
        snapshot.identity, current.identity
    ):
        raise NativeCanonicalExcelExportError(f"{label} changed during strict replay")


def _assert_role_containment(
    relative_paths: Sequence[str],
    *,
    role: str,
    policy: NativeCanonicalExcelPolicy,
) -> None:
    role_root = PurePosixPath(policy.role_directories[role])
    for relative in relative_paths:
        try:
            PurePosixPath(relative).relative_to(role_root)
        except ValueError as exc:
            raise NativeCanonicalExcelExportError(
                f"paired {role} artifacts must stay under {policy.role_directories[role]}"
            ) from exc


def _receipt_input_identity(
    value: object,
    *,
    label: str,
    accepted_contract: Mapping[str, object],
) -> ArtifactIdentity:
    record = _exact_mapping(
        value,
        {
            "path",
            "sha256",
            "size_bytes",
            "claim_boundary",
            "format_version",
            "policy",
            "run_id",
            "status",
        },
        f"provenance {label} input",
    )
    identity = ArtifactIdentity(
        path=_canonical_relative_path(record.get("path"), f"provenance {label} path"),
        sha256=record.get("sha256"),
        size_bytes=record.get("size_bytes"),
    )
    _validate_identity(identity, f"provenance {label} input")
    for key, expected in accepted_contract.items():
        if record.get(key) != expected:
            raise NativeCanonicalExcelExportError(f"provenance {label} {key} is not canonical")
    if not isinstance(record.get("run_id"), str) or not record["run_id"]:
        raise NativeCanonicalExcelExportError(f"provenance {label} run_id is invalid")
    return identity


def _validate_receipt_envelope(
    receipt: object,
    *,
    policy: NativeCanonicalExcelPolicy,
    workbook_relative: str,
    provenance_relative: str,
    workbook_sha256: str,
    workbook_size_bytes: int,
) -> tuple[
    Mapping[str, Any],
    ArtifactIdentity,
    ArtifactIdentity,
    tuple[dict[str, object], ...],
]:
    record = _exact_mapping(
        receipt,
        {
            "artifact_type",
            "claim_boundary",
            "code",
            "content_ledger",
            "dataset_role",
            "format_version",
            "inputs",
            "isolation",
            "policy",
            "provenance",
            "run_id",
            "source",
            "status",
            "summary",
            "workbook",
        },
        "native canonical Excel provenance",
    )
    if (
        record.get("artifact_type") != _RECEIPT_TYPE
        or record.get("claim_boundary") != policy.claim_boundary
        or type(record.get("format_version")) is not int
        or record.get("format_version") != 1
        or record.get("status") != _EXPORT_STATUS
    ):
        raise NativeCanonicalExcelExportError("native canonical Excel provenance identity drifted")
    role = record.get("dataset_role")
    if not isinstance(role, str) or role not in policy.allowed_roles:
        raise NativeCanonicalExcelExportError("native canonical Excel provenance role is invalid")
    source = _mapping(record.get("source"), "native canonical Excel provenance source")
    if source.get("dataset_role") != role:
        raise NativeCanonicalExcelExportError("provenance source/role binding drifted")
    if not isinstance(record.get("run_id"), str) or not record["run_id"]:
        raise NativeCanonicalExcelExportError("native canonical Excel provenance run_id is invalid")
    _mapping(record.get("content_ledger"), "provenance content ledger")
    _mapping(record.get("summary"), "provenance summary")
    if record.get("isolation") != _receipt_isolation():
        raise NativeCanonicalExcelExportError("provenance isolation receipt drifted")
    if record.get("policy") != _receipt_policy(policy):
        raise NativeCanonicalExcelExportError("provenance policy receipt drifted")

    code = _exact_mapping(
        record.get("code"),
        {"exporter_implementation", "mapping_producer", "native_rows_producer"},
        "provenance code receipt",
    )
    if not isinstance(code.get("exporter_implementation"), list):
        raise NativeCanonicalExcelExportError("provenance implementation ledger is invalid")
    implementation = _validate_implementation_ledger(code["exporter_implementation"])
    _mapping(code.get("mapping_producer"), "provenance mapping producer")
    _mapping(code.get("native_rows_producer"), "provenance native-row producer")

    inputs = _exact_mapping(
        record.get("inputs"),
        {"exact_pair_verified", "mapping", "native_rows"},
        "provenance inputs",
    )
    if inputs.get("exact_pair_verified") is not True:
        raise NativeCanonicalExcelExportError("provenance exact-pair receipt is invalid")
    mapping_identity = _receipt_input_identity(
        inputs.get("mapping"),
        label="mapping",
        accepted_contract=policy.accepted_mapping,
    )
    rows_identity = _receipt_input_identity(
        inputs.get("native_rows"),
        label="native-row",
        accepted_contract=policy.accepted_rows,
    )
    if mapping_identity.path == rows_identity.path:
        raise NativeCanonicalExcelExportError("provenance input paths must be distinct")

    workbook = _exact_mapping(
        record.get("workbook"),
        {"filename", "formula_count", "path", "sha256", "sheet_names", "size_bytes"},
        "provenance workbook receipt",
    )
    provenance = _exact_mapping(
        record.get("provenance"),
        {"filename", "path"},
        "provenance self receipt",
    )
    workbook_filename = PurePosixPath(workbook_relative).name
    provenance_filename = PurePosixPath(provenance_relative).name
    if (
        workbook.get("filename") != workbook_filename
        or provenance.get("filename") != provenance_filename
        or workbook.get("path") != workbook_relative
        or provenance.get("path") != provenance_relative
        or workbook.get("sha256") != workbook_sha256
        or workbook.get("size_bytes") != workbook_size_bytes
        or type(workbook.get("formula_count")) is not int
        or workbook.get("formula_count") != 0
        or workbook.get("sheet_names") != list(SHEET_NAMES)
    ):
        raise NativeCanonicalExcelExportError("provenance output-pair receipt drifted")
    return record, mapping_identity, rows_identity, implementation


def load_registered_native_canonical_excel(
    *,
    project_root: Path,
    workbook_path: Path,
    workbook_expected_sha256: str,
    provenance_path: Path,
    provenance_expected_sha256: str,
) -> NativeCanonicalExcelExportResult:
    """Strict-replay a completed canonical workbook/provenance pair."""

    project_root = project_root.resolve()
    if _SHA256.fullmatch(workbook_expected_sha256) is None:
        raise NativeCanonicalExcelExportError("trusted workbook SHA-256 is invalid")
    if _SHA256.fullmatch(provenance_expected_sha256) is None:
        raise NativeCanonicalExcelExportError("trusted provenance SHA-256 is invalid")
    workbook_relative, workbook_absolute = _strict_relative_artifact_path(
        project_root, workbook_path, "workbook path"
    )
    provenance_relative, provenance_absolute = _strict_relative_artifact_path(
        project_root, provenance_path, "provenance path"
    )
    if (
        PurePosixPath(workbook_relative).parent != PurePosixPath(provenance_relative).parent
        or workbook_relative == provenance_relative
    ):
        raise NativeCanonicalExcelExportError("completed outputs must be distinct siblings")
    if (
        workbook_absolute.suffix.casefold() != ".xlsx"
        or provenance_absolute.suffix.casefold() != ".json"
    ):
        raise NativeCanonicalExcelExportError("completed output extensions must be .xlsx and .json")

    workbook_snapshot = _read_regular_snapshot(workbook_absolute, "canonical workbook")
    provenance_snapshot = _read_regular_snapshot(provenance_absolute, "canonical provenance")
    if _same_regular_inode(workbook_snapshot.identity, provenance_snapshot.identity):
        raise NativeCanonicalExcelExportError(
            "completed workbook and provenance must be distinct regular files"
        )
    if _sha256(workbook_snapshot.payload) != workbook_expected_sha256:
        raise NativeCanonicalExcelExportError("workbook does not match trusted SHA-256")
    if _sha256(provenance_snapshot.payload) != provenance_expected_sha256:
        raise NativeCanonicalExcelExportError("provenance does not match trusted SHA-256")
    try:
        receipt = json.loads(provenance_snapshot.payload)
    except json.JSONDecodeError as exc:
        raise NativeCanonicalExcelExportError("canonical provenance is not JSON") from exc
    if provenance_snapshot.payload != _canonical_json_bytes(receipt):
        raise NativeCanonicalExcelExportError("provenance is not canonical JSON")

    policy_relative, policy_absolute = _strict_relative_artifact_path(
        project_root, EXPORT_POLICY_RELATIVE_PATH, "export policy path"
    )
    if policy_relative != EXPORT_POLICY_RELATIVE_PATH.as_posix():
        raise NativeCanonicalExcelExportError("export policy path is not canonical")
    policy_snapshot = _read_regular_snapshot(policy_absolute, "canonical export policy")
    policy = load_native_canonical_excel_policy(policy_absolute, project_root)
    if _sha256(policy_snapshot.payload) != policy.sha256:
        raise NativeCanonicalExcelExportError("canonical export policy changed during load")
    (
        receipt_record,
        mapping_identity,
        rows_identity,
        receipt_implementation,
    ) = _validate_receipt_envelope(
        receipt,
        policy=policy,
        workbook_relative=workbook_relative,
        provenance_relative=provenance_relative,
        workbook_sha256=workbook_expected_sha256,
        workbook_size_bytes=len(workbook_snapshot.payload),
    )
    current_implementation = _implementation_ledger(project_root)
    if receipt_implementation != current_implementation:
        raise NativeCanonicalExcelExportError("provenance implementation ledger drifted")

    mapping_relative, mapping_absolute = _strict_relative_artifact_path(
        project_root, Path(mapping_identity.path), "receipt-bound mapping path"
    )
    rows_relative, rows_absolute = _strict_relative_artifact_path(
        project_root, Path(rows_identity.path), "receipt-bound native-row path"
    )
    role = str(receipt_record["dataset_role"])
    _assert_role_containment(
        (workbook_relative, provenance_relative, mapping_relative, rows_relative),
        role=role,
        policy=policy,
    )
    mapping_snapshot = _read_regular_snapshot(mapping_absolute, "receipt-bound mapping")
    rows_snapshot = _read_regular_snapshot(rows_absolute, "receipt-bound native rows")
    for snapshot, identity, label in (
        (mapping_snapshot, mapping_identity, "mapping"),
        (rows_snapshot, rows_identity, "native-row"),
    ):
        if (
            _sha256(snapshot.payload) != identity.sha256
            or len(snapshot.payload) != identity.size_bytes
        ):
            raise NativeCanonicalExcelExportError(
                f"receipt-bound {label} identity does not match its file"
            )

    try:
        mapping_payload = load_registered_native_canonical_mapping(
            mapping_absolute,
            project_root=project_root,
            expected_sha256=mapping_identity.sha256,
            policy_path=None,
            rows_policy_path=None,
        )
        rows_payload = load_registered_native_statement_rows(
            rows_absolute,
            project_root=project_root,
            expected_sha256=rows_identity.sha256,
            policy_path=None,
        )
    except Exception as exc:
        raise NativeCanonicalExcelExportError(
            "receipt-bound mapping/native rows failed strict replay"
        ) from exc
    source, loaded_role = _validate_payload_identity(
        mapping_payload,
        rows_payload,
        mapping_identity=mapping_identity,
        rows_identity=rows_identity,
        policy=policy,
    )
    inputs = _mapping(receipt_record["inputs"], "provenance inputs")
    code = _mapping(receipt_record["code"], "provenance code receipt")
    if (
        loaded_role != role
        or source != receipt_record.get("source")
        or mapping_payload.get("run_id") != receipt_record.get("run_id")
        or inputs.get("mapping") != _receipt_input_record(mapping_identity, mapping_payload)
        or inputs.get("native_rows") != _receipt_input_record(rows_identity, rows_payload)
        or code.get("mapping_producer") != mapping_payload.get("code")
        or code.get("native_rows_producer") != rows_payload.get("code")
    ):
        raise NativeCanonicalExcelExportError(
            "completed pair run/source/role/input receipt drifted"
        )

    rebuilt = _build_prevalidated_native_canonical_excel_artifacts(
        mapping_payload,
        rows_payload,
        mapping_identity=mapping_identity,
        rows_identity=rows_identity,
        workbook_relative_path=workbook_relative,
        provenance_relative_path=provenance_relative,
        policy=policy,
        implementation_ledger=current_implementation,
    )
    if (
        rebuilt.workbook_bytes != workbook_snapshot.payload
        or rebuilt.provenance_bytes != provenance_snapshot.payload
    ):
        raise NativeCanonicalExcelExportError(
            "completed canonical Excel pair does not match deterministic replay"
        )
    for snapshot, label in (
        (workbook_snapshot, "canonical workbook"),
        (provenance_snapshot, "canonical provenance"),
        (mapping_snapshot, "receipt-bound mapping"),
        (rows_snapshot, "receipt-bound native rows"),
        (policy_snapshot, "canonical export policy"),
    ):
        _assert_snapshot_stable(snapshot, label)
    if _implementation_ledger(project_root) != current_implementation:
        raise NativeCanonicalExcelExportError("export implementation changed during strict replay")
    return NativeCanonicalExcelExportResult(
        workbook_path=workbook_absolute,
        provenance_path=provenance_absolute,
        workbook_sha256=rebuilt.workbook_sha256,
        provenance_sha256=rebuilt.provenance_sha256,
        workbook_size_bytes=len(rebuilt.workbook_bytes),
        provenance_size_bytes=len(rebuilt.provenance_bytes),
        summary=rebuilt.summary,
    )


def export_registered_native_canonical_excel(
    *,
    project_root: Path,
    mapping_path: Path,
    mapping_expected_sha256: str,
    rows_path: Path,
    rows_expected_sha256: str,
    workbook_path: Path,
    provenance_path: Path,
    export_policy_path: Path = EXPORT_POLICY_RELATIVE_PATH,
    mapping_policy_path: Path | None = None,
    rows_policy_path: Path | None = None,
) -> NativeCanonicalExcelExportResult:
    """Strict-load a trusted mapping/native-row pair and exclusively publish its Excel pair."""

    project_root = project_root.resolve()
    mapping_path = _project_path(project_root, mapping_path)
    rows_path = _project_path(project_root, rows_path)
    workbook_path = _project_path(project_root, workbook_path)
    provenance_path = _project_path(project_root, provenance_path)
    export_policy_path = _project_path(project_root, export_policy_path)
    mapping_policy_path = (
        None if mapping_policy_path is None else _project_path(project_root, mapping_policy_path)
    )
    rows_policy_path = (
        None if rows_policy_path is None else _project_path(project_root, rows_policy_path)
    )
    if _SHA256.fullmatch(mapping_expected_sha256) is None:
        raise NativeCanonicalExcelExportError("trusted mapping SHA-256 is invalid")
    if _SHA256.fullmatch(rows_expected_sha256) is None:
        raise NativeCanonicalExcelExportError("trusted native-row SHA-256 is invalid")
    if workbook_path.suffix.casefold() != ".xlsx" or provenance_path.suffix.casefold() != ".json":
        raise NativeCanonicalExcelExportError("output extensions must be .xlsx and .json")
    if workbook_path.parent != provenance_path.parent or workbook_path == provenance_path:
        raise NativeCanonicalExcelExportError("export outputs must be distinct siblings")
    if os.path.lexists(workbook_path) or os.path.lexists(provenance_path):
        raise NativeCanonicalExcelExportError("canonical Excel export refuses to overwrite")
    policy = load_native_canonical_excel_policy(export_policy_path, project_root)
    try:
        initial_mapping = mapping_path.read_bytes()
        initial_rows = rows_path.read_bytes()
    except OSError as exc:
        raise NativeCanonicalExcelExportError("cannot read trusted export inputs") from exc
    if _sha256(initial_mapping) != mapping_expected_sha256:
        raise NativeCanonicalExcelExportError("mapping input does not match trusted SHA-256")
    if _sha256(initial_rows) != rows_expected_sha256:
        raise NativeCanonicalExcelExportError("native-row input does not match trusted SHA-256")
    try:
        mapping_payload = load_registered_native_canonical_mapping(
            mapping_path,
            project_root=project_root,
            expected_sha256=mapping_expected_sha256,
            policy_path=mapping_policy_path,
            rows_policy_path=rows_policy_path,
        )
        rows_payload = load_registered_native_statement_rows(
            rows_path,
            project_root=project_root,
            expected_sha256=rows_expected_sha256,
            policy_path=rows_policy_path,
        )
    except Exception as exc:
        raise NativeCanonicalExcelExportError("paired inputs failed strict load") from exc
    source = _mapping(rows_payload.get("source"), "native-row source")
    role = source.get("dataset_role")
    if not isinstance(role, str) or role not in policy.allowed_roles:
        raise NativeCanonicalExcelExportError("paired input role is not exportable")
    role_root = (project_root / policy.role_directories[role]).resolve()
    _relative(project_root, role_root, "dataset-role directory")
    for path in (mapping_path, rows_path, workbook_path, provenance_path):
        try:
            path.relative_to(role_root)
        except ValueError as exc:
            raise NativeCanonicalExcelExportError(
                f"paired {role} input/output must stay under {policy.role_directories[role]}"
            ) from exc
    _assert_no_symlink_components(project_root, workbook_path.parent)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    _assert_no_symlink_components(project_root, workbook_path.parent)
    implementation = _implementation_ledger(project_root)
    artifacts = _build_prevalidated_native_canonical_excel_artifacts(
        mapping_payload,
        rows_payload,
        mapping_identity=ArtifactIdentity(
            path=_relative(project_root, mapping_path, "mapping input"),
            sha256=mapping_expected_sha256,
            size_bytes=len(initial_mapping),
        ),
        rows_identity=ArtifactIdentity(
            path=_relative(project_root, rows_path, "native-row input"),
            sha256=rows_expected_sha256,
            size_bytes=len(initial_rows),
        ),
        workbook_relative_path=_relative(project_root, workbook_path, "workbook output"),
        provenance_relative_path=_relative(project_root, provenance_path, "provenance output"),
        policy=policy,
        implementation_ledger=implementation,
    )
    try:
        final_mapping = mapping_path.read_bytes()
        final_rows = rows_path.read_bytes()
        final_policy = export_policy_path.read_bytes()
    except OSError as exc:
        raise NativeCanonicalExcelExportError("export input disappeared during build") from exc
    if (
        final_mapping != initial_mapping
        or final_rows != initial_rows
        or _sha256(final_policy) != policy.sha256
        or _implementation_ledger(project_root) != implementation
    ):
        raise NativeCanonicalExcelExportError("export inputs changed during build")
    published_identities = _publish_pair(
        workbook_path,
        provenance_path,
        artifacts.workbook_bytes,
        artifacts.provenance_bytes,
    )
    try:
        return load_registered_native_canonical_excel(
            project_root=project_root,
            workbook_path=Path(_relative(project_root, workbook_path, "published workbook")),
            workbook_expected_sha256=artifacts.workbook_sha256,
            provenance_path=Path(_relative(project_root, provenance_path, "published provenance")),
            provenance_expected_sha256=artifacts.provenance_sha256,
        )
    except BaseException as replay_error:
        _rollback_owned_pair(
            workbook_path,
            provenance_path,
            published_identities,
            replay_error,
        )
        raise NativeCanonicalExcelExportError(
            "published canonical Excel pair failed strict replay and was rolled back"
        ) from replay_error


__all__ = [
    "CELL_HEADERS",
    "EXPORT_POLICY_RELATIVE_PATH",
    "NEW_ITEM_PROPOSAL_HEADERS",
    "RUN_METADATA_HEADERS",
    "SCHEMA_COVERAGE_HEADERS",
    "SHEET_NAMES",
    "SOURCE_ROW_HEADERS",
    "VALIDATION_HEADERS",
    "ArtifactIdentity",
    "NativeCanonicalExcelArtifacts",
    "NativeCanonicalExcelExportError",
    "NativeCanonicalExcelExportResult",
    "NativeCanonicalExcelPolicy",
    "export_registered_native_canonical_excel",
    "load_registered_native_canonical_excel",
    "load_native_canonical_excel_policy",
]
