"""Deterministic, review-only SHB maturity workbook pair.

This module is deliberately outside every canonical export path.  It accepts
one exact schema *candidate*, one exact page-local statement context, and the
exact replay-stable E-0042 numeric verification bytes.  The resulting XLSX is
a human-review convenience view: candidate ReportNormIds remain candidates,
the source-only total remains unmapped, and neither workbook nor provenance
receipt grants canonicalization or export authority.
"""

from __future__ import annotations

import copy
import hashlib
import json
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from io import BytesIO
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bctc_ai.export.canonical_xlsx import (
    append_literal_row,
    deterministic_workbook_bytes,
    set_literal_cell,
    workbook_has_formula,
)
from bctc_ai.source_structure.contracts_v1 import (
    canonical_clone_v1,
    canonical_json_sha256_v1,
)

FORMAT_VERSION = "SHB_MATURITY_REVIEW_WORKBOOK_V1"
PROVENANCE_FORMAT_VERSION = "SHB_MATURITY_REVIEW_WORKBOOK_PROVENANCE_V1"
ARTIFACT_ROLE = "REVIEW_ONLY_NON_CANONICAL_NON_EXPORT_AUTHORITY"
STATE = "DETERMINISTIC_REVIEW_ONLY_PAIR_BUILT"
CLAIM_BOUNDARY = (
    "EXACT_SHB_PAGE24_SCHEMA_CANDIDATE_STATEMENT_CONTEXT_AND_E0042_NUMERIC_"
    "VERIFICATION_REVIEW_VIEW_ONLY_NO_ACCEPTED_SCHEMA_MAPPING_CANONICALIZATION_"
    "VALUE_MATERIALIZATION_ACCOUNTING_TRUTH_PRODUCTION_OR_EXPORT_AUTHORITY"
)
SHEET_NAMES = ("REVIEW_ONLY", "CELL_PROVENANCE", "METADATA_CLAIMS")

E0042_RELATIVE_PATH = "docs/experiments/E-0042-shb-maturity-numeric-verification.json"
E0042_SHA256 = "929c1c81b0e08e14b5908087d866dc7bacc67c19cc62eb832353c5efb6c1801e"
E0042_SIZE_BYTES = 18_835
E0042_VERIFICATION_ID = (
    "sgnpvv1:verification:92b2d1d0ad293fb5ee2953128db9fb93c1c7f588eefff1bc00cfdeae16b61f1d"
)
SCHEMA_CANDIDATE_ID = (
    "slascv1:candidate:c297f71128bef07be383e684a9ff7ea33b11bf96e4fff154b2e05b39e72ef223"
)
STATEMENT_CONTEXT_ID = (
    "sscxtv1:context:a2d480f3bece8e0a29e0a935dbd4be00e4168159a6ec7d3d2946ab17d0b0ab8e"
)
SEMANTIC_GRAPH_ID = "slagv2:graph:47ec2635a8b57ee0773f26612d97dc7ce1a700993b169c25d7286f9b74be28d7"
SEMANTIC_GRAPH_SHA256 = "afbb553b45b3b776f36cea0696d04ead3031ac576cdb35450e11edd6e77854e6"
SOURCE_LOCAL_PAGE_ID = "ssv2:page:736b745df05b5c1f0ef81a5e985e38a44ef9b92612e9574c1b87ebb3e3b21ca1"
SOURCE_PROJECTION_SHA256 = "1036a24b4fbf8dde6f6b20341cee6d640f7c12cc22d83f67a798af2152e06ff7"
SEMANTIC_PAGE_BINDING_SHA256 = "e89153d4d78d337e438b90157ea330e8a84890f58847615ed34e304eff2a3a52"

_WORKBOOK_CREATOR = "bctc-ai/shb-maturity-review-v1"
_WARNING = "REVIEW ONLY — NOT CANONICAL — NOT EXPORT AUTHORITY"
_CANDIDATE_STATUS = "CANDIDATE_ONLY_NOT_MAPPED"
_UNIT_DISPLAY = "triệu đồng"
_UNIT_SOURCE_SURFACE = "Triệu đồng"
_NUMBER_FORMAT = "@"

_SOURCE_CONTEXT: dict[str, Any] = {
    "bank": "SHB",
    "physical_page": 24,
    "document_id": "sha256:3a66122194e4dd2e0ca18d584beeacb81279cf71e276eface59d17e72813dcfd",
    "source_pdf_ref": {
        "path": "vietstock_bctc/SHB/2026/BCTC Hợp nhất quý 2 năm 2026.pdf",
        "sha256": "3a66122194e4dd2e0ca18d584beeacb81279cf71e276eface59d17e72813dcfd",
        "size_bytes": 6_465_872,
    },
    "target_page_render_ref": {
        "path": (
            "output/development/bank-corpus-wave-1-role-b-page-reader-v1/full-v3/"
            "objects/sha256/43/43067bf4cb05b4ea8c7b526111bc170a3ef969f7aa79fd59a4f036201947772e.png"
        ),
        "sha256": "43067bf4cb05b4ea8c7b526111bc170a3ef969f7aa79fd59a4f036201947772e",
        "size_bytes": 932_776,
    },
    "display_context_basis": (
        "PINNED_EXACT_SHB_SEMANTIC_GRAPH_LINEAGE_FOR_REVIEW_DISPLAY_ONLY_"
        "NOT_REAUTHENTICATED_SOURCE_OR_EXPORT_AUTHORITY"
    ),
}

_PERIODS: tuple[dict[str, Any], ...] = (
    {
        "axis_id": "axis-0",
        "axis_ordinal": 0,
        "role": "CURRENT",
        "visible_date": "30/06/2026",
    },
    {
        "axis_id": "axis-1",
        "axis_ordinal": 1,
        "role": "COMPARATIVE",
        "visible_date": "31/12/2025",
    },
)


@dataclass(frozen=True, slots=True)
class _RowSpec:
    row_ordinal: int
    typed_role: str
    source_label: str
    candidate_report_norm_id: int | None
    candidate_canonical_name: str | None
    candidate_disposition: str
    candidate_graph_node_id: str


_ROWS: tuple[_RowSpec, ...] = (
    _RowSpec(
        0,
        "SHORT_TERM",
        "Nợ ngắn hạn",
        753,
        "+ Ngắn hạn",
        "VALUE_ROW_SCHEMA_CANDIDATE",
        "slagv2:node:e9c809acd1491657a7d3a3eeb97dad555adaf44f6c7c224f86a44bfc50773c89",
    ),
    _RowSpec(
        1,
        "MEDIUM_TERM",
        "Nợ trung hạn",
        754,
        "+ Trung hạn",
        "VALUE_ROW_SCHEMA_CANDIDATE",
        "slagv2:node:fd9612e657e44ecb9a8ea7001e01f8c29417fbef6d18d7b01c856d1821cb937f",
    ),
    _RowSpec(
        2,
        "LONG_TERM",
        "Nợ dài hạn",
        755,
        "+ Dài hạn",
        "VALUE_ROW_SCHEMA_CANDIDATE",
        "slagv2:node:0d999a5ce4e7243cb56ef16ba42b715c1e3eaaf9a4e95adae512311514a9be7f",
    ),
    _RowSpec(
        3,
        "TOTAL",
        "(unlabeled source total)",
        None,
        None,
        "SOURCE_ONLY_VALIDATION",
        "slagv2:node:e2c3b675b30fdf58b591cdfda0551bbac0498358f90e929f8f18bc1d904ee442",
    ),
)

_SAFETY = {
    "review_only": True,
    "schema_ids_are_candidates_only": True,
    "total_candidate_report_norm_id_is_null": True,
    "source_and_independent_values_preserved_as_text": True,
    "accepted_schema_mapping_authority": False,
    "canonicalization_authority": False,
    "value_materialization_authority": False,
    "accounting_truth_authority": False,
    "production_authority": False,
    "export_authority": False,
}


class ShbMaturityReviewWorkbookV1Error(ValueError):
    """Exact review inputs drifted or the review-only pair could not be built."""


@dataclass(frozen=True, slots=True)
class ShbMaturityReviewWorkbookArtifactsV1:
    workbook_bytes: bytes
    provenance_bytes: bytes
    workbook_sha256: str
    provenance_sha256: str
    projection_sha256: str
    sheet_names: tuple[str, ...]


def _error(message: str) -> ShbMaturityReviewWorkbookV1Error:
    return ShbMaturityReviewWorkbookV1Error(message)


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _json_text(value: object) -> str:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError, UnicodeError) as exc:
        raise _error("review workbook value is not finite JSON") from exc


def _pretty_json_bytes(value: object) -> bytes:
    try:
        rendered = json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            indent=2,
            sort_keys=True,
        )
    except (TypeError, ValueError, UnicodeError) as exc:
        raise _error("review provenance is not finite JSON") from exc
    return (rendered + "\n").encode("utf-8")


def _closed_json_object(payload: bytes, label: str) -> dict[str, Any]:
    if type(payload) is not bytes or not payload:
        raise _error(f"{label} must be non-empty exact bytes")

    def reject_constant(value: str) -> None:
        raise _error(f"{label} contains non-finite constant {value}")

    def reject_duplicate(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                raise _error(f"{label} contains duplicate key {key!r}")
            result[key] = value
        return result

    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=reject_duplicate,
            parse_constant=reject_constant,
        )
    except (UnicodeDecodeError, json.JSONDecodeError, RecursionError) as exc:
        raise _error(f"cannot decode {label} as strict JSON") from exc
    if type(value) is not dict:
        raise _error(f"{label} must be one JSON object")
    return cast(dict[str, Any], value)


def _exact_object(value: object, label: str) -> dict[str, Any]:
    if type(value) is not dict:
        raise _error(f"{label} must be one exact JSON object")
    try:
        clone = canonical_clone_v1(value)
    except ValueError as exc:
        raise _error(f"{label} is not an exact finite JSON tree") from exc
    return cast(dict[str, Any], clone)


def _validate_minted_id(
    value: Mapping[str, Any], *, id_key: str, expected_id: str, prefix: str, label: str
) -> None:
    clone = copy.deepcopy(dict(value))
    actual_id = clone.pop(id_key, None)
    if actual_id != expected_id:
        raise _error(f"{label} identity is not the exact SHB review input")
    rebuilt = f"{prefix}{canonical_json_sha256_v1(clone)}"
    if actual_id != rebuilt:
        raise _error(f"{label} minted identity does not match its payload")


def _validate_candidate(value: object) -> dict[str, Any]:
    candidate = _exact_object(value, "schema candidate")
    _validate_minted_id(
        candidate,
        id_key="candidate_set_id",
        expected_id=SCHEMA_CANDIDATE_ID,
        prefix="slascv1:candidate:",
        label="schema candidate",
    )
    if (
        candidate.get("status") != "CANDIDATE_SET_READY"
        or candidate.get("semantic_graph_id") != SEMANTIC_GRAPH_ID
        or candidate.get("semantic_graph_sha256") != SEMANTIC_GRAPH_SHA256
        or candidate.get("family_id") != "LOAN_MATURITY_BUCKETS"
        or candidate.get("source_semantics")
        != {"statement_type": None, "report_scope": None, "canonical_period_type": None}
    ):
        raise _error("schema candidate SHB graph or non-authority boundary drifted")
    readiness = candidate.get("readiness")
    safety = candidate.get("safety")
    if (
        type(readiness) is not dict
        or readiness.get("schema_mapping_ready") is not False
        or readiness.get("canonicalization_eligible") is not False
        or readiness.get("export_eligible") is not False
        or type(safety) is not dict
        or safety.get("schema_mapping_authority") is not False
        or safety.get("canonicalization_authority") is not False
        or safety.get("export_authority") is not False
    ):
        raise _error("schema candidate unexpectedly grants mapping or export authority")
    roles = candidate.get("role_candidates")
    if type(roles) is not list:
        raise _error("schema candidate role inventory is absent")
    by_role = {item.get("typed_role"): item for item in roles if type(item) is dict}
    if len(by_role) != len(roles):
        raise _error("schema candidate repeats or malforms a typed role")
    for spec in _ROWS:
        role = by_role.get(spec.typed_role)
        expected_ids = (
            [] if spec.candidate_report_norm_id is None else [spec.candidate_report_norm_id]
        )
        if (
            type(role) is not dict
            or role.get("graph_node_kind") != "LOGICAL_ROW"
            or role.get("graph_node_id") != spec.candidate_graph_node_id
            or role.get("candidate_report_norm_ids") != expected_ids
            or role.get("canonical_name") != spec.candidate_canonical_name
            or role.get("disposition") != spec.candidate_disposition
        ):
            raise _error(f"schema candidate role {spec.typed_role} drifted")
    if set(by_role) != {
        "OWNER_LABEL",
        "BRANCH_LABEL",
        "SHORT_TERM",
        "MEDIUM_TERM",
        "LONG_TERM",
        "TOTAL",
    }:
        raise _error("schema candidate is not the exact six-role maturity slice")
    return candidate


def _validate_context(value: object) -> dict[str, Any]:
    context = _exact_object(value, "statement context")
    _validate_minted_id(
        context,
        id_key="context_id",
        expected_id=STATEMENT_CONTEXT_ID,
        prefix="sscxtv1:context:",
        label="statement context",
    )
    if (
        context.get("status") != "RESOLVED_VISIBLE_PAGE_STATEMENT_CONTEXT"
        or context.get("source_local_page_id") != SOURCE_LOCAL_PAGE_ID
        or context.get("source_projection_sha256") != SOURCE_PROJECTION_SHA256
        or context.get("semantic_page_binding_sha256") != SEMANTIC_PAGE_BINDING_SHA256
        or context.get("statement_type") != "TM"
        or context.get("report_scope") != "CONSOLIDATED"
        or context.get("continuation") is not True
    ):
        raise _error("statement context is not exact SHB page-24 TM consolidated continuation")
    evidence = context.get("heading_evidence")
    readiness = context.get("readiness")
    safety = context.get("safety")
    if (
        type(evidence) is not dict
        or evidence.get("raw_transformer_text_utf8")
        != "THUYẾT MINH BÁO CÁO TÀI CHÍNH HỢP NHẤT (TIẾP THEO)"
        or type(evidence.get("crop_ref")) is not dict
        or evidence["crop_ref"].get("sha256")
        != "c92d9a8a0093bf06c3fb8f41d96a216f106641a171e6858b5ffbc5126e145b2c"
        or type(readiness) is not dict
        or readiness.get("schema_mapping_ready") is not False
        or readiness.get("export_eligible") is not False
        or type(safety) is not dict
        or safety.get("schema_mapping_authority") is not False
        or safety.get("canonicalization_authority") is not False
        or safety.get("export_authority") is not False
    ):
        raise _error("statement heading evidence or non-authority boundary drifted")
    return context


def _validate_verification(payload: bytes) -> dict[str, Any]:
    if len(payload) != E0042_SIZE_BYTES or _sha256(payload) != E0042_SHA256:
        raise _error("E-0042 verification exact byte identity drifted")
    verification = _closed_json_object(payload, "E-0042 verification")
    inputs = verification.get("inputs")
    metrics = verification.get("metrics")
    safety = verification.get("safety")
    if (
        verification.get("verification_id") != E0042_VERIFICATION_ID
        or verification.get("status") != "COMPLETE_WITH_EXACT_EIGHT_CELL_AGREEMENT"
        or verification.get("authority")
        != "BOUNDED_FROZEN_HISTORICAL_NUMERIC_CANDIDATE_VERIFICATION_ONLY"
        or type(inputs) is not dict
        or inputs.get("semantic_graph")
        != {"graph_id": SEMANTIC_GRAPH_ID, "sha256": SEMANTIC_GRAPH_SHA256}
        or inputs.get("source_projection_sha256") != SOURCE_PROJECTION_SHA256
        or inputs.get("semantic_page_binding_sha256") != SEMANTIC_PAGE_BINDING_SHA256
        or type(metrics) is not dict
        or metrics.get("cell_count") != 8
        or metrics.get("exact_eight_cell_agreement") is not True
        or metrics.get("reader_score_decision_use_count") != 0
        or type(safety) is not dict
        or safety.get("report_norm_id_or_schema_authority") is not False
        or safety.get("period_or_unit_or_scope_authority") is not False
        or safety.get("statement_or_export_authority") is not False
    ):
        raise _error("E-0042 verification status, lineage, or safety boundary drifted")
    cells = verification.get("cells")
    if type(cells) is not list or len(cells) != 8:
        raise _error("E-0042 verification does not contain exactly eight cells")
    expected_coordinates = {
        (row.row_ordinal, period["axis_ordinal"]) for row in _ROWS for period in _PERIODS
    }
    observed_coordinates: set[tuple[int, int]] = set()
    observed_ids: set[str] = set()
    for cell in cells:
        if type(cell) is not dict:
            raise _error("E-0042 verification cell is not an object")
        coordinate = (cell.get("row_ordinal"), cell.get("axis_ordinal"))
        cell_id = cell.get("cell_id")
        primary = cell.get("primary")
        challenger = cell.get("challenger")
        if (
            coordinate not in expected_coordinates
            or type(cell_id) is not str
            or cell_id in observed_ids
            or coordinate in observed_coordinates
            or cell.get("page") != 1
            or type(primary) is not dict
            or type(challenger) is not dict
            or primary.get("observation") != "VALUE"
            or challenger.get("proposal_status") != "NUMERIC_CHARACTERS_ONLY_PROPOSAL"
            or primary.get("value") != challenger.get("parsed_value")
            or primary.get("raw_text") != challenger.get("raw_text")
            or cell.get("normalized_numeric_value") != primary.get("value")
            or cell.get("verification_status") != "VERIFIED_OBSERVED_VALUE"
            or cell.get("decision") != "ACCEPT_EXACT_VALUE_AND_SIGN_AGREEMENT"
            or cell.get("final_value_status") != "OBSERVED_VALUE"
        ):
            raise _error("E-0042 cell topology or exact-agreement evidence drifted")
        observed_coordinates.add(cast(tuple[int, int], coordinate))
        observed_ids.add(cell_id)
    if observed_coordinates != expected_coordinates:
        raise _error("E-0042 cell coordinate denominator drifted")
    return verification


def _cell_by_coordinate(verification: Mapping[str, Any]) -> dict[tuple[int, int], dict[str, Any]]:
    return {
        (cell["row_ordinal"], cell["axis_ordinal"]): cell
        for cell in cast(Sequence[dict[str, Any]], verification["cells"])
    }


def _build_projection(
    candidate: Mapping[str, Any],
    context: Mapping[str, Any],
    verification: Mapping[str, Any],
) -> dict[str, Any]:
    cells = _cell_by_coordinate(verification)
    review_rows: list[dict[str, Any]] = []
    provenance_cells: list[dict[str, Any]] = []
    for spec in _ROWS:
        periods: list[dict[str, Any]] = []
        for period in _PERIODS:
            cell = cells[(spec.row_ordinal, period["axis_ordinal"])]
            primary = cast(Mapping[str, Any], cell["primary"])
            challenger = cast(Mapping[str, Any], cell["challenger"])
            record = {
                "cell_id": cell["cell_id"],
                "physical_page": _SOURCE_CONTEXT["physical_page"],
                "crop_local_page": cell["page"],
                "row_ordinal": spec.row_ordinal,
                "typed_role": spec.typed_role,
                "source_label": spec.source_label,
                "candidate_report_norm_id": spec.candidate_report_norm_id,
                "candidate_canonical_name": spec.candidate_canonical_name,
                "candidate_disposition": spec.candidate_disposition,
                "candidate_status": _CANDIDATE_STATUS,
                "candidate_row_graph_node_id": spec.candidate_graph_node_id,
                "axis_id": period["axis_id"],
                "axis_ordinal": period["axis_ordinal"],
                "period_role": period["role"],
                "period_visible_date": period["visible_date"],
                "unit_display": _UNIT_DISPLAY,
                "unit_source_surface": _UNIT_SOURCE_SURFACE,
                "source_value": {
                    "raw_text": primary["raw_text"],
                    "normalized_text": primary["normalized_text"],
                    "value": primary["value"],
                    "observation": primary["observation"],
                    "sign_evidence": primary["sign_evidence"],
                },
                "independent_value": {
                    "raw_text": challenger["raw_text"],
                    "parsed_value": challenger["parsed_value"],
                    "parsed_observation": challenger["parsed_observation"],
                    "proposal_status": challenger["proposal_status"],
                    "reader_score_diagnostic_only": challenger["reader_score"],
                    "sign_evidence": challenger["sign_evidence"],
                },
                "verification": {
                    "status": cell["verification_status"],
                    "decision": cell["decision"],
                    "final_value_status": cell["final_value_status"],
                    "normalized_numeric_value": cell["normalized_numeric_value"],
                },
                "source_provenance": {
                    "source_line_index": cell["source_line_index"],
                    "crop_path": cell["crop_path"],
                    "crop_sha256": cell["crop_sha256"],
                    "crop_size_bytes": cell["crop_size_bytes"],
                    "source_atom_id": cell["source_atom_id"],
                    "source_evidence_node_id": cell["source_evidence_node_id"],
                    "source_value_graph_node_id": cell["source_graph_node_id"],
                },
            }
            periods.append(record)
            provenance_cells.append(copy.deepcopy(record))
        review_rows.append(
            {
                "row_ordinal": spec.row_ordinal,
                "typed_role": spec.typed_role,
                "source_label": spec.source_label,
                "candidate_report_norm_id": spec.candidate_report_norm_id,
                "candidate_canonical_name": spec.candidate_canonical_name,
                "candidate_disposition": spec.candidate_disposition,
                "candidate_status": _CANDIDATE_STATUS,
                "candidate_row_graph_node_id": spec.candidate_graph_node_id,
                "periods": periods,
                "reviewer_decision": "UNREVIEWED",
                "reviewer_notes": None,
            }
        )
    payload = {
        "format_version": FORMAT_VERSION,
        "artifact_role": ARTIFACT_ROLE,
        "state": STATE,
        "claim_boundary": CLAIM_BOUNDARY,
        "source_context": copy.deepcopy(_SOURCE_CONTEXT),
        "statement_context": {
            "statement_type": context["statement_type"],
            "report_scope": context["report_scope"],
            "continuation": context["continuation"],
            "heading_surface": context["heading_evidence"]["raw_transformer_text_utf8"],
        },
        "review_table_context": {
            "family_id": "LOAN_MATURITY_BUCKETS",
            "periods": copy.deepcopy(list(_PERIODS)),
            "unit_display": _UNIT_DISPLAY,
            "unit_source_surface": _UNIT_SOURCE_SURFACE,
        },
        "input_identities": {
            "schema_candidate": {
                "candidate_set_id": candidate["candidate_set_id"],
                "canonical_json_sha256": canonical_json_sha256_v1(candidate),
            },
            "statement_context": {
                "context_id": context["context_id"],
                "canonical_json_sha256": canonical_json_sha256_v1(context),
            },
            "e0042_numeric_verification": {
                "path": E0042_RELATIVE_PATH,
                "sha256": E0042_SHA256,
                "size_bytes": E0042_SIZE_BYTES,
                "verification_id": verification["verification_id"],
            },
        },
        "shared_lineage": {
            "semantic_graph_id": SEMANTIC_GRAPH_ID,
            "semantic_graph_sha256": SEMANTIC_GRAPH_SHA256,
            "source_local_page_id": SOURCE_LOCAL_PAGE_ID,
            "source_projection_sha256": SOURCE_PROJECTION_SHA256,
            "semantic_page_binding_sha256": SEMANTIC_PAGE_BINDING_SHA256,
        },
        "review_rows": review_rows,
        "provenance_cells": provenance_cells,
        "metrics": {
            "review_row_count": 4,
            "provenance_cell_count": 8,
            "candidate_report_norm_id_count": 3,
            "source_only_null_candidate_row_count": 1,
            "verified_observed_cell_count": 8,
        },
        "safety": copy.deepcopy(_SAFETY),
    }
    payload["projection_id"] = f"shbmrwv1:projection:{canonical_json_sha256_v1(payload)}"
    return cast(dict[str, Any], canonical_clone_v1(payload))


def _style_header(sheet: Any, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _set_widths(sheet: Any, widths: Mapping[int, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_warning(sheet: Any, *, last_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    set_literal_cell(sheet.cell(1, 1), _WARNING)
    cell = sheet.cell(1, 1)
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.fill = PatternFill("solid", fgColor="C00000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24


def _write_review_sheet(workbook: Workbook, projection: Mapping[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "REVIEW_ONLY"
    _write_warning(sheet, last_column=18)
    metadata_rows = (
        ("Bank", "SHB", "Physical page", 24, "Statement", "TM", "Scope", "CONSOLIDATED"),
        (
            "Continuation",
            True,
            "Current period",
            "30/06/2026",
            "Comparative period",
            "31/12/2025",
            "Unit",
            _UNIT_DISPLAY,
        ),
        (
            "Candidate warning",
            "ReportNormIds below are candidates only; no row is an accepted mapping.",
            "Total warning",
            "TOTAL intentionally has a null candidate ReportNormId.",
            None,
            None,
            None,
            None,
        ),
    )
    for values in metadata_rows:
        append_literal_row(sheet, values)
    append_literal_row(sheet, (None,))
    headers = (
        "Row",
        "TypedRole",
        "SourceLabel",
        "CandidateReportNormId",
        "CandidateCanonicalName",
        "CandidateDisposition",
        "MappingStatus",
        "30/06/2026 Source",
        "30/06/2026 Independent",
        "30/06/2026 Verification",
        "31/12/2025 Source",
        "31/12/2025 Independent",
        "31/12/2025 Verification",
        "Unit",
        "PhysicalPage",
        "ReviewerDecision",
        "ReviewerNotes",
        "Authority",
    )
    append_literal_row(sheet, headers)
    header_row = sheet.max_row
    for row in cast(Sequence[Mapping[str, Any]], projection["review_rows"]):
        current, comparative = cast(Sequence[Mapping[str, Any]], row["periods"])
        append_literal_row(
            sheet,
            (
                row["row_ordinal"],
                row["typed_role"],
                row["source_label"],
                row["candidate_report_norm_id"],
                row["candidate_canonical_name"],
                row["candidate_disposition"],
                row["candidate_status"],
                current["source_value"]["raw_text"],
                current["independent_value"]["raw_text"],
                current["verification"]["status"],
                comparative["source_value"]["raw_text"],
                comparative["independent_value"]["raw_text"],
                comparative["verification"]["status"],
                row["periods"][0]["unit_display"],
                _SOURCE_CONTEXT["physical_page"],
                row["reviewer_decision"],
                row["reviewer_notes"],
                ARTIFACT_ROLE,
            ),
        )
    _style_header(sheet, header_row)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:R{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "C00000"
    _set_widths(
        sheet,
        {
            1: 8,
            2: 18,
            3: 28,
            4: 24,
            5: 24,
            6: 30,
            7: 28,
            8: 20,
            9: 20,
            10: 28,
            11: 20,
            12: 20,
            13: 28,
            14: 16,
            15: 14,
            16: 20,
            17: 30,
            18: 38,
        },
    )
    for row_number in range(header_row + 1, sheet.max_row + 1):
        for column in (8, 9, 11, 12):
            sheet.cell(row_number, column).number_format = _NUMBER_FORMAT
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet.cell(row_number, 2).value == "TOTAL":
            for cell in sheet[row_number]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFF2CC")


def _write_provenance_sheet(workbook: Workbook, projection: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("CELL_PROVENANCE")
    _write_warning(sheet, last_column=31)
    headers = (
        "CellId",
        "PhysicalPage",
        "CropLocalPage",
        "RowOrdinal",
        "TypedRole",
        "SourceLabel",
        "CandidateReportNormId",
        "CandidateCanonicalName",
        "CandidateDisposition",
        "CandidateStatus",
        "AxisId",
        "AxisOrdinal",
        "PeriodRole",
        "PeriodVisibleDate",
        "Unit",
        "SourceRawText",
        "SourceNormalizedText",
        "SourceObservation",
        "IndependentRawText",
        "IndependentParsedValue",
        "IndependentProposalStatus",
        "IndependentReaderScoreDiagnosticOnly",
        "VerificationStatus",
        "VerificationDecision",
        "FinalValueStatus",
        "SourceLineIndex",
        "CropPath",
        "CropSha256",
        "CropSizeBytes",
        "SourceAtomId",
        "GraphProvenanceJson",
    )
    append_literal_row(sheet, headers)
    header_row = sheet.max_row
    for cell in cast(Sequence[Mapping[str, Any]], projection["provenance_cells"]):
        source = cast(Mapping[str, Any], cell["source_value"])
        independent = cast(Mapping[str, Any], cell["independent_value"])
        verification = cast(Mapping[str, Any], cell["verification"])
        provenance = cast(Mapping[str, Any], cell["source_provenance"])
        append_literal_row(
            sheet,
            (
                cell["cell_id"],
                cell["physical_page"],
                cell["crop_local_page"],
                cell["row_ordinal"],
                cell["typed_role"],
                cell["source_label"],
                cell["candidate_report_norm_id"],
                cell["candidate_canonical_name"],
                cell["candidate_disposition"],
                cell["candidate_status"],
                cell["axis_id"],
                cell["axis_ordinal"],
                cell["period_role"],
                cell["period_visible_date"],
                cell["unit_display"],
                source["raw_text"],
                source["normalized_text"],
                source["observation"],
                independent["raw_text"],
                independent["parsed_value"],
                independent["proposal_status"],
                independent["reader_score_diagnostic_only"],
                verification["status"],
                verification["decision"],
                verification["final_value_status"],
                provenance["source_line_index"],
                provenance["crop_path"],
                provenance["crop_sha256"],
                provenance["crop_size_bytes"],
                provenance["source_atom_id"],
                _json_text(
                    {
                        "candidate_row_graph_node_id": cell["candidate_row_graph_node_id"],
                        "source_evidence_node_id": provenance["source_evidence_node_id"],
                        "source_value_graph_node_id": provenance["source_value_graph_node_id"],
                    }
                ),
            ),
        )
    _style_header(sheet, header_row)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:AE{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "5B9BD5"
    _set_widths(
        sheet,
        {
            1: 34,
            5: 18,
            6: 28,
            7: 24,
            8: 24,
            9: 30,
            10: 28,
            14: 18,
            15: 16,
            16: 20,
            17: 22,
            19: 22,
            20: 22,
            21: 38,
            22: 24,
            23: 28,
            24: 40,
            25: 22,
            27: 46,
            28: 68,
            30: 76,
            31: 100,
        },
    )
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in (16, 17, 19, 20):
            sheet.cell(row[0].row, column).number_format = _NUMBER_FORMAT


def _write_metadata_sheet(workbook: Workbook, projection: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("METADATA_CLAIMS")
    _write_warning(sheet, last_column=3)
    append_literal_row(sheet, ("Key", "Value", "Meaning"))
    header_row = sheet.max_row
    input_identities = cast(Mapping[str, Any], projection["input_identities"])
    lineage = cast(Mapping[str, Any], projection["shared_lineage"])
    metadata = (
        ("format_version", FORMAT_VERSION, "Review pair contract"),
        ("artifact_role", ARTIFACT_ROLE, "Never a canonical/export artifact"),
        ("state", STATE, "Deterministic bytes built after exact input gates"),
        ("claim_boundary", CLAIM_BOUNDARY, "Maximum permissible claim"),
        ("projection_id", projection["projection_id"], "Content-addressed review projection"),
        ("bank", "SHB", "Pinned review display context"),
        ("physical_page", 24, "PDF physical page, not crop-local page 1"),
        ("statement_type", "TM", "Exact page-local statement context"),
        ("report_scope", "CONSOLIDATED", "Exact page-local statement context"),
        ("continuation", True, "Visible heading includes (TIẾP THEO)"),
        ("periods", _json_text(["30/06/2026", "31/12/2025"]), "Visible period surfaces"),
        ("unit", _UNIT_DISPLAY, "Visible source unit; no canonicalization authority"),
        ("schema_candidate", _json_text(input_identities["schema_candidate"]), "Exact input"),
        ("statement_context", _json_text(input_identities["statement_context"]), "Exact input"),
        (
            "e0042_numeric_verification",
            _json_text(input_identities["e0042_numeric_verification"]),
            "Exact committed bytes",
        ),
        ("shared_lineage", _json_text(lineage), "Cross-input identity gate"),
        ("candidate_ids", _json_text([753, 754, 755, None]), "Candidates; TOTAL is null"),
        ("accepted_schema_mapping_authority", False, "Explicit denial"),
        ("canonicalization_authority", False, "Explicit denial"),
        ("value_materialization_authority", False, "Explicit denial"),
        ("accounting_truth_authority", False, "Explicit denial"),
        ("production_authority", False, "Explicit denial"),
        ("export_authority", False, "Explicit denial"),
    )
    for item in metadata:
        append_literal_row(sheet, item)
    _style_header(sheet, header_row)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:C{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "A5A5A5"
    _set_widths(sheet, {1: 38, 2: 110, 3: 58})
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_workbook_bytes(projection: Mapping[str, Any]) -> bytes:
    workbook = Workbook()
    try:
        _write_review_sheet(workbook, projection)
        _write_provenance_sheet(workbook, projection)
        _write_metadata_sheet(workbook, projection)
        if tuple(workbook.sheetnames) != SHEET_NAMES or workbook_has_formula(workbook):
            raise _error("review workbook sheet contract or literal-cell safety drifted")
        payload = deterministic_workbook_bytes(workbook, creator=_WORKBOOK_CREATOR)
    finally:
        workbook.close()
    try:
        reopened = load_workbook(BytesIO(payload), data_only=False, read_only=False)
    except Exception as exc:
        raise _error("cannot reopen deterministic review workbook") from exc
    try:
        if tuple(reopened.sheetnames) != SHEET_NAMES or workbook_has_formula(reopened):
            raise _error("serialized review workbook contains formula or sheet drift")
        total_row = next(
            row for row in reopened["REVIEW_ONLY"].iter_rows() if row[1].value == "TOTAL"
        )
        if total_row[3].value is not None:
            raise _error("serialized TOTAL candidate ReportNormId must remain null")
    finally:
        reopened.close()
    return payload


def build_shb_maturity_review_workbook_v1(
    schema_candidate: object,
    statement_context: object,
    e0042_verification_bytes: bytes,
) -> ShbMaturityReviewWorkbookArtifactsV1:
    """Build deterministic review-only XLSX and companion provenance JSON bytes.

    Inputs must be the exact frozen SHB candidate/context objects and exact
    corrected E-0042 bytes.  No file is written and no canonical export object
    is produced.
    """

    candidate = _validate_candidate(schema_candidate)
    context = _validate_context(statement_context)
    verification = _validate_verification(e0042_verification_bytes)
    projection = _build_projection(candidate, context, verification)
    projection_sha256 = canonical_json_sha256_v1(projection)
    workbook_bytes = _build_workbook_bytes(projection)
    workbook_sha256 = _sha256(workbook_bytes)
    provenance = {
        "format_version": PROVENANCE_FORMAT_VERSION,
        "artifact_role": ARTIFACT_ROLE,
        "state": STATE,
        "claim_boundary": CLAIM_BOUNDARY,
        "projection_sha256": projection_sha256,
        "projection": projection,
        "workbook": {
            "sha256": workbook_sha256,
            "size_bytes": len(workbook_bytes),
            "sheet_names": list(SHEET_NAMES),
            "formula_count": 0,
            "creator": _WORKBOOK_CREATOR,
        },
        "safety": copy.deepcopy(_SAFETY),
    }
    provenance["provenance_id"] = f"shbmrwpv1:provenance:{canonical_json_sha256_v1(provenance)}"
    provenance_bytes = _pretty_json_bytes(provenance)
    return ShbMaturityReviewWorkbookArtifactsV1(
        workbook_bytes=workbook_bytes,
        provenance_bytes=provenance_bytes,
        workbook_sha256=workbook_sha256,
        provenance_sha256=_sha256(provenance_bytes),
        projection_sha256=projection_sha256,
        sheet_names=SHEET_NAMES,
    )


__all__ = [
    "ARTIFACT_ROLE",
    "CLAIM_BOUNDARY",
    "E0042_RELATIVE_PATH",
    "E0042_SHA256",
    "E0042_SIZE_BYTES",
    "E0042_VERIFICATION_ID",
    "FORMAT_VERSION",
    "PROVENANCE_FORMAT_VERSION",
    "SCHEMA_CANDIDATE_ID",
    "SHEET_NAMES",
    "STATEMENT_CONTEXT_ID",
    "ShbMaturityReviewWorkbookArtifactsV1",
    "ShbMaturityReviewWorkbookV1Error",
    "build_shb_maturity_review_workbook_v1",
]
