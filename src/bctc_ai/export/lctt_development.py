"""Coverage-first LCTT development workbook export."""

from __future__ import annotations

import hashlib
import json
import os
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bctc_ai.core.contracts import ObservationKind
from bctc_ai.core.text import normalize_text
from bctc_ai.mapping.lctt import CashFlowMethod
from bctc_ai.mapping.lctt_item_mapping import (
    LCTTItemMappingResult,
    LCTTSchemaStatus,
    LCTTSourceRowStatus,
    validate_lctt_item_mapping_result,
)
from bctc_ai.tables.lctt_word_box import ParsedLCTTWordBoxDocument

LCTT_DEVELOPMENT_SHEETS = ("LCTT", "PROVENANCE", "RUN_METADATA")
LCTT_DEVELOPMENT_SCHEMA_COUNT = 107
LCTT_DEVELOPMENT_SOURCE_ROW_COUNT = 43
LCTT_DEVELOPMENT_MAPPED_SCHEMA_COUNT = 40
LCTT_DEVELOPMENT_MAPPED_CELL_COUNT = 80
LCTT_DEVELOPMENT_MAPPED_VALUE_COUNT = 67

_FIXED_TIMESTAMP = datetime(2000, 1, 1)


class LCTTDevelopmentExportError(ValueError):
    """Raised when an LCTT development export would overclaim or drift."""


@dataclass(frozen=True)
class LCTTDevelopmentArtifact:
    workbook_bytes: bytes
    workbook_sha256: str
    mapped_cell_count: int
    mapped_value_count: int
    provenance_row_count: int
    fully_verified: bool = False


@dataclass(frozen=True)
class LCTTDevelopmentExportResult:
    workbook_path: str
    workbook_sha256: str
    workbook_size_bytes: int
    mapped_cell_count: int
    mapped_value_count: int
    provenance_row_count: int
    fully_verified: bool = False


@dataclass(frozen=True)
class _ValidatedInputs:
    axes_by_role: dict[str, Any]
    axis_index_by_role: dict[str, int]
    rows_by_id: dict[str, Any]
    source_by_id: dict[str, Any]
    schema_by_id: dict[int, Any]


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _json_cell(payload: object) -> str:
    return json.dumps(
        payload,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _bbox_payload(bbox: Any) -> dict[str, object] | None:
    return None if bbox is None else asdict(bbox)


def _validate_inputs(
    parsed: ParsedLCTTWordBoxDocument,
    mapping: LCTTItemMappingResult,
) -> _ValidatedInputs:
    if not isinstance(parsed, ParsedLCTTWordBoxDocument):
        raise LCTTDevelopmentExportError("parsed LCTT document model is required")
    if not isinstance(mapping, LCTTItemMappingResult):
        raise LCTTDevelopmentExportError("LCTT item mapping model is required")
    try:
        validate_lctt_item_mapping_result(mapping)
    except ValueError as exc:
        raise LCTTDevelopmentExportError("LCTT mapping result is invalid") from exc
    if (
        parsed.scope != "CONSOLIDATED"
        or parsed.method is not CashFlowMethod.DIRECT
        or len(parsed.pages) != 2
        or len(parsed.rows) != LCTT_DEVELOPMENT_SOURCE_ROW_COUNT
        or parsed.cell_slot_count != 86
        or parsed.value_cell_count != 71
        or parsed.dash_cell_count != 9
        or parsed.blank_cell_count != 6
    ):
        raise LCTTDevelopmentExportError("LCTT parser coverage is not exportable")
    if (
        mapping.statement_type != "LCTT"
        or mapping.status != "PARTIAL_AUTOMATIC_MAPPING_WITH_UNRESOLVED_ITEMS"
        or not mapping.automatic_selection_allowed
        or mapping.schema_item_count != LCTT_DEVELOPMENT_SCHEMA_COUNT
        or mapping.schema_status_reconciled_count != LCTT_DEVELOPMENT_SCHEMA_COUNT
        or mapping.mapped_schema_count != LCTT_DEVELOPMENT_MAPPED_SCHEMA_COUNT
        or mapping.candidate_linked_schema_count != 1
        or mapping.label_conflict_schema_count != 1
        or mapping.ambiguous_schema_count != 5
        or mapping.not_observed_schema_count != 4
        or mapping.not_applicable_schema_count != 57
        or mapping.fully_verified_schema_count != 0
        or mapping.source_row_count != LCTT_DEVELOPMENT_SOURCE_ROW_COUNT
        or mapping.mapped_source_row_count != LCTT_DEVELOPMENT_MAPPED_SCHEMA_COUNT
        or mapping.candidate_linked_source_row_count != 1
        or mapping.label_conflict_source_row_count != 1
        or mapping.source_only_row_count != 2
    ):
        raise LCTTDevelopmentExportError("LCTT mapping coverage is not exportable")

    axes_by_role = {axis.current_or_comparative: axis for axis in parsed.axes}
    axis_index_by_role = {
        axis.current_or_comparative: index for index, axis in enumerate(parsed.axes)
    }
    if (
        len(parsed.axes) != 2
        or set(axes_by_role) != {"CURRENT", "COMPARATIVE"}
        or any(
            axis.period_type != "DURATION"
            or axis.duration_months != 3
            or axis.canonical_unit != "VND"
            or axis.unit_multiplier != 1_000_000
            for axis in parsed.axes
        )
    ):
        raise LCTTDevelopmentExportError("LCTT duration/unit axes drifted")

    rows_by_id = {row.row_id: row for row in parsed.rows}
    source_by_id = {item.row_id: item for item in mapping.source_dispositions}
    schema_by_id = {item.report_norm_id: item for item in mapping.schema_dispositions}
    if (
        len(rows_by_id) != LCTT_DEVELOPMENT_SOURCE_ROW_COUNT
        or set(rows_by_id) != set(source_by_id)
        or len(schema_by_id) != LCTT_DEVELOPMENT_SCHEMA_COUNT
    ):
        raise LCTTDevelopmentExportError("LCTT parser/mapping identity binding drifted")
    for order, row in enumerate(parsed.rows):
        source = source_by_id[row.row_id]
        if source.order != order or source.page_tag != row.page_tag or len(row.row.cells) != 2:
            raise LCTTDevelopmentExportError("LCTT source order/cell binding drifted")
        if source.status == LCTTSourceRowStatus.MAPPED_AUTOMATIC.value:
            if len(source.candidate_report_norm_ids) != 1:
                raise LCTTDevelopmentExportError("mapped LCTT source row is not one-to-one")
            schema_item = schema_by_id[source.candidate_report_norm_ids[0]]
            if (
                schema_item.status != LCTTSchemaStatus.MAPPED_AUTOMATIC.value
                or schema_item.candidate_source_row_ids != (row.row_id,)
            ):
                raise LCTTDevelopmentExportError("LCTT source/schema mapping cross-link drifted")
        elif source.status == LCTTSourceRowStatus.LABEL_CONFLICT_CANDIDATE_NOT_AUTOMATIC.value:
            if source.candidate_report_norm_ids != (4140,):
                raise LCTTDevelopmentExportError("LCTT label-conflict source identity drifted")
            schema_item = schema_by_id[4140]
            if (
                schema_item.status != LCTTSchemaStatus.LABEL_CONFLICT_CANDIDATE_NOT_AUTOMATIC.value
                or schema_item.candidate_source_row_ids != (row.row_id,)
            ):
                raise LCTTDevelopmentExportError("LCTT label-conflict cross-link drifted")
        elif source.status != LCTTSourceRowStatus.SOURCE_ONLY_PDF_ROW.value:
            raise LCTTDevelopmentExportError("unresolved LCTT source row reached export")
    return _ValidatedInputs(
        axes_by_role=axes_by_role,
        axis_index_by_role=axis_index_by_role,
        rows_by_id=rows_by_id,
        source_by_id=source_by_id,
        schema_by_id=schema_by_id,
    )


def _style_headers(sheet: Any, *, start_column: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(start_column, sheet.max_column + 1):
        cell = sheet.cell(1, column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _has_formula(workbook: Any) -> bool:
    return any(
        cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def _load_template(template_bytes: bytes, mapping: LCTTItemMappingResult) -> tuple[Any, tuple]:
    try:
        workbook = load_workbook(BytesIO(template_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise LCTTDevelopmentExportError("cannot load LCTT workbook template") from exc
    sheet = workbook.active
    if (
        workbook.sheetnames != ["Sheet1"]
        or sheet.max_row != 108
        or sheet.max_column != 3
        or tuple(sheet.cell(1, column).value for column in range(1, 4))
        != (None, "ReportNormId", "ReportNormName")
    ):
        workbook.close()
        raise LCTTDevelopmentExportError("LCTT workbook template identity drifted")
    schema_by_order = {item.display_order: item for item in mapping.schema_dispositions}
    snapshot = []
    for row_index in range(1, 109):
        cells = tuple(sheet.cell(row_index, column) for column in range(1, 4))
        snapshot.append(tuple((cell.value, cell.data_type, cell.style_id) for cell in cells))
        if row_index == 1:
            continue
        display_order, report_norm_id, canonical_name = (cell.value for cell in cells)
        disposition = schema_by_order.get(display_order)
        if disposition is None or (
            disposition.report_norm_id,
            normalize_text(disposition.canonical_name),
        ) != (report_norm_id, normalize_text(canonical_name)):
            workbook.close()
            raise LCTTDevelopmentExportError("LCTT mapping/template schema binding drifted")
    if _has_formula(workbook):
        workbook.close()
        raise LCTTDevelopmentExportError("formulas are forbidden in LCTT development export")
    sheet.title = "LCTT"
    return workbook, tuple(snapshot)


def _canonical_value(cell: Any, axis: Any) -> int | None:
    if cell.observation is ObservationKind.VALUE:
        if not isinstance(cell.value, Decimal) or cell.value != cell.value.to_integral_value():
            raise LCTTDevelopmentExportError("LCTT visible value is not an integral decimal")
        return int(cell.value) * axis.unit_multiplier
    if (
        cell.observation not in {ObservationKind.DASH, ObservationKind.BLANK}
        or cell.value is not None
    ):
        raise LCTTDevelopmentExportError("unsupported LCTT cell observation reached export")
    return None


def _write_main_sheet(workbook: Any, context: _ValidatedInputs) -> tuple[int, int]:
    sheet = workbook["LCTT"]
    headers = (
        "CurrentValueVND",
        "CurrentStatus",
        "CurrentPeriodStart",
        "CurrentPeriodEnd",
        "CurrentUnit",
        "ComparativeValueVND",
        "ComparativeStatus",
        "ComparativePeriodStart",
        "ComparativePeriodEnd",
        "ComparativeUnit",
    )
    for column, header in enumerate(headers, start=4):
        sheet.cell(1, column, header)
    _style_headers(sheet, start_column=4)
    sheet.freeze_panes = "D2"

    mapped_cells = 0
    mapped_values = 0
    for row_index in range(2, 109):
        report_norm_id = sheet.cell(row_index, 2).value
        disposition = context.schema_by_id[report_norm_id]
        for role, value_column in (("CURRENT", 4), ("COMPARATIVE", 9)):
            axis = context.axes_by_role[role]
            sheet.cell(row_index, value_column + 2, axis.period_start.isoformat())
            sheet.cell(row_index, value_column + 3, axis.period_end.isoformat())
            sheet.cell(row_index, value_column + 4, axis.canonical_unit)
            if disposition.status != LCTTSchemaStatus.MAPPED_AUTOMATIC.value:
                sheet.cell(row_index, value_column + 1, disposition.status)
                continue
            source_row_id = disposition.candidate_source_row_ids[0]
            row = context.rows_by_id[source_row_id]
            cell = row.row.cells[context.axis_index_by_role[role]]
            value = _canonical_value(cell, axis)
            sheet.cell(row_index, value_column, value)
            sheet.cell(row_index, value_column + 1, cell.observation.value)
            if value is not None:
                sheet.cell(row_index, value_column).number_format = "#,##0;[Red](#,##0);-"
                mapped_values += 1
            mapped_cells += 1
    for column in range(4, 14):
        sheet.column_dimensions[get_column_letter(column)].width = 22
    sheet.auto_filter.ref = "A1:M108"
    if (
        mapped_cells != LCTT_DEVELOPMENT_MAPPED_CELL_COUNT
        or mapped_values != LCTT_DEVELOPMENT_MAPPED_VALUE_COUNT
    ):
        raise LCTTDevelopmentExportError("LCTT mapped cell/value denominator drifted")
    return mapped_cells, mapped_values


def _write_provenance(
    workbook: Any,
    parsed: ParsedLCTTWordBoxDocument,
    context: _ValidatedInputs,
) -> int:
    sheet = workbook.create_sheet("PROVENANCE")
    headers = (
        "RowId",
        "GlobalOrder",
        "PageTag",
        "PageRowOrdinal",
        "SourceLabel",
        "NoteReference",
        "SourceRowStatus",
        "CandidateReportNormIdsJson",
        "CurrentRawText",
        "CurrentValueReportedUnit",
        "CurrentStatus",
        "CurrentValueVND",
        "ComparativeRawText",
        "ComparativeValueReportedUnit",
        "ComparativeStatus",
        "ComparativeValueVND",
        "PeriodStartCurrent",
        "PeriodEndCurrent",
        "PeriodStartComparative",
        "PeriodEndComparative",
        "PeriodType",
        "CanonicalUnit",
        "UnitMultiplier",
        "LabelBboxJson",
        "CurrentValueBboxJson",
        "ComparativeValueBboxJson",
        "LabelLineIndicesJson",
        "CurrentValueLineIndicesJson",
        "ComparativeValueLineIndicesJson",
        "SupportingReaderIdsJson",
        "MappingReason",
        "FullyVerified",
    )
    sheet.append(headers)
    current_axis = context.axes_by_role["CURRENT"]
    comparative_axis = context.axes_by_role["COMPARATIVE"]
    current_index = context.axis_index_by_role["CURRENT"]
    comparative_index = context.axis_index_by_role["COMPARATIVE"]
    for global_order, row in enumerate(parsed.rows):
        source = context.source_by_id[row.row_id]
        current = row.row.cells[current_index]
        comparative = row.row.cells[comparative_index]
        sheet.append(
            (
                row.row_id,
                global_order,
                row.page_tag,
                row.ordinal,
                row.row.label,
                row.row.note_reference,
                source.status,
                _json_cell(list(source.candidate_report_norm_ids)),
                current.raw_text,
                str(current.value) if current.value is not None else None,
                current.observation.value,
                _canonical_value(current, current_axis),
                comparative.raw_text,
                str(comparative.value) if comparative.value is not None else None,
                comparative.observation.value,
                _canonical_value(comparative, comparative_axis),
                current_axis.period_start.isoformat(),
                current_axis.period_end.isoformat(),
                comparative_axis.period_start.isoformat(),
                comparative_axis.period_end.isoformat(),
                current_axis.period_type,
                current_axis.canonical_unit,
                current_axis.unit_multiplier,
                _json_cell(_bbox_payload(row.label_bbox)),
                _json_cell(_bbox_payload(row.value_bboxes[current_index])),
                _json_cell(_bbox_payload(row.value_bboxes[comparative_index])),
                _json_cell(list(row.label_line_indices)),
                _json_cell(list(row.value_line_indices[current_index])),
                _json_cell(list(row.value_line_indices[comparative_index])),
                _json_cell(list(source.supporting_reader_ids)),
                source.reason,
                False,
            )
        )
    _style_headers(sheet)
    for column in (5, 24, 25, 26, 31):
        sheet.column_dimensions[get_column_letter(column)].width = 52
    if sheet.max_row != LCTT_DEVELOPMENT_SOURCE_ROW_COUNT + 1:
        raise LCTTDevelopmentExportError("LCTT provenance row denominator drifted")
    return LCTT_DEVELOPMENT_SOURCE_ROW_COUNT


def _write_metadata(
    workbook: Any,
    *,
    parsed: ParsedLCTTWordBoxDocument,
    mapping: LCTTItemMappingResult,
    mapped_cell_count: int,
    mapped_value_count: int,
) -> None:
    sheet = workbook.create_sheet("RUN_METADATA")
    sheet.append(("Key", "Value"))
    metadata: dict[str, object] = {
        "dataset_role": "DEVELOPMENT",
        "fully_verified": False,
        "mapping.ambiguous_count": mapping.ambiguous_schema_count,
        "mapping.automatic_selection_allowed": mapping.automatic_selection_allowed,
        "mapping.label_conflict_count": mapping.label_conflict_schema_count,
        "mapping.mapped_count": mapping.mapped_schema_count,
        "mapping.not_applicable_count": mapping.not_applicable_schema_count,
        "mapping.not_observed_count": mapping.not_observed_schema_count,
        "mapping.status": mapping.status,
        "numeric.independently_verified": False,
        "numeric.reader": "PP_OCRV6_WORD_BOX_SINGLE_READER",
        "parser.blank_cell_count": parsed.blank_cell_count,
        "parser.dash_cell_count": parsed.dash_cell_count,
        "parser.source_row_count": len(parsed.rows),
        "parser.value_cell_count": parsed.value_cell_count,
        "report_scope": parsed.scope,
        "schema.reconciled_count": mapping.schema_status_reconciled_count,
        "source.source_only_count": mapping.source_only_row_count,
        "statement_type": "LCTT",
        "target.mapped_cell_count": mapped_cell_count,
        "target.mapped_value_count": mapped_value_count,
    }
    for key, value in sorted(metadata.items()):
        sheet.append((key, value))
    _style_headers(sheet)
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 68


def _deterministic_workbook_bytes(workbook: Any) -> bytes:
    workbook.properties.creator = "bctc-ai"
    workbook.properties.lastModifiedBy = "bctc-ai"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP
    raw = BytesIO()
    workbook.save(raw)
    source = BytesIO(raw.getvalue())
    target = BytesIO()
    with (
        zipfile.ZipFile(source, "r") as archive,
        zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as output,
    ):
        for name in sorted(archive.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            info.flag_bits = 0x800
            output.writestr(info, archive.read(name), compresslevel=9)
    return target.getvalue()


def _verify_serialized_workbook(
    workbook_bytes: bytes,
    *,
    template_snapshot: tuple,
    mapped_value_count: int,
) -> None:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    except Exception as exc:
        raise LCTTDevelopmentExportError("serialized LCTT workbook cannot be reopened") from exc
    try:
        if tuple(workbook.sheetnames) != LCTT_DEVELOPMENT_SHEETS or _has_formula(workbook):
            raise LCTTDevelopmentExportError("serialized LCTT workbook contract drifted")
        sheet = workbook["LCTT"]
        actual_snapshot = tuple(
            tuple(
                (
                    sheet.cell(row_index, column).value,
                    sheet.cell(row_index, column).data_type,
                    sheet.cell(row_index, column).style_id,
                )
                for column in range(1, 4)
            )
            for row_index in range(1, 109)
        )
        if actual_snapshot != template_snapshot:
            raise LCTTDevelopmentExportError("LCTT template columns A:C were not preserved")
        exported = sum(
            sheet.cell(row_index, column).value is not None
            for row_index in range(2, 109)
            for column in (4, 9)
        )
        if exported != mapped_value_count or workbook["PROVENANCE"].max_row != 44:
            raise LCTTDevelopmentExportError("serialized LCTT coverage drifted")
    finally:
        workbook.close()


def build_lctt_development_artifact(
    *,
    template_path: Path,
    parsed: ParsedLCTTWordBoxDocument,
    mapping: LCTTItemMappingResult,
) -> LCTTDevelopmentArtifact:
    """Build deterministic workbook bytes without writing an output artifact."""

    try:
        template_bytes = Path(template_path).read_bytes()
    except OSError as exc:
        raise LCTTDevelopmentExportError("cannot read LCTT workbook template") from exc
    context = _validate_inputs(parsed, mapping)
    workbook, snapshot = _load_template(template_bytes, mapping)
    try:
        mapped_cell_count, mapped_value_count = _write_main_sheet(workbook, context)
        provenance_row_count = _write_provenance(workbook, parsed, context)
        _write_metadata(
            workbook,
            parsed=parsed,
            mapping=mapping,
            mapped_cell_count=mapped_cell_count,
            mapped_value_count=mapped_value_count,
        )
        if _has_formula(workbook):
            raise LCTTDevelopmentExportError("formulas are forbidden in LCTT development export")
        workbook_bytes = _deterministic_workbook_bytes(workbook)
    finally:
        workbook.close()
    _verify_serialized_workbook(
        workbook_bytes,
        template_snapshot=snapshot,
        mapped_value_count=mapped_value_count,
    )
    return LCTTDevelopmentArtifact(
        workbook_bytes=workbook_bytes,
        workbook_sha256=_sha256(workbook_bytes),
        mapped_cell_count=mapped_cell_count,
        mapped_value_count=mapped_value_count,
        provenance_row_count=provenance_row_count,
    )


def export_lctt_development(
    *,
    template_path: Path,
    workbook_path: Path,
    parsed: ParsedLCTTWordBoxDocument,
    mapping: LCTTItemMappingResult,
) -> LCTTDevelopmentExportResult:
    """Write one development workbook, refusing any overwrite or symlink target."""

    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if os.path.lexists(workbook_path):
        raise LCTTDevelopmentExportError("LCTT development export refuses to overwrite output")
    artifact = build_lctt_development_artifact(
        template_path=template_path,
        parsed=parsed,
        mapping=mapping,
    )
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(workbook_path, flags, 0o644)
    except OSError as exc:
        raise LCTTDevelopmentExportError(
            f"LCTT output already exists or cannot be created: {workbook_path}"
        ) from exc
    try:
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(artifact.workbook_bytes)
            stream.flush()
            os.fsync(stream.fileno())
    except Exception:
        workbook_path.unlink(missing_ok=True)
        raise
    return LCTTDevelopmentExportResult(
        workbook_path=workbook_path.as_posix(),
        workbook_sha256=artifact.workbook_sha256,
        workbook_size_bytes=len(artifact.workbook_bytes),
        mapped_cell_count=artifact.mapped_cell_count,
        mapped_value_count=artifact.mapped_value_count,
        provenance_row_count=artifact.provenance_row_count,
    )


__all__ = [
    "LCTT_DEVELOPMENT_MAPPED_CELL_COUNT",
    "LCTT_DEVELOPMENT_MAPPED_SCHEMA_COUNT",
    "LCTT_DEVELOPMENT_MAPPED_VALUE_COUNT",
    "LCTT_DEVELOPMENT_SCHEMA_COUNT",
    "LCTT_DEVELOPMENT_SHEETS",
    "LCTT_DEVELOPMENT_SOURCE_ROW_COUNT",
    "LCTTDevelopmentArtifact",
    "LCTTDevelopmentExportError",
    "LCTTDevelopmentExportResult",
    "build_lctt_development_artifact",
    "export_lctt_development",
]
