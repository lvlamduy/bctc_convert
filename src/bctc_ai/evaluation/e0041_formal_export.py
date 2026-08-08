"""Formal E-0041 capture and deterministic two-file hash sealing.

This integration is intentionally narrower than the development core.  It
authenticates one fixed E-0040 mapping chain, opens the E-0037 postjoin only
after that chain has minted an immutable carrier, authenticates every geometry
dependency before opening the template, and then invokes the pinned E-0041
core from in-memory snapshots.  Formal publication is exclusive and never
overwrites an existing pair or seal.
"""

from __future__ import annotations

import hashlib
import json
import os
import platform
import re
import stat
import subprocess
import sys
import zlib
from collections import Counter
from collections.abc import Callable, Mapping, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from importlib.metadata import version as distribution_version
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any, cast

import yaml
from openpyxl import load_workbook
from yaml.constructor import ConstructorError
from yaml.events import AliasEvent

from bctc_ai.evaluation import e0041_post_mapping_export as _core


class E0041FormalExportError(RuntimeError):
    """Raised when the formal E-0041 contract cannot be proved safely."""


CONTROL_RELATIVE_PATH = Path("config/experiments/e0041-mbb-cdkt-formal-export.yaml")
OUTPUT_DIRECTORY_RELATIVE_PATH = Path(
    "output/calibration/e0041-mbb-cdkt-post-mapping-development-excel"
)
WORKBOOK_RELATIVE_PATH = OUTPUT_DIRECTORY_RELATIVE_PATH / "mbb-cdkt-development.xlsx"
PROVENANCE_RELATIVE_PATH = OUTPUT_DIRECTORY_RELATIVE_PATH / "provenance.json"
SEAL_RELATIVE_PATH = Path(
    "docs/experiments/E-0041-mbb-cdkt-post-mapping-development-excel-seal.json"
)

PAIR_STATE = "FORMAL_POST_MAPPING_EXPORT_PAIR_READY_FOR_HASH_SEAL"
SEAL_STATE = "FORMAL_POST_MAPPING_EXPORT_PAIR_HASH_SEALED"
CORE_RELATIVE_PATH = Path("src/bctc_ai/evaluation/e0041_post_mapping_export.py")
CORE_SHA256 = "c40dd4f31d80fd18e330984c7f6c43c1f45d8b96a05d1a19d2476c440a7e14f5"
CORE_SIZE_BYTES = 116_528
LEGACY_GEOMETRY_ROOT = PurePosixPath("/workspace/bctc-ai")

_SHA256 = re.compile(r"[0-9a-f]{64}")
_GIT_COMMIT = re.compile(r"[0-9a-f]{40}")
_GIT_OBJECT = re.compile(r"(?:[0-9a-f]{40}|[0-9a-f]{64})")
_MAX_JSON_BYTES = 32 * 1024 * 1024
_MAX_YAML_BYTES = 1024 * 1024
_NORMALIZED_REGISTRY_SHA256 = "e834efd4f6e70c03e607d834a17adc69e0fa0868658767637c5db3cf8c06be6a"
_NORMALIZED_REGISTRY_SIZE = 161_120
_ASSET_INVENTORY_SHA256 = "7a50f4737c8381bc73a04ad4fce9abd200b64b917e9282a0197bf56d9a044cc0"
_ASSET_INVENTORY_SIZE = 31_130
_ASSET_COUNT = 134
_ASSET_TOTAL_SIZE = 4_489_853
_PROJECTION_SHA256 = "4ba86ea84b102932b5ae2952f2998c205907d58095793535eacfd100c501db8d"
_PROJECTION_SIZE = 442_178
_WORKBOOK_SHA256 = "60c5f6d01ba0a11276da6b6fc7f85d4d4f4893b6ee70d0ca54bd14764a922caf"
_WORKBOOK_SIZE = 43_746
_TEMPLATE_FIDELITY_SHA256 = "38c74d1af3be5a70e4dd27d18030e4950f96469a194c6d06f9843a1467831408"
_PHYSICAL_EQUATIONS_SHA256 = "a611078b4734d1e57026d58db5aced4a0b342114ba57aefdff29868032b3b42b"
_RUNTIME_VERSIONS = {
    "python": "3.11.10",
    "openpyxl": "3.1.5",
    "pyyaml": "6.0.3",
    "et_xmlfile": "2.0.0",
    "zlib": "1.3.1",
}
_VALIDATION_ORDER = [
    "CONTROL_IMPLEMENTATION_RUNTIME_AND_CLEAN_GIT",
    "E0040_MAPPING_SEAL",
    "E0040_S3_REGISTRATION",
    "SHARED_S3_REGISTRY_FROZEN_BASELINE",
    "E0040_MAPPING",
    "AUTHENTICATED_E0040_RESULT_CARRIER",
    "E0037_POSTJOIN",
    "RECONSTRUCTED_GEOMETRY_REGISTRY",
    "EXACT_134_GEOMETRY_ASSET_INVENTORY_AND_ROOT_NORMALIZATION",
    "CDKT_TEMPLATE",
    "POST_MAPPING_PROJECTION_WORKBOOK_AND_PROVENANCE",
]
_FORBIDDEN_MODULE_TOKENS = (
    "e0038",
    "e0039",
    "review",
    "history",
    "mongodb",
    "qwen",
    "holdout",
)
_TRACKED_INPUT_NAMES = frozenset(
    {
        "e0040_mapping_seal",
        "e0040_s3_registration",
        "shared_s3_registry_frozen_baseline",
        "e0037_postjoin",
        "geometry_crop_policy",
        "geometry_row_contract",
        "cdkt_template",
    }
)

_CLAIM_BOUNDARY = (
    "E-0041 formally captures one calibration-only post-mapping workbook and its "
    "provenance after authenticating the exact E-0040 mapping chain, joining the "
    "E-0037 period/unit/numeric evidence only by row_id, and verifying the exact "
    "geometry and template inputs. It does not establish schema authority, mapping "
    "accuracy, review approval, accounting correctness, human gold, holdout "
    "performance, or production readiness."
)


class _UniqueKeySafeLoader(yaml.SafeLoader):
    def compose_node(self, parent: yaml.Node | None, index: int | None) -> yaml.Node:
        if self.check_event(AliasEvent):
            event = self.peek_event()
            raise ConstructorError(
                "while composing E-0041 formal control",
                getattr(event, "start_mark", None),
                "YAML aliases are forbidden",
                getattr(event, "start_mark", None),
            )
        return super().compose_node(parent, index)


def _construct_unique_mapping(
    loader: _UniqueKeySafeLoader,
    node: yaml.MappingNode,
    deep: bool = False,
) -> dict[Any, Any]:
    result: dict[Any, Any] = {}
    for key_node, value_node in node.value:
        key = loader.construct_object(key_node, deep=deep)
        try:
            duplicate = key in result
        except TypeError as exc:
            raise ConstructorError(
                "while constructing E-0041 control",
                node.start_mark,
                "found an unhashable key",
                key_node.start_mark,
            ) from exc
        if duplicate:
            raise ConstructorError(
                "while constructing E-0041 control",
                node.start_mark,
                f"found duplicate key {key!r}",
                key_node.start_mark,
            )
        result[key] = loader.construct_object(value_node, deep=deep)
    return result


_UniqueKeySafeLoader.add_constructor(
    yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
    _construct_unique_mapping,
)


@dataclass(frozen=True, slots=True)
class _StableFile:
    name: str
    path: Path
    payload: bytes
    identity: tuple[int, int, int, int, int, int]
    artifact: dict[str, Any]
    require_head_blob: bool


@dataclass(frozen=True, slots=True)
class _GitSnapshot:
    commit: str
    upstream_commit: str


@dataclass(frozen=True, slots=True)
class _Prerequisites:
    control: dict[str, Any]
    control_file: _StableFile
    implementations: dict[str, _StableFile]
    runtime_files: dict[str, _StableFile]


@dataclass(frozen=True, slots=True)
class _Materials:
    prerequisites: _Prerequisites
    ordered_files: tuple[_StableFile, ...]
    seal_file: _StableFile
    registration_file: _StableFile
    shared_registry_file: _StableFile
    mapping_file: _StableFile
    postjoin_file: _StableFile
    geometry_file: _StableFile
    template_file: _StableFile
    asset_files: tuple[_StableFile, ...]
    asset_inventory: tuple[dict[str, Any], ...]
    normalized_registry_artifact: dict[str, Any]


@dataclass(frozen=True, slots=True)
class FormalExportBuild:
    """In-memory formal candidate; no output has been published."""

    projection: dict[str, Any]
    projection_bytes: bytes
    workbook_bytes: bytes
    provenance: dict[str, Any]
    provenance_bytes: bytes
    materials: _Materials


StableReader = Callable[[Path, Path, int, str], bytes]


def _fail(message: str, error: BaseException | None = None) -> E0041FormalExportError:
    if error is None:
        return E0041FormalExportError(message)
    return E0041FormalExportError(f"{message}: {error}")


def _reject_nonfinite(value: str) -> None:
    raise ValueError(f"non-finite JSON constant: {value}")


def _reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _decode_json(payload: bytes, name: str) -> dict[str, Any]:
    if type(payload) is not bytes or not payload or len(payload) > _MAX_JSON_BYTES:
        raise _fail(f"{name} byte size is invalid")
    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=_reject_duplicate_keys,
            parse_constant=_reject_nonfinite,
        )
    except (UnicodeDecodeError, json.JSONDecodeError, ValueError, RecursionError) as exc:
        raise _fail(f"cannot decode {name} as strict JSON", exc) from exc
    if not isinstance(value, dict):
        raise _fail(f"{name} must be a JSON object")
    return value


def _decode_control(payload: bytes) -> dict[str, Any]:
    if type(payload) is not bytes or not payload or len(payload) > _MAX_YAML_BYTES:
        raise _fail("E-0041 formal control byte size is invalid")
    try:
        value = yaml.load(payload.decode("utf-8"), Loader=_UniqueKeySafeLoader)
    except (UnicodeDecodeError, yaml.YAMLError, ValueError, RecursionError) as exc:
        raise _fail("cannot decode E-0041 formal control", exc) from exc
    if not isinstance(value, dict):
        raise _fail("E-0041 formal control must be a YAML object")
    return value


def _compact_bytes(value: object) -> bytes:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    except (TypeError, ValueError, RecursionError) as exc:
        raise _fail("value is not canonical finite JSON", exc) from exc


def _encoded_json(value: Mapping[str, Any]) -> bytes:
    try:
        return (
            json.dumps(
                value,
                ensure_ascii=False,
                sort_keys=True,
                indent=2,
                allow_nan=False,
            )
            + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError, RecursionError) as exc:
        raise _fail("value is not canonical finite JSON", exc) from exc


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _stat_identity(value: os.stat_result) -> tuple[int, int, int, int, int, int]:
    return (
        value.st_dev,
        value.st_ino,
        value.st_mode,
        value.st_size,
        value.st_mtime_ns,
        value.st_ctime_ns,
    )


def _mapping(value: object, name: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise _fail(f"{name} must be an object")
    return cast(dict[str, Any], value)


def _sequence(value: object, name: str) -> list[Any]:
    if not isinstance(value, list):
        raise _fail(f"{name} must be an array")
    return value


def _artifact_identity(value: object, name: str) -> dict[str, Any]:
    record = _mapping(value, name)
    result = {key: record.get(key) for key in ("path", "sha256", "size_bytes")}
    path = result["path"]
    digest = result["sha256"]
    size = result["size_bytes"]
    if (
        type(path) is not str
        or not path
        or PurePosixPath(path).is_absolute()
        or ".." in PurePosixPath(path).parts
        or type(digest) is not str
        or _SHA256.fullmatch(digest) is None
        or type(size) is not int
        or size <= 0
    ):
        raise _fail(f"{name} artifact identity is invalid")
    return result


def _canonical_path(project_root: Path, relative: str | Path, name: str) -> Path:
    raw = Path(relative)
    if raw.is_absolute() or not raw.parts or ".." in raw.parts:
        raise _fail(f"{name} must be a safe project-relative path")
    return project_root / raw


def _default_reader(project_root: Path, path: Path, maximum_size: int, name: str) -> bytes:
    try:
        return _core._stable_read(project_root, path, maximum_size=maximum_size, name=name)
    except Exception as exc:
        if isinstance(exc, E0041FormalExportError):
            raise
        raise _fail(f"cannot stable-read {name}", exc) from exc


def _read_file(
    project_root: Path,
    record: Mapping[str, Any],
    name: str,
    *,
    reader: StableReader,
    maximum_size: int,
    require_head_blob: bool,
) -> _StableFile:
    normalized = _artifact_identity(record, name)
    path = _canonical_path(project_root, cast(str, normalized["path"]), name)
    try:
        before = os.stat(path, follow_symlinks=False)
    except OSError as exc:
        raise _fail(f"cannot stat {name}", exc) from exc
    if not stat.S_ISREG(before.st_mode):
        raise _fail(f"{name} is not a regular file")
    payload = reader(project_root, path, maximum_size, name)
    try:
        after = os.stat(path, follow_symlinks=False)
    except OSError as exc:
        raise _fail(f"cannot restat {name}", exc) from exc
    if _stat_identity(before) != _stat_identity(after):
        raise _fail(f"{name} identity changed across stable read")
    if len(payload) != normalized["size_bytes"] or _sha256(payload) != normalized["sha256"]:
        raise _fail(f"{name} byte identity drifted")
    return _StableFile(
        name=name,
        path=path,
        payload=payload,
        identity=_stat_identity(after),
        artifact=dict(normalized),
        require_head_blob=require_head_blob,
    )


def _record_for_payload(path: Path, payload: bytes) -> dict[str, Any]:
    return {"path": path.as_posix(), "sha256": _sha256(payload), "size_bytes": len(payload)}


def _assert_process_isolation() -> None:
    contaminated = sorted(
        name
        for name in sys.modules
        if name.startswith("bctc_ai.")
        and name != __name__
        and any(token in name.casefold() for token in _FORBIDDEN_MODULE_TOKENS)
    )
    if contaminated:
        raise _fail(
            "formal E-0041 process is contaminated by forbidden preloaded modules: "
            + ", ".join(contaminated)
        )


def _assert_no_absolute_paths(value: object, name: str, *, depth: int = 0) -> None:
    if depth > 128:
        raise _fail(f"{name} nesting is excessive")
    if isinstance(value, dict):
        for key, item in value.items():
            if not isinstance(key, str):
                raise _fail(f"{name} contains a non-string key")
            _assert_no_absolute_paths(item, name, depth=depth + 1)
    elif isinstance(value, list):
        for item in value:
            _assert_no_absolute_paths(item, name, depth=depth + 1)
    elif isinstance(value, str):
        if value.startswith("/") or re.match(r"^[A-Za-z]:[\\/]", value):
            raise _fail(f"{name} contains an absolute path")


def _sanitized_git_environment() -> dict[str, str]:
    environment = {key: value for key, value in os.environ.items() if not key.startswith("GIT_")}
    environment.update(
        {
            "GIT_CONFIG_NOSYSTEM": "1",
            "GIT_CONFIG_GLOBAL": os.devnull,
            "GIT_OPTIONAL_LOCKS": "0",
            "GIT_LITERAL_PATHSPECS": "1",
        }
    )
    return environment


def _git_process(
    project_root: Path,
    *arguments: str,
    check: bool = True,
) -> subprocess.CompletedProcess[bytes]:
    command = [
        "git",
        "-c",
        "core.fsmonitor=false",
        "-c",
        "core.untrackedCache=false",
        "-c",
        "diff.external=",
        *arguments,
    ]
    try:
        result = subprocess.run(
            command,
            cwd=project_root,
            env=_sanitized_git_environment(),
            stdin=subprocess.DEVNULL,
            capture_output=True,
            check=False,
        )
    except OSError as exc:
        raise _fail("cannot execute sanitized Git command", exc) from exc
    if check and result.returncode != 0:
        stderr = result.stderr.decode("utf-8", errors="replace").strip()
        raise _fail(f"sanitized Git command failed ({' '.join(arguments)}): {stderr}")
    return result


def _git_text(project_root: Path, *arguments: str) -> str:
    payload = _git_process(project_root, *arguments).stdout
    try:
        return payload.decode("utf-8").strip()
    except UnicodeDecodeError as exc:
        raise _fail("sanitized Git output is not UTF-8", exc) from exc


def _clean_git_snapshot(project_root: Path) -> _GitSnapshot:
    if not project_root.is_absolute() or project_root != project_root.resolve():
        raise _fail("formal E-0041 project root must be a canonical absolute path")
    if _git_text(project_root, "rev-parse", "--is-inside-work-tree") != "true":
        raise _fail("formal E-0041 root is not inside a Git worktree")
    try:
        reported_root = Path(_git_text(project_root, "rev-parse", "--show-toplevel")).resolve()
    except (OSError, RuntimeError) as exc:
        raise _fail("cannot resolve sanitized Git worktree root", exc) from exc
    if reported_root != project_root or _git_text(project_root, "rev-parse", "--show-prefix"):
        raise _fail("formal E-0041 project root is not the exact Git worktree root")
    commit = _git_text(project_root, "rev-parse", "--verify", "HEAD^{commit}")
    if _GIT_COMMIT.fullmatch(commit) is None:
        raise _fail("formal E-0041 HEAD commit is invalid")
    status = _git_process(
        project_root,
        "status",
        "--porcelain=v1",
        "-z",
        "--untracked-files=all",
        "--ignore-submodules=none",
    )
    if status.stdout:
        raise _fail("formal E-0041 capture requires a clean Git worktree and index")
    for arguments, label in (
        (
            ("diff", "--no-ext-diff", "--ignore-submodules=none", "--quiet", "--"),
            "working tree",
        ),
        (
            (
                "diff",
                "--cached",
                "--no-ext-diff",
                "--ignore-submodules=none",
                "--quiet",
                "HEAD",
                "--",
            ),
            "index",
        ),
        (("diff-files", "--ignore-submodules=none", "--quiet", "--"), "working index"),
        (
            ("diff-index", "--cached", "--ignore-submodules=none", "--quiet", "HEAD", "--"),
            "HEAD index",
        ),
    ):
        result = _git_process(project_root, *arguments, check=False)
        if result.returncode != 0:
            raise _fail(f"formal E-0041 sanitized Git {label} is not clean")
    if _git_process(project_root, "ls-files", "-u", "-z").stdout:
        raise _fail("formal E-0041 index contains unmerged entries")
    upstream = _git_text(project_root, "rev-parse", "--verify", "@{upstream}^{commit}")
    if _GIT_COMMIT.fullmatch(upstream) is None or upstream != commit:
        raise _fail("formal E-0041 HEAD is not identical to its configured upstream")
    divergence = _git_text(
        project_root,
        "rev-list",
        "--left-right",
        "--count",
        "HEAD...@{upstream}",
    ).split()
    if divergence != ["0", "0"]:
        raise _fail("formal E-0041 HEAD/upstream divergence is nonzero")
    return _GitSnapshot(commit=commit, upstream_commit=upstream)


def _assert_head_blob_binding(project_root: Path, stable: _StableFile) -> None:
    if not stable.require_head_blob:
        return
    relative = stable.path.relative_to(project_root).as_posix()
    stage_output = _git_process(
        project_root,
        "ls-files",
        "--stage",
        "-z",
        "--",
        relative,
    ).stdout
    entries = [item for item in stage_output.split(b"\0") if item]
    if len(entries) != 1:
        raise _fail(f"tracked ledger has no unique normal index entry: {relative}")
    try:
        metadata, encoded_path = entries[0].split(b"\t", 1)
        mode, index_blob, stage = metadata.decode("ascii").split()
        listed_path = encoded_path.decode("utf-8")
    except (ValueError, UnicodeDecodeError) as exc:
        raise _fail(f"cannot parse index entry for {relative}", exc) from exc
    if (
        listed_path != relative
        or stage != "0"
        or mode not in {"100644", "100755"}
        or _GIT_OBJECT.fullmatch(index_blob) is None
    ):
        raise _fail(f"tracked ledger index entry is unsafe: {relative}")
    verbose = _git_text(project_root, "ls-files", "-v", "--", relative)
    fsmonitor = _git_text(project_root, "ls-files", "-f", "--", relative)
    if verbose != f"H {relative}" or fsmonitor != f"H {relative}":
        raise _fail(f"tracked ledger has non-normal index flags: {relative}")
    debug = _git_text(project_root, "ls-files", "--debug", "--", relative)
    if not debug.endswith("flags: 0"):
        raise _fail(f"tracked ledger index flags are nonzero: {relative}")
    tree = _git_process(project_root, "ls-tree", "-z", "HEAD", "--", relative).stdout
    tree_entries = [item for item in tree.split(b"\0") if item]
    if len(tree_entries) != 1:
        raise _fail(f"tracked ledger is not a unique HEAD tree entry: {relative}")
    try:
        tree_metadata, tree_path = tree_entries[0].split(b"\t", 1)
        head_mode, object_type, head_blob = tree_metadata.decode("ascii").split()
        decoded_tree_path = tree_path.decode("utf-8")
    except (ValueError, UnicodeDecodeError) as exc:
        raise _fail(f"cannot parse HEAD tree entry for {relative}", exc) from exc
    if (
        decoded_tree_path != relative
        or object_type != "blob"
        or head_mode != mode
        or head_blob != index_blob
    ):
        raise _fail(f"tracked ledger HEAD/index binding drifted: {relative}")
    head_payload = _git_process(project_root, "cat-file", "blob", head_blob).stdout
    if head_payload != stable.payload:
        raise _fail(f"tracked ledger working bytes differ from HEAD blob: {relative}")


def _assert_runtime_versions() -> None:
    try:
        observed = {
            "python": platform.python_version(),
            "openpyxl": distribution_version("openpyxl"),
            "pyyaml": distribution_version("PyYAML"),
            "et_xmlfile": distribution_version("et_xmlfile"),
            "zlib": zlib.ZLIB_RUNTIME_VERSION,
        }
    except Exception as exc:
        raise _fail("cannot resolve formal runtime versions", exc) from exc
    if observed != _RUNTIME_VERSIONS:
        raise _fail(f"formal runtime version drifted: {observed!r}")


def _validate_control(control: Mapping[str, Any]) -> None:
    expected_keys = {
        "version",
        "experiment_id",
        "dataset_role",
        "design",
        "state",
        "phase_outputs",
        "input_authority",
        "implementation",
        "runtime_authority",
        "validation_order",
        "formal_oracle",
        "accounting_validation",
        "publication",
        "forbidden_inputs",
        "limitations",
        "claim_boundary",
    }
    if set(control) != expected_keys:
        raise _fail("E-0041 formal control keyset drifted")
    if (
        control.get("version") != 1
        or control.get("experiment_id") != "E-0041"
        or control.get("dataset_role") != "CALIBRATION"
        or control.get("design")
        != "AUTHENTICATED_E0040_POSTJOIN_GEOMETRY_TEMPLATE_FORMAL_PAIR_THEN_REPLAY_SEAL"
        or control.get("state") != "READY_FOR_FORMAL_POST_MAPPING_EXPORT_PAIR_THEN_HASH_SEAL"
        or control.get("validation_order") != _VALIDATION_ORDER
        or control.get("claim_boundary") != _CLAIM_BOUNDARY
    ):
        raise _fail("E-0041 formal control identity drifted")
    outputs = _mapping(control.get("phase_outputs"), "formal phase outputs")
    if outputs != {
        "export_pair": {
            "directory": OUTPUT_DIRECTORY_RELATIVE_PATH.as_posix(),
            "workbook": WORKBOOK_RELATIVE_PATH.name,
            "provenance_json": PROVENANCE_RELATIVE_PATH.name,
            "provenance_state": PAIR_STATE,
            "pair_hash_sealed": False,
            "exact_inventory_file_count": 2,
        },
        "export_seal": {
            "path": SEAL_RELATIVE_PATH.as_posix(),
            "required_state": SEAL_STATE,
            "exact_inventory_file_count": 2,
        },
    }:
        raise _fail("E-0041 formal phase-output contract drifted")
    inputs = _mapping(control.get("input_authority"), "formal input authority")
    required_inputs = {
        "e0040_mapping_seal",
        "e0040_s3_registration",
        "shared_s3_registry_frozen_baseline",
        "e0040_mapping",
        "e0037_postjoin",
        "reconstructed_geometry_registry",
        "geometry_crop_policy",
        "geometry_row_contract",
        "cdkt_template",
    }
    if set(inputs) != required_inputs:
        raise _fail("E-0041 formal input authority keyset drifted")
    input_record_keys = {
        "e0040_mapping_seal": {"path", "sha256", "size_bytes"},
        "e0040_s3_registration": {"path", "sha256", "size_bytes"},
        "shared_s3_registry_frozen_baseline": {
            "path",
            "sha256",
            "size_bytes",
            "authority",
        },
        "e0040_mapping": {"path", "sha256", "size_bytes"},
        "e0037_postjoin": {"path", "sha256", "size_bytes", "authority"},
        "reconstructed_geometry_registry": {
            "path",
            "sha256",
            "size_bytes",
            "authority",
        },
        "geometry_crop_policy": {"path", "sha256", "size_bytes"},
        "geometry_row_contract": {"path", "sha256", "size_bytes"},
        "cdkt_template": {"path", "sha256", "size_bytes"},
    }
    if any(
        set(_mapping(inputs[name], f"formal input {name}")) != keys
        for name, keys in input_record_keys.items()
    ):
        raise _fail("E-0041 formal input record nested keyset drifted")
    if (
        inputs["shared_s3_registry_frozen_baseline"].get("authority")
        != "FROZEN_DURABILITY_BASELINE_ONLY_NOT_MAPPING_AUTHORITY"
        or inputs["e0037_postjoin"].get("authority")
        != "PERIOD_UNIT_SCOPE_AND_NESTED_NUMERIC_EVIDENCE_ONLY"
        or inputs["reconstructed_geometry_registry"].get("authority")
        != "RECONSTRUCTED_GEOMETRY_WITH_CELL_SHA_PARITY"
    ):
        raise _fail("E-0041 formal input authority annotation drifted")
    for name in required_inputs:
        _artifact_identity(inputs[name], f"formal input {name}")
    expected_inputs = {
        "e0040_mapping_seal": {
            "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-seal.json",
            "sha256": "68306f7f540faa77d6e2e383927eae23fc3724cfdc8c53cded978a86f3a00b29",
            "size_bytes": 7_611,
        },
        "e0040_s3_registration": {
            "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-s3-registration.json",
            "sha256": "f38d9a1bbed4ec48e2156d441e5c76c6e6d82b0771208de3eef92d96173dd4b5",
            "size_bytes": 13_360,
        },
        "shared_s3_registry_frozen_baseline": {
            "path": "data/registered/s3_artifact_snapshot_registry.jsonl",
            "sha256": "25da6b205a775d87eca8e4ffe55e3f762ee64e92cbb7190c2834708a7de0d78d",
            "size_bytes": 6_050,
        },
        "e0040_mapping": {
            "path": "output/calibration/e0040-mbb-cdkt-formal-mapping/mapping_only.json",
            "sha256": "8def983007fc3aacf59351395426d5246ad3f28d605442f590de55eaf396cb0d",
            "size_bytes": 1_157_172,
        },
        "e0037_postjoin": {
            "path": "docs/experiments/E-0037-mbb-cdkt-sealed-evidence-mapping.json",
            "sha256": "a44146ff98ac9b33dd7f04037e69ba258ef7361dc158ecbd51a6688d7fbb6f7b",
            "size_bytes": 1_045_610,
        },
        "reconstructed_geometry_registry": {
            "path": (
                "output/calibration/e0041-mbb-cdkt-reconstructed-geometry/"
                "65fa9b7c0de1/crop_registry.json"
            ),
            "sha256": "65fa9b7c0de1f0db26ae57a46dae2bb64c2475e3a87e5194461f208fc786cbef",
            "size_bytes": 217_837,
        },
        "geometry_crop_policy": {
            "path": "config/tables/numeric-cell-crops-v2.yaml",
            "sha256": "f1dcbe668b70a9d71f08483f46aad01fc7cd719db96f8bd202d32554c5a7907b",
            "size_bytes": 1_099,
        },
        "geometry_row_contract": {
            "path": "docs/experiments/E-0033-mbb-cdkt-note-row-split-immutable.json",
            "sha256": "d9c0ecf44f6a0f652e6c991d3ab95b7ab0e821068366764e39a3f0de7f0711fb",
            "size_bytes": 190_290,
        },
        "cdkt_template": {
            "path": "template/Bank_CDKT_ReportNormId.xlsx",
            "sha256": "a07ff47f7c41011fe4ca5a66681106d476586ded9013b5874cbb9f67a6ad8486",
            "size_bytes": 10_945,
        },
    }
    if any(
        _artifact_identity(inputs[name], f"formal input {name}") != expected
        for name, expected in expected_inputs.items()
    ):
        raise _fail("E-0041 pinned formal input identity drifted")
    implementations = _mapping(control.get("implementation"), "formal implementation ledger")
    required_implementations = {
        "post_mapping_export_core",
        "e0040_challenger_carrier",
        "arithmetic_validation",
        "geometry_reconstruction_builder",
        "formal_export_integration",
        "capture_export_script",
        "capture_seal_script",
    }
    if set(implementations) != required_implementations:
        raise _fail("E-0041 formal implementation keyset drifted")
    for name in required_implementations:
        if set(_mapping(implementations[name], f"formal implementation {name}")) != {
            "path",
            "sha256",
            "size_bytes",
        }:
            raise _fail(f"formal implementation {name} nested keyset drifted")
        record = _artifact_identity(implementations[name], f"formal implementation {name}")
        if record["sha256"] == "0" * 64 or record["size_bytes"] <= 0:
            raise _fail(f"formal implementation {name} is still a placeholder")
    if implementations["post_mapping_export_core"] != {
        "path": CORE_RELATIVE_PATH.as_posix(),
        "sha256": CORE_SHA256,
        "size_bytes": CORE_SIZE_BYTES,
    }:
        raise _fail("pinned E-0041 post-mapping core identity drifted")
    fixed_implementations = {
        "e0040_challenger_carrier": {
            "path": "src/bctc_ai/mapping/e0040_calibration_challenger.py",
            "sha256": "c379ccf784868ec5b2f40714be00c402147b2b9a94e06b917a3e2cd6b926609b",
            "size_bytes": 47_333,
        },
        "arithmetic_validation": {
            "path": "src/bctc_ai/validation/arithmetic.py",
            "sha256": "3841fa0c40064554134f0efcaef6b4746b1b5eb56ec4c935bcfd20fc8a666c17",
            "size_bytes": 2_937,
        },
        "geometry_reconstruction_builder": {
            "path": "src/bctc_ai/evaluation/numeric_cell_crops.py",
            "sha256": "fe2aa1d33d422888f2efa0399503743f978b1af1605e4f99ac2a872ba249cb29",
            "size_bytes": 19_014,
        },
    }
    if any(implementations[name] != expected for name, expected in fixed_implementations.items()):
        raise _fail("pinned E-0041 transitive implementation identity drifted")
    expected_new_paths = {
        "formal_export_integration": "src/bctc_ai/evaluation/e0041_formal_export.py",
        "capture_export_script": "scripts/experiments/capture_e0041_mbb_cdkt_formal_export.py",
        "capture_seal_script": ("scripts/experiments/capture_e0041_mbb_cdkt_formal_export_seal.py"),
    }
    if any(implementations[name].get("path") != path for name, path in expected_new_paths.items()):
        raise _fail("E-0041 new mechanism implementation path drifted")
    runtime = _mapping(control.get("runtime_authority"), "formal runtime authority")
    if set(runtime) != {"artifacts", "versions"} or runtime.get("versions") != _RUNTIME_VERSIONS:
        raise _fail("E-0041 formal runtime authority drifted")
    runtime_artifacts = _mapping(runtime.get("artifacts"), "formal runtime artifacts")
    if set(runtime_artifacts) != {"project_metadata", "dependency_lock"}:
        raise _fail("E-0041 formal runtime artifact keyset drifted")
    for name, record in runtime_artifacts.items():
        if set(_mapping(record, f"formal runtime {name}")) != {
            "path",
            "sha256",
            "size_bytes",
        }:
            raise _fail(f"formal runtime {name} nested keyset drifted")
        _artifact_identity(record, f"formal runtime {name}")
    if runtime_artifacts != {
        "project_metadata": {
            "path": "pyproject.toml",
            "sha256": "59ad1366246975921df36b6d1f9bf2a91680a7b2c02d3777e772b0c3aa4ba045",
            "size_bytes": 1_094,
        },
        "dependency_lock": {
            "path": "uv.lock",
            "sha256": "3edf1c8471e283e74b721140b417456afc713b7d859d33bb600939eb18b79a9a",
            "size_bytes": 86_842,
        },
    }:
        raise _fail("E-0041 pinned runtime artifact identity drifted")
    oracle = _mapping(control.get("formal_oracle"), "formal oracle")
    expected_oracle = {
        "source_row_count": 64,
        "physical_cell_count": 128,
        "schema_row_count": 77,
        "final_selected_mapping_count": 61,
        "source_only_mapping_count": 3,
        "physical_cell_status_counts": {
            "VALUE": 111,
            "DASH": 5,
            "BLANK": 5,
            "UNRESOLVED": 7,
        },
        "selected_target_cell_count": 122,
        "exported_numeric_cell_count": 111,
        "strict_validation_finding_counts": {"PASS": 30, "NOT_TESTABLE": 6, "FAIL": 0},
        "geometry_normalized_registry_sha256": _NORMALIZED_REGISTRY_SHA256,
        "geometry_normalized_registry_size_bytes": _NORMALIZED_REGISTRY_SIZE,
        "geometry_asset_count": _ASSET_COUNT,
        "geometry_asset_total_size_bytes": _ASSET_TOTAL_SIZE,
        "geometry_asset_inventory_sha256": _ASSET_INVENTORY_SHA256,
        "geometry_asset_inventory_size_bytes": _ASSET_INVENTORY_SIZE,
        "projection_sha256": _PROJECTION_SHA256,
        "projection_size_bytes": _PROJECTION_SIZE,
        "workbook_sha256": _WORKBOOK_SHA256,
        "workbook_size_bytes": _WORKBOOK_SIZE,
        "template_identity_value_style_sha256": _TEMPLATE_FIDELITY_SHA256,
        "preserved_template_range": "A1:C78",
        "formula_count": 0,
        "post_mapping_core_sha256": CORE_SHA256,
    }
    if oracle != expected_oracle:
        raise _fail("E-0041 formal oracle drifted")
    validation = _mapping(control.get("accounting_validation"), "accounting validation")
    strict = _mapping(
        validation.get("strict_physical_visible_row_equations"),
        "strict physical equations",
    )
    equations = strict.get("equations")
    if (
        validation.get("mode") != "DIAGNOSTIC_ONLY_NO_REPAIR"
        or validation.get("strict_operand_statuses") != ["VALUE", "ZERO"]
        or validation.get("dash_as_zero_allowed") is not False
        or validation.get("blank_as_zero_allowed") is not False
        or validation.get("failed_check_may_change_mapping_status_or_value") is not False
        or strict.get("authority") != "NEWLY_FROZEN_CALIBRATION_MECHANISM_ASSERTION"
        or strict.get("mapping_independent_physical_row_ids") is not True
        or strict.get("expected_family_count") != 18
        or strict.get("expected_finding_count") != 36
        or not isinstance(equations, list)
        or len(equations) != 18
        or _sha256(_compact_bytes(equations)) != _PHYSICAL_EQUATIONS_SHA256
        or validation.get("secondary_schema_hierarchy")
        != {"enabled": False, "separate_denominator": True}
    ):
        raise _fail("E-0041 formal accounting-validation contract drifted")
    publication = control.get("publication")
    if publication != {
        "canonical_paths_only": True,
        "exact_output_inventory": True,
        "exclusive_no_overwrite": True,
        "hardened_core_pair_publication_primitives_with_formal_exact_empty_gate": True,
        "atomic_pair_publication": False,
        "clean_git_and_head_blob_recheck_immediately_before_publication": True,
        "no_validation_after_publisher_returns": True,
        "deterministic_double_build_byte_equality": True,
        "seal_replays_pair_at_same_clean_commit": True,
        "seal_publication_is_final_action": True,
    }:
        raise _fail("E-0041 formal publication contract drifted")
    if control.get("forbidden_inputs") != [
        "SUPERSEDED_MAPPING_OR_REVIEW_ARTIFACTS",
        "REVIEW_OR_STEWARD_ANSWERS",
        "HISTORY_OR_MONGODB_VALUES",
        "QWEN_RAW_REJECTED_OUTPUT_OR_TOKEN_STREAM",
        "HOLDOUT_FEATURES_LABELS_OR_RESULTS",
    ]:
        raise _fail("E-0041 forbidden-input contract drifted")
    limitations = control.get("limitations")
    if limitations != [
        "CALIBRATION_ONLY_SINGLE_BANK_SINGLE_PERIOD_DOCUMENT",
        "MACHINE_MAPPING_IS_NOT_SCHEMA_MAPPING_ACCURACY_OR_REVIEW_AUTHORITY",
        "ARITHMETIC_IS_DIAGNOSTIC_ONLY_AND_CANNOT_REPAIR_MAPPING_STATUS_OR_VALUE",
        "RECONSTRUCTED_GEOMETRY_IS_BOUNDED_TO_THE_PINNED_128_CELL_REGISTRY",
        "SHARED_S3_REGISTRY_IS_DURABILITY_BASELINE_ONLY",
        "NO_HUMAN_GOLD_OR_BANK_OR_PERIOD_DISJOINT_VALIDATION",
        "NOT_HOLDOUT_OR_PRODUCTION_APPROVED",
        "PAIR_PUBLICATION_NOT_CRASH_ATOMIC_PARTIAL_PAIR_REQUIRES_MANUAL_DISPOSITION",
    ]:
        raise _fail("E-0041 formal limitations drifted")
    for ledger in (inputs, implementations, runtime_artifacts):
        for raw in ledger.values():
            path = _artifact_identity(raw, "formal path scan")["path"]
            if any(token in cast(str, path).casefold() for token in _FORBIDDEN_MODULE_TOKENS):
                raise _fail("E-0041 formal ledger contains a forbidden path")


def _load_prerequisites(
    project_root: Path,
    config_path: Path,
    *,
    reader: StableReader,
    git_snapshot: _GitSnapshot | None,
) -> _Prerequisites:
    if config_path != CONTROL_RELATIVE_PATH or config_path.is_absolute():
        raise _fail(f"formal control must use canonical path {CONTROL_RELATIVE_PATH}")
    control_path = project_root / CONTROL_RELATIVE_PATH
    try:
        before = os.stat(control_path, follow_symlinks=False)
    except OSError as exc:
        raise _fail("cannot stat E-0041 formal control", exc) from exc
    if not stat.S_ISREG(before.st_mode):
        raise _fail("E-0041 formal control is not regular")
    control_payload = reader(project_root, control_path, _MAX_YAML_BYTES, "E-0041 formal control")
    after = os.stat(control_path, follow_symlinks=False)
    if _stat_identity(before) != _stat_identity(after) or len(control_payload) != after.st_size:
        raise _fail("E-0041 formal control identity changed during stable read")
    control = _decode_control(control_payload)
    _validate_control(control)
    control_file = _StableFile(
        name="E-0041 formal control",
        path=control_path,
        payload=control_payload,
        identity=_stat_identity(after),
        artifact=_record_for_payload(CONTROL_RELATIVE_PATH, control_payload),
        require_head_blob=True,
    )
    if git_snapshot is not None:
        _assert_head_blob_binding(project_root, control_file)
    implementations: dict[str, _StableFile] = {}
    for name, raw in _mapping(control["implementation"], "implementation ledger").items():
        stable = _read_file(
            project_root,
            _mapping(raw, f"implementation {name}"),
            f"implementation {name}",
            reader=reader,
            maximum_size=4 * 1024 * 1024,
            require_head_blob=True,
        )
        if git_snapshot is not None:
            _assert_head_blob_binding(project_root, stable)
        implementations[name] = stable
    runtime_files: dict[str, _StableFile] = {}
    runtime_records = _mapping(
        _mapping(control["runtime_authority"], "runtime authority")["artifacts"],
        "runtime artifacts",
    )
    for name, raw in runtime_records.items():
        stable = _read_file(
            project_root,
            _mapping(raw, f"runtime {name}"),
            f"runtime {name}",
            reader=reader,
            maximum_size=4 * 1024 * 1024,
            require_head_blob=True,
        )
        if git_snapshot is not None:
            _assert_head_blob_binding(project_root, stable)
        runtime_files[name] = stable
    _assert_runtime_versions()
    return _Prerequisites(
        control=control,
        control_file=control_file,
        implementations=implementations,
        runtime_files=runtime_files,
    )


def _validate_seal_before_later_opens(payload: Mapping[str, Any]) -> None:
    inventory = _mapping(payload.get("inventory"), "E-0040 seal inventory")
    files = _sequence(inventory.get("files"), "E-0040 seal inventory files")
    if (
        payload.get("format_version") != 1
        or payload.get("experiment_id") != "E-0040"
        or payload.get("dataset_role") != "CALIBRATION"
        or payload.get("state") != "E0040_GENERIC_CHALLENGER_MAPPING_HASH_SEALED"
        or payload.get("seal_git_dirty") is not False
        or payload.get("mapping_status") != "CHALLENGER_COMPLETE_61_SELECTED_3_SOURCE_ONLY"
        or inventory.get("file_count") != 1
        or len(files) != 1
        or _artifact_identity(files[0], "E-0040 sealed mapping")
        != {
            "path": "output/calibration/e0040-mbb-cdkt-formal-mapping/mapping_only.json",
            "sha256": "8def983007fc3aacf59351395426d5246ad3f28d605442f590de55eaf396cb0d",
            "size_bytes": 1_157_172,
        }
    ):
        raise _fail("E-0040 mapping seal authority drifted before registration open")
    metrics = _mapping(payload.get("metrics"), "E-0040 seal metrics")
    if (
        metrics.get("source_row_count") != 64
        or metrics.get("schema_node_count") != 77
        or metrics.get("final_selected_count") != 61
        or metrics.get("source_only_structural_count") != 3
        or metrics.get("selected_anchor_count") != 43
        or metrics.get("selected_path_count") != 18
        or metrics.get("all_intervals_exhaustive") is not True
        or metrics.get("all_pruning_counts_zero") is not True
    ):
        raise _fail("E-0040 seal metric authority drifted")


def _validate_registration_before_mapping_open(
    payload: Mapping[str, Any],
    *,
    expected_shared_registry: Mapping[str, Any],
) -> None:
    local = _mapping(payload.get("local_artifacts"), "E-0040 registration local artifacts")
    shared = _mapping(payload.get("shared_registry"), "E-0040 shared-registry receipt")
    remote = _mapping(payload.get("remote_verification"), "E-0040 remote verification")
    if (
        payload.get("format_version") != 1
        or payload.get("experiment_id") != "E-0040"
        or payload.get("dataset_role") != "CALIBRATION"
        or payload.get("state") != "E0040_FORMAL_MAPPING_IMMUTABLY_REGISTERED_IN_S3_POST_SEAL"
        or payload.get("policy") != "IMMUTABLE_POST_SEAL_S3_REGISTRATION_V1"
        or remote.get("status") != "PASS"
        or set(local) != {"mapping_only", "mapping_seal"}
        or _artifact_identity(local["mapping_seal"], "registered E-0040 seal")
        != {
            "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-seal.json",
            "sha256": "68306f7f540faa77d6e2e383927eae23fc3724cfdc8c53cded978a86f3a00b29",
            "size_bytes": 7_611,
        }
        or _artifact_identity(local["mapping_only"], "registered E-0040 mapping")
        != {
            "path": "output/calibration/e0040-mbb-cdkt-formal-mapping/mapping_only.json",
            "sha256": "8def983007fc3aacf59351395426d5246ad3f28d605442f590de55eaf396cb0d",
            "size_bytes": 1_157_172,
        }
        or shared
        != {
            "appended": False,
            "not_registered_there": True,
            "path": expected_shared_registry["path"],
            "sha256": expected_shared_registry["sha256"],
            "size_bytes": expected_shared_registry["size_bytes"],
            "unchanged": True,
        }
    ):
        raise _fail("E-0040 S3 registration authority drifted before mapping open")


def _normalize_project_path(value: object, name: str) -> str:
    if not isinstance(value, str) or not value or "\\" in value:
        raise _fail(f"{name} is not a nonempty POSIX path")
    raw = PurePosixPath(value)
    if raw.is_absolute():
        try:
            raw = raw.relative_to(LEGACY_GEOMETRY_ROOT)
        except ValueError as exc:
            raise _fail(f"{name} has a foreign absolute checkout root", exc) from exc
    if not raw.parts or raw == PurePosixPath(".") or ".." in raw.parts:
        raise _fail(f"{name} is unsafe")
    return raw.as_posix()


def _geometry_asset_declarations(
    registry: Mapping[str, Any],
    registry_relative_path: Path,
) -> dict[str, dict[str, Any]]:
    declarations: dict[str, dict[str, Any]] = {}

    def add(
        raw_path: object,
        raw_digest: object,
        role: str,
        *,
        relative_to_registry: bool = False,
        expected_size: object | None = None,
    ) -> None:
        if not isinstance(raw_digest, str) or _SHA256.fullmatch(raw_digest) is None:
            raise _fail(f"geometry asset {role} digest is invalid")
        if relative_to_registry:
            if not isinstance(raw_path, str):
                raise _fail(f"geometry asset {role} path is invalid")
            suffix = PurePosixPath(raw_path)
            if suffix.is_absolute() or not suffix.parts or ".." in suffix.parts:
                raise _fail(f"geometry asset {role} relative path is unsafe")
            path = (registry_relative_path.parent / Path(*suffix.parts)).as_posix()
        else:
            path = _normalize_project_path(raw_path, f"geometry asset {role}")
        if expected_size is not None and (type(expected_size) is not int or expected_size <= 0):
            raise _fail(f"geometry asset {role} size is invalid")
        existing = declarations.get(path)
        if existing is None:
            declarations[path] = {
                "path": path,
                "sha256": raw_digest,
                "expected_size": expected_size,
                "roles": {role},
            }
            return
        if existing["sha256"] != raw_digest:
            raise _fail(f"geometry asset path has conflicting digests: {path}")
        prior_size = existing["expected_size"]
        if expected_size is not None and prior_size not in {None, expected_size}:
            raise _fail(f"geometry asset path has conflicting sizes: {path}")
        if prior_size is None:
            existing["expected_size"] = expected_size
        cast(set[str], existing["roles"]).add(role)

    crop_policy = _mapping(registry.get("crop_policy"), "geometry crop policy")
    row_contract = _mapping(registry.get("row_contract"), "geometry row contract")
    add(crop_policy.get("path"), crop_policy.get("sha256"), "crop_policy")
    add(row_contract.get("path"), row_contract.get("sha256"), "row_contract")
    pages = _sequence(registry.get("pages"), "geometry pages")
    if len(pages) != 2:
        raise _fail("geometry page count drifted")
    page_numbers: set[int] = set()
    for index, raw in enumerate(pages):
        page = _mapping(raw, f"geometry page {index}")
        page_number = page.get("page")
        if type(page_number) is not int or page_number in page_numbers:
            raise _fail("geometry page identity is invalid")
        page_numbers.add(page_number)
        add(page.get("ocr_path"), page.get("ocr_sha256"), f"page_{page_number}_ocr")
        add(page.get("render_path"), page.get("render_sha256"), f"page_{page_number}_render")
    if page_numbers != {3, 4}:
        raise _fail("geometry page set drifted")
    cells = _sequence(registry.get("cells"), "geometry cells")
    if len(cells) != 128:
        raise _fail("geometry cell count drifted")
    cell_ids: set[str] = set()
    for index, raw in enumerate(cells):
        cell = _mapping(raw, f"geometry cell {index}")
        cell_id = cell.get("cell_id")
        if not isinstance(cell_id, str) or not cell_id or cell_id in cell_ids:
            raise _fail("geometry cell identity is duplicate or invalid")
        cell_ids.add(cell_id)
        add(
            cell.get("crop_path"),
            cell.get("crop_sha256"),
            "cell_crop",
            relative_to_registry=True,
            expected_size=cell.get("crop_size_bytes"),
        )
        add(cell.get("source_ocr_path"), cell.get("source_ocr_sha256"), "cell_source_ocr")
        add(
            cell.get("source_render_path"),
            cell.get("source_render_sha256"),
            "cell_source_render",
        )
    if len(declarations) != _ASSET_COUNT:
        raise _fail("geometry unique asset denominator drifted")
    return declarations


def _load_geometry_assets(
    project_root: Path,
    registry: Mapping[str, Any],
    registry_relative_path: Path,
    *,
    reader: StableReader,
    git_snapshot: _GitSnapshot | None,
    control_inputs: Mapping[str, Any],
) -> tuple[tuple[_StableFile, ...], tuple[dict[str, Any], ...]]:
    declarations = _geometry_asset_declarations(registry, registry_relative_path)
    asset_files: list[_StableFile] = []
    inventory: list[dict[str, Any]] = []
    observed_sizes: dict[str, int] = {}
    for path in sorted(declarations):
        absolute = project_root / path
        try:
            observed = os.stat(absolute, follow_symlinks=False)
        except OSError as exc:
            raise _fail(f"cannot preflight geometry asset {path}", exc) from exc
        if not stat.S_ISREG(observed.st_mode) or observed.st_size <= 0:
            raise _fail(f"geometry asset is not a nonempty regular file: {path}")
        expected_size = declarations[path]["expected_size"]
        if expected_size is not None and observed.st_size != expected_size:
            raise _fail(f"geometry crop preflight size drifted: {path}")
        if observed.st_size > 16 * 1024 * 1024:
            raise _fail(f"geometry asset exceeds its per-file size budget: {path}")
        observed_sizes[path] = observed.st_size
    if sum(observed_sizes.values()) != _ASSET_TOTAL_SIZE:
        raise _fail("geometry asset aggregate size drifted before asset reads")
    for path in sorted(declarations):
        declaration = declarations[path]
        expected_size = declaration["expected_size"]
        record = {
            "path": path,
            "sha256": declaration["sha256"],
            "size_bytes": expected_size if expected_size is not None else 1,
        }
        # Source, policy, and row-contract sizes are authenticated from the
        # descriptor read itself; crop sizes are also pinned in the registry.
        observed_size = observed_sizes[path]
        record["size_bytes"] = observed_size
        require_head = path in {
            cast(
                str,
                _artifact_identity(control_inputs["geometry_crop_policy"], "crop policy")["path"],
            ),
            cast(
                str,
                _artifact_identity(control_inputs["geometry_row_contract"], "row contract")["path"],
            ),
        }
        stable = _read_file(
            project_root,
            record,
            f"geometry asset {path}",
            reader=reader,
            maximum_size=observed_size,
            require_head_blob=require_head,
        )
        if expected_size is not None and len(stable.payload) != expected_size:
            raise _fail(f"geometry crop size drifted: {path}")
        if git_snapshot is not None:
            _assert_head_blob_binding(project_root, stable)
        asset_files.append(stable)
        inventory.append(
            {
                **stable.artifact,
                "roles": sorted(cast(set[str], declaration["roles"])),
            }
        )
    inventory_bytes = _compact_bytes(inventory)
    if (
        len(inventory) != _ASSET_COUNT
        or sum(cast(int, item["size_bytes"]) for item in inventory) != _ASSET_TOTAL_SIZE
        or len(inventory_bytes) != _ASSET_INVENTORY_SIZE
        or _sha256(inventory_bytes) != _ASSET_INVENTORY_SHA256
    ):
        raise _fail("exact geometry asset inventory oracle drifted")
    policy_identity = _artifact_identity(control_inputs["geometry_crop_policy"], "crop policy")
    row_identity = _artifact_identity(control_inputs["geometry_row_contract"], "row contract")
    by_path = {item["path"]: item for item in inventory}
    if by_path.get(policy_identity["path"], {}) | {} != {
        **policy_identity,
        "roles": ["crop_policy"],
    } or by_path.get(row_identity["path"], {}) | {} != {
        **row_identity,
        "roles": ["row_contract"],
    }:
        raise _fail("geometry policy/row-contract ledger linkage drifted")
    return tuple(asset_files), tuple(inventory)


def _load_materials(
    project_root: Path,
    config_path: Path,
    *,
    reader: StableReader,
    git_snapshot: _GitSnapshot | None,
) -> _Materials:
    prerequisites = _load_prerequisites(
        project_root,
        config_path,
        reader=reader,
        git_snapshot=git_snapshot,
    )
    inputs = _mapping(prerequisites.control["input_authority"], "formal inputs")
    ordered: list[_StableFile] = [
        prerequisites.control_file,
        *prerequisites.implementations.values(),
        *prerequisites.runtime_files.values(),
    ]

    seal_file = _read_file(
        project_root,
        inputs["e0040_mapping_seal"],
        "E-0040 mapping seal",
        reader=reader,
        maximum_size=_MAX_JSON_BYTES,
        require_head_blob=True,
    )
    if git_snapshot is not None:
        _assert_head_blob_binding(project_root, seal_file)
    _validate_seal_before_later_opens(_decode_json(seal_file.payload, seal_file.name))
    ordered.append(seal_file)

    registration_file = _read_file(
        project_root,
        inputs["e0040_s3_registration"],
        "E-0040 S3 registration",
        reader=reader,
        maximum_size=_MAX_JSON_BYTES,
        require_head_blob=True,
    )
    if git_snapshot is not None:
        _assert_head_blob_binding(project_root, registration_file)
    registration = _decode_json(registration_file.payload, registration_file.name)
    shared_identity = _artifact_identity(
        inputs["shared_s3_registry_frozen_baseline"],
        "shared S3 registry",
    )
    _validate_registration_before_mapping_open(
        registration,
        expected_shared_registry=shared_identity,
    )
    ordered.append(registration_file)

    shared_registry_file = _read_file(
        project_root,
        inputs["shared_s3_registry_frozen_baseline"],
        "shared S3 registry frozen baseline",
        reader=reader,
        maximum_size=1024 * 1024,
        require_head_blob=True,
    )
    if git_snapshot is not None:
        _assert_head_blob_binding(project_root, shared_registry_file)
    ordered.append(shared_registry_file)

    mapping_file = _read_file(
        project_root,
        inputs["e0040_mapping"],
        "E-0040 formal mapping",
        reader=reader,
        maximum_size=_MAX_JSON_BYTES,
        require_head_blob=False,
    )
    ordered.append(mapping_file)
    # Minting the carrier is deliberately after the mapping file's first open.
    _core.authenticate_e0040_result_carrier(
        mapping_bytes=mapping_file.payload,
        seal_bytes=seal_file.payload,
        registration_bytes=registration_file.payload,
    )

    postjoin_file = _read_file(
        project_root,
        inputs["e0037_postjoin"],
        "E-0037 postjoin",
        reader=reader,
        maximum_size=_MAX_JSON_BYTES,
        require_head_blob=True,
    )
    if git_snapshot is not None:
        _assert_head_blob_binding(project_root, postjoin_file)
    ordered.append(postjoin_file)

    geometry_file = _read_file(
        project_root,
        inputs["reconstructed_geometry_registry"],
        "E-0041 reconstructed geometry registry",
        reader=reader,
        maximum_size=_MAX_JSON_BYTES,
        require_head_blob=False,
    )
    geometry = _decode_json(geometry_file.payload, geometry_file.name)
    normalized = _core._normalize_authenticated_geometry_registry(geometry)
    normalized_bytes = _compact_bytes(normalized)
    if (
        len(normalized_bytes) != _NORMALIZED_REGISTRY_SIZE
        or _sha256(normalized_bytes) != _NORMALIZED_REGISTRY_SHA256
    ):
        raise _fail("normalized geometry registry identity drifted")
    ordered.append(geometry_file)
    geometry_relative = Path(cast(str, geometry_file.artifact["path"]))
    asset_files, asset_inventory = _load_geometry_assets(
        project_root,
        geometry,
        geometry_relative,
        reader=reader,
        git_snapshot=git_snapshot,
        control_inputs=inputs,
    )
    ordered.extend(asset_files)

    template_file = _read_file(
        project_root,
        inputs["cdkt_template"],
        "CDKT template",
        reader=reader,
        maximum_size=16 * 1024 * 1024,
        require_head_blob=True,
    )
    if git_snapshot is not None:
        _assert_head_blob_binding(project_root, template_file)
    ordered.append(template_file)
    return _Materials(
        prerequisites=prerequisites,
        ordered_files=tuple(ordered),
        seal_file=seal_file,
        registration_file=registration_file,
        shared_registry_file=shared_registry_file,
        mapping_file=mapping_file,
        postjoin_file=postjoin_file,
        geometry_file=geometry_file,
        template_file=template_file,
        asset_files=asset_files,
        asset_inventory=asset_inventory,
        normalized_registry_artifact={
            "sha256": _NORMALIZED_REGISTRY_SHA256,
            "size_bytes": _NORMALIZED_REGISTRY_SIZE,
        },
    )


@contextmanager
def _authenticated_core_asset_cache(
    project_root: Path,
    asset_files: Sequence[_StableFile],
) -> Any:
    payloads = {stable.path.relative_to(project_root): stable.payload for stable in asset_files}
    requested: list[Path] = []
    original = _core._stable_read

    def cached_reader(
        supplied_root: Path,
        path: Path,
        *,
        maximum_size: int,
        name: str,
    ) -> bytes:
        if supplied_root != project_root or not path.is_absolute():
            raise _fail(f"core requested an untrusted cached path for {name}")
        try:
            relative = path.relative_to(project_root)
        except ValueError as exc:
            raise _fail(f"core cached path escapes project root for {name}", exc) from exc
        payload = payloads.get(relative)
        if payload is None:
            raise _fail(f"core requested an unauthenticated geometry asset: {relative}")
        if len(payload) > maximum_size:
            raise _fail(f"core geometry cache size budget failed: {relative}")
        requested.append(relative)
        return payload

    if _core._stable_read is not original:
        raise _fail("E-0041 core stable reader was replaced before formal invocation")
    _core._stable_read = cached_reader
    try:
        yield requested
    finally:
        if _core._stable_read is not cached_reader:
            _core._stable_read = original
            raise _fail("E-0041 core stable reader changed during formal invocation")
        _core._stable_read = original


def _validate_projection(projection: Mapping[str, Any], projection_bytes: bytes) -> None:
    if (
        projection.get("format_version") != 1
        or projection.get("experiment_id") != "E-0041"
        or projection.get("dataset_role") != "CALIBRATION"
        or projection.get("state") != "POST_MAPPING_DEVELOPMENT_EXCEL_ASSEMBLED"
        or len(projection_bytes) != _PROJECTION_SIZE
        or _sha256(projection_bytes) != _PROJECTION_SHA256
    ):
        raise _fail("formal E-0041 projection identity drifted")
    metrics = _mapping(projection.get("metrics"), "projection metrics")
    if metrics != {
        "source_row_count": 64,
        "physical_cell_count": 128,
        "schema_row_count": 77,
        "physical_cell_status_counts": {
            "BLANK": 5,
            "DASH": 5,
            "UNRESOLVED": 7,
            "VALUE": 111,
        },
        "schema_period_status_counts": {
            "AMBIGUOUS": 6,
            "BLANK": 5,
            "DASH": 5,
            "UNRESOLVED": 27,
            "VALUE": 111,
        },
        "selected_target_cell_count": 122,
        "exported_numeric_cell_count": 111,
    }:
        raise _fail("formal E-0041 projection metrics drifted")
    mapping_authority = _mapping(projection.get("mapping_authority"), "mapping authority")
    if (
        mapping_authority.get("source") != "E0040_GENERIC_CALIBRATION_CHALLENGER"
        or mapping_authority.get("final_selected_pair_count") != 61
        or mapping_authority.get("source_only_row_count") != 3
        or mapping_authority.get("unselected_row_count") != 0
        or mapping_authority.get("unapproved_changed_alias_report_norm_ids") != []
        or mapping_authority.get("id_scoped_alias_report_norm_ids") != []
    ):
        raise _fail("formal E-0041 mapping authority metrics drifted")
    validation = _mapping(projection.get("validation"), "projection validation")
    strict = _mapping(
        validation.get("strict_physical_visible_row_equations"),
        "projection strict validation",
    )
    findings = _sequence(strict.get("findings"), "strict validation findings")
    observed_counts = Counter(
        cast(str, _mapping(item, "strict validation finding").get("result")) for item in findings
    )
    if (
        strict.get("family_count") != 18
        or strict.get("finding_count") != 36
        or len(findings) != 36
        or observed_counts != Counter({"PASS": 30, "NOT_TESTABLE": 6})
        or observed_counts.get("FAIL", 0) != 0
        or validation.get("pre_validation_value_status_sha256")
        != validation.get("post_validation_value_status_sha256")
    ):
        raise _fail("formal E-0041 diagnostic oracle drifted")
    if (
        len(_sequence(projection.get("physical_cells"), "physical cells")) != 128
        or len(_sequence(projection.get("schema_rows"), "schema rows")) != 77
    ):
        raise _fail("formal E-0041 projection denominator drifted")


def _validate_workbook(
    workbook_bytes: bytes,
    receipt: Mapping[str, Any],
    *,
    expected_template_snapshot: Mapping[str, Any],
) -> None:
    if len(workbook_bytes) != _WORKBOOK_SIZE or _sha256(workbook_bytes) != _WORKBOOK_SHA256:
        raise _fail("formal E-0041 workbook byte identity drifted")
    if receipt != {
        "template_identity_value_style_sha256": _TEMPLATE_FIDELITY_SHA256,
        "reopened_template_identity_value_style_sha256": _TEMPLATE_FIDELITY_SHA256,
        "exact_template_identity_value_style_fidelity": True,
        "source_sheet": "Sheet1",
        "preserved_range": "A1:C78",
        "sheet_names": ["Sheet1", "PROVENANCE", "VALIDATION_DIAGNOSTICS", "RUN_METADATA"],
        "workbook_sha256": _WORKBOOK_SHA256,
        "workbook_size_bytes": _WORKBOOK_SIZE,
        "deterministic_core_properties_sha256": (
            "a025959e8b178cfc6c6aae8f2d49d86fa305d3e36e165c8cbbc16923068668e4"
        ),
    }:
        raise _fail("formal E-0041 workbook receipt drifted")
    if _sha256(_compact_bytes(expected_template_snapshot)) != _TEMPLATE_FIDELITY_SHA256:
        raise _fail("formal E-0041 template A:C snapshot drifted")
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    except Exception as exc:
        raise _fail("cannot reopen formal E-0041 workbook", exc) from exc
    try:
        formula_count = sum(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        snapshot = _core._template_snapshot(workbook, "Sheet1", 78)
        if (
            formula_count != 0
            or workbook.sheetnames
            != ["Sheet1", "PROVENANCE", "VALIDATION_DIAGNOSTICS", "RUN_METADATA"]
            or snapshot != expected_template_snapshot
        ):
            raise _fail("formal E-0041 workbook formula/A:C fidelity gate failed")
    finally:
        workbook.close()


def _provenance_payload(
    materials: _Materials,
    *,
    capture_git_commit: str,
    projection_bytes: bytes,
    workbook_bytes: bytes,
    workbook_receipt: Mapping[str, Any],
) -> dict[str, Any]:
    prerequisites = materials.prerequisites
    metrics = {
        "source_row_count": 64,
        "physical_cell_count": 128,
        "schema_row_count": 77,
        "final_selected_mapping_count": 61,
        "source_only_mapping_count": 3,
        "physical_cell_status_counts": {
            "BLANK": 5,
            "DASH": 5,
            "UNRESOLVED": 7,
            "VALUE": 111,
        },
        "selected_target_cell_count": 122,
        "exported_numeric_cell_count": 111,
        "strict_validation_finding_counts": {"FAIL": 0, "NOT_TESTABLE": 6, "PASS": 30},
    }
    input_ledger = {
        "control": prerequisites.control_file.artifact,
        "e0040_mapping_seal": materials.seal_file.artifact,
        "e0040_s3_registration": materials.registration_file.artifact,
        "shared_s3_registry_frozen_baseline": materials.shared_registry_file.artifact,
        "e0040_mapping": materials.mapping_file.artifact,
        "e0037_postjoin": materials.postjoin_file.artifact,
        "reconstructed_geometry_registry": materials.geometry_file.artifact,
        "geometry_crop_policy": _artifact_identity(
            prerequisites.control["input_authority"]["geometry_crop_policy"],
            "geometry crop policy",
        ),
        "geometry_row_contract": _artifact_identity(
            prerequisites.control["input_authority"]["geometry_row_contract"],
            "geometry row contract",
        ),
        "cdkt_template": materials.template_file.artifact,
    }
    payload: dict[str, Any] = {
        "format_version": 1,
        "experiment_id": "E-0041",
        "dataset_role": "CALIBRATION",
        "state": PAIR_STATE,
        "capture_git_commit": capture_git_commit,
        "capture_git_dirty": False,
        "pair_hash_sealed": False,
        "outputs": {
            "workbook": _record_for_payload(WORKBOOK_RELATIVE_PATH, workbook_bytes),
            "provenance_path": PROVENANCE_RELATIVE_PATH.as_posix(),
            "publication_order": [WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name],
            "provenance_is_completion_marker": True,
            "atomic_pair_publication": False,
        },
        "projection_receipt": {
            "sha256": _sha256(projection_bytes),
            "size_bytes": len(projection_bytes),
        },
        "workbook_receipt": dict(workbook_receipt),
        "input_hash_ledger": input_ledger,
        "implementation_hash_ledger": {
            name: stable.artifact for name, stable in prerequisites.implementations.items()
        },
        "runtime_hash_ledger": {
            name: stable.artifact for name, stable in prerequisites.runtime_files.items()
        },
        "runtime_versions": dict(_RUNTIME_VERSIONS),
        "geometry_authority": {
            "normalized_registry": materials.normalized_registry_artifact,
            "asset_inventory": {
                "file_count": len(materials.asset_inventory),
                "total_size_bytes": sum(
                    cast(int, item["size_bytes"]) for item in materials.asset_inventory
                ),
                "canonical_sha256": _sha256(_compact_bytes(materials.asset_inventory)),
                "canonical_size_bytes": len(_compact_bytes(materials.asset_inventory)),
                "files": list(materials.asset_inventory),
            },
            "diagnostic_source_image_paths_opened_or_emitted": False,
            "legacy_checkout_root_removed_from_authoritative_paths": True,
        },
        "metrics": metrics,
        "access_contract": {
            "validation_order": list(_VALIDATION_ORDER),
            "e0040_seal_validated_before_registration_open": True,
            "e0040_registration_and_shared_registry_validated_before_mapping_open": True,
            "authenticated_mapping_carrier_minted_before_postjoin_open": True,
            "geometry_registry_and_exact_assets_validated_before_template_open": True,
            "two_builds_use_separate_decoded_in_memory_snapshots": True,
            "projection_workbook_and_provenance_double_build_byte_equal": True,
            "mapping_challenger_replayed_or_repaired": False,
            "superseded_mapping_or_review_artifact_opened": False,
            "review_or_steward_answers_opened": False,
            "history_or_mongodb_opened": False,
            "qwen_raw_or_token_output_opened": False,
            "holdout_artifact_opened": False,
            "forbidden_module_contamination_guard_passed": True,
        },
        "authority": {
            "exact_input_and_workbook_hash_identity": True,
            "canonical_unsealed_provenance_bytes": True,
            "provenance_self_hash_or_pair_seal": False,
            "authenticated_e0040_machine_mapping": True,
            "period_unit_scope_and_nested_numeric_evidence": True,
            "exact_reconstructed_geometry": True,
            "template_identity_fidelity": True,
            "schema_authority": False,
            "mapping_accuracy": False,
            "review_or_steward_approval": False,
            "accounting_correctness": False,
            "human_gold": False,
            "holdout_or_production": False,
        },
        "limitations": list(cast(list[str], prerequisites.control["limitations"])),
        "claim_boundary": _CLAIM_BOUNDARY,
    }
    _assert_no_absolute_paths(payload, "formal provenance")
    if "provenance_sha256" in _encoded_json(payload).decode("utf-8"):
        raise _fail("formal provenance must not contain its own hash")
    return payload


def _validate_new_provenance(
    payload: Mapping[str, Any],
    encoded: bytes,
    *,
    materials: _Materials,
    commit: str,
    projection_bytes: bytes,
    workbook_bytes: bytes,
    workbook_receipt: Mapping[str, Any],
) -> None:
    workbook_record = _validate_existing_provenance(payload, encoded, expected_commit=commit)
    if workbook_record != _record_for_payload(WORKBOOK_RELATIVE_PATH, workbook_bytes):
        raise _fail("formal provenance/workbook cross-link drifted")
    expected_inputs = {
        "control": materials.prerequisites.control_file.artifact,
        "e0040_mapping_seal": materials.seal_file.artifact,
        "e0040_s3_registration": materials.registration_file.artifact,
        "shared_s3_registry_frozen_baseline": materials.shared_registry_file.artifact,
        "e0040_mapping": materials.mapping_file.artifact,
        "e0037_postjoin": materials.postjoin_file.artifact,
        "reconstructed_geometry_registry": materials.geometry_file.artifact,
        "geometry_crop_policy": _artifact_identity(
            materials.prerequisites.control["input_authority"]["geometry_crop_policy"],
            "geometry crop policy",
        ),
        "geometry_row_contract": _artifact_identity(
            materials.prerequisites.control["input_authority"]["geometry_row_contract"],
            "geometry row contract",
        ),
        "cdkt_template": materials.template_file.artifact,
    }
    expected_implementations = {
        name: stable.artifact for name, stable in materials.prerequisites.implementations.items()
    }
    expected_runtime = {
        name: stable.artifact for name, stable in materials.prerequisites.runtime_files.items()
    }
    geometry = _mapping(payload.get("geometry_authority"), "formal geometry authority")
    expected_inventory = {
        "file_count": _ASSET_COUNT,
        "total_size_bytes": _ASSET_TOTAL_SIZE,
        "canonical_sha256": _ASSET_INVENTORY_SHA256,
        "canonical_size_bytes": _ASSET_INVENTORY_SIZE,
        "files": list(materials.asset_inventory),
    }
    if (
        payload.get("projection_receipt")
        != {"sha256": _sha256(projection_bytes), "size_bytes": len(projection_bytes)}
        or payload.get("workbook_receipt") != workbook_receipt
        or payload.get("input_hash_ledger") != expected_inputs
        or payload.get("implementation_hash_ledger") != expected_implementations
        or payload.get("runtime_hash_ledger") != expected_runtime
        or payload.get("runtime_versions") != _RUNTIME_VERSIONS
        or geometry
        != {
            "normalized_registry": materials.normalized_registry_artifact,
            "asset_inventory": expected_inventory,
            "diagnostic_source_image_paths_opened_or_emitted": False,
            "legacy_checkout_root_removed_from_authoritative_paths": True,
        }
        or payload.get("metrics")
        != {
            "source_row_count": 64,
            "physical_cell_count": 128,
            "schema_row_count": 77,
            "final_selected_mapping_count": 61,
            "source_only_mapping_count": 3,
            "physical_cell_status_counts": {
                "BLANK": 5,
                "DASH": 5,
                "UNRESOLVED": 7,
                "VALUE": 111,
            },
            "selected_target_cell_count": 122,
            "exported_numeric_cell_count": 111,
            "strict_validation_finding_counts": {
                "FAIL": 0,
                "NOT_TESTABLE": 6,
                "PASS": 30,
            },
        }
        or payload.get("access_contract")
        != {
            "validation_order": list(_VALIDATION_ORDER),
            "e0040_seal_validated_before_registration_open": True,
            "e0040_registration_and_shared_registry_validated_before_mapping_open": True,
            "authenticated_mapping_carrier_minted_before_postjoin_open": True,
            "geometry_registry_and_exact_assets_validated_before_template_open": True,
            "two_builds_use_separate_decoded_in_memory_snapshots": True,
            "projection_workbook_and_provenance_double_build_byte_equal": True,
            "mapping_challenger_replayed_or_repaired": False,
            "superseded_mapping_or_review_artifact_opened": False,
            "review_or_steward_answers_opened": False,
            "history_or_mongodb_opened": False,
            "qwen_raw_or_token_output_opened": False,
            "holdout_artifact_opened": False,
            "forbidden_module_contamination_guard_passed": True,
        }
        or payload.get("authority")
        != {
            "exact_input_and_workbook_hash_identity": True,
            "canonical_unsealed_provenance_bytes": True,
            "provenance_self_hash_or_pair_seal": False,
            "authenticated_e0040_machine_mapping": True,
            "period_unit_scope_and_nested_numeric_evidence": True,
            "exact_reconstructed_geometry": True,
            "template_identity_fidelity": True,
            "schema_authority": False,
            "mapping_accuracy": False,
            "review_or_steward_approval": False,
            "accounting_correctness": False,
            "human_gold": False,
            "holdout_or_production": False,
        }
        or payload.get("limitations") != materials.prerequisites.control["limitations"]
    ):
        raise _fail("formal provenance full contract or cross-ledger validation failed")


def _build_once(
    project_root: Path,
    materials: _Materials,
    *,
    capture_git_commit: str,
) -> tuple[dict[str, Any], bytes, bytes, dict[str, Any], dict[str, Any], bytes]:
    # Each deterministic build starts from independently decoded immutable byte
    # snapshots.  Nothing produced by build one is passed to build two.
    postjoin = _decode_json(bytes(materials.postjoin_file.payload), "E-0037 postjoin snapshot")
    geometry = _decode_json(bytes(materials.geometry_file.payload), "geometry registry snapshot")
    carrier = _core.authenticate_e0040_result_carrier(
        mapping_bytes=bytes(materials.mapping_file.payload),
        seal_bytes=bytes(materials.seal_file.payload),
        registration_bytes=bytes(materials.registration_file.payload),
    )
    workbook, template_rows, template_snapshot = _core._load_template(
        bytes(materials.template_file.payload),
        sheet_name="Sheet1",
        schema_row_count=77,
    )
    workbook.close()
    input_records = {
        "e0037_postjoin": materials.postjoin_file.artifact,
        "mapping_challenger": materials.mapping_file.artifact,
        "geometry_registry": materials.geometry_file.artifact,
    }
    with _authenticated_core_asset_cache(project_root, materials.asset_files) as requested:
        projection = _core.assemble_post_mapping_projection(
            postjoin_payload=postjoin,
            mapping_payload=carrier,
            geometry_registry=geometry,
            geometry_registry_path=materials.geometry_file.path,
            template_rows=template_rows,
            project_root=project_root,
            input_records=input_records,
            validation_config=_decode_json(
                _compact_bytes(materials.prerequisites.control["accounting_validation"]),
                "accounting validation snapshot",
            ),
            expected_row_count=64,
            expected_cell_count=128,
        )
    expected_core_assets = {
        stable.path.relative_to(project_root)
        for stable in materials.asset_files
        if stable.artifact["path"]
        not in {
            materials.prerequisites.control["input_authority"]["geometry_crop_policy"]["path"],
            materials.prerequisites.control["input_authority"]["geometry_row_contract"]["path"],
        }
    }
    if set(requested) != expected_core_assets or len(requested) != 132:
        raise _fail("E-0041 core did not consume the exact authenticated geometry asset set")
    projection_bytes = _compact_bytes(projection)
    _validate_projection(projection, projection_bytes)
    workbook_bytes, workbook_receipt = _core.build_development_workbook(
        template_bytes=bytes(materials.template_file.payload),
        projection=projection,
        source_sheet="Sheet1",
        schema_row_count=77,
    )
    _validate_workbook(
        workbook_bytes,
        workbook_receipt,
        expected_template_snapshot=template_snapshot,
    )
    provenance = _provenance_payload(
        materials,
        capture_git_commit=capture_git_commit,
        projection_bytes=projection_bytes,
        workbook_bytes=workbook_bytes,
        workbook_receipt=workbook_receipt,
    )
    provenance_bytes = _encoded_json(provenance)
    if (
        _encoded_json(_decode_json(provenance_bytes, "formal provenance round trip"))
        != provenance_bytes
    ):
        raise _fail("formal provenance is not canonical JSON")
    _validate_new_provenance(
        provenance,
        provenance_bytes,
        materials=materials,
        commit=capture_git_commit,
        projection_bytes=projection_bytes,
        workbook_bytes=workbook_bytes,
        workbook_receipt=workbook_receipt,
    )
    return (
        projection,
        projection_bytes,
        workbook_bytes,
        workbook_receipt,
        provenance,
        provenance_bytes,
    )


def _build_e0041_formal_export_verified(
    project_root: Path,
    *,
    capture_git_commit: str,
    config_path: Path = CONTROL_RELATIVE_PATH,
    reader: StableReader,
    git_snapshot: _GitSnapshot | None,
) -> FormalExportBuild:
    """Internal build using a caller that already acquired the Git gate."""

    _assert_process_isolation()
    if type(capture_git_commit) is not str or _GIT_COMMIT.fullmatch(capture_git_commit) is None:
        raise _fail("formal E-0041 capture Git commit is invalid")
    root = project_root.resolve()
    if git_snapshot is not None and git_snapshot.commit != capture_git_commit:
        raise _fail("formal E-0041 supplied Git snapshot/commit differs")
    materials = _load_materials(
        root,
        config_path,
        reader=reader,
        git_snapshot=git_snapshot,
    )
    first = _build_once(root, materials, capture_git_commit=capture_git_commit)
    second = _build_once(root, materials, capture_git_commit=capture_git_commit)
    if (
        first[1] != second[1]
        or first[2] != second[2]
        or first[5] != second[5]
        or first[0] != second[0]
        or first[4] != second[4]
    ):
        raise _fail("formal projection/workbook/provenance double build is not byte-identical")
    return FormalExportBuild(
        projection=first[0],
        projection_bytes=first[1],
        workbook_bytes=first[2],
        provenance=first[4],
        provenance_bytes=first[5],
        materials=materials,
    )


def build_e0041_formal_export(
    project_root: Path,
    *,
    capture_git_commit: str,
    config_path: Path = CONTROL_RELATIVE_PATH,
) -> FormalExportBuild:
    """Build twice after independently acquiring the strong clean-Git gate."""

    root = project_root.resolve()
    snapshot = _clean_git_snapshot(root)
    if capture_git_commit != snapshot.commit:
        raise _fail("requested formal build commit is not the verified clean HEAD/upstream")
    return _build_e0041_formal_export_verified(
        root,
        capture_git_commit=capture_git_commit,
        config_path=config_path,
        reader=_default_reader,
        git_snapshot=snapshot,
    )


def _recheck_all_materials(
    project_root: Path,
    materials: _Materials,
    *,
    reader: StableReader,
) -> None:
    for stable in materials.ordered_files:
        fresh = _read_file(
            project_root,
            stable.artifact,
            f"final recheck {stable.name}",
            reader=reader,
            maximum_size=max(len(stable.payload), 1),
            require_head_blob=stable.require_head_blob,
        )
        if fresh.payload != stable.payload or fresh.identity != stable.identity:
            raise _fail(f"formal input changed before publication: {stable.name}")


def _recheck_ignored_materials(
    project_root: Path,
    materials: _Materials,
    *,
    reader: StableReader,
) -> None:
    for stable in materials.ordered_files:
        if stable.require_head_blob:
            continue
        fresh = _read_file(
            project_root,
            stable.artifact,
            f"terminal ignored-input recheck {stable.name}",
            reader=reader,
            maximum_size=max(len(stable.payload), 1),
            require_head_blob=False,
        )
        if fresh.payload != stable.payload or fresh.identity != stable.identity:
            raise _fail(f"ignored formal input changed at publication boundary: {stable.name}")


def _assert_all_head_bindings(project_root: Path, materials: _Materials) -> None:
    for stable in materials.ordered_files:
        _assert_head_blob_binding(project_root, stable)


def _output_inventory(
    project_root: Path,
    *,
    expected: tuple[str, ...],
    allow_absent: bool,
) -> None:
    root, _root_identity = _core._open_trusted_root(project_root, "formal output inventory")
    descriptor: int | None = None
    try:
        try:
            descriptor, _chain = _core._open_directory_chain(
                root,
                OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
                "formal output inventory",
                create=False,
            )
        except Exception as exc:
            if allow_absent and isinstance(exc, _core.E0041PostMappingExportError):
                # Confirm absence through the canonical parent, rather than
                # treating every traversal error as an absent directory.
                parent, _ = _core._open_directory_chain(
                    root,
                    OUTPUT_DIRECTORY_RELATIVE_PATH.parts[:-1],
                    "formal output inventory parent",
                    create=False,
                )
                try:
                    try:
                        os.stat(
                            OUTPUT_DIRECTORY_RELATIVE_PATH.name,
                            dir_fd=parent,
                            follow_symlinks=False,
                        )
                    except FileNotFoundError:
                        return
                finally:
                    os.close(parent)
            raise _fail("cannot inspect canonical formal output inventory", exc) from exc
        observed = tuple(sorted(os.listdir(descriptor)))
        if observed != tuple(sorted(expected)):
            raise _fail(
                f"formal output inventory drifted: expected {sorted(expected)!r}, "
                f"observed {list(observed)!r}"
            )
        for filename in expected:
            item = os.stat(filename, dir_fd=descriptor, follow_symlinks=False)
            if not stat.S_ISREG(item.st_mode):
                raise _fail(f"formal output inventory member is not regular: {filename}")
    finally:
        if descriptor is not None:
            os.close(descriptor)
        os.close(root)


def _publish_formal_pair(
    project_root: Path,
    build: FormalExportBuild,
) -> dict[str, Any]:
    """Publish workbook first and provenance completion marker last.

    The operation is deliberately not claimed to be crash atomic.  Its held
    directory descriptor, exact-empty initial inventory, exclusive creates,
    identity-bound rollback, and final batch revalidation prevent a successful
    return with an unrelated or substituted file.
    """

    root, root_identity = _core._open_trusted_root(project_root, "formal pair publication")
    output: int | None = None
    created: list[tuple[str, os.stat_result, bytes]] = []
    try:
        output, held_chain = _core._open_directory_chain(
            root,
            OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
            "formal pair output directory",
            create=True,
        )
        if tuple(os.listdir(output)) != ():
            raise _fail("formal pair output directory must start exactly empty")
        workbook_identity = _core._write_exclusive_at(
            output,
            WORKBOOK_RELATIVE_PATH.name,
            build.workbook_bytes,
        )
        created.append((WORKBOOK_RELATIVE_PATH.name, workbook_identity, build.workbook_bytes))
        provenance_identity = _core._write_exclusive_at(
            output,
            PROVENANCE_RELATIVE_PATH.name,
            build.provenance_bytes,
        )
        created.append((PROVENANCE_RELATIVE_PATH.name, provenance_identity, build.provenance_bytes))
        os.fsync(output)
        expected_inventory = tuple(
            sorted((WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name))
        )
        if tuple(sorted(os.listdir(output))) != expected_inventory:
            raise _fail("formal pair held-directory final inventory is not exact")

        fresh_root, fresh_output, fresh_chain = _core._open_fresh_directory_chain(
            project_root,
            root_identity,
            OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
            "formal pair final canonical revalidation",
        )
        try:
            if (
                fresh_chain != held_chain
                or _core._directory_identity(os.fstat(fresh_output))
                != _core._directory_identity(os.fstat(output))
                or tuple(sorted(os.listdir(fresh_output))) != expected_inventory
            ):
                raise _fail("formal pair output directory detached from canonical path")
            canonical_workbook = os.stat(
                WORKBOOK_RELATIVE_PATH.name,
                dir_fd=fresh_output,
                follow_symlinks=False,
            )
            canonical_provenance = os.stat(
                PROVENANCE_RELATIVE_PATH.name,
                dir_fd=fresh_output,
                follow_symlinks=False,
            )
            if not _core._same_regular_file(workbook_identity, canonical_workbook) or not (
                _core._same_regular_file(provenance_identity, canonical_provenance)
            ):
                raise _fail("formal pair canonical inode identity drifted")
            _core._read_exact_batch_at(fresh_output, created)
        finally:
            os.close(fresh_output)
            os.close(fresh_root)
    except BaseException as publication_error:
        rollback_errors: list[BaseException] = []
        if output is not None:
            for filename, identity, _payload in reversed(created):
                try:
                    _core._rollback_created_at(output, filename, identity)
                except BaseException as rollback_error:
                    rollback_errors.append(rollback_error)
        if rollback_errors:
            details = "; ".join(str(error) for error in rollback_errors)
            raise _fail(
                f"formal pair publication rollback was incomplete: {details}"
            ) from publication_error
        raise
    finally:
        if output is not None:
            os.close(output)
        os.close(root)
    return build.provenance


def _read_unpinned_file(
    project_root: Path,
    relative_path: Path,
    name: str,
    *,
    reader: StableReader,
    maximum_size: int,
) -> _StableFile:
    path = _canonical_path(project_root, relative_path, name)
    try:
        before = os.stat(path, follow_symlinks=False)
    except OSError as exc:
        raise _fail(f"cannot stat {name}", exc) from exc
    if not stat.S_ISREG(before.st_mode) or before.st_size <= 0 or before.st_size > maximum_size:
        raise _fail(f"{name} is not a bounded regular file")
    payload = reader(project_root, path, maximum_size, name)
    after = os.stat(path, follow_symlinks=False)
    if _stat_identity(before) != _stat_identity(after) or len(payload) != before.st_size:
        raise _fail(f"{name} identity changed during stable read")
    return _StableFile(
        name=name,
        path=path,
        payload=payload,
        identity=_stat_identity(after),
        artifact=_record_for_payload(relative_path, payload),
        require_head_blob=False,
    )


def _validate_existing_provenance(
    payload: Mapping[str, Any],
    encoded: bytes,
    *,
    expected_commit: str,
) -> dict[str, Any]:
    expected_keys = {
        "format_version",
        "experiment_id",
        "dataset_role",
        "state",
        "capture_git_commit",
        "capture_git_dirty",
        "pair_hash_sealed",
        "outputs",
        "projection_receipt",
        "workbook_receipt",
        "input_hash_ledger",
        "implementation_hash_ledger",
        "runtime_hash_ledger",
        "runtime_versions",
        "geometry_authority",
        "metrics",
        "access_contract",
        "authority",
        "limitations",
        "claim_boundary",
    }
    if (
        set(payload) != expected_keys
        or payload.get("format_version") != 1
        or payload.get("experiment_id") != "E-0041"
        or payload.get("dataset_role") != "CALIBRATION"
        or payload.get("state") != PAIR_STATE
        or payload.get("capture_git_commit") != expected_commit
        or payload.get("capture_git_dirty") is not False
        or payload.get("pair_hash_sealed") is not False
        or payload.get("claim_boundary") != _CLAIM_BOUNDARY
        or _encoded_json(payload) != encoded
    ):
        raise _fail("published formal provenance envelope or canonical bytes drifted")
    _assert_no_absolute_paths(payload, "published formal provenance")
    outputs = _mapping(payload.get("outputs"), "published formal outputs")
    if (
        set(outputs)
        != {
            "workbook",
            "provenance_path",
            "publication_order",
            "provenance_is_completion_marker",
            "atomic_pair_publication",
        }
        or outputs.get("provenance_path") != PROVENANCE_RELATIVE_PATH.as_posix()
        or outputs.get("publication_order")
        != [WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name]
        or outputs.get("provenance_is_completion_marker") is not True
        or outputs.get("atomic_pair_publication") is not False
    ):
        raise _fail("published formal provenance output contract drifted")
    workbook_raw = _mapping(outputs.get("workbook"), "published formal workbook")
    workbook = _artifact_identity(workbook_raw, "published formal workbook")
    if workbook_raw != workbook or workbook != {
        "path": WORKBOOK_RELATIVE_PATH.as_posix(),
        "sha256": _WORKBOOK_SHA256,
        "size_bytes": _WORKBOOK_SIZE,
    }:
        raise _fail("published formal workbook identity drifted")
    if b"provenance_sha256" in encoded:
        raise _fail("published formal provenance contains a forbidden self-hash")
    return workbook


def _validate_published_workbook_structure(payload: bytes) -> None:
    try:
        workbook = load_workbook(BytesIO(payload), data_only=False, read_only=False)
    except Exception as exc:
        raise _fail("cannot open published E-0041 workbook", exc) from exc
    try:
        if workbook.sheetnames != [
            "Sheet1",
            "PROVENANCE",
            "VALIDATION_DIAGNOSTICS",
            "RUN_METADATA",
        ]:
            raise _fail("published E-0041 workbook sheet order drifted")
        formula_count = sum(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        snapshot = _core._template_snapshot(workbook, "Sheet1", 78)
        if formula_count != 0 or _sha256(_compact_bytes(snapshot)) != _TEMPLATE_FIDELITY_SHA256:
            raise _fail("published E-0041 workbook formula/A:C fidelity drifted")
        properties = workbook.properties
        if (
            properties.creator != "bctc-ai/E-0041"
            or properties.lastModifiedBy != "bctc-ai/E-0041"
            or properties.created != _core._DETERMINISTIC_CORE_TIMESTAMP
            or properties.modified != _core._DETERMINISTIC_CORE_TIMESTAMP
            or properties.version != "1"
            or properties.revision != "1"
        ):
            raise _fail("published E-0041 workbook core properties drifted")
    finally:
        workbook.close()


def _assert_stable_unchanged(
    project_root: Path,
    stable: _StableFile,
    *,
    reader: StableReader,
) -> None:
    fresh = _read_file(
        project_root,
        stable.artifact,
        f"final recheck {stable.name}",
        reader=reader,
        maximum_size=len(stable.payload),
        require_head_blob=False,
    )
    if fresh.payload != stable.payload or fresh.identity != stable.identity:
        raise _fail(f"published pair member changed during sealing: {stable.name}")


def _recheck_pair_batch(
    project_root: Path,
    *,
    workbook_file: _StableFile,
    provenance_file: _StableFile,
) -> None:
    root, root_identity = _core._open_trusted_root(project_root, "formal pair batch recheck")
    output: int | None = None
    try:
        output, held_chain = _core._open_directory_chain(
            root,
            OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
            "formal pair batch recheck",
            create=False,
        )
        expected_inventory = tuple(
            sorted((WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name))
        )
        if tuple(sorted(os.listdir(output))) != expected_inventory:
            raise _fail("formal pair batch inventory drifted before seal publication")
        workbook_link = os.stat(
            WORKBOOK_RELATIVE_PATH.name,
            dir_fd=output,
            follow_symlinks=False,
        )
        provenance_link = os.stat(
            PROVENANCE_RELATIVE_PATH.name,
            dir_fd=output,
            follow_symlinks=False,
        )
        if (
            _stat_identity(workbook_link) != workbook_file.identity
            or _stat_identity(provenance_link) != provenance_file.identity
        ):
            raise _fail("formal pair canonical identities drifted before batch recheck")
        fresh_root, fresh_output, fresh_chain = _core._open_fresh_directory_chain(
            project_root,
            root_identity,
            OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
            "formal pair batch fresh-root recheck",
        )
        try:
            if (
                fresh_chain != held_chain
                or _core._directory_identity(os.fstat(fresh_output))
                != _core._directory_identity(os.fstat(output))
                or tuple(sorted(os.listdir(fresh_output))) != expected_inventory
            ):
                raise _fail("formal pair detached during final batch recheck")
            _core._read_exact_batch_at(
                fresh_output,
                [
                    (WORKBOOK_RELATIVE_PATH.name, workbook_link, workbook_file.payload),
                    (PROVENANCE_RELATIVE_PATH.name, provenance_link, provenance_file.payload),
                ],
            )
        finally:
            os.close(fresh_output)
            os.close(fresh_root)
    finally:
        if output is not None:
            os.close(output)
        os.close(root)


def _seal_payload(
    build: FormalExportBuild,
    *,
    commit: str,
    workbook_file: _StableFile,
    provenance_file: _StableFile,
) -> dict[str, Any]:
    return {
        "format_version": 1,
        "experiment_id": "E-0041",
        "dataset_role": "CALIBRATION",
        "state": SEAL_STATE,
        "seal_git_commit": commit,
        "seal_git_dirty": False,
        "pair_capture_git_commit": commit,
        "inventory": {
            "file_count": 2,
            "files": [workbook_file.artifact, provenance_file.artifact],
        },
        "projection_receipt": {
            "sha256": _sha256(build.projection_bytes),
            "size_bytes": len(build.projection_bytes),
        },
        "metrics": build.provenance["metrics"],
        "input_hash_ledger": {
            "control": build.materials.prerequisites.control_file.artifact,
            "workbook": workbook_file.artifact,
            "provenance": provenance_file.artifact,
            "deterministic_replay_inputs": build.provenance["input_hash_ledger"],
            "deterministic_replay_implementation": build.provenance["implementation_hash_ledger"],
            "deterministic_replay_runtime": build.provenance["runtime_hash_ledger"],
            "deterministic_replay_runtime_versions": build.provenance["runtime_versions"],
            "geometry_asset_inventory": build.provenance["geometry_authority"]["asset_inventory"],
        },
        "replay": {
            "formal_replay_invocation_count": 1,
            "rebuilt_projection_matches_provenance_sha256_and_size_receipt": True,
            "workbook_exact_byte_equality": True,
            "provenance_exact_canonical_byte_equality": True,
            "same_clean_git_commit": True,
            "published_pair_rewritten": False,
        },
        "access_contract": {
            "canonical_provenance_envelope_and_workbook_link_validated_before_workbook_open": True,
            "exact_two_file_inventory_validated_before_replay_and_seal_publication": True,
            "all_replay_inputs_rechecked_and_terminal_clean_git_gate_passed": True,
            "superseded_mapping_or_review_artifact_opened": False,
            "review_or_steward_answers_opened": False,
            "history_or_mongodb_opened": False,
            "qwen_raw_or_token_output_opened": False,
            "holdout_artifact_opened": False,
        },
        "authority": {
            "exact_two_file_hash_identity": True,
            "deterministic_projection_workbook_provenance_replay": True,
            "schema_authority": False,
            "mapping_accuracy": False,
            "review_or_steward_approval": False,
            "accounting_correctness": False,
            "human_gold": False,
            "holdout_or_production": False,
        },
        "limitations": build.provenance["limitations"],
        "claim_boundary": (
            "This artifact hash-seals exactly the two-file E-0041 calibration export pair "
            "after exact deterministic projection, workbook, and provenance replay at the "
            "same clean commit. It adds no schema, mapping-accuracy, review, accounting, "
            "human-gold, holdout, or production authority."
        ),
    }


def _validate_seal_payload(
    payload: Mapping[str, Any],
    encoded: bytes,
    *,
    build: FormalExportBuild,
    commit: str,
    workbook_file: _StableFile,
    provenance_file: _StableFile,
) -> None:
    expected_keys = {
        "format_version",
        "experiment_id",
        "dataset_role",
        "state",
        "seal_git_commit",
        "seal_git_dirty",
        "pair_capture_git_commit",
        "inventory",
        "projection_receipt",
        "metrics",
        "input_hash_ledger",
        "replay",
        "access_contract",
        "authority",
        "limitations",
        "claim_boundary",
    }
    inventory = _mapping(payload.get("inventory"), "formal seal inventory")
    expected_inventory = {
        "file_count": 2,
        "files": [workbook_file.artifact, provenance_file.artifact],
    }
    expected_replay = {
        "formal_replay_invocation_count": 1,
        "rebuilt_projection_matches_provenance_sha256_and_size_receipt": True,
        "workbook_exact_byte_equality": True,
        "provenance_exact_canonical_byte_equality": True,
        "same_clean_git_commit": True,
        "published_pair_rewritten": False,
    }
    expected_access = {
        "canonical_provenance_envelope_and_workbook_link_validated_before_workbook_open": True,
        "exact_two_file_inventory_validated_before_replay_and_seal_publication": True,
        "all_replay_inputs_rechecked_and_terminal_clean_git_gate_passed": True,
        "superseded_mapping_or_review_artifact_opened": False,
        "review_or_steward_answers_opened": False,
        "history_or_mongodb_opened": False,
        "qwen_raw_or_token_output_opened": False,
        "holdout_artifact_opened": False,
    }
    expected_authority = {
        "exact_two_file_hash_identity": True,
        "deterministic_projection_workbook_provenance_replay": True,
        "schema_authority": False,
        "mapping_accuracy": False,
        "review_or_steward_approval": False,
        "accounting_correctness": False,
        "human_gold": False,
        "holdout_or_production": False,
    }
    expected_ledger = {
        "control": build.materials.prerequisites.control_file.artifact,
        "workbook": workbook_file.artifact,
        "provenance": provenance_file.artifact,
        "deterministic_replay_inputs": build.provenance["input_hash_ledger"],
        "deterministic_replay_implementation": build.provenance["implementation_hash_ledger"],
        "deterministic_replay_runtime": build.provenance["runtime_hash_ledger"],
        "deterministic_replay_runtime_versions": build.provenance["runtime_versions"],
        "geometry_asset_inventory": build.provenance["geometry_authority"]["asset_inventory"],
    }
    if (
        set(payload) != expected_keys
        or payload.get("format_version") != 1
        or payload.get("experiment_id") != "E-0041"
        or payload.get("dataset_role") != "CALIBRATION"
        or payload.get("state") != SEAL_STATE
        or payload.get("seal_git_commit") != commit
        or payload.get("seal_git_dirty") is not False
        or payload.get("pair_capture_git_commit") != commit
        or inventory != expected_inventory
        or payload.get("projection_receipt")
        != {"sha256": _PROJECTION_SHA256, "size_bytes": _PROJECTION_SIZE}
        or payload.get("metrics") != build.provenance["metrics"]
        or payload.get("input_hash_ledger") != expected_ledger
        or payload.get("replay") != expected_replay
        or payload.get("access_contract") != expected_access
        or payload.get("authority") != expected_authority
        or payload.get("limitations") != build.provenance["limitations"]
        or payload.get("claim_boundary")
        != (
            "This artifact hash-seals exactly the two-file E-0041 calibration export pair "
            "after exact deterministic projection, workbook, and provenance replay at the "
            "same clean commit. It adds no schema, mapping-accuracy, review, accounting, "
            "human-gold, holdout, or production authority."
        )
        or _encoded_json(payload) != encoded
    ):
        raise _fail("formal E-0041 seal full contract validation failed")


def _publish_formal_seal(
    project_root: Path,
    payload: dict[str, Any],
    *,
    workbook_file: _StableFile,
    provenance_file: _StableFile,
) -> dict[str, Any]:
    encoded = _encoded_json(payload)
    root, root_identity = _core._open_trusted_root(project_root, "formal seal publication")
    pair: int | None = None
    parent: int | None = None
    created_identity: os.stat_result | None = None
    try:
        pair, held_pair_chain = _core._open_directory_chain(
            root,
            OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
            "formal seal held pair",
            create=False,
        )
        pair_inventory = tuple(sorted((WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name)))
        if tuple(sorted(os.listdir(pair))) != pair_inventory:
            raise _fail("formal seal held pair inventory drifted")
        workbook_identity = os.stat(
            WORKBOOK_RELATIVE_PATH.name,
            dir_fd=pair,
            follow_symlinks=False,
        )
        provenance_identity = os.stat(
            PROVENANCE_RELATIVE_PATH.name,
            dir_fd=pair,
            follow_symlinks=False,
        )
        if (
            _stat_identity(workbook_identity) != workbook_file.identity
            or _stat_identity(provenance_identity) != provenance_file.identity
        ):
            raise _fail("formal seal held pair identity drifted")
        _core._read_exact_batch_at(
            pair,
            [
                (WORKBOOK_RELATIVE_PATH.name, workbook_identity, workbook_file.payload),
                (PROVENANCE_RELATIVE_PATH.name, provenance_identity, provenance_file.payload),
            ],
        )

        parent, held_seal_chain = _core._open_directory_chain(
            root,
            SEAL_RELATIVE_PATH.parts[:-1],
            "formal seal parent",
            create=False,
        )
        created_identity = _core._write_exclusive_at(parent, SEAL_RELATIVE_PATH.name, encoded)
        os.fsync(parent)
        fresh_root, fresh_root_identity = _core._open_trusted_root(
            project_root,
            "formal seal final shared-root revalidation",
        )
        try:
            if fresh_root_identity != root_identity:
                raise _fail("formal seal trusted root changed during publication")
            fresh_pair: int | None = None
            fresh_parent: int | None = None
            try:
                fresh_pair, fresh_pair_chain = _core._open_directory_chain(
                    fresh_root,
                    OUTPUT_DIRECTORY_RELATIVE_PATH.parts,
                    "formal seal fresh pair",
                    create=False,
                )
                fresh_parent, fresh_seal_chain = _core._open_directory_chain(
                    fresh_root,
                    SEAL_RELATIVE_PATH.parts[:-1],
                    "formal seal fresh parent",
                    create=False,
                )
                canonical = os.stat(
                    SEAL_RELATIVE_PATH.name,
                    dir_fd=fresh_parent,
                    follow_symlinks=False,
                )
                if (
                    fresh_pair_chain != held_pair_chain
                    or _core._directory_identity(os.fstat(fresh_pair))
                    != _core._directory_identity(os.fstat(pair))
                    or tuple(sorted(os.listdir(fresh_pair))) != pair_inventory
                    or fresh_seal_chain != held_seal_chain
                    or _core._directory_identity(os.fstat(fresh_parent))
                    != _core._directory_identity(os.fstat(parent))
                    or not _core._same_regular_file(created_identity, canonical)
                ):
                    raise _fail("formal pair or seal detached from the shared canonical root")
                _core._read_exact_batch_at(
                    fresh_pair,
                    [
                        (
                            WORKBOOK_RELATIVE_PATH.name,
                            workbook_identity,
                            workbook_file.payload,
                        ),
                        (
                            PROVENANCE_RELATIVE_PATH.name,
                            provenance_identity,
                            provenance_file.payload,
                        ),
                    ],
                )
                _core._read_exact_at(
                    fresh_parent,
                    SEAL_RELATIVE_PATH.name,
                    created_identity,
                    encoded,
                )
            finally:
                if fresh_parent is not None:
                    os.close(fresh_parent)
                if fresh_pair is not None:
                    os.close(fresh_pair)
        finally:
            os.close(fresh_root)
    except BaseException as publication_error:
        if parent is not None and created_identity is not None:
            try:
                _core._rollback_created_at(parent, SEAL_RELATIVE_PATH.name, created_identity)
            except BaseException as rollback_error:
                raise _fail(
                    f"formal seal publication rollback was incomplete: {rollback_error}"
                ) from publication_error
        raise
    finally:
        if parent is not None:
            os.close(parent)
        if pair is not None:
            os.close(pair)
        os.close(root)
    return payload


def dry_run_e0041_formal_export(
    project_root: Path,
    *,
    config_path: Path = CONTROL_RELATIVE_PATH,
) -> FormalExportBuild:
    """Build the complete formal candidate without publishing any output."""

    _assert_process_isolation()
    root = project_root.resolve()
    snapshot = _clean_git_snapshot(root)
    return _build_e0041_formal_export_verified(
        root,
        capture_git_commit=snapshot.commit,
        config_path=config_path,
        reader=_default_reader,
        git_snapshot=snapshot,
    )


def capture_e0041_formal_export(
    project_root: Path,
    *,
    config_path: Path = CONTROL_RELATIVE_PATH,
) -> dict[str, Any]:
    """Exclusively publish the formal workbook/provenance pair from a clean commit."""

    _assert_process_isolation()
    root = project_root.resolve()
    initial_git = _clean_git_snapshot(root)
    _output_inventory(root, expected=(), allow_absent=True)
    build = _build_e0041_formal_export_verified(
        root,
        capture_git_commit=initial_git.commit,
        config_path=config_path,
        reader=_default_reader,
        git_snapshot=initial_git,
    )
    _assert_process_isolation()
    final_git = _clean_git_snapshot(root)
    if final_git != initial_git:
        raise _fail("formal E-0041 Git HEAD/upstream changed during capture")
    _assert_all_head_bindings(root, build.materials)
    _recheck_all_materials(root, build.materials, reader=_default_reader)
    ultimate_git = _clean_git_snapshot(root)
    if ultimate_git != initial_git:
        raise _fail("formal E-0041 Git state changed during final input rechecks")
    _recheck_all_materials(root, build.materials, reader=_default_reader)
    terminal_git = _clean_git_snapshot(root)
    if terminal_git != initial_git:
        raise _fail("formal E-0041 Git state changed at terminal publication gate")
    _assert_all_head_bindings(root, build.materials)
    _recheck_ignored_materials(root, build.materials, reader=_default_reader)
    # This publisher owns the final exact-empty check under its held dirfd.
    # No validation or observer runs after it returns.
    return _publish_formal_pair(root, build)


def capture_e0041_formal_export_seal(
    project_root: Path,
    *,
    config_path: Path = CONTROL_RELATIVE_PATH,
) -> dict[str, Any]:
    """Replay and exclusively hash-seal the exact two-file formal export pair."""

    _assert_process_isolation()
    root = project_root.resolve()
    initial_git = _clean_git_snapshot(root)
    _output_inventory(
        root,
        expected=(WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name),
        allow_absent=False,
    )
    provenance_file = _read_unpinned_file(
        root,
        PROVENANCE_RELATIVE_PATH,
        "published E-0041 provenance",
        reader=_default_reader,
        maximum_size=16 * 1024 * 1024,
    )
    provenance = _decode_json(provenance_file.payload, provenance_file.name)
    workbook_record = _validate_existing_provenance(
        provenance,
        provenance_file.payload,
        expected_commit=initial_git.commit,
    )
    workbook_file = _read_file(
        root,
        workbook_record,
        "published E-0041 workbook",
        reader=_default_reader,
        maximum_size=16 * 1024 * 1024,
        require_head_blob=False,
    )
    # Structure/formula validation occurs before replay, but expected values
    # remain frozen constants rather than values derived from provenance.
    if workbook_file.artifact != {
        "path": WORKBOOK_RELATIVE_PATH.as_posix(),
        "sha256": _WORKBOOK_SHA256,
        "size_bytes": _WORKBOOK_SIZE,
    }:
        raise _fail("published E-0041 workbook identity is not the frozen oracle")
    _validate_published_workbook_structure(workbook_file.payload)
    build = _build_e0041_formal_export_verified(
        root,
        capture_git_commit=initial_git.commit,
        config_path=config_path,
        reader=_default_reader,
        git_snapshot=initial_git,
    )
    if (
        build.workbook_bytes != workbook_file.payload
        or build.provenance_bytes != provenance_file.payload
        or build.provenance != provenance
        or _compact_bytes(build.projection) != build.projection_bytes
    ):
        raise _fail("published E-0041 pair differs from deterministic formal replay")
    seal_payload = _seal_payload(
        build,
        commit=initial_git.commit,
        workbook_file=workbook_file,
        provenance_file=provenance_file,
    )
    encoded_seal = _encoded_json(seal_payload)
    if _encoded_json(_decode_json(encoded_seal, "formal seal round trip")) != encoded_seal:
        raise _fail("formal E-0041 seal is not canonical JSON")
    _assert_no_absolute_paths(seal_payload, "formal E-0041 seal")
    _validate_seal_payload(
        seal_payload,
        encoded_seal,
        build=build,
        commit=initial_git.commit,
        workbook_file=workbook_file,
        provenance_file=provenance_file,
    )
    _assert_process_isolation()
    final_git = _clean_git_snapshot(root)
    if final_git != initial_git:
        raise _fail("formal E-0041 Git HEAD/upstream changed during sealing")
    _assert_all_head_bindings(root, build.materials)
    _recheck_all_materials(root, build.materials, reader=_default_reader)
    ultimate_git = _clean_git_snapshot(root)
    if ultimate_git != initial_git:
        raise _fail("formal E-0041 Git state changed during final seal input rechecks")
    _recheck_all_materials(root, build.materials, reader=_default_reader)
    terminal_git = _clean_git_snapshot(root)
    if terminal_git != initial_git:
        raise _fail("formal E-0041 Git state changed at terminal seal gate")
    _assert_all_head_bindings(root, build.materials)
    _recheck_ignored_materials(root, build.materials, reader=_default_reader)
    _output_inventory(
        root,
        expected=(WORKBOOK_RELATIVE_PATH.name, PROVENANCE_RELATIVE_PATH.name),
        allow_absent=False,
    )
    _assert_stable_unchanged(root, provenance_file, reader=_default_reader)
    _assert_stable_unchanged(root, workbook_file, reader=_default_reader)
    # Exclusive seal publication is intentionally the final action.  The
    # publisher performs its own canonical byte/inode/root checks internally.
    return _publish_formal_seal(
        root,
        seal_payload,
        workbook_file=workbook_file,
        provenance_file=provenance_file,
    )
