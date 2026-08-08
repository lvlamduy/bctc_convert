from __future__ import annotations

import copy
import errno
import hashlib
import json
import os
import re
import stat
import subprocess
import zipfile
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any, cast

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.cell_style import StyleArray

from bctc_ai.mapping.e0040_calibration_challenger import E0040ChallengerResult
from bctc_ai.validation.arithmetic import NumericOperand, check_parent_children, check_sum


class E0041PostMappingExportError(RuntimeError):
    pass


CONTROL_RELATIVE_PATH = Path(
    "config/experiments/e0041-mbb-cdkt-post-mapping-development-excel.yaml"
)

_SHA256 = re.compile(r"[0-9a-f]{64}")
_ROW_ID = re.compile(r"page-(?P<page>\d{4})-row-(?P<row>\d{3})-label")
_CELL_ID = re.compile(r"page-(?P<page>\d{4})-row-(?P<row>\d{3})-axis-(?P<axis>[12])")
_SELECTED_MAPPING_STATUSES = frozenset({"RESOLVED_ANCHOR", "RESOLVED_PATH"})
_UNSELECTED_MAPPING_STATUSES = frozenset(
    {"BEST_PATH_SKIPPED", "NO_ADMISSIBLE_PAIR", "AMBIGUOUS_ACROSS_PATHS"}
)
_SOURCE_ONLY_MAPPING_STATUSES = frozenset(
    {"SOURCE_ONLY_STRUCTURAL_ROW", "OUT_OF_SCOPE_FOR_TARGET_TEMPLATE"}
)
_FINAL_STATUSES = frozenset(
    {
        "VALUE",
        "ZERO",
        "DASH",
        "BLANK",
        "NOT_OBSERVED",
        "AMBIGUOUS",
        "UNRESOLVED",
    }
)
_STRICT_NUMERIC_STATUSES = frozenset({"VALUE", "ZERO"})
_MAX_SAFE_EXCEL_INTEGER = 2**53
_WORKBOOK_NUMBER_FORMAT = "#,##0;[Red](#,##0);-"
_E0040_POLICY_SHA256 = "eba3a1380f44f34958398edb13076dc3a87da95fc5ff347968a1a2c023e3995a"
_E0040_MAPPER_POLICY_SHA256 = "2f18880339b8e2c04ec3ba900919f174f8af478515adfbfb0e43ff80ddd13268"
_E0040_BASE_PROJECTION_SHA256 = "7025ca729c01a3f2030af38e9745a0ee8d72b1dad60c8d4e3b7cf749e5eb860c"
_E0040_RESULT_PROJECTION_SHA256 = "5c3c4a09650beda8eca21e5a00fe459e052ae7cc8d735359bc41a58a391da9b0"
_E0040_MAPPING_ARTIFACT = {
    "path": "output/calibration/e0040-mbb-cdkt-formal-mapping/mapping_only.json",
    "sha256": "8def983007fc3aacf59351395426d5246ad3f28d605442f590de55eaf396cb0d",
    "size_bytes": 1_157_172,
}
_E0040_MAPPING_SEAL_ARTIFACT = {
    "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-seal.json",
    "sha256": "68306f7f540faa77d6e2e383927eae23fc3724cfdc8c53cded978a86f3a00b29",
    "size_bytes": 7_611,
}
_E0040_S3_REGISTRATION_ARTIFACT = {
    "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-s3-registration.json",
    "sha256": "f38d9a1bbed4ec48e2156d441e5c76c6e6d82b0771208de3eef92d96173dd4b5",
    "size_bytes": 13_360,
}
_E0040_CHALLENGER_RESULT_SHA256 = "2e49d8623692fde9fd4a5a87f9c2e2159941b0f3ded7b7b16dddac2ab1e85fbd"
_E0040_CHALLENGER_RESULT_SIZE_BYTES = 700_869
_E0040_CAPTURE_GIT_COMMIT = "18aca8942faf5d47e1ac5f049045d7a7a297b5fc"
_E0041_GEOMETRY_REGISTRY_ARTIFACT = {
    "path": (
        "output/calibration/e0041-mbb-cdkt-reconstructed-geometry/65fa9b7c0de1/crop_registry.json"
    ),
    "sha256": "65fa9b7c0de1f0db26ae57a46dae2bb64c2475e3a87e5194461f208fc786cbef",
    "size_bytes": 217_837,
}
_E0041_NORMALIZED_GEOMETRY_REGISTRY_SHA256 = (
    "e834efd4f6e70c03e607d834a17adc69e0fa0868658767637c5db3cf8c06be6a"
)
_LEGACY_GEOMETRY_PROJECT_ROOT = PurePosixPath("/workspace/bctc-ai")
_GEOMETRY_ROOTED_PATH_FIELDS = frozenset(
    {"path", "ocr_path", "render_path", "source_ocr_path", "source_render_path"}
)
_E0041_PHYSICAL_EQUATIONS_SHA256 = (
    "a611078b4734d1e57026d58db5aced4a0b342114ba57aefdff29868032b3b42b"
)
_DETERMINISTIC_CORE_TIMESTAMP = datetime(2000, 1, 1, 0, 0, 0)
_DETERMINISTIC_CORE_ACTOR = "bctc-ai/E-0041"
_DETERMINISTIC_CORE_PROPERTIES_XML = (
    b'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/'
    b'metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" '
    b'xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/'
    b'XMLSchema-instance"><dc:creator>bctc-ai/E-0041</dc:creator><dcterms:created '
    b'xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:created>'
    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z'
    b"</dcterms:modified><cp:lastModifiedBy>bctc-ai/E-0041</cp:lastModifiedBy>"
    b"<cp:version>1</cp:version><cp:revision>1</cp:revision></cp:coreProperties>"
)
_DETERMINISTIC_CORE_PROPERTIES_SHA256 = (
    "a025959e8b178cfc6c6aae8f2d49d86fa305d3e36e165c8cbbc16923068668e4"
)


@dataclass(frozen=True, slots=True)
class E0040ArtifactRecord:
    path: str
    sha256: str
    size_bytes: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "sha256": self.sha256,
            "size_bytes": self.size_bytes,
        }


@dataclass(frozen=True, slots=True)
class AuthenticatedE0040ResultCarrier:
    """Immutable E-0040 result minted from the three exact formal artifacts."""

    mapping_bytes: bytes
    seal_bytes: bytes
    registration_bytes: bytes
    mapping_artifact: E0040ArtifactRecord
    seal_artifact: E0040ArtifactRecord
    registration_artifact: E0040ArtifactRecord
    challenger_result_sha256: str
    capture_git_commit: str


def _fail(message: str, error: BaseException | None = None) -> E0041PostMappingExportError:
    if error is None:
        return E0041PostMappingExportError(message)
    return E0041PostMappingExportError(f"{message}: {error}")


def _canonical_bytes(value: object) -> bytes:
    try:
        return json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise _fail("value is not canonical JSON", exc) from exc


def _canonical_sha256(value: object) -> str:
    return hashlib.sha256(_canonical_bytes(value)).hexdigest()


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _json_pairs(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise _fail(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _reject_constant(value: str) -> None:
    raise _fail(f"non-finite JSON constant: {value}")


def _load_json_bytes(payload: bytes, name: str) -> dict[str, Any]:
    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=_json_pairs,
            parse_constant=_reject_constant,
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise _fail(f"cannot decode {name}", exc) from exc
    if not isinstance(value, dict):
        raise _fail(f"{name} must be a JSON object")
    return value


def _mapping(value: object, name: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise _fail(f"{name} must be an object")
    return cast(dict[str, Any], value)


def _sequence(value: object, name: str) -> list[Any]:
    if not isinstance(value, list):
        raise _fail(f"{name} must be an array")
    return value


def _positive_int(value: object, name: str) -> int:
    if type(value) is not int or value <= 0:
        raise _fail(f"{name} must be a positive integer")
    return value


def _artifact_record(value: object, name: str) -> dict[str, Any]:
    record = _mapping(value, name)
    if set(record) != {"path", "sha256", "size_bytes"}:
        raise _fail(f"{name} must contain exact artifact identity fields")
    path = record.get("path")
    digest = record.get("sha256")
    size = record.get("size_bytes")
    if (
        not isinstance(path, str)
        or not path
        or PurePosixPath(path).is_absolute()
        or ".." in PurePosixPath(path).parts
        or not isinstance(digest, str)
        or _SHA256.fullmatch(digest) is None
        or type(size) is not int
        or size <= 0
    ):
        raise _fail(f"{name} has an invalid artifact identity")
    return {"path": path, "sha256": digest, "size_bytes": size}


def _artifact_identity(value: object, name: str) -> dict[str, Any]:
    """Extract only byte-identity fields from an annotated control record."""

    record = _mapping(value, name)
    return _artifact_record(
        {key: record.get(key) for key in ("path", "sha256", "size_bytes")},
        name,
    )


def _load_exact_e0040_artifact(
    payload: bytes,
    expected: Mapping[str, Any],
    name: str,
) -> dict[str, Any]:
    if type(payload) is not bytes:
        raise _fail(f"{name} must be exact bytes")
    if len(payload) != expected["size_bytes"] or _sha256_bytes(payload) != expected["sha256"]:
        raise _fail(f"{name} byte identity drifted")
    return _load_json_bytes(payload, name)


def _frozen_artifact_record(value: Mapping[str, Any]) -> E0040ArtifactRecord:
    return E0040ArtifactRecord(
        path=cast(str, value["path"]),
        sha256=cast(str, value["sha256"]),
        size_bytes=cast(int, value["size_bytes"]),
    )


def _authenticate_e0040_artifact_chain(
    *,
    mapping_bytes: bytes,
    seal_bytes: bytes,
    registration_bytes: bytes,
) -> bytes:
    """Authenticate the exact formal chain and return canonical result bytes."""

    # Validate the already-read seal and durability registration before decoding
    # mapping content. A future formal reader must preserve that same file-open
    # order when it supplies these bytes.
    seal = _load_exact_e0040_artifact(
        seal_bytes,
        _E0040_MAPPING_SEAL_ARTIFACT,
        "E-0040 formal mapping seal",
    )
    registration = _load_exact_e0040_artifact(
        registration_bytes,
        _E0040_S3_REGISTRATION_ARTIFACT,
        "E-0040 formal S3 registration",
    )
    mapping = _load_exact_e0040_artifact(
        mapping_bytes,
        _E0040_MAPPING_ARTIFACT,
        "E-0040 formal mapping",
    )

    challenger_result = mapping.get("challenger_result")
    if type(challenger_result) is not dict:
        raise _fail("E-0040 formal mapping lacks a JSON-native challenger result")
    challenger_result = cast(dict[str, Any], challenger_result)
    challenger_bytes = _canonical_bytes(challenger_result)
    challenger_digest = _sha256_bytes(challenger_bytes)
    mapping_receipts = _mapping(mapping.get("result_receipts"), "E-0040 result receipts")
    if (
        challenger_digest != _E0040_CHALLENGER_RESULT_SHA256
        or len(challenger_bytes) != _E0040_CHALLENGER_RESULT_SIZE_BYTES
        or mapping_receipts.get("challenger_result_sha256") != challenger_digest
        or mapping_receipts.get("challenger_result_size_bytes") != len(challenger_bytes)
    ):
        raise _fail("E-0040 formal challenger result identity drifted")

    seal_ledger = _mapping(seal.get("input_hash_ledger"), "E-0040 seal input ledger")
    seal_inventory = _mapping(seal.get("inventory"), "E-0040 seal inventory")
    inventory_files = _sequence(seal_inventory.get("files"), "E-0040 seal inventory files")
    seal_receipts = _mapping(seal.get("result_receipts"), "E-0040 seal result receipts")
    if (
        _artifact_record(seal_ledger.get("mapping_only"), "E-0040 seal mapping")
        != _E0040_MAPPING_ARTIFACT
        or seal_inventory.get("file_count") != 1
        or len(inventory_files) != 1
        or _artifact_record(inventory_files[0], "E-0040 seal inventory mapping")
        != _E0040_MAPPING_ARTIFACT
        or seal_receipts != mapping_receipts
        or mapping.get("capture_git_commit") != _E0040_CAPTURE_GIT_COMMIT
        or seal.get("mapping_capture_git_commit") != _E0040_CAPTURE_GIT_COMMIT
        or seal.get("seal_git_commit") != _E0040_CAPTURE_GIT_COMMIT
    ):
        raise _fail("E-0040 mapping/seal linkage drifted")

    local_artifacts = _mapping(
        registration.get("local_artifacts"),
        "E-0040 registration local artifacts",
    )
    summary = _mapping(
        registration.get("formal_result_summary"),
        "E-0040 registration formal result summary",
    )
    seal_linkage = _mapping(
        registration.get("seal_linkage"),
        "E-0040 registration seal linkage",
    )
    required_linkage_flags = (
        "mapping_canonical_bytes_validated",
        "mapping_inventory_identity_matches",
        "mapping_ledger_identity_matches",
        "mapping_metrics_match_seal",
        "mapping_result_receipts_match_seal",
        "result_projection_matches_mapping",
        "s3_source_git_commit_matches_seal_artifact_commit",
    )
    if (
        set(local_artifacts) != {"mapping_only", "mapping_seal"}
        or _artifact_record(local_artifacts.get("mapping_only"), "registered E-0040 mapping")
        != _E0040_MAPPING_ARTIFACT
        or _artifact_record(local_artifacts.get("mapping_seal"), "registered E-0040 seal")
        != _E0040_MAPPING_SEAL_ARTIFACT
        or summary.get("challenger_result_sha256") != challenger_digest
        or summary.get("final_result_sha256") != mapping_receipts.get("final_result_sha256")
        or summary.get("final_selected_pairs_sha256")
        != mapping_receipts.get("final_selected_pairs_sha256")
        or summary.get("result_projection_sha256") != _E0040_RESULT_PROJECTION_SHA256
        or seal_linkage.get("mapping_capture_git_commit") != _E0040_CAPTURE_GIT_COMMIT
        or any(seal_linkage.get(name) is not True for name in required_linkage_flags)
    ):
        raise _fail("E-0040 registration linkage drifted")

    return challenger_bytes


def authenticate_e0040_result_carrier(
    *,
    mapping_bytes: bytes,
    seal_bytes: bytes,
    registration_bytes: bytes,
) -> AuthenticatedE0040ResultCarrier:
    """Mint a carrier only after authenticating the exact formal E-0040 chain."""

    challenger_bytes = _authenticate_e0040_artifact_chain(
        mapping_bytes=mapping_bytes,
        seal_bytes=seal_bytes,
        registration_bytes=registration_bytes,
    )
    return AuthenticatedE0040ResultCarrier(
        mapping_bytes=mapping_bytes,
        seal_bytes=seal_bytes,
        registration_bytes=registration_bytes,
        mapping_artifact=_frozen_artifact_record(_E0040_MAPPING_ARTIFACT),
        seal_artifact=_frozen_artifact_record(_E0040_MAPPING_SEAL_ARTIFACT),
        registration_artifact=_frozen_artifact_record(_E0040_S3_REGISTRATION_ARTIFACT),
        challenger_result_sha256=_sha256_bytes(challenger_bytes),
        capture_git_commit=_E0040_CAPTURE_GIT_COMMIT,
    )


def _project_path(project_root: Path, value: str | Path, name: str) -> Path:
    raw = Path(value)
    if not project_root.is_absolute():
        raise _fail(f"trusted project root for {name} must be absolute")
    if ".." in raw.parts:
        raise _fail(f"{name} contains parent traversal")
    path = raw if raw.is_absolute() else project_root / raw
    try:
        path.relative_to(project_root)
    except ValueError as exc:
        raise _fail(f"{name} escapes the project root", exc) from exc
    return path


def _directory_open_flags() -> int:
    if not hasattr(os, "O_DIRECTORY") or not hasattr(os, "O_NOFOLLOW"):
        raise _fail("secure directory-descriptor operations are unavailable")
    return os.O_RDONLY | os.O_CLOEXEC | os.O_DIRECTORY | os.O_NOFOLLOW


def _open_trusted_root(project_root: Path, name: str) -> tuple[int, tuple[int, int, int]]:
    if not project_root.is_absolute():
        raise _fail(f"trusted project root for {name} must be absolute")
    try:
        descriptor = os.open(project_root, _directory_open_flags())
    except OSError as exc:
        raise _fail(f"cannot open trusted project root for {name}", exc) from exc
    opened = os.fstat(descriptor)
    if not stat.S_ISDIR(opened.st_mode):
        os.close(descriptor)
        raise _fail(f"trusted project root for {name} is not a directory")
    return descriptor, (opened.st_dev, opened.st_ino, opened.st_mode)


def _relative_under_root(project_root: Path, path: Path, name: str) -> Path:
    if not project_root.is_absolute() or not path.is_absolute():
        raise _fail(f"{name} requires absolute trusted paths")
    try:
        relative = path.relative_to(project_root)
    except ValueError as exc:
        raise _fail(f"{name} escapes the trusted project root", exc) from exc
    if not relative.parts or ".." in relative.parts:
        raise _fail(f"{name} has an unsafe project-relative path")
    return relative


def _directory_identity(value: os.stat_result) -> tuple[int, int, int]:
    return value.st_dev, value.st_ino, value.st_mode


def _file_identity(value: os.stat_result) -> tuple[int, int, int, int, int, int]:
    return (
        value.st_dev,
        value.st_ino,
        value.st_mode,
        value.st_size,
        value.st_mtime_ns,
        value.st_ctime_ns,
    )


def _open_directory_chain(
    root_descriptor: int,
    parts: Sequence[str],
    name: str,
    *,
    create: bool,
) -> tuple[int, tuple[tuple[int, int, int], ...]]:
    current = os.dup(root_descriptor)
    identities: list[tuple[int, int, int]] = []
    try:
        for part in parts:
            if not part or part in {".", ".."} or "/" in part:
                raise _fail(f"unsafe directory component for {name}")
            try:
                following = os.open(part, _directory_open_flags(), dir_fd=current)
            except OSError as exc:
                if not create or exc.errno != errno.ENOENT:
                    raise _fail(f"cannot traverse trusted directory chain for {name}", exc) from exc
                try:
                    os.mkdir(part, 0o755, dir_fd=current)
                    os.fsync(current)
                    following = os.open(part, _directory_open_flags(), dir_fd=current)
                except OSError as create_exc:
                    raise _fail(
                        f"cannot create trusted directory component for {name}",
                        create_exc,
                    ) from create_exc
            os.close(current)
            current = following
            opened = os.fstat(current)
            if not stat.S_ISDIR(opened.st_mode):
                raise _fail(f"trusted directory component for {name} is not a directory")
            identities.append(_directory_identity(opened))
        return current, tuple(identities)
    except BaseException:
        os.close(current)
        raise


def _open_fresh_directory_chain(
    project_root: Path,
    expected_root_identity: tuple[int, int, int],
    parts: Sequence[str],
    name: str,
) -> tuple[int, int, tuple[tuple[int, int, int], ...]]:
    fresh_root, fresh_root_identity = _open_trusted_root(project_root, name)
    if fresh_root_identity != expected_root_identity:
        os.close(fresh_root)
        raise _fail(f"trusted project root changed during {name}")
    try:
        directory, chain = _open_directory_chain(
            fresh_root,
            parts,
            name,
            create=False,
        )
    except BaseException:
        os.close(fresh_root)
        raise
    return fresh_root, directory, chain


def _stable_read(
    project_root: Path,
    path: Path,
    *,
    maximum_size: int,
    name: str,
) -> bytes:
    relative = _relative_under_root(project_root, path, name)
    root, root_identity = _open_trusted_root(project_root, name)
    parent: int | None = None
    descriptor: int | None = None
    try:
        parent, parent_chain = _open_directory_chain(
            root,
            relative.parts[:-1],
            name,
            create=False,
        )
        flags = (
            os.O_RDONLY
            | os.O_CLOEXEC
            | os.O_NOFOLLOW
            | getattr(os, "O_NONBLOCK", 0)
            | getattr(os, "O_BINARY", 0)
        )
        try:
            descriptor = os.open(relative.parts[-1], flags, dir_fd=parent)
        except OSError as exc:
            raise _fail(f"cannot open trusted file for {name}", exc) from exc
        opened = os.fstat(descriptor)
        if not stat.S_ISREG(opened.st_mode):
            raise _fail(f"{name} must be a regular non-symlink file")
        if opened.st_size <= 0 or opened.st_size > maximum_size:
            raise _fail(f"{name} violates its size budget")
        chunks: list[bytes] = []
        remaining = opened.st_size
        while remaining:
            chunk = os.read(descriptor, min(remaining, 1024 * 1024))
            if not chunk:
                raise _fail(f"short read from {name}")
            chunks.append(chunk)
            remaining -= len(chunk)
        growth = os.read(descriptor, 1)
        after = os.fstat(descriptor)
        linked = os.stat(relative.parts[-1], dir_fd=parent, follow_symlinks=False)
        identity = _file_identity(opened)
        if identity != _file_identity(after) or identity != _file_identity(linked) or growth:
            raise _fail(f"{name} changed during descriptor read")
        os.close(descriptor)
        descriptor = None
        os.close(parent)
        parent = None

        fresh_root, fresh_parent, fresh_chain = _open_fresh_directory_chain(
            project_root,
            root_identity,
            relative.parts[:-1],
            f"{name} post-read revalidation",
        )
        try:
            fresh_file = os.stat(
                relative.parts[-1],
                dir_fd=fresh_parent,
                follow_symlinks=False,
            )
        except OSError as exc:
            raise _fail(f"cannot revalidate canonical file for {name}", exc) from exc
        finally:
            os.close(fresh_parent)
            os.close(fresh_root)
        if fresh_chain != parent_chain or _file_identity(fresh_file) != identity:
            raise _fail(f"canonical parent/file identity changed after reading {name}")
        return b"".join(chunks)
    finally:
        if descriptor is not None:
            os.close(descriptor)
        if parent is not None:
            os.close(parent)
        os.close(root)


def _read_verified_artifact(
    project_root: Path,
    record: Mapping[str, Any],
    *,
    maximum_size: int,
    name: str,
) -> tuple[Path, bytes]:
    normalized = _artifact_record(dict(record), name)
    path = _project_path(project_root, normalized["path"], name)
    payload = _stable_read(project_root, path, maximum_size=maximum_size, name=name)
    if len(payload) != normalized["size_bytes"] or _sha256_bytes(payload) != normalized["sha256"]:
        raise _fail(f"{name} byte identity drifted")
    return path, payload


def _load_control(project_root: Path, config_path: Path) -> tuple[dict[str, Any], Path, bytes]:
    path = _project_path(project_root, config_path, "E-0041 control")
    payload = _stable_read(
        project_root,
        path,
        maximum_size=128 * 1024,
        name="E-0041 control",
    )
    try:
        control = yaml.safe_load(payload.decode("utf-8"))
    except (UnicodeDecodeError, yaml.YAMLError) as exc:
        raise _fail("cannot decode E-0041 control", exc) from exc
    if not isinstance(control, dict):
        raise _fail("E-0041 control must be an object")
    if (
        control.get("version") != 1
        or control.get("experiment_id") != "E-0041"
        or control.get("dataset_role") != "CALIBRATION"
        or control.get("state") != "MECHANISM_READY_FORMAL_CAPTURE_BLOCKED_PENDING_E0040_SEAL"
    ):
        raise _fail("E-0041 control identity drifted")
    counts = _mapping(control.get("fixed_cardinality"), "fixed cardinality")
    if counts != {
        "source_row_count": 64,
        "physical_cell_count": 128,
        "cells_per_source_row": 2,
        "schema_row_count": 77,
        "current_or_comparative": ["CURRENT", "COMPARATIVE"],
    }:
        raise _fail("E-0041 fixed cardinality drifted")
    status_contract = _mapping(control.get("cell_status_contract"), "cell status contract")
    if set(status_contract.get("statuses", [])) != _FINAL_STATUSES:
        raise _fail("E-0041 final status vocabulary drifted")
    if set(status_contract.get("strict_numeric_arithmetic_operands", [])) != (
        _STRICT_NUMERIC_STATUSES
    ):
        raise _fail("E-0041 strict arithmetic status contract drifted")
    return cast(dict[str, Any], control), path, payload


def _clean_git_commit(project_root: Path) -> str:
    status_result = subprocess.run(
        ["git", "-c", "status.showUntrackedFiles=all", "status", "--porcelain=v1"],
        cwd=project_root,
        check=False,
        capture_output=True,
        text=True,
        env={"PATH": os.environ.get("PATH", "")},
    )
    if status_result.returncode != 0 or status_result.stdout:
        raise _fail("formal E-0041 capture requires a clean Git worktree")
    head_result = subprocess.run(
        ["git", "rev-parse", "--verify", "HEAD^{commit}"],
        cwd=project_root,
        check=False,
        capture_output=True,
        text=True,
        env={"PATH": os.environ.get("PATH", "")},
    )
    commit = head_result.stdout.strip()
    if head_result.returncode != 0 or re.fullmatch(r"[0-9a-f]{40}", commit) is None:
        raise _fail("cannot resolve the E-0041 capture commit")
    return commit


def _row_coordinates(row_id: str) -> tuple[int, int]:
    match = _ROW_ID.fullmatch(row_id)
    if match is None:
        raise _fail(f"invalid row_id: {row_id}")
    return int(match.group("page")), int(match.group("row"))


def _cell_coordinates(cell_id: str) -> tuple[int, int, int]:
    match = _CELL_ID.fullmatch(cell_id)
    if match is None:
        raise _fail(f"invalid cell_id: {cell_id}")
    return (
        int(match.group("page")),
        int(match.group("row")),
        int(match.group("axis")) - 1,
    )


def _unique_by_key(
    records: Sequence[object],
    *,
    key: str,
    name: str,
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for index, raw in enumerate(records):
        record = _mapping(raw, f"{name}[{index}]")
        value = record.get(key)
        if not isinstance(value, str) or not value:
            raise _fail(f"{name}[{index}] has no {key}")
        if value in result:
            raise _fail(f"duplicate {key} in {name}: {value}")
        result[value] = record
    return result


def _validate_postjoin(
    payload: Mapping[str, Any],
    *,
    row_count: int,
    cell_count: int,
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]]]:
    if (
        payload.get("format_version") != 1
        or payload.get("experiment_id") != "E-0037"
        or payload.get("dataset_role") != "CALIBRATION"
        or payload.get("state") != "SEALED_MAPPING_POSTJOIN_ASSEMBLY_COMPLETE"
    ):
        raise _fail("E-0037 postjoin identity drifted")
    access = _mapping(payload.get("access_order"), "E-0037 access order")
    if (
        access.get("mapping_only_seal_validated_before_postjoin_open") is not True
        or access.get("mapping_only_hash_validated_before_postjoin_open") is not True
        or access.get("mapper_invocation_count") != 0
        or access.get("mapping_result_repaired_or_rerun") is not False
        or access.get("review_or_history_opened") is not False
    ):
        raise _fail("E-0037 postjoin access isolation drifted")
    rows = _sequence(payload.get("rows"), "E-0037 rows")
    cells = _sequence(payload.get("cells"), "E-0037 cells")
    if len(rows) != row_count or len(cells) != cell_count:
        raise _fail("E-0037 row/cell cardinality drifted")
    rows_by_id = _unique_by_key(rows, key="row_id", name="E-0037 rows")
    cells_by_id = _unique_by_key(cells, key="cell_id", name="E-0037 cells")
    per_row: Counter[str] = Counter()
    axes_per_row: dict[str, set[int]] = {}
    for row_id, row in rows_by_id.items():
        page, ordinal = _row_coordinates(row_id)
        if row.get("page") != page or row.get("row_ordinal") != ordinal:
            raise _fail(f"E-0037 row coordinates drifted: {row_id}")
        proposals = _mapping(row.get("semantic_proposals"), f"semantic proposals {row_id}")
        if not isinstance(proposals.get("ppocrv6_source"), str):
            raise _fail(f"E-0037 row lacks source-label provenance: {row_id}")
        # Deliberately do not inspect row["mapping"]. It is the superseded E-0037
        # decision and is outside this consumer's authority.
    for cell_id, cell in cells_by_id.items():
        page, ordinal, axis = _cell_coordinates(cell_id)
        row_id = f"page-{page:04d}-row-{ordinal:03d}-label"
        if (
            cell.get("row_id") != row_id
            or cell.get("page") != page
            or cell.get("row_ordinal") != ordinal
            or cell.get("axis_ordinal") != axis
            or row_id not in rows_by_id
        ):
            raise _fail(f"E-0037 cell coordinates drifted: {cell_id}")
        per_row[row_id] += 1
        axes_per_row.setdefault(row_id, set()).add(axis)
    if any(per_row[row_id] != 2 or axes_per_row[row_id] != {0, 1} for row_id in rows_by_id):
        raise _fail("E-0037 does not retain exactly two physical axes per row")
    nodes = _sequence(
        _mapping(payload.get("schema_projection"), "E-0037 schema projection").get("nodes"),
        "E-0037 schema nodes",
    )
    return (
        rows_by_id,
        cells_by_id,
        [_mapping(node, f"E-0037 schema node {index}") for index, node in enumerate(nodes)],
    )


def _validate_mapper_result(
    result: Mapping[str, Any],
    *,
    expected_row_ids: set[str],
    expected_schema_ids: set[int],
) -> tuple[dict[str, dict[str, Any]], dict[int, dict[str, Any]]]:
    if result.get("status") != "RESOLVED" or result.get("automatic_selection_allowed") is not True:
        raise _fail("mapping challenger result is not resolved")
    row_records = _sequence(result.get("row_mappings"), "challenger row mappings")
    rows_by_id = _unique_by_key(row_records, key="row_id", name="challenger row mappings")
    if set(rows_by_id) != expected_row_ids:
        raise _fail("postjoin/challenger row_id sets differ")
    selected_ids: dict[int, str] = {}
    for row_id, row in rows_by_id.items():
        status_value = row.get("status")
        selected = row.get("selected_report_norm_id")
        candidates = row.get("candidate_report_norm_ids")
        if not isinstance(candidates, list) or any(type(item) is not int for item in candidates):
            raise _fail(f"challenger candidates are invalid: {row_id}")
        if len(candidates) != len(set(candidates)) or any(
            item not in expected_schema_ids for item in candidates
        ):
            raise _fail(f"challenger candidates drifted: {row_id}")
        if status_value in _SELECTED_MAPPING_STATUSES:
            if type(selected) is not int or selected not in candidates:
                raise _fail(f"selected challenger row is malformed: {row_id}")
            if selected in selected_ids:
                raise _fail(
                    f"duplicate selected ReportNormId {selected}: "
                    f"{selected_ids[selected]} and {row_id}"
                )
            selected_ids[selected] = row_id
        elif status_value in _UNSELECTED_MAPPING_STATUSES:
            if selected is not None:
                raise _fail(f"unselected challenger row carries a target: {row_id}")
        elif status_value in _SOURCE_ONLY_MAPPING_STATUSES:
            if selected is not None or candidates:
                raise _fail(f"source-only challenger row carries schema authority: {row_id}")
        else:
            raise _fail(f"unsupported challenger row status: {status_value}")

    disposition_records = _sequence(
        result.get("schema_dispositions"), "challenger schema dispositions"
    )
    dispositions: dict[int, dict[str, Any]] = {}
    for index, raw in enumerate(disposition_records):
        disposition = _mapping(raw, f"schema disposition {index}")
        report_norm_id = disposition.get("report_norm_id")
        if type(report_norm_id) is not int or report_norm_id in dispositions:
            raise _fail("duplicate or invalid schema disposition ReportNormId")
        dispositions[report_norm_id] = disposition
    if set(dispositions) != expected_schema_ids:
        raise _fail("challenger schema disposition/template identity differs")
    for report_norm_id, disposition in dispositions.items():
        status_value = disposition.get("status")
        selected_row_id = disposition.get("selected_row_id")
        candidates = disposition.get("candidate_row_ids")
        if not isinstance(candidates, list) or any(
            not isinstance(item, str) or item not in expected_row_ids for item in candidates
        ):
            raise _fail(f"invalid schema candidate-row set: {report_norm_id}")
        if status_value == "MAPPED":
            if (
                not isinstance(selected_row_id, str)
                or selected_row_id not in rows_by_id
                or rows_by_id[selected_row_id].get("selected_report_norm_id") != report_norm_id
            ):
                raise _fail(f"schema/row selection parity failed: {report_norm_id}")
        elif status_value in {
            "UNMATCHED_SCHEMA_NODE",
            "UNMATCHED_SCHEMA_NODE_WITH_SKIPPED_CANDIDATES",
            "NOT_OBSERVED",
        }:
            if selected_row_id is not None:
                raise _fail(
                    f"unmatched schema disposition carries a selected row: {report_norm_id}"
                )
        else:
            raise _fail(f"unsupported schema disposition status: {status_value}")

    return rows_by_id, dispositions


def _validate_e0038_mapping_challenger(
    payload: Mapping[str, Any],
    *,
    expected_row_ids: set[str],
    expected_schema_ids: set[int],
) -> tuple[
    dict[str, dict[str, Any]],
    dict[int, dict[str, Any]],
    set[int],
    dict[str, Any],
]:
    bundle = _mapping(payload.get("exact_mapping_bundle"), "mapping challenger bundle")
    receipt = _mapping(bundle.get("alias_overlay_receipt"), "alias overlay receipt")
    exact = _mapping(bundle.get("exact_search"), "exact search")
    result = _mapping(
        exact.get("mapping_result_without_internal_alias_authority"),
        "mapping challenger result",
    )
    if exact.get("status") != "EXACT_SEARCH_COMPLETE":
        raise _fail("mapping challenger exact search is incomplete")
    rows_by_id, dispositions = _validate_mapper_result(
        result,
        expected_row_ids=expected_row_ids,
        expected_schema_ids=expected_schema_ids,
    )
    changed = receipt.get("changed_report_norm_ids")
    if not isinstance(changed, list) or any(type(item) is not int for item in changed):
        raise _fail("alias receipt changed-ID set is invalid")
    unapproved_alias_ids = (
        set(changed) if receipt.get("review_or_steward_approved") is not True else set()
    )
    if not unapproved_alias_ids.issubset(expected_schema_ids):
        raise _fail("alias receipt names an out-of-template ID")
    authority = {
        "source": "E0038_EXACT_MAPPING_CHALLENGER",
        "alias_overlay_receipt": copy.deepcopy(receipt),
    }
    return rows_by_id, dispositions, unapproved_alias_ids, authority


def _row_target_pairs(value: object, name: str) -> tuple[tuple[str, int], ...]:
    records = _sequence(value, name)
    result: list[tuple[str, int]] = []
    for index, raw in enumerate(records):
        if (
            not isinstance(raw, list)
            or len(raw) != 2
            or not isinstance(raw[0], str)
            or type(raw[1]) is not int
        ):
            raise _fail(f"{name}[{index}] is not a row/target pair")
        result.append((raw[0], raw[1]))
    if len(result) != len(set(result)):
        raise _fail(f"{name} contains duplicate pairs")
    return tuple(result)


def _e0040_challenger_json(
    challenger: E0040ChallengerResult | AuthenticatedE0040ResultCarrier,
) -> tuple[dict[str, Any], dict[str, dict[str, Any]] | None]:
    if type(challenger) is E0040ChallengerResult:
        live = cast(E0040ChallengerResult, challenger)
        return _load_json_bytes(_canonical_bytes(live.to_dict()), "E-0040 challenger"), None
    if type(challenger) is not AuthenticatedE0040ResultCarrier:
        raise _fail("E-0040 mapping authority must be a direct result or authenticated carrier")
    carrier = cast(AuthenticatedE0040ResultCarrier, challenger)
    challenger_bytes = _authenticate_e0040_artifact_chain(
        mapping_bytes=carrier.mapping_bytes,
        seal_bytes=carrier.seal_bytes,
        registration_bytes=carrier.registration_bytes,
    )
    if any(
        type(record) is not E0040ArtifactRecord
        for record in (
            carrier.mapping_artifact,
            carrier.seal_artifact,
            carrier.registration_artifact,
        )
    ):
        raise _fail("E-0040 authenticated result carrier has mutable artifact records")
    artifacts = {
        "mapping": carrier.mapping_artifact.to_dict(),
        "seal": carrier.seal_artifact.to_dict(),
        "registration": carrier.registration_artifact.to_dict(),
    }
    if (
        artifacts["mapping"] != _E0040_MAPPING_ARTIFACT
        or artifacts["seal"] != _E0040_MAPPING_SEAL_ARTIFACT
        or artifacts["registration"] != _E0040_S3_REGISTRATION_ARTIFACT
        or carrier.challenger_result_sha256 != _E0040_CHALLENGER_RESULT_SHA256
        or carrier.capture_git_commit != _E0040_CAPTURE_GIT_COMMIT
    ):
        raise _fail("E-0040 authenticated result carrier identity drifted")
    normalized = _load_json_bytes(challenger_bytes, "authenticated E-0040 challenger")
    canonical = _canonical_bytes(normalized)
    if (
        canonical != challenger_bytes
        or len(canonical) != _E0040_CHALLENGER_RESULT_SIZE_BYTES
        or _sha256_bytes(canonical) != carrier.challenger_result_sha256
    ):
        raise _fail("E-0040 authenticated challenger result identity drifted")
    return normalized, artifacts


def _validate_e0040_mapping_challenger(
    challenger: E0040ChallengerResult | AuthenticatedE0040ResultCarrier,
    *,
    expected_row_ids: set[str],
    expected_schema_ids: set[int],
) -> tuple[
    dict[str, dict[str, Any]],
    dict[int, dict[str, Any]],
    set[int],
    dict[str, Any],
]:
    # A live frozen result proves direct invocation. The JSON-native alternative
    # is accepted only after the exact mapping/seal/registration byte chain has
    # minted its carrier; a detached raw mapping remains outside this boundary.
    normalized, authenticated_artifacts = _e0040_challenger_json(challenger)
    required = {
        "policy_sha256",
        "mapper_policy_sha256",
        "mapper_invocation_count",
        "normalization",
        "collision_audit",
        "combined_parent_overrides",
        "source_only_structural_rows",
        "baseline_selected_pairs",
        "final_selected_pairs",
        "newly_selected_pairs",
        "baseline_result",
        "final_result",
    }
    if set(normalized) != required:
        raise _fail("E-0040 challenger is detached from its authority receipts")
    policy_sha = normalized.get("policy_sha256")
    mapper_sha = normalized.get("mapper_policy_sha256")
    if (
        policy_sha != _E0040_POLICY_SHA256
        or mapper_sha != _E0040_MAPPER_POLICY_SHA256
        or normalized.get("mapper_invocation_count") != 2
    ):
        raise _fail("E-0040 policy or mapper receipt is invalid")

    receipt = _mapping(normalized.get("normalization"), "E-0040 normalization receipt")
    collision = _mapping(normalized.get("collision_audit"), "E-0040 collision audit")
    if (
        receipt.get("statement_type") != "CDKT"
        or receipt.get("bank_scope") != "ALL_BANKS"
        or receipt.get("base_projection_sha256") != _E0040_BASE_PROJECTION_SHA256
        or receipt.get("result_projection_sha256") != _E0040_RESULT_PROJECTION_SHA256
        or receipt.get("id_scoped_alias_invocation_count") != 0
        or receipt.get("bank_page_or_row_rule_invocation_count") != 0
        or receipt.get("input_alias_authority") != "CANONICAL_AND_STRUCTURAL_ALIASES_ONLY"
        or receipt.get("mapper_carrier_alias_authority") != "CANONICAL_AND_STRUCTURAL_ALIASES_ONLY"
        or type(receipt.get("changed_schema_node_count")) is not int
        or receipt.get("changed_schema_node_count", 0) <= 0
        or type(receipt.get("derived_key_count")) is not int
        or receipt.get("derived_key_count", 0) <= 0
        or collision.get("statement_type") != "CDKT"
        or collision.get("node_count") != len(expected_schema_ids)
        or collision.get("new_collision_pairs") != []
    ):
        raise _fail("E-0040 generic normalization authority is unsafe")
    for hash_field in ("base_projection_sha256", "result_projection_sha256"):
        digest = receipt.get(hash_field)
        if not isinstance(digest, str) or _SHA256.fullmatch(digest) is None:
            raise _fail(f"E-0040 normalization {hash_field} is invalid")

    result = _mapping(normalized.get("final_result"), "E-0040 final mapper result")
    if (
        result.get("schema_projection_sha256") != receipt.get("result_projection_sha256")
        or result.get("schema_alias_authority") != receipt.get("mapper_carrier_alias_authority")
        or result.get("policy_sha256") != mapper_sha
    ):
        raise _fail("E-0040 final result is detached from its receipts")
    search = _mapping(result.get("search"), "E-0040 final search receipt")
    if any(
        search.get(field) != 0
        for field in (
            "pruned_states",
            "main_search_pruned_states",
            "counterfactual_search_pruned_states",
        )
    ):
        raise _fail("E-0040 final search was pruned")
    intervals = _sequence(result.get("intervals"), "E-0040 final intervals")
    for index, raw in enumerate(intervals):
        interval = _mapping(raw, f"E-0040 final interval {index}")
        if (
            interval.get("search_exhaustive") is not True
            or interval.get("automatic_selection_allowed") is not True
            or interval.get("main_search_pruned_states") != 0
            or interval.get("counterfactual_search_pruned_states") != 0
        ):
            raise _fail("E-0040 final interval lacks exhaustive selection authority")

    rows_by_id, dispositions = _validate_mapper_result(
        result,
        expected_row_ids=expected_row_ids,
        expected_schema_ids=expected_schema_ids,
    )
    final_pairs = _row_target_pairs(
        normalized.get("final_selected_pairs"), "E-0040 final selected pairs"
    )
    selected_from_result = tuple(
        (row_id, cast(int, row["selected_report_norm_id"]))
        for row_id, row in rows_by_id.items()
        if row.get("selected_report_norm_id") is not None
    )
    if final_pairs != selected_from_result or len(final_pairs) != 61:
        raise _fail("E-0040 final 61-row selection receipt differs from mapper output")
    selected_status_counts = Counter(
        row["status"]
        for row in rows_by_id.values()
        if row.get("selected_report_norm_id") is not None
    )
    if selected_status_counts != {"RESOLVED_ANCHOR": 43, "RESOLVED_PATH": 18}:
        raise _fail("E-0040 selected anchor/path cardinality drifted")

    anchors = _sequence(result.get("anchors"), "E-0040 final anchors")
    selected_anchors: dict[str, dict[str, Any]] = {}
    for index, raw in enumerate(anchors):
        anchor = _mapping(raw, f"E-0040 final anchor {index}")
        selected = anchor.get("selected_report_norm_id")
        if selected is None:
            continue
        row_id = anchor.get("row_id")
        margin = anchor.get("counterfactual_margin")
        if (
            not isinstance(row_id, str)
            or row_id in selected_anchors
            or rows_by_id.get(row_id, {}).get("status") != "RESOLVED_ANCHOR"
            or rows_by_id[row_id].get("selected_report_norm_id") != selected
            or anchor.get("constraint_report_norm_id") != selected
            or anchor.get("selection_allowed") is not True
            or type(margin) not in {int, float}
            or margin < 0.15
        ):
            raise _fail("E-0040 selected anchor lacks a decisive counterfactual gate")
        selected_anchors[row_id] = anchor
    if len(selected_anchors) != 43:
        raise _fail("E-0040 selected anchor receipt count drifted")

    path_counterfactuals: dict[str, dict[str, Any]] = {}
    best_path_rows: set[str] = set()
    for index, interval in enumerate(intervals):
        best_path = _mapping(interval.get("best_path"), f"E-0040 interval {index} best path")
        for match_index, raw in enumerate(
            _sequence(best_path.get("matches"), f"E-0040 interval {index} matches")
        ):
            match = _mapping(raw, f"E-0040 interval {index} match {match_index}")
            row_id = match.get("row_id")
            if not isinstance(row_id, str) or row_id in best_path_rows:
                raise _fail("E-0040 best-path row receipt is invalid")
            best_path_rows.add(row_id)
        for cf_index, raw in enumerate(
            _sequence(
                interval.get("counterfactuals"),
                f"E-0040 interval {index} counterfactuals",
            )
        ):
            counterfactual = _mapping(raw, f"E-0040 interval {index} counterfactual {cf_index}")
            row_id = counterfactual.get("row_id")
            selected = counterfactual.get("selected_report_norm_id")
            margin = counterfactual.get("exclusion_margin")
            if (
                not isinstance(row_id, str)
                or row_id in path_counterfactuals
                or rows_by_id.get(row_id, {}).get("status") != "RESOLVED_PATH"
                or rows_by_id[row_id].get("selected_report_norm_id") != selected
                or counterfactual.get("stable") is not True
                or type(margin) not in {int, float}
                or margin < 0.15
            ):
                raise _fail("E-0040 selected path lacks a decisive counterfactual gate")
            path_counterfactuals[row_id] = counterfactual
    selected_path_ids = {
        row_id for row_id, row in rows_by_id.items() if row.get("status") == "RESOLVED_PATH"
    }
    if (
        len(path_counterfactuals) != 18
        or set(path_counterfactuals) != selected_path_ids
        or best_path_rows != selected_path_ids
    ):
        raise _fail("E-0040 selected path/counterfactual receipt coverage drifted")
    baseline_pairs = _row_target_pairs(
        normalized.get("baseline_selected_pairs"), "E-0040 baseline selected pairs"
    )
    new_pairs = _row_target_pairs(
        normalized.get("newly_selected_pairs"), "E-0040 newly selected pairs"
    )
    if (
        len(baseline_pairs) != 59
        or len(new_pairs) != 2
        or not set(baseline_pairs).issubset(final_pairs)
        or tuple(item for item in final_pairs if item not in set(baseline_pairs)) != new_pairs
    ):
        raise _fail("E-0040 baseline/new selection receipts are inconsistent")
    overrides = _sequence(
        normalized.get("combined_parent_overrides"), "E-0040 combined-parent overrides"
    )
    override_pairs = tuple(
        (
            cast(str, _mapping(raw, f"E-0040 override {index}")["row_id"]),
            cast(int, _mapping(raw, f"E-0040 override {index}")["target_report_norm_id"]),
        )
        for index, raw in enumerate(overrides)
    )
    if override_pairs != new_pairs:
        raise _fail("E-0040 combined-parent proof does not bind new selections")

    source_only_records = _sequence(
        normalized.get("source_only_structural_rows"), "E-0040 source-only rows"
    )
    source_only_by_id = _unique_by_key(
        source_only_records,
        key="row_id",
        name="E-0040 source-only rows",
    )
    if len(source_only_by_id) != 3 or set(source_only_by_id) != (
        expected_row_ids - {row_id for row_id, _ in final_pairs}
    ):
        raise _fail("E-0040 does not prove the exact 61 mapped / 3 source-only partition")
    if Counter(item.get("observed_role") for item in source_only_by_id.values()) != {
        "SECTION": 2,
        "TOTAL": 1,
    }:
        raise _fail("E-0040 source-only structural-role receipt drifted")
    normalized_rows = copy.deepcopy(rows_by_id)
    for row_id, source_only in source_only_by_id.items():
        row = normalized_rows[row_id]
        if (
            source_only.get("disposition") != "SOURCE_ONLY_STRUCTURAL_ROW_HYPOTHESIS_UNMATCHED"
            or source_only.get("selected_report_norm_id") is not None
            or source_only.get("final_mapping_status") != row.get("status")
            or row.get("selected_report_norm_id") is not None
            or row.get("candidate_report_norm_ids")
        ):
            raise _fail(f"E-0040 source-only row is not fail-closed: {row_id}")
        row["status"] = "SOURCE_ONLY_STRUCTURAL_ROW"
        row["reason"] = f"{row.get('reason')}; SOURCE_ONLY_STRUCTURAL_ROW_HYPOTHESIS_UNMATCHED"

    authority = {
        "source": "E0040_GENERIC_CALIBRATION_CHALLENGER",
        "policy_sha256": policy_sha,
        "mapper_policy_sha256": mapper_sha,
        "normalization": copy.deepcopy(receipt),
        "collision_audit_sha256": _canonical_sha256(collision),
        "final_selected_pair_count": 61,
        "source_only_row_count": 3,
        "id_scoped_alias_report_norm_ids": [],
        "challenger_result_sha256": _canonical_sha256(normalized),
    }
    if authenticated_artifacts is not None:
        authority["authenticated_formal_artifacts"] = copy.deepcopy(authenticated_artifacts)
        authority["capture_git_commit"] = cast(
            AuthenticatedE0040ResultCarrier,
            challenger,
        ).capture_git_commit
    return normalized_rows, dispositions, set(), authority


def _validate_mapping_challenger(
    payload: object,
    *,
    expected_row_ids: set[str],
    expected_schema_ids: set[int],
) -> tuple[
    dict[str, dict[str, Any]],
    dict[int, dict[str, Any]],
    set[int],
    dict[str, Any],
]:
    if type(payload) in {E0040ChallengerResult, AuthenticatedE0040ResultCarrier}:
        return _validate_e0040_mapping_challenger(
            payload,
            expected_row_ids=expected_row_ids,
            expected_schema_ids=expected_schema_ids,
        )
    if not isinstance(payload, Mapping):
        raise _fail("mapping challenger must be an authenticated result carrier")
    if "exact_mapping_bundle" in payload:
        return _validate_e0038_mapping_challenger(
            payload,
            expected_row_ids=expected_row_ids,
            expected_schema_ids=expected_schema_ids,
        )
    if "final_result" in payload:
        raise _fail("deserialized E-0040 mappings lack direct-call authority")
    raise _fail("mapping challenger lacks a recognized authority envelope")


def _template_snapshot(workbook: Any, sheet_name: str, maximum_row: int) -> dict[str, Any]:
    if sheet_name not in workbook.sheetnames:
        raise _fail(f"template sheet is absent: {sheet_name}")
    sheet = workbook[sheet_name]
    cells: list[dict[str, Any]] = []
    for row in range(1, maximum_row + 1):
        for column in range(1, 4):
            cell = sheet.cell(row=row, column=column)
            cells.append(
                {
                    "coordinate": cell.coordinate,
                    "value": cell.value,
                    "data_type": cell.data_type,
                    # openpyxl may materialize an all-zero default StyleArray
                    # when unrelated cells are edited. None and that array are
                    # the same Excel style, so normalize both before fidelity checks.
                    "style_array": list(cell._style if cell._style is not None else StyleArray()),
                    "number_format": cell.number_format,
                    "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
                    "comment": cell.comment.text if cell.comment else None,
                }
            )
    return {
        "sheet_name": sheet_name,
        "sheet_index": workbook.sheetnames.index(sheet_name),
        "cells": cells,
        "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
        "row_dimensions": {
            str(row): {
                "height": sheet.row_dimensions[row].height,
                "hidden": bool(sheet.row_dimensions[row].hidden),
            }
            for row in range(1, maximum_row + 1)
        },
        "column_dimensions": {
            column: {
                "width": sheet.column_dimensions[column].width,
                "hidden": bool(sheet.column_dimensions[column].hidden),
            }
            for column in ("A", "B", "C")
        },
    }


def _load_template(
    template_bytes: bytes,
    *,
    sheet_name: str,
    schema_row_count: int,
) -> tuple[Any, list[dict[str, Any]], dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(template_bytes), data_only=False)
    except Exception as exc:
        raise _fail("cannot open CDKT template workbook", exc) from exc
    if workbook.sheetnames != [sheet_name]:
        workbook.close()
        raise _fail("CDKT template workbook sheet identity drifted")
    sheet = workbook[sheet_name]
    if sheet.max_row != schema_row_count + 1 or sheet.max_column != 3:
        workbook.close()
        raise _fail("CDKT template dimensions drifted")
    if sheet["B1"].value != "ReportNormId" or sheet["C1"].value != "ReportNormName":
        workbook.close()
        raise _fail("CDKT template headers drifted")
    rows: list[dict[str, Any]] = []
    seen: set[int] = set()
    for excel_row in range(2, schema_row_count + 2):
        display_order = sheet.cell(excel_row, 1).value
        report_norm_id = sheet.cell(excel_row, 2).value
        name = sheet.cell(excel_row, 3).value
        if (
            display_order != excel_row - 2
            or type(report_norm_id) is not int
            or report_norm_id in seen
            or not isinstance(name, str)
            or not name
        ):
            workbook.close()
            raise _fail(f"invalid CDKT template row {excel_row}")
        seen.add(report_norm_id)
        rows.append(
            {
                "excel_row": excel_row,
                "display_order": display_order,
                "report_norm_id": report_norm_id,
                "report_norm_name": name,
            }
        )
    snapshot = _template_snapshot(workbook, sheet_name, schema_row_count + 1)
    return workbook, rows, snapshot


def _safe_registry_crop_path(registry_path: Path, value: object, project_root: Path) -> Path:
    if not isinstance(value, str) or not value:
        raise _fail("geometry registry crop_path is invalid")
    relative = PurePosixPath(value)
    if relative.is_absolute() or ".." in relative.parts:
        raise _fail("geometry registry crop_path is unsafe")
    path = registry_path.parent / Path(*relative.parts)
    try:
        path.relative_to(project_root)
    except ValueError as exc:
        raise _fail("geometry crop escapes project root", exc) from exc
    return path


def _normalize_authenticated_geometry_registry(
    registry: Mapping[str, Any],
) -> dict[str, Any]:
    """Copy the pinned registry while removing its legacy checkout root.

    Only path fields in E-0041's authenticated geometry/source authority are
    rewritten. Diagnostic ``source_image_path`` values are outside that
    authority and deliberately remain byte-for-byte unchanged.
    """

    def normalize(value: object, field_name: str | None = None) -> Any:
        if isinstance(value, dict):
            return {key: normalize(item, key) for key, item in value.items()}
        if isinstance(value, list):
            return [normalize(item, field_name) for item in value]
        if field_name not in _GEOMETRY_ROOTED_PATH_FIELDS:
            return copy.deepcopy(value)
        if not isinstance(value, str) or not value or "\\" in value:
            raise _fail(f"authenticated geometry {field_name} is not a POSIX path")
        path = PurePosixPath(value)
        if path.is_absolute():
            try:
                path = path.relative_to(_LEGACY_GEOMETRY_PROJECT_ROOT)
            except ValueError as exc:
                raise _fail(
                    f"authenticated geometry {field_name} has a foreign absolute prefix",
                    exc,
                ) from exc
        if not path.parts or path == PurePosixPath(".") or ".." in path.parts:
            raise _fail(f"authenticated geometry {field_name} is unsafe")
        return path.as_posix()

    if not isinstance(registry, dict):
        raise _fail("authenticated geometry registry must be a JSON object")
    normalized = cast(dict[str, Any], normalize(registry))
    if _canonical_sha256(normalized) != _E0041_NORMALIZED_GEOMETRY_REGISTRY_SHA256:
        raise _fail("normalized authenticated geometry registry identity drifted")
    return normalized


def _validate_geometry_registry(
    project_root: Path,
    registry_path: Path,
    registry: Mapping[str, Any],
    *,
    cells_by_id: Mapping[str, Mapping[str, Any]],
    expected_authority: str,
) -> dict[str, dict[str, Any]]:
    if (
        registry.get("format_version") != 2
        or registry.get("policy") != "FIXED_GRID_NUMERIC_CELL_CROPS_V2"
        or registry.get("geometry_authority") != "E0033_PP_OCRV6_FIXED_GRID"
        or expected_authority != "RECONSTRUCTED_GEOMETRY_WITH_CELL_SHA_PARITY"
    ):
        raise _fail("reconstructed geometry registry identity drifted")
    isolation = _mapping(registry.get("reference_isolation"), "geometry isolation")
    required_isolation = {
        "accounting_validation_invoked",
        "historical_or_mongodb_values_loaded",
        "human_review_loaded",
        "schema_mapping_invoked",
        "template_labels_or_report_norm_ids_loaded",
    }
    if set(isolation) != required_isolation or any(
        isolation.get(name) is not False for name in required_isolation
    ):
        raise _fail("geometry reconstruction used a forbidden authority")
    records = _sequence(registry.get("cells"), "geometry registry cells")
    source_geometry_by_id = _unique_by_key(records, key="cell_id", name="geometry registry cells")
    geometry_by_id = copy.deepcopy(source_geometry_by_id)
    if set(geometry_by_id) != set(cells_by_id):
        raise _fail("geometry/postjoin cell ID sets differ")
    verified_sources: dict[tuple[str, str], bool] = {}
    for cell_id, geometry in geometry_by_id.items():
        page, row, axis = _cell_coordinates(cell_id)
        if (
            geometry.get("page") != page
            or geometry.get("row_ordinal") != row
            or geometry.get("axis_ordinal") != axis
        ):
            raise _fail(f"geometry coordinates drifted: {cell_id}")
        bbox = geometry.get("crop_bbox")
        if (
            not isinstance(bbox, list)
            or len(bbox) != 4
            or any(type(value) not in {int, float} for value in bbox)
            or not (bbox[0] < bbox[2] and bbox[1] < bbox[3])
        ):
            raise _fail(f"geometry crop bbox is invalid: {cell_id}")
        source_row_ids = geometry.get("source_row_ids")
        if (
            not isinstance(source_row_ids, list)
            or not source_row_ids
            or any(not isinstance(value, str) or not value for value in source_row_ids)
        ):
            raise _fail(f"geometry source-row provenance is absent: {cell_id}")
        digest = geometry.get("crop_sha256")
        size = geometry.get("crop_size_bytes")
        if (
            not isinstance(digest, str)
            or _SHA256.fullmatch(digest) is None
            or type(size) is not int
            or size <= 0
        ):
            raise _fail(f"geometry crop identity is invalid: {cell_id}")
        numeric = _mapping(
            cells_by_id[cell_id].get("numeric_evidence"),
            f"E-0037 numeric evidence {cell_id}",
        )
        if numeric.get("cell_id") != cell_id or numeric.get("crop_sha256") != digest:
            raise _fail(f"reconstructed/original sealed crop hash parity failed: {cell_id}")
        crop_path = _safe_registry_crop_path(registry_path, geometry.get("crop_path"), project_root)
        crop_bytes = _stable_read(
            project_root,
            crop_path,
            maximum_size=16 * 1024 * 1024,
            name=f"crop {cell_id}",
        )
        if len(crop_bytes) != size or _sha256_bytes(crop_bytes) != digest:
            raise _fail(f"reconstructed crop bytes drifted: {cell_id}")
        for path_field, hash_field in (
            ("source_render_path", "source_render_sha256"),
            ("source_ocr_path", "source_ocr_sha256"),
        ):
            source_path_raw = geometry.get(path_field)
            source_digest = geometry.get(hash_field)
            if (
                not isinstance(source_path_raw, str)
                or not isinstance(source_digest, str)
                or _SHA256.fullmatch(source_digest) is None
            ):
                raise _fail(f"geometry source identity is invalid: {cell_id}/{path_field}")
            key = (source_path_raw, source_digest)
            if key not in verified_sources:
                source_path = _project_path(project_root, source_path_raw, path_field)
                source_bytes = _stable_read(
                    project_root,
                    source_path,
                    maximum_size=128 * 1024 * 1024,
                    name=f"geometry {path_field}",
                )
                if _sha256_bytes(source_bytes) != source_digest:
                    raise _fail(f"geometry source bytes drifted: {path_field}")
                verified_sources[key] = True
        geometry["_resolved_crop_path"] = crop_path.relative_to(project_root).as_posix()
    return geometry_by_id


def _decimal(value: object, name: str) -> Decimal:
    if not isinstance(value, str) or not value:
        raise _fail(f"{name} must be a decimal string")
    try:
        number = Decimal(value)
    except InvalidOperation as exc:
        raise _fail(f"{name} is not numeric", exc) from exc
    if not number.is_finite():
        raise _fail(f"{name} is not finite")
    return number


def _source_numeric_disposition(cell: Mapping[str, Any]) -> tuple[str, str | None, Decimal | None]:
    numeric = _mapping(cell.get("numeric_evidence"), "nested numeric evidence")
    primary = _mapping(numeric.get("primary"), "nested primary numeric evidence")
    if primary.get("raw_text") != cell.get("visible_raw_value"):
        raise _fail(f"visible/nested raw numeric text differs: {cell.get('cell_id')}")
    verification = cell.get("numeric_verification_status")
    observation = cell.get("source_observation")
    if verification == "VERIFIED_OBSERVED_VALUE":
        if observation not in {"VALUE", "ZERO"}:
            raise _fail("verified numeric value has incompatible observation")
        number = _decimal(numeric.get("normalized_numeric_value"), "normalized numeric value")
        return ("ZERO" if number == 0 else "VALUE", cast(str, primary["raw_text"]), number)
    if verification == "VERIFIED_OBSERVED_DASH":
        if observation != "DASH" or primary.get("raw_text") != "-":
            raise _fail("verified dash lost its visible dash evidence")
        return "DASH", "-", None
    if verification == "UNRESOLVED_BLANK_PENDING_ROW_SEMANTICS":
        if observation != "BLANK" or primary.get("raw_text") != "":
            raise _fail("blank numeric evidence drifted")
        return "BLANK", "", None
    return "UNRESOLVED", cast(str | None, primary.get("raw_text")), None


def _normalize_cell(
    cell: Mapping[str, Any],
    row: Mapping[str, Any],
    challenger: Mapping[str, Any],
    geometry: Mapping[str, Any],
    *,
    unapproved_alias_ids: set[int],
    registry_record: Mapping[str, Any],
    postjoin_record: Mapping[str, Any],
    mapping_record: Mapping[str, Any],
) -> dict[str, Any]:
    source_status, raw_value, evidence_value = _source_numeric_disposition(cell)
    mapping_status = challenger.get("status")
    proposed_id = challenger.get("selected_report_norm_id")
    candidate_ids = list(cast(list[int], challenger.get("candidate_report_norm_ids")))
    alias_dependent = type(proposed_id) is int and proposed_id in unapproved_alias_ids
    source_only = mapping_status in _SOURCE_ONLY_MAPPING_STATUSES
    if alias_dependent:
        final_status = "AMBIGUOUS"
        report_norm_id = None
        reason = "selected mapping depends on an unapproved ID-scoped alias"
    elif mapping_status in _SELECTED_MAPPING_STATUSES:
        final_status = source_status
        report_norm_id = cast(int, proposed_id)
        reason = "mapping selected independently before numeric/period postjoin"
    elif source_only:
        final_status = "UNRESOLVED"
        report_norm_id = None
        reason = "visible source-only row is retained but cannot populate the target template"
    elif candidate_ids or mapping_status == "BEST_PATH_SKIPPED":
        final_status = "AMBIGUOUS"
        report_norm_id = None
        reason = "mapping challenger retained candidates without selection authority"
    else:
        final_status = "UNRESOLVED"
        report_norm_id = None
        reason = "mapping challenger found no admissible target"
    if final_status not in _FINAL_STATUSES:
        raise _fail("derived an invalid final cell status")

    axis = _mapping(cell.get("period_axis"), "period axis")
    role = axis.get("current_or_comparative")
    if role not in {"CURRENT", "COMPARATIVE"}:
        raise _fail("period axis role is invalid")
    multiplier = _positive_int(axis.get("unit_multiplier"), "unit multiplier")
    if multiplier != cell.get("unit_multiplier"):
        raise _fail("axis/cell unit multiplier differs")
    evidence_canonical_value: int | None = None
    if source_status in _STRICT_NUMERIC_STATUSES:
        if evidence_value is None:
            raise _fail("strict source numeric status has no numeric evidence")
        canonical = evidence_value * Decimal(multiplier)
        if canonical != canonical.to_integral_value():
            raise _fail("canonical-unit evidence value is not integral")
        evidence_canonical_value = int(canonical)
        if abs(evidence_canonical_value) > _MAX_SAFE_EXCEL_INTEGER:
            raise _fail("canonical-unit evidence is not exactly representable in Excel")
    canonical_value: int | None = None
    displayed_value: str | None = None
    if report_norm_id is not None and final_status in _STRICT_NUMERIC_STATUSES:
        if evidence_value is None or evidence_canonical_value is None:
            raise _fail("strict numeric status has no numeric evidence")
        canonical_value = evidence_canonical_value
        displayed_value = format(evidence_value, "f")

    proposals = _mapping(row.get("semantic_proposals"), "source row semantic proposals")
    structure = _mapping(row.get("source_structure"), "source row structure")
    crop_path = geometry.get("_resolved_crop_path")
    if not isinstance(crop_path, str):
        raise _fail("validated crop path is absent")
    return {
        "cell_id": cell["cell_id"],
        "row_id": cell["row_id"],
        "page": cell["page"],
        "row_ordinal": cell["row_ordinal"],
        "axis_ordinal": cell["axis_ordinal"],
        "current_or_comparative": role,
        "period_start": axis.get("period_start"),
        "period_end": axis.get("period_end"),
        "period_type": axis.get("period_type"),
        "raw_period_header": axis.get("raw_period_header"),
        "canonical_unit": axis.get("canonical_unit"),
        "displayed_unit": axis.get("raw_unit_text"),
        "unit_multiplier": multiplier,
        "scope": structure.get("report_scope", cell.get("report_scope")),
        "source_row_ids": copy.deepcopy(geometry.get("source_row_ids")),
        "source_label": proposals.get("ppocrv6_source"),
        "row_role": structure.get("row_role"),
        "source_observation": cell.get("source_observation"),
        "numeric_verification_status": cell.get("numeric_verification_status"),
        "source_numeric_status": source_status,
        "visible_raw_value": raw_value,
        "evidence_displayed_value": (
            format(evidence_value, "f") if evidence_value is not None else None
        ),
        "evidence_canonical_value": evidence_canonical_value,
        "exported_displayed_value": displayed_value,
        "exported_canonical_value": canonical_value,
        "status": final_status,
        "mapping_status": mapping_status,
        "mapping_reason": challenger.get("reason"),
        "mapping_disposition_reason": reason,
        "report_norm_id": report_norm_id,
        "proposed_report_norm_id": proposed_id,
        "candidate_report_norm_ids": candidate_ids,
        "unapproved_alias_dependency": alias_dependent,
        "source_only": source_only,
        "crop": {
            "authority": "RECONSTRUCTED_GEOMETRY_WITH_CELL_SHA_PARITY",
            "registry": copy.deepcopy(dict(registry_record)),
            "path": crop_path,
            "bbox": copy.deepcopy(geometry.get("crop_bbox")),
            "sha256": geometry.get("crop_sha256"),
            "size_bytes": geometry.get("crop_size_bytes"),
            "source_render_path": geometry.get("source_render_path"),
            "source_render_sha256": geometry.get("source_render_sha256"),
            "source_ocr_path": geometry.get("source_ocr_path"),
            "source_ocr_sha256": geometry.get("source_ocr_sha256"),
            "value_line_indices": copy.deepcopy(geometry.get("value_line_indices")),
        },
        "numeric_evidence_sha256": _canonical_sha256(cell.get("numeric_evidence")),
        "source_artifact_refs": {
            "e0037_postjoin": copy.deepcopy(dict(postjoin_record)),
            "mapping_challenger": copy.deepcopy(dict(mapping_record)),
        },
        "validation_refs": [],
    }


def _schema_status_from_disposition(disposition: Mapping[str, Any], alias_ids: set[int]) -> str:
    report_norm_id = cast(int, disposition["report_norm_id"])
    status_value = disposition.get("status")
    if status_value == "MAPPED":
        if report_norm_id in alias_ids:
            return "AMBIGUOUS"
        return "UNRESOLVED"  # replaced by the selected physical cell below
    if status_value == "UNMATCHED_SCHEMA_NODE_WITH_SKIPPED_CANDIDATES":
        return "AMBIGUOUS"
    if status_value == "NOT_OBSERVED":
        return "NOT_OBSERVED"
    if status_value == "UNMATCHED_SCHEMA_NODE":
        # E-0038 says only that no row was selected. It does not provide the
        # explicit exhaustive-absence authority required for NOT_OBSERVED.
        return "UNRESOLVED"
    raise _fail(f"unsupported schema disposition: {status_value}")


def _assemble_schema_rows(
    template_rows: Sequence[Mapping[str, Any]],
    cells: Sequence[dict[str, Any]],
    dispositions: Mapping[int, Mapping[str, Any]],
    alias_ids: set[int],
) -> list[dict[str, Any]]:
    selected: dict[tuple[int, str], dict[str, Any]] = {}
    for cell in cells:
        report_norm_id = cell["report_norm_id"]
        if report_norm_id is None:
            continue
        key = (cast(int, report_norm_id), cast(str, cell["current_or_comparative"]))
        if key in selected:
            raise _fail(f"duplicate ReportNormId/period target: {key}")
        selected[key] = cell
    result: list[dict[str, Any]] = []
    for template_row in template_rows:
        report_norm_id = cast(int, template_row["report_norm_id"])
        disposition = dispositions[report_norm_id]
        periods: dict[str, dict[str, Any]] = {}
        for role in ("CURRENT", "COMPARATIVE"):
            cell = selected.get((report_norm_id, role))
            if cell is None:
                status_value = _schema_status_from_disposition(disposition, alias_ids)
                periods[role] = {
                    "status": status_value,
                    "canonical_value": None,
                    "cell_id": None,
                    "source_row_id": disposition.get("selected_row_id"),
                    "candidate_row_ids": copy.deepcopy(disposition.get("candidate_row_ids", [])),
                    "period_end": None,
                }
            else:
                periods[role] = {
                    "status": cell["status"],
                    "canonical_value": cell["exported_canonical_value"],
                    "cell_id": cell["cell_id"],
                    "source_row_id": cell["row_id"],
                    "candidate_row_ids": [cell["row_id"]],
                    "period_end": cell["period_end"],
                }
        result.append(
            {
                **dict(template_row),
                "current": periods["CURRENT"],
                "comparative": periods["COMPARATIVE"],
            }
        )
    return result


def _schema_operand(
    by_id: Mapping[int, Mapping[str, Any]], report_norm_id: int, role: str
) -> NumericOperand:
    schema_row = by_id.get(report_norm_id)
    period = None
    if schema_row is not None:
        period = schema_row["current" if role == "CURRENT" else "comparative"]
    value: Decimal | None = None
    if isinstance(period, dict) and period.get("status") in _STRICT_NUMERIC_STATUSES:
        raw = period.get("canonical_value")
        if type(raw) is not int:
            raise _fail("strict arithmetic operand has a non-integer canonical value")
        value = Decimal(raw)
    cell_id = period.get("cell_id") if isinstance(period, dict) else None
    return NumericOperand(
        operand_id=(
            cell_id if isinstance(cell_id, str) else f"report-norm-id-{report_norm_id}:{role}"
        ),
        value=value,
        cell_id=cell_id if isinstance(cell_id, str) else None,
    )


def _physical_operand(cell: Mapping[str, Any]) -> NumericOperand:
    value: Decimal | None = None
    if cell.get("source_numeric_status") in _STRICT_NUMERIC_STATUSES:
        raw = cell.get("evidence_canonical_value")
        if type(raw) is not int:
            raise _fail("strict physical arithmetic operand is not an integer")
        value = Decimal(raw)
    return NumericOperand(
        operand_id=cast(str, cell["cell_id"]),
        value=value,
        cell_id=cast(str, cell["cell_id"]),
    )


def _finding_dict(
    finding_id: str,
    finding: Any,
    role: str,
    *,
    diagnostic_set: str,
    equation: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "finding_id": finding_id,
        "diagnostic_set": diagnostic_set,
        "period_role": role,
        "check_type": finding.check_type,
        "equation": copy.deepcopy(dict(equation)),
        "result": finding.result.value,
        "expected": str(finding.expected) if finding.expected is not None else None,
        "observed": str(finding.observed) if finding.observed is not None else None,
        "residual": str(finding.residual) if finding.residual is not None else None,
        "tolerance": str(finding.tolerance),
        "operand_ids": list(finding.operand_ids),
        "remediation": list(finding.remediation),
        "may_generate_value": False,
    }


def _diagnose_schema_hierarchy(
    schema_rows: Sequence[dict[str, Any]],
    hierarchy_nodes: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    by_id = {cast(int, row["report_norm_id"]): row for row in schema_rows}
    findings: list[dict[str, Any]] = []
    for node in hierarchy_nodes:
        parent_id = node.get("report_norm_id")
        children = node.get("child_report_norm_ids")
        if type(parent_id) is not int or not isinstance(children, list) or not children:
            continue
        if any(type(child) is not int for child in children):
            raise _fail("schema hierarchy child ID is invalid")
        for role in ("CURRENT", "COMPARATIVE"):
            finding = check_parent_children(
                _schema_operand(by_id, parent_id, role),
                [_schema_operand(by_id, child, role) for child in children],
                tolerance=Decimal(0),
            )
            finding_id = f"SCHEMA_PARENT_{parent_id}_{role}"
            findings.append(
                _finding_dict(
                    finding_id,
                    finding,
                    role,
                    diagnostic_set="SECONDARY_SCHEMA_HIERARCHY",
                    equation={
                        "lhs_report_norm_id": parent_id,
                        "rhs_report_norm_ids": list(children),
                    },
                )
            )
    return findings


def _diagnose_physical_visible_equations(
    cells: Sequence[Mapping[str, Any]],
    validation_config: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], int]:
    equation_control = _mapping(
        validation_config.get("strict_physical_visible_row_equations"),
        "strict physical visible-row equations",
    )
    expected_families = equation_control.get("expected_family_count")
    if (
        validation_config.get("mode") != "DIAGNOSTIC_ONLY_NO_REPAIR"
        or set(validation_config.get("strict_operand_statuses", [])) != _STRICT_NUMERIC_STATUSES
        or validation_config.get("dash_as_zero_allowed") is not False
        or validation_config.get("blank_as_zero_allowed") is not False
        or validation_config.get("failed_check_may_change_mapping_status_or_value") is not False
        or equation_control.get("authority") != "NEWLY_FROZEN_CALIBRATION_MECHANISM_ASSERTION"
        or equation_control.get("mapping_independent_physical_row_ids") is not True
        or expected_families != 18
        or equation_control.get("expected_finding_count") != 36
    ):
        raise _fail("strict physical equation family denominator drifted")
    equations = _sequence(equation_control.get("equations"), "physical equations")
    if (
        len(equations) != expected_families
        or _canonical_sha256(equations) != _E0041_PHYSICAL_EQUATIONS_SHA256
    ):
        raise _fail("strict physical equation list does not contain 18 families")
    by_row_role: dict[tuple[str, str], Mapping[str, Any]] = {}
    for cell in cells:
        key = (cast(str, cell["row_id"]), cast(str, cell["current_or_comparative"]))
        if key in by_row_role:
            raise _fail(f"duplicate physical arithmetic operand: {key}")
        by_row_role[key] = cell

    findings: list[dict[str, Any]] = []
    seen_families: set[str] = set()
    for index, raw in enumerate(equations):
        equation = _mapping(raw, f"physical equation {index}")
        if any("report_norm_id" in key.lower() for key in equation):
            raise _fail("physical arithmetic equations may not use ReportNormId")
        family = equation.get("family_id")
        lhs = equation.get("lhs_row_id")
        rhs = equation.get("rhs_row_ids")
        tolerance = equation.get("tolerance_vnd")
        if (
            not isinstance(family, str)
            or not family
            or family in seen_families
            or not isinstance(lhs, str)
            or _ROW_ID.fullmatch(lhs) is None
            or not isinstance(rhs, list)
            or not rhs
            or any(not isinstance(item, str) or _ROW_ID.fullmatch(item) is None for item in rhs)
            or len(rhs) != len(set(rhs))
            or lhs in rhs
            or type(tolerance) is not int
            or tolerance < 0
        ):
            raise _fail(f"physical equation {index} is invalid")
        seen_families.add(family)
        for role in ("CURRENT", "COMPARATIVE"):
            try:
                lhs_cell = by_row_role[(lhs, role)]
                rhs_cells = [by_row_role[(row_id, role)] for row_id in rhs]
            except KeyError as exc:
                raise _fail(f"physical equation {family} names an absent row", exc) from exc
            finding = check_sum(
                _physical_operand(lhs_cell),
                [_physical_operand(cell) for cell in rhs_cells],
                check_type="PHYSICAL_VISIBLE_ROW_EQUATION",
                tolerance=Decimal(tolerance),
            )
            findings.append(
                _finding_dict(
                    f"{family}_{role}",
                    finding,
                    role,
                    diagnostic_set="STRICT_PHYSICAL_VISIBLE_ROW_EQUATIONS",
                    equation=equation,
                )
            )
    if len(findings) != expected_families * 2:
        raise _fail("strict physical arithmetic finding denominator drifted")
    return findings, expected_families


def _attach_validation_refs(
    cells: Sequence[dict[str, Any]], findings: Sequence[Mapping[str, Any]]
) -> None:
    refs: dict[str, list[str]] = {}
    for finding in findings:
        finding_id = cast(str, finding["finding_id"])
        for operand_id in cast(list[str], finding["operand_ids"]):
            if _CELL_ID.fullmatch(operand_id) is not None:
                refs.setdefault(operand_id, []).append(finding_id)
    for cell in cells:
        cell["validation_refs"] = sorted(refs.get(cast(str, cell["cell_id"]), []))


def assemble_post_mapping_projection(
    *,
    postjoin_payload: Mapping[str, Any],
    mapping_payload: (Mapping[str, Any] | E0040ChallengerResult | AuthenticatedE0040ResultCarrier),
    geometry_registry: Mapping[str, Any],
    geometry_registry_path: Path,
    template_rows: Sequence[Mapping[str, Any]],
    project_root: Path,
    input_records: Mapping[str, Mapping[str, Any]],
    validation_config: Mapping[str, Any],
    expected_row_count: int = 64,
    expected_cell_count: int = 128,
) -> dict[str, Any]:
    """Join post-mapping evidence without consulting E-0037 mapping fields.

    This function is intentionally public for adversarial mechanism tests. The
    formal capture additionally authenticates every input byte before invoking it.
    """

    rows_by_id, cells_by_id, postjoin_nodes = _validate_postjoin(
        postjoin_payload,
        row_count=expected_row_count,
        cell_count=expected_cell_count,
    )
    template_ids = {cast(int, row["report_norm_id"]) for row in template_rows}
    challenger_rows, dispositions, alias_ids, mapping_authority = _validate_mapping_challenger(
        mapping_payload,
        expected_row_ids=set(rows_by_id),
        expected_schema_ids=template_ids,
    )
    registry_record = input_records["geometry_registry"]
    registry_identity = _artifact_identity(registry_record, "geometry registry input record")
    registry_for_validation = (
        _normalize_authenticated_geometry_registry(geometry_registry)
        if registry_identity == _E0041_GEOMETRY_REGISTRY_ARTIFACT
        else geometry_registry
    )
    geometry_by_id = _validate_geometry_registry(
        project_root,
        geometry_registry_path,
        registry_for_validation,
        cells_by_id=cells_by_id,
        expected_authority="RECONSTRUCTED_GEOMETRY_WITH_CELL_SHA_PARITY",
    )
    postjoin_record = input_records["e0037_postjoin"]
    mapping_record = input_records["mapping_challenger"]
    physical_cells = [
        _normalize_cell(
            cells_by_id[cell_id],
            rows_by_id[cells_by_id[cell_id]["row_id"]],
            challenger_rows[cells_by_id[cell_id]["row_id"]],
            geometry_by_id[cell_id],
            unapproved_alias_ids=alias_ids,
            registry_record=registry_record,
            postjoin_record=postjoin_record,
            mapping_record=mapping_record,
        )
        for cell_id in sorted(
            cells_by_id,
            key=lambda value: _cell_coordinates(value),
        )
    ]
    if len(physical_cells) != expected_cell_count:
        raise _fail("physical cell denominator drifted")
    schema_rows = _assemble_schema_rows(template_rows, physical_cells, dispositions, alias_ids)
    value_status_projection = {
        "schema_rows": copy.deepcopy(schema_rows),
        "physical_cells": [
            {
                "cell_id": cell["cell_id"],
                "report_norm_id": cell["report_norm_id"],
                "status": cell["status"],
                "exported_canonical_value": cell["exported_canonical_value"],
            }
            for cell in physical_cells
        ],
    }
    before_validation = _canonical_sha256(value_status_projection)
    strict_findings, strict_family_count = _diagnose_physical_visible_equations(
        physical_cells, validation_config
    )
    secondary_control = _mapping(
        validation_config.get("secondary_schema_hierarchy"),
        "secondary schema hierarchy control",
    )
    # E-0037's authenticated schema projection is the only possible secondary
    # hierarchy authority.  It remains a separate diagnostic denominator.
    secondary_findings = (
        _diagnose_schema_hierarchy(schema_rows, postjoin_nodes)
        if secondary_control.get("enabled") is True
        else []
    )
    after_validation = _canonical_sha256(value_status_projection)
    if before_validation != after_validation:
        raise _fail("accounting diagnostics mutated mapping, status, or values")
    _attach_validation_refs(physical_cells, [*strict_findings, *secondary_findings])
    status_counts = dict(sorted(Counter(cell["status"] for cell in physical_cells).items()))
    schema_status_counts = dict(
        sorted(
            Counter(
                period["status"]
                for row in schema_rows
                for period in (row["current"], row["comparative"])
            ).items()
        )
    )
    return {
        "format_version": 1,
        "experiment_id": "E-0041",
        "dataset_role": "CALIBRATION",
        "state": "POST_MAPPING_DEVELOPMENT_EXCEL_ASSEMBLED",
        "input_hash_ledger": copy.deepcopy(dict(input_records)),
        "access_contract": {
            "join_key": "row_id",
            "exact_postjoin_challenger_row_id_set": True,
            "e0037_mapping_fields_read_or_used": False,
            "e0037_selected_output_fields_read_or_used": False,
            "numeric_period_history_or_review_used_to_choose_mapping": False,
            "mapping_challenger_invocation_count": 0,
            "human_review_opened": False,
            "history_or_mongodb_opened": False,
        },
        "mapping_authority": {
            **copy.deepcopy(mapping_authority),
            "unapproved_changed_alias_report_norm_ids": sorted(alias_ids),
            "unselected_row_count": sum(
                row.get("status") in _UNSELECTED_MAPPING_STATUSES
                for row in challenger_rows.values()
            ),
            "source_only_row_count": sum(
                row.get("status") in _SOURCE_ONLY_MAPPING_STATUSES
                for row in challenger_rows.values()
            ),
        },
        "physical_cells": physical_cells,
        "schema_rows": schema_rows,
        "validation": {
            "mode": "DIAGNOSTIC_ONLY_NO_REPAIR",
            "strict_numeric_operand_statuses": sorted(_STRICT_NUMERIC_STATUSES),
            "dash_or_blank_as_zero": False,
            "pre_validation_value_status_sha256": before_validation,
            "post_validation_value_status_sha256": after_validation,
            "strict_physical_visible_row_equations": {
                "authority": "NEWLY_FROZEN_CALIBRATION_MECHANISM_ASSERTION",
                "mapping_independent_physical_row_ids": True,
                "family_count": strict_family_count,
                "finding_count": len(strict_findings),
                "findings": strict_findings,
                "finding_counts": dict(
                    sorted(Counter(finding["result"] for finding in strict_findings).items())
                ),
            },
            "secondary_schema_hierarchy": {
                "enabled": secondary_control.get("enabled") is True,
                "separate_denominator": True,
                "finding_count": len(secondary_findings),
                "findings": secondary_findings,
                "finding_counts": dict(
                    sorted(Counter(finding["result"] for finding in secondary_findings).items())
                ),
            },
            # Workbook compatibility view: strict findings only.  Secondary
            # diagnostics are never mixed into this denominator.
            "findings": strict_findings,
            "finding_counts": dict(
                sorted(Counter(finding["result"] for finding in strict_findings).items())
            ),
        },
        "metrics": {
            "source_row_count": len(rows_by_id),
            "physical_cell_count": len(physical_cells),
            "schema_row_count": len(schema_rows),
            "physical_cell_status_counts": status_counts,
            "schema_period_status_counts": schema_status_counts,
            "selected_target_cell_count": sum(
                cell["report_norm_id"] is not None for cell in physical_cells
            ),
            "exported_numeric_cell_count": sum(
                cell["exported_canonical_value"] is not None for cell in physical_cells
            ),
        },
        "claim_boundary": (
            "E-0041 is calibration-only post-mapping evidence. It retains exactly "
            "128 source cells, exports only independently selected non-alias-dependent "
            "VALUE/ZERO cells, and never uses arithmetic to repair a mapping or value."
        ),
    }


def _json_text(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, allow_nan=False, sort_keys=True)


def _set_literal_cell(cell: Any, value: object) -> None:
    """Write strings as literal cells, including OCR text beginning with =,+,-,@."""

    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def _append_literal_row(sheet: Any, values: Sequence[object]) -> None:
    row = 1 if sheet.max_row == 1 and sheet["A1"].value is None else sheet.max_row + 1
    for column, value in enumerate(values, start=1):
        _set_literal_cell(sheet.cell(row, column), value)


def _write_provenance_sheet(workbook: Any, cells: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet("PROVENANCE")
    headers = [
        "CellId",
        "RowId",
        "Page",
        "RowOrdinal",
        "AxisOrdinal",
        "PeriodRole",
        "PeriodEnd",
        "Scope",
        "SourceRowIds",
        "SourceLabel",
        "SourceObservation",
        "SourceNumericStatus",
        "VisibleRawValue",
        "EvidenceDisplayedValue",
        "EvidenceCanonicalValue",
        "ExportedCanonicalValue",
        "Status",
        "MappingStatus",
        "ReportNormId",
        "ProposedReportNormId",
        "CandidateReportNormIds",
        "CropPath",
        "CropBbox",
        "CropSha256",
        "SourceRenderSha256",
        "SourceOcrSha256",
        "Unit",
        "UnitMultiplier",
        "NumericVerificationStatus",
        "ValidationRefs",
    ]
    _append_literal_row(sheet, headers)
    for cell in cells:
        crop = cast(Mapping[str, Any], cell["crop"])
        _append_literal_row(
            sheet,
            [
                cell["cell_id"],
                cell["row_id"],
                cell["page"],
                cell["row_ordinal"],
                cell["axis_ordinal"],
                cell["current_or_comparative"],
                cell["period_end"],
                cell["scope"],
                _json_text(cell["source_row_ids"]),
                cell["source_label"],
                cell["source_observation"],
                cell["source_numeric_status"],
                cell["visible_raw_value"],
                cell["evidence_displayed_value"],
                cell["evidence_canonical_value"],
                cell["exported_canonical_value"],
                cell["status"],
                cell["mapping_status"],
                cell["report_norm_id"],
                cell["proposed_report_norm_id"],
                _json_text(cell["candidate_report_norm_ids"]),
                crop["path"],
                _json_text(crop["bbox"]),
                crop["sha256"],
                crop["source_render_sha256"],
                crop["source_ocr_sha256"],
                cell["canonical_unit"],
                cell["unit_multiplier"],
                cell["numeric_verification_status"],
                _json_text(cell["validation_refs"]),
            ],
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _write_validation_sheet(workbook: Any, validation: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("VALIDATION_DIAGNOSTICS")
    headers = [
        "FindingId",
        "DiagnosticSet",
        "PeriodRole",
        "CheckType",
        "Equation",
        "Result",
        "Expected",
        "Observed",
        "Residual",
        "Tolerance",
        "OperandIds",
        "Remediation",
        "MayGenerateValue",
    ]
    _append_literal_row(sheet, headers)
    for finding in cast(Sequence[Mapping[str, Any]], validation["findings"]):
        _append_literal_row(
            sheet,
            [
                finding["finding_id"],
                finding["diagnostic_set"],
                finding["period_role"],
                finding["check_type"],
                _json_text(finding["equation"]),
                finding["result"],
                finding["expected"],
                finding["observed"],
                finding["residual"],
                finding["tolerance"],
                _json_text(finding["operand_ids"]),
                _json_text(finding["remediation"]),
                False,
            ],
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _write_metadata_sheet(workbook: Any, projection: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet("RUN_METADATA")
    _append_literal_row(sheet, ["Key", "Value"])
    values = {
        "experiment_id": projection["experiment_id"],
        "dataset_role": projection["dataset_role"],
        "state": projection["state"],
        "input_hash_ledger": projection["input_hash_ledger"],
        "metrics": projection["metrics"],
        "validation_mode": cast(Mapping[str, Any], projection["validation"])["mode"],
        "claim_boundary": projection["claim_boundary"],
    }
    for key, value in values.items():
        _append_literal_row(
            sheet,
            [key, _json_text(value) if isinstance(value, (dict, list)) else value],
        )


def _freeze_workbook_core_properties(workbook: Any) -> None:
    properties = workbook.properties
    properties.creator = _DETERMINISTIC_CORE_ACTOR
    properties.lastModifiedBy = _DETERMINISTIC_CORE_ACTOR
    properties.created = _DETERMINISTIC_CORE_TIMESTAMP
    properties.modified = _DETERMINISTIC_CORE_TIMESTAMP
    properties.version = "1"
    properties.revision = "1"
    properties.title = None
    properties.subject = None
    properties.description = None
    properties.identifier = None
    properties.language = None
    properties.keywords = None
    properties.category = None
    properties.contentStatus = None
    properties.lastPrinted = None


def _normalize_xlsx_zip(payload: bytes) -> bytes:
    if _sha256_bytes(_DETERMINISTIC_CORE_PROPERTIES_XML) != _DETERMINISTIC_CORE_PROPERTIES_SHA256:
        raise _fail("deterministic core-properties constant drifted")
    source = BytesIO(payload)
    destination = BytesIO()
    try:
        with (
            zipfile.ZipFile(source, "r") as archive,
            zipfile.ZipFile(
                destination,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            ) as output,
        ):
            names = archive.namelist()
            if len(names) != len(set(names)):
                raise _fail("generated workbook contains duplicate ZIP members")
            if "docProps/core.xml" not in names:
                raise _fail("generated workbook lacks core properties")
            for name in sorted(names):
                info = archive.getinfo(name)
                normalized = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                normalized.compress_type = zipfile.ZIP_DEFLATED
                normalized.external_attr = info.external_attr
                normalized.create_system = info.create_system
                member = archive.read(name)
                if name == "docProps/core.xml":
                    member = _DETERMINISTIC_CORE_PROPERTIES_XML
                output.writestr(normalized, member)
    except (OSError, zipfile.BadZipFile) as exc:
        raise _fail("cannot normalize generated workbook", exc) from exc
    return destination.getvalue()


def build_development_workbook(
    *,
    template_bytes: bytes,
    projection: Mapping[str, Any],
    source_sheet: str = "Sheet1",
    schema_row_count: int = 77,
) -> tuple[bytes, dict[str, Any]]:
    workbook, template_rows, before = _load_template(
        template_bytes,
        sheet_name=source_sheet,
        schema_row_count=schema_row_count,
    )
    if [row["report_norm_id"] for row in template_rows] != [
        row["report_norm_id"]
        for row in cast(Sequence[Mapping[str, Any]], projection["schema_rows"])
    ]:
        workbook.close()
        raise _fail("projection/template workbook order differs")
    sheet = workbook[source_sheet]
    headers = [
        "CurrentValueVND",
        "ComparativeValueVND",
        "CurrentStatus",
        "ComparativeStatus",
        "CurrentPeriodEnd",
        "ComparativePeriodEnd",
        "CanonicalUnit",
        "Scope",
    ]
    header_style = copy.copy(sheet["B1"]._style)
    for column, header in enumerate(headers, start=4):
        cell = sheet.cell(1, column)
        _set_literal_cell(cell, header)
        cell._style = copy.copy(header_style)
    projection_rows = cast(Sequence[Mapping[str, Any]], projection["schema_rows"])
    cells = cast(Sequence[Mapping[str, Any]], projection["physical_cells"])
    scope_by_id: dict[int, str | None] = {}
    unit_by_id: dict[int, str | None] = {}
    for cell in cells:
        report_norm_id = cell.get("report_norm_id") or cell.get("proposed_report_norm_id")
        if type(report_norm_id) is int:
            scope_by_id.setdefault(report_norm_id, cast(str | None, cell.get("scope")))
            unit_by_id.setdefault(report_norm_id, cast(str | None, cell.get("canonical_unit")))
    for row in projection_rows:
        excel_row = cast(int, row["excel_row"])
        report_norm_id = cast(int, row["report_norm_id"])
        current = cast(Mapping[str, Any], row["current"])
        comparative = cast(Mapping[str, Any], row["comparative"])
        values = [
            current["canonical_value"],
            comparative["canonical_value"],
            current["status"],
            comparative["status"],
            current["period_end"],
            comparative["period_end"],
            unit_by_id.get(report_norm_id),
            scope_by_id.get(report_norm_id),
        ]
        for column, value in enumerate(values, start=4):
            _set_literal_cell(sheet.cell(excel_row, column), value)
        sheet.cell(excel_row, 4).number_format = _WORKBOOK_NUMBER_FORMAT
        sheet.cell(excel_row, 5).number_format = _WORKBOOK_NUMBER_FORMAT
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = f"A1:K{schema_row_count + 1}"
    _write_provenance_sheet(workbook, cells)
    _write_validation_sheet(workbook, cast(Mapping[str, Any], projection["validation"]))
    _write_metadata_sheet(workbook, projection)
    # New sheets use independent presentation styles; A:C remain untouched.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for support_name in ("PROVENANCE", "VALIDATION_DIAGNOSTICS", "RUN_METADATA"):
        support = workbook[support_name]
        for cell in support[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    _freeze_workbook_core_properties(workbook)
    after = _template_snapshot(workbook, source_sheet, schema_row_count + 1)
    if after != before:
        workbook.close()
        raise _fail("template A:C value/style identity changed before save")
    buffer = BytesIO()
    try:
        workbook.save(buffer)
    except Exception as exc:
        workbook.close()
        raise _fail("cannot serialize development workbook", exc) from exc
    workbook.close()
    output = _normalize_xlsx_zip(buffer.getvalue())
    try:
        reopened = load_workbook(BytesIO(output), data_only=False, read_only=False)
    except Exception as exc:
        raise _fail("cannot reopen development workbook", exc) from exc
    if reopened.sheetnames != [
        source_sheet,
        "PROVENANCE",
        "VALIDATION_DIAGNOSTICS",
        "RUN_METADATA",
    ]:
        reopened.close()
        raise _fail("development workbook sheet order drifted")
    reopened_snapshot = _template_snapshot(reopened, source_sheet, schema_row_count + 1)
    reopened_properties = reopened.properties
    if (
        reopened_properties.creator != _DETERMINISTIC_CORE_ACTOR
        or reopened_properties.lastModifiedBy != _DETERMINISTIC_CORE_ACTOR
        or reopened_properties.created != _DETERMINISTIC_CORE_TIMESTAMP
        or reopened_properties.modified != _DETERMINISTIC_CORE_TIMESTAMP
        or reopened_properties.version != "1"
        or reopened_properties.revision != "1"
    ):
        reopened.close()
        raise _fail("development workbook core properties are nondeterministic")
    reopened.close()
    if reopened_snapshot != before:
        raise _fail("template A:C value/style identity changed after save")
    receipt = {
        "template_identity_value_style_sha256": _canonical_sha256(before),
        "reopened_template_identity_value_style_sha256": _canonical_sha256(reopened_snapshot),
        "exact_template_identity_value_style_fidelity": True,
        "source_sheet": source_sheet,
        "preserved_range": f"A1:C{schema_row_count + 1}",
        "sheet_names": [
            source_sheet,
            "PROVENANCE",
            "VALIDATION_DIAGNOSTICS",
            "RUN_METADATA",
        ],
        "workbook_sha256": _sha256_bytes(output),
        "workbook_size_bytes": len(output),
        "deterministic_core_properties_sha256": (_DETERMINISTIC_CORE_PROPERTIES_SHA256),
    }
    return output, receipt


def _same_regular_file(left: os.stat_result, right: os.stat_result) -> bool:
    return (
        stat.S_ISREG(left.st_mode)
        and stat.S_ISREG(right.st_mode)
        and _file_identity(left) == _file_identity(right)
    )


def _same_regular_inode(left: os.stat_result, right: os.stat_result) -> bool:
    return (
        stat.S_ISREG(left.st_mode)
        and stat.S_ISREG(right.st_mode)
        and (left.st_dev, left.st_ino) == (right.st_dev, right.st_ino)
    )


def _write_exclusive_at(
    parent_descriptor: int,
    filename: str,
    payload: bytes,
) -> os.stat_result:
    flags = (
        os.O_WRONLY
        | os.O_CREAT
        | os.O_EXCL
        | os.O_CLOEXEC
        | os.O_NOFOLLOW
        | getattr(os, "O_BINARY", 0)
    )
    try:
        descriptor = os.open(filename, flags, 0o644, dir_fd=parent_descriptor)
    except OSError as exc:
        if exc.errno == errno.EEXIST:
            raise _fail(f"refusing to overwrite E-0041 output: {filename}", exc) from exc
        raise _fail(f"cannot exclusively create E-0041 output {filename}", exc) from exc
    created_identity: os.stat_result | None = None
    descriptor_open = True
    try:
        created_identity = os.fstat(descriptor)
        view = memoryview(payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                raise _fail(f"short write to E-0041 output {filename}")
            view = view[written:]
        os.fsync(descriptor)
        descriptor_identity = os.fstat(descriptor)
        os.close(descriptor)
        descriptor_open = False
        linked_identity = os.stat(filename, dir_fd=parent_descriptor, follow_symlinks=False)
        if not _same_regular_file(
            descriptor_identity, linked_identity
        ) or linked_identity.st_size != len(payload):
            raise _fail(f"E-0041 output identity changed after write: {filename}")
        return linked_identity
    except BaseException as operation_error:
        cleanup_errors: list[BaseException] = []
        if descriptor_open:
            try:
                os.close(descriptor)
            except BaseException as close_error:
                cleanup_errors.append(close_error)
            descriptor_open = False
        if created_identity is None:
            cleanup_errors.append(
                _fail(f"cannot identify partial E-0041 output for rollback: {filename}")
            )
        else:
            try:
                current = os.stat(filename, dir_fd=parent_descriptor, follow_symlinks=False)
            except FileNotFoundError:
                pass
            except BaseException as stat_error:
                cleanup_errors.append(stat_error)
            else:
                if not _same_regular_inode(created_identity, current):
                    cleanup_errors.append(
                        _fail(f"refusing unsafe partial-write rollback: {filename}")
                    )
                else:
                    try:
                        os.unlink(filename, dir_fd=parent_descriptor)
                        os.fsync(parent_descriptor)
                    except BaseException as unlink_error:
                        cleanup_errors.append(unlink_error)
        if cleanup_errors:
            details = "; ".join(str(error) for error in cleanup_errors)
            raise _fail(
                f"E-0041 output creation failed with incomplete self-rollback: "
                f"{filename}: {details}"
            ) from operation_error
        raise


def _rollback_created_at(
    parent_descriptor: int,
    filename: str,
    identity: os.stat_result,
) -> None:
    try:
        current = os.stat(filename, dir_fd=parent_descriptor, follow_symlinks=False)
    except FileNotFoundError:
        return
    if not _same_regular_inode(current, identity):
        raise _fail(f"refusing unsafe rollback of changed E-0041 output: {filename}")
    os.unlink(filename, dir_fd=parent_descriptor)
    os.fsync(parent_descriptor)


def _read_exact_at(
    parent_descriptor: int,
    filename: str,
    expected_identity: os.stat_result,
    expected_payload: bytes,
) -> None:
    flags = os.O_RDONLY | os.O_CLOEXEC | os.O_NOFOLLOW | getattr(os, "O_NONBLOCK", 0)
    try:
        descriptor = os.open(filename, flags, dir_fd=parent_descriptor)
    except OSError as exc:
        raise _fail(f"cannot reopen E-0041 output {filename}", exc) from exc
    try:
        before = os.fstat(descriptor)
        if not _same_regular_file(before, expected_identity):
            raise _fail(f"E-0041 output identity drifted: {filename}")
        chunks: list[bytes] = []
        remaining = before.st_size
        while remaining:
            block = os.read(descriptor, min(remaining, 1024 * 1024))
            if not block:
                raise _fail(f"short canonical read of E-0041 output {filename}")
            chunks.append(block)
            remaining -= len(block)
        growth = os.read(descriptor, 1)
        after = os.fstat(descriptor)
    finally:
        os.close(descriptor)
    linked = os.stat(filename, dir_fd=parent_descriptor, follow_symlinks=False)
    if (
        growth
        or not _same_regular_file(before, after)
        or not _same_regular_file(after, linked)
        or b"".join(chunks) != expected_payload
    ):
        raise _fail(f"E-0041 output failed canonical byte revalidation: {filename}")


def _read_exact_batch_at(
    parent_descriptor: int,
    expected: Sequence[tuple[str, os.stat_result, bytes]],
) -> None:
    """Observe all pair members as one final cross-file validation batch."""

    flags = os.O_RDONLY | os.O_CLOEXEC | os.O_NOFOLLOW | getattr(os, "O_NONBLOCK", 0)
    descriptors: list[int] = []
    try:
        for filename, _identity, _payload in expected:
            try:
                descriptors.append(os.open(filename, flags, dir_fd=parent_descriptor))
            except OSError as exc:
                raise _fail(f"cannot reopen E-0041 output {filename}", exc) from exc
        before = [os.fstat(descriptor) for descriptor in descriptors]
        if any(
            not _same_regular_file(observed, identity)
            for observed, (_filename, identity, _payload) in zip(before, expected, strict=True)
        ):
            raise _fail("E-0041 final output batch identity drifted before reading")

        observed_payloads: list[bytes] = []
        growth: list[bytes] = []
        for descriptor, observed in zip(descriptors, before, strict=True):
            chunks: list[bytes] = []
            remaining = observed.st_size
            while remaining:
                block = os.read(descriptor, min(remaining, 1024 * 1024))
                if not block:
                    raise _fail("short canonical read of E-0041 final output batch")
                chunks.append(block)
                remaining -= len(block)
            observed_payloads.append(b"".join(chunks))
            growth.append(os.read(descriptor, 1))

        after_read = [os.fstat(descriptor) for descriptor in descriptors]
        linked = [
            os.stat(filename, dir_fd=parent_descriptor, follow_symlinks=False)
            for filename, _identity, _payload in expected
        ]
        final = [os.fstat(descriptor) for descriptor in descriptors]
        if any(growth) or any(
            observed_payload != payload
            or not _same_regular_file(identity, first)
            or not _same_regular_file(first, after)
            or not _same_regular_file(after, linked_identity)
            or not _same_regular_file(linked_identity, last)
            for observed_payload, first, after, linked_identity, last, (
                _filename,
                identity,
                payload,
            ) in zip(
                observed_payloads,
                before,
                after_read,
                linked,
                final,
                expected,
                strict=True,
            )
        ):
            raise _fail("E-0041 final output batch failed canonical byte revalidation")
    finally:
        for descriptor in reversed(descriptors):
            os.close(descriptor)


def _publish_pair(
    project_root: Path,
    output_directory: Path,
    workbook_name: str,
    provenance_name: str,
    workbook_bytes: bytes,
    provenance_bytes: bytes,
) -> tuple[Path, Path]:
    if (
        Path(workbook_name).name != workbook_name
        or Path(provenance_name).name != provenance_name
        or workbook_name == provenance_name
    ):
        raise _fail("E-0041 output names must be distinct plain filenames")
    relative = _relative_under_root(project_root, output_directory, "E-0041 output directory")
    workbook_path = output_directory / workbook_name
    provenance_path = output_directory / provenance_name
    root, root_identity = _open_trusted_root(project_root, "E-0041 pair publication")
    output_descriptor: int | None = None
    created: list[tuple[str, os.stat_result, bytes]] = []
    try:
        output_descriptor, held_chain = _open_directory_chain(
            root,
            relative.parts,
            "E-0041 output directory",
            create=True,
        )
        initial_inventory = tuple(sorted(os.listdir(output_descriptor)))
        if workbook_name in initial_inventory or provenance_name in initial_inventory:
            raise _fail("refusing to overwrite E-0041 output pair")
        provenance_identity = _write_exclusive_at(
            output_descriptor,
            provenance_name,
            provenance_bytes,
        )
        created.append((provenance_name, provenance_identity, provenance_bytes))
        workbook_identity = _write_exclusive_at(
            output_descriptor,
            workbook_name,
            workbook_bytes,
        )
        created.append((workbook_name, workbook_identity, workbook_bytes))
        os.fsync(output_descriptor)
        expected_inventory = tuple(sorted((*initial_inventory, provenance_name, workbook_name)))
        if tuple(sorted(os.listdir(output_descriptor))) != expected_inventory:
            raise _fail("E-0041 held output inventory drifted")

        fresh_output, fresh_chain = _open_directory_chain(
            root,
            relative.parts,
            "E-0041 canonical output revalidation",
            create=False,
        )
        try:
            if (
                fresh_chain != held_chain
                or _directory_identity(os.fstat(fresh_output))
                != _directory_identity(os.fstat(output_descriptor))
                or tuple(sorted(os.listdir(fresh_output))) != expected_inventory
            ):
                raise _fail("E-0041 output directory detached from canonical path")
            canonical_provenance = os.stat(
                provenance_name,
                dir_fd=fresh_output,
                follow_symlinks=False,
            )
            canonical_workbook = os.stat(
                workbook_name,
                dir_fd=fresh_output,
                follow_symlinks=False,
            )
            if not _same_regular_file(provenance_identity, canonical_provenance) or not (
                _same_regular_file(workbook_identity, canonical_workbook)
            ):
                raise _fail("E-0041 output pair inode identity drifted")
            _read_exact_at(
                fresh_output,
                provenance_name,
                provenance_identity,
                provenance_bytes,
            )
            _read_exact_at(
                fresh_output,
                workbook_name,
                workbook_identity,
                workbook_bytes,
            )
        finally:
            os.close(fresh_output)

        os.fsync(output_descriptor)
        final_root, final_output, final_chain = _open_fresh_directory_chain(
            project_root,
            root_identity,
            relative.parts,
            "E-0041 final canonical root revalidation",
        )
        try:
            if (
                final_chain != held_chain
                or _directory_identity(os.fstat(final_output))
                != _directory_identity(os.fstat(output_descriptor))
                or tuple(sorted(os.listdir(final_output))) != expected_inventory
            ):
                raise _fail("E-0041 final output directory detached from canonical path")
            # Open and observe both members before deciding success. Per-file
            # sequential revalidation leaves the first member unauthenticated if
            # it changes while the second member is being read.
            _read_exact_batch_at(final_output, created)
        finally:
            os.close(final_output)
            os.close(final_root)
    except BaseException as publication_error:
        rollback_errors: list[BaseException] = []
        if output_descriptor is not None:
            for filename, identity, _payload in reversed(created):
                try:
                    _rollback_created_at(output_descriptor, filename, identity)
                except BaseException as rollback_error:
                    rollback_errors.append(rollback_error)
        if rollback_errors:
            details = "; ".join(str(error) for error in rollback_errors)
            raise _fail(f"E-0041 pair rollback was incomplete: {details}") from publication_error
        raise
    finally:
        if output_descriptor is not None:
            os.close(output_descriptor)
        os.close(root)
    return workbook_path, provenance_path


def capture_e0041_post_mapping_development_excel(
    project_root: Path,
    *,
    config_path: Path = CONTROL_RELATIVE_PATH,
) -> dict[str, Any]:
    """Fail closed until E0040 has an authenticated artifact and seal.

    ``assemble_post_mapping_projection`` and ``build_development_workbook`` are
    deliberately usable for in-memory calibration development.  Formal capture
    must not silently fall back to the legacy 58-row E0038 baseline.
    """

    del project_root, config_path
    raise _fail(
        "formal E-0041 capture is blocked pending an authenticated E-0040 "
        "artifact and seal; E-0038 is regression baseline only"
    )
