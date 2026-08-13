from __future__ import annotations

import copy
import hashlib
import importlib.util
import json
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from bctc_ai.export.canonical_xlsx import workbook_has_formula
from bctc_ai.export.shb_maturity_review_workbook_v1 import (
    E0042_RELATIVE_PATH,
    build_shb_maturity_review_workbook_v1,
)
from bctc_ai.source_structure.contracts_v1 import canonical_json_sha256_v1

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_SEAL_PATH = _PROJECT_ROOT / "docs/experiments/E-0043-shb-maturity-review-workbook-seal.json"
_SCRIPT = _PROJECT_ROOT / "scripts/experiments/export_shb_maturity_review_workbook_v1.py"
_SPEC = importlib.util.spec_from_file_location("e0043_review_pair_replay", _SCRIPT)
assert _SPEC is not None and _SPEC.loader is not None
_REPLAY = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(_REPLAY)

_CAPTURE_COMMIT = "2f5329867be722e8f33db0a61ec34dee6a7bb26f"
_SEAL_CANONICAL_JSON_SHA256 = "37bf959f5731844b9680f2e5753a132494c39a60da99077af9a1c0bc58e6a142"
_SNAPSHOT_ID = "20260813T144429Z-e0042-shb-maturity-review-workbook-pair-4d5df001433c"
_MANIFEST_SHA256 = "16fea9099bcce9bbdcce1294148ba88c72b80566424aa789506dd904307716f8"
_RUN_RECORD_SHA256 = "5f9a6ea538c71808c5587bc0dacc997a29b902deb75574cbef14d1543654ab8e"
_PROVENANCE_SHA256 = "4d5df001433c4779be995930307bf03acad08b0a765266152b0f4792995c50e1"
_WORKBOOK_SHA256 = "9b4dd0cffd3d088f52dfd9f71548910f3ae05383f056ad2ee68e441899243d9c"
_EXPECTED_SAFETY = {
    "accepted_schema_mapping_authority": False,
    "accounting_truth_authority": False,
    "canonicalization_authority": False,
    "export_authority": False,
    "production_authority": False,
    "review_only": True,
    "schema_ids_are_candidates_only": True,
    "source_and_independent_values_preserved_as_text": True,
    "total_candidate_report_norm_id_is_null": True,
    "value_materialization_authority": False,
}


def _reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_bytes(), object_pairs_hook=_reject_duplicate_keys)
    assert isinstance(value, dict)
    return value


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _assert_exact_static_seal_contract(seal: dict[str, Any]) -> None:
    """Pin every seal field while keeping critical authority checks readable."""

    assert set(seal) == {
        "access_contract",
        "authority",
        "capture_git",
        "claim_boundary",
        "dataset_role",
        "experiment_id",
        "format_version",
        "input_identities",
        "inventory",
        "replay",
        "review_projection_contract",
        "s3_registration",
        "safety",
        "state",
    }
    assert seal["experiment_id"] == "E-0043"
    assert seal["format_version"] == 1
    assert seal["dataset_role"] == "DEVELOPMENT_REVIEW_ONLY"
    assert seal["state"] == "E0043_SHB_MATURITY_REVIEW_ONLY_PAIR_HASH_SEALED"
    assert seal["claim_boundary"] == (
        "This development artifact hash-seals exactly the current two-file SHB "
        "page-24 review-only workbook pair after byte-equal deterministic replay "
        "from the bound E-0042 verification, schema-candidate, statement-context, "
        "and semantic-graph lineage. It adds no accepted ReportNormId mapping, "
        "schema, canonicalization, value-materialization, accounting-truth, "
        "human-review, holdout, production, or export authority."
    )
    assert seal["access_contract"] == {
        "canonical_export_path_invoked": False,
        "deterministic_pair_rebuilt_in_new_output_directory": True,
        "exact_two_file_inventory_validated_before_publication": True,
        "history_or_mongodb_opened": False,
        "holdout_artifact_opened": False,
        "human_review_answer_opened": False,
        "published_pair_opened_only_for_bounded_identity_and_contract_validation": True,
        "schema_mapping_acceptance_invoked": False,
    }
    assert seal["authority"] == {
        "accepted_schema_mapping": False,
        "accounting_truth": False,
        "canonicalization": False,
        "deterministic_review_pair_byte_identity": True,
        "exact_two_file_hash_identity": True,
        "export": False,
        "holdout_or_production": False,
        "human_gold_or_review_approval": False,
        "schema_authority": False,
        "value_materialization": False,
    }
    assert seal["capture_git"] == {
        "commit": _CAPTURE_COMMIT,
        "dirty": False,
        "required_replay_relationship": (
            "EXACT_COMMIT_OR_CLEAN_DESCENDANT_WITH_UNCHANGED_BOUND_INPUTS_AND_IMPLEMENTATION"
        ),
    }
    assert seal["replay"] == {
        "generated_pair_exact_byte_equality": True,
        "provenance_exact_byte_equality": True,
        "replay_git_commit": _CAPTURE_COMMIT,
        "replay_git_dirty": False,
        "workbook_exact_byte_equality": True,
    }
    assert seal["input_identities"] == {
        "e0042_numeric_verification": {
            "path": "docs/experiments/E-0042-shb-maturity-numeric-verification.json",
            "sha256": "929c1c81b0e08e14b5908087d866dc7bacc67c19cc62eb832353c5efb6c1801e",
            "size_bytes": 18835,
            "verification_id": (
                "sgnpvv1:verification:92b2d1d0ad293fb5ee2953128db9fb93c1c7f588eefff1bc00cfdeae16b61f1d"
            ),
        },
        "schema_candidate": {
            "candidate_set_id": (
                "slascv1:candidate:c297f71128bef07be383e684a9ff7ea33b11bf96e4fff154b2e05b39e72ef223"
            ),
            "canonical_json_sha256": (
                "c567007cee02036d77ed4a0bf5dbf797f503694b66713ad2edac06b2b73f4214"
            ),
        },
        "semantic_graph": {
            "graph_id": (
                "slagv2:graph:47ec2635a8b57ee0773f26612d97dc7ce1a700993b169c25d7286f9b74be28d7"
            ),
            "sha256": "afbb553b45b3b776f36cea0696d04ead3031ac576cdb35450e11edd6e77854e6",
        },
        "semantic_page_binding_sha256": (
            "e89153d4d78d337e438b90157ea330e8a84890f58847615ed34e304eff2a3a52"
        ),
        "source_projection": {
            "local_page_id": (
                "ssv2:page:736b745df05b5c1f0ef81a5e985e38a44ef9b92612e9574c1b87ebb3e3b21ca1"
            ),
            "sha256": "1036a24b4fbf8dde6f6b20341cee6d640f7c12cc22d83f67a798af2152e06ff7",
        },
        "statement_context": {
            "canonical_json_sha256": (
                "212df1f51a830a35c2633e02b007580102450f4dbb3bb847584fb8a5a56f285b"
            ),
            "context_id": (
                "sscxtv1:context:a2d480f3bece8e0a29e0a935dbd4be00e4168159a6ec7d3d2946ab17d0b0ab8e"
            ),
        },
    }
    assert seal["inventory"] == {
        "file_count": 2,
        "files": [
            {
                "path": (
                    "output/development/e0042-shb-maturity-review-workbook-v1/provenance.json"
                ),
                "role": "REVIEW_ONLY_MACHINE_PROVENANCE",
                "sha256": _PROVENANCE_SHA256,
                "size_bytes": 44308,
            },
            {
                "path": (
                    "output/development/e0042-shb-maturity-review-workbook-v1/"
                    "shb-maturity-review-only.xlsx"
                ),
                "role": "REVIEW_ONLY_WORKBOOK",
                "sha256": _WORKBOOK_SHA256,
                "size_bytes": 13671,
            },
        ],
    }
    assert seal["review_projection_contract"] == {
        "artifact_role": "REVIEW_ONLY_NON_CANONICAL_NON_EXPORT_AUTHORITY",
        "candidate_report_norm_ids": [753, 754, 755],
        "cell_ids": [
            "page-0001-row-000-axis-1",
            "page-0001-row-000-axis-2",
            "page-0001-row-001-axis-1",
            "page-0001-row-001-axis-2",
            "page-0001-row-002-axis-1",
            "page-0001-row-002-axis-2",
            "page-0001-row-003-axis-1",
            "page-0001-row-003-axis-2",
        ],
        "formula_count": 0,
        "projection_id": (
            "shbmrwv1:projection:3e4bd5f00370c9ce86dee626467e58a41951721904326f91b1db5b3f4b17f129"
        ),
        "projection_sha256": ("3612f3480561d03183caa54e23d11645faba110c46438474e6a5119319930777"),
        "provenance_cell_count": 8,
        "provenance_id": (
            "shbmrwpv1:provenance:3e0876482ed5be902ed9a263175154fa141e68f946609fb32d2753104ec7d4a5"
        ),
        "review_row_count": 4,
        "sheet_names": ["REVIEW_ONLY", "CELL_PROVENANCE", "METADATA_CLAIMS"],
        "source_only_total_candidate_report_norm_id": None,
        "verified_observed_cell_count": 8,
    }
    assert seal["safety"] == _EXPECTED_SAFETY
    assert seal["s3_registration"] == {
        "bucket": "test-s3-duylv",
        "independent_hydrate": {
            "exact_file_hashes_and_sizes": True,
            "status": "PASS",
        },
        "inventory": {
            "logical_bytes": 57979,
            "logical_file_count": 2,
            "unique_bytes": 57979,
            "unique_object_count": 2,
        },
        "manifest": {
            "key": (f"bctc-ai/artifact-snapshots/{_SNAPSHOT_ID}/manifest-{_MANIFEST_SHA256}.json"),
            "sha256": _MANIFEST_SHA256,
            "size_bytes": 3940,
        },
        "objects": [
            {
                "key": f"bctc-ai/objects/sha256/4d/{_PROVENANCE_SHA256}",
                "sha256": _PROVENANCE_SHA256,
                "size_bytes": 44308,
                "upload_disposition": "UPLOADED",
            },
            {
                "key": f"bctc-ai/objects/sha256/9b/{_WORKBOOK_SHA256}",
                "sha256": _WORKBOOK_SHA256,
                "size_bytes": 13671,
                "upload_disposition": "UPLOADED",
            },
        ],
        "policy_id": "S3_BOUNDED_ARTIFACT_SNAPSHOT_V1",
        "restore": {
            "all_incremental_objects_restore_verified": True,
            "status": "PASS",
        },
        "run_record": {
            "key": (f"bctc-ai/artifact-runs/{_SNAPSHOT_ID}/run-{_RUN_RECORD_SHA256}.json"),
            "sha256": _RUN_RECORD_SHA256,
            "size_bytes": 1433,
        },
        "snapshot_id": _SNAPSHOT_ID,
        "source_git_commit": _CAPTURE_COMMIT,
        "upload": {
            "reused_object_count": 0,
            "unique_object_count": 2,
            "uploaded_object_count": 2,
        },
    }
    assert canonical_json_sha256_v1(seal) == _SEAL_CANONICAL_JSON_SHA256


def test_e0043_seal_replays_and_validates_exact_review_only_pair() -> None:
    seal = _read_json(_SEAL_PATH)
    _assert_exact_static_seal_contract(seal)
    candidate, context = _REPLAY._build_exact_inputs()
    verification_bytes = (_PROJECT_ROOT / E0042_RELATIVE_PATH).read_bytes()
    rebuilt = build_shb_maturity_review_workbook_v1(
        candidate,
        context,
        verification_bytes,
    )

    files = {entry["role"]: entry for entry in seal["inventory"]["files"]}
    workbook_entry = files["REVIEW_ONLY_WORKBOOK"]
    provenance_entry = files["REVIEW_ONLY_MACHINE_PROVENANCE"]
    workbook_bytes = (_PROJECT_ROOT / workbook_entry["path"]).read_bytes()
    provenance_bytes = (_PROJECT_ROOT / provenance_entry["path"]).read_bytes()

    assert workbook_bytes == rebuilt.workbook_bytes
    assert provenance_bytes == rebuilt.provenance_bytes
    assert (_sha256(workbook_bytes), len(workbook_bytes)) == (
        workbook_entry["sha256"],
        workbook_entry["size_bytes"],
    )
    assert (_sha256(provenance_bytes), len(provenance_bytes)) == (
        provenance_entry["sha256"],
        provenance_entry["size_bytes"],
    )

    registration = seal["s3_registration"]
    assert registration["snapshot_id"] == (
        "20260813T144429Z-e0042-shb-maturity-review-workbook-pair-4d5df001433c"
    )
    assert registration["manifest"]["sha256"] == (
        "16fea9099bcce9bbdcce1294148ba88c72b80566424aa789506dd904307716f8"
    )
    assert registration["run_record"]["sha256"] == (
        "5f9a6ea538c71808c5587bc0dacc997a29b902deb75574cbef14d1543654ab8e"
    )
    assert registration["inventory"] == {
        "logical_bytes": workbook_entry["size_bytes"] + provenance_entry["size_bytes"],
        "logical_file_count": 2,
        "unique_bytes": workbook_entry["size_bytes"] + provenance_entry["size_bytes"],
        "unique_object_count": 2,
    }
    assert {(entry["sha256"], entry["size_bytes"]) for entry in registration["objects"]} == {
        (workbook_entry["sha256"], workbook_entry["size_bytes"]),
        (provenance_entry["sha256"], provenance_entry["size_bytes"]),
    }
    assert registration["upload"] == {
        "reused_object_count": 0,
        "unique_object_count": 2,
        "uploaded_object_count": 2,
    }
    assert registration["restore"] == {
        "all_incremental_objects_restore_verified": True,
        "status": "PASS",
    }
    assert registration["independent_hydrate"] == {
        "exact_file_hashes_and_sizes": True,
        "status": "PASS",
    }

    provenance = json.loads(
        provenance_bytes,
        object_pairs_hook=_reject_duplicate_keys,
    )
    projection = provenance["projection"]
    contract = seal["review_projection_contract"]
    identities = seal["input_identities"]
    assert provenance["provenance_id"] == contract["provenance_id"]
    assert provenance["projection_sha256"] == contract["projection_sha256"]
    assert projection["projection_id"] == contract["projection_id"]
    assert provenance["workbook"] == {
        "creator": "bctc-ai/shb-maturity-review-v1",
        "formula_count": 0,
        "sha256": workbook_entry["sha256"],
        "sheet_names": contract["sheet_names"],
        "size_bytes": workbook_entry["size_bytes"],
    }
    assert projection["input_identities"] == {
        "e0042_numeric_verification": identities["e0042_numeric_verification"],
        "schema_candidate": identities["schema_candidate"],
        "statement_context": identities["statement_context"],
    }
    assert projection["shared_lineage"] == {
        "semantic_graph_id": identities["semantic_graph"]["graph_id"],
        "semantic_graph_sha256": identities["semantic_graph"]["sha256"],
        "semantic_page_binding_sha256": identities["semantic_page_binding_sha256"],
        "source_local_page_id": identities["source_projection"]["local_page_id"],
        "source_projection_sha256": identities["source_projection"]["sha256"],
    }
    assert projection["metrics"] == {
        "candidate_report_norm_id_count": 3,
        "provenance_cell_count": 8,
        "review_row_count": 4,
        "source_only_null_candidate_row_count": 1,
        "verified_observed_cell_count": 8,
    }
    assert [cell["cell_id"] for cell in projection["provenance_cells"]] == contract["cell_ids"]
    assert [row["candidate_report_norm_id"] for row in projection["review_rows"]] == [
        753,
        754,
        755,
        None,
    ]
    assert provenance["safety"] == seal["safety"]
    assert all(
        provenance["safety"][key] is False
        for key in (
            "accepted_schema_mapping_authority",
            "accounting_truth_authority",
            "canonicalization_authority",
            "export_authority",
            "production_authority",
            "value_materialization_authority",
        )
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    try:
        assert workbook.sheetnames == contract["sheet_names"]
        assert not workbook_has_formula(workbook)
        assert workbook["CELL_PROVENANCE"].max_row - 2 == 8
    finally:
        workbook.close()


def test_e0043_static_contract_rejects_coordinated_critical_seal_tamper() -> None:
    tampered = copy.deepcopy(_read_json(_SEAL_PATH))
    tampered["authority"]["export"] = True
    tampered["authority"]["holdout_or_production"] = True
    tampered["claim_boundary"] = "PRODUCTION_EXPORT_AUTHORITY"
    tampered["access_contract"]["human_review_answer_opened"] = True
    tampered["review_projection_contract"]["candidate_report_norm_ids"] = [999]
    tampered["review_projection_contract"]["source_only_total_candidate_report_norm_id"] = 999
    tampered["review_projection_contract"]["formula_count"] = 999
    tampered["s3_registration"]["bucket"] = "attacker-bucket"
    tampered["s3_registration"]["source_git_commit"] = "0" * 40
    tampered["s3_registration"]["manifest"]["key"] = "attacker/manifest.json"
    tampered["s3_registration"]["run_record"]["key"] = "attacker/run.json"

    with pytest.raises(AssertionError):
        _assert_exact_static_seal_contract(tampered)
