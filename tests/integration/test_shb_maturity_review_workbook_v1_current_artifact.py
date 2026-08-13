from __future__ import annotations

import copy
import hashlib
import importlib.util
import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bctc_ai.export.canonical_xlsx import workbook_has_formula
from bctc_ai.export.shb_maturity_review_workbook_v1 import (
    ARTIFACT_ROLE,
    E0042_RELATIVE_PATH,
    SHEET_NAMES,
    ShbMaturityReviewWorkbookV1Error,
    build_shb_maturity_review_workbook_v1,
)
from bctc_ai.source_structure.contracts_v1 import canonical_json_sha256_v1

_SCRIPT = (
    Path(__file__).resolve().parents[2]
    / "scripts/experiments/export_shb_maturity_review_workbook_v1.py"
)
_SPEC = importlib.util.spec_from_file_location("export_shb_maturity_review_workbook_v1", _SCRIPT)
assert _SPEC is not None and _SPEC.loader is not None
_CLI = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(_CLI)


@pytest.fixture(scope="module")
def exact_inputs(project_root):
    candidate, context = _CLI._build_exact_inputs()
    verification = (project_root / E0042_RELATIVE_PATH).read_bytes()
    return candidate, context, verification


def test_exact_real_inputs_build_deterministic_review_pair(exact_inputs) -> None:
    candidate, context, verification = exact_inputs
    first = build_shb_maturity_review_workbook_v1(candidate, context, verification)
    second = build_shb_maturity_review_workbook_v1(candidate, context, verification)

    assert first == second
    assert first.workbook_sha256 == hashlib.sha256(first.workbook_bytes).hexdigest()
    assert first.provenance_sha256 == hashlib.sha256(first.provenance_bytes).hexdigest()
    provenance = json.loads(first.provenance_bytes)
    assert provenance["artifact_role"] == ARTIFACT_ROLE
    assert provenance["workbook"] == {
        "creator": "bctc-ai/shb-maturity-review-v1",
        "formula_count": 0,
        "sha256": first.workbook_sha256,
        "sheet_names": list(SHEET_NAMES),
        "size_bytes": len(first.workbook_bytes),
    }
    assert provenance["safety"]["export_authority"] is False
    assert provenance["safety"]["accepted_schema_mapping_authority"] is False

    workbook = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        assert not workbook_has_formula(workbook)
        review = workbook["REVIEW_ONLY"]
        assert review["A1"].value == "REVIEW ONLY — NOT CANONICAL — NOT EXPORT AUTHORITY"
        rows = [
            row
            for row in review.iter_rows()
            if row[1].value in {"SHORT_TERM", "MEDIUM_TERM", "LONG_TERM", "TOTAL"}
        ]
        assert [row[1].value for row in rows] == ["SHORT_TERM", "MEDIUM_TERM", "LONG_TERM", "TOTAL"]
        assert [row[3].value for row in rows] == [753, 754, 755, None]
        assert [row[7].value for row in rows] == [
            "225.268.906",
            "162.845.429",
            "271.496.634",
            "659.610.969",
        ]
        assert [row[8].value for row in rows] == [
            "225.268.906",
            "162.845.429",
            "271.496.634",
            "659.610.969",
        ]
        assert [row[10].value for row in rows] == [
            "215.455.247",
            "156.575.830",
            "242.830.903",
            "614.861.980",
        ]
        assert [row[11].value for row in rows] == [
            "215.455.247",
            "156.575.830",
            "242.830.903",
            "614.861.980",
        ]
        assert all(row[13].value == "triệu đồng" for row in rows)
        assert all(row[14].value == 24 for row in rows)
        cells = workbook["CELL_PROVENANCE"]
        assert cells.max_row == 10
        assert {cells.cell(row, 22).value for row in range(3, 11)} != {None}
        assert all(cells.cell(row, 23).value == "VERIFIED_OBSERVED_VALUE" for row in range(3, 11))
        assert all(len(cells.cell(row, 28).value) == 64 for row in range(3, 11))
    finally:
        workbook.close()


@pytest.mark.parametrize("kind", ["candidate", "context"])
def test_semantic_tamper_fails_even_after_coordinated_identity_rehash(
    exact_inputs, kind: str
) -> None:
    candidate, context, verification = exact_inputs
    candidate = copy.deepcopy(candidate)
    context = copy.deepcopy(context)
    if kind == "candidate":
        candidate["role_candidates"][2]["candidate_report_norm_ids"] = [1944]
        payload = copy.deepcopy(candidate)
        payload.pop("candidate_set_id")
        candidate["candidate_set_id"] = f"slascv1:candidate:{canonical_json_sha256_v1(payload)}"
    else:
        context["report_scope"] = "SEPARATE"
        payload = copy.deepcopy(context)
        payload.pop("context_id")
        context["context_id"] = f"sscxtv1:context:{canonical_json_sha256_v1(payload)}"

    with pytest.raises(ShbMaturityReviewWorkbookV1Error, match="identity"):
        build_shb_maturity_review_workbook_v1(candidate, context, verification)
