from __future__ import annotations

import zipfile
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from bctc_ai.export.canonical_xlsx import (
    CORE_PROPERTIES_MEMBER,
    ZIP_MEMBER_TIMESTAMP,
    CanonicalXlsxError,
    append_literal_row,
    canonicalize_xlsx_zip,
    deterministic_workbook_bytes,
    freeze_core_properties,
    set_literal_cell,
    workbook_has_formula,
)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SOURCE"
    append_literal_row(sheet, ("RawText", "Value"))
    append_literal_row(sheet, ("=2+2", "123456789012345678901234567890"))
    try:
        return deterministic_workbook_bytes(workbook, creator="bctc-ai/unit-test")
    finally:
        workbook.close()


def test_deterministic_workbook_bytes_are_canonical_literal_and_idempotent():
    first = _workbook_bytes()
    second = _workbook_bytes()
    assert first == second

    with zipfile.ZipFile(BytesIO(first), "r") as archive:
        assert archive.namelist() == sorted(archive.namelist())
        assert archive.namelist().count(CORE_PROPERTIES_MEMBER) == 1
        assert archive.comment == b""
        for info in archive.infolist():
            assert info.date_time == ZIP_MEMBER_TIMESTAMP
            assert info.compress_type == zipfile.ZIP_DEFLATED
            assert info.create_system == 3
            assert info.extra == b""
            assert info.comment == b""
        core = archive.read(CORE_PROPERTIES_MEMBER)

    assert canonicalize_xlsx_zip(first, core_properties_xml=core) == first
    reopened = load_workbook(BytesIO(first), data_only=False)
    try:
        cell = reopened["SOURCE"]["A2"]
        assert cell.value == "=2+2"
        assert cell.data_type == "s"
        assert not workbook_has_formula(reopened)
        assert reopened["SOURCE"]["B2"].value == "123456789012345678901234567890"
    finally:
        reopened.close()


def test_literal_cell_rejects_text_that_excel_would_silently_truncate():
    workbook = Workbook()
    try:
        with pytest.raises(CanonicalXlsxError, match="cell text limit"):
            set_literal_cell(workbook.active["A1"], "x" * 33_000)
    finally:
        workbook.close()


@pytest.mark.parametrize("unsafe_name", ["../escape.xml", "a//b.xml", "a/./b.xml"])
def test_canonicalizer_rejects_missing_duplicate_and_unsafe_members(unsafe_name: str):
    workbook = Workbook()
    core = freeze_core_properties(workbook, creator="bctc-ai/unit-test")
    workbook.close()

    missing = BytesIO()
    with zipfile.ZipFile(missing, "w") as archive:
        archive.writestr("xl/workbook.xml", b"<workbook/>")
    with pytest.raises(CanonicalXlsxError, match="core-properties"):
        canonicalize_xlsx_zip(missing.getvalue(), core_properties_xml=core)

    duplicate = BytesIO()
    with pytest.warns(UserWarning, match="Duplicate name"):
        with zipfile.ZipFile(duplicate, "w") as archive:
            archive.writestr(CORE_PROPERTIES_MEMBER, core)
            archive.writestr(CORE_PROPERTIES_MEMBER, core)
    with pytest.raises(CanonicalXlsxError, match="duplicate"):
        canonicalize_xlsx_zip(duplicate.getvalue(), core_properties_xml=core)

    unsafe = BytesIO()
    with zipfile.ZipFile(unsafe, "w") as archive:
        archive.writestr(CORE_PROPERTIES_MEMBER, core)
        archive.writestr(unsafe_name, b"unsafe")
    with pytest.raises(CanonicalXlsxError, match="unsafe"):
        canonicalize_xlsx_zip(unsafe.getvalue(), core_properties_xml=core)
