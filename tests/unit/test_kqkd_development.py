from __future__ import annotations

import hashlib
import json
from dataclasses import replace
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bctc_ai.export.kqkd_development import (
    KQKDDevelopmentExportError,
    build_kqkd_development_artifacts,
    export_kqkd_development,
)
from bctc_ai.mapping.kqkd_item_mapping import (
    KQKD_POLICY_RELATIVE_PATH,
    load_kqkd_mapping_policy,
    map_kqkd_items,
)
from bctc_ai.reconciliation.kqkd_numeric import verify_kqkd_numeric_page
from bctc_ai.schema.hierarchy import apply_hierarchy_reference, load_hierarchy_reference
from bctc_ai.schema.registry import load_all
from bctc_ai.tables.kqkd_word_box import (
    load_kqkd_word_box_policy,
    parse_kqkd_word_box_page,
)

_WORD_BOX_FIXTURE = Path("tests/golden/kqkd/mbb-q1-2026-page-0006-ppocrv6-word-box.json")
_SOURCE_PDF = Path("vietstock_bctc/MBB/2026/BCTC Hợp nhất quý 1 năm 2026.pdf")
_TEMPLATE = Path("template/Bank_KQKD_ReportNormId.xlsx")
_DEEPSEEK_LABELS = (
    "_____ Thu nhập lãi và các khoản thu nhập tương tự",
    "Chi phí lãi và các khoản chi phí tương tự",
    "Thụ nhập lại thuận",
    "Thủ nhập từ hoạt động dịch vụ",
    "Chi phí hoạt động dịch vụ",
    "Lãi thuần từ hoạt động dịch vụ",
    "Lãi thuần từ hoạt động kinh doanh ngoại hối",
    "Lãi thuần từ mua bán chứng khoán kinh doanh",
    "Lãi thuần từ mua bán chứng khoán đầu tư",
    "Lãi thuần từ hoạt động kinh doanh khác",
    "Thu nhập từ góp vốn, mua cổ phần",
    "TỔNG THU NHẬP HOẠT ĐỘNG",
    "TỔNG CHI PHÍ HOẠT ĐỘNG",
    "Lợi nhuận thuần từ hoạt động kinh doanh trước chi phí dự phòng rủi ro",
    "Chi phí dự phòng rủi ro",
    "TỔNG LỢI NHUẬN TRƯỚC THUẾ",
    "Chi phí thuê TNDN hiện hành",
    "Chi phí thuế TNDN hoãn lại",
    "Chi phí thuê TNDN trong kỳ",
    "LƠI NHUẬN SAU THUẾ",
    "Lợi ích của cổ đông không kiểm soát",
    "LỢI NHUẬN SAU THUẾ CỦA NGÂN HÀNG",
)


@pytest.fixture(scope="module")
def development_inputs(project_root: Path):
    parsed = parse_kqkd_word_box_page(
        project_root / _WORD_BOX_FIXTURE,
        load_kqkd_word_box_policy(project_root / "config/tables/kqkd-word-box-v1.yaml"),
        page_tag="page-0006",
    )
    _workbooks, schema = load_all(project_root / "template", project_root)
    _registry, hierarchy = load_hierarchy_reference(
        project_root / "config/schemas/hierarchy_reference.yaml",
        project_root,
        schema,
    )
    apply_hierarchy_reference(schema, hierarchy)
    mapping = map_kqkd_items(
        parsed.rows,
        labels_by_reader={
            "deepseek": {
                row.row_id: label for row, label in zip(parsed.rows, _DEEPSEEK_LABELS, strict=True)
            },
            "ppocr": {row.row_id: row.row.label for row in parsed.rows},
        },
        schema=schema,
        policy=load_kqkd_mapping_policy(project_root / KQKD_POLICY_RELATIVE_PATH),
        report_scope=parsed.scope,
    )
    numeric = verify_kqkd_numeric_page(
        parsed,
        project_root / _SOURCE_PDF,
        page_number=6,
    )
    return parsed, mapping, numeric


def _headers(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def test_real_coverage_export_has_42_target_values_and_separate_evidence_scopes(
    tmp_path: Path,
    project_root: Path,
    development_inputs,
):
    parsed, mapping, numeric = development_inputs
    workbook_path = tmp_path / "mbb-kqkd-development.xlsx"
    coverage_path = tmp_path / "coverage.json"
    result = export_kqkd_development(
        template_path=project_root / _TEMPLATE,
        workbook_path=workbook_path,
        coverage_path=coverage_path,
        parsed=parsed,
        mapping=mapping,
        numeric=numeric,
        run_metadata={"bank": "MBB", "quarter": "2026-Q1"},
    )

    assert result.target_value_count == 42
    assert result.accounting_check_count == 32
    assert result.fully_verified is False
    assert hashlib.sha256(workbook_path.read_bytes()).hexdigest() == result.workbook_sha256
    coverage = json.loads(coverage_path.read_text(encoding="utf-8"))
    assert coverage["workbook"] == {
        "filename": workbook_path.name,
        "sha256": result.workbook_sha256,
        "size_bytes": workbook_path.stat().st_size,
    }
    assert coverage["fully_verified"] is False
    assert coverage["authority"] == {
        "accounting": {"passed": True, "scope": "ACCOUNTING_EQUATIONS_ONLY"},
        "fully_verified": False,
        "mapping": {
            "automatic_selection_allowed": True,
            "passed": True,
            "scope": "SCHEMA_ROW_SELECTION_ONLY",
        },
        "numeric": {
            "mapping_authority": False,
            "passed": True,
            "scope": "NUMERIC_TRANSCRIPTION_ONLY",
        },
    }
    assert coverage["coverage"] == {
        "accounting_check_count": 32,
        "accounting_passed_check_count": 32,
        "mapped_schema_count": 21,
        "mapped_ytd_provenance_value_count": 42,
        "not_observed_schema_count": 3,
        "numeric_verified_cell_count": 88,
        "provenance_cell_count": 88,
        "reconciled_schema_count": 24,
        "source_only_row_count": 1,
        "source_row_count": 22,
        "target_value_count": 42,
        "ytd_provenance_cell_count": 44,
    }
    assert coverage["not_observed_report_norm_ids"] == [4381, 4394, 4395]
    assert coverage["source_only_rows"] == [
        {
            "row_id": "page-0006:row-0012",
            "row_ordinal": 12,
            "source_label": "TONG THU NHAP HOAT DÔNG",
        }
    ]
    assert "coverage_sha256" not in coverage

    template = load_workbook(project_root / _TEMPLATE, data_only=False)
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        assert workbook.sheetnames == [
            "KQKD",
            "PROVENANCE",
            "VALIDATION_DIAGNOSTICS",
            "RUN_METADATA",
        ]
        template_main = template.active
        main = workbook["KQKD"]
        for row in range(1, 26):
            for column in range(1, 4):
                assert main.cell(row, column).value == template_main.cell(row, column).value
                assert main.cell(row, column).style_id == template_main.cell(row, column).style_id

        by_id = {main.cell(row, 2).value: row for row in range(2, 26)}
        assert main.cell(by_id[4399], 4).value == 28_982_071_000_000
        assert main.cell(by_id[4399], 9).value == 19_590_312_000_000
        assert (
            sum(
                main.cell(row, column).value is not None
                for row in range(2, 26)
                for column in (4, 9)
            )
            == 42
        )
        for report_norm_id in (4394, 4395, 4381):
            row = by_id[report_norm_id]
            assert main.cell(row, 4).value is None
            assert main.cell(row, 5).value == "NOT_OBSERVED_IN_THIS_PDF"
            assert main.cell(row, 9).value is None
            assert main.cell(row, 10).value == "NOT_OBSERVED_IN_THIS_PDF"

        provenance = workbook["PROVENANCE"]
        provenance_headers = _headers(provenance)
        records = [
            {
                header: provenance.cell(row, column).value
                for header, column in provenance_headers.items()
            }
            for row in range(2, provenance.max_row + 1)
        ]
        assert len(records) == 88
        assert sum(record["AxisGroup"] == "YTD" for record in records) == 44
        source_only = [record for record in records if record["RowId"] == "page-0006:row-0012"]
        assert len(source_only) == 4
        assert all(record["SourceRowStatus"] == "SOURCE_ONLY_PDF_ROW" for record in source_only)
        assert all(record["ReportNormId"] is None for record in source_only)

        diagnostics = workbook["VALIDATION_DIAGNOSTICS"]
        diagnostic_headers = _headers(diagnostics)
        assert diagnostics.max_row == 33
        assert all(
            diagnostics.cell(row, diagnostic_headers["Status"]).value == "PASS"
            for row in range(2, diagnostics.max_row + 1)
        )
        assert not any(
            cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()
        template.close()


def test_build_is_byte_deterministic_and_coverage_has_no_self_hash(
    project_root: Path,
    development_inputs,
):
    parsed, mapping, numeric = development_inputs
    kwargs = {
        "template_path": project_root / _TEMPLATE,
        "workbook_name": "mbb-kqkd-development.xlsx",
        "parsed": parsed,
        "mapping": mapping,
        "numeric": numeric,
        "run_metadata": {"z": [2, 1], "a": {"scope": "development"}},
    }
    first = build_kqkd_development_artifacts(**kwargs)
    second = build_kqkd_development_artifacts(**kwargs)

    assert first.workbook_bytes == second.workbook_bytes
    assert first.coverage_bytes == second.coverage_bytes
    assert first.workbook_sha256 == hashlib.sha256(first.workbook_bytes).hexdigest()
    assert first.coverage_sha256 == hashlib.sha256(first.coverage_bytes).hexdigest()
    coverage = json.loads(first.coverage_bytes)
    assert coverage["workbook"]["sha256"] == first.workbook_sha256
    assert "coverage_sha256" not in first.coverage_bytes.decode("utf-8")
    workbook = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    workbook.close()


def test_export_refuses_overwrite_and_cross_bound_numeric_result(
    tmp_path: Path,
    project_root: Path,
    development_inputs,
):
    parsed, mapping, numeric = development_inputs
    workbook_path = tmp_path / "mbb-kqkd-development.xlsx"
    coverage_path = tmp_path / "coverage.json"
    workbook_path.write_bytes(b"KEEP")
    with pytest.raises(KQKDDevelopmentExportError, match="overwrite"):
        export_kqkd_development(
            template_path=project_root / _TEMPLATE,
            workbook_path=workbook_path,
            coverage_path=coverage_path,
            parsed=parsed,
            mapping=mapping,
            numeric=numeric,
        )
    assert workbook_path.read_bytes() == b"KEEP"
    assert not coverage_path.exists()

    with pytest.raises(KQKDDevelopmentExportError, match="another OCR input"):
        build_kqkd_development_artifacts(
            template_path=project_root / _TEMPLATE,
            workbook_name="mbb-kqkd-development.xlsx",
            parsed=parsed,
            mapping=mapping,
            numeric=replace(numeric, source_ocr_sha256="0" * 64),
        )
