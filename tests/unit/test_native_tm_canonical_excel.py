from __future__ import annotations

import copy
import hashlib
import inspect
import json
import os
import shutil
import subprocess
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from bctc_ai.export import native_tm_canonical_excel as exporter


def _canonical(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")


def _write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(payload)


def _git(root: Path, *arguments: str) -> str:
    return subprocess.run(
        ["git", *arguments],
        cwd=root,
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()


_STRICT_LOADER_STUB = """
from __future__ import annotations

import hashlib
import json
from pathlib import Path


def _load(path: Path, project_root: Path, expected_sha256: str):
    candidate = path if path.is_absolute() else project_root / path
    payload = candidate.read_bytes()
    if hashlib.sha256(payload).hexdigest() != expected_sha256:
        raise RuntimeError("trusted SHA drifted")
    value = json.loads(payload)
    encoded = (json.dumps(value, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\\n").encode("utf-8")
    if encoded != payload:
        raise RuntimeError("artifact is not canonical JSON")
    return value
"""

_MAPPING_STUB = (
    _STRICT_LOADER_STUB
    + """

def load_registered_native_tm_canonical_mapping(path: Path, *, project_root: Path, expected_sha256: str):
    return _load(path, project_root, expected_sha256)
"""
)

_OBSERVATIONS_STUB = (
    _STRICT_LOADER_STUB
    + """

def load_registered_native_tm_observations(path: Path, *, project_root: Path, expected_sha256: str):
    return _load(path, project_root, expected_sha256)
"""
)


def _copy_committed_runtime(source_root: Path, target_root: Path) -> str:
    shutil.copytree(source_root / "src/bctc_ai", target_root / "src/bctc_ai")
    for relative in (
        exporter.EXPORT_POLICY_RELATIVE_PATH,
        Path("pyproject.toml"),
        Path("uv.lock"),
    ):
        destination = target_root / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_root / relative, destination)
    (target_root / "src/bctc_ai/mapping/native_tm_canonical.py").write_text(
        _MAPPING_STUB, encoding="utf-8"
    )
    (target_root / "src/bctc_ai/rows/native_tm_observations.py").write_text(
        _OBSERVATIONS_STUB, encoding="utf-8"
    )
    (target_root / ".gitignore").write_text("output/**\n", encoding="utf-8")
    _git(target_root, "init", "--quiet")
    _git(target_root, "config", "user.email", "tests@example.invalid")
    _git(target_root, "config", "user.name", "Native TM Excel Tests")
    _git(target_root, "add", ".")
    _git(target_root, "commit", "--quiet", "-m", "synthetic exporter producer")
    return _git(target_root, "rev-parse", "HEAD")


def _schema_payload() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    schema: list[dict[str, Any]] = []
    contexts: list[dict[str, Any]] = []
    dispositions: list[dict[str, Any]] = []
    for index in range(1713):
        identifier = 10_000 + index
        name = "=Mapped total" if index == 0 else f"TM item {index:04d}"
        schema.append(
            {
                "schema_id": identifier,
                "canonical_name": name,
                "statement_type": "TM",
                "display_order": index,
                "parent_id": None,
                "children": [],
                "hierarchy_level": 0,
                "notes_section": "SYNTHETIC TM",
            }
        )
        contexts.append(
            {
                "report_norm_id": identifier,
                "display_order": index,
                "context_status": "RESOLVED",
                "mapping_eligible": True,
                "parent_report_norm_id": None,
                "note_family_root_id": identifier,
            }
        )
        if index < 2:
            mapping_disposition = "EXISTING_ITEM"
            outcome = "OBSERVED_VALUE"
            reason = "SYNTHETIC_MAPPED_EXISTING_ITEM"
            row_ids = [f"row-{index}"]
            observation_ids = [f"observation-{index}"]
        elif index == 2:
            mapping_disposition = "LOCALLY_NOT_OBSERVED"
            outcome = "NOT_OBSERVED"
            reason = "ABSENT_FROM_LOCALLY_COMPLETE_BOUNDED_TABLE_SUBTREE"
            row_ids = []
            observation_ids = []
        else:
            mapping_disposition = "UNRESOLVED"
            outcome = "UNRESOLVED"
            reason = "UNASSESSED_OUTSIDE_BOUNDED_TABLE_SUBTREE"
            row_ids = []
            observation_ids = []
        dispositions.append(
            {
                "report_norm_id": identifier,
                "canonical_name": name,
                "display_order": index,
                "context_status": "RESOLVED",
                "mapping_eligible": True,
                "parent_report_norm_id": None,
                "note_family_root_id": identifier,
                "mapping_disposition": mapping_disposition,
                "terminal_outcome": outcome,
                "reason": reason,
                "source_row_ids": row_ids,
                "source_observation_ids": observation_ids,
            }
        )
    return schema, contexts, dispositions


def _source_collections() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    page = {
        "page_id": "page-1",
        "page": 1,
        "record_type": "PAGE_CONTEXT",
        "source_object_id": "PAGE_CONTEXT::page-1",
        "source_disposition": "PRESERVED_SOURCE_OBJECT",
    }
    context = {
        "context_id": "context-1",
        "page": 38,
        "source_table_id": "table-1",
        "row_ids": ["row-0", "row-1"],
        "dimension_ids": ["dimension-0", "dimension-1"],
        "observation_ids": ["observation-0", "observation-1"],
        "record_type": "CONTEXT",
        "source_object_id": "CONTEXT::context-1",
        "source_disposition": "PRESERVED_SOURCE_OBJECT",
        "large_receipt_bound_context": "x" * 35_000,
    }
    rows = [
        {
            "row_id": f"row-{index}",
            "context_id": "context-1",
            "page": 38,
            "source_table_id": "table-1",
            "observation_ids": [f"observation-{index}"],
            "label": f"Raw label {index}",
            "record_type": "ROW",
            "source_object_id": f"ROW::row-{index}",
            "source_disposition": "QUANTITATIVE_TM",
        }
        for index in range(2)
    ]
    dimensions = [
        {
            "dimension_id": f"dimension-{index}",
            "context_id": "context-1",
            "page": 38,
            "source_table_id": "table-1",
            "axis_ordinal": index,
            "binding_status": "RESOLVED",
            "unit": "VND",
            "unit_multiplier": 1_000_000,
            "period_type": "SNAPSHOT",
            "period_start": "2026-03-31",
            "period_end": "2026-03-31",
            "period_materialization": {"resolution_status": "SOURCE_BINDING_RESOLVED"},
            "unit_materialization": {"resolution_status": "SOURCE_BINDING_RESOLVED"},
            "record_type": "DIMENSION",
            "source_object_id": f"DIMENSION::dimension-{index}",
            "source_disposition": "QUANTITATIVE_TM",
        }
        for index in range(2)
    ]
    observations = [
        {
            "observation_id": f"observation-{index}",
            "row_id": f"row-{index}",
            "dimension_id": f"dimension-{index}",
            "context_id": "context-1",
            "page": 38,
            "source_table_id": "table-1",
            "value_text": str(100 + index),
            "source_status": "OBSERVED_VALUE",
            "parsed": {
                "value": str(100 + index),
                "normalized_text": str(100 + index),
            },
            "source_slot_record_sha256": f"{index + 1}" * 64,
            "record_type": "OBSERVATION",
            "source_object_id": f"observation-{index}",
            "source_disposition": "QUANTITATIVE_TM",
        }
        for index in range(2)
    ]
    nested_evidence = {
        "record_type": "SOURCE_EVIDENCE",
        "source_object_id": "evidence-nested",
        "source_disposition": "PRESERVED_SOURCE_OBJECT",
        "value": "nested",
    }
    top_evidence = {
        "record_type": "SOURCE_EVIDENCE",
        "source_object_id": "evidence-top",
        "source_disposition": "PRESERVED_SOURCE_OBJECT",
        "nested": nested_evidence,
    }
    collections = {
        "page_inventory": [page],
        "contexts": [context],
        "rows": rows,
        "dimensions": dimensions,
        "observations": observations,
        "source_evidence": {"synthetic": [top_evidence]},
        "source_references": {},
    }
    return collections, [
        page,
        context,
        *rows,
        *dimensions,
        *observations,
        top_evidence,
        nested_evidence,
    ]


def _source_dispositions(
    objects: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    upstream = [
        {
            "source_object_type": record["record_type"],
            "source_object_id": record["source_object_id"],
            "source_disposition": record["source_disposition"],
        }
        for record in objects
    ]
    upstream.sort(key=lambda record: (record["source_object_type"], record["source_object_id"]))
    mapping: list[dict[str, Any]] = []
    for record in upstream:
        object_type = record["source_object_type"]
        object_id = record["source_object_id"]
        mapped_index: int | None = None
        if object_type == "OBSERVATION":
            mapped_index = int(object_id.rsplit("-", 1)[1])
        elif object_type == "ROW":
            mapped_index = int(object_id.rsplit("-", 1)[1])
        mapped = mapped_index is not None
        mapping.append(
            {
                "source_object_type": object_type,
                "source_object_id": object_id,
                "upstream_source_disposition": record["source_disposition"],
                "mapping_disposition": "MAPPED_EXISTING_ITEM" if mapped else "UNRESOLVED",
                "reason": (
                    "=literal formula-looking provenance"
                    if object_id == "observation-0"
                    else (
                        "SYNTHETIC_MAPPED_EXISTING_ITEM"
                        if mapped
                        else "UNASSESSED_OUTSIDE_BOUNDED_TABLE_SUBTREE"
                    )
                ),
                "context_id": "context-1" if mapped else None,
                "row_id": f"row-{mapped_index}" if mapped else None,
                "dimension_id": (
                    f"dimension-{mapped_index}" if object_type == "OBSERVATION" else None
                ),
                "target_report_norm_id": 10_000 + mapped_index if mapped else None,
                "candidate_report_norm_ids": [10_000 + mapped_index] if mapped else [],
                "match_basis": "CANONICAL_RETRIEVAL_KEY_EXACT" if mapped else None,
                "matched_retrieval_key": f"raw label {mapped_index}" if mapped else None,
                "alias_authority_type": None,
                "alias_authority_evidence_sha256": None,
            }
        )
    return upstream, mapping


def _build_project(source_root: Path, project_root: Path) -> dict[str, Any]:
    project_root.mkdir()
    producer_commit = _copy_committed_runtime(source_root, project_root)
    source_bytes = b"synthetic native TM PDF bytes\n"
    source_relative = "output/development/synthetic-input/source.pdf"
    _write(project_root / source_relative, source_bytes)
    source = {
        "dataset_role": "LOGIC_DEVELOPMENT",
        "relative_path": source_relative,
        "sha256": hashlib.sha256(source_bytes).hexdigest(),
        "size_bytes": len(source_bytes),
    }

    discovery_relative = "output/development/synthetic-input/discovery.json"
    discovery_payload = {"run_id": "discovery", "source": source}
    discovery_bytes = _canonical(discovery_payload)
    _write(project_root / discovery_relative, discovery_bytes)
    discovery_receipt = {
        "path": discovery_relative,
        "sha256": hashlib.sha256(discovery_bytes).hexdigest(),
        "size_bytes": len(discovery_bytes),
    }

    native_relative = "output/development/synthetic-input/native-document.json"
    native_payload = {
        "format_version": "REGISTERED_NATIVE_TM_FULL_DOCUMENT_ARTIFACT_RESULT_V1",
        "policy": "REGISTERED_NATIVE_TM_FULL_DOCUMENT_ARTIFACT_V1",
        "claim_boundary": "SOURCE_VISIBLE_NATIVE_TM_INVENTORY_ONLY",
        "status": "PARTIAL_NATIVE_TM_FULL_DOCUMENT_ARTIFACT",
        "run_id": "synthetic-native-document",
        "source": source,
        "statement_discovery": discovery_receipt,
        "code": {"commit": "a" * 40, "dirty": False, "implementation": []},
    }
    native_bytes = _canonical(native_payload)
    _write(project_root / native_relative, native_bytes)
    native_receipt = {
        "path": native_relative,
        "sha256": hashlib.sha256(native_bytes).hexdigest(),
        "size_bytes": len(native_bytes),
        "format_version": native_payload["format_version"],
        "policy": native_payload["policy"],
        "claim_boundary": native_payload["claim_boundary"],
        "status": native_payload["status"],
        "run_id": native_payload["run_id"],
        "producer_git_commit": native_payload["code"]["commit"],
    }

    collections, source_objects = _source_collections()
    upstream, mapping_source = _source_dispositions(source_objects)
    observations_relative = "output/development/synthetic-input/observations.json"
    observations_payload = {
        "format_version": "REGISTERED_NATIVE_TM_OBSERVATIONS_RESULT_V1",
        "policy": "REGISTERED_NATIVE_TM_OBSERVATIONS_V1",
        "claim_boundary": "SOURCE_ONLY_NATIVE_TM_OBSERVATION_FLATTENING",
        "status": "COMPLETE_NATIVE_TM_SOURCE_OBJECT_ACCOUNTING",
        "run_id": "synthetic-observations",
        "source": source,
        "native_tm_document": native_receipt,
        "code": {"commit": "b" * 40, "dirty": False, "implementation": []},
        "report_scope_binding": {
            "binding_status": "RESOLVED_UNANIMOUS_SOURCE_VISIBLE_TM_HEADERS",
            "scope": "CONSOLIDATED",
        },
        **collections,
        "source_accounting": {
            "counts": {"source_disposition_count": len(upstream)},
            "source_object_accounting_complete": True,
        },
        "source_dispositions": upstream,
    }
    observations_bytes = _canonical(observations_payload)
    _write(project_root / observations_relative, observations_bytes)
    observations_receipt = {
        "path": observations_relative,
        "sha256": hashlib.sha256(observations_bytes).hexdigest(),
        "size_bytes": len(observations_bytes),
        "format_version": observations_payload["format_version"],
        "policy": observations_payload["policy"],
        "claim_boundary": observations_payload["claim_boundary"],
        "status": observations_payload["status"],
        "run_id": observations_payload["run_id"],
        "producer_git_commit": observations_payload["code"]["commit"],
    }

    schema, contexts, schema_dispositions = _schema_payload()
    canonical_observations = [
        {
            "observation_id": f"observation-{index}",
            "row_id": f"row-{index}",
            "dimension_id": f"dimension-{index}",
            "context_id": "context-1",
            "report_norm_id": 10_000 + index,
            "terminal_outcome": "OBSERVED_VALUE",
            "reported_value": str(100 + index),
            "unit": "VND",
            "unit_multiplier": 1_000_000,
            "canonical_value": str((100 + index) * 1_000_000),
            "period_type": "SNAPSHOT",
            "period_start": "2026-03-31",
            "period_end": "2026-03-31",
            "as_of_date": "2026-03-31",
            "presentation_scope": "CONSOLIDATED",
            "match_basis": "CANONICAL_RETRIEVAL_KEY_EXACT",
            "source_record_sha256": f"{index + 1}" * 64,
        }
        for index in range(2)
    ]
    reason_counts = {
        "ABSENT_FROM_LOCALLY_COMPLETE_BOUNDED_TABLE_SUBTREE": 1,
        "SYNTHETIC_MAPPED_EXISTING_ITEM": 2,
        "UNASSESSED_OUTSIDE_BOUNDED_TABLE_SUBTREE": 1710,
    }
    mapping_relative = "output/development/synthetic-input/mapping.json"
    mapping_payload = {
        "format_version": "REGISTERED_NATIVE_TM_CANONICAL_MAPPING_RESULT_V1",
        "policy": "REGISTERED_NATIVE_TM_CANONICAL_MAPPING_V1",
        "claim_boundary": "BOUNDED_SOURCE_EVIDENCE_TM_CANONICAL_MAPPING_ONLY",
        "status": "COMPLETE_NATIVE_TM_CANONICAL_DISPOSITION_ACCOUNTING",
        "run_id": "synthetic-mapping",
        "source": source,
        "native_tm_observations": observations_receipt,
        "schema": {"statement_type": "TM", "item_count": 1713},
        "code": {"commit": "c" * 40, "dirty": False, "implementation": []},
        "authority": {},
        "isolation": {},
        "non_decision_features": {},
        "inputs": {},
        "producer_snapshots": {
            "policy": {"payload": {}},
            "tm_schema": {
                "record_count": len(schema),
                "payload_sha256": exporter._record_sha256(schema),
                "records": schema,
            },
            "tm_context": {
                "record_count": len(contexts),
                "payload_sha256": exporter._record_sha256(contexts),
                "projection_sha256": "d" * 64,
                "records": contexts,
            },
            "accepted_typed_aliases": {
                "record_count": 0,
                "payload_sha256": exporter._record_sha256([]),
                "projection_sha256": "e" * 64,
                "records": [],
            },
        },
        "routing_contract": {"equation_used_for_target_selection": False},
        "root_assessments": [
            {
                "inferred_root_report_norm_id": 10_000,
                "accepted_context_id": "context-1",
                "status": "ACCEPTED",
                "terminal_row_id": "row-0",
                "equation_check_ids": ["equation-1"],
            }
        ],
        "accepted_subtrees": [
            {
                "inferred_root_report_norm_id": 10_000,
                "context_id": "context-1",
                "terminal_row_id": "row-0",
                "equation_check_ids": ["equation-1"],
                "local_completeness_status": "COMPLETE_BOUNDED_TABLE_SUBTREE",
            }
        ],
        "source_accounting": {
            "upstream_source_object_count": len(upstream),
            "mapping_source_disposition_count": len(mapping_source),
            "upstream_source_dispositions_sha256": exporter._record_sha256(upstream),
            "mapping_source_dispositions_sha256": exporter._record_sha256(mapping_source),
            "exactly_one_mapping_disposition_per_upstream_source_object": True,
            "source_object_accounting_complete": True,
        },
        "source_dispositions": mapping_source,
        "canonical_observations": canonical_observations,
        "schema_dispositions": schema_dispositions,
        "equation_checks": [
            {
                "equation_check_id": "equation-1",
                "status": "EXACT",
                "inferred_root_report_norm_id": 10_000,
                "context_id": "context-1",
                "used_as_post_lineage_veto": True,
                "used_for_target_selection": False,
            }
        ],
        "coverage": {
            "statement_type": "TM",
            "schema_item_count": 1713,
            "schema_disposition_count": 1713,
            "terminal_outcome_counts": {
                "OBSERVED_VALUE": 2,
                "OBSERVED_ZERO": 0,
                "DASH": 0,
                "BLANK": 0,
                "NOT_OBSERVED": 1,
                "NOT_APPLICABLE": 0,
                "AMBIGUOUS": 0,
                "UNRESOLVED": 1710,
            },
            "reason_counts": reason_counts,
            "exactly_one_terminal_outcome_per_schema_id": True,
            "workbook_display_order_complete": True,
        },
        "completion": {
            "accepted_root_count": 1,
            "source_accounting_complete": True,
            "tm_schema_disposition_accounting_complete": True,
            "document_complete": False,
        },
    }
    mapping_bytes = _canonical(mapping_payload)
    _write(project_root / mapping_relative, mapping_bytes)
    return {
        "project_root": project_root,
        "producer_commit": producer_commit,
        "mapping_relative": mapping_relative,
        "mapping_path": project_root / mapping_relative,
        "mapping_sha256": hashlib.sha256(mapping_bytes).hexdigest(),
        "mapping_payload": mapping_payload,
        "observations_relative": observations_relative,
        "observations_path": project_root / observations_relative,
        "observations_sha256": hashlib.sha256(observations_bytes).hexdigest(),
        "observations_payload": observations_payload,
        "transitive_identities": (
            exporter.ArtifactIdentity(
                native_relative, hashlib.sha256(native_bytes).hexdigest(), len(native_bytes)
            ),
            exporter.ArtifactIdentity(
                source_relative, hashlib.sha256(source_bytes).hexdigest(), len(source_bytes)
            ),
            exporter.ArtifactIdentity(
                discovery_relative,
                hashlib.sha256(discovery_bytes).hexdigest(),
                len(discovery_bytes),
            ),
        ),
        "source_count": len(upstream),
    }


@pytest.fixture(scope="module")
def synthetic(tmp_path_factory: pytest.TempPathFactory, project_root: Path) -> dict[str, Any]:
    return _build_project(project_root, tmp_path_factory.mktemp("tm-excel") / "project")


@pytest.fixture(scope="module")
def policy(synthetic: dict[str, Any]):
    return exporter.load_native_tm_canonical_excel_policy(
        exporter.EXPORT_POLICY_RELATIVE_PATH, synthetic["project_root"]
    )


def _producer(synthetic: dict[str, Any]) -> dict[str, object]:
    return exporter._current_exporter_producer(synthetic["project_root"])


def _build_artifacts(
    synthetic: dict[str, Any],
    policy: exporter.NativeTMCanonicalExcelPolicy,
    *,
    mapping_payload: dict[str, Any] | None = None,
    observations_payload: dict[str, Any] | None = None,
    workbook: str = "output/development/export/prebuilt.xlsx",
    provenance: str = "output/development/export/prebuilt.json",
):
    mapping_payload = copy.deepcopy(mapping_payload or synthetic["mapping_payload"])
    observations_payload = observations_payload or synthetic["observations_payload"]
    observations_bytes = _canonical(observations_payload)
    observations_identity = exporter.ArtifactIdentity(
        synthetic["observations_relative"],
        hashlib.sha256(observations_bytes).hexdigest(),
        len(observations_bytes),
    )
    mapping_payload["native_tm_observations"] = exporter._expected_upstream_envelope(
        observations_identity, observations_payload
    )
    mapping_bytes = _canonical(mapping_payload)
    return exporter._build_prevalidated_native_tm_canonical_excel_artifacts(
        mapping_payload,
        observations_payload,
        mapping_identity=exporter.ArtifactIdentity(
            synthetic["mapping_relative"],
            hashlib.sha256(mapping_bytes).hexdigest(),
            len(mapping_bytes),
        ),
        observations_identity=observations_identity,
        transitive_identities=synthetic["transitive_identities"],
        workbook_relative_path=workbook,
        provenance_relative_path=provenance,
        policy=policy,
        exporter_producer=_producer(synthetic),
    )


@pytest.fixture(scope="module")
def artifacts(synthetic: dict[str, Any], policy):
    return _build_artifacts(synthetic, policy)


def _mock_strict_loaders(monkeypatch: pytest.MonkeyPatch, synthetic: dict[str, Any]) -> None:
    monkeypatch.setattr(
        exporter,
        "load_registered_native_tm_canonical_mapping",
        lambda path, *, project_root, expected_sha256: copy.deepcopy(synthetic["mapping_payload"]),
    )
    monkeypatch.setattr(
        exporter,
        "load_registered_native_tm_observations",
        lambda path, *, project_root, expected_sha256: copy.deepcopy(
            synthetic["observations_payload"]
        ),
    )


def test_policy_and_public_api_are_bank_agnostic(synthetic: dict[str, Any], policy) -> None:
    assert policy.schema_disposition_count == 1713
    assert not hasattr(policy, "source_disposition_count")
    parameters = inspect.signature(exporter.export_registered_native_tm_canonical_excel).parameters
    assert set(parameters) == {
        "project_root",
        "mapping_path",
        "mapping_expected_sha256",
        "workbook_path",
        "provenance_path",
    }
    assert (
        "11032"
        not in (synthetic["project_root"] / exporter.EXPORT_POLICY_RELATIVE_PATH).read_text()
    )


def test_workbook_is_complete_auditable_and_formula_free(
    synthetic: dict[str, Any], artifacts
) -> None:
    assert artifacts.summary["tm_schema_disposition_count"] == 1713
    assert artifacts.summary["source_object_disposition_count"] == synthetic["source_count"]
    assert artifacts.summary["source_object_sheet_row_count"] > synthetic["source_count"]
    workbook = load_workbook(BytesIO(artifacts.workbook_bytes), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == exporter.SHEET_NAMES
        assert workbook["SCHEMA_DISPOSITIONS"].max_row == 1714
        assert workbook["SOURCE_DISPOSITIONS"].max_row == synthetic["source_count"] + 1
        assert workbook["CANONICAL_OBSERVATIONS"].max_row == 3
        assert workbook["CANONICAL_OBSERVATIONS"]["D2"].value == "=Mapped total"
        assert workbook["CANONICAL_OBSERVATIONS"]["D2"].data_type == "s"
        headers = [cell.value for cell in workbook["CANONICAL_OBSERVATIONS"][1]]
        assert "SourceRawLabel" in headers
        assert "SourceNormalizedLabel" in headers
        source_headers = [cell.value for cell in workbook["SOURCE_DISPOSITIONS"][1]]
        assert "SourceObjectSha256" in source_headers
        assert "UpstreamSourceDispositionJson" in source_headers
        assert not exporter.workbook_has_formula(workbook)
        receipt = json.loads(artifacts.provenance_bytes)
        assert set(receipt["content_ledger"]["observations_sections"]) == {
            "page_inventory",
            "contexts",
            "rows",
            "dimensions",
            "observations",
            "source_evidence",
            "source_dispositions",
            "source_accounting",
        }
        assert (
            receipt["content_ledger"]["reconstructed_source_objects"]["record_count"]
            == synthetic["source_count"]
        )
    finally:
        workbook.close()


def test_build_is_byte_deterministic(synthetic: dict[str, Any], policy, artifacts) -> None:
    rebuilt = _build_artifacts(synthetic, policy)
    assert rebuilt.workbook_bytes == artifacts.workbook_bytes
    assert rebuilt.provenance_bytes == artifacts.provenance_bytes


@pytest.mark.parametrize(
    "mutation",
    [
        "SOURCE_COUNT",
        "UPSTREAM_ACCOUNTING",
        "DUPLICATE_CITATION",
        "VALUE",
        "STATUS",
        "ZERO_STATUS",
        "UNIT",
        "PERIOD",
        "SCOPE",
        "LOCALITY",
        "ORDER",
    ],
)
def test_semantic_and_accounting_mutations_fail_closed(
    synthetic: dict[str, Any], policy, mutation: str
) -> None:
    mapping = copy.deepcopy(synthetic["mapping_payload"])
    observations = copy.deepcopy(synthetic["observations_payload"])
    if mutation == "SOURCE_COUNT":
        mapping["source_dispositions"].pop()
    elif mutation == "UPSTREAM_ACCOUNTING":
        observations["source_accounting"]["counts"]["source_disposition_count"] += 1
    elif mutation == "DUPLICATE_CITATION":
        mapping["schema_dispositions"][1]["source_observation_ids"] = ["observation-0"]
    elif mutation == "VALUE":
        mapping["canonical_observations"][0]["reported_value"] = "999"
    elif mutation == "STATUS":
        observations["observations"][0]["source_status"] = "OBSERVED_ZERO"
    elif mutation == "ZERO_STATUS":
        observations["observations"][0]["value_text"] = "0.00"
        observations["observations"][0]["parsed"].update(
            {"value": "0.00", "normalized_text": "0.00"}
        )
        mapping["canonical_observations"][0].update({"reported_value": "0", "canonical_value": "0"})
    elif mutation == "UNIT":
        mapping["canonical_observations"][0]["unit"] = "USD"
    elif mutation == "PERIOD":
        mapping["canonical_observations"][0]["period_end"] = "2025-12-31"
    elif mutation == "SCOPE":
        observations["report_scope_binding"]["scope"] = "SEPARATE"
    elif mutation == "LOCALITY":
        observations["observations"][0]["page"] = 39
    else:
        mapping["canonical_observations"].reverse()
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        _build_artifacts(
            synthetic,
            policy,
            mapping_payload=mapping,
            observations_payload=observations,
            workbook="output/development/export/rejected.xlsx",
            provenance="output/development/export/rejected.json",
        )


def test_scaled_zero_source_is_valid_when_status_is_observed_zero(
    synthetic: dict[str, Any], policy
) -> None:
    mapping = copy.deepcopy(synthetic["mapping_payload"])
    observations = copy.deepcopy(synthetic["observations_payload"])
    observations["observations"][0].update(
        {
            "value_text": "0.00",
            "source_status": "OBSERVED_ZERO",
            "parsed": {"value": "0.00", "normalized_text": "0.00"},
        }
    )
    mapping["canonical_observations"][0].update(
        {
            "reported_value": "0",
            "canonical_value": "0",
            "terminal_outcome": "OBSERVED_ZERO",
        }
    )
    mapping["schema_dispositions"][0]["terminal_outcome"] = "OBSERVED_ZERO"
    mapping["coverage"]["terminal_outcome_counts"].update({"OBSERVED_VALUE": 1, "OBSERVED_ZERO": 1})
    built = _build_artifacts(
        synthetic,
        policy,
        mapping_payload=mapping,
        observations_payload=observations,
        workbook="output/development/export/zero-scale.xlsx",
        provenance="output/development/export/zero-scale.json",
    )
    assert built.summary["schema_terminal_outcome_counts"]["OBSERVED_ZERO"] == 1


def test_zero_accepted_root_and_mismatch_equation_are_valid_generic_cases(
    synthetic: dict[str, Any], policy
) -> None:
    mapping = copy.deepcopy(synthetic["mapping_payload"])
    observations = copy.deepcopy(synthetic["observations_payload"])
    for disposition in mapping["schema_dispositions"]:
        disposition.update(
            {
                "mapping_disposition": "UNRESOLVED",
                "terminal_outcome": "UNRESOLVED",
                "reason": "UNRESOLVED_LOCAL_COMPLETENESS_OR_EQUATION",
                "source_row_ids": [],
                "source_observation_ids": [],
            }
        )
    for record in mapping["source_dispositions"]:
        if record["mapping_disposition"] == "MAPPED_EXISTING_ITEM":
            record.update(
                {
                    "mapping_disposition": "UNRESOLVED",
                    "target_report_norm_id": None,
                    "candidate_report_norm_ids": [],
                    "match_basis": None,
                    "matched_retrieval_key": None,
                }
            )
    mapping["canonical_observations"] = []
    mapping["accepted_subtrees"] = []
    mapping["root_assessments"] = [
        {
            "inferred_root_report_norm_id": 10_000,
            "status": "UNRESOLVED_LOCAL_COMPLETENESS_OR_EQUATION",
            "equation_check_ids": ["equation-1"],
        }
    ]
    mapping["equation_checks"][0]["status"] = "MISMATCH"
    mapping["completion"]["accepted_root_count"] = 0
    mapping["coverage"]["terminal_outcome_counts"] = {
        outcome: 1713 if outcome == "UNRESOLVED" else 0 for outcome in exporter._TERMINAL_OUTCOMES
    }
    mapping["coverage"]["reason_counts"] = {"UNRESOLVED_LOCAL_COMPLETENESS_OR_EQUATION": 1713}
    mapping["source_accounting"]["mapping_source_dispositions_sha256"] = exporter._record_sha256(
        mapping["source_dispositions"]
    )
    built = _build_artifacts(
        synthetic,
        policy,
        mapping_payload=mapping,
        observations_payload=observations,
        workbook="output/development/export/zero.xlsx",
        provenance="output/development/export/zero.json",
    )
    assert built.summary["canonical_observation_count"] == 0


def test_mapping_authentication_precedes_any_receipt_path_read(
    synthetic: dict[str, Any], policy, monkeypatch: pytest.MonkeyPatch
) -> None:
    opened_transitive = False

    def explode_loader(*args, **kwargs):
        raise RuntimeError("untrusted mapping rejected")

    def observe_transitive(*args, **kwargs):
        nonlocal opened_transitive
        opened_transitive = True
        raise AssertionError("transitive path was opened before mapping authentication")

    monkeypatch.setattr(exporter, "load_registered_native_tm_canonical_mapping", explode_loader)
    monkeypatch.setattr(exporter, "_open_identity_guard", observe_transitive)
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError, match="authoritative producer"):
        exporter._strict_load_input_bundle(
            project_root=synthetic["project_root"],
            mapping_path=synthetic["mapping_path"],
            mapping_expected_sha256=synthetic["mapping_sha256"],
            policy=policy,
        )
    assert opened_transitive is False


def test_completed_loader_opens_no_lineage_path_before_producer_replay_authentication(
    synthetic: dict[str, Any], artifacts, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook = Path("output/development/export/prebuilt.xlsx")
    provenance = Path("output/development/export/prebuilt.json")
    _write(synthetic["project_root"] / workbook, artifacts.workbook_bytes)
    _write(synthetic["project_root"] / provenance, artifacts.provenance_bytes)
    opened_lineage = False

    def reject_replay(*args, **kwargs):
        assert opened_lineage is False
        raise exporter.NativeTMCanonicalExcelExportError("producer authentication rejected")

    def observe_lineage(*args, **kwargs):
        nonlocal opened_lineage
        opened_lineage = True
        raise AssertionError("lineage opened before historical producer replay")

    monkeypatch.setattr(exporter, "_producer_commit_replay", reject_replay)
    monkeypatch.setattr(exporter, "_open_identity_guard", observe_lineage)
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError, match="authentication rejected"):
        exporter.load_registered_native_tm_canonical_excel(
            project_root=synthetic["project_root"],
            workbook_path=workbook,
            workbook_expected_sha256=artifacts.workbook_sha256,
            provenance_path=provenance,
            provenance_expected_sha256=artifacts.provenance_sha256,
        )
    assert opened_lineage is False


def test_completed_pair_uses_old_producer_not_current_semantics(
    tmp_path: Path, project_root: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    synthetic = _build_project(project_root, tmp_path / "project")
    policy = exporter.load_native_tm_canonical_excel_policy(
        exporter.EXPORT_POLICY_RELATIVE_PATH, synthetic["project_root"]
    )
    artifacts = _build_artifacts(synthetic, policy)
    workbook = Path("output/development/export/prebuilt.xlsx")
    provenance = Path("output/development/export/prebuilt.json")
    _write(synthetic["project_root"] / workbook, artifacts.workbook_bytes)
    _write(synthetic["project_root"] / provenance, artifacts.provenance_bytes)
    producer_module = synthetic["project_root"] / "src/bctc_ai/export/native_tm_canonical_excel.py"
    producer_module.write_text(producer_module.read_text() + "\n# future exporter change\n")
    _git(
        synthetic["project_root"],
        "add",
        producer_module.relative_to(synthetic["project_root"]).as_posix(),
    )
    _git(synthetic["project_root"], "commit", "--quiet", "-m", "future exporter")

    def explode(*args, **kwargs):
        raise AssertionError("current exporter/upstream semantics were executed")

    monkeypatch.setattr(exporter, "_policy_from_bytes", explode)
    monkeypatch.setattr(
        exporter, "_build_prevalidated_native_tm_canonical_excel_artifacts", explode
    )
    monkeypatch.setattr(exporter, "_strict_load_input_bundle", explode)
    monkeypatch.setattr(exporter, "load_registered_native_tm_canonical_mapping", explode)
    monkeypatch.setattr(exporter, "load_registered_native_tm_observations", explode)
    result = exporter.load_registered_native_tm_canonical_excel(
        project_root=synthetic["project_root"],
        workbook_path=workbook,
        workbook_expected_sha256=artifacts.workbook_sha256,
        provenance_path=provenance,
        provenance_expected_sha256=artifacts.provenance_sha256,
    )
    assert result.summary["tm_schema_disposition_count"] == 1713


@pytest.mark.parametrize(
    "mutation", ["COMMIT", "IMPLEMENTATION", "MANIFEST_SUPERSET", "POLICY_SHA"]
)
def test_tampered_exporter_producer_receipt_fails_before_replay(
    tmp_path: Path, project_root: Path, mutation: str
) -> None:
    synthetic = _build_project(project_root, tmp_path / "project")
    policy = exporter.load_native_tm_canonical_excel_policy(
        exporter.EXPORT_POLICY_RELATIVE_PATH, synthetic["project_root"]
    )
    artifacts = _build_artifacts(synthetic, policy)
    receipt = json.loads(artifacts.provenance_bytes)
    if mutation == "COMMIT":
        receipt["code"]["commit"] = "f" * 40
    elif mutation == "IMPLEMENTATION":
        receipt["code"]["implementation"][0]["sha256"] = "0" * 64
    elif mutation == "MANIFEST_SUPERSET":
        relative = exporter.EXPORT_POLICY_RELATIVE_PATH.as_posix()
        payload = exporter._git_file_bytes(
            synthetic["project_root"], receipt["code"]["commit"], relative
        )
        receipt["code"]["implementation"].append(
            {
                "path": relative,
                "sha256": hashlib.sha256(payload).hexdigest(),
                "size_bytes": len(payload),
            }
        )
        receipt["code"]["implementation"].sort(key=lambda record: record["path"])
    else:
        receipt["policy"]["export_policy_sha256"] = "0" * 64
    provenance_bytes = _canonical(receipt)
    workbook = Path("output/development/export/prebuilt.xlsx")
    provenance = Path("output/development/export/tampered.json")
    _write(synthetic["project_root"] / workbook, artifacts.workbook_bytes)
    _write(synthetic["project_root"] / provenance, provenance_bytes)
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter.load_registered_native_tm_canonical_excel(
            project_root=synthetic["project_root"],
            workbook_path=workbook,
            workbook_expected_sha256=artifacts.workbook_sha256,
            provenance_path=provenance,
            provenance_expected_sha256=hashlib.sha256(provenance_bytes).hexdigest(),
        )


@pytest.mark.parametrize(
    "payload",
    [
        b"",
        b"bad",
        exporter._REPLAY_MAGIC + b"\0" * 16,
        exporter._frame_replayed_pair(b"w", b"p") + b"x",
    ],
)
def test_replay_pair_frame_is_exact(payload: bytes) -> None:
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter._parse_replayed_pair(payload)


def test_public_export_is_exclusive_and_strict_replays(
    tmp_path: Path, project_root: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    synthetic = _build_project(project_root, tmp_path / "project")
    _mock_strict_loaders(monkeypatch, synthetic)
    workbook = Path("output/development/export/public.xlsx")
    provenance = Path("output/development/export/public.json")
    result = exporter.export_registered_native_tm_canonical_excel(
        project_root=synthetic["project_root"],
        mapping_path=synthetic["mapping_path"],
        mapping_expected_sha256=synthetic["mapping_sha256"],
        workbook_path=workbook,
        provenance_path=provenance,
    )
    assert result.workbook_path.is_file()
    assert result.provenance_path.is_file()
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter.export_registered_native_tm_canonical_excel(
            project_root=synthetic["project_root"],
            mapping_path=synthetic["mapping_path"],
            mapping_expected_sha256=synthetic["mapping_sha256"],
            workbook_path=workbook,
            provenance_path=provenance,
        )


@pytest.mark.parametrize("existing_kind", ["FILE", "SYMLINK"])
def test_final_name_concurrent_publishers_are_preserved(tmp_path: Path, existing_kind: str) -> None:
    root = tmp_path / "project"
    root.mkdir()
    workbook = root / "output/development/export/pair.xlsx"
    provenance = root / "output/development/export/pair.json"
    workbook.parent.mkdir(parents=True)
    if existing_kind == "FILE":
        workbook.write_bytes(b"foreign")
    else:
        target = root / "foreign"
        target.write_bytes(b"foreign")
        workbook.symlink_to(target)
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter._publish_pair(root, workbook, provenance, b"workbook", b"provenance")
    assert os.path.lexists(workbook)
    assert not provenance.exists()


def test_second_output_o_excl_rolls_back_workbook_and_preserves_provenance(
    tmp_path: Path,
) -> None:
    root = tmp_path / "project"
    root.mkdir()
    workbook = root / "output/development/export/pair.xlsx"
    provenance = root / "output/development/export/pair.json"
    provenance.parent.mkdir(parents=True)
    provenance.write_bytes(b"foreign-provenance")
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter._publish_pair(root, workbook, provenance, b"workbook", b"provenance")
    assert not workbook.exists()
    assert provenance.read_bytes() == b"foreign-provenance"


def test_parent_symlink_is_rejected_without_outside_write(tmp_path: Path) -> None:
    root = tmp_path / "project"
    outside = tmp_path / "outside"
    root.mkdir()
    outside.mkdir()
    (root / "output").symlink_to(outside, target_is_directory=True)
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter._publish_pair(
            root,
            root / "output/development/pair.xlsx",
            root / "output/development/pair.json",
            b"workbook",
            b"provenance",
        )
    assert list(outside.iterdir()) == []


def test_pair_detects_parent_swap_between_writes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root = tmp_path / "project"
    root.mkdir()
    parent = root / "output/development/export"
    workbook = parent / "pair.xlsx"
    provenance = parent / "pair.json"
    original = exporter._publication._write_exclusive
    calls = 0

    def swap_after_first(project_root: Path, path: Path, payload: bytes):
        nonlocal calls
        guard = original(project_root, path, payload)
        calls += 1
        if calls == 1:
            parent.rename(parent.with_name("old-export"))
            parent.mkdir()
        return guard

    monkeypatch.setattr(exporter._publication, "_write_exclusive", swap_after_first)
    with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
        exporter._publish_pair(root, workbook, provenance, b"workbook", b"provenance")
    assert not workbook.exists()
    assert not provenance.exists()
    assert not (parent.with_name("old-export") / workbook.name).exists()


def test_foreign_replacement_before_rollback_survives(tmp_path: Path) -> None:
    root = tmp_path / "project"
    root.mkdir()
    workbook = root / "output/development/export/pair.xlsx"
    provenance = root / "output/development/export/pair.json"
    pair = exporter._publish_pair(root, workbook, provenance, b"workbook", b"provenance")
    workbook.unlink()
    workbook.write_bytes(b"foreign")
    try:
        with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
            exporter._rollback_pair(pair, RuntimeError("replay failed"))
    finally:
        exporter._close_pair(pair)
    assert workbook.read_bytes() == b"foreign"
    assert not provenance.exists()


def test_foreign_replacement_after_rollback_capture_survives(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root = tmp_path / "project"
    root.mkdir()
    output = root / "output/development/export/pair.xlsx"
    guard = exporter._publication._write_exclusive(root, output, b"owned")
    original_rename = exporter._publication.os.rename

    def replace_after_capture(src, dst, *args, **kwargs):
        result = original_rename(src, dst, *args, **kwargs)
        if src == output.name:
            output.write_bytes(b"foreign-after-capture")
        return result

    monkeypatch.setattr(exporter._publication.os, "rename", replace_after_capture)
    try:
        exporter._publication._rollback_publication(guard, RuntimeError("replay failed"))
    finally:
        exporter._publication._close_guard_best_effort(guard)
    assert output.read_bytes() == b"foreign-after-capture"


@pytest.mark.parametrize("mutation", ["NAME", "PARENT"])
def test_final_pair_guard_revalidation_detects_name_and_parent_replacement(
    tmp_path: Path, mutation: str
) -> None:
    root = tmp_path / "project"
    root.mkdir()
    parent = root / "output/development/export"
    workbook = parent / "pair.xlsx"
    provenance = parent / "pair.json"
    pair = exporter._publish_pair(root, workbook, provenance, b"workbook", b"provenance")
    if mutation == "NAME":
        workbook.unlink()
        workbook.write_bytes(b"foreign")
    else:
        parent.rename(parent.with_name("old-export"))
        parent.mkdir()
    try:
        with pytest.raises(exporter.NativeTMCanonicalExcelExportError):
            exporter._revalidate_published_pair(root, pair, b"workbook", b"provenance")
    finally:
        try:
            exporter._rollback_pair(pair, RuntimeError("revalidation failed"))
        except exporter.NativeTMCanonicalExcelExportError:
            pass
        exporter._close_pair(pair)
    if mutation == "NAME":
        assert workbook.read_bytes() == b"foreign"


@pytest.mark.parametrize("target_index", range(5))
def test_each_transitive_input_name_swap_fails_stability_recheck(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
    target_index: int,
) -> None:
    synthetic = _build_project(project_root, tmp_path / "project")
    policy = exporter.load_native_tm_canonical_excel_policy(
        exporter.EXPORT_POLICY_RELATIVE_PATH, synthetic["project_root"]
    )
    _mock_strict_loaders(monkeypatch, synthetic)
    bundle = exporter._strict_load_input_bundle(
        project_root=synthetic["project_root"],
        mapping_path=synthetic["mapping_path"],
        mapping_expected_sha256=synthetic["mapping_sha256"],
        policy=policy,
    )
    target = bundle.guards[target_index].path
    original = target.with_name(target.name + ".owned")
    target.rename(original)
    target.write_bytes(b"foreign")
    try:
        with pytest.raises(
            exporter.NativeTMCanonicalExcelExportError, match="changed during export"
        ):
            exporter._assert_bundle_stable(bundle)
        assert target.read_bytes() == b"foreign"
    finally:
        exporter._close_input_bundle(bundle)


def test_real_registered_vpb_projection_counts_and_sheets(project_root: Path) -> None:
    mapping_path = (
        project_root
        / "output/development/vpb-q1-2026-native-tm-canonical-v1/native-tm-canonical-mapping.json"
    )
    observations_path = (
        project_root
        / "output/development/vpb-q1-2026-native-tm-observations-v1/native-tm-observations.json"
    )
    if not mapping_path.is_file() or not observations_path.is_file():
        pytest.skip("registered VPB integration artifacts are not present")
    mapping_bytes = mapping_path.read_bytes()
    observations_bytes = observations_path.read_bytes()
    mapping = json.loads(mapping_bytes)
    observations = json.loads(observations_bytes)
    policy = exporter.load_native_tm_canonical_excel_policy(
        exporter.EXPORT_POLICY_RELATIVE_PATH, project_root
    )
    producer_schema_count = mapping["producer_snapshots"]["tm_schema"]["record_count"]
    if producer_schema_count != policy.schema_disposition_count:
        pytest.skip("registered VPB mapping predates the current TM schema denominator")
    native_document = observations["native_tm_document"]
    source = observations["source"]
    native_payload = json.loads((project_root / native_document["path"]).read_bytes())
    transitive = (
        exporter.ArtifactIdentity(
            native_document["path"], native_document["sha256"], native_document["size_bytes"]
        ),
        exporter.ArtifactIdentity(source["relative_path"], source["sha256"], source["size_bytes"]),
        exporter._identity_from_record(native_payload["statement_discovery"], "discovery"),
    )
    producer = {
        "commit": "0" * 40,
        "dirty": False,
        "implementation": [
            dict(record) for record in exporter._implementation_ledger(project_root)
        ],
    }
    built = exporter._build_prevalidated_native_tm_canonical_excel_artifacts(
        mapping,
        observations,
        mapping_identity=exporter.ArtifactIdentity(
            mapping_path.relative_to(project_root).as_posix(),
            hashlib.sha256(mapping_bytes).hexdigest(),
            len(mapping_bytes),
        ),
        observations_identity=exporter.ArtifactIdentity(
            observations_path.relative_to(project_root).as_posix(),
            hashlib.sha256(observations_bytes).hexdigest(),
            len(observations_bytes),
        ),
        transitive_identities=transitive,
        workbook_relative_path="output/development/integration/vpb.xlsx",
        provenance_relative_path="output/development/integration/vpb.json",
        policy=policy,
        exporter_producer=producer,
    )
    assert built.summary["tm_schema_disposition_count"] == 1713
    assert built.summary["source_object_disposition_count"] == 11032
    assert built.summary["source_object_sheet_row_count"] == 11037
    assert built.summary["canonical_observation_count"] == 8
    assert built.summary["schema_terminal_outcome_counts"] == {
        "OBSERVED_VALUE": 4,
        "OBSERVED_ZERO": 0,
        "DASH": 0,
        "BLANK": 0,
        "NOT_OBSERVED": 4,
        "NOT_APPLICABLE": 0,
        "AMBIGUOUS": 0,
        "UNRESOLVED": 1693,
    }
    assert [record["inferred_root_report_norm_id"] for record in mapping["root_assessments"]] == [
        561
    ]
    expected_tuples = [
        (
            561,
            "4065152",
            "4065152000000",
            "VND",
            1_000_000,
            "2026-03-31",
            "CONSOLIDATED",
            "OBSERVED_VALUE",
        ),
        (
            561,
            "2774182",
            "2774182000000",
            "VND",
            1_000_000,
            "2025-12-31",
            "CONSOLIDATED",
            "OBSERVED_VALUE",
        ),
        (
            562,
            "2970048",
            "2970048000000",
            "VND",
            1_000_000,
            "2026-03-31",
            "CONSOLIDATED",
            "OBSERVED_VALUE",
        ),
        (
            562,
            "2292077",
            "2292077000000",
            "VND",
            1_000_000,
            "2025-12-31",
            "CONSOLIDATED",
            "OBSERVED_VALUE",
        ),
        (
            563,
            "1094895",
            "1094895000000",
            "VND",
            1_000_000,
            "2026-03-31",
            "CONSOLIDATED",
            "OBSERVED_VALUE",
        ),
        (
            563,
            "481921",
            "481921000000",
            "VND",
            1_000_000,
            "2025-12-31",
            "CONSOLIDATED",
            "OBSERVED_VALUE",
        ),
        (565, "209", "209000000", "VND", 1_000_000, "2026-03-31", "CONSOLIDATED", "OBSERVED_VALUE"),
        (565, "184", "184000000", "VND", 1_000_000, "2025-12-31", "CONSOLIDATED", "OBSERVED_VALUE"),
    ]
    assert [
        (
            record["report_norm_id"],
            record["reported_value"],
            record["canonical_value"],
            record["unit"],
            record["unit_multiplier"],
            record["period_end"],
            record["presentation_scope"],
            record["terminal_outcome"],
        )
        for record in mapping["canonical_observations"]
    ] == expected_tuples
    workbook = load_workbook(BytesIO(built.workbook_bytes), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == exporter.SHEET_NAMES
        assert workbook["SOURCE_OBJECTS"].max_row == 11038
        assert not exporter.workbook_has_formula(workbook)
    finally:
        workbook.close()
