from __future__ import annotations

from openpyxl import load_workbook

from bctc_ai.core.contracts import EvidenceStatus, ObservationKind, PipelineRecord, ValueStatus
from bctc_ai.export.workbook import export_workbook, verify_export
from bctc_ai.schema.registry import load_all


def test_workbook_preserves_schema_order_and_support_sheets(tmp_path, project_root):
    output = tmp_path / "result.xlsx"
    records = [
        PipelineRecord(
            document_id="sha256:sample",
            statement_type="CDKT",
            schema_id=4302,
            canonical_name="TÀI SẢN",
            raw_value="1.234",
            normalized_value="1234",
            current_or_comparative="CURRENT",
            status=EvidenceStatus.AUTO_VERIFIED_MEDIUM,
            unit="VND",
            period_end="2026-06-30",
        ),
        PipelineRecord(
            document_id="sha256:sample",
            statement_type="CDKT",
            schema_id=4310,
            canonical_name="Tiền mặt, vàng bạc, đá quý",
            raw_value="khó đọc",
            normalized_value=None,
            current_or_comparative="CURRENT",
            status=EvidenceStatus.REVIEW_REQUIRED,
            rejection_reason="independent OCR disagreement",
        ),
    ]
    result = export_workbook(
        project_root,
        output,
        records,
        run_metadata={"run_id": "test", "source_sha256": "sample"},
        questions_path=project_root / "questions_for_user.jsonl",
        schema_additions_path=project_root / "proposed_schema_additions.jsonl",
    )
    workbooks, items = load_all(project_root / "template", project_root)
    verify_export(output, workbooks, items)
    assert result.exported_value_count == 1
    assert result.review_count == 1
    assert result.schema_counts == {"CDKT": 99, "KQKD": 25, "LCTT": 110, "TM": 1721}

    workbook = load_workbook(output, read_only=True, data_only=True)
    tm = workbook["TM"]
    assert tm.cell(tm.max_row, 2).value == 1944
    assert (
        tm.cell(tm.max_row, 3).value == "Cho vay giao dịch ký quỹ và ứng trước tiền bán chứng khoán"
    )
    workbook.close()


def test_workbook_exports_observed_zero_but_never_out_of_scope_value(tmp_path, project_root):
    output = tmp_path / "value-status.xlsx"
    records = [
        PipelineRecord(
            document_id="sha256:sample",
            statement_type="CDKT",
            schema_id=4310,
            canonical_name="Tiền mặt, vàng bạc, đá quý",
            raw_value="-",
            normalized_value="0",
            current_or_comparative="CURRENT",
            status=EvidenceStatus.AUTO_VERIFIED_MEDIUM,
            value_status=ValueStatus.OBSERVED_ZERO,
            observation=ObservationKind.DASH,
        ),
        PipelineRecord(
            document_id="sha256:sample",
            statement_type="CDKT",
            schema_id=4366,
            canonical_name='Chi tiết Tài sản "Có" khác',
            raw_value="115.147.331",
            normalized_value="115147331",
            current_or_comparative="CURRENT",
            status=EvidenceStatus.AUTO_VERIFIED_MEDIUM,
            value_status=ValueStatus.OUT_OF_SCOPE_FOR_TARGET_TEMPLATE,
            observation=ObservationKind.VALUE,
        ),
    ]
    result = export_workbook(
        project_root,
        output,
        records,
        run_metadata={"run_id": "value-status"},
    )
    assert result.exported_value_count == 1

    workbook = load_workbook(output, read_only=True, data_only=True)
    sheet = workbook["CDKT"]
    by_id = {sheet.cell(row=row, column=2).value: row for row in range(2, sheet.max_row + 1)}
    zero_row = by_id[4310]
    outside_row = by_id[4366]
    assert sheet.cell(zero_row, 4).value == 0
    assert sheet.cell(zero_row, 10).value == ValueStatus.OBSERVED_ZERO.value
    assert sheet.cell(outside_row, 4).value is None
    assert sheet.cell(outside_row, 10).value == ValueStatus.OUT_OF_SCOPE_FOR_TARGET_TEMPLATE.value
    workbook.close()


def test_workbook_exports_tm_1944_value_on_its_appended_row(tmp_path, project_root):
    output = tmp_path / "tm-1944.xlsx"
    record = PipelineRecord(
        document_id="sha256:sample",
        statement_type="TM",
        schema_id=1944,
        canonical_name="Cho vay giao dịch ký quỹ và ứng trước tiền bán chứng khoán",
        raw_value="1.234",
        normalized_value="1234",
        current_or_comparative="CURRENT",
        status=EvidenceStatus.AUTO_VERIFIED_MEDIUM,
        value_status=ValueStatus.OBSERVED_VALUE,
        observation=ObservationKind.VALUE,
    )
    result = export_workbook(
        project_root,
        output,
        [record],
        run_metadata={"run_id": "tm-1944"},
    )
    assert result.exported_value_count == 1
    workbook = load_workbook(output, read_only=True, data_only=True)
    tm = workbook["TM"]
    assert tm.cell(tm.max_row, 2).value == 1944
    assert tm.cell(tm.max_row, 4).value == 1234
    workbook.close()
