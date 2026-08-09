from __future__ import annotations

import copy
import hashlib
import json
import shutil
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from bctc_ai.export import native_canonical_excel as export_module
from bctc_ai.export.native_canonical_excel import (
    EXPORT_POLICY_RELATIVE_PATH,
    ArtifactIdentity,
    NativeCanonicalExcelExportError,
    export_registered_native_canonical_excel,
    load_native_canonical_excel_policy,
    load_registered_native_canonical_excel,
)

_DOCUMENT_SHA256 = "a" * 64


def _canonical(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")


def _compact_hash(value: object) -> str:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        allow_nan=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _bbox(x0: float, y0: float, x1: float, y1: float) -> dict[str, float]:
    return {"x0": x0, "y0": y0, "x1": x1, "y1": y1}


def _cell(
    row_id: str,
    axis_id: str,
    raw_text: str,
    normalized_text: str,
    value: str | None,
    observation: str,
    source_status: str,
    *,
    index: int,
) -> dict[str, Any]:
    bbox = _bbox(300 + index * 5, 100 + index * 7, 322 + index * 5, 108 + index * 7)
    return {
        "axis_id": axis_id,
        "raw_text": raw_text,
        "normalized_text": normalized_text,
        "value": value,
        "observation": observation,
        "source_status": source_status,
        "sign_evidence": "dash" if observation == "DASH" else None,
        "parse_reason": "synthetic invalid token" if source_status == "UNRESOLVED" else None,
        "bbox": bbox,
        "run_id": f"cell-run-{index}",
        "axis_distance": index / 10,
        "provenance": {
            "document_sha256": _DOCUMENT_SHA256,
            "page": 2,
            "table_id": "synthetic-cdkt",
            "row_id": row_id,
            "column_id": axis_id,
            "value_bbox": bbox,
        },
    }


def _row(
    ordinal: int,
    label: str,
    cells: list[dict[str, Any]],
    *,
    row_type: str = "DATA_ROW",
    within_span: bool = True,
    indentation: float = 0.0,
) -> dict[str, Any]:
    row_id = f"synthetic-cdkt:row-{ordinal:04d}"
    labels = [] if not label else [_bbox(10, ordinal * 18, 240, ordinal * 18 + 10)]
    return {
        "row_id": row_id,
        "page": 2,
        "row_type": row_type,
        "source_status": "OBSERVED_ROW",
        "raw_label": label,
        "normalized_label": label,
        "label_bboxes": labels,
        "raw_note_reference": None,
        "note_reference": None,
        "note_bbox": None,
        "cells": cells,
        "y0": float(ordinal * 18),
        "y1": float(ordinal * 18 + 10),
        "indentation": indentation,
        "within_financial_table_span": within_span,
        "warnings": (
            ["outside financial table span retained for audit"]
            if not within_span
            else (["numeric row has no attached label"] if not label and cells else [])
        ),
        "provenance": {
            "document_sha256": _DOCUMENT_SHA256,
            "page": 2,
            "table_id": "synthetic-cdkt",
            "row_id": row_id,
            "label_bboxes": labels,
            "note_bbox": None,
        },
    }


def _source_join(row: dict[str, Any], rows_sha256: str) -> dict[str, Any]:
    return {
        "native_rows_sha256": rows_sha256,
        "row_id": row["row_id"],
        "source_row_sha256": _compact_hash(row),
        "source_cells_sha256": _compact_hash(row["cells"]),
        "source_cell_ids": [f"{row['row_id']}:{cell['axis_id']}" for cell in row["cells"]],
        "source_cell_count": len(row["cells"]),
    }


def _schema_item(
    schema_id: int,
    name: str,
    display_order: int,
    *,
    parent_id: int | None,
    children: list[int],
    level: int,
    scope: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "schema_id": schema_id,
        "canonical_name": name,
        "normalized_name": name,
        "statement_type": "CDKT",
        "display_order": display_order,
        "notes_section": "TÀI SẢN",
        "parent_id": parent_id,
        "children": children,
        "siblings": [],
        "previous_id": None if display_order == 0 else schema_id - 1,
        "next_id": None,
        "allowed_period_type": ["SNAPSHOT"],
        "allowed_unit": ["VND"],
        "allowed_sign": ["ANY"],
        "scope": list(scope or ["CONSOLIDATED", "SEPARATE"]),
        "cash_flow_branch": None,
        "hierarchy_level": level,
        "hierarchy_source": "SYNTHETIC_PRODUCER_SNAPSHOT",
        "structural_aliases": [],
        "source_workbook": "template/synthetic.xlsx",
        "source_row": display_order + 2,
    }


def _schema_disposition(
    item: dict[str, Any],
    outcome: str,
    *,
    source_row_id: str | None = None,
    candidate_rows: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "report_norm_id": item["schema_id"],
        "canonical_name": item["canonical_name"],
        "statement": item["statement_type"],
        "display_order": item["display_order"],
        "parent_report_norm_id": item["parent_id"],
        "hierarchy_level": item["hierarchy_level"],
        "applicable_scope": item["scope"],
        "document_reporting_scope": "UNKNOWN",
        "source_statement_scope": "MAIN_STATEMENT",
        "block_exhaustive": True,
        "block_evidence_sha256": "b" * 64,
        "terminal_outcome": outcome,
        "observation_basis": "SOURCE_ROW" if source_row_id else "EXHAUSTIVE_BLOCK",
        "source_row_id": source_row_id,
        "candidate_source_row_ids": list(candidate_rows or []),
        "source_scope_evidence": {
            "presentation_scope": "MAIN_STATEMENT",
            "pages": [2],
        },
    }


def _disposition(
    row: dict[str, Any],
    source_order: int,
    status: str,
    rows_sha256: str,
    *,
    selected_id: int | None = None,
    selected_name: str | None = None,
    candidate_ids: list[int] | None = None,
    equation_indexes: list[int] | None = None,
    conflict_indexes: list[int] | None = None,
) -> dict[str, Any]:
    return {
        "row_id": row["row_id"],
        "source_order": source_order,
        "page": row["page"],
        "page_row_order": source_order,
        "statement": "CDKT",
        "scope": "MAIN_STATEMENT",
        "within_financial_table_span": row["within_financial_table_span"],
        "row_type": row["row_type"],
        "raw_label": row["raw_label"],
        "normalized_label": row["normalized_label"],
        "indentation": row["indentation"],
        "disposition": status,
        "selected_report_norm_id": selected_id,
        "selected_canonical_name": selected_name,
        "candidate_report_norm_ids": list(candidate_ids or []),
        "match_basis": "CANONICAL_EXACT" if selected_id is not None else None,
        "matched_schema_label": selected_name,
        "alias_authority_type": None,
        "alias_authority_evidence_sha256": None,
        "source_parent_row_id": None,
        "schema_parent_report_norm_id": 100 if selected_id is not None else None,
        "schema_hierarchy_level": 1 if selected_id is not None else None,
        "schema_display_order": 1 if selected_id is not None else None,
        "equation_evidence_indexes": list(equation_indexes or []),
        "conflict_evidence_indexes": list(conflict_indexes or []),
        "source_cell_join": _source_join(row, rows_sha256),
        "reason": f"synthetic {status.lower()} evidence",
    }


def _paired_payloads(
    *,
    rows_path: str = "output/calibration/native-rows.json",
) -> tuple[dict[str, Any], dict[str, Any]]:
    headers = []
    for index, axis_id in enumerate(("value-1", "value-2", "value-3", "value-4"), start=1):
        headers.append(
            {
                "axis_id": axis_id,
                "raw_header": f"Column {index}",
                "header_bbox": _bbox(280 + index * 40, 30, 310 + index * 40, 45),
                "unit": "VND",
                "unit_multiplier": 1_000_000,
                "unit_bbox": _bbox(250, 15, 280, 25),
                "period_start": "2026-03-31",
                "period_end": "2026-03-31",
                "period_type": "SNAPSHOT",
                "duration_months": None,
                "current_or_comparative": "CURRENT" if index == 1 else "COMPARATIVE",
                "restated": False,
                "confidence": 1.0,
                "evidence": ["synthetic visible header"],
            }
        )

    row1_id = "synthetic-cdkt:row-0001"
    row2_id = "synthetic-cdkt:row-0002"
    row3_id = "synthetic-cdkt:row-0003"
    row4_id = "synthetic-cdkt:row-0004"
    row6_id = "synthetic-cdkt:row-0006"
    inside = [
        _row(
            1,
            "Existing item",
            [
                _cell(row1_id, "value-1", "123", "123", "123", "VALUE", "OBSERVED_VALUE", index=1),
                _cell(row1_id, "value-2", "0", "0", "0", "ZERO", "OBSERVED_ZERO", index=2),
                _cell(row1_id, "value-3", "-", "-", None, "DASH", "DASH", index=3),
                _cell(row1_id, "value-4", "", "", None, "BLANK", "BLANK", index=4),
            ],
        ),
        _row(
            2,
            "=New source-visible item",
            [_cell(row2_id, "value-1", "42", "42", "42", "VALUE", "OBSERVED_VALUE", index=5)],
        ),
        _row(
            3,
            "Ambiguous wording",
            [_cell(row3_id, "value-1", "=2+2", "=2+2", None, "INVALID", "UNRESOLVED", index=6)],
        ),
        _row(
            4,
            "",
            [_cell(row4_id, "value-1", "-", "-", None, "DASH", "DASH", index=7)],
        ),
        _row(5, "Repeated heading", [], row_type="SECTION_HEADER"),
    ]
    outside = [
        _row(
            6,
            "Outside retained evidence",
            [_cell(row6_id, "value-1", "9", "9", "9", "VALUE", "OBSERVED_VALUE", index=8)],
            within_span=False,
        )
    ]
    page = {
        "page": 2,
        "statement_type": "CDKT",
        "scope": "MAIN_STATEMENT",
        "discovery_contract": {
            "page": 2,
            "statement_type": "CDKT",
            "scope": "MAIN_STATEMENT",
            "mapping_eligible": True,
            "locally_accepted": True,
        },
        "text_quality": "USABLE_TEXT_LAYER",
        "corruption_markers": [],
        "width_points": 600.0,
        "height_points": 800.0,
        "rotation": 0,
        "native_word_count": 30,
        "native_words_sha256": "c" * 64,
        "geometry": {
            "authority": "PYMUPDF_NATIVE_TEXT_WORDS",
            "axes": [
                {
                    "axis_id": header["axis_id"],
                    "role": "VALUE",
                    "right_edge": 350.0 + index * 40,
                    "left_edge": 320.0 + index * 40,
                    "sample_count": 2,
                    "source": "native",
                }
                for index, header in enumerate(headers)
            ],
        },
        "headers": headers,
        "rows": inside,
        "outside_financial_table_span_rows": outside,
        "reconstructed_row_count": 6,
        "financial_table_span_row_count": 5,
        "outside_financial_table_span_row_count": 1,
    }
    source = {
        "document_id": f"sha256:{_DOCUMENT_SHA256}",
        "relative_path": "data/synthetic.pdf",
        "sha256": _DOCUMENT_SHA256,
        "size_bytes": 123,
        "bank": "SYNTHETIC",
        "year": 2026,
        "dataset_role": "CALIBRATION",
        "registry_state": "REGISTERED",
        "hash_verified_stable": True,
        "immutable_role_assignment": True,
    }
    rows_payload = {
        "format_version": "REGISTERED_NATIVE_STATEMENT_ROWS_RESULT_V1",
        "policy": "REGISTERED_NATIVE_STATEMENT_ROWS_V1",
        "claim_boundary": "UNMAPPED_SOURCE_ROWS_AND_CELLS_ONLY",
        "status": "ACCEPTED_NATIVE_STATEMENT_ROWS",
        "run_id": "synthetic-native-rows",
        "source": source,
        "statement_discovery": {
            "path": "output/calibration/discovery.json",
            "sha256": "d" * 64,
        },
        "code": {"commit": "e" * 40, "dirty": False, "implementation": []},
        "authority": {"schema_mapper": None},
        "isolation": {"schema_inputs_loaded": False},
        "inputs": {"runtime_read_ledger": []},
        "selection": {"selected_pages": [2]},
        "summary": {"page_count": 1, "financial_table_span_row_count": 5},
        "pages": [page],
    }
    rows_bytes = _canonical(rows_payload)
    rows_sha256 = hashlib.sha256(rows_bytes).hexdigest()

    source_dispositions = [
        _disposition(
            inside[0],
            1,
            "EXISTING_ITEM",
            rows_sha256,
            selected_id=101,
            selected_name="Existing item",
            candidate_ids=[101],
            equation_indexes=[0],
        ),
        _disposition(inside[1], 2, "NEW_ITEM_PROPOSAL", rows_sha256),
        _disposition(
            inside[2],
            3,
            "AMBIGUOUS",
            rows_sha256,
            candidate_ids=[102],
            conflict_indexes=[0],
        ),
        _disposition(inside[3], 4, "UNRESOLVED", rows_sha256, equation_indexes=[0]),
        _disposition(inside[4], 5, "STRUCTURAL", rows_sha256),
        _disposition(outside[0], 6, "UNRESOLVED", rows_sha256),
    ]
    schema_items = [
        _schema_item(100, "Assets", 0, parent_id=None, children=[101, 102, 103], level=0),
        _schema_item(101, "Existing item", 1, parent_id=100, children=[], level=1),
        _schema_item(102, "Ambiguous target", 2, parent_id=100, children=[], level=1),
        _schema_item(103, "Not observed item", 3, parent_id=100, children=[], level=1),
    ]
    schema_dispositions = [
        _schema_disposition(schema_items[0], "BLANK"),
        _schema_disposition(schema_items[1], "OBSERVED_VALUE", source_row_id=inside[0]["row_id"]),
        _schema_disposition(schema_items[2], "AMBIGUOUS", candidate_rows=[inside[2]["row_id"]]),
        _schema_disposition(schema_items[3], "NOT_OBSERVED"),
    ]
    proposal = {
        "proposal_key": "schema-gap-synthetic",
        "status": "PROPOSED_NEEDS_USER_REVIEW",
        "report_norm_id": None,
        "canonical_label": "=New source-visible item",
        "statement": "CDKT",
        "section": "Assets",
        "parent": {
            "kind": "EXISTING_ITEM",
            "proposal_key": None,
            "report_norm_id": 100,
            "canonical_name": "Assets",
        },
        "hierarchy_level": 1,
        "display_order_anchors": {
            "insert_after_report_norm_id": 101,
            "insert_before_report_norm_id": 102,
        },
        "source_evidence": {
            "row_id": inside[1]["row_id"],
            "page": 2,
            "visible_label": inside[1]["raw_label"],
            "values": [{"axis_id": "value-1", "value": "42"}],
        },
        "reason_existing_items_are_insufficient": "no equivalent accounting concept",
        "nearest_existing_candidates_diagnostic_only": [],
        "equation_evidence": {"status": "NOT_APPLICABLE"},
        "possible_aliases": ["New visible item"],
        "allocation_authority": False,
    }
    accepted_aliases = [
        {
            "statement_type": "CDKT",
            "report_norm_id": 101,
            "alias": "Existing wording",
            "authority_type": "AUDITED_SCHEMA_ALIAS",
            "authority_evidence_sha256": "f" * 64,
            "authority_evidence": {"source": "synthetic audit"},
        }
    ]
    producer_snapshots = {
        "policy": {
            "path": "config/mapping/native-canonical-v1.yaml",
            "source_sha256": "1" * 64,
            "canonical_payload_sha256": "2" * 64,
            "payload": {"policy": "REGISTERED_NATIVE_CANONICAL_MAPPING_V1"},
        },
        "schema": {"items_sha256": _compact_hash(schema_items), "items": schema_items},
        "accepted_aliases": {
            "records_sha256": _compact_hash(accepted_aliases),
            "records": accepted_aliases,
        },
        "coverage": {
            "payload_sha256": "3" * 64,
            "payload": {"status": "ACTIVE", "padding": "x" * 35_000},
        },
        "cash_flow_rules": {
            "payload_sha256": "4" * 64,
            "payload": {"method": "UNKNOWN"},
        },
    }
    disposition_counts = {
        status: sum(record["disposition"] == status for record in source_dispositions)
        for status in (
            "EXISTING_ITEM",
            "NEW_ITEM_PROPOSAL",
            "AMBIGUOUS",
            "UNRESOLVED",
            "STRUCTURAL",
        )
    }
    mapping_payload = {
        "format_version": "REGISTERED_NATIVE_CANONICAL_MAPPING_RESULT_V1",
        "policy": "REGISTERED_NATIVE_CANONICAL_MAPPING_V1",
        "claim_boundary": "SOURCE_ROW_CANONICAL_DISPOSITIONS_AND_SCHEMA_GAP_PROPOSALS",
        "status": "ACCEPTED_NATIVE_CANONICAL_MAPPING",
        "run_id": "synthetic-native-canonical",
        "source": source,
        "native_rows": {
            "path": rows_path,
            "sha256": rows_sha256,
            "size_bytes": len(rows_bytes),
            "format_version": rows_payload["format_version"],
            "policy": rows_payload["policy"],
            "claim_boundary": rows_payload["claim_boundary"],
            "status": rows_payload["status"],
            "run_id": rows_payload["run_id"],
            "producer_git_commit": rows_payload["code"]["commit"],
            "denominator": "ALL_RECONSTRUCTED_SOURCE_ROWS",
        },
        "schema": {"global_item_count": len(schema_items), "global_high_water_mark": 103},
        "producer_snapshots": producer_snapshots,
        "code": {"commit": "5" * 40, "dirty": False, "implementation": []},
        "authority": {"source_rows_and_cells": "TRUSTED_SHA256_JOIN"},
        "isolation": {
            "new_item_proposals_allocate_report_norm_id": False,
            "same_run_alias_proposals_mapping_eligible": False,
        },
        "inputs": {"runtime_read_ledger": []},
        "lctt_method": {
            "method": "UNKNOWN",
            "reason": "no LCTT block",
            "semantic_high_confidence_allowed": False,
        },
        "path_summaries": {
            statement: {
                "statement": statement,
                "source_row_count": 5 if statement == "CDKT" else 0,
                "maximum_cardinality": 1 if statement == "CDKT" else 0,
                "projection_item_count": len(schema_items) if statement == "CDKT" else 0,
                "projection_sha256": str(index) * 64,
            }
            for index, statement in enumerate(("CDKT", "KQKD", "LCTT", "TM"), start=6)
        },
        "source_dispositions": source_dispositions,
        "new_item_proposals": [proposal],
        "alias_proposals": [
            {
                "proposal_key": "alias-synthetic",
                "row_id": inside[2]["row_id"],
                "candidate_report_norm_id": 102,
                "proposed_alias": "Ambiguous wording",
                "proposal_type": "SOURCE_WORDING",
                "mapping_eligible_this_run": False,
                "status": "PROPOSED_NEEDS_REVIEW",
            }
        ],
        "equations": [
            {
                "equation_type": "SYNTHETIC_CORROBORATION",
                "row_id": inside[0]["row_id"],
                "affected_row_ids": [inside[0]["row_id"], inside[3]["row_id"]],
                "status": "PASS",
                "selection_authority": False,
            }
        ],
        "conflicts": [
            {
                "conflict_type": "SYNTHETIC_LOCAL_CONFLICT",
                "affected_row_ids": [inside[2]["row_id"]],
                "candidate_report_norm_ids": [102],
                "resolution": "DEMOTE_AFFECTED_ROW_TO_AMBIGUOUS",
            }
        ],
        "schema_dispositions": schema_dispositions,
        "mandatory_search": {
            "status": "PASS",
            "statement": "ALL",
            "evaluated_schema_item_count": len(schema_items),
        },
        "summary": {
            "visible_source_items": 6,
            "source_disposition_counts": disposition_counts,
            "new_schema_item_proposals": 1,
            "universal_schema_item_count": len(schema_items),
        },
    }
    return mapping_payload, rows_payload


def _identities(
    mapping_payload: dict[str, Any], rows_payload: dict[str, Any]
) -> tuple[ArtifactIdentity, ArtifactIdentity]:
    mapping_bytes = _canonical(mapping_payload)
    rows_bytes = _canonical(rows_payload)
    return (
        ArtifactIdentity(
            "output/calibration/native-canonical.json",
            hashlib.sha256(mapping_bytes).hexdigest(),
            len(mapping_bytes),
        ),
        ArtifactIdentity(
            "output/calibration/native-rows.json",
            hashlib.sha256(rows_bytes).hexdigest(),
            len(rows_bytes),
        ),
    )


def _headers(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def _build_artifacts(
    mapping_payload: dict[str, Any],
    rows_payload: dict[str, Any],
    *,
    mapping_identity: ArtifactIdentity,
    rows_identity: ArtifactIdentity,
    workbook_filename: str,
    policy,
    implementation_ledger,
):
    return export_module._build_prevalidated_native_canonical_excel_artifacts(
        mapping_payload,
        rows_payload,
        mapping_identity=mapping_identity,
        rows_identity=rows_identity,
        workbook_relative_path=f"output/calibration/export/{workbook_filename}",
        provenance_relative_path=(
            "output/calibration/export/"
            f"{Path(workbook_filename).with_suffix('.provenance.json').name}"
        ),
        policy=policy,
        implementation_ledger=implementation_ledger,
    )


_TEST_IMPLEMENTATION = ({"path": "src/exporter.py", "sha256": "7" * 64, "size_bytes": 10},)


def _write_completed_fixture(
    root: Path,
    source_project_root: Path,
    mapping_payload: dict[str, Any],
    rows_payload: dict[str, Any],
) -> dict[str, Any]:
    root.mkdir(parents=True, exist_ok=True)
    policy_path = root / EXPORT_POLICY_RELATIVE_PATH
    policy_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_project_root / EXPORT_POLICY_RELATIVE_PATH, policy_path)
    mapping_identity, rows_identity = _identities(mapping_payload, rows_payload)
    mapping_path = root / mapping_identity.path
    rows_path = root / rows_identity.path
    mapping_path.parent.mkdir(parents=True, exist_ok=True)
    mapping_path.write_bytes(_canonical(mapping_payload))
    rows_path.write_bytes(_canonical(rows_payload))
    policy = load_native_canonical_excel_policy(policy_path, root)
    workbook_relative = Path("output/calibration/export/native-canonical.xlsx")
    provenance_relative = Path("output/calibration/export/native-canonical.provenance.json")
    artifacts = export_module._build_prevalidated_native_canonical_excel_artifacts(
        mapping_payload,
        rows_payload,
        mapping_identity=mapping_identity,
        rows_identity=rows_identity,
        workbook_relative_path=workbook_relative.as_posix(),
        provenance_relative_path=provenance_relative.as_posix(),
        policy=policy,
        implementation_ledger=_TEST_IMPLEMENTATION,
    )
    workbook_path = root / workbook_relative
    provenance_path = root / provenance_relative
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_bytes(artifacts.workbook_bytes)
    provenance_path.write_bytes(artifacts.provenance_bytes)
    return {
        "mapping_payload": mapping_payload,
        "rows_payload": rows_payload,
        "mapping_identity": mapping_identity,
        "rows_identity": rows_identity,
        "workbook_relative": workbook_relative,
        "provenance_relative": provenance_relative,
        "workbook_path": workbook_path,
        "provenance_path": provenance_path,
        "workbook_sha256": artifacts.workbook_sha256,
        "provenance_sha256": artifacts.provenance_sha256,
        "artifacts": artifacts,
    }


def _mock_strict_pair_dependencies(
    monkeypatch: pytest.MonkeyPatch,
    mapping_payload: dict[str, Any],
    rows_payload: dict[str, Any],
) -> None:
    def fake_mapping_loader(
        path: Path,
        *,
        project_root: Path,
        expected_sha256: str,
        policy_path: Path | None = None,
        rows_policy_path: Path | None = None,
    ) -> dict[str, Any]:
        del policy_path, rows_policy_path
        assert path.relative_to(project_root).as_posix() == (
            mapping_payload["native_rows"]["path"].replace("native-rows", "native-canonical")
        )
        assert hashlib.sha256(path.read_bytes()).hexdigest() == expected_sha256
        return mapping_payload

    def fake_rows_loader(
        path: Path,
        *,
        project_root: Path,
        expected_sha256: str,
        policy_path: Path | None = None,
    ) -> dict[str, Any]:
        del policy_path
        assert path.relative_to(project_root).as_posix() == mapping_payload["native_rows"]["path"]
        assert hashlib.sha256(path.read_bytes()).hexdigest() == expected_sha256
        return rows_payload

    monkeypatch.setattr(
        export_module,
        "load_registered_native_canonical_mapping",
        fake_mapping_loader,
    )
    monkeypatch.setattr(
        export_module,
        "load_registered_native_statement_rows",
        fake_rows_loader,
    )
    monkeypatch.setattr(
        export_module,
        "_implementation_ledger",
        lambda _root: _TEST_IMPLEMENTATION,
    )


def test_paired_workbook_is_deterministic_lossless_and_formula_free(project_root: Path):
    mapping_payload, rows_payload = _paired_payloads()
    mapping_identity, rows_identity = _identities(mapping_payload, rows_payload)
    policy = load_native_canonical_excel_policy(
        project_root / EXPORT_POLICY_RELATIVE_PATH, project_root
    )
    kwargs = {
        "mapping_identity": mapping_identity,
        "rows_identity": rows_identity,
        "workbook_filename": "native-canonical.xlsx",
        "policy": policy,
        "implementation_ledger": (
            {"path": "src/exporter.py", "sha256": "7" * 64, "size_bytes": 10},
        ),
    }
    first = _build_artifacts(mapping_payload, rows_payload, **kwargs)
    second = _build_artifacts(mapping_payload, rows_payload, **kwargs)
    assert first.workbook_bytes == second.workbook_bytes
    assert first.provenance_bytes == second.provenance_bytes

    workbook = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    try:
        assert workbook.sheetnames == [
            "SOURCE_ROWS",
            "CELLS",
            "SCHEMA_COVERAGE",
            "NEW_ITEM_PROPOSALS",
            "VALIDATION",
            "RUN_METADATA",
        ]
        source_rows = workbook["SOURCE_ROWS"]
        source_headers = _headers(source_rows)
        assert source_rows.max_row == 7
        assert [
            source_rows.cell(row, source_headers["Disposition"]).value
            for row in range(2, source_rows.max_row + 1)
        ] == [
            "EXISTING_ITEM",
            "NEW_ITEM_PROPOSAL",
            "AMBIGUOUS",
            "UNRESOLVED",
            "STRUCTURAL",
            "UNRESOLVED",
        ]
        assert source_rows.cell(3, source_headers["RawLabel"]).value == ("=New source-visible item")
        assert source_rows.cell(3, source_headers["RawLabel"]).data_type == "s"
        assert source_rows.cell(7, source_headers["SourceBucket"]).value == (
            "OUTSIDE_FINANCIAL_TABLE_SPAN"
        )
        assert source_rows.cell(7, source_headers["WithinFinancialTableSpan"]).value is False

        cells = workbook["CELLS"]
        cell_headers = _headers(cells)
        statuses = [
            cells.cell(row, cell_headers["SourceStatus"]).value
            for row in range(2, cells.max_row + 1)
        ]
        assert {"OBSERVED_VALUE", "OBSERVED_ZERO", "DASH", "BLANK", "UNRESOLVED"}.issubset(statuses)
        by_raw = {
            cells.cell(row, cell_headers["RawText"]).value: row
            for row in range(2, cells.max_row + 1)
        }
        assert cells.cell(by_raw["123"], cell_headers["ValueText"]).value == "123"
        assert cells.cell(by_raw["123"], cell_headers["ValueJson"]).value == '"123"'
        assert cells.cell(by_raw["0"], cell_headers["ValueJson"]).value == '"0"'
        assert cells.cell(by_raw["-"], cell_headers["ValueText"]).value is None
        assert cells.cell(by_raw["=2+2"], cell_headers["ValueJson"]).value == "null"
        assert cells.cell(by_raw["=2+2"], cell_headers["RawText"]).data_type == "s"

        coverage = workbook["SCHEMA_COVERAGE"]
        coverage_headers = _headers(coverage)
        assert [
            coverage.cell(row, coverage_headers["ReportNormId"]).value
            for row in range(2, coverage.max_row + 1)
        ] == [100, 101, 102, 103]
        by_id = {
            coverage.cell(row, coverage_headers["ReportNormId"]).value: row
            for row in range(2, coverage.max_row + 1)
        }
        assert coverage.cell(by_id[101], coverage_headers["ParentReportNormId"]).value == 100
        assert coverage.cell(by_id[102], coverage_headers["TerminalOutcome"]).value == ("AMBIGUOUS")
        assert coverage.cell(by_id[103], coverage_headers["TerminalOutcome"]).value == (
            "NOT_OBSERVED"
        )

        proposals = workbook["NEW_ITEM_PROPOSALS"]
        proposal_headers = _headers(proposals)
        assert proposals.max_row == 2
        assert proposals.cell(2, proposal_headers["ProposalKey"]).value == ("schema-gap-synthetic")
        assert proposals.cell(2, proposal_headers["AllocatedReportNormId"]).value is None
        assert proposals.cell(2, proposal_headers["ParentReportNormId"]).value == 100

        validation = workbook["VALIDATION"]
        validation_headers = _headers(validation)
        coverage_rows = [
            row
            for row in range(2, validation.max_row + 1)
            if validation.cell(row, validation_headers["RecordType"]).value
            == "PRODUCER_COVERAGE_SNAPSHOT"
        ]
        assert len(coverage_rows) == 2
        assert [
            validation.cell(row, validation_headers["PartNumber"]).value for row in coverage_rows
        ] == [1, 2]
        assert not any(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()

    receipt = json.loads(first.provenance_bytes)
    assert receipt["inputs"]["mapping"]["sha256"] == mapping_identity.sha256
    assert receipt["inputs"]["native_rows"]["sha256"] == rows_identity.sha256
    assert receipt["inputs"]["exact_pair_verified"] is True
    assert receipt["summary"]["source_row_count"] == 6
    assert receipt["summary"]["source_cell_count"] == 8
    assert receipt["summary"]["source_disposition_counts"] == {
        "AMBIGUOUS": 1,
        "EXISTING_ITEM": 1,
        "NEW_ITEM_PROPOSAL": 1,
        "STRUCTURAL": 1,
        "UNRESOLVED": 2,
    }
    assert receipt["summary"]["formula_count"] == 0
    assert receipt["summary"]["imputed_cell_count"] == 0
    assert receipt["summary"]["allocated_report_norm_id_count"] == 0
    assert receipt["isolation"]["current_mutable_schema_loaded"] is False
    assert receipt["workbook"]["sha256"] == hashlib.sha256(first.workbook_bytes).hexdigest()


def test_future_producer_schema_snapshot_is_used_without_current_schema(project_root: Path):
    mapping_payload, rows_payload = _paired_payloads()
    future = _schema_item(
        999,
        "Future bank-specific item",
        4,
        parent_id=100,
        children=[],
        level=1,
        scope=["CONSOLIDATED"],
    )
    items = mapping_payload["producer_snapshots"]["schema"]["items"]
    items.append(future)
    items[0]["children"].append(999)
    mapping_payload["producer_snapshots"]["schema"]["items_sha256"] = _compact_hash(items)
    mapping_payload["schema_dispositions"].append(_schema_disposition(future, "NOT_APPLICABLE"))
    mapping_payload["schema"]["global_item_count"] = 5
    mapping_payload["schema"]["global_high_water_mark"] = 999
    mapping_payload["summary"]["universal_schema_item_count"] = 5
    mapping_identity, rows_identity = _identities(mapping_payload, rows_payload)
    policy = load_native_canonical_excel_policy(
        project_root / EXPORT_POLICY_RELATIVE_PATH, project_root
    )
    artifact = _build_artifacts(
        mapping_payload,
        rows_payload,
        mapping_identity=mapping_identity,
        rows_identity=rows_identity,
        workbook_filename="future.xlsx",
        policy=policy,
        implementation_ledger=(),
    )
    workbook = load_workbook(BytesIO(artifact.workbook_bytes), data_only=False)
    try:
        sheet = workbook["SCHEMA_COVERAGE"]
        headers = _headers(sheet)
        last = sheet.max_row
        assert sheet.cell(last, headers["ReportNormId"]).value == 999
        assert sheet.cell(last, headers["CanonicalName"]).value == ("Future bank-specific item")
        assert sheet.cell(last, headers["ApplicableScopeJson"]).value == ('["CONSOLIDATED"]')
        assert sheet.cell(last, headers["TerminalOutcome"]).value == "NOT_APPLICABLE"
    finally:
        workbook.close()


def test_pair_join_schema_and_proposal_mutations_fail_closed(project_root: Path):
    mapping_payload, rows_payload = _paired_payloads()
    policy = load_native_canonical_excel_policy(
        project_root / EXPORT_POLICY_RELATIVE_PATH, project_root
    )

    tampered_rows = copy.deepcopy(rows_payload)
    tampered_rows["pages"][0]["rows"][0]["cells"][0]["source_status"] = "OBSERVED_ZERO"
    tampered_rows_bytes = _canonical(tampered_rows)
    tampered_rows_identity = ArtifactIdentity(
        "output/calibration/native-rows.json",
        hashlib.sha256(tampered_rows_bytes).hexdigest(),
        len(tampered_rows_bytes),
    )
    tampered_mapping = copy.deepcopy(mapping_payload)
    tampered_mapping["native_rows"]["sha256"] = tampered_rows_identity.sha256
    tampered_mapping["native_rows"]["size_bytes"] = tampered_rows_identity.size_bytes
    tampered_mapping_identity, _ = _identities(tampered_mapping, rows_payload)
    with pytest.raises(NativeCanonicalExcelExportError, match="cell hash join"):
        _build_artifacts(
            tampered_mapping,
            tampered_rows,
            mapping_identity=tampered_mapping_identity,
            rows_identity=tampered_rows_identity,
            workbook_filename="tampered.xlsx",
            policy=policy,
            implementation_ledger=(),
        )

    unknown_candidate = copy.deepcopy(mapping_payload)
    unknown_candidate["source_dispositions"][2]["candidate_report_norm_ids"] = [999]
    mapping_identity, rows_identity = _identities(unknown_candidate, rows_payload)
    with pytest.raises(NativeCanonicalExcelExportError, match="unknown schema"):
        _build_artifacts(
            unknown_candidate,
            rows_payload,
            mapping_identity=mapping_identity,
            rows_identity=rows_identity,
            workbook_filename="unknown.xlsx",
            policy=policy,
            implementation_ledger=(),
        )

    allocated = copy.deepcopy(mapping_payload)
    allocated["new_item_proposals"][0]["report_norm_id"] = 104
    mapping_identity, rows_identity = _identities(allocated, rows_payload)
    with pytest.raises(NativeCanonicalExcelExportError, match="allocated a ReportNormId"):
        _build_artifacts(
            allocated,
            rows_payload,
            mapping_identity=mapping_identity,
            rows_identity=rows_identity,
            workbook_filename="allocated.xlsx",
            policy=policy,
            implementation_ledger=(),
        )


def test_strict_publication_requires_both_hashes_same_role_and_never_overwrites(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    root = tmp_path.resolve()
    policy_path = root / EXPORT_POLICY_RELATIVE_PATH
    policy_path.parent.mkdir(parents=True)
    shutil.copyfile(project_root / EXPORT_POLICY_RELATIVE_PATH, policy_path)
    mapping_payload, rows_payload = _paired_payloads()
    mapping_bytes = _canonical(mapping_payload)
    rows_bytes = _canonical(rows_payload)
    mapping_sha256 = hashlib.sha256(mapping_bytes).hexdigest()
    rows_sha256 = hashlib.sha256(rows_bytes).hexdigest()
    mapping_path = root / "output/calibration/native-canonical.json"
    rows_path = root / "output/calibration/native-rows.json"
    mapping_path.parent.mkdir(parents=True)
    mapping_path.write_bytes(mapping_bytes)
    rows_path.write_bytes(rows_bytes)
    calls: list[tuple[str, str]] = []

    def fake_mapping_loader(
        path: Path,
        *,
        project_root: Path,
        expected_sha256: str,
        policy_path: Path | None = None,
        rows_policy_path: Path | None = None,
    ) -> dict[str, Any]:
        del project_root, policy_path, rows_policy_path
        assert path == mapping_path
        calls.append(("mapping", expected_sha256))
        return mapping_payload

    def fake_rows_loader(
        path: Path,
        *,
        project_root: Path,
        expected_sha256: str,
        policy_path: Path | None = None,
    ) -> dict[str, Any]:
        del project_root, policy_path
        assert path == rows_path
        calls.append(("rows", expected_sha256))
        return rows_payload

    implementation = ({"path": "src/exporter.py", "sha256": "7" * 64, "size_bytes": 10},)
    monkeypatch.setattr(
        export_module, "load_registered_native_canonical_mapping", fake_mapping_loader
    )
    monkeypatch.setattr(export_module, "load_registered_native_statement_rows", fake_rows_loader)
    monkeypatch.setattr(export_module, "_implementation_ledger", lambda _root: implementation)
    workbook_path = root / "output/calibration/export/native-canonical.xlsx"
    provenance_path = root / "output/calibration/export/native-canonical.provenance.json"
    result = export_registered_native_canonical_excel(
        project_root=root,
        mapping_path=mapping_path,
        mapping_expected_sha256=mapping_sha256,
        rows_path=rows_path,
        rows_expected_sha256=rows_sha256,
        workbook_path=workbook_path,
        provenance_path=provenance_path,
        export_policy_path=policy_path,
    )
    assert calls == [
        ("mapping", mapping_sha256),
        ("rows", rows_sha256),
        ("mapping", mapping_sha256),
        ("rows", rows_sha256),
    ]
    assert result.workbook_sha256 == hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    assert result.provenance_sha256 == hashlib.sha256(provenance_path.read_bytes()).hexdigest()
    pair = (workbook_path.read_bytes(), provenance_path.read_bytes())

    with pytest.raises(NativeCanonicalExcelExportError, match="overwrite"):
        export_registered_native_canonical_excel(
            project_root=root,
            mapping_path=mapping_path,
            mapping_expected_sha256=mapping_sha256,
            rows_path=rows_path,
            rows_expected_sha256=rows_sha256,
            workbook_path=workbook_path,
            provenance_path=provenance_path,
            export_policy_path=policy_path,
        )
    assert (workbook_path.read_bytes(), provenance_path.read_bytes()) == pair

    with pytest.raises(NativeCanonicalExcelExportError, match="must stay under"):
        export_registered_native_canonical_excel(
            project_root=root,
            mapping_path=mapping_path,
            mapping_expected_sha256=mapping_sha256,
            rows_path=rows_path,
            rows_expected_sha256=rows_sha256,
            workbook_path=root / "output/development/cross-role.xlsx",
            provenance_path=root / "output/development/cross-role.json",
            export_policy_path=policy_path,
        )

    with pytest.raises(NativeCanonicalExcelExportError, match="trusted native-row"):
        export_registered_native_canonical_excel(
            project_root=root,
            mapping_path=mapping_path,
            mapping_expected_sha256=mapping_sha256,
            rows_path=rows_path,
            rows_expected_sha256="bad",
            workbook_path=root / "output/calibration/bad.xlsx",
            provenance_path=root / "output/calibration/bad.json",
            export_policy_path=policy_path,
        )


def test_publication_rolls_back_workbook_if_completion_marker_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    workbook_path = tmp_path / "canonical.xlsx"
    provenance_path = tmp_path / "canonical.json"
    original = export_module._write_exclusive
    calls = 0

    def fail_second(path: Path, payload: bytes):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise NativeCanonicalExcelExportError("synthetic provenance failure")
        return original(path, payload)

    monkeypatch.setattr(export_module, "_write_exclusive", fail_second)
    with pytest.raises(NativeCanonicalExcelExportError, match="synthetic provenance"):
        export_module._publish_pair(workbook_path, provenance_path, b"workbook", b"provenance")
    assert not workbook_path.exists()
    assert not provenance_path.exists()


def test_deprecated_proposal_summary_and_public_accepted_builder_are_rejected(
    project_root: Path,
):
    assert not hasattr(export_module, "build_native_canonical_excel_artifacts")
    assert "build_native_canonical_excel_artifacts" not in export_module.__all__
    mapping_payload, rows_payload = _paired_payloads()
    mapping_payload["summary"]["new_schema_items_discovered"] = 1
    mapping_identity, rows_identity = _identities(mapping_payload, rows_payload)
    policy = load_native_canonical_excel_policy(
        project_root / EXPORT_POLICY_RELATIVE_PATH, project_root
    )
    with pytest.raises(NativeCanonicalExcelExportError, match="deprecated"):
        _build_artifacts(
            mapping_payload,
            rows_payload,
            mapping_identity=mapping_identity,
            rows_identity=rows_identity,
            workbook_filename="deprecated.xlsx",
            policy=policy,
            implementation_ledger=_TEST_IMPLEMENTATION,
        )


def test_completed_pair_strict_replay_supports_future_schema_and_relocated_root(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    mapping_payload, rows_payload = _paired_payloads()
    future = _schema_item(
        999,
        "Future bank-specific item",
        4,
        parent_id=100,
        children=[],
        level=1,
        scope=["CONSOLIDATED"],
    )
    items = mapping_payload["producer_snapshots"]["schema"]["items"]
    items.append(future)
    items[0]["children"].append(999)
    mapping_payload["producer_snapshots"]["schema"]["items_sha256"] = _compact_hash(items)
    mapping_payload["schema_dispositions"].append(_schema_disposition(future, "NOT_APPLICABLE"))
    mapping_payload["schema"]["global_item_count"] = 5
    mapping_payload["schema"]["global_high_water_mark"] = 999
    mapping_payload["summary"]["universal_schema_item_count"] = 5
    first = _write_completed_fixture(
        tmp_path / "root-one", project_root, mapping_payload, rows_payload
    )
    second = _write_completed_fixture(
        tmp_path / "relocated-root", project_root, mapping_payload, rows_payload
    )
    _mock_strict_pair_dependencies(monkeypatch, mapping_payload, rows_payload)
    assert first["workbook_path"].read_bytes() == second["workbook_path"].read_bytes()
    assert first["provenance_path"].read_bytes() == second["provenance_path"].read_bytes()

    for root, fixture in (
        (tmp_path / "root-one", first),
        (tmp_path / "relocated-root", second),
    ):
        loaded = load_registered_native_canonical_excel(
            project_root=root,
            workbook_path=fixture["workbook_relative"],
            workbook_expected_sha256=fixture["workbook_sha256"],
            provenance_path=fixture["provenance_relative"],
            provenance_expected_sha256=fixture["provenance_sha256"],
        )
        assert loaded.summary["producer_schema_item_count"] == 5
        assert loaded.workbook_path == fixture["workbook_path"]

    with pytest.raises(NativeCanonicalExcelExportError, match="relative"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "root-one",
            workbook_path=first["workbook_path"],
            workbook_expected_sha256=first["workbook_sha256"],
            provenance_path=first["provenance_relative"],
            provenance_expected_sha256=first["provenance_sha256"],
        )
    with pytest.raises(NativeCanonicalExcelExportError, match="distinct siblings"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "root-one",
            workbook_path=first["workbook_relative"],
            workbook_expected_sha256=first["workbook_sha256"],
            provenance_path=Path("output/calibration/elsewhere/native-canonical.provenance.json"),
            provenance_expected_sha256=first["provenance_sha256"],
        )


def test_completed_pair_receipt_workbook_role_and_symlink_mutations_fail_closed(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    mapping_payload, rows_payload = _paired_payloads()
    _mock_strict_pair_dependencies(monkeypatch, mapping_payload, rows_payload)

    def fixture(name: str) -> dict[str, Any]:
        return _write_completed_fixture(
            tmp_path / name,
            project_root,
            copy.deepcopy(mapping_payload),
            copy.deepcopy(rows_payload),
        )

    extra = fixture("extra-field")
    extra_receipt = json.loads(extra["provenance_path"].read_bytes())
    extra_receipt["unexpected"] = True
    extra_bytes = _canonical(extra_receipt)
    extra["provenance_path"].write_bytes(extra_bytes)
    with pytest.raises(NativeCanonicalExcelExportError, match="fields drifted"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "extra-field",
            workbook_path=extra["workbook_relative"],
            workbook_expected_sha256=extra["workbook_sha256"],
            provenance_path=extra["provenance_relative"],
            provenance_expected_sha256=hashlib.sha256(extra_bytes).hexdigest(),
        )

    missing = fixture("missing-field")
    missing_receipt = json.loads(missing["provenance_path"].read_bytes())
    missing_receipt.pop("summary")
    missing_bytes = _canonical(missing_receipt)
    missing["provenance_path"].write_bytes(missing_bytes)
    with pytest.raises(NativeCanonicalExcelExportError, match="fields drifted"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "missing-field",
            workbook_path=missing["workbook_relative"],
            workbook_expected_sha256=missing["workbook_sha256"],
            provenance_path=missing["provenance_relative"],
            provenance_expected_sha256=hashlib.sha256(missing_bytes).hexdigest(),
        )

    trusted_hash = fixture("trusted-hash")
    with pytest.raises(NativeCanonicalExcelExportError, match="trusted SHA-256"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "trusted-hash",
            workbook_path=trusted_hash["workbook_relative"],
            workbook_expected_sha256="9" * 64,
            provenance_path=trusted_hash["provenance_relative"],
            provenance_expected_sha256=trusted_hash["provenance_sha256"],
        )

    ledger = fixture("ledger")
    ledger_receipt = json.loads(ledger["provenance_path"].read_bytes())
    ledger_receipt["code"]["exporter_implementation"][0]["sha256"] = "8" * 64
    ledger_bytes = _canonical(ledger_receipt)
    ledger["provenance_path"].write_bytes(ledger_bytes)
    with pytest.raises(NativeCanonicalExcelExportError, match="implementation ledger drifted"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "ledger",
            workbook_path=ledger["workbook_relative"],
            workbook_expected_sha256=ledger["workbook_sha256"],
            provenance_path=ledger["provenance_relative"],
            provenance_expected_sha256=hashlib.sha256(ledger_bytes).hexdigest(),
        )

    run = fixture("run")
    run_receipt = json.loads(run["provenance_path"].read_bytes())
    run_receipt["run_id"] = "different-run"
    run_bytes = _canonical(run_receipt)
    run["provenance_path"].write_bytes(run_bytes)
    with pytest.raises(NativeCanonicalExcelExportError, match="run/source/role"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "run",
            workbook_path=run["workbook_relative"],
            workbook_expected_sha256=run["workbook_sha256"],
            provenance_path=run["provenance_relative"],
            provenance_expected_sha256=hashlib.sha256(run_bytes).hexdigest(),
        )

    workbook = fixture("workbook")
    changed_workbook = workbook["workbook_path"].read_bytes() + b"tampered"
    workbook["workbook_path"].write_bytes(changed_workbook)
    workbook_receipt = json.loads(workbook["provenance_path"].read_bytes())
    workbook_receipt["workbook"]["sha256"] = hashlib.sha256(changed_workbook).hexdigest()
    workbook_receipt["workbook"]["size_bytes"] = len(changed_workbook)
    workbook_receipt_bytes = _canonical(workbook_receipt)
    workbook["provenance_path"].write_bytes(workbook_receipt_bytes)
    with pytest.raises(NativeCanonicalExcelExportError, match="deterministic replay"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "workbook",
            workbook_path=workbook["workbook_relative"],
            workbook_expected_sha256=hashlib.sha256(changed_workbook).hexdigest(),
            provenance_path=workbook["provenance_relative"],
            provenance_expected_sha256=hashlib.sha256(workbook_receipt_bytes).hexdigest(),
        )

    role = fixture("role")
    development = tmp_path / "role/output/development/export"
    development.mkdir(parents=True)
    moved_workbook = development / role["workbook_path"].name
    moved_provenance = development / role["provenance_path"].name
    shutil.copyfile(role["workbook_path"], moved_workbook)
    shutil.copyfile(role["provenance_path"], moved_provenance)
    moved_receipt = json.loads(moved_provenance.read_bytes())
    moved_receipt["workbook"]["path"] = moved_workbook.relative_to(tmp_path / "role").as_posix()
    moved_receipt["provenance"]["path"] = moved_provenance.relative_to(tmp_path / "role").as_posix()
    moved_provenance_bytes = _canonical(moved_receipt)
    moved_provenance.write_bytes(moved_provenance_bytes)
    with pytest.raises(NativeCanonicalExcelExportError, match="must stay under"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "role",
            workbook_path=moved_workbook.relative_to(tmp_path / "role"),
            workbook_expected_sha256=role["workbook_sha256"],
            provenance_path=moved_provenance.relative_to(tmp_path / "role"),
            provenance_expected_sha256=hashlib.sha256(moved_provenance_bytes).hexdigest(),
        )

    symlink = fixture("symlink")
    target = symlink["workbook_path"].with_name("target.xlsx")
    symlink["workbook_path"].rename(target)
    symlink["workbook_path"].symlink_to(target.name)
    with pytest.raises(NativeCanonicalExcelExportError, match="regular file"):
        load_registered_native_canonical_excel(
            project_root=tmp_path / "symlink",
            workbook_path=symlink["workbook_relative"],
            workbook_expected_sha256=symlink["workbook_sha256"],
            provenance_path=symlink["provenance_relative"],
            provenance_expected_sha256=symlink["provenance_sha256"],
        )


def test_completed_pair_detects_during_replay_mutation(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    mapping_payload, rows_payload = _paired_payloads()
    completed = _write_completed_fixture(tmp_path, project_root, mapping_payload, rows_payload)
    _mock_strict_pair_dependencies(monkeypatch, mapping_payload, rows_payload)
    original_build = export_module._build_prevalidated_native_canonical_excel_artifacts

    def mutate_during_rebuild(*args, **kwargs):
        rebuilt = original_build(*args, **kwargs)
        completed["workbook_path"].write_bytes(rebuilt.workbook_bytes + b"transient")
        completed["workbook_path"].write_bytes(rebuilt.workbook_bytes)
        return rebuilt

    monkeypatch.setattr(
        export_module,
        "_build_prevalidated_native_canonical_excel_artifacts",
        mutate_during_rebuild,
    )
    with pytest.raises(NativeCanonicalExcelExportError, match="changed during strict replay"):
        load_registered_native_canonical_excel(
            project_root=tmp_path,
            workbook_path=completed["workbook_relative"],
            workbook_expected_sha256=completed["workbook_sha256"],
            provenance_path=completed["provenance_relative"],
            provenance_expected_sha256=completed["provenance_sha256"],
        )


def test_publisher_rolls_back_both_owned_files_when_postpublication_replay_fails(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    mapping_payload, rows_payload = _paired_payloads()
    completed = _write_completed_fixture(tmp_path, project_root, mapping_payload, rows_payload)
    completed["workbook_path"].unlink()
    completed["provenance_path"].unlink()
    _mock_strict_pair_dependencies(monkeypatch, mapping_payload, rows_payload)

    def fail_replay(**_kwargs):
        raise NativeCanonicalExcelExportError("synthetic strict replay failure")

    monkeypatch.setattr(
        export_module,
        "load_registered_native_canonical_excel",
        fail_replay,
    )
    with pytest.raises(NativeCanonicalExcelExportError, match="rolled back"):
        export_registered_native_canonical_excel(
            project_root=tmp_path,
            mapping_path=tmp_path / completed["mapping_identity"].path,
            mapping_expected_sha256=completed["mapping_identity"].sha256,
            rows_path=tmp_path / completed["rows_identity"].path,
            rows_expected_sha256=completed["rows_identity"].sha256,
            workbook_path=completed["workbook_path"],
            provenance_path=completed["provenance_path"],
            export_policy_path=tmp_path / EXPORT_POLICY_RELATIVE_PATH,
        )
    assert not completed["workbook_path"].exists()
    assert not completed["provenance_path"].exists()
