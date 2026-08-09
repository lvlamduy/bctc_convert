from __future__ import annotations

import datetime as datetime_module
import json
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from bctc_ai.export.cdkt_development import (
    CDKTDevelopmentExportError,
    build_cdkt_development_artifacts,
    export_cdkt_development,
)

_TEMPLATE = Path("template/Bank_CDKT_ReportNormId.v2.xlsx")
_SEALED = Path(
    "output/calibration/e0041-mbb-cdkt-post-mapping-development-excel/mbb-cdkt-development.xlsx"
)


def _build(project_root: Path):
    return build_cdkt_development_artifacts(
        template_path=project_root / _TEMPLATE,
        sealed_workbook_path=project_root / _SEALED,
        workbook_name="mbb-cdkt-development.xlsx",
    )


def _headers(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def test_cdkt_business_update_reconciles_97_items_and_preserves_source_statuses(
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from openpyxl.writer import excel as excel_writer

    save_clock = iter(
        (
            datetime_module.datetime(2030, 1, 2, 3, 4, 5),
            datetime_module.datetime(2040, 6, 7, 8, 9, 10),
        )
    )

    class FakeDateTime:
        @classmethod
        def now(cls, tz=None):
            value = next(save_clock)
            return value.replace(tzinfo=tz) if tz is not None else value

    monkeypatch.setattr(
        excel_writer,
        "datetime",
        SimpleNamespace(datetime=FakeDateTime, timezone=datetime_module.timezone),
    )
    first = _build(project_root)
    second = _build(project_root)
    assert first.workbook_bytes == second.workbook_bytes
    assert first.coverage_bytes == second.coverage_bytes
    assert (first.observed_value_count, first.derived_value_count, first.exported_value_count) == (
        132,
        2,
        134,
    )
    assert first.fully_verified is False

    coverage = json.loads(first.coverage_bytes)
    assert coverage["authority"] == {
        "components": {
            "off_balance_page_5": ("HASH_BOUND_SOURCE_PDF_RENDER_OCR_PLUS_AUDITED_SCHEMA_ALIASES"),
            "statement_pages_3_4": (
                "USER_CONFIRMED_MAPPING_AND_VISIBLE_VALUE_PLUS_ACCOUNTING_DERIVATION"
            ),
        },
        "fully_verified": False,
        "production_approved": False,
        "scope": "MIXED_USER_CONFIRMED_AND_AUDITED_SOURCE_EVIDENCE",
    }
    assert coverage["coverage"] == {
        "derived_value_count": 2,
        "exported_value_count": 134,
        "mapped_schema_count": 73,
        "not_observed_schema_count": 24,
        "observed_value_count": 132,
        "physical_cell_count": 150,
        "physical_status_counts": {"BLANK": 13, "DASH": 5, "VALUE": 132},
        "reconciled_schema_count": 97,
        "source_row_count": 75,
        "unresolved_schema_count": 0,
    }
    assert coverage["decisions"]["Q010"]["canonical_value_vnd"] == 2_320_000_000
    assert coverage["decisions"]["Q005"] == {
        "status": "TARGET_STATEMENT_BOUNDARY_CLARIFIED_NO_VALUE_BACKFILL",
        "target_status": "NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES",
    }
    assert coverage["not_observed_on_target_statement_pages_report_norm_ids"] == [
        4303,
        4309,
        4326,
        4329,
        4333,
        4340,
        4341,
        4344,
        4345,
        4347,
        4352,
        4359,
        4360,
        4369,
        4370,
        4373,
        4374,
    ]
    assert coverage["decisions"]["MBB_PROVISION_SCHEMA_EVOLUTION"] == {
        "status": "SOURCE_VISIBLE_BROADER_ITEMS_REASSIGNED",
        "reassignments": [
            {
                "from_report_norm_id": 4347,
                "source_row_id": "page-0003-row-006-label",
                "to_report_norm_id": 6035,
            },
            {
                "from_report_norm_id": 4352,
                "source_row_id": "page-0003-row-017-label",
                "to_report_norm_id": 6036,
            },
        ],
    }
    assert coverage["off_balance_page5"] == {
        "accounting_checks_passed": [
            "6041=SUM(6042,6043,6044,6045)/CURRENT",
            "6041=SUM(6042,6043,6044,6045)/COMPARATIVE",
        ],
        "mapped_report_norm_ids": list(range(6038, 6049)),
        "physical_cell_count": 22,
        "policy_path": "config/tables/cdkt-off-balance-page5-v1.yaml",
        "policy_sha256": "3c0075df66647a3854b68c16dea703397ac37b3f5c5ce2d0d73cb2c736c671af",
        "source_ocr_path": (
            "output/calibration/recovery-e0027-mbb-q1-2026-role-c-20260807/"
            "ppocrv6-page-0005/ocr_result.json"
        ),
        "source_ocr_sha256": ("27e5cc72f71a4b759bd0a72e28a9178aa55faaf04aa2cd67812322d83b591d68"),
        "source_pdf_path": "vietstock_bctc/MBB/2026/BCTC Hợp nhất quý 1 năm 2026.pdf",
        "source_pdf_sha256": ("eebeda2ebc09b0d4203259e92cda0169b46fde555557f150a314c72517fc1c83"),
        "source_render_path": (
            "output/calibration/recovery-e0027-mbb-q1-2026-20260807/"
            "eebeda2ebc09b0d42032/renders/page-0005.png"
        ),
        "source_render_sha256": (
            "7f2574bf11ad7df3d93dc6256c8aa631f6851f8e0056e7bed3c0195d8eeccc6a"
        ),
        "source_row_count": 11,
    }
    assert coverage["cross_statement_note_links"] == [
        {
            "cdkt_report_norm_id": 4344,
            "note_pdf_page": 30,
            "note_report_norm_id": 576,
            "note_statement": "TM",
            "relation": "RELATED_NOTE_DETAIL_NOT_PROVEN_IDENTICAL",
            "value_backfill_allowed": False,
        },
        {
            "cdkt_report_norm_id": 4326,
            "note_pdf_page": 30,
            "note_report_norm_id": 585,
            "note_statement": "TM",
            "relation": "RELATED_NOTE_DISCLOSURE_NO_BACKFILL",
            "value_backfill_allowed": False,
        },
        {
            "cdkt_report_norm_id": 4345,
            "note_pdf_page": 30,
            "note_report_norm_id": 5718,
            "note_statement": "TM",
            "relation": "BROADER_COMBINED_PROVISION_NOT_EQUIVALENT",
            "value_backfill_allowed": False,
        },
    ]
    assert coverage["formula_checks"] == {
        "4305_equals_4304_plus_5712": {"passed": True},
        "4325_components": {"passed": True, "status": "DERIVED_FROM_COMPONENTS"},
        "5712_equals_4325_plus_4306": {
            "passed": True,
            "status": "VISIBLE_TOTAL_WITH_4306_NOT_OBSERVED",
        },
    }

    template = load_workbook(project_root / _TEMPLATE, data_only=False)
    workbook = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    try:
        assert workbook.properties.created == datetime_module.datetime(2000, 1, 1)
        assert workbook.properties.modified == datetime_module.datetime(2000, 1, 1)
        assert workbook.sheetnames == [
            "CDKT",
            "PROVENANCE",
            "VALIDATION_DIAGNOSTICS",
            "RUN_METADATA",
        ]
        main = workbook["CDKT"]
        for row in range(1, 99):
            for column in range(1, 4):
                assert main.cell(row, column).value == template.active.cell(row, column).value
                assert main.cell(row, column).style_id == template.active.cell(row, column).style_id
        by_id = {main.cell(row, 2).value: row for row in range(2, 99)}
        assert len(by_id) == 97
        assert main.cell(by_id[4350], 3).value == "Chứng khoán đầu tư sẵn sàng để bán"
        assert main.cell(by_id[4363], 4).value == 2_320_000_000
        assert main.cell(by_id[4363], 6).value == "VALUE"
        assert main.cell(by_id[4306], 6).value == "NOT_OBSERVED_IN_THIS_PDF"
        assert main.cell(by_id[4344], 6).value == ("NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES")
        assert main.cell(by_id[4326], 6).value == ("NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES")
        assert main.cell(by_id[4345], 6).value == ("NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES")
        for report_norm_id in (4347, 4352):
            assert main.cell(by_id[report_norm_id], 4).value is None
            assert main.cell(by_id[report_norm_id], 5).value is None
            assert main.cell(by_id[report_norm_id], 6).value == (
                "NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES"
            )
            assert main.cell(by_id[report_norm_id], 7).value == (
                "NOT_OBSERVED_ON_TARGET_STATEMENT_PAGES"
            )
        assert tuple(main.cell(by_id[6035], column).value for column in (4, 5)) == (
            -34_663_000_000,
            -39_393_000_000,
        )
        assert tuple(main.cell(by_id[6036], column).value for column in (4, 5)) == (
            -561_362_000_000,
            -232_739_000_000,
        )
        assert tuple(main.cell(by_id[6035], column).value for column in (6, 7)) == (
            "VALUE",
            "VALUE",
        )
        assert tuple(main.cell(by_id[6036], column).value for column in (6, 7)) == (
            "VALUE",
            "VALUE",
        )
        assert main.cell(by_id[6037], 6).value == "NOT_OBSERVED_IN_THIS_PDF"
        assert main.cell(by_id[6037], 7).value == "NOT_OBSERVED_IN_THIS_PDF"
        for report_norm_id in (6038, 6039):
            assert main.cell(by_id[report_norm_id], 4).value is None
            assert main.cell(by_id[report_norm_id], 5).value is None
            assert main.cell(by_id[report_norm_id], 6).value == "BLANK"
            assert main.cell(by_id[report_norm_id], 7).value == "BLANK"
        expected_off_balance_values = {
            6040: (1_681_823_000_000, 1_684_717_000_000),
            6041: (723_980_330_000_000, 618_888_427_000_000),
            6042: (1_302_737_000_000, 9_738_358_000_000),
            6043: (2_160_046_000_000, 8_752_345_000_000),
            6044: (359_933_489_000_000, 299_830_234_000_000),
            6045: (360_584_058_000_000, 300_567_490_000_000),
            6046: (71_763_365_000_000, 59_728_018_000_000),
            6047: (186_067_393_000_000, 190_317_517_000_000),
            6048: (117_681_586_000_000, 127_878_633_000_000),
        }
        for report_norm_id, expected_values in expected_off_balance_values.items():
            assert tuple(main.cell(by_id[report_norm_id], column).value for column in (4, 5)) == (
                expected_values
            )
            assert tuple(main.cell(by_id[report_norm_id], column).value for column in (6, 7)) == (
                "VALUE",
                "VALUE",
            )
        for report_norm_id in range(6049, 6054):
            assert main.cell(by_id[report_norm_id], 6).value == "NOT_OBSERVED_IN_THIS_PDF"
            assert main.cell(by_id[report_norm_id], 7).value == "NOT_OBSERVED_IN_THIS_PDF"
        assert main.cell(by_id[4325], 4).value == 149_745_325_000_000
        assert main.cell(by_id[4325], 6).value == "DERIVED_FROM_COMPONENTS"
        assert main.cell(by_id[5712], 4).value == 149_745_325_000_000
        assert main.cell(by_id[5712], 5).value == 142_022_525_000_000
        assert (
            main.cell(by_id[4304], 4).value + main.cell(by_id[5712], 4).value
            == main.cell(by_id[4305], 4).value
        )
        assert (
            sum(
                main.cell(row, column).value is not None
                for row in range(2, 99)
                for column in (4, 5)
            )
            == 134
        )

        provenance = workbook["PROVENANCE"]
        headers = _headers(provenance)
        records = [
            {name: provenance.cell(row, column).value for name, column in headers.items()}
            for row in range(2, provenance.max_row + 1)
        ]
        assert provenance.max_row == 151
        p4r0 = [record for record in records if record["RowId"] == "page-0004-row-000-label"]
        p4r13 = [record for record in records if record["RowId"] == "page-0004-row-013-label"]
        p4r23 = [record for record in records if record["RowId"] == "page-0004-row-023-label"]
        assert {record["ReportNormId"] for record in p4r0} == {4304}
        assert {record["Status"] for record in p4r0} == {"BLANK"}
        assert {record["ReportNormId"] for record in p4r13} == {5712}
        assert {record["Status"] for record in p4r13} == {"BLANK"}
        assert {record["ReportNormId"] for record in p4r23} == {5712}
        assert {record["Status"] for record in p4r23} == {"VALUE"}
        provision_reassignments = {
            record["RowId"]: record
            for record in records
            if record["RowId"] in {"page-0003-row-006-label", "page-0003-row-017-label"}
        }
        assert {
            row_id: {record["ReportNormId"] for record in records if record["RowId"] == row_id}
            for row_id in provision_reassignments
        } == {
            "page-0003-row-006-label": {6035},
            "page-0003-row-017-label": {6036},
        }
        assert all(
            record["MappingStatus"] == "SCHEMA_EVOLUTION_REASSIGNED_FROM_NARROWER_ITEM"
            and record["CandidateReportNormIds"] == f"[{record['ReportNormId']}]"
            and record["Status"] == "VALUE"
            and record["NumericVerificationStatus"] == "VERIFIED_OBSERVED_VALUE"
            and record["CropSha256"]
            and record["SourceRenderSha256"]
            and record["SourceOcrSha256"]
            for record in records
            if record["RowId"] in {"page-0003-row-006-label", "page-0003-row-017-label"}
        )
        page5 = [record for record in records if record["Page"] == 5]
        assert len(page5) == 22
        assert {record["ReportNormId"] for record in page5} == set(range(6038, 6049))
        assert {record["Status"] for record in page5} == {"BLANK", "VALUE"}
        assert sum(record["Status"] == "BLANK" for record in page5) == 4
        assert sum(record["Status"] == "VALUE" for record in page5) == 18
        assert all(
            record["SourceRenderSha256"]
            == "7f2574bf11ad7df3d93dc6256c8aa631f6851f8e0056e7bed3c0195d8eeccc6a"
            and record["SourceOcrSha256"]
            == "27e5cc72f71a4b759bd0a72e28a9178aa55faaf04aa2cd67812322d83b591d68"
            and record["CropBbox"]
            and record["SourceRowIds"]
            for record in page5
        )
        confirmed = next(
            record
            for record in records
            if record["RowId"] == "page-0004-row-011-label" and record["AxisOrdinal"] == 0
        )
        assert confirmed["ExportedCanonicalValue"] == 2_320_000_000
        assert confirmed["NumericVerificationStatus"] == "USER_CONFIRMED_VISIBLE_VALUE"
        assert all(record["Scope"] == "CONSOLIDATED" for record in records)

        assert not any(
            cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()
        template.close()


def test_cdkt_export_refuses_overwrite_and_schema_identity_drift(
    tmp_path: Path,
    project_root: Path,
):
    workbook_path = tmp_path / "mbb-cdkt-development.xlsx"
    coverage_path = tmp_path / "coverage.json"
    result = export_cdkt_development(
        template_path=project_root / _TEMPLATE,
        sealed_workbook_path=project_root / _SEALED,
        workbook_path=workbook_path,
        coverage_path=coverage_path,
    )
    assert workbook_path.stat().st_size == result.workbook_size_bytes
    assert coverage_path.stat().st_size == result.coverage_size_bytes
    with pytest.raises(CDKTDevelopmentExportError, match="already exists"):
        export_cdkt_development(
            template_path=project_root / _TEMPLATE,
            sealed_workbook_path=project_root / _SEALED,
            workbook_path=workbook_path,
            coverage_path=coverage_path,
        )

    template = load_workbook(project_root / _TEMPLATE)
    vpb_row = next(
        row
        for row in range(2, template.active.max_row + 1)
        if template.active.cell(row, 2).value == 6035
    )
    template.active.cell(vpb_row, 2, 999_999)
    drift_path = tmp_path / "drift.xlsx"
    template.save(drift_path)
    template.close()
    with pytest.raises(CDKTDevelopmentExportError, match="business-schema identities"):
        build_cdkt_development_artifacts(
            template_path=drift_path,
            sealed_workbook_path=project_root / _SEALED,
            workbook_name="mbb-cdkt-development.xlsx",
        )


def test_cdkt_export_durably_fsyncs_directory_after_pair_rollback(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bctc_ai.export import cdkt_development as module

    workbook_path = tmp_path / "mbb-cdkt-development.xlsx"
    coverage_path = tmp_path / "coverage.json"
    real_write_exclusive = module._write_exclusive
    write_count = 0
    directory_events: list[tuple[Path, bool]] = []

    def record_directory_fsync(directory: Path) -> None:
        directory_events.append((directory, workbook_path.exists()))

    def fail_second_write(path: Path, data: bytes) -> None:
        nonlocal write_count
        write_count += 1
        if write_count == 2:
            raise OSError("injected second-member failure")
        real_write_exclusive(path, data)

    monkeypatch.setattr(module, "_fsync_directory", record_directory_fsync)
    monkeypatch.setattr(module, "_write_exclusive", fail_second_write)
    with pytest.raises(OSError, match="injected second-member failure"):
        export_cdkt_development(
            template_path=project_root / _TEMPLATE,
            sealed_workbook_path=project_root / _SEALED,
            workbook_path=workbook_path,
            coverage_path=coverage_path,
        )

    assert not workbook_path.exists()
    assert not coverage_path.exists()
    assert directory_events == [(tmp_path, True), (tmp_path, False)]
