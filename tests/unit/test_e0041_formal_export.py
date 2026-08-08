from __future__ import annotations

import copy
import hashlib
import inspect
import json
import os
import shutil
import subprocess
import sys
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

import pytest
import yaml
from openpyxl import load_workbook

from bctc_ai.evaluation import e0041_formal_export as formal

_COMMIT = "a199bd08891d8d5c79c2927410e89c66efcf6036"


@pytest.fixture(scope="module")
def materials(project_root: Path):
    return formal._load_materials(
        project_root,
        formal.CONTROL_RELATIVE_PATH,
        reader=formal._default_reader,
        git_snapshot=None,
    )


@pytest.fixture(scope="module")
def candidate(project_root: Path, materials):
    built = formal._build_once(project_root, materials, capture_git_commit=_COMMIT)
    return formal.FormalExportBuild(
        projection=built[0],
        projection_bytes=built[1],
        workbook_bytes=built[2],
        provenance=built[4],
        provenance_bytes=built[5],
        materials=materials,
    )


def test_control_has_no_superseded_path_and_pins_runtime_and_core(project_root: Path):
    control = yaml.safe_load((project_root / formal.CONTROL_RELATIVE_PATH).read_text())
    ledger_text = json.dumps(
        {
            "inputs": control["input_authority"],
            "implementation": control["implementation"],
            "runtime": control["runtime_authority"],
        },
        sort_keys=True,
    ).casefold()

    assert "e0038" not in ledger_text
    assert control["implementation"]["post_mapping_export_core"] == {
        "path": formal.CORE_RELATIVE_PATH.as_posix(),
        "sha256": formal.CORE_SHA256,
        "size_bytes": formal.CORE_SIZE_BYTES,
    }
    assert control["runtime_authority"]["versions"] == formal._RUNTIME_VERSIONS
    assert control["publication"]["atomic_pair_publication"] is False
    assert control["phase_outputs"]["export_pair"]["pair_hash_sealed"] is False


def test_control_rejects_unknown_nested_keys_and_execution_path_substitution(project_root: Path):
    control = yaml.safe_load((project_root / formal.CONTROL_RELATIVE_PATH).read_text())
    unknown = copy.deepcopy(control)
    unknown["input_authority"]["e0040_mapping_seal"]["review_path"] = "/tmp/answers.json"
    with pytest.raises(formal.E0041FormalExportError, match="nested keyset"):
        formal._validate_control(unknown)

    substituted = copy.deepcopy(control)
    substituted["implementation"]["formal_export_integration"]["path"] = "README.md"
    with pytest.raises(formal.E0041FormalExportError, match="implementation path"):
        formal._validate_control(substituted)

    duplicate = b"version: 1\nversion: 2\n"
    with pytest.raises(formal.E0041FormalExportError, match="decode"):
        formal._decode_control(duplicate)
    alias = b"version: &v 1\nexperiment_id: *v\n"
    with pytest.raises(formal.E0041FormalExportError, match="decode"):
        formal._decode_control(alias)


def test_material_open_order_is_exact_and_diagnostic_paths_are_never_opened(project_root: Path):
    opened: list[str] = []

    def spy(root: Path, path: Path, maximum_size: int, name: str) -> bytes:
        opened.append(name)
        return formal._default_reader(root, path, maximum_size, name)

    loaded = formal._load_materials(
        project_root,
        formal.CONTROL_RELATIVE_PATH,
        reader=spy,
        git_snapshot=None,
    )
    positions = {
        name: opened.index(name)
        for name in (
            "E-0041 formal control",
            "E-0040 mapping seal",
            "E-0040 S3 registration",
            "shared S3 registry frozen baseline",
            "E-0040 formal mapping",
            "E-0037 postjoin",
            "E-0041 reconstructed geometry registry",
            "CDKT template",
        )
    }
    assert list(positions.values()) == sorted(positions.values())
    implementation_positions = [
        index for index, name in enumerate(opened) if name.startswith("implementation ")
    ]
    runtime_positions = [index for index, name in enumerate(opened) if name.startswith("runtime ")]
    asset_positions = [
        index for index, name in enumerate(opened) if name.startswith("geometry asset ")
    ]
    assert max((*implementation_positions, *runtime_positions)) < positions["E-0040 mapping seal"]
    assert positions["E-0041 reconstructed geometry registry"] < min(asset_positions)
    assert max(asset_positions) < positions["CDKT template"]
    assert len(asset_positions) == 134
    assert not any("source_image_path" in name for name in opened)
    assert len(loaded.asset_inventory) == 134


@pytest.mark.parametrize(
    ("corrupt_name", "must_not_open"),
    [
        ("E-0040 mapping seal", "E-0040 S3 registration"),
        ("E-0040 S3 registration", "E-0040 formal mapping"),
    ],
)
def test_corrupt_authority_stops_before_later_opens(
    project_root: Path,
    corrupt_name: str,
    must_not_open: str,
):
    opened: list[str] = []

    def corrupting_reader(root: Path, path: Path, maximum_size: int, name: str) -> bytes:
        opened.append(name)
        payload = formal._default_reader(root, path, maximum_size, name)
        return payload[:-1] + bytes([payload[-1] ^ 1]) if name == corrupt_name else payload

    with pytest.raises(formal.E0041FormalExportError, match="byte identity drifted"):
        formal._load_materials(
            project_root,
            formal.CONTROL_RELATIVE_PATH,
            reader=corrupting_reader,
            git_snapshot=None,
        )
    assert must_not_open not in opened


def test_exact_geometry_inventory_and_root_normalization(materials):
    inventory_bytes = formal._compact_bytes(materials.asset_inventory)

    assert len(materials.asset_inventory) == 134
    assert sum(item["size_bytes"] for item in materials.asset_inventory) == 4_489_853
    assert (len(inventory_bytes), hashlib.sha256(inventory_bytes).hexdigest()) == (
        31_130,
        "7a50f4737c8381bc73a04ad4fce9abd200b64b917e9282a0197bf56d9a044cc0",
    )
    assert all(not Path(item["path"]).is_absolute() for item in materials.asset_inventory)
    role_counts = {
        role: sum(role in item["roles"] for item in materials.asset_inventory)
        for role in {role for item in materials.asset_inventory for role in item["roles"]}
    }
    assert role_counts == {
        "cell_crop": 128,
        "cell_source_ocr": 2,
        "cell_source_render": 2,
        "crop_policy": 1,
        "page_3_ocr": 1,
        "page_3_render": 1,
        "page_4_ocr": 1,
        "page_4_render": 1,
        "row_contract": 1,
    }


def test_projection_workbook_and_provenance_are_exact_and_deterministic(
    project_root: Path,
    materials,
    candidate: formal.FormalExportBuild,
):
    replay = formal._build_once(project_root, materials, capture_git_commit=_COMMIT)

    assert (
        len(candidate.projection_bytes),
        hashlib.sha256(candidate.projection_bytes).hexdigest(),
    ) == (
        442_178,
        "4ba86ea84b102932b5ae2952f2998c205907d58095793535eacfd100c501db8d",
    )
    assert (
        len(candidate.workbook_bytes),
        hashlib.sha256(candidate.workbook_bytes).hexdigest(),
    ) == (
        43_746,
        "60c5f6d01ba0a11276da6b6fc7f85d4d4f4893b6ee70d0ca54bd14764a922caf",
    )
    assert candidate.projection_bytes == replay[1]
    assert candidate.workbook_bytes == replay[2]
    assert candidate.provenance_bytes == replay[5]
    assert candidate.provenance["pair_hash_sealed"] is False
    assert b"provenance_sha256" not in candidate.provenance_bytes
    assert b"/workspace/" not in candidate.provenance_bytes


def test_workbook_has_no_formulas_and_preserves_template_identity(candidate):
    workbook = load_workbook(formal.BytesIO(candidate.workbook_bytes), data_only=False)
    try:
        assert (
            sum(
                cell.data_type == "f"
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
            )
            == 0
        )
        snapshot = formal._core._template_snapshot(workbook, "Sheet1", 78)
        assert hashlib.sha256(formal._compact_bytes(snapshot)).hexdigest() == (
            "38c74d1af3be5a70e4dd27d18030e4950f96469a194c6d06f9843a1467831408"
        )
    finally:
        workbook.close()


def test_coordinated_provenance_mutations_are_rejected(candidate):
    extra_geometry_claim = copy.deepcopy(candidate.provenance)
    extra_geometry_claim["geometry_authority"]["extra_claim"] = True
    with pytest.raises(formal.E0041FormalExportError, match="full contract"):
        formal._validate_new_provenance(
            extra_geometry_claim,
            formal._encoded_json(extra_geometry_claim),
            materials=candidate.materials,
            commit=_COMMIT,
            projection_bytes=candidate.projection_bytes,
            workbook_bytes=candidate.workbook_bytes,
            workbook_receipt=candidate.provenance["workbook_receipt"],
        )

    extra_workbook_claim = copy.deepcopy(candidate.provenance)
    extra_workbook_claim["outputs"]["workbook"]["verified"] = True
    with pytest.raises(formal.E0041FormalExportError, match="workbook identity"):
        formal._validate_existing_provenance(
            extra_workbook_claim,
            formal._encoded_json(extra_workbook_claim),
            expected_commit=_COMMIT,
        )


def test_relocation_and_two_filesystem_clocks_do_not_change_bytes(
    tmp_path: Path,
    project_root: Path,
    materials,
):
    relocated = tmp_path / "relocated-checkout"
    for stable in materials.ordered_files:
        relative = stable.path.relative_to(project_root)
        destination = relocated / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(stable.path, destination)
    first_materials = formal._load_materials(
        relocated,
        formal.CONTROL_RELATIVE_PATH,
        reader=formal._default_reader,
        git_snapshot=None,
    )
    first = formal._build_once(relocated, first_materials, capture_git_commit=_COMMIT)
    for stable in first_materials.ordered_files:
        os.utime(stable.path, ns=(1_600_000_000_000_000_000, 1_600_000_000_000_000_000))
    second_materials = formal._load_materials(
        relocated,
        formal.CONTROL_RELATIVE_PATH,
        reader=formal._default_reader,
        git_snapshot=None,
    )
    second = formal._build_once(relocated, second_materials, capture_git_commit=_COMMIT)

    assert first[1] == second[1]
    assert first[2] == second[2]
    assert first[5] == second[5]


def _fake_pair_build() -> SimpleNamespace:
    return SimpleNamespace(
        workbook_bytes=b"formal-workbook-candidate",
        provenance_bytes=b'{"formal":"provenance"}\n',
        provenance={"state": formal.PAIR_STATE},
    )


def test_formal_pair_publisher_requires_exact_empty_and_never_overwrites(tmp_path: Path):
    (tmp_path / formal.OUTPUT_DIRECTORY_RELATIVE_PATH.parent).mkdir(parents=True)
    build = _fake_pair_build()
    result = formal._publish_formal_pair(tmp_path, build)

    assert result == build.provenance
    assert sorted((tmp_path / formal.OUTPUT_DIRECTORY_RELATIVE_PATH).iterdir()) == sorted(
        [tmp_path / formal.WORKBOOK_RELATIVE_PATH, tmp_path / formal.PROVENANCE_RELATIVE_PATH]
    )
    with pytest.raises(Exception, match="start exactly empty"):
        formal._publish_formal_pair(tmp_path, build)

    hostile_root = tmp_path / "hostile"
    output = hostile_root / formal.OUTPUT_DIRECTORY_RELATIVE_PATH
    output.mkdir(parents=True)
    (output / "ignored-extra.bin").write_bytes(b"hostile")
    with pytest.raises(formal.E0041FormalExportError, match="start exactly empty"):
        formal._publish_formal_pair(hostile_root, build)
    assert sorted(path.name for path in output.iterdir()) == ["ignored-extra.bin"]


def test_formal_pair_publisher_detects_inventory_race_and_rolls_back(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    (tmp_path / formal.OUTPUT_DIRECTORY_RELATIVE_PATH.parent).mkdir(parents=True)
    original = formal._core._write_exclusive_at
    injected = False

    def racing_write(parent_descriptor: int, filename: str, payload: bytes):
        nonlocal injected
        identity = original(parent_descriptor, filename, payload)
        if not injected:
            injected = True
            descriptor = os.open(
                "racing-extra.bin",
                os.O_WRONLY | os.O_CREAT | os.O_EXCL,
                0o644,
                dir_fd=parent_descriptor,
            )
            os.write(descriptor, b"race")
            os.close(descriptor)
        return identity

    monkeypatch.setattr(formal._core, "_write_exclusive_at", racing_write)
    with pytest.raises(formal.E0041FormalExportError, match="inventory"):
        formal._publish_formal_pair(tmp_path, _fake_pair_build())
    output = tmp_path / formal.OUTPUT_DIRECTORY_RELATIVE_PATH
    assert sorted(path.name for path in output.iterdir()) == ["racing-extra.bin"]


def test_formal_pair_writes_workbook_then_completion_marker_and_rolls_back_second_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    (tmp_path / formal.OUTPUT_DIRECTORY_RELATIVE_PATH.parent).mkdir(parents=True)
    original = formal._core._write_exclusive_at
    writes: list[str] = []

    def failing_second(parent_descriptor: int, filename: str, payload: bytes):
        writes.append(filename)
        if len(writes) == 2:
            raise formal.E0041FormalExportError("synthetic second-write failure")
        return original(parent_descriptor, filename, payload)

    monkeypatch.setattr(formal._core, "_write_exclusive_at", failing_second)
    with pytest.raises(formal.E0041FormalExportError, match="second-write"):
        formal._publish_formal_pair(tmp_path, _fake_pair_build())
    assert writes == [formal.WORKBOOK_RELATIVE_PATH.name, formal.PROVENANCE_RELATIVE_PATH.name]
    assert list((tmp_path / formal.OUTPUT_DIRECTORY_RELATIVE_PATH).iterdir()) == []


def test_seal_validator_rejects_coordinated_mutation(candidate):
    workbook = formal._StableFile(
        "workbook",
        Path("/tmp/workbook"),
        candidate.workbook_bytes,
        (0, 0, stat_mode := 0, len(candidate.workbook_bytes), 0, 0),
        formal._record_for_payload(formal.WORKBOOK_RELATIVE_PATH, candidate.workbook_bytes),
        False,
    )
    provenance = formal._StableFile(
        "provenance",
        Path("/tmp/provenance"),
        candidate.provenance_bytes,
        (0, 0, stat_mode, len(candidate.provenance_bytes), 0, 0),
        formal._record_for_payload(formal.PROVENANCE_RELATIVE_PATH, candidate.provenance_bytes),
        False,
    )
    seal = formal._seal_payload(
        candidate,
        commit=_COMMIT,
        workbook_file=workbook,
        provenance_file=provenance,
    )
    seal["replay"]["extra_claim"] = True
    with pytest.raises(formal.E0041FormalExportError, match="full contract"):
        formal._validate_seal_payload(
            seal,
            formal._encoded_json(seal),
            build=candidate,
            commit=_COMMIT,
            workbook_file=workbook,
            provenance_file=provenance,
        )


def _temporary_pair(root: Path) -> tuple[formal._StableFile, formal._StableFile]:
    output = root / formal.OUTPUT_DIRECTORY_RELATIVE_PATH
    output.mkdir(parents=True)
    (root / formal.SEAL_RELATIVE_PATH.parent).mkdir(parents=True)
    (root / formal.WORKBOOK_RELATIVE_PATH).write_bytes(b"workbook-for-seal-publisher")
    (root / formal.PROVENANCE_RELATIVE_PATH).write_bytes(b"provenance-for-seal-publisher")
    workbook = formal._read_unpinned_file(
        root,
        formal.WORKBOOK_RELATIVE_PATH,
        "workbook",
        reader=formal._default_reader,
        maximum_size=1024,
    )
    provenance = formal._read_unpinned_file(
        root,
        formal.PROVENANCE_RELATIVE_PATH,
        "provenance",
        reader=formal._default_reader,
        maximum_size=1024,
    )
    return workbook, provenance


def test_formal_seal_publisher_is_exclusive_and_requires_exact_pair_inventory(tmp_path: Path):
    workbook, provenance = _temporary_pair(tmp_path)
    payload = {"state": formal.SEAL_STATE, "bounded": True}
    formal._publish_formal_seal(
        tmp_path,
        payload,
        workbook_file=workbook,
        provenance_file=provenance,
    )
    sealed = (tmp_path / formal.SEAL_RELATIVE_PATH).read_bytes()
    with pytest.raises(Exception, match="overwrite"):
        formal._publish_formal_seal(
            tmp_path,
            payload,
            workbook_file=workbook,
            provenance_file=provenance,
        )
    assert (tmp_path / formal.SEAL_RELATIVE_PATH).read_bytes() == sealed

    hostile = tmp_path / "hostile-seal"
    hostile_workbook, hostile_provenance = _temporary_pair(hostile)
    (hostile / formal.OUTPUT_DIRECTORY_RELATIVE_PATH / "extra.bin").write_bytes(b"extra")
    with pytest.raises(formal.E0041FormalExportError, match="inventory"):
        formal._publish_formal_seal(
            hostile,
            payload,
            workbook_file=hostile_workbook,
            provenance_file=hostile_provenance,
        )
    assert not (hostile / formal.SEAL_RELATIVE_PATH).exists()


def test_formal_seal_shared_root_swap_rolls_back_created_seal(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    real_root = tmp_path / "real"
    alternate_root = tmp_path / "alternate"
    workbook, provenance = _temporary_pair(real_root)
    _temporary_pair(alternate_root)
    original = formal._core._open_trusted_root
    calls = 0

    def swapping_root(project_root: Path, name: str):
        nonlocal calls
        calls += 1
        return original(alternate_root if calls == 2 else project_root, name)

    monkeypatch.setattr(formal._core, "_open_trusted_root", swapping_root)
    with pytest.raises(formal.E0041FormalExportError, match="trusted root changed"):
        formal._publish_formal_seal(
            real_root,
            {"state": formal.SEAL_STATE},
            workbook_file=workbook,
            provenance_file=provenance,
        )
    assert not (real_root / formal.SEAL_RELATIVE_PATH).exists()


def test_terminal_ignored_input_mutation_is_rejected(tmp_path: Path, materials):
    path = tmp_path / "ignored.bin"
    path.write_bytes(b"first")
    observed = path.stat()
    stable = formal._StableFile(
        name="ignored mutation",
        path=path,
        payload=b"first",
        identity=formal._stat_identity(observed),
        artifact=formal._record_for_payload(Path("ignored.bin"), b"first"),
        require_head_blob=False,
    )
    narrowed = replace(materials, ordered_files=(stable,))
    path.write_bytes(b"other")
    with pytest.raises(formal.E0041FormalExportError, match="byte identity drifted"):
        formal._recheck_ignored_materials(tmp_path, narrowed, reader=formal._default_reader)


def test_public_build_has_no_git_snapshot_or_reader_bypass():
    parameters = inspect.signature(formal.build_e0041_formal_export).parameters
    assert "_git_snapshot" not in parameters
    assert "_reader" not in parameters


def test_sanitized_git_gate_ignores_hostile_environment_and_rejects_index_flags(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    remote = tmp_path / "remote.git"
    work = tmp_path / "work"

    def git(cwd: Path, *arguments: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            ["git", *arguments],
            cwd=cwd,
            env={key: value for key, value in os.environ.items() if not key.startswith("GIT_")},
            capture_output=True,
            text=True,
            check=True,
        )

    git(tmp_path, "init", "--bare", str(remote))
    git(tmp_path, "init", str(work))
    git(work, "config", "user.name", "Formal Test")
    git(work, "config", "user.email", "formal@example.invalid")
    tracked = work / "tracked.txt"
    tracked.write_text("tracked bytes\n", encoding="utf-8")
    git(work, "add", "tracked.txt")
    git(work, "commit", "-m", "initial")
    git(work, "branch", "-M", "main")
    git(work, "remote", "add", "origin", str(remote))
    git(work, "push", "-u", "origin", "main")

    hostile_global = tmp_path / "hostile.gitconfig"
    hostile_global.write_text("[status]\n\tshowUntrackedFiles = no\n", encoding="utf-8")
    monkeypatch.setenv("GIT_DIR", "/definitely/not/the/test/repository")
    monkeypatch.setenv("GIT_WORK_TREE", "/definitely/not/the/test/worktree")
    monkeypatch.setenv("GIT_INDEX_FILE", "/definitely/not/the/test/index")
    monkeypatch.setenv("GIT_CONFIG_GLOBAL", str(hostile_global))

    snapshot = formal._clean_git_snapshot(work.resolve())
    assert snapshot.commit == snapshot.upstream_commit
    payload = tracked.read_bytes()
    observed = tracked.stat()
    stable = formal._StableFile(
        name="tracked test ledger",
        path=tracked,
        payload=payload,
        identity=formal._stat_identity(observed),
        artifact=formal._record_for_payload(Path("tracked.txt"), payload),
        require_head_blob=True,
    )
    formal._assert_head_blob_binding(work.resolve(), stable)

    git(work, "update-index", "--assume-unchanged", "tracked.txt")
    with pytest.raises(formal.E0041FormalExportError, match="index flags"):
        formal._assert_head_blob_binding(work.resolve(), stable)
    git(work, "update-index", "--no-assume-unchanged", "tracked.txt")
    git(work, "update-index", "--skip-worktree", "tracked.txt")
    with pytest.raises(formal.E0041FormalExportError, match="index"):
        formal._assert_head_blob_binding(work.resolve(), stable)


def test_fresh_subprocess_process_isolation(project_root: Path):
    clean_code = (
        "from bctc_ai.evaluation.e0041_formal_export import _assert_process_isolation;"
        "_assert_process_isolation();print('PASS')"
    )
    clean = subprocess.run(
        [sys.executable, "-c", clean_code],
        cwd=project_root,
        env={**os.environ, "PYTHONPATH": str(project_root / "src")},
        capture_output=True,
        text=True,
        check=False,
    )
    assert clean.returncode == 0 and clean.stdout.strip() == "PASS"

    contaminated_code = (
        "import sys;from types import ModuleType;"
        "from bctc_ai.evaluation.e0041_formal_export import _assert_process_isolation;"
        "sys.modules['bctc_ai.evaluation.review_answers']=ModuleType('review_answers');"
        "\ntry:_assert_process_isolation()\nexcept RuntimeError:print('BLOCKED')\n"
    )
    contaminated = subprocess.run(
        [sys.executable, "-c", contaminated_code],
        cwd=project_root,
        env={**os.environ, "PYTHONPATH": str(project_root / "src")},
        capture_output=True,
        text=True,
        check=False,
    )
    assert contaminated.returncode == 0 and contaminated.stdout.strip() == "BLOCKED"
