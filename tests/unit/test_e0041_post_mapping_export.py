from __future__ import annotations

import copy
import datetime as datetime_module
import hashlib
import json
import os
import zipfile
from collections import Counter
from dataclasses import FrozenInstanceError, replace
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
import yaml
from openpyxl import load_workbook

from bctc_ai.evaluation import e0041_post_mapping_export as export_module
from bctc_ai.evaluation.e0041_post_mapping_export import (
    AuthenticatedE0040ResultCarrier,
    E0041PostMappingExportError,
    _load_template,
    _normalize_authenticated_geometry_registry,
    _publish_pair,
    _stable_read,
    _validate_mapping_challenger,
    assemble_post_mapping_projection,
    authenticate_e0040_result_carrier,
    build_development_workbook,
)
from bctc_ai.mapping.e0040_calibration_challenger import (
    align_e0040_calibration_challenger,
    load_e0040_policy,
    projection_from_sealed_mapping_payload,
    source_rows_from_sealed_mapping_payload,
)
from bctc_ai.mapping.ordered_subgraph_v2 import load_ordered_subgraph_v2_policy

_CONTROL = Path("config/experiments/e0041-mbb-cdkt-post-mapping-development-excel.yaml")
_POSTJOIN = Path("docs/experiments/E-0037-mbb-cdkt-sealed-evidence-mapping.json")
_E0037_MAPPING = Path("output/calibration/e0037-mbb-cdkt-sealed-evidence-mapping/mapping_only.json")
_E0038_MAPPING = Path("output/calibration/e0038-mbb-cdkt-exact-mapping/mapping_only.json")
_E0040_POLICY = Path("config/mapping/e0040-cdkt-semantic-normalization.yaml")
_MAPPER_POLICY = Path("config/mapping/ordered-subgraph-v2-exact-e0038.yaml")
_TEMPLATE = Path("template/Bank_CDKT_ReportNormId.xlsx")
_E0040_MAPPING = Path("output/calibration/e0040-mbb-cdkt-formal-mapping/mapping_only.json")
_E0040_SEAL = Path("docs/experiments/E-0040-mbb-cdkt-formal-mapping-seal.json")
_E0040_REGISTRATION = Path("docs/experiments/E-0040-mbb-cdkt-formal-mapping-s3-registration.json")
_GEOMETRY = Path(
    "output/calibration/e0041-mbb-cdkt-reconstructed-geometry/65fa9b7c0de1/crop_registry.json"
)


@pytest.fixture(scope="module")
def e0040_result(project_root: Path):
    sealed = json.loads((project_root / _E0037_MAPPING).read_text(encoding="utf-8"))
    return align_e0040_calibration_challenger(
        source_rows_from_sealed_mapping_payload(sealed),
        projection_from_sealed_mapping_payload(sealed),
        policy=load_e0040_policy(project_root / _E0040_POLICY),
        mapper_policy=load_ordered_subgraph_v2_policy(project_root / _MAPPER_POLICY),
    )


@pytest.fixture(scope="module")
def authenticated_e0040_carrier(project_root: Path):
    # Preserve the formal trust order: authenticate seal and registration bytes
    # before opening the mapping payload that contains the result.
    seal_bytes = (project_root / _E0040_SEAL).read_bytes()
    registration_bytes = (project_root / _E0040_REGISTRATION).read_bytes()
    mapping_bytes = (project_root / _E0040_MAPPING).read_bytes()
    return authenticate_e0040_result_carrier(
        mapping_bytes=mapping_bytes,
        seal_bytes=seal_bytes,
        registration_bytes=registration_bytes,
    )


@pytest.fixture(scope="module")
def template_material(project_root: Path):
    payload = (project_root / _TEMPLATE).read_bytes()
    workbook, rows, snapshot = _load_template(
        payload,
        sheet_name="Sheet1",
        schema_row_count=77,
    )
    workbook.close()
    return payload, rows, snapshot


@pytest.fixture
def assembly_factory(
    tmp_path: Path,
    project_root: Path,
    template_material,
):
    counter = 0

    def assemble(mapping_authority, *, hostile_source_label: bool = False):
        nonlocal counter
        counter += 1
        run_root = tmp_path / f"run-{counter}"
        registry_directory = run_root / "registry"
        registry_directory.mkdir(parents=True)
        crop_payload = b"synthetic-crop-for-e0041-mechanism-test"
        source_payload = b"synthetic-registered-render-and-ocr-source"
        crop_path = registry_directory / "crop.bin"
        source_path = run_root / "registered-source.bin"
        crop_path.write_bytes(crop_payload)
        source_path.write_bytes(source_payload)
        crop_sha256 = hashlib.sha256(crop_payload).hexdigest()
        source_sha256 = hashlib.sha256(source_payload).hexdigest()

        postjoin = json.loads((project_root / _POSTJOIN).read_text(encoding="utf-8"))
        # These superseded fields are hostile on purpose. E0041 must never read
        # them to make or populate a mapping.
        for index, row in enumerate(postjoin["rows"]):
            row["mapping"] = {"selected_report_norm_id": 900000 + index}
            if hostile_source_label and index == 0:
                row["semantic_proposals"]["ppocrv6_source"] = "=1+1"
        for index, cell in enumerate(postjoin["cells"]):
            cell["selected_report_norm_id"] = 800000 + index
            cell["candidate_report_norm_ids"] = [800000 + index]
            cell["output_status"] = "VALUE"
            cell["selected_raw_value"] = "999999"
            cell["selected_normalized_value"] = "999999"
            cell["displayed_unit_value"] = "999999"
            cell["canonical_unit_value"] = 999999
            cell["numeric_evidence"]["crop_sha256"] = crop_sha256

        relative_source = source_path.relative_to(run_root).as_posix()
        geometry_cells = [
            {
                "cell_id": cell["cell_id"],
                "page": cell["page"],
                "row_ordinal": cell["row_ordinal"],
                "axis_ordinal": cell["axis_ordinal"],
                "crop_bbox": [1, 2, 101, 42],
                "crop_path": "crop.bin",
                "crop_sha256": crop_sha256,
                "crop_size_bytes": len(crop_payload),
                "source_row_ids": [f"registered-source:{cell['row_id']}"],
                "source_render_path": relative_source,
                "source_render_sha256": source_sha256,
                "source_ocr_path": relative_source,
                "source_ocr_sha256": source_sha256,
                "value_line_indices": [],
            }
            for cell in postjoin["cells"]
        ]
        geometry = {
            "format_version": 2,
            "policy": "FIXED_GRID_NUMERIC_CELL_CROPS_V2",
            "geometry_authority": "E0033_PP_OCRV6_FIXED_GRID",
            "reference_isolation": {
                "accounting_validation_invoked": False,
                "historical_or_mongodb_values_loaded": False,
                "human_review_loaded": False,
                "schema_mapping_invoked": False,
                "template_labels_or_report_norm_ids_loaded": False,
            },
            "cells": geometry_cells,
        }
        geometry_before = copy.deepcopy(geometry)
        control = yaml.safe_load((project_root / _CONTROL).read_text(encoding="utf-8"))
        synthetic_record = {"path": "synthetic", "sha256": "0" * 64, "size_bytes": 1}
        projection = assemble_post_mapping_projection(
            postjoin_payload=postjoin,
            mapping_payload=mapping_authority,
            geometry_registry=geometry,
            geometry_registry_path=registry_directory / "crop_registry.json",
            template_rows=template_material[1],
            project_root=run_root,
            input_records={
                "e0037_postjoin": synthetic_record,
                "mapping_challenger": synthetic_record,
                "geometry_registry": synthetic_record,
            },
            validation_config=control["accounting_validation"],
        )
        return projection, geometry == geometry_before

    return assemble


def _expected_sets(e0040_result) -> tuple[set[str], set[int]]:
    rows = {item.row_id for item in e0040_result.final_result.row_mappings}
    schema = {item.report_norm_id for item in e0040_result.final_result.schema_dispositions}
    return rows, schema


def test_control_freezes_physical_equations_and_does_not_overclaim_atomic_publication(
    project_root: Path,
):
    control = yaml.safe_load((project_root / _CONTROL).read_text(encoding="utf-8"))
    equations = control["accounting_validation"]["strict_physical_visible_row_equations"]

    assert equations["authority"] == "NEWLY_FROZEN_CALIBRATION_MECHANISM_ASSERTION"
    assert equations["mapping_independent_physical_row_ids"] is True
    assert equations["expected_family_count"] == 18
    assert equations["expected_finding_count"] == 36
    assert len(equations["equations"]) == 18
    assert all(
        not any("report_norm_id" in key.casefold() for key in equation)
        for equation in equations["equations"]
    )
    assert control["outputs"]["publication"] == (
        "FD_RELATIVE_EXCLUSIVE_NO_OVERWRITE_PAIR_WITH_HELD_DIRFD_ROLLBACK_"
        "AND_FRESH_CANONICAL_CHAIN_REVALIDATION"
    )
    assert control["outputs"]["publication_path_safety"] == {
        "trusted_project_root_dirfd": "O_DIRECTORY_NOFOLLOW",
        "output_chain_traversal": "COMPONENT_BY_COMPONENT_DIRFD_RELATIVE_NOFOLLOW",
        "writes_and_rollback": "HELD_OUTPUT_DIRFD_RELATIVE",
        "success_gate": "FRESH_CANONICAL_CHAIN_PARENT_AND_FILE_INODE_REVALIDATION",
    }
    assert control["state"] == ("MECHANISM_READY_FORMAL_CAPTURE_BLOCKED_PENDING_E0040_SEAL")
    assert control["formal_capture_gate"] == {
        "status": "BLOCKED_PENDING_AUTHENTICATED_E0040_ARTIFACT_AND_SEAL",
        "required_mapping_authority": "E0040_GENERIC_CALIBRATION_CHALLENGER",
        "authenticated_e0040_artifact_bound": False,
        "authenticated_e0040_seal_bound": False,
        "e0038_may_be_used_as_formal_mapping_authority": False,
    }
    assert control["outputs"]["formal_publication_allowed_before_e0040_seal"] is False
    assert "atomic_pair_publication" not in control["outputs"]
    assert control["workbook"]["deterministic_core_properties"] == {
        "creator": "bctc-ai/E-0041",
        "last_modified_by": "bctc-ai/E-0041",
        "created_utc": "2000-01-01T00:00:00Z",
        "modified_utc": "2000-01-01T00:00:00Z",
        "version": "1",
        "revision": "1",
        "optional_free_text_fields_cleared": True,
        "canonical_core_xml_sha256": (
            "a025959e8b178cfc6c6aae8f2d49d86fa305d3e36e165c8cbbc16923068668e4"
        ),
    }
    for name in ("e0037_postjoin", "cdkt_template"):
        record = control["inputs"][name]
        payload = (project_root / record["path"]).read_bytes()
        assert (record["sha256"], record["size_bytes"]) == (
            hashlib.sha256(payload).hexdigest(),
            len(payload),
        )


def test_e0040_direct_result_adapter_rejects_detached_and_forged_envelopes(e0040_result):
    expected_rows, expected_schema = _expected_sets(e0040_result)
    rows, _dispositions, aliases, authority = _validate_mapping_challenger(
        e0040_result,
        expected_row_ids=expected_rows,
        expected_schema_ids=expected_schema,
    )

    assert Counter(item["status"] for item in rows.values()) == {
        "RESOLVED_ANCHOR": 43,
        "RESOLVED_PATH": 18,
        "SOURCE_ONLY_STRUCTURAL_ROW": 3,
    }
    assert aliases == set()
    assert authority["final_selected_pair_count"] == 61
    assert authority["source_only_row_count"] == 3

    # A fully coordinated JSON envelope is still detached raw data, not direct
    # call provenance from the approved challenger implementation.
    coordinated_forgery = e0040_result.to_dict()
    coordinated_forgery["normalization"]["id_scoped_alias_invocation_count"] = 0
    coordinated_forgery["collision_audit"]["new_collision_pairs"] = ()
    with pytest.raises(E0041PostMappingExportError, match="direct-call authority"):
        _validate_mapping_challenger(
            coordinated_forgery,
            expected_row_ids=expected_rows,
            expected_schema_ids=expected_schema,
        )

    alias_forgery = replace(
        e0040_result,
        normalization=replace(
            e0040_result.normalization,
            id_scoped_alias_invocation_count=1,
        ),
    )
    with pytest.raises(E0041PostMappingExportError, match="authority is unsafe"):
        _validate_mapping_challenger(
            alias_forgery,
            expected_row_ids=expected_rows,
            expected_schema_ids=expected_schema,
        )

    forged_search = replace(e0040_result.final_result.search, pruned_states=1)
    pruning_forgery = replace(
        e0040_result,
        final_result=replace(e0040_result.final_result, search=forged_search),
    )
    with pytest.raises(E0041PostMappingExportError, match="was pruned"):
        _validate_mapping_challenger(
            pruning_forgery,
            expected_row_ids=expected_rows,
            expected_schema_ids=expected_schema,
        )


def test_authenticated_e0040_json_carrier_is_consumed_without_mapper_replay(
    authenticated_e0040_carrier: AuthenticatedE0040ResultCarrier,
    assembly_factory,
):
    carrier = authenticated_e0040_carrier
    result = json.loads(carrier.mapping_bytes)["challenger_result"]
    expected_rows = {item["row_id"] for item in result["final_result"]["row_mappings"]}
    expected_schema = {
        item["report_norm_id"] for item in result["final_result"]["schema_dispositions"]
    }
    rows, _dispositions, aliases, authority = _validate_mapping_challenger(
        carrier,
        expected_row_ids=expected_rows,
        expected_schema_ids=expected_schema,
    )

    assert all(
        type(payload) is bytes
        for payload in (carrier.mapping_bytes, carrier.seal_bytes, carrier.registration_bytes)
    )
    assert not hasattr(carrier, "__dict__")
    assert not hasattr(carrier.mapping_artifact, "__dict__")
    with pytest.raises(FrozenInstanceError):
        carrier.capture_git_commit = "0" * 40
    assert carrier.challenger_result_sha256 == (
        "2e49d8623692fde9fd4a5a87f9c2e2159941b0f3ded7b7b16dddac2ab1e85fbd"
    )
    assert Counter(item["status"] for item in rows.values()) == {
        "RESOLVED_ANCHOR": 43,
        "RESOLVED_PATH": 18,
        "SOURCE_ONLY_STRUCTURAL_ROW": 3,
    }
    assert aliases == set()
    assert authority["capture_git_commit"] == "18aca8942faf5d47e1ac5f049045d7a7a297b5fc"
    assert authority["authenticated_formal_artifacts"] == {
        "mapping": {
            "path": "output/calibration/e0040-mbb-cdkt-formal-mapping/mapping_only.json",
            "sha256": "8def983007fc3aacf59351395426d5246ad3f28d605442f590de55eaf396cb0d",
            "size_bytes": 1_157_172,
        },
        "seal": {
            "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-seal.json",
            "sha256": "68306f7f540faa77d6e2e383927eae23fc3724cfdc8c53cded978a86f3a00b29",
            "size_bytes": 7_611,
        },
        "registration": {
            "path": "docs/experiments/E-0040-mbb-cdkt-formal-mapping-s3-registration.json",
            "sha256": "f38d9a1bbed4ec48e2156d441e5c76c6e6d82b0771208de3eef92d96173dd4b5",
            "size_bytes": 13_360,
        },
    }

    projection, geometry_unchanged = assembly_factory(carrier)
    assert geometry_unchanged is True
    assert projection["access_contract"]["mapping_challenger_invocation_count"] == 0
    assert projection["mapping_authority"]["challenger_result_sha256"] == (
        carrier.challenger_result_sha256
    )

    with pytest.raises(E0041PostMappingExportError, match="direct-call authority"):
        _validate_mapping_challenger(
            result,
            expected_row_ids=expected_rows,
            expected_schema_ids=expected_schema,
        )


def test_authenticated_e0040_carrier_rejects_byte_and_receipt_mutations(
    project_root: Path,
    authenticated_e0040_carrier: AuthenticatedE0040ResultCarrier,
):
    carrier = authenticated_e0040_carrier
    result = json.loads(carrier.mapping_bytes)["challenger_result"]
    expected_rows = {item["row_id"] for item in result["final_result"]["row_mappings"]}
    expected_schema = {
        item["report_norm_id"] for item in result["final_result"]["schema_dispositions"]
    }
    mutations = (
        replace(
            carrier,
            mapping_artifact=replace(
                carrier.mapping_artifact,
                size_bytes=carrier.mapping_artifact.size_bytes - 1,
            ),
        ),
        replace(
            carrier,
            seal_artifact=replace(carrier.seal_artifact, sha256="0" * 64),
        ),
        replace(
            carrier,
            registration_artifact=replace(
                carrier.registration_artifact,
                path="docs/experiments/forged-registration.json",
            ),
        ),
        replace(carrier, challenger_result_sha256="0" * 64),
        replace(carrier, capture_git_commit="0" * 40),
    )
    for mutation in mutations:
        with pytest.raises(E0041PostMappingExportError, match="carrier|challenger"):
            _validate_mapping_challenger(
                mutation,
                expected_row_ids=expected_rows,
                expected_schema_ids=expected_schema,
            )

    unauthenticated = AuthenticatedE0040ResultCarrier(
        mapping_bytes=carrier.mapping_bytes,
        seal_bytes=b"{}",
        registration_bytes=carrier.registration_bytes,
        mapping_artifact=carrier.mapping_artifact,
        seal_artifact=carrier.seal_artifact,
        registration_artifact=carrier.registration_artifact,
        challenger_result_sha256=carrier.challenger_result_sha256,
        capture_git_commit=carrier.capture_git_commit,
    )
    with pytest.raises(E0041PostMappingExportError, match="seal byte identity drifted"):
        _validate_mapping_challenger(
            unauthenticated,
            expected_row_ids=expected_rows,
            expected_schema_ids=expected_schema,
        )

    artifact_bytes = {
        "mapping_bytes": (project_root / _E0040_MAPPING).read_bytes(),
        "seal_bytes": (project_root / _E0040_SEAL).read_bytes(),
        "registration_bytes": (project_root / _E0040_REGISTRATION).read_bytes(),
    }
    for name, payload in artifact_bytes.items():
        forged = {**artifact_bytes, name: payload[:-1] + bytes([payload[-1] ^ 1])}
        with pytest.raises(E0041PostMappingExportError, match="byte identity drifted"):
            authenticate_e0040_result_carrier(**forged)

    class BytesSubclass(bytes):
        pass

    with pytest.raises(E0041PostMappingExportError, match="must be exact bytes"):
        authenticate_e0040_result_carrier(
            mapping_bytes=BytesSubclass(artifact_bytes["mapping_bytes"]),
            seal_bytes=artifact_bytes["seal_bytes"],
            registration_bytes=artifact_bytes["registration_bytes"],
        )


def test_authenticated_geometry_registry_paths_are_relocatable_and_digest_pinned(
    project_root: Path,
    tmp_path: Path,
):
    registry = json.loads((project_root / _GEOMETRY).read_text(encoding="utf-8"))
    original = copy.deepcopy(registry)
    normalized = _normalize_authenticated_geometry_registry(registry)
    normalized_bytes = json.dumps(
        normalized,
        ensure_ascii=False,
        allow_nan=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")

    assert registry == original
    assert len(normalized_bytes) == 161_120
    assert hashlib.sha256(normalized_bytes).hexdigest() == (
        "e834efd4f6e70c03e607d834a17adc69e0fa0868658767637c5db3cf8c06be6a"
    )
    assert _normalize_authenticated_geometry_registry(normalized) == normalized

    rooted_fields = {"path", "ocr_path", "render_path", "source_ocr_path", "source_render_path"}
    rooted_values: list[str] = []

    def collect(value, field_name: str | None = None):
        if isinstance(value, dict):
            for key, item in value.items():
                collect(item, key)
        elif isinstance(value, list):
            for item in value:
                collect(item, field_name)
        elif field_name in rooted_fields:
            rooted_values.append(value)

    collect(normalized)
    assert len(rooted_values) == 262
    assert len(set(rooted_values)) == 6
    assert all(
        not Path(value).is_absolute() and ".." not in Path(value).parts for value in rooted_values
    )
    relocated_root = tmp_path / "relocated-checkout"
    assert relocated_root / normalized["cells"][0]["source_render_path"] == (
        relocated_root / "output/calibration/recovery-e0027-mbb-q1-2026-20260807/"
        "eebeda2ebc09b0d42032/renders/page-0003.png"
    )

    assets = {
        normalized["crop_policy"]["path"],
        normalized["row_contract"]["path"],
        *(page[field] for page in normalized["pages"] for field in ("ocr_path", "render_path")),
        *((_GEOMETRY.parent / cell["crop_path"]).as_posix() for cell in normalized["cells"]),
    }
    assert len(assets) == 134
    assert sum((project_root / path).stat().st_size for path in assets) == 4_489_853


def test_authenticated_geometry_registry_rejects_foreign_roots_and_traversal(
    project_root: Path,
):
    registry = json.loads((project_root / _GEOMETRY).read_text(encoding="utf-8"))
    foreign = copy.deepcopy(registry)
    foreign["cells"][0]["source_render_path"] = "/foreign/bctc-ai/source.png"
    with pytest.raises(E0041PostMappingExportError, match="foreign absolute prefix"):
        _normalize_authenticated_geometry_registry(foreign)

    traversal = copy.deepcopy(registry)
    traversal["pages"][0]["ocr_path"] = "/workspace/bctc-ai/../foreign/ocr.json"
    with pytest.raises(E0041PostMappingExportError, match="unsafe"):
        _normalize_authenticated_geometry_registry(traversal)


def test_projection_retains_128_physical_cells_and_matches_strict_oracle(
    assembly_factory,
    e0040_result,
):
    projection, geometry_unchanged = assembly_factory(e0040_result)
    metrics = projection["metrics"]

    assert geometry_unchanged is True
    assert metrics["source_row_count"] == 64
    assert metrics["physical_cell_count"] == 128
    assert metrics["selected_target_cell_count"] == 122
    assert metrics["exported_numeric_cell_count"] == 111
    assert metrics["physical_cell_status_counts"] == {
        "BLANK": 5,
        "DASH": 5,
        "UNRESOLVED": 7,
        "VALUE": 111,
    }
    cells = projection["physical_cells"]
    assert len({cell["cell_id"] for cell in cells}) == 128
    assert all(
        cell["exported_canonical_value"] is None for cell in cells if cell["status"] == "DASH"
    )
    assert all(cell["visible_raw_value"] == "-" for cell in cells if cell["status"] == "DASH")
    assert all(
        cell["exported_canonical_value"] is None
        for cell in cells
        if cell["status"] in {"BLANK", "AMBIGUOUS", "UNRESOLVED"}
    )

    source_only = [cell for cell in cells if cell["source_only"]]
    assert len(source_only) == 6
    assert all(cell["report_norm_id"] is None for cell in source_only)
    source_only_total = [
        cell for cell in source_only if cell["row_id"] == "page-0004-row-023-label"
    ]
    assert len(source_only_total) == 2
    assert all(cell["source_numeric_status"] == "VALUE" for cell in source_only_total)
    assert all(cell["evidence_canonical_value"] is not None for cell in source_only_total)
    assert all(cell["exported_canonical_value"] is None for cell in source_only_total)
    assert all(
        any(ref.startswith("P4_R023_VISIBLE_CHILDREN_ONLY") for ref in cell["validation_refs"])
        for cell in source_only_total
    )

    validation = projection["validation"]
    assert (
        validation["pre_validation_value_status_sha256"]
        == (validation["post_validation_value_status_sha256"])
    )
    assert validation["finding_counts"] == {"NOT_TESTABLE": 6, "PASS": 30}
    assert len(validation["findings"]) == 36
    assert validation["secondary_schema_hierarchy"] == {
        "enabled": False,
        "separate_denominator": True,
        "finding_count": 0,
        "findings": [],
        "finding_counts": {},
    }
    assert {
        finding["finding_id"]
        for finding in validation["findings"]
        if finding["result"] == "NOT_TESTABLE"
    } == {
        "P3_R018_CURRENT",
        "P3_R018_COMPARATIVE",
        "P3_R038_TOTAL_CURRENT",
        "P3_R038_TOTAL_COMPARATIVE",
        "P4_R007_CURRENT",
        "P4_R007_COMPARATIVE",
    }
    assert not any(finding["result"] == "FAIL" for finding in validation["findings"])


def test_physical_equation_injection_is_rejected(
    project_root: Path,
    assembly_factory,
    e0040_result,
):
    projection, _ = assembly_factory(e0040_result)
    control = yaml.safe_load((project_root / _CONTROL).read_text(encoding="utf-8"))
    forged = copy.deepcopy(control["accounting_validation"])
    forged["strict_physical_visible_row_equations"]["equations"][0]["rhs_row_ids"] = [
        "page-0003-row-007-label"
    ]

    with pytest.raises(E0041PostMappingExportError, match="does not contain 18 families"):
        export_module._diagnose_physical_visible_equations(
            projection["physical_cells"],
            forged,
        )


def test_e0038_alias_and_six_unselected_rows_fail_closed_without_changing_arithmetic(
    project_root: Path,
    assembly_factory,
    e0040_result,
):
    e0038 = json.loads((project_root / _E0038_MAPPING).read_text(encoding="utf-8"))
    e0038_projection, _ = assembly_factory(e0038)
    e0040_projection, _ = assembly_factory(e0040_result)
    exact = e0038["exact_mapping_bundle"]["exact_search"][
        "mapping_result_without_internal_alias_authority"
    ]
    alias_rows = {
        row["row_id"]
        for row in exact["row_mappings"]
        if row["selected_report_norm_id"] in {4375, 5699}
    }
    unselected_rows = {
        row["row_id"] for row in exact["row_mappings"] if row["selected_report_norm_id"] is None
    }

    assert len(alias_rows) == 2
    assert len(unselected_rows) == 6
    by_row: dict[str, list[dict[str, object]]] = {}
    for cell in e0038_projection["physical_cells"]:
        by_row.setdefault(cell["row_id"], []).append(cell)
    assert all(
        cell["status"] == "AMBIGUOUS"
        and cell["report_norm_id"] is None
        and cell["exported_canonical_value"] is None
        for row_id in alias_rows
        for cell in by_row[row_id]
    )
    assert all(
        cell["status"] in {"AMBIGUOUS", "UNRESOLVED"}
        and cell["report_norm_id"] is None
        and cell["exported_canonical_value"] is None
        for row_id in unselected_rows
        for cell in by_row[row_id]
    )
    assert e0038_projection["metrics"]["selected_target_cell_count"] == 112
    assert (
        e0038_projection["validation"]["findings"] == (e0040_projection["validation"]["findings"])
    )


def test_duplicate_selected_target_is_rejected(project_root: Path):
    payload = json.loads((project_root / _E0038_MAPPING).read_text(encoding="utf-8"))
    result = payload["exact_mapping_bundle"]["exact_search"][
        "mapping_result_without_internal_alias_authority"
    ]
    selected = [row for row in result["row_mappings"] if row["selected_report_norm_id"]]
    selected[1]["selected_report_norm_id"] = selected[0]["selected_report_norm_id"]
    selected[1]["candidate_report_norm_ids"] = [selected[0]["selected_report_norm_id"]]

    with pytest.raises(E0041PostMappingExportError, match="duplicate selected ReportNormId"):
        _validate_mapping_challenger(
            payload,
            expected_row_ids={row["row_id"] for row in result["row_mappings"]},
            expected_schema_ids={row["report_norm_id"] for row in result["schema_dispositions"]},
        )


def test_workbook_clones_template_and_writes_hostile_ocr_as_literal_text(
    assembly_factory,
    e0040_result,
    template_material,
    monkeypatch: pytest.MonkeyPatch,
):
    projection, _ = assembly_factory(e0040_result, hostile_source_label=True)
    from openpyxl.writer import excel as excel_writer

    clock = [datetime_module.datetime(2030, 1, 2, 3, 4, 5)]

    class FakeDateTime:
        @classmethod
        def now(cls, tz=None):
            value = clock[0]
            return value.replace(tzinfo=tz) if tz is not None else value

    monkeypatch.setattr(
        excel_writer,
        "datetime",
        SimpleNamespace(datetime=FakeDateTime, timezone=datetime_module.timezone),
    )
    workbook_bytes, receipt = build_development_workbook(
        template_bytes=template_material[0],
        projection=projection,
    )
    clock[0] = datetime_module.datetime(2040, 6, 7, 8, 9, 10)
    repeated_bytes, repeated_receipt = build_development_workbook(
        template_bytes=template_material[0],
        projection=projection,
    )

    assert workbook_bytes == repeated_bytes
    assert receipt == repeated_receipt
    assert receipt["exact_template_identity_value_style_fidelity"] is True
    assert receipt["deterministic_core_properties_sha256"] == (
        "a025959e8b178cfc6c6aae8f2d49d86fa305d3e36e165c8cbbc16923068668e4"
    )
    with zipfile.ZipFile(BytesIO(workbook_bytes)) as archive:
        core = archive.read("docProps/core.xml")
    assert hashlib.sha256(core).hexdigest() == (
        "a025959e8b178cfc6c6aae8f2d49d86fa305d3e36e165c8cbbc16923068668e4"
    )
    assert core == export_module._DETERMINISTIC_CORE_PROPERTIES_XML
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    try:
        assert workbook.sheetnames == [
            "Sheet1",
            "PROVENANCE",
            "VALIDATION_DIAGNOSTICS",
            "RUN_METADATA",
        ]
        assert workbook.properties.created == datetime_module.datetime(2000, 1, 1)
        assert workbook.properties.modified == datetime_module.datetime(2000, 1, 1)
        assert workbook.properties.creator == "bctc-ai/E-0041"
        assert workbook.properties.lastModifiedBy == "bctc-ai/E-0041"
        assert workbook.properties.version == "1"
        assert workbook.properties.revision == "1"
        assert not any(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        provenance = workbook["PROVENANCE"]
        headers = {cell.value: cell.column for cell in provenance[1]}
        hostile = provenance.cell(2, headers["SourceLabel"])
        assert hostile.value == "=1+1"
        assert hostile.data_type == "s"
        dash_cells = [
            provenance.cell(row, headers["VisibleRawValue"])
            for row in range(2, provenance.max_row + 1)
            if provenance.cell(row, headers["Status"]).value == "DASH"
        ]
        assert len(dash_cells) == 5
        assert all(cell.value == "-" and cell.data_type == "s" for cell in dash_cells)

        source = workbook["Sheet1"]
        dash_schema_rows = [
            row
            for row in range(2, source.max_row + 1)
            if source.cell(row, 6).value == "DASH" or source.cell(row, 7).value == "DASH"
        ]
        assert dash_schema_rows
        for row in dash_schema_rows:
            if source.cell(row, 6).value == "DASH":
                assert source.cell(row, 4).value is None
            if source.cell(row, 7).value == "DASH":
                assert source.cell(row, 5).value is None
    finally:
        workbook.close()


def test_stable_read_rejects_parent_swap_to_symlink(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    parent = tmp_path / "input"
    detached = tmp_path / "detached-input"
    replacement = tmp_path / "replacement"
    parent.mkdir()
    replacement.mkdir()
    target = parent / "payload.bin"
    target.write_bytes(b"same-bytes")
    (replacement / "payload.bin").write_bytes(b"same-bytes")
    original_read = export_module.os.read
    raced = False

    def racing_read(descriptor: int, size: int):
        nonlocal raced
        payload = original_read(descriptor, size)
        if payload and not raced:
            raced = True
            parent.rename(detached)
            parent.symlink_to(replacement.name, target_is_directory=True)
        return payload

    monkeypatch.setattr(export_module.os, "read", racing_read)
    with pytest.raises(E0041PostMappingExportError, match="trusted directory chain"):
        _stable_read(tmp_path, target, maximum_size=1024, name="race fixture")
    assert (detached / "payload.bin").read_bytes() == b"same-bytes"


def test_pair_publication_is_exclusive_and_never_overwrites(tmp_path: Path):
    output = tmp_path / "output"
    workbook, provenance = _publish_pair(
        tmp_path,
        output,
        "development.xlsx",
        "provenance.json",
        b"workbook-v1",
        b"provenance-v1",
    )
    with pytest.raises(E0041PostMappingExportError, match="refusing to overwrite"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook-v2",
            b"provenance-v2",
        )

    assert workbook.read_bytes() == b"workbook-v1"
    assert provenance.read_bytes() == b"provenance-v1"
    assert (
        export_module._sha256_bytes(provenance.read_bytes())
        == hashlib.sha256(b"provenance-v1").hexdigest()
    )


def test_pair_publication_rolls_back_both_files_after_output_directory_swap(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    detached = tmp_path / "detached-output"
    original_write = export_module._write_exclusive_at

    def racing_write(parent_descriptor: int, filename: str, payload: bytes):
        identity = original_write(parent_descriptor, filename, payload)
        if filename == "provenance.json":
            output.rename(detached)
            output.mkdir()
        return identity

    monkeypatch.setattr(export_module, "_write_exclusive_at", racing_write)
    with pytest.raises(E0041PostMappingExportError, match="detached from canonical path"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            b"provenance",
        )

    assert list(output.iterdir()) == []
    assert list(detached.iterdir()) == []
    assert not list(tmp_path.rglob("*.xlsx"))
    assert not list(tmp_path.rglob("*.json"))


def test_pair_publication_rolls_back_same_inode_size_mutation_after_parent_swap(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    detached = tmp_path / "detached-output"
    original_write = export_module._write_exclusive_at

    def racing_write(parent_descriptor: int, filename: str, payload: bytes):
        identity = original_write(parent_descriptor, filename, payload)
        if filename == "development.xlsx":
            descriptor = os.open("provenance.json", os.O_WRONLY, dir_fd=parent_descriptor)
            try:
                os.ftruncate(descriptor, 3)
                os.fsync(descriptor)
            finally:
                os.close(descriptor)
            output.rename(detached)
            output.mkdir()
        return identity

    monkeypatch.setattr(export_module, "_write_exclusive_at", racing_write)
    with pytest.raises(E0041PostMappingExportError, match="detached from canonical path"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            b"provenance-original",
        )

    assert list(output.iterdir()) == []
    assert list(detached.iterdir()) == []


def test_pair_publication_rejects_same_inode_forgery_after_final_stat(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    original_stat = export_module.os.stat
    stat_counts: Counter[str] = Counter()
    original_provenance = b"provenance"
    forged_provenance = b"FORGED!!!?"
    assert len(forged_provenance) == len(original_provenance)

    def racing_stat(path, *args, **kwargs):
        result = original_stat(path, *args, **kwargs)
        name = os.fsdecode(path) if isinstance(path, (str, bytes)) else None
        if name in {"development.xlsx", "provenance.json"}:
            stat_counts[name] += 1
            if name == "provenance.json" and stat_counts[name] == 4:
                (output / name).write_bytes(forged_provenance)
        return result

    monkeypatch.setattr(export_module.os, "stat", racing_stat)
    with pytest.raises(
        E0041PostMappingExportError,
        match="output identity drifted|canonical byte revalidation",
    ):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            original_provenance,
        )

    assert stat_counts["provenance.json"] >= 5
    assert list(output.iterdir()) == []


def test_exclusive_write_self_rolls_back_post_write_linked_stat_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    original_stat = export_module.os.stat
    raced = False

    def racing_stat(path, *args, **kwargs):
        nonlocal raced
        name = os.fsdecode(path) if isinstance(path, (str, bytes)) else None
        if name == "development.xlsx" and kwargs.get("dir_fd") is not None and not raced:
            raced = True
            descriptor = os.open(name, os.O_WRONLY, dir_fd=kwargs["dir_fd"])
            try:
                os.ftruncate(descriptor, 1)
                os.fsync(descriptor)
            finally:
                os.close(descriptor)
        return original_stat(path, *args, **kwargs)

    monkeypatch.setattr(export_module.os, "stat", racing_stat)
    with pytest.raises(E0041PostMappingExportError, match="identity changed after write"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            b"provenance",
        )

    assert raced is True
    assert list(output.iterdir()) == []


def test_pair_rollback_continues_after_foreign_replacement(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    original_write = export_module._write_exclusive_at
    replacement = b"foreign-workbook"

    def replacing_write(parent_descriptor: int, filename: str, payload: bytes):
        identity = original_write(parent_descriptor, filename, payload)
        if filename == "development.xlsx":
            os.unlink(filename, dir_fd=parent_descriptor)
            descriptor = os.open(
                filename,
                os.O_WRONLY | os.O_CREAT | os.O_EXCL,
                0o644,
                dir_fd=parent_descriptor,
            )
            try:
                os.write(descriptor, replacement)
                os.fsync(descriptor)
            finally:
                os.close(descriptor)
        return identity

    monkeypatch.setattr(export_module, "_write_exclusive_at", replacing_write)
    with pytest.raises(E0041PostMappingExportError, match="pair rollback was incomplete"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            b"provenance",
        )

    assert (output / "development.xlsx").read_bytes() == replacement
    assert not (output / "provenance.json").exists()


def test_final_pair_batch_detects_first_member_mutation_during_second_read(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    original_read = export_module.os.read
    seen_workbook_descriptors: set[int] = set()
    workbook_read_count = 0
    original_provenance = b"provenance"
    forged_provenance = b"FORGED!!!?"
    assert len(forged_provenance) == len(original_provenance)

    def racing_read(descriptor: int, size: int):
        nonlocal workbook_read_count
        target = Path(os.readlink(f"/proc/self/fd/{descriptor}")).name
        if target == "development.xlsx" and descriptor not in seen_workbook_descriptors:
            seen_workbook_descriptors.add(descriptor)
            workbook_read_count += 1
            if workbook_read_count == 2:
                (output / "provenance.json").write_bytes(forged_provenance)
        return original_read(descriptor, size)

    monkeypatch.setattr(export_module.os, "read", racing_read)
    with pytest.raises(E0041PostMappingExportError, match="final output batch"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            original_provenance,
        )

    assert workbook_read_count == 2
    assert list(output.iterdir()) == []


def test_second_pair_write_failure_rolls_back_detached_provenance(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    output = tmp_path / "output"
    detached = tmp_path / "detached-output"
    original_write = export_module._write_exclusive_at

    def failing_second_write(parent_descriptor: int, filename: str, payload: bytes):
        if filename == "development.xlsx":
            raise OSError("injected second-write failure")
        identity = original_write(parent_descriptor, filename, payload)
        output.rename(detached)
        output.mkdir()
        return identity

    monkeypatch.setattr(export_module, "_write_exclusive_at", failing_second_write)
    with pytest.raises(OSError, match="injected second-write failure"):
        _publish_pair(
            tmp_path,
            output,
            "development.xlsx",
            "provenance.json",
            b"workbook",
            b"provenance",
        )

    assert list(output.iterdir()) == []
    assert list(detached.iterdir()) == []
    assert not list(tmp_path.rglob("*.xlsx"))
    assert not list(tmp_path.rglob("*.json"))


def test_formal_capture_blocks_before_any_read_or_legacy_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    def forbidden(*_args, **_kwargs):
        raise AssertionError("formal capture touched input or publication code")

    monkeypatch.setattr(export_module, "_load_control", forbidden)
    monkeypatch.setattr(export_module, "_clean_git_commit", forbidden)
    monkeypatch.setattr(export_module, "_read_verified_artifact", forbidden)
    monkeypatch.setattr(export_module, "_publish_pair", forbidden)

    with pytest.raises(
        E0041PostMappingExportError,
        match="blocked pending an authenticated E-0040 artifact and seal",
    ):
        export_module.capture_e0041_post_mapping_development_excel(tmp_path)
    assert list(tmp_path.iterdir()) == []
