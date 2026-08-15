from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, replace
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from bctc_ai.export.tm_consolidated_development import (
    TM_CONSOLIDATED_POLICY_RELATIVE_PATH,
    TM_CONSOLIDATED_SCHEMA_COUNT,
    TM_CONSOLIDATED_SCHEMA_PROJECTION_SHA256,
    TM_CONSOLIDATED_SHEETS,
    TM_CONSOLIDATED_TEMPLATE_SHA256,
    TMConsolidatedDevelopmentExportError,
    TMConsolidatedExportPolicy,
    TMConsolidatedOwnerInput,
    TMConsolidatedOwnerPolicy,
    audit_tm_consolidated_owner_result_contracts,
    bind_tm_consolidated_owner_results,
    build_tm_consolidated_development_artifacts,
    export_tm_consolidated_development,
    load_tm_consolidated_export_policy,
)
from bctc_ai.schema.registry import SchemaItem


@dataclass(frozen=True)
class _Disposition:
    report_norm_id: int
    display_order: int
    canonical_name: str
    status: str
    source_row_ids: tuple[str, ...]
    reason: str


@dataclass(frozen=True)
class _Assignment:
    report_norm_id: int
    source_row_id: str
    cell_index: int
    axis_key: str
    observation: str
    value: Decimal | None
    period_start: str
    period_end: str
    period_type: str
    period_role: str
    unit: str
    unit_multiplier: int
    mapping_basis: str
    source_bbox: tuple[float, float, float, float]
    is_derived: bool = False
    is_imputed: bool = False
    source_image_path: str = "/tmp/first-run/page-0045.png"


@dataclass(frozen=True)
class _ValidationCheck:
    check_id: str
    status: str
    expected_value: Decimal | None
    observed_value: Decimal | None
    residual: Decimal | None


@dataclass(frozen=True)
class _MappingResult:
    statement_type: str
    report_scope: str
    mapping_authority_scope: str
    mapping_authority_granted: bool
    schema_item_count: int
    status_reconciled_schema_count: int
    mapped_schema_count: int
    ambiguous_schema_count: int
    unresolved_schema_count: int
    not_observed_schema_count: int
    not_applicable_schema_count: int
    unassessed_schema_count: int
    schema_dispositions: tuple[_Disposition, ...]
    source_dispositions: tuple[object, ...]
    mapped_assignments: tuple[_Assignment, ...]
    mapped_assignment_count: int
    accounting_checks: tuple[_ValidationCheck, ...] = ()


@dataclass(frozen=True)
class _NestedSource:
    row_id: str
    mapped_assignments: tuple[_Assignment, ...]


@dataclass(frozen=True)
class _NarrativeAssignment:
    source_row_ids: tuple[str, ...]
    value_index: int
    report_norm_id: int
    observation: str
    value: Decimal
    period_start: str
    period_end: str
    period_type: str
    period_role: str
    unit: str
    unit_multiplier: int
    mapping_basis: str
    raw_text: str


@dataclass(frozen=True)
class _NestedMappingResult:
    statement_type: str
    report_scope: str
    mapping_authority_scope: str
    mapping_authority_granted: bool
    schema_item_count: int
    status_reconciled_schema_count: int
    mapped_schema_count: int
    ambiguous_schema_count: int
    unresolved_schema_count: int
    not_observed_schema_count: int
    not_applicable_schema_count: int
    unassessed_schema_count: int
    schema_dispositions: tuple[_Disposition, ...]
    source_dispositions: tuple[_NestedSource, ...]
    narrative_assignments: tuple[_NarrativeAssignment, ...]
    mapped_value_count: int


@dataclass(frozen=True)
class _Fixture:
    template_path: Path
    schema: tuple[SchemaItem, ...]
    policy: TMConsolidatedExportPolicy
    owner_a: _MappingResult
    owner_b: _MappingResult

    @property
    def inputs(self) -> tuple[TMConsolidatedOwnerInput, ...]:
        return (
            TMConsolidatedOwnerInput("page-0045", self.owner_a),
            TMConsolidatedOwnerInput("residual", self.owner_b),
        )


def _projection_sha256(schema: tuple[SchemaItem, ...]) -> str:
    payload = [
        (
            item.schema_id,
            item.display_order,
            item.canonical_name,
            item.parent_id,
            tuple(item.children),
        )
        for item in schema
    ]
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _result_type(value: object) -> str:
    return f"{type(value).__module__}.{type(value).__qualname__}"


@pytest.fixture
def development_fixture(tmp_path: Path) -> _Fixture:
    schema = tuple(
        SchemaItem(
            schema_id=100 + order,
            canonical_name=f"TM item {order}",
            normalized_name=f"tm item {order}",
            statement_type="TM",
            display_order=order,
        )
        for order in range(8)
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append((None, "ReportNormId", "ReportNormName"))
    for item in schema:
        sheet.append((item.display_order, item.schema_id, item.canonical_name))
    template_path = tmp_path / "tm-template.xlsx"
    workbook.save(template_path)
    workbook.close()

    statuses_a = (
        "MAPPED_AUTOMATIC_SCOPED",
        "MAPPED_AUTOMATIC_SCOPED",
        "MAPPED_AUTOMATIC_SCOPED",
        "MAPPED_AUTOMATIC_SCOPED",
        "UNASSESSED",
        "UNASSESSED",
        "UNASSESSED",
        "UNASSESSED",
    )
    statuses_b = (
        "UNASSESSED",
        "UNASSESSED",
        "UNASSESSED",
        "UNASSESSED",
        "AMBIGUOUS_MAPPING",
        "UNRESOLVED",
        "NOT_OBSERVED_IN_THIS_PDF",
        "SCHEMA_ITEM_NOT_APPLICABLE",
    )

    def dispositions(statuses: tuple[str, ...], owner: str) -> tuple[_Disposition, ...]:
        return tuple(
            _Disposition(
                report_norm_id=item.schema_id,
                display_order=item.display_order,
                canonical_name=item.canonical_name,
                status=status,
                source_row_ids=(f"{owner}:row-{item.display_order}",)
                if status == "MAPPED_AUTOMATIC_SCOPED"
                else (),
                reason=f"{owner} owns {status}",
            )
            for item, status in zip(schema, statuses, strict=True)
        )

    assignments = tuple(
        _Assignment(
            report_norm_id=100 + index,
            source_row_id=f"page-0045:row-{index}",
            cell_index=index,
            axis_key=f"axis-{index}",
            observation=observation,
            value=value,
            period_start="2026-03-31",
            period_end="2026-03-31",
            period_type="SNAPSHOT",
            period_role="CURRENT",
            unit="VND",
            unit_multiplier=1_000_000,
            mapping_basis="VISIBLE_SOURCE_CELL",
            source_bbox=(10.0, 20.0 + index, 30.0, 40.0 + index),
        )
        for index, (observation, value) in enumerate(
            (
                ("VALUE", Decimal("12.5")),
                ("ZERO", Decimal(0)),
                ("DASH", None),
                ("BLANK", None),
            )
        )
    )
    owner_a = _MappingResult(
        statement_type="TM",
        report_scope="CONSOLIDATED",
        mapping_authority_scope="TEST_PAGE45_OWNER",
        mapping_authority_granted=True,
        schema_item_count=8,
        status_reconciled_schema_count=4,
        mapped_schema_count=4,
        ambiguous_schema_count=0,
        unresolved_schema_count=0,
        not_observed_schema_count=0,
        not_applicable_schema_count=0,
        unassessed_schema_count=4,
        schema_dispositions=dispositions(statuses_a, "page-0045"),
        source_dispositions=(),
        mapped_assignments=assignments,
        mapped_assignment_count=4,
        accounting_checks=(
            _ValidationCheck(
                check_id="VISIBLE_VALUES_BALANCE",
                status="PASS",
                expected_value=Decimal("12.5"),
                observed_value=Decimal("12.5"),
                residual=Decimal(0),
            ),
            _ValidationCheck(
                check_id="DASH_CELL_EQUATION",
                status="NOT_TESTABLE_DASH_IS_NOT_ZERO",
                expected_value=None,
                observed_value=None,
                residual=None,
            ),
        ),
    )
    owner_b = _MappingResult(
        statement_type="TM",
        report_scope="CONSOLIDATED",
        mapping_authority_scope="TEST_RESIDUAL_OWNER",
        mapping_authority_granted=True,
        schema_item_count=8,
        status_reconciled_schema_count=4,
        mapped_schema_count=0,
        ambiguous_schema_count=1,
        unresolved_schema_count=1,
        not_observed_schema_count=1,
        not_applicable_schema_count=1,
        unassessed_schema_count=4,
        schema_dispositions=dispositions(statuses_b, "residual"),
        source_dispositions=(),
        mapped_assignments=(),
        mapped_assignment_count=0,
    )
    owner_type = _result_type(owner_a)
    policy = TMConsolidatedExportPolicy(
        source_path=tmp_path / "test-policy.yaml",
        policy_sha256="a" * 64,
        statement_type="TM",
        bank="MBB",
        report_scope="CONSOLIDATED",
        dataset_role="DEVELOPMENT",
        schema_item_count=8,
        schema_workbook_path=Path("template/test.xlsx"),
        schema_workbook_sha256=hashlib.sha256(template_path.read_bytes()).hexdigest(),
        schema_projection_sha256=_projection_sha256(schema),
        output_sheets=TM_CONSOLIDATED_SHEETS,
        owners=(
            TMConsolidatedOwnerPolicy("page-0045", owner_type, "mapped_assignments"),
            TMConsolidatedOwnerPolicy("residual", owner_type, "no_observations"),
        ),
    )
    return _Fixture(template_path, schema, policy, owner_a, owner_b)


def _build(fixture: _Fixture, inputs: tuple[TMConsolidatedOwnerInput, ...] | None = None):
    return build_tm_consolidated_development_artifacts(
        template_path=fixture.template_path,
        workbook_name="mbb-tm-consolidated-development.xlsx",
        schema=fixture.schema,
        owner_inputs=inputs or fixture.inputs,
        policy=fixture.policy,
        run_metadata={"quarter": "2026-Q1", "bank": "MBB"},
    )


def _headers(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def test_production_policy_pins_1712_and_all_27_required_owners(project_root: Path) -> None:
    policy = load_tm_consolidated_export_policy(project_root / TM_CONSOLIDATED_POLICY_RELATIVE_PATH)

    assert policy.schema_item_count == TM_CONSOLIDATED_SCHEMA_COUNT == 1_712
    assert policy.schema_workbook_sha256 == TM_CONSOLIDATED_TEMPLATE_SHA256
    assert policy.schema_projection_sha256 == TM_CONSOLIDATED_SCHEMA_PROJECTION_SHA256
    assert policy.output_sheets == TM_CONSOLIDATED_SHEETS
    assert len(policy.owners) == 27
    assert len({owner.owner_key for owner in policy.owners}) == 27
    assert {"page-0045", "residual"} <= {owner.owner_key for owner in policy.owners}
    assert next(owner for owner in policy.owners if owner.owner_key == "page-0045").result_type == (
        "bctc_ai.mapping.tm_note_page45_mapping.TMPage45MappingResult"
    )
    contracts = audit_tm_consolidated_owner_result_contracts(policy)
    assert tuple(contracts) == tuple(owner.owner_key for owner in policy.owners)
    assert len(contracts) == 27


def test_owner_result_binder_orders_and_type_checks_actual_result_surface(
    development_fixture: _Fixture,
) -> None:
    bound = bind_tm_consolidated_owner_results(
        {
            "residual": development_fixture.owner_b,
            "page-0045": development_fixture.owner_a,
        },
        development_fixture.policy,
    )

    assert tuple(item.owner_key for item in bound) == ("page-0045", "residual")
    assert bound[0].result is development_fixture.owner_a
    assert bound[1].result is development_fixture.owner_b


def test_full_schema_and_long_form_observations_preserve_all_statuses(
    development_fixture: _Fixture,
) -> None:
    artifact = _build(development_fixture)

    assert artifact.schema_item_count == 8
    assert artifact.observation_count == artifact.provenance_count == 4
    assert artifact.status_counts == {
        "AMBIGUOUS": 1,
        "MAPPED": 4,
        "NA": 1,
        "NOT_OBSERVED": 1,
        "UNRESOLVED": 1,
    }
    assert not artifact.fully_verified
    workbook = load_workbook(BytesIO(artifact.workbook_bytes), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == TM_CONSOLIDATED_SHEETS
        main = workbook["TM"]
        main_headers = _headers(main)
        assert main.max_row == 9
        assert [main.cell(row, main_headers["SchemaStatus"]).value for row in range(2, 10)] == [
            "MAPPED",
            "MAPPED",
            "MAPPED",
            "MAPPED",
            "AMBIGUOUS",
            "UNRESOLVED",
            "NOT_OBSERVED",
            "NA",
        ]
        assert [main.cell(row, main_headers["ObservationCount"]).value for row in range(2, 10)] == [
            1,
            1,
            1,
            1,
            0,
            0,
            0,
            0,
        ]

        observations = workbook["OBSERVATIONS"]
        observation_headers = _headers(observations)
        assert [
            observations.cell(row, observation_headers["ValueStatus"]).value for row in range(2, 6)
        ] == ["VALUE", "ZERO", "DASH", "BLANK"]
        assert observations.cell(2, observation_headers["ReportedValue"]).value == 12.5
        assert observations.cell(2, observation_headers["CanonicalValue"]).value == 12_500_000
        assert observations.cell(3, observation_headers["ReportedValue"]).value == 0
        assert observations.cell(3, observation_headers["CanonicalValue"]).value == 0
        assert all(
            observations.cell(row, observation_headers["ReportedValue"]).value is None
            and observations.cell(row, observation_headers["CanonicalValue"]).value is None
            for row in (4, 5)
        )
        assert all(
            observations.cell(row, observation_headers["Scope"]).value == "CONSOLIDATED"
            for row in range(2, 6)
        )
        assert all(
            observations.cell(row, observation_headers["PeriodType"]).value == "SNAPSHOT"
            for row in range(2, 6)
        )
        provenance = workbook["PROVENANCE"]
        provenance_headers = _headers(provenance)
        assert provenance.max_row == 5
        assert [
            provenance.cell(row, provenance_headers["ProvenanceKey"]).value for row in range(2, 6)
        ] == [
            observations.cell(row, observation_headers["ProvenanceKey"]).value
            for row in range(2, 6)
        ]
        assert not any(
            cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        validation = workbook["VALIDATION"]
        validation_headers = _headers(validation)
        page_rows = [
            {
                "owner": validation.cell(row, validation_headers["OwnerKey"]).value,
                "family": validation.cell(row, validation_headers["ValidationFamily"]).value,
                "id": validation.cell(row, validation_headers["CheckId"]).value,
                "status": validation.cell(row, validation_headers["Status"]).value,
            }
            for row in range(2, validation.max_row + 1)
            if validation.cell(row, validation_headers["ValidationFamily"]).value
            == "accounting_checks"
        ]
        assert page_rows == [
            {
                "owner": "page-0045",
                "family": "accounting_checks",
                "id": "VISIBLE_VALUES_BALANCE",
                "status": "PASS",
            },
            {
                "owner": "page-0045",
                "family": "accounting_checks",
                "id": "DASH_CELL_EQUATION",
                "status": "NOT_TESTABLE_DASH_IS_NOT_ZERO",
            },
        ]
    finally:
        workbook.close()

    payload = json.loads(artifact.provenance_bytes)
    assert payload["fully_verified"] is False
    assert payload["summary"]["schema_item_count"] == 8
    assert payload["summary"]["value_status_counts"] == {
        "BLANK": 1,
        "DASH": 1,
        "VALUE": 1,
        "ZERO": 1,
    }
    assert len(payload["observations"]) == len(payload["provenance"]) == 4
    assert all(record["evidence_detail"] for record in payload["provenance"])


def test_build_is_byte_deterministic_and_paired_export_refuses_overwrite(
    tmp_path: Path,
    development_fixture: _Fixture,
) -> None:
    first = _build(development_fixture)
    second = _build(development_fixture)

    assert first.workbook_bytes == second.workbook_bytes
    assert first.provenance_bytes == second.provenance_bytes
    assert hashlib.sha256(first.workbook_bytes).hexdigest() == first.workbook_sha256
    assert hashlib.sha256(first.provenance_bytes).hexdigest() == first.provenance_sha256

    workbook_path = tmp_path / "out" / "tm.xlsx"
    provenance_path = tmp_path / "out" / "tm.provenance.json"
    result = export_tm_consolidated_development(
        template_path=development_fixture.template_path,
        workbook_path=workbook_path,
        provenance_path=provenance_path,
        schema=development_fixture.schema,
        owner_inputs=development_fixture.inputs,
        policy=development_fixture.policy,
    )
    assert result.observation_count == result.provenance_count == 4
    assert workbook_path.read_bytes()
    assert provenance_path.read_bytes()
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="refuses to overwrite"):
        export_tm_consolidated_development(
            template_path=development_fixture.template_path,
            workbook_path=workbook_path,
            provenance_path=provenance_path,
            schema=development_fixture.schema,
            owner_inputs=development_fixture.inputs,
            policy=development_fixture.policy,
        )


def test_temporary_render_directory_does_not_change_artifact_bytes(
    development_fixture: _Fixture,
) -> None:
    moved_assignments = tuple(
        replace(
            assignment,
            source_image_path=f"/different/cache/root/{Path(assignment.source_image_path).name}",
        )
        for assignment in development_fixture.owner_a.mapped_assignments
    )
    moved_owner = replace(
        development_fixture.owner_a,
        mapped_assignments=moved_assignments,
    )

    baseline = _build(development_fixture)
    moved = _build(
        development_fixture,
        (
            TMConsolidatedOwnerInput("page-0045", moved_owner),
            development_fixture.inputs[1],
        ),
    )

    assert moved.workbook_bytes == baseline.workbook_bytes
    assert moved.provenance_bytes == baseline.provenance_bytes


def test_missing_or_duplicate_owner_fails_before_output(development_fixture: _Fixture) -> None:
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="inventory incomplete"):
        _build(development_fixture, development_fixture.inputs[:1])

    duplicated = (
        development_fixture.inputs[0],
        development_fixture.inputs[0],
        development_fixture.inputs[1],
    )
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="duplicate TM owner"):
        _build(development_fixture, duplicated)


def test_duplicate_schema_id_or_owner_fails_closed(development_fixture: _Fixture) -> None:
    duplicate_id_dispositions = list(development_fixture.owner_b.schema_dispositions)
    duplicate_id_dispositions[-1] = replace(
        duplicate_id_dispositions[-1],
        report_norm_id=106,
        display_order=6,
        canonical_name="TM item 6",
    )
    duplicate_id_owner = replace(
        development_fixture.owner_b,
        schema_dispositions=tuple(duplicate_id_dispositions),
    )
    with pytest.raises(
        TMConsolidatedDevelopmentExportError, match="duplicate/invalid TM disposition"
    ):
        _build(
            development_fixture,
            (
                development_fixture.inputs[0],
                TMConsolidatedOwnerInput("residual", duplicate_id_owner),
            ),
        )

    duplicate_owner_dispositions = list(development_fixture.owner_b.schema_dispositions)
    duplicate_owner_dispositions[0] = replace(
        duplicate_owner_dispositions[0],
        status="NOT_OBSERVED_IN_THIS_PDF",
        reason="duplicate ownership mutation",
    )
    duplicate_owner_dispositions[6] = replace(
        duplicate_owner_dispositions[6],
        status="UNASSESSED",
        reason="status-count preserving mutation",
    )
    duplicate_owner = replace(
        development_fixture.owner_b,
        schema_dispositions=tuple(duplicate_owner_dispositions),
    )
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="duplicate owners"):
        _build(
            development_fixture,
            (
                development_fixture.inputs[0],
                TMConsolidatedOwnerInput("residual", duplicate_owner),
            ),
        )


def test_duplicate_provenance_and_status_drift_fail_closed(
    development_fixture: _Fixture,
) -> None:
    duplicate_assignment_owner = replace(
        development_fixture.owner_a,
        mapped_assignments=(
            *development_fixture.owner_a.mapped_assignments,
            development_fixture.owner_a.mapped_assignments[0],
        ),
        mapped_assignment_count=5,
    )
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="duplicate or unlinked"):
        _build(
            development_fixture,
            (
                TMConsolidatedOwnerInput("page-0045", duplicate_assignment_owner),
                development_fixture.inputs[1],
            ),
        )


def test_page_validation_fail_or_duplicate_id_fails_closed(
    development_fixture: _Fixture,
) -> None:
    failed = replace(
        development_fixture.owner_a,
        accounting_checks=(
            replace(development_fixture.owner_a.accounting_checks[0], status="FAIL"),
        ),
    )
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="unsafe TM page validation"):
        _build(
            development_fixture,
            (
                TMConsolidatedOwnerInput("page-0045", failed),
                development_fixture.inputs[1],
            ),
        )

    duplicated = replace(
        development_fixture.owner_a,
        accounting_checks=(
            development_fixture.owner_a.accounting_checks[0],
            development_fixture.owner_a.accounting_checks[0],
        ),
    )
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="duplicate TM page validation"):
        _build(
            development_fixture,
            (
                TMConsolidatedOwnerInput("page-0045", duplicated),
                development_fixture.inputs[1],
            ),
        )

    drifted_dispositions = list(development_fixture.owner_a.schema_dispositions)
    drifted_dispositions[0] = replace(drifted_dispositions[0], status="UNRESOLVED")
    drifted_owner = replace(
        development_fixture.owner_a,
        schema_dispositions=tuple(drifted_dispositions),
    )
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="status counts drifted"):
        _build(
            development_fixture,
            (
                TMConsolidatedOwnerInput("page-0045", drifted_owner),
                development_fixture.inputs[1],
            ),
        )


def test_derived_or_imputed_value_requires_explicit_method_tag(
    development_fixture: _Fixture,
) -> None:
    assignments = list(development_fixture.owner_a.mapped_assignments)
    assignments[0] = replace(assignments[0], is_derived=True)
    untagged = replace(development_fixture.owner_a, mapped_assignments=tuple(assignments))
    with pytest.raises(TMConsolidatedDevelopmentExportError, match="not explicitly tagged"):
        _build(
            development_fixture,
            (
                TMConsolidatedOwnerInput("page-0045", untagged),
                development_fixture.inputs[1],
            ),
        )

    assignments[0] = replace(assignments[0], mapping_basis="SUM_SOURCE_ROWS")
    tagged = replace(development_fixture.owner_a, mapped_assignments=tuple(assignments))
    artifact = _build(
        development_fixture,
        (
            TMConsolidatedOwnerInput("page-0045", tagged),
            development_fixture.inputs[1],
        ),
    )
    payload = json.loads(artifact.provenance_bytes)
    assert payload["observations"][0]["derivation_method"] == "SUM_SOURCE_ROWS"
    assert payload["observations"][0]["derivation_source_ids"] == ["page-0045:row-0"]


def test_nested_adapter_includes_page44_top_level_narrative_assignments(
    development_fixture: _Fixture,
) -> None:
    nested = _NestedMappingResult(
        statement_type="TM",
        report_scope="CONSOLIDATED",
        mapping_authority_scope="TEST_PAGE44_NESTED_OWNER",
        mapping_authority_granted=True,
        schema_item_count=8,
        status_reconciled_schema_count=4,
        mapped_schema_count=4,
        ambiguous_schema_count=0,
        unresolved_schema_count=0,
        not_observed_schema_count=0,
        not_applicable_schema_count=0,
        unassessed_schema_count=4,
        schema_dispositions=development_fixture.owner_a.schema_dispositions,
        source_dispositions=(
            _NestedSource(
                row_id="page-0044:grid-row",
                mapped_assignments=development_fixture.owner_a.mapped_assignments[:3],
            ),
        ),
        narrative_assignments=(
            _NarrativeAssignment(
                source_row_ids=("page-0044:narrative-line-1",),
                value_index=0,
                report_norm_id=103,
                observation="VALUE",
                value=Decimal("5"),
                period_start="2026-03-31",
                period_end="2026-03-31",
                period_type="SNAPSHOT",
                period_role="REPORT_DATE",
                unit="VND",
                unit_multiplier=1,
                mapping_basis="VISIBLE_NARRATIVE_QUANTITY_DIRECT_MAPPING",
                raw_text="Visible narrative quantity: 5 VND",
            ),
        ),
        mapped_value_count=3,
    )
    policy = replace(
        development_fixture.policy,
        owners=(
            TMConsolidatedOwnerPolicy("page-0045", _result_type(nested), "nested_assignments"),
            development_fixture.policy.owners[1],
        ),
    )

    artifact = build_tm_consolidated_development_artifacts(
        template_path=development_fixture.template_path,
        workbook_name="page44-narrative.xlsx",
        schema=development_fixture.schema,
        owner_inputs=(
            TMConsolidatedOwnerInput("page-0045", nested),
            development_fixture.inputs[1],
        ),
        policy=policy,
    )

    payload = json.loads(artifact.provenance_bytes)
    assert artifact.observation_count == artifact.provenance_count == 4
    assert any(
        observation["report_norm_id"] == 103
        and observation["value_status"] == "VALUE"
        and observation["reported_value"] == "5"
        for observation in payload["observations"]
    )


@pytest.fixture(scope="module")
def actual_page45_inputs(project_root: Path, tmp_path_factory: pytest.TempPathFactory):
    from bctc_ai.mapping.tm_note_page45_mapping import (
        TM_PAGE45_MAPPING_POLICY_RELATIVE_PATH,
        load_tm_page45_mapping_policy,
        reconcile_tm_page45_items,
    )
    from bctc_ai.rendering.pdf import render_pages
    from bctc_ai.schema.hierarchy import apply_hierarchy_reference, load_hierarchy_reference
    from bctc_ai.schema.registry import load_all
    from bctc_ai.tables.tm_note_page45 import load_tm_page45_policy, parse_tm_page45

    source_pdf = project_root / "vietstock_bctc/MBB/2026/BCTC Hợp nhất quý 1 năm 2026.pdf"
    render_root = tmp_path_factory.mktemp("tm-export-page45")
    render = Path(render_pages(source_pdf, render_root, dpi=300, page_numbers={45})[0].path)
    parsed = parse_tm_page45(
        project_root / "tests/golden/tm/mbb-q1-2026-page-0045-ppocrv6-word-box.json",
        render,
        load_tm_page45_policy(project_root / "config/tables/tm-note-page45-v1.yaml"),
    )
    _workbooks, schema = load_all(project_root / "template", project_root)
    _hierarchy_payload, hierarchy = load_hierarchy_reference(
        project_root / "config/schemas/hierarchy_reference.yaml",
        project_root,
        schema,
    )
    apply_hierarchy_reference(schema, hierarchy)
    result = reconcile_tm_page45_items(
        parsed,
        schema=schema,
        policy=load_tm_page45_mapping_policy(project_root / TM_PAGE45_MAPPING_POLICY_RELATIVE_PATH),
        source_pdf_path=source_pdf,
        schema_workbook_path=project_root / "template/Bank_TM_ReportNormId.v2.xlsx",
    )
    return tuple(schema), result


def test_actual_page45_production_result_exports_exact_value_dash_blank_surface(
    project_root: Path,
    actual_page45_inputs,
) -> None:
    schema, page45_result = actual_page45_inputs
    tm_schema = tuple(
        sorted(
            (item for item in schema if item.statement_type == "TM"),
            key=lambda item: item.display_order,
        )
    )
    mapped_ids = {
        disposition.report_norm_id
        for disposition in page45_result.schema_dispositions
        if disposition.status == "MAPPED_AUTOMATIC_SCOPED"
    }
    residual_dispositions = tuple(
        _Disposition(
            report_norm_id=item.schema_id,
            display_order=item.display_order,
            canonical_name=item.canonical_name,
            status="UNASSESSED" if item.schema_id in mapped_ids else "NOT_OBSERVED_IN_THIS_PDF",
            source_row_ids=(),
            reason="test-only complement around the actual production page-45 owner",
        )
        for item in tm_schema
    )
    residual = _MappingResult(
        statement_type="TM",
        report_scope="CONSOLIDATED",
        mapping_authority_scope="TEST_COMPLEMENT_OF_ACTUAL_PAGE45",
        mapping_authority_granted=True,
        schema_item_count=1_712,
        status_reconciled_schema_count=1_699,
        mapped_schema_count=0,
        ambiguous_schema_count=0,
        unresolved_schema_count=0,
        not_observed_schema_count=1_699,
        not_applicable_schema_count=0,
        unassessed_schema_count=13,
        schema_dispositions=residual_dispositions,
        source_dispositions=(),
        mapped_assignments=(),
        mapped_assignment_count=0,
    )
    policy = TMConsolidatedExportPolicy(
        source_path=project_root / "config/export/tm-consolidated-development-v1.yaml",
        policy_sha256="b" * 64,
        statement_type="TM",
        bank="MBB",
        report_scope="CONSOLIDATED",
        dataset_role="DEVELOPMENT",
        schema_item_count=1_712,
        schema_workbook_path=Path("template/Bank_TM_ReportNormId.v2.xlsx"),
        schema_workbook_sha256=TM_CONSOLIDATED_TEMPLATE_SHA256,
        schema_projection_sha256=TM_CONSOLIDATED_SCHEMA_PROJECTION_SHA256,
        output_sheets=TM_CONSOLIDATED_SHEETS,
        owners=(
            TMConsolidatedOwnerPolicy(
                "page-0045", _result_type(page45_result), "mapped_assignments"
            ),
            TMConsolidatedOwnerPolicy("residual", _result_type(residual), "no_observations"),
        ),
    )

    artifact = build_tm_consolidated_development_artifacts(
        template_path=project_root / "template/Bank_TM_ReportNormId.v2.xlsx",
        workbook_name="actual-page45-in-memory.xlsx",
        schema=schema,
        owner_inputs=(
            TMConsolidatedOwnerInput("page-0045", page45_result),
            TMConsolidatedOwnerInput("residual", residual),
        ),
        policy=policy,
    )

    assert artifact.schema_item_count == 1_712
    assert artifact.observation_count == artifact.provenance_count == 22
    payload = json.loads(artifact.provenance_bytes)
    assert payload["summary"]["value_status_counts"] == {
        "BLANK": 2,
        "DASH": 8,
        "VALUE": 12,
    }
    observations_by_id = {}
    for observation in payload["observations"]:
        observations_by_id.setdefault(observation["report_norm_id"], []).append(observation)
    assert not observations_by_id.keys() & {5946, 5949}
    assert {item["value_status"] for item in observations_by_id[5950]} == {"BLANK"}
    assert {
        item["value_status"]
        for report_norm_id in (5953, 5954, 5955, 5958)
        for item in observations_by_id[report_norm_id]
    } == {"DASH"}
