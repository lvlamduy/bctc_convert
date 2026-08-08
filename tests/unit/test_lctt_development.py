from __future__ import annotations

import datetime as datetime_module
import hashlib
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from bctc_ai.export.lctt_development import (
    LCTTDevelopmentExportError,
    build_lctt_development_artifact,
    export_lctt_development,
)
from bctc_ai.mapping.lctt_item_mapping import (
    LCTT_POLICY_RELATIVE_PATH,
    load_lctt_direct_mapping_policy,
    reconcile_lctt_direct_items,
)
from bctc_ai.schema.hierarchy import apply_hierarchy_reference, load_hierarchy_reference
from bctc_ai.schema.registry import load_all
from bctc_ai.tables.lctt_word_box import (
    LCTTPageInput,
    load_lctt_word_box_policy,
    parse_lctt_word_box_document,
)

_OCR_ROOT = Path("output/calibration/recovery-e0027-mbb-q1-2026-role-c-20260807")
_RENDER_ROOT = Path(
    "output/calibration/recovery-e0027-mbb-q1-2026-20260807/eebeda2ebc09b0d42032/renders"
)
_TEMPLATE = Path("template/Bank_LCTT_ReportNormId.v2.xlsx")
_DEEPSEEK_LABELS = (
    "| LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG KINH DOANH",
    "| Thu lãi và các khoản thu tương tự nhận được",
    "| Chi lài và các khoản chi tương tự đã trả",
    "Thu nhập từ hoạt động dịch vụ nhận được",
    "Chênh lệch số tiền thực thu/(chỉ) từ hoạt động kinh doanh (ngoại tệ, vàng bạc, chứng khoán)",
    "| Thu nhập/(Chi phí) khác",
    "Tiền thu các khoản nợ đã được xử lý, xoá, bù đắp bằng nguồn rủi ro",
    "Tiền chi trả cho nhân viên và hoạt động quản lý, công vụ",
    "Tiền thuế thu nhập doanh nghiệp thực nộp trong kỳ",
    "Lưu chuyển tiền thuần từ hoạt động kinh doanh trước những thay đổi về tài sản và vốn lưu động",
    "Những thay đổi về tài sản hoạt động",
    "(Tăng)/Giảm các khoản tiền gửi và cho vay các TCTD khác",
    "(Tăng)/Giảm các khoản về kinh doanh chứng khoán",
    "(Tăng)/Giảm các công cụ tài chính phái sinh và các tài sản tài chính khác",
    "(Tăng)/Giảm các khoản cho vay khách hàng và mua nợ",
    "Giảm nguồn dự phòng để xử lý rủi ro, xử lý, bù đắp tổn thất các khoản (tín dụng, "
    "chứng khoán, đầu tư, phải thu khác)",
    "(Tăng)/Giảm khác về tài sản hoạt động",
    "Những thay đổi về công nghệ hoạt động",
    "| [Tăng/(Giảm)] các khoản nợ Chính phủ và NHNN",
    "Tăng/(Giảm) các khoản tiền gửi, tiền vay các TCTD khác",
    "Tăng/(giảm) tiền gửi của khách hàng",
    "Tăng/(Giảm) phát hành giấy tờ có giá",
    "TẠNG/(Giám) vốn tài trợ, uỷ thác đầu tư, cho vay mà TCTD chịu rủi ro",
    "Tăng/(Giảm) các công cụ tài chính phái sinh và các tài sản tài chính khác",
    "Tăng/(Giảm) khác về công nợ hoạt động",
    "| Chi tiết các quỹ của TCTD",
    "| Lưu chuyển tiền thuần từ hoạt động kinh doanh",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG ĐẦU TƯ",
    "Mua sắm tài sản cố định",
    "Tiền thu từ thanh lý, nhượng bán tài sản cố định",
    "| Tiền thu/(chi) bất động sản đầu tư",
    "| Tiền thu/(chi) đầu tư, góp vốn vào các đơn vị khác",
    "Tên thủ công của lợi nhuận được chia từ các khoản đầu tư, góp vốn dài hạn",
    "//Lưu chuyển tiền thuần sử dụng trong hoạt động đầu tư",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG TÀI CHÍNH",
    "|Tăng vốn cổ phần từ góp vốn và/hoặc phát hành cổ phiếu",
    "Tiền thu từ phát hành giấy tờ có giá dài hạn đủ điều kiện tính vào vốn tự có và "
    "các khoản vốn vay dài hạn khác",
    "Tiền chi thanh toán giấy tờ có giá dài hạn có đủ điều kiện tính vào vốn tự có và "
    "các khoản vốn vay dài hạn khác",
    "Cố tức trả cho cổ đông",
    "| Lưu chuyển tiền thuần từ hoạt động tài chính",
    "| Lưu chuyển tiền thuần trong kỳ",
    "| Tiền và các khoản tương đương tiền tại thời điểm đầu kỳ",
    "# Tiền và các khoản tương đương tiền tại thời điểm cuối kỳ",
)


@pytest.fixture(scope="module")
def development_inputs(project_root: Path):
    inputs = tuple(
        LCTTPageInput(
            result_path=(project_root / _OCR_ROOT / f"ppocrv6-page-{page:04d}" / "ocr_result.json"),
            source_image_path=project_root / _RENDER_ROOT / f"page-{page:04d}.png",
            page_tag=f"page-{page:04d}",
        )
        for page in (7, 8)
    )
    if not all(page.result_path.is_file() and page.source_image_path.is_file() for page in inputs):
        pytest.skip("MBB Q1/2026 LCTT source-visible artifacts are not local")
    parsed = parse_lctt_word_box_document(
        inputs,
        load_lctt_word_box_policy(project_root / "config/tables/lctt-word-box-v1.yaml"),
    )
    _workbooks, schema = load_all(project_root / "template", project_root)
    _registry, hierarchy = load_hierarchy_reference(
        project_root / "config/schemas/hierarchy_reference.yaml",
        project_root,
        schema,
    )
    apply_hierarchy_reference(schema, hierarchy)
    semantic_rows = tuple(
        {
            "row_id": row.row_id,
            "page_tag": row.page_tag,
            "proposal_text": label,
        }
        for row, label in zip(parsed.rows, _DEEPSEEK_LABELS, strict=True)
    )
    mapping = reconcile_lctt_direct_items(
        parsed.rows,
        schema=schema,
        policy=load_lctt_direct_mapping_policy(project_root / LCTT_POLICY_RELATIVE_PATH),
        report_scope=parsed.scope,
        cash_flow_method=parsed.method,
        independent_semantic_rows=semantic_rows,
        independent_source_reader_id="deepseek-ocr2-logical-row",
    )
    return parsed, mapping


def _headers(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def test_real_lctt_export_preserves_template_and_exposes_43_source_rows(
    tmp_path: Path,
    project_root: Path,
    development_inputs,
) -> None:
    parsed, mapping = development_inputs
    workbook_path = tmp_path / "mbb-lctt-development.xlsx"
    result = export_lctt_development(
        template_path=project_root / _TEMPLATE,
        workbook_path=workbook_path,
        parsed=parsed,
        mapping=mapping,
    )

    assert result.mapped_cell_count == 86
    assert result.mapped_value_count == 71
    assert result.provenance_row_count == 43
    assert result.fully_verified is False
    assert hashlib.sha256(workbook_path.read_bytes()).hexdigest() == result.workbook_sha256

    template = load_workbook(project_root / _TEMPLATE, data_only=False)
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        assert workbook.sheetnames == ["LCTT", "PROVENANCE", "RUN_METADATA"]
        main = workbook["LCTT"]
        for row in range(1, 111):
            for column in range(1, 4):
                assert main.cell(row, column).value == template.active.cell(row, column).value
                assert main.cell(row, column).style_id == template.active.cell(row, column).style_id

        by_id = {main.cell(row, 2).value: row for row in range(2, 111)}
        assert main.cell(by_id[4123], 4).value == 26_904_675_000_000
        assert main.cell(by_id[4123], 9).value == 18_186_686_000_000
        assert main.cell(by_id[4123], 5).value == "VALUE"
        assert main.cell(by_id[4123], 10).value == "VALUE"
        assert main.cell(by_id[4104], 5).value == "BLANK"
        assert main.cell(by_id[4104], 10).value == "BLANK"
        for report_norm_id in (
            4143,
            4144,
            4145,
            4146,
            4120,
            4121,
            4151,
            4152,
            4117,
        ):
            assert main.cell(by_id[report_norm_id], 5).value == "NOT_OBSERVED_IN_THIS_PDF"
            assert main.cell(by_id[report_norm_id], 10).value == "NOT_OBSERVED_IN_THIS_PDF"
        assert main.cell(by_id[4140], 4).value == -37_183_000_000
        assert main.cell(by_id[4140], 9).value == 334_598_000_000
        assert main.cell(by_id[4140], 5).value == "VALUE"
        assert main.cell(by_id[4140], 10).value == "VALUE"
        assert main.cell(by_id[6034], 4).value is None
        assert main.cell(by_id[6034], 9).value is None
        assert main.cell(by_id[6034], 5).value == "DASH"
        assert main.cell(by_id[6034], 10).value == "DASH"
        assert main.cell(by_id[5714], 4).value == 490_000_000
        assert main.cell(by_id[5714], 9).value == -71_299_000_000
        assert main.cell(by_id[5714], 5).value == "VALUE"
        assert main.cell(by_id[5714], 10).value == "VALUE"
        assert main.cell(by_id[4155], 5).value == "SCHEMA_ITEM_NOT_APPLICABLE"
        assert (
            sum(
                main.cell(row, column).value is not None
                for row in range(2, 111)
                for column in (4, 9)
            )
            == 71
        )

        provenance = workbook["PROVENANCE"]
        assert provenance.max_row == 44
        headers = _headers(provenance)
        records = [
            {name: provenance.cell(row, column).value for name, column in headers.items()}
            for row in range(2, provenance.max_row + 1)
        ]
        assert {record["SourceRowStatus"] for record in records} == {"MAPPED_AUTOMATIC"}
        business = {
            record["RowId"]: record
            for record in records
            if record["BusinessResolutionKey"] is not None
        }
        assert set(business) == {
            "page-0007:row-0024",
            "page-0007:row-0031",
            "page-0007:row-0032",
        }
        assert business["page-0007:row-0024"]["CandidateReportNormIdsJson"] == "[4140]"
        assert business["page-0007:row-0024"]["CurrentValueVND"] == -37_183_000_000
        assert business["page-0007:row-0024"]["ComparativeValueVND"] == 334_598_000_000
        assert business["page-0007:row-0024"]["BusinessResolutionKey"] == (
            "USER_Q018_CONTEXTUAL_WORDING_TO_4140"
        )
        assert (
            "USER_AUTHORIZED_CONTEXTUAL_WORDING_MAPPING"
            in business["page-0007:row-0024"]["MappingReason"]
        )
        assert business["page-0007:row-0031"]["CandidateReportNormIdsJson"] == "[6034]"
        assert business["page-0007:row-0031"]["CurrentStatus"] == "DASH"
        assert business["page-0007:row-0031"]["ComparativeStatus"] == "DASH"
        assert business["page-0007:row-0032"]["CandidateReportNormIdsJson"] == "[5714]"
        assert business["page-0007:row-0032"]["CurrentValueVND"] == 490_000_000
        assert business["page-0007:row-0032"]["ComparativeValueVND"] == -71_299_000_000
        assert all(record["FullyVerified"] is False for record in records)
        metadata = {
            workbook["RUN_METADATA"].cell(row, 1).value: workbook["RUN_METADATA"].cell(row, 2).value
            for row in range(2, workbook["RUN_METADATA"].max_row + 1)
        }
        assert metadata["mapping.investment_property_net_id"] == 6034
        assert metadata["mapping.investment_property_component_ids"] == "[4144,4145,4146]"
        assert metadata["mapping.investment_property_formula_policy"] == (
            "VALIDATION_ONLY_NO_DERIVATION_OR_IMPUTATION"
        )
        assert not any(
            cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()
        template.close()


def test_lctt_build_is_deterministic_and_export_refuses_overwrite(
    tmp_path: Path,
    project_root: Path,
    development_inputs,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from openpyxl.writer import excel as excel_writer

    save_clock = iter(
        (
            datetime_module.datetime(2030, 1, 2, 3, 4, 5),
            datetime_module.datetime(2040, 6, 7, 8, 9, 10),
        )
    )

    class FakeDateTime:
        @classmethod
        def now(cls, tz=None):
            value = next(save_clock)
            return value.replace(tzinfo=tz) if tz is not None else value

    monkeypatch.setattr(
        excel_writer,
        "datetime",
        SimpleNamespace(datetime=FakeDateTime, timezone=datetime_module.timezone),
    )
    parsed, mapping = development_inputs
    kwargs = {
        "template_path": project_root / _TEMPLATE,
        "parsed": parsed,
        "mapping": mapping,
    }
    first = build_lctt_development_artifact(**kwargs)
    second = build_lctt_development_artifact(**kwargs)
    assert first.workbook_bytes == second.workbook_bytes
    assert first.workbook_sha256 == hashlib.sha256(first.workbook_bytes).hexdigest()
    workbook = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    try:
        assert workbook.properties.created == datetime_module.datetime(2000, 1, 1)
        assert workbook.properties.modified == datetime_module.datetime(2000, 1, 1)
    finally:
        workbook.close()

    destination = tmp_path / "existing.xlsx"
    destination.write_bytes(b"KEEP")
    with pytest.raises(LCTTDevelopmentExportError, match="overwrite"):
        export_lctt_development(
            workbook_path=destination,
            **kwargs,
        )
    assert destination.read_bytes() == b"KEEP"
