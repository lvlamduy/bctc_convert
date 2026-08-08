from __future__ import annotations

import copy
import hashlib
import json
import shutil
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from bctc_ai.export import native_rows as export_module
from bctc_ai.export.native_rows import (
    EXPORT_POLICY_RELATIVE_PATH,
    NativeRowsExcelExportError,
    NativeRowsInputIdentity,
    build_native_rows_excel_artifacts,
    export_registered_native_rows_excel,
    load_native_rows_excel_policy,
)

_DOCUMENT_SHA256 = "a" * 64


def _bbox(x0: float, y0: float, x1: float, y1: float) -> dict[str, float]:
    return {"x0": x0, "y0": y0, "x1": x1, "y1": y1}


def _cell(
    row_id: str,
    axis_id: str,
    raw_text: str,
    normalized_text: str,
    value: str | None,
    observation: str,
    source_status: str,
    *,
    index: int,
    parse_reason: str | None = None,
) -> dict[str, Any]:
    box = _bbox(300 + index * 20, 100 + index * 8, 318 + index * 20, 106 + index * 8)
    return {
        "axis_id": axis_id,
        "raw_text": raw_text,
        "normalized_text": normalized_text,
        "value": value,
        "observation": observation,
        "source_status": source_status,
        "sign_evidence": "dash" if observation == "DASH" else None,
        "parse_reason": parse_reason,
        "bbox": box,
        "run_id": f"run-{index}",
        "axis_distance": float(index) / 10,
        "provenance": {
            "document_sha256": _DOCUMENT_SHA256,
            "page": 3,
            "table_id": "source-table",
            "row_id": row_id,
            "column_id": axis_id,
            "value_bbox": box,
        },
    }


def _row(
    ordinal: int,
    *,
    row_type: str,
    raw_label: str,
    normalized_label: str,
    cells: list[dict[str, Any]],
    within_span: bool,
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    row_id = f"source-table:row-{ordinal:04d}"
    label_boxes = [] if not raw_label else [_bbox(10, ordinal * 20, 180, ordinal * 20 + 8)]
    return {
        "row_id": row_id,
        "page": 3,
        "row_type": row_type,
        "source_status": "OBSERVED_ROW",
        "raw_label": raw_label,
        "normalized_label": normalized_label,
        "label_bboxes": label_boxes,
        "raw_note_reference": None,
        "note_reference": None,
        "note_bbox": None,
        "cells": cells,
        "y0": float(ordinal * 20),
        "y1": float(ordinal * 20 + 8),
        "indentation": 10.0 if raw_label else 0.0,
        "within_financial_table_span": within_span,
        "warnings": list(warnings or []),
        "provenance": {
            "document_sha256": _DOCUMENT_SHA256,
            "page": 3,
            "table_id": "source-table",
            "row_id": row_id,
            "label_bboxes": label_boxes,
            "note_bbox": None,
        },
    }


def _payload() -> dict[str, Any]:
    outside_id = "source-table:row-0001"
    unlabeled_id = "source-table:row-0003"
    data_id = "source-table:row-0004"
    outside = _row(
        1,
        row_type="NON_DATA",
        raw_label="Outside audit row",
        normalized_label="Outside audit row",
        cells=[
            _cell(
                outside_id,
                "value-1",
                "5",
                "5",
                "5",
                "VALUE",
                "OBSERVED_VALUE",
                index=1,
            )
        ],
        within_span=False,
        warnings=["outside financial table span retained for audit"],
    )
    section = _row(
        2,
        row_type="SECTION_HEADER",
        raw_label="SECTION",
        normalized_label="SECTION",
        cells=[],
        within_span=True,
        warnings=["label-only row retained for ordered context"],
    )
    unlabeled = _row(
        3,
        row_type="DATA_ROW",
        raw_label="",
        normalized_label="",
        cells=[
            _cell(
                unlabeled_id,
                "value-1",
                "-",
                "-",
                None,
                "DASH",
                "DASH",
                index=2,
            ),
            _cell(
                unlabeled_id,
                "value-2",
                "=2+2",
                "=2+2",
                None,
                "INVALID",
                "UNRESOLVED",
                index=3,
                parse_reason="unsupported characters",
            ),
        ],
        within_span=True,
        warnings=["numeric row has no attached label"],
    )
    data = _row(
        4,
        row_type="DATA_ROW",
        raw_label="=literal source label",
        normalized_label="=literal source label",
        cells=[
            _cell(
                data_id,
                "value-1",
                "0",
                "0",
                "0",
                "ZERO",
                "OBSERVED_ZERO",
                index=4,
            ),
            _cell(
                data_id,
                "value-2",
                "123456789012345678901234567890",
                "123456789012345678901234567890",
                "123456789012345678901234567890",
                "VALUE",
                "OBSERVED_VALUE",
                index=5,
            ),
        ],
        within_span=True,
    )
    headers = [
        {
            "axis_id": "value-1",
            "raw_header": "Current",
            "header_bbox": _bbox(280, 40, 340, 55),
            "unit": "VND",
            "unit_multiplier": 1_000_000,
            "unit_bbox": _bbox(250, 20, 280, 30),
            "period_start": "2026-03-31",
            "period_end": "2026-03-31",
            "period_type": "SNAPSHOT",
            "duration_months": None,
            "current_or_comparative": "CURRENT",
            "restated": False,
            "confidence": 1.0,
            "evidence": ["visible header"],
        },
        {
            "axis_id": "value-2",
            "raw_header": "Comparative",
            "header_bbox": _bbox(360, 40, 430, 55),
            "unit": "VND",
            "unit_multiplier": 1_000_000,
            "unit_bbox": _bbox(250, 20, 280, 30),
            "period_start": "2025-12-31",
            "period_end": "2025-12-31",
            "period_type": "SNAPSHOT",
            "duration_months": None,
            "current_or_comparative": "COMPARATIVE",
            "restated": False,
            "confidence": 1.0,
            "evidence": ["visible header"],
        },
    ]
    page = {
        "page": 3,
        "statement_type": "CDKT",
        "scope": "MAIN_STATEMENT",
        "discovery_contract": {
            "page": 3,
            "statement_type": "CDKT",
            "scope": "MAIN_STATEMENT",
            "mapping_eligible": True,
            "locally_accepted": True,
        },
        "text_quality": "USABLE_TEXT_LAYER",
        "corruption_markers": [],
        "width_points": 600.0,
        "height_points": 800.0,
        "rotation": 0,
        "native_word_count": 20,
        "native_words_sha256": "b" * 64,
        "geometry": {
            "authority": "PYMUPDF_NATIVE_TEXT_WORDS",
            "data_start_y": 60.0,
            "data_end_y": 700.0,
            "label_right_boundary": 250.0,
            "edge_tolerance": 2.0,
            "axes": [
                {
                    "axis_id": "value-1",
                    "role": "VALUE",
                    "right_edge": 340.0,
                    "left_edge": 280.0,
                    "sample_count": 4,
                    "source": "native",
                },
                {
                    "axis_id": "value-2",
                    "role": "VALUE",
                    "right_edge": 430.0,
                    "left_edge": 360.0,
                    "sample_count": 4,
                    "source": "native",
                },
            ],
            "unit_run_ids": ["unit-1"],
            "warnings": [],
        },
        "headers": headers,
        "rows": [section, unlabeled, data],
        "outside_financial_table_span_rows": [outside],
        "reconstructed_row_count": 4,
        "financial_table_span_row_count": 3,
        "outside_financial_table_span_row_count": 1,
    }
    return {
        "format_version": "REGISTERED_NATIVE_STATEMENT_ROWS_RESULT_V1",
        "policy": "REGISTERED_NATIVE_STATEMENT_ROWS_V1",
        "claim_boundary": "UNMAPPED_SOURCE_ROWS_AND_CELLS_ONLY",
        "status": "ACCEPTED_NATIVE_STATEMENT_ROWS",
        "run_id": "native-unit-001",
        "source": {
            "document_id": f"sha256:{_DOCUMENT_SHA256}",
            "relative_path": "data/source.pdf",
            "sha256": _DOCUMENT_SHA256,
            "size_bytes": 123,
            "bank": "TEST",
            "year": 2026,
            "dataset_role": "CALIBRATION",
            "registry_state": "REGISTERED",
            "hash_verified_stable": True,
            "immutable_role_assignment": True,
        },
        "statement_discovery": {
            "path": "output/calibration/discovery.json",
            "sha256": "c" * 64,
            "size_bytes": 456,
            "format_version": "REGISTERED_NATIVE_TEXT_STATEMENT_DISCOVERY_RESULT_V1",
            "status": "ACCEPTED_NATIVE_TEXT_STATEMENT_DISCOVERY",
            "run_id": "discovery-unit-001",
            "producer_git_commit": "d" * 40,
        },
        "code": {"commit": "e" * 40, "dirty": False, "implementation": []},
        "authority": {
            "geometry": "PYMUPDF_NATIVE_TEXT_WORDS",
            "evidence_source": "PYMUPDF_NATIVE_TEXT_GEOMETRY",
            "row_reconstruction": "PDF_STATEMENT_ROWS",
            "financial_table_span": "FINANCIAL_TABLE_SPAN",
            "header_binding": "AXIS_LOCAL_VISIBLE_HEADER_BINDING",
            "semantic_reader": None,
            "schema_mapper": None,
        },
        "isolation": {
            "prior_answer_artifacts_loaded": False,
            "historical_values_loaded": False,
            "role_a_outputs_loaded": False,
            "schema_inputs_loaded": False,
            "template_inputs_loaded": False,
            "bank_identity_used_for_row_reconstruction": False,
            "filename_identity_used_for_row_reconstruction": False,
            "page_number_rules_used_for_row_reconstruction": False,
            "runtime_input_policy": "EXACT_DECLARED_PROJECT_INPUT_LEDGER",
        },
        "inputs": {"runtime_read_ledger": [], "runtime_read_ledger_sha256": "f" * 64},
        "selection": {
            "policy": "ACCEPTED_DISCOVERY_CONTRACTS_ONLY",
            "selected_pages": [3],
            "selected_page_count": 1,
            "statement_page_counts": {"CDKT": 1},
            "notes_pages_selected": 0,
        },
        "summary": {
            "page_count": 1,
            "pages_sha256": "1" * 64,
            "financial_table_span_row_count": 3,
            "cell_count": 4,
            "section_header_count": 1,
            "unlabeled_numeric_row_count": 1,
            "cell_source_status_counts": {
                "DASH": 1,
                "OBSERVED_VALUE": 1,
                "OBSERVED_ZERO": 1,
                "UNRESOLVED": 1,
            },
            "schema_items_created": 0,
            "schema_items_mapped": 0,
        },
        "pages": [page],
    }


def _canonical(payload: dict[str, Any]) -> bytes:
    return (
        json.dumps(payload, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")


def _headers(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def test_workbook_is_source_faithful_deterministic_and_hash_bound(project_root: Path):
    payload = _payload()
    encoded = _canonical(payload)
    digest = hashlib.sha256(encoded).hexdigest()
    policy = load_native_rows_excel_policy(project_root / EXPORT_POLICY_RELATIVE_PATH, project_root)
    kwargs = {
        "input_identity": NativeRowsInputIdentity(
            "output/calibration/native-rows.json", digest, len(encoded)
        ),
        "workbook_filename": "native-rows.xlsx",
        "policy": policy,
        "implementation_ledger": (
            {"path": "src/exporter.py", "sha256": "2" * 64, "size_bytes": 10},
        ),
    }
    first = build_native_rows_excel_artifacts(payload, **kwargs)
    second = build_native_rows_excel_artifacts(payload, **kwargs)
    assert first.workbook_bytes == second.workbook_bytes
    assert first.provenance_bytes == second.provenance_bytes

    workbook = load_workbook(BytesIO(first.workbook_bytes), data_only=False)
    try:
        assert workbook.sheetnames == ["PAGES", "ROWS", "CELLS", "HEADERS", "RUN_METADATA"]
        rows = workbook["ROWS"]
        row_headers = _headers(rows)
        assert [
            rows.cell(row, row_headers["SourceRowOrdinal"]).value
            for row in range(2, rows.max_row + 1)
        ] == [1, 2, 3, 4]
        assert rows.cell(2, row_headers["SourceBucket"]).value == ("OUTSIDE_FINANCIAL_TABLE_SPAN")
        assert rows.cell(3, row_headers["RowType"]).value == "SECTION_HEADER"
        assert rows.cell(3, row_headers["CellCount"]).value == 0
        assert rows.cell(3, row_headers["Y0"]).value == "40.0"
        assert (
            "numeric row has no attached label" in rows.cell(4, row_headers["WarningsJson"]).value
        )
        assert rows.cell(5, row_headers["RawLabel"]).value == "=literal source label"
        assert rows.cell(5, row_headers["RawLabel"]).data_type == "s"

        cells = workbook["CELLS"]
        cell_headers = _headers(cells)
        by_raw = {
            cells.cell(row, cell_headers["RawText"]).value: row
            for row in range(2, cells.max_row + 1)
        }
        dash = by_raw["-"]
        zero = by_raw["0"]
        hostile = by_raw["=2+2"]
        huge = by_raw["123456789012345678901234567890"]
        assert cells.cell(dash, cell_headers["Observation"]).value == "DASH"
        assert cells.cell(dash, cell_headers["ValueText"]).value is None
        assert cells.cell(zero, cell_headers["Observation"]).value == "ZERO"
        assert cells.cell(zero, cell_headers["ValueText"]).value == "0"
        assert cells.cell(huge, cell_headers["ValueText"]).value == (
            "123456789012345678901234567890"
        )
        assert cells.cell(hostile, cell_headers["RawText"]).data_type == "s"
        assert cells.cell(dash, cell_headers["AxisDistance"]).value == "0.2"
        assert cells.cell(dash, cell_headers["PeriodEnd"]).value == "2026-03-31"
        assert cells.cell(dash, cell_headers["UnitMultiplier"]).value == 1_000_000
        assert cells.cell(dash, cell_headers["Scope"]).value == "MAIN_STATEMENT"
        pages = workbook["PAGES"]
        assert pages.cell(2, _headers(pages)["WidthPoints"]).value == "600.0"
        headers = workbook["HEADERS"]
        assert headers.cell(2, _headers(headers)["Confidence"]).value == "1.0"
        assert not any(
            "Schema" in str(cell.value) or "ReportNorm" in str(cell.value)
            for sheet in workbook.worksheets
            for cell in sheet[1]
        )
        assert not any(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()

    receipt = json.loads(first.provenance_bytes)
    assert receipt["input"]["sha256"] == digest
    assert receipt["workbook"]["sha256"] == hashlib.sha256(first.workbook_bytes).hexdigest()
    assert receipt["summary"]["financial_table_span_cell_count"] == 4
    assert receipt["summary"]["outside_financial_table_span_cell_count"] == 1
    assert receipt["summary"]["all_source_cell_count"] == 5
    assert receipt["summary"]["financial_table_span_row_count"] == 3
    assert receipt["summary"]["outside_financial_table_span_row_count"] == 1
    assert receipt["summary"]["all_source_row_count"] == 4
    assert receipt["isolation"]["strict_source_contract_revalidated"] is True
    assert {
        "SOURCE_PDF",
        "SOURCE_REGISTRY",
        "DATASET_ROLE_REGISTRY",
        "ACCEPTED_STATEMENT_DISCOVERY",
        "NATIVE_STATEMENT_ROWS_POLICY",
        "NATIVE_TEXT_QUALITY_CONFIG",
        "GEOMETRY_CONFIG",
        "PRODUCER_GIT_COMMIT",
    }.issubset(receipt["isolation"]["runtime_input_allowlist"])
    assert b"provenance_sha256" not in first.provenance_bytes
    assert b"/workspace/" not in first.provenance_bytes


def test_bounded_page_false_is_preserved_and_oversize_source_text_fails_closed(
    project_root: Path,
):
    payload = _payload()
    payload["pages"][0]["discovery_contract"]["locally_accepted"] = False
    encoded = _canonical(payload)
    policy = load_native_rows_excel_policy(project_root / EXPORT_POLICY_RELATIVE_PATH, project_root)
    kwargs = {
        "input_identity": NativeRowsInputIdentity(
            "output/calibration/native-rows.json",
            hashlib.sha256(encoded).hexdigest(),
            len(encoded),
        ),
        "workbook_filename": "native-rows.xlsx",
        "policy": policy,
        "implementation_ledger": (),
    }
    artifact = build_native_rows_excel_artifacts(payload, **kwargs)
    workbook = load_workbook(BytesIO(artifact.workbook_bytes), data_only=False)
    try:
        pages = workbook["PAGES"]
        assert pages.cell(2, _headers(pages)["LocallyAccepted"]).value is False
    finally:
        workbook.close()

    oversized = copy.deepcopy(payload)
    oversized["pages"][0]["rows"][2]["raw_label"] = "x" * 33_000
    oversized["pages"][0]["rows"][2]["normalized_label"] = "x" * 33_000
    oversized_encoded = _canonical(oversized)
    oversized_kwargs = {
        **kwargs,
        "input_identity": NativeRowsInputIdentity(
            "output/calibration/native-rows.json",
            hashlib.sha256(oversized_encoded).hexdigest(),
            len(oversized_encoded),
        ),
    }
    with pytest.raises(NativeRowsExcelExportError, match="cell text limit"):
        build_native_rows_excel_artifacts(oversized, **oversized_kwargs)

    with pytest.raises(NativeRowsExcelExportError, match="canonical relative POSIX"):
        build_native_rows_excel_artifacts(
            payload,
            **{
                **kwargs,
                "input_identity": NativeRowsInputIdentity(
                    "../native-rows.json",
                    kwargs["input_identity"].sha256,
                    kwargs["input_identity"].size_bytes,
                ),
            },
        )
    with pytest.raises(NativeRowsExcelExportError, match="implementation path"):
        build_native_rows_excel_artifacts(
            payload,
            **{
                **kwargs,
                "implementation_ledger": (
                    {"path": "/workspace/exporter.py", "sha256": "2" * 64, "size_bytes": 1},
                ),
            },
        )


def test_export_requires_trusted_hash_same_role_and_never_overwrites(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    root = tmp_path.resolve()
    policy_path = root / EXPORT_POLICY_RELATIVE_PATH
    policy_path.parent.mkdir(parents=True)
    shutil.copyfile(project_root / EXPORT_POLICY_RELATIVE_PATH, policy_path)
    copied_policy = root / "config/export/copied-policy.yaml"
    shutil.copyfile(policy_path, copied_policy)
    with pytest.raises(NativeRowsExcelExportError, match="canonical policy"):
        load_native_rows_excel_policy(copied_policy, root)
    payload = _payload()
    encoded = _canonical(payload)
    digest = hashlib.sha256(encoded).hexdigest()
    rows_path = root / "output/calibration/native-rows.json"
    rows_path.parent.mkdir(parents=True)
    rows_path.write_bytes(encoded)
    calls: list[str] = []

    def fake_loader(
        path: Path,
        *,
        project_root: Path,
        expected_sha256: str,
        policy_path: Path | None = None,
    ) -> dict[str, Any]:
        del project_root, policy_path
        assert path == rows_path
        calls.append(expected_sha256)
        return payload

    implementation = ({"path": "src/exporter.py", "sha256": "2" * 64, "size_bytes": 10},)
    monkeypatch.setattr(export_module, "load_registered_native_statement_rows", fake_loader)
    monkeypatch.setattr(export_module, "_implementation_ledger", lambda _root: implementation)
    workbook_path = root / "output/calibration/export/native-rows.xlsx"
    provenance_path = root / "output/calibration/export/native-rows.provenance.json"
    result = export_registered_native_rows_excel(
        project_root=root,
        rows_path=rows_path,
        expected_sha256=digest,
        workbook_path=workbook_path,
        provenance_path=provenance_path,
        policy_path=policy_path,
    )
    assert calls == [digest]
    assert result.workbook_sha256 == hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    assert result.provenance_sha256 == hashlib.sha256(provenance_path.read_bytes()).hexdigest()
    original_pair = (workbook_path.read_bytes(), provenance_path.read_bytes())

    with pytest.raises(NativeRowsExcelExportError, match="overwrite"):
        export_registered_native_rows_excel(
            project_root=root,
            rows_path=rows_path,
            expected_sha256=digest,
            workbook_path=workbook_path,
            provenance_path=provenance_path,
            policy_path=policy_path,
        )
    assert (workbook_path.read_bytes(), provenance_path.read_bytes()) == original_pair

    with pytest.raises(NativeRowsExcelExportError, match="must stay under"):
        export_registered_native_rows_excel(
            project_root=root,
            rows_path=rows_path,
            expected_sha256=digest,
            workbook_path=root / "output/development/cross-role.xlsx",
            provenance_path=root / "output/development/cross-role.json",
            policy_path=policy_path,
        )


def test_export_pair_bytes_do_not_depend_on_absolute_project_root(
    tmp_path: Path,
    project_root: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    payload = _payload()
    encoded = _canonical(payload)
    digest = hashlib.sha256(encoded).hexdigest()
    implementation = ({"path": "src/exporter.py", "sha256": "2" * 64, "size_bytes": 10},)

    def fake_loader(
        path: Path,
        *,
        project_root: Path,
        expected_sha256: str,
        policy_path: Path | None = None,
    ) -> dict[str, Any]:
        del project_root, policy_path
        assert path.read_bytes() == encoded
        assert expected_sha256 == digest
        return payload

    monkeypatch.setattr(export_module, "load_registered_native_statement_rows", fake_loader)
    monkeypatch.setattr(export_module, "_implementation_ledger", lambda _root: implementation)
    pairs: list[tuple[bytes, bytes]] = []
    for name in ("first-root", "second-root"):
        root = tmp_path / name
        policy_path = root / EXPORT_POLICY_RELATIVE_PATH
        policy_path.parent.mkdir(parents=True)
        shutil.copyfile(project_root / EXPORT_POLICY_RELATIVE_PATH, policy_path)
        rows_path = root / "output/calibration/native-rows.json"
        rows_path.parent.mkdir(parents=True)
        rows_path.write_bytes(encoded)
        workbook_path = root / "output/calibration/export/native-rows.xlsx"
        provenance_path = root / "output/calibration/export/native-rows.provenance.json"
        export_registered_native_rows_excel(
            project_root=root,
            rows_path=rows_path,
            expected_sha256=digest,
            workbook_path=workbook_path,
            provenance_path=provenance_path,
            policy_path=policy_path,
        )
        pairs.append((workbook_path.read_bytes(), provenance_path.read_bytes()))
    assert pairs[0] == pairs[1]
    assert str(tmp_path).encode() not in pairs[0][1]


def test_pair_rolls_back_workbook_when_completion_marker_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    workbook = tmp_path / "native.xlsx"
    provenance = tmp_path / "native.json"
    original = export_module._write_exclusive
    calls = 0
    fsynced: list[Path] = []

    def fail_second(path: Path, payload: bytes):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise NativeRowsExcelExportError("synthetic provenance failure")
        return original(path, payload)

    monkeypatch.setattr(export_module, "_write_exclusive", fail_second)
    monkeypatch.setattr(export_module, "_fsync_directory", fsynced.append)
    with pytest.raises(NativeRowsExcelExportError, match="synthetic"):
        export_module._publish_pair(workbook, provenance, b"workbook", b"provenance")
    assert not workbook.exists()
    assert not provenance.exists()
    assert fsynced == [tmp_path]
